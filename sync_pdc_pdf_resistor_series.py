from __future__ import annotations

import argparse
import datetime as dt
import shutil
import sqlite3
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import pandas as pd

import component_matcher as cm
from incremental_semiconductor_cache_update import refresh_search_sidecar_rows
from sync_selected_cache_rows import stream_replace_prepared_rows


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "components.db"
SOURCE_DIR = Path(r"C:\Users\zjh\Desktop\被动产品线资料\信昌PDC\Resistor电阻")
RESISTOR_DIR = ROOT / "Resistor"

BRAND = "PSA(信昌电陶)"
TODAY = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
DATA_SOURCE_TAG = "PDC official PDF generated"


E24_BASE = [Decimal(x) for x in "10 11 12 13 15 16 18 20 22 24 27 30 33 36 39 43 47 51 56 62 68 75 82 91".split()]
E96_BASE = [
    Decimal(x)
    for x in (
        "100 102 105 107 110 113 115 118 121 124 127 130 133 137 140 143 147 150 154 158 162 165 "
        "169 174 178 182 187 191 196 200 205 210 215 221 226 232 237 243 249 255 261 267 274 280 "
        "287 294 301 309 316 324 332 340 348 357 365 374 383 392 402 412 422 432 442 453 464 475 "
        "487 499 511 523 536 549 562 576 590 604 619 634 649 665 681 698 715 732 750 768 787 806 "
        "825 845 866 887 909 931 953 976"
    ).split()
]

SIZE_DIMENSIONS = {
    "0A": ("01005", "0.40", "0.20", "0.13"),
    "01": ("0201", "0.60", "0.30", "0.23"),
    "02": ("0402", "1.00", "0.50", "0.35"),
    "03": ("0603", "1.60", "0.80", "0.45"),
    "05": ("0805", "2.00", "1.25", "0.50"),
    "06": ("1206", "3.10", "1.60", "0.55"),
    "12": ("1210", "3.10", "2.60", "0.55"),
    "18": ("1218", "3.05", "4.60", "0.55"),
    "20": ("2010", "5.00", "2.50", "0.60"),
    "25": ("2512", "6.40", "3.20", "0.60"),
}

PACK_BY_SIZE = {
    "0A": "S",
    "01": "T",
    "02": "V",
    "03": "T",
    "05": "T",
    "06": "T",
    "12": "T",
    "18": "P",
    "20": "P",
    "25": "P",
}

TOL_TEXT = {
    "B": "±0.10%",
    "C": "±0.25%",
    "D": "±0.50%",
    "F": "±1%",
    "G": "±2%",
    "J": "±5%",
    "K": "±10%",
    "L": "±15%",
    "M": "±20%",
}


@dataclass(frozen=True)
class RangeRule:
    tol_code: str
    min_ohm: Decimal
    max_ohm: Decimal
    e_series: str
    tcr: str


@dataclass(frozen=True)
class SizeRule:
    size_code: str
    inch: str
    power: str
    max_voltage: str
    overload_voltage: str
    pack: str = ""


@dataclass(frozen=True)
class SeriesRule:
    series: str
    source_pdf: str
    series_desc: str
    device_type: str
    special_usage: str
    sizes: tuple[SizeRule, ...]
    ranges_by_size: dict[str, tuple[RangeRule, ...]]
    special_suffixes: tuple[str, ...] = ("",)
    model_style: str = "standard"


def d(value: str | int | float) -> Decimal:
    return Decimal(str(value))


def clean_text(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"nan", "none"} else text


def decimal_text(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def resistance_values(min_ohm: Decimal, max_ohm: Decimal, e_series: str) -> list[Decimal]:
    bases = E96_BASE if e_series.upper() == "E96" else E24_BASE
    values: set[Decimal] = set()
    for exponent in range(-4, 9):
        scale = Decimal(10) ** exponent
        for base in bases:
            value = (base * scale).normalize()
            if min_ohm <= value <= max_ohm:
                values.add(value)
    if min_ohm == 0:
        values.add(Decimal(0))
    return sorted(values)


def mohm_values(min_mohm: int, max_mohm: int) -> list[Decimal]:
    values: set[Decimal] = set()
    for value in [10, 11, 12, 13, 15, 16, 18, 20, 22, 24, 27, 30, 33, 36, 39, 43, 47, 50, 51, 56, 62, 68, 75, 82, 91, 99]:
        if min_mohm <= value <= max_mohm:
            values.add(Decimal(value))
    for base in [100, 110, 120, 130, 150, 160, 180, 200, 220, 240, 270, 300, 330, 360, 390, 430, 470, 510, 560, 620, 680, 750, 820, 910]:
        if min_mohm <= base <= max_mohm:
            values.add(Decimal(base))
    return sorted(values)


def resistor_code(value: Decimal, digits: int) -> str:
    if value == 0:
        return "0" * digits
    if digits == 4 and value < 100:
        text = decimal_text(value)
        if "." in text:
            left, right = text.split(".", 1)
            return f"{left}R{right}".ljust(digits, "0")[:digits]
        return f"{text}R0".ljust(digits, "0")[:digits]
    if value < 10:
        places = 2 if digits == 4 else 1
        text = f"{value:.{places}f}".rstrip("0").rstrip(".")
        left, _, right = text.partition(".")
        return f"{left}R{right}".ljust(digits, "0")

    target_min = Decimal(100) if digits == 4 else Decimal(10)
    target_max = Decimal(1000) if digits == 4 else Decimal(100)
    coefficient = value
    exponent = 0
    while coefficient >= target_max:
        coefficient = coefficient / Decimal(10)
        exponent += 1
    while coefficient < target_min:
        coefficient = coefficient * Decimal(10)
        exponent -= 1
    coeff = int(coefficient.to_integral_value())
    if exponent < 0:
        return decimal_text(value).replace(".", "R").ljust(digits, "0")[:digits]
    return f"{coeff:0{digits - 1}d}{exponent}"


def low_ohm_code_mohm(mohm: Decimal, width: int = 3) -> str:
    if mohm == Decimal("0.5"):
        return "R0L5"
    if mohm == Decimal("0.75"):
        return "R0L75"
    if mohm == Decimal("2.5"):
        return "R2L5"
    return f"R{int(mohm):0{width}d}"


def tolerance_code_digits(tol_code: str, size_code: str = "") -> int:
    if tol_code in {"B", "C", "D", "F"}:
        return 4
    if size_code in {"18", "20", "25"} and tol_code == "J":
        return 4
    return 3


def source_url(pdf_name: str) -> str:
    return str(SOURCE_DIR / pdf_name)


def base_db_row() -> dict[str, str]:
    return {
        "品牌": BRAND,
        "安装方式": "贴片",
        "生产状态": "量产",
        "数据来源": DATA_SOURCE_TAG,
        "数据状态": "官方PDF命名规则生成",
        "校验时间": TODAY,
    }


def size_dimensions(size: SizeRule) -> tuple[str, str, str]:
    for code, dims in SIZE_DIMENSIONS.items():
        if dims[0] == size.inch:
            return dims[1], dims[2], dims[3]
    if size.inch == "0612":
        return "1.60", "3.20", "0.60"
    if size.inch == "1225":
        return "3.10", "6.30", "0.60"
    return "", "", ""


def standard_model(rule: SeriesRule, size: SizeRule, range_rule: RangeRule, value: Decimal, suffix: str) -> str:
    pack = size.pack or PACK_BY_SIZE.get(size.size_code, "T")
    rcode = resistor_code(value, tolerance_code_digits(range_rule.tol_code, size.size_code))
    if rule.model_style == "wcf":
        watt_code = "J" if size.size_code == "25" else "H"
        tcr_code = "N" if "100" in range_rule.tcr else "L"
        return f"WCF{size.size_code}{range_rule.tol_code}{pack}{watt_code}{rcode}{tcr_code}M"
    if rule.model_style == "fpf_l":
        rcode = low_ohm_code_mohm(value, width=3)
        return f"FPF{size.size_code}{range_rule.tol_code}{pack}-{rcode}{suffix}"
    if rule.model_style == "fbf":
        rcode = low_ohm_code_mohm(value, width=3)
        return f"FBF{size.size_code}{range_rule.tol_code}{pack}-{rcode}{suffix}"
    if rule.model_style == "fof":
        watt_code = {"02": "E", "03": "F", "05": "G", "06": "H", "25": "J"}[size.size_code]
        rcode = low_ohm_code_mohm(value, width=3)
        tcr_code = "N" if value < Decimal(10) else "P"
        return f"FOF{size.size_code}{range_rule.tol_code}{pack}{watt_code}{rcode}{tcr_code}SS"
    if rule.model_style == "faf":
        pack = "T" if size.size_code not in {"20", "25"} else "P"
        power_code = {
            "01": "T",
            "02": "A",
            "03": "B",
            "05": "C",
            "06": "C",
            "12": "D",
            "20": "F",
            "25": "G",
        }.get(size.size_code, "-")
        return f"FAF{size.size_code}{range_rule.tol_code}{pack}{power_code}{resistor_code(value, 4)}Q"
    return f"{rule.series}{size.size_code}{range_rule.tol_code}{pack}-{resistor_code(value, tolerance_code_digits(range_rule.tol_code, size.size_code))}{suffix}"


def db_row_for_resistor(rule: SeriesRule, size: SizeRule, range_rule: RangeRule, value: Decimal, model: str) -> dict[str, str]:
    length, width, height = size_dimensions(size)
    resistance_text = decimal_text(value)
    unit = "Ω"
    display_value = resistance_text
    if value >= Decimal("1000000"):
        unit = "MΩ"
        display_value = decimal_text(value / Decimal("1000000"))
    elif value >= Decimal("1000"):
        unit = "KΩ"
        display_value = decimal_text(value / Decimal("1000"))
    row = base_db_row()
    row.update(
        {
            "型号": model,
            "系列": rule.series,
            "系列说明": rule.series_desc,
            "器件类型": rule.device_type,
            "特殊用途": rule.special_usage,
            "封装代码": size.inch,
            "尺寸（inch）": size.inch,
            "尺寸（mm）": f"{length}×{width}" if length and width else "",
            "长度（mm）": length,
            "宽度（mm）": width,
            "高度（mm）": height,
            "阻值@25C": resistance_text,
            "阻值单位": "Ω",
            "阻值误差": TOL_TEXT.get(range_rule.tol_code, ""),
            "工作温度": "-55~155°C",
            "备注1": f"{display_value}{unit} {TOL_TEXT.get(range_rule.tol_code, '')} {size.power} TCR {range_rule.tcr}",
            "规格摘要": f"{rule.series} {size.inch} {display_value}{unit} {TOL_TEXT.get(range_rule.tol_code, '')} {size.power}",
            "官网链接": source_url(rule.source_pdf),
            "校验备注": f"{rule.source_pdf}: {size.inch} {size.power} {range_rule.min_ohm}Ω~{range_rule.max_ohm}Ω {range_rule.e_series}",
            "_model_rule_authority": f"PDC {rule.source_pdf}",
            "_resistance_ohm": resistance_text,
        }
    )
    if size.max_voltage:
        row["备注2"] = f"Max working {size.max_voltage}; overload {size.overload_voltage}".strip()
    return row


def generate_standard_rows(rule: SeriesRule) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for size in rule.sizes:
        for range_rule in rule.ranges_by_size.get(size.size_code, ()):
            values = mohm_values(int(range_rule.min_ohm), int(range_rule.max_ohm)) if rule.model_style in {"fbf", "fpf_l", "fof"} else resistance_values(range_rule.min_ohm, range_rule.max_ohm, range_rule.e_series)
            for value in values:
                ohm_value = value / Decimal(1000) if rule.model_style in {"fbf", "fpf_l", "fof"} else value
                for suffix in rule.special_suffixes:
                    model = standard_model(rule, size, range_rule, value, suffix)
                    rows.append(db_row_for_resistor(rule, size, range_rule, ohm_value, model))
    return rows


def common_sizes(series: str, specs: list[tuple[str, str, str, str, str]]) -> tuple[SizeRule, ...]:
    return tuple(SizeRule(code, inch, power, v, ov, PACK_BY_SIZE.get(code, "T")) for code, inch, power, v, ov in specs)


def build_rules() -> list[SeriesRule]:
    fpf_sizes = common_sizes("FPF", [
        ("02", "0402", "1/8W", "50V", "100V"),
        ("03", "0603", "1/8W", "50V", "100V"),
        ("05", "0805", "1/4W", "150V", "300V"),
        ("06", "1206", "1/2W", "200V", "400V"),
        ("12", "1210", "1/2W", "200V", "400V"),
        ("20", "2010", "1W", "200V", "400V"),
        ("25", "2512", "2W", "300V", "600V"),
    ])
    fps_sizes = tuple(SizeRule(s.size_code, s.inch, s.power, s.max_voltage, s.overload_voltage, s.pack) for s in fpf_sizes if s.size_code != "02")
    hv_sizes = common_sizes("FVF", [
        ("03", "0603", "1/10W", "200V", "400V"),
        ("05", "0805", "1/8W", "400V", "800V"),
        ("06", "1206", "1/4W", "800V", "1600V"),
        ("20", "2010", "1/2W", "2000V", "3000V"),
        ("25", "2512", "1W", "3000V", "4000V"),
    ])
    av_sizes = common_sizes("AVF", [
        ("06", "1206", "1/4W", "800V", "1600V"),
        ("20", "2010", "1/2W", "2000V", "3000V"),
        ("25", "2512", "1W", "3000V", "4000V"),
    ])

    rules: list[SeriesRule] = []
    rules.append(
        SeriesRule(
            "FCF",
            "FCF.pdf",
            "PDC FCF 通用厚膜贴片电阻",
            "厚膜电阻",
            "通用厚膜",
            common_sizes("FCF", [
                ("0A", "01005", "1/32W", "15V", "30V"),
                ("01", "0201", "1/20W", "25V", "50V"),
                ("02", "0402", "1/16W", "50V", "100V"),
                ("03", "0603", "1/10W", "75V", "150V"),
                ("05", "0805", "1/8W", "150V", "300V"),
                ("06", "1206", "1/4W", "200V", "400V"),
                ("12", "1210", "1/3W", "200V", "400V"),
                ("18", "1218", "1W", "200V", "400V"),
                ("20", "2010", "3/4W", "200V", "400V"),
                ("25", "2512", "1W", "250V", "500V"),
            ]),
            {},
        )
    )
    general_ranges = {
        "0A": (RangeRule("F", d("100"), d("1000000"), "E96", "±200ppm/°C"), RangeRule("J", d("10"), d("1000000"), "E24", "±300ppm/°C")),
        "01": (RangeRule("F", d("10"), d("10000000"), "E96", "±200ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
        "02": (RangeRule("F", d("1"), d("10000000"), "E96", "±200ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
        "03": (RangeRule("F", d("1"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
        "05": (RangeRule("F", d("1"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
        "06": (RangeRule("F", d("1"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
        "12": (RangeRule("F", d("1"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
        "18": (RangeRule("F", d("1"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
        "20": (RangeRule("F", d("1"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
        "25": (RangeRule("F", d("1"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±200ppm/°C")),
    }
    rules[-1].ranges_by_size.update(general_ranges)  # type: ignore[misc]

    rules.extend(
        [
            SeriesRule(
                "WCF",
                "WCF.pdf",
                "PDC WCF 宽端子高功率厚膜贴片电阻（AEC-Q200）",
                "厚膜电阻",
                "宽端子 | 高功率 | AEC-Q200",
                (SizeRule("06", "0612", "1W", "200V", "400V", "T"), SizeRule("25", "1225", "2W", "200V", "400V", "P")),
                {
                    "06": (RangeRule("F", d("1"), d("4.64"), "E96", "±200ppm/°C"), RangeRule("F", d("4.7"), d("1000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("4.3"), "E24", "±200ppm/°C"), RangeRule("J", d("4.7"), d("1000000"), "E24", "±100ppm/°C")),
                    "25": (RangeRule("F", d("1"), d("4.64"), "E96", "±200ppm/°C"), RangeRule("F", d("4.7"), d("1000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("4.3"), "E24", "±200ppm/°C"), RangeRule("J", d("4.7"), d("1000000"), "E24", "±100ppm/°C")),
                },
                model_style="wcf",
            ),
            SeriesRule(
                "FHF",
                "FHF.pdf",
                "PDC FHF 高阻值厚膜贴片电阻",
                "厚膜电阻",
                "高阻值 | 厚膜",
                common_sizes("FHF", [("02", "0402", "1/16W", "50V", "100V"), ("03", "0603", "1/10W", "50V", "100V"), ("05", "0805", "1/8W", "150V", "300V"), ("06", "1206", "1/4W", "200V", "400V")]),
                {
                    "02": (RangeRule("F", d("10000000"), d("30000000"), "E24", "±300ppm/°C"), RangeRule("J", d("10000000"), d("30000000"), "E24", "±300ppm/°C")),
                    "03": (RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("11000000"), d("100000000"), "E24", "±200ppm/°C")),
                    "05": (RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("11000000"), d("100000000"), "E24", "±200ppm/°C")),
                    "06": (RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("11000000"), d("100000000"), "E24", "±200ppm/°C")),
                },
            ),
            SeriesRule(
                "FNF",
                "FNF.pdf",
                "PDC FNF 抗浪涌厚膜贴片电阻（AEC-Q200）",
                "厚膜电阻",
                "抗浪涌 | AEC-Q200",
                common_sizes("FNF", [("03", "0603", "1/10W", "50V", "100V"), ("05", "0805", "1/8W", "150V", "300V"), ("06", "1206", "1/4W", "200V", "400V"), ("12", "1210", "1/3W", "200V", "400V"), ("20", "2010", "3/4W", "200V", "400V"), ("25", "2512", "1W", "200V", "400V")]),
                {code: tuple(RangeRule(t, d("1"), d("1000000"), "E24", "±100ppm/°C") for t in ("J", "K", "L", "M")) for code in ("03", "05", "06", "12", "20", "25")},
                ("", "-M"),
            ),
            SeriesRule(
                "FPF",
                "FPF.pdf",
                "PDC FPF 高功率厚膜贴片电阻（AEC-Q200）",
                "厚膜电阻",
                "高功率 | AEC-Q200",
                fpf_sizes,
                {s.size_code: (RangeRule("F", d("1"), d("1000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("1000000"), "E24", "±200ppm/°C")) for s in fpf_sizes},
                ("", "-M"),
            ),
            SeriesRule(
                "FPS",
                "FPS.pdf",
                "PDC FPS 高功率抗浪涌厚膜贴片电阻（AEC-Q200）",
                "厚膜电阻",
                "高功率 | 抗浪涌 | AEC-Q200",
                fps_sizes,
                {s.size_code: (RangeRule("F", d("1"), d("1000000"), "E96", "±100ppm/°C"), RangeRule("J", d("1"), d("1000000"), "E24", "±200ppm/°C")) for s in fps_sizes},
                ("", "-M", "-MB", "-MD"),
            ),
            SeriesRule(
                "FVF",
                "FVF.pdf",
                "PDC FVF 高压厚膜贴片电阻（AEC-Q200）",
                "厚膜电阻",
                "高压 | AEC-Q200",
                hv_sizes,
                {
                    "03": (RangeRule("F", d("47"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("100000"), d("10000000"), "E24", "±200ppm/°C")),
                    "05": (RangeRule("F", d("47"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("100000"), d("22000000"), "E24", "±200ppm/°C")),
                    "06": (RangeRule("F", d("47"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("47"), d("100000000"), "E24", "±200ppm/°C")),
                    "20": (RangeRule("F", d("47"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("47"), d("100000000"), "E24", "±200ppm/°C")),
                    "25": (RangeRule("F", d("47"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("47"), d("100000000"), "E24", "±200ppm/°C")),
                },
                ("", "-M"),
            ),
            SeriesRule(
                "FVS",
                "FVS.pdf",
                "PDC FVS 高压安规厚膜贴片电阻（AEC-Q200）",
                "厚膜电阻",
                "高压 | 安规 | AEC-Q200",
                hv_sizes,
                {
                    "03": (RangeRule("F", d("100000"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("100000"), d("10000000"), "E24", "±200ppm/°C")),
                    "05": (RangeRule("F", d("100000"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("J", d("100000"), d("22000000"), "E24", "±200ppm/°C")),
                    "06": (RangeRule("F", d("100000"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("100000"), d("100000000"), "E24", "±200ppm/°C")),
                    "20": (RangeRule("F", d("100000"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("100000"), d("100000000"), "E24", "±200ppm/°C")),
                    "25": (RangeRule("F", d("100000"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("100000"), d("100000000"), "E24", "±200ppm/°C")),
                },
                ("", "-M"),
            ),
            SeriesRule("AVS", "AVS.pdf", "PDC AVS 高压安规抗硫化厚膜贴片电阻（AEC-Q200）", "厚膜电阻", "高压 | 安规 | 抗硫化 | AEC-Q200", av_sizes, {s.size_code: (RangeRule("F", d("100000"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("100000"), d("100000000"), "E24", "±200ppm/°C")) for s in av_sizes}, ("-M",)),
            SeriesRule("AVF", "AVF.pdf", "PDC AVF 高压通用抗硫化厚膜贴片电阻（AEC-Q200）", "厚膜电阻", "高压 | 抗硫化 | AEC-Q200", av_sizes, {s.size_code: (RangeRule("F", d("100000"), d("10000000"), "E96", "±100ppm/°C"), RangeRule("F", d("11000000"), d("22000000"), "E24", "±200ppm/°C"), RangeRule("J", d("100000"), d("100000000"), "E24", "±200ppm/°C")) for s in av_sizes}, ("-M",)),
        ]
    )

    fwf_sizes = common_sizes("FWF", [("02", "0402", "1/16W", "50V", "100V"), ("03", "0603", "1/10W", "75V", "150V"), ("05", "0805", "1/8W", "150V", "300V"), ("06", "1206", "1/4W", "200V", "400V"), ("12", "1210", "1/2W", "200V", "400V"), ("20", "2010", "3/4W", "200V", "400V"), ("25", "2512", "1W", "250V", "500V")])
    rules.append(SeriesRule("FWF", "FWF.pdf", "PDC FWF 车规通用厚膜贴片电阻（AEC-Q200）", "厚膜电阻", "车规 | 厚膜 | AEC-Q200", fwf_sizes, {s.size_code: (RangeRule("F", d("1"), d("10000000"), "E96", "±100/±200ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±100/±200ppm/°C")) for s in fwf_sizes}, ("", "-W")))
    fwfs_sizes = common_sizes("FWF", [("01", "0201", "1/20W", "25V", "50V"), ("02", "0402", "1/10W", "50V", "100V"), ("03", "0603", "1/8W", "75V", "150V"), ("05", "0805", "1/4W", "150V", "300V"), ("06", "1206", "1/4W", "200V", "400V"), ("12", "1210", "1/2W", "200V", "400V"), ("20", "2010", "3/4W", "200V", "400V"), ("25", "2512", "1W", "250V", "500V")])
    rules.append(SeriesRule("FWF", "FWF-S.pdf", "PDC FWF-S 车规抗硫化厚膜贴片电阻（AEC-Q200）", "厚膜电阻", "车规 | 抗硫化 | AEC-Q200", fwfs_sizes, {s.size_code: (RangeRule("F", d("1"), d("10000000"), "E96", "±100/±200ppm/°C"), RangeRule("J", d("1"), d("10000000"), "E24", "±100/±200ppm/°C")) for s in fwfs_sizes}, ("-S",)))

    low_sizes = common_sizes("FBF", [("03", "0603", "1/8W", "", ""), ("05", "0805", "1/4W", "", ""), ("06", "1206", "1/3W", "", ""), ("12", "1210", "2/3W", "", ""), ("20", "2010", "3/4W", "", ""), ("25", "2512", "1W", "", "")])
    rules.append(SeriesRule("FBF", "FBF.pdf", "PDC FBF 低阻厚膜电流检测贴片电阻", "厚膜电阻", "电流检测 | 低阻 | 厚膜", low_sizes, {s.size_code: (RangeRule("F", d("10" if s.size_code != "03" else "40"), d("910"), "E24", "±100/±200ppm/°C"), RangeRule("J", d("10" if s.size_code != "03" else "40"), d("910"), "E24", "±100/±200ppm/°C")) for s in low_sizes}, ("",), "fbf"))
    rules.append(SeriesRule("FBF", "FBF_AECQ.pdf", "PDC FBF-M 车规抗硫化低阻厚膜电流检测电阻", "厚膜电阻", "电流检测 | 低阻 | 抗硫化 | AEC-Q200", low_sizes, {s.size_code: (RangeRule("F", d("50"), d("910"), "E24", "±100/±200ppm/°C"), RangeRule("J", d("50"), d("910"), "E24", "±100/±200ppm/°C")) for s in low_sizes}, ("-M", "-ME"), "fbf"))

    rules.append(SeriesRule("FPF", "FPF-L.pdf", "PDC FPF-L 高功率低阻厚膜电流检测电阻（AEC-Q200）", "厚膜电阻", "电流检测 | 低阻 | 高功率 | AEC-Q200", tuple(s for s in fpf_sizes if s.size_code != "02"), {s.size_code: (RangeRule("F", d("50"), d("910"), "E24", "±100/±250ppm/°C"), RangeRule("J", d("50"), d("910"), "E24", "±100/±250ppm/°C")) for s in fpf_sizes if s.size_code != "02"}, ("", "-M"), "fpf_l"))
    fof_sizes = common_sizes("FOF", [("02", "0402", "1/3W", "", ""), ("03", "0603", "1/2W", "", ""), ("05", "0805", "3/4W", "", ""), ("06", "1206", "1W", "", ""), ("25", "2512", "2W", "", "")])
    fof_ranges = {"02": (RangeRule("F", d("5"), d("25"), "E24", "±100ppm/°C"),), "03": (RangeRule("F", d("5"), d("75"), "E24", "±50/±100ppm/°C"),), "05": (RangeRule("F", d("3"), d("500"), "E24", "±50/±100ppm/°C"),), "06": (RangeRule("F", d("3"), d("700"), "E24", "±50/±100ppm/°C"),), "25": (RangeRule("F", d("2"), d("700"), "E24", "±50/±100ppm/°C"),)}
    rules.append(SeriesRule("FOF", "FOF.pdf", "PDC FOF 金属箔高功率抗硫化电流检测电阻", "合金电阻", "电流检测 | 金属箔 | 抗硫化", fof_sizes, fof_ranges, ("",), "fof"))

    faf_sizes = common_sizes("FAF", [("01", "0201", "1/32W", "15V", "30V"), ("02", "0402", "1/16W", "50V", "100V"), ("03", "0603", "1/16W", "50V", "100V"), ("05", "0805", "1/10W", "100V", "200V"), ("06", "1206", "1/8W", "200V", "400V"), ("12", "1210", "1/4W", "200V", "400V"), ("20", "2010", "1/2W", "200V", "400V"), ("25", "2512", "3/4W", "200V", "400V")])
    faf_ranges = {
        "01": (RangeRule("F", d("100"), d("12000"), "E96", "±25ppm/°C"),),
        "02": (RangeRule("F", d("10"), d("255000"), "E96", "±25ppm/°C"),),
        "03": (RangeRule("F", d("3.9"), d("1000000"), "E96", "±25ppm/°C"),),
        "05": (RangeRule("F", d("4.7"), d("2000000"), "E96", "±25ppm/°C"),),
        "06": (RangeRule("F", d("1"), d("2490000"), "E96", "±25ppm/°C"),),
        "12": (RangeRule("F", d("4.7"), d("2490000"), "E96", "±25ppm/°C"),),
        "20": (RangeRule("F", d("4.7"), d("3000000"), "E96", "±25ppm/°C"),),
        "25": (RangeRule("F", d("1"), d("3000000"), "E96", "±25ppm/°C"),),
    }
    rules.append(SeriesRule("FAF", "FAF.pdf", "PDC FAF 高精密薄膜贴片电阻", "薄膜电阻", "高精密 | 薄膜", faf_sizes, faf_ranges, ("",), "faf"))
    return rules


def generated_rows() -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for rule in build_rules():
        rows.extend(generate_standard_rows(rule))
    frame = pd.DataFrame(rows)
    return frame.drop_duplicates(subset=["品牌", "型号"], keep="first").reset_index(drop=True)


def db_columns(conn: sqlite3.Connection) -> list[str]:
    return [row[1] for row in conn.execute('PRAGMA table_info("components")').fetchall()]


def backup_database() -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = DB_PATH.with_name(f"{DB_PATH.name}.pdc_pdf_resistors_{timestamp}.bak")
    shutil.copy2(DB_PATH, backup_path)
    return backup_path


def merge_to_db_columns(conn: sqlite3.Connection, frame: pd.DataFrame) -> pd.DataFrame:
    columns = db_columns(conn)
    rows = []
    for row in frame.to_dict("records"):
        merged = {column: "" for column in columns}
        for column, value in row.items():
            if column in merged:
                merged[column] = clean_text(value)
        rows.append(merged)
    return pd.DataFrame(rows, columns=columns)


def apply_to_database(frame: pd.DataFrame, dry_run: bool, no_backup: bool) -> Path | None:
    if dry_run:
        return None
    backup_path = None if no_backup else backup_database()
    with sqlite3.connect(DB_PATH, timeout=120) as conn:
        conn.execute("PRAGMA busy_timeout = 120000")
        db_frame = merge_to_db_columns(conn, frame)
        pairs = db_frame[["品牌", "型号"]].drop_duplicates().itertuples(index=False)
        for brand, model in pairs:
            conn.execute("DELETE FROM components WHERE 品牌 = ? AND 型号 = ?", (brand, model))
        db_frame.to_sql("components", conn, if_exists="append", index=False, chunksize=10, method="multi")
        conn.commit()
    return backup_path


def workbook_row(row: dict[str, str], columns: list[str]) -> list[str]:
    mapped = dict(row)
    mapped["阻值"] = clean_text(row.get("阻值@25C", ""))
    mapped["功率"] = clean_text(row.get("备注1", "")).split(" TCR ")[0].split()[-1] if clean_text(row.get("备注1", "")) else ""
    mapped["TCR（ppm）"] = clean_text(row.get("备注1", "")).split("TCR ", 1)[-1] if "TCR " in clean_text(row.get("备注1", "")) else ""
    mapped["最高工作电压"] = ""
    mapped["最大过载电压"] = ""
    note2 = clean_text(row.get("备注2", ""))
    if "Max working" in note2:
        parts = note2.replace("Max working", "").replace("overload", "").replace(";", "").split()
        if parts:
            mapped["最高工作电压"] = parts[0]
        if len(parts) > 1:
            mapped["最大过载电压"] = parts[1]
    return [mapped.get(column, "") for column in columns]


def update_workbook(path: Path, frame: pd.DataFrame, dry_run: bool) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    try:
        ws = wb[wb.sheetnames[0]]
        columns = [cell.value for cell in ws[1]]
        source_idx = columns.index("数据来源") + 1 if "数据来源" in columns else None
        rows_to_delete: list[int] = []
        if source_idx:
            for row_idx in range(2, ws.max_row + 1):
                if clean_text(ws.cell(row_idx, source_idx).value) == DATA_SOURCE_TAG:
                    rows_to_delete.append(row_idx)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx, 1)
        count = 0
        for row in frame.to_dict("records"):
            ws.append(workbook_row(row, columns))
            count += 1
        if not dry_run:
            wb.save(path)
        return count
    finally:
        wb.close()


def update_source_workbooks(frame: pd.DataFrame, dry_run: bool) -> dict[str, int]:
    targets = {
        "厚膜电阻": RESISTOR_DIR / "厚膜电阻.xlsx",
        "薄膜电阻": RESISTOR_DIR / "薄膜电阻.xlsx",
        "合金电阻": RESISTOR_DIR / "合金电阻.xlsx",
    }
    counts: dict[str, int] = {}
    for device_type, path in targets.items():
        subset = frame[frame["器件类型"] == device_type].copy()
        if subset.empty or not path.exists():
            continue
        counts[path.name] = update_workbook(path, subset, dry_run=dry_run)
    return counts


def refresh_cache(frame: pd.DataFrame, skip_cache: bool) -> dict[str, int]:
    if skip_cache:
        return {}
    with sqlite3.connect(DB_PATH, timeout=120) as conn:
        models = frame["型号"].drop_duplicates().tolist()
        parts = []
        for idx in range(0, len(models), 800):
            chunk = models[idx : idx + 800]
            placeholders = ",".join("?" for _ in chunk)
            parts.append(pd.read_sql_query(f"SELECT * FROM components WHERE 型号 IN ({placeholders})", conn, params=chunk))
    selected = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
    prepared = cm.prepare_search_dataframe(cm.deduplicate_component_rows(selected))
    removed, inserted = stream_replace_prepared_rows(prepared)
    sidecar_counts = refresh_search_sidecar_rows(prepared)
    return {
        "selected_db_rows": int(len(selected)),
        "prepared_removed": int(removed),
        "prepared_inserted": int(inserted),
        "search_core_rows": int(sidecar_counts.get(cm.COMPONENTS_SEARCH_CORE_TABLE, 0)),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync PDC resistor model rows from official PDF rating tables.")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-db", action="store_true")
    parser.add_argument("--skip-workbooks", action="store_true")
    parser.add_argument("--skip-cache", action="store_true")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()

    frame = generated_rows()
    print(f"generated_rows={len(frame)}")
    print(frame.groupby(["系列", "器件类型"]).size().sort_values(ascending=False).head(40).to_string())

    backup_path = None
    if not args.skip_db:
        backup_path = apply_to_database(frame, dry_run=args.dry_run, no_backup=args.no_backup)
        if backup_path:
            print(f"backup_path={backup_path}")

    if not args.skip_workbooks:
        for name, count in update_source_workbooks(frame, dry_run=args.dry_run).items():
            print(f"workbook_rows[{name}]={count}")

    if not args.dry_run and not args.skip_db:
        for key, value in refresh_cache(frame, skip_cache=args.skip_cache).items():
            print(f"{key}={value}")


if __name__ == "__main__":
    main()
