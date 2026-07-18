from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import mock

import pandas as pd
from openpyxl import Workbook

import component_matcher as cm
import sync_epson_parametric_products as epson_sync


class EpsonTimingIntegrationTests(unittest.TestCase):
    def test_decimal_text_preserves_integer_trailing_zeroes(self):
        self.assertEqual(epson_sync.decimal_text("50"), "50")
        self.assertEqual(epson_sync.decimal_text("100"), "100")
        self.assertEqual(epson_sync.decimal_text("32.768000"), "32.768")

    def test_build_crystal_product_row_uses_official_product_number(self):
        product = {
            "pn": "Q13FC13500002",
            "model": "FC-135",
            "dimensionL": "3.2",
            "dimensionW": "1.5",
            "dimensionH": "0.9",
            "frequencyMin": "32.768000",
            "loadCapPf": "7",
            "frequencyTol25C": "+/-20",
            "operatingTempRange": "-40 to +85",
            "frequencyAging": "3",
            "turnoverTemp": "+25&deg;C +/-5&deg;C",
            "parabolicCoef": "-0.04",
            "esrMax": "70",
            "driveLevel": "0.5",
            "pdf_switch": "1",
            "aecq200": "No",
        }
        row = epson_sync.build_product_row(
            product,
            "xtal_32khz.json",
            epson_sync.SOURCE_SPECS["xtal_32khz.json"],
            "2026-07-16 12:00:00",
        )

        self.assertEqual(row["品牌"], "爱普生Epson")
        self.assertEqual(row["型号"], "Q13FC13500002")
        self.assertEqual(row["系列"], "FC-135")
        self.assertEqual(row["器件类型"], "晶振")
        self.assertEqual(row["尺寸（inch）"], "3215")
        self.assertEqual(row["容值"], "32.768")
        self.assertEqual(row["容值单位"], "kHz")
        self.assertEqual(row["容值误差"], "20")
        self.assertEqual(row["负载电容（pF）"], "7")
        self.assertEqual(row["25℃老化（ppm）"], "±3ppm")
        self.assertEqual(row["拐点温度"], "+25℃ ±5℃")
        self.assertEqual(row["抛物线系数（ppm/℃²）"], "-0.04ppm/℃²")
        self.assertEqual(row["型号粒度"], "官方逐料号")

    def test_build_mhz_crystal_row_maps_temperature_aging_and_overtone(self):
        product = {
            "pn": "Q22FA12800007",
            "model": "FA-128",
            "dimensionL": "2.0",
            "dimensionW": "1.6",
            "dimensionH": "0.5",
            "frequencyMin": "38.400000",
            "loadCap": "10",
            "frequencyTol25C": "+/-15",
            "frequencyTolTempRange": "+/-15",
            "25CAging": "+/-1ppm",
            "overtoneOrder": "Fundamental",
            "tempRange": "-20 to +85",
        }
        row = epson_sync.build_product_row(
            product,
            "xtal_mhz.json",
            epson_sync.SOURCE_SPECS["xtal_mhz.json"],
            "2026-07-17 00:00:00",
        )

        self.assertEqual(row["频率温度特性（ppm）"], "±15ppm")
        self.assertEqual(row["25℃老化（ppm）"], "±1ppm")
        self.assertEqual(row["泛音阶次"], "基频（Fundamental）")
        self.assertEqual(row["工作温度"], "-20~+85°C")

    def test_build_rtc_product_row_maps_official_fields(self):
        product = {
            "pn": "Q418025510002",
            "model": "RX-8025SA",
            "description": "Real time clock module",
            "dimensionL": "10.1",
            "dimensionW": "5.0",
            "dimensionH": "1.3",
            "interface": "I2C-Bus",
            "frequncyTol": "+/-5",
            "operTemperature": "-40 to +85",
            "backUpCurrent": "0.48",
            "clockVoltage": "1.15 to 5.5",
            "operVoltage": "1.7 to 5.5",
            "frequencyOutput": "32768",
            "pkgType": "SOP 14-pin",
            "monthlyDeviation_sec": "13",
        }
        row = epson_sync.build_product_row(
            product,
            "rtc.json",
            epson_sync.SOURCE_SPECS["rtc.json"],
            "2026-07-17 12:00:00",
        )

        self.assertEqual(row["型号"], "Q418025510002")
        self.assertEqual(row["系列"], "RX-8025SA")
        self.assertEqual(row["器件类型"], "实时时钟模块")
        self.assertEqual(row["容值"], "32.768")
        self.assertEqual(row["容值单位"], "kHz")
        self.assertEqual(row["容值误差"], "5")
        self.assertEqual(row["工作温度"], "-40~+85°C")
        self.assertEqual(row["接口类型"], "I2C-Bus")
        self.assertEqual(row["计时电压（V）"], "1.15~5.5")
        self.assertEqual(row["备用电流（µA）"], "0.48")
        self.assertEqual(row["月偏差（s）"], "13")
        self.assertEqual(row["封装代码"], "SOP 14-pin")

    def test_rx8025t_uc_channel_model_is_an_exact_rtc_row(self):
        rows = epson_sync.build_rx8025t_variant_rows("2026-07-17 12:00:00")
        uc = next(row for row in rows if row["型号"] == "RX8025T-UC")

        self.assertEqual(uc["器件类型"], "实时时钟模块")
        self.assertEqual(uc["容值"], "32.768")
        self.assertEqual(uc["容值单位"], "kHz")
        self.assertEqual(uc["容值误差"], "5")
        self.assertEqual(uc["工作温度"], "-30~+70°C")
        self.assertEqual(uc["接口类型"], "I²C")
        self.assertEqual(uc["计时电压（V）"], "1.8~5.5")
        self.assertEqual(uc["型号粒度"], "渠道确认精确型号")
        self.assertIn("全球参数选型未列出", uc["生产状态"])

        prepared = cm.prepare_search_dataframe(
            cm.normalize_imported_component_dataframe(pd.DataFrame(rows))
        )
        mode, spec = cm.detect_query_mode_and_spec(prepared, "RX8025T-UC")

        self.assertEqual(mode, "料号")
        self.assertIsNotNone(spec)
        self.assertEqual(spec["型号"], "RX8025T-UC")
        self.assertEqual(spec["器件类型"], "实时时钟模块")
        self.assertEqual(spec["接口类型"], "I²C")
        self.assertEqual(spec["计时电压（V）"], "1.8~5.5")
        self.assertEqual(spec["备用电流（µA）"], "0.8 Typ.")

    def test_rtc_series_alias_requires_full_product_number_confirmation(self):
        product = {
            "pn": "Q418025510002",
            "model": "RX-8025SA",
            "interface": "I2C-Bus",
            "frequncyTol": "+/-5",
            "operTemperature": "-40 to +85",
            "clockVoltage": "1.15 to 5.5",
            "operVoltage": "1.7 to 5.5",
            "pkgType": "SOP 14-pin",
        }
        official_row = epson_sync.build_product_row(
            product,
            "rtc.json",
            epson_sync.SOURCE_SPECS["rtc.json"],
            "2026-07-17 12:00:00",
        )
        aliases = epson_sync.build_rtc_series_alias_rows([official_row])

        self.assertEqual(len(aliases), 1)
        self.assertEqual(aliases[0]["型号"], "RX-8025SA")
        self.assertEqual(aliases[0]["官方规格编号"], "")
        self.assertIn("具体PN需确认", aliases[0]["型号粒度"])

    def test_tsx3225_legacy_and_packaging_aliases_are_searchable_rows(self):
        rows = epson_sync.build_tsx3225_legacy_alias_rows(
            "2026-07-17 22:00:00",
            existing_rows=[],
        )

        self.assertEqual(
            {row["型号"] for row in rows},
            {"X1E0000210139", "X1E000021013900"},
        )
        for row in rows:
            self.assertEqual(row["系列"], "TSX-3225")
            self.assertEqual(row["器件类型"], "晶振")
            self.assertEqual(row["尺寸（inch）"], "3225")
            self.assertEqual(row["容值"], "25")
            self.assertEqual(row["容值误差"], "10")
            self.assertEqual(row["负载电容（pF）"], "12")
            self.assertEqual(row["工作温度"], "-40~+85°C")
            self.assertIn("历史/包装料号", row["数据状态"])

    def test_compound_epson_series_and_product_number_exposes_exact_token(self):
        tokens = cm.extract_model_like_tokens("SG2520HGN_X1G0058910005")

        self.assertEqual(tokens[0], "SG2520HGN_X1G0058910005")
        self.assertIn("SG2520HGN", tokens)
        self.assertIn("X1G0058910005", tokens)

    def test_compound_epson_query_prefers_official_product_number_over_series(self):
        exact_row = cm.prepare_search_dataframe(cm.normalize_imported_component_dataframe(pd.DataFrame(
            [
                {
                    "品牌": "爱普生Epson",
                    "型号": "X1G0058910005",
                    "系列": "SG2520HGN",
                    "器件类型": "振荡器",
                    "尺寸（inch）": "2520",
                    "容值": "100",
                    "容值单位": "MHz",
                    "容值误差": "25",
                    "输出类型": "HCSL",
                }
            ]
        )))
        with mock.patch.object(
            cm,
            "load_component_rows_by_query_model_tokens",
            return_value=(exact_row, ["SG2520HGN", "X1G0058910005"], "X1G0058910005"),
        ):
            mode, spec = cm.detect_query_mode_and_spec(
                pd.DataFrame(),
                "SG2520HGN_X1G0058910005",
            )

        self.assertEqual(mode, "料号")
        self.assertEqual(spec["型号"], "X1G0058910005")
        self.assertEqual(spec["容值"], "100")
        self.assertEqual(spec["输出类型"], "HCSL")

    def test_structured_bom_text_with_package_underscore_remains_a_spec_query(self):
        query = "1u,10V Capacitor C_0402 TCC0402X5R105K100AT"

        mode, spec = cm.detect_query_mode_and_spec(pd.DataFrame(), query)

        self.assertEqual(mode, "规格")
        self.assertEqual(spec["尺寸（inch）"], "0402")
        self.assertEqual(spec["容值_pf"], 1_000_000.0)
        self.assertEqual(cm.clean_voltage(spec["耐压（V）"]), "10")

    def test_fc2012an_series_alias_keeps_official_common_details(self):
        product_row = {
            "品牌": "爱普生Epson",
            "型号": "X1A0001710001",
            "系列": "FC2012AN",
            "器件类型": "晶振",
            "尺寸（inch）": "2012",
            "容值": "32.768",
            "容值单位": "kHz",
            "容值误差": "20",
            "负载电容（pF）": "12.5",
            "工作温度": "-40~+105°C",
            "ESR": "60kΩ Max",
            "驱动电平": "0.5µW Max",
            "25℃老化（ppm）": "±3ppm",
            "拐点温度": "+25℃ ±5℃",
            "抛物线系数（ppm/℃²）": "-0.04ppm/℃²",
            "泛音阶次": "基频（Fundamental）",
        }
        aliases = epson_sync.build_fc2012an_series_alias_rows(
            "2026-07-17 23:00:00",
            [product_row],
        )

        self.assertEqual(len(aliases), 1)
        alias = aliases[0]
        self.assertEqual(alias["型号"], "FC2012AN")
        self.assertEqual(alias["负载电容（pF）"], "")
        self.assertEqual(alias["ESR"], "60kΩ Max")
        self.assertEqual(alias["25℃老化（ppm）"], "±3ppm")
        self.assertEqual(alias["泛音阶次"], "基频（Fundamental）")
        self.assertIn("具体PN需确认", alias["型号粒度"])

    def test_exact_timing_database_row_precedes_generic_model_rule(self):
        prepared = cm.prepare_search_dataframe(
            cm.normalize_imported_component_dataframe(
                pd.DataFrame(
                    [
                        {
                            "品牌": "爱普生Epson",
                            "型号": "FC2012AN",
                            "系列": "FC2012AN",
                            "器件类型": "晶振",
                            "尺寸（inch）": "2012",
                            "容值": "32.768",
                            "容值单位": "kHz",
                            "容值误差": "20",
                            "工作温度": "-40~+105°C",
                            "ESR": "60kΩ Max",
                            "驱动电平": "0.5µW Max",
                            "泛音阶次": "基频（Fundamental）",
                        }
                    ]
                )
            )
        )

        mode, spec = cm.detect_query_mode_and_spec(prepared, "FC2012AN")

        self.assertEqual(mode, "料号")
        self.assertEqual(spec["ESR"], "60kΩ Max")
        self.assertEqual(spec["驱动电平"], "0.5µW Max")
        self.assertEqual(spec["泛音阶次"], "基频（Fundamental）")

    def test_partial_crystal_match_lists_missing_and_different_parameters(self):
        spec = {
            "型号": "FC2012AN",
            "器件类型": "晶振",
            "尺寸（inch）": "2012",
            "容值": "32.768",
            "容值单位": "kHz",
            "容值误差": "20",
            "工作温度": "-40~+105°C",
            "ESR": "60kΩ Max",
            "驱动电平": "0.5µW Max",
            "25℃老化（ppm）": "±3ppm",
            "拐点温度": "+25℃ ±5℃",
            "抛物线系数（ppm/℃²）": "-0.04ppm/℃²",
            "泛音阶次": "基频（Fundamental）",
        }
        candidate = pd.Series(
            {
                "推荐等级": "部分参数匹配",
                "品牌": "NDK",
                "型号": "NX2012SA",
                "尺寸（inch）": "2012",
                "容值": "32.768",
                "容值单位": "kHz",
                "容值误差": "20",
                "工作温度": "-40~+105°C",
                "负载电容（pF）": "12.5",
                "ESR": "80kΩ Max",
            }
        )

        detail = cm.build_match_confirmation_detail(candidate, spec)

        self.assertIn("原型号资料缺少：负载电容(CL)", detail)
        self.assertIn("候选资料缺少", detail)
        self.assertIn("候选ESR 80kΩ Max高于原型号60kΩ Max", detail)
        self.assertIn("振荡电路负阻裕量", detail)

    def test_partial_hcsl_oscillator_match_calls_out_jitter_and_pin_checks(self):
        spec = {
            "型号": "X1G0058910005",
            "器件类型": "振荡器",
            "尺寸（inch）": "2520",
            "容值": "100",
            "容值单位": "MHz",
            "容值误差": "25",
            "耐压（V）": "3.135~3.465",
            "工作温度": "-40~+85°C",
            "输出类型": "HCSL",
            "占空比": "45~55%",
            "25℃老化（ppm）": "包含在总频差内（10年）",
        }
        candidate = pd.Series(
            {
                "推荐等级": "部分参数匹配",
                "品牌": "Abracon",
                "型号": "AK2AAIGHDF1-100.0000T",
                "尺寸（inch）": "2520",
                "容值": "100",
                "容值单位": "MHz",
                "容值误差": "25",
                "电源电压": "3.3",
                "工作温度": "-40~+85°C",
                "输出类型": "HCSL",
                "占空比": "45~55%",
            }
        )

        detail = cm.build_match_confirmation_detail(candidate, spec)

        self.assertIn("候选资料缺少：老化、相位噪声/抖动", detail)
        self.assertIn("输出摆幅与终端负载", detail)
        self.assertIn("OE/ST功能及脚位", detail)
        self.assertIn("相位抖动/相位噪声", detail)

    def test_bom_export_keeps_candidate_notes_with_their_brand_slots(self):
        upload_df = pd.DataFrame(
            {
                "原型号": ["FC2012AN", "RX8025T-UC"],
                "备注1": ["客户原备注", ""],
            }
        )
        result_df = pd.DataFrame(
            {
                "状态": ["需确认", "无匹配"],
                "推荐理由": ["存在部分参数匹配", "暂无跨品牌候选"],
                "备注1": [
                    "旧版首选候选备注",
                    "需向客户/工程确认：封装与引脚定义、I²C地址与寄存器兼容性",
                ],
                "自有品牌": ["NDK", ""],
                "自有型号": ["NX2012SA", ""],
                "自有匹配说明": ["只匹配了已识别参数，缺失参数需人工确认", ""],
                "自有匹配备注": [
                    "候选原备注；原型号资料缺少：负载电容(CL)；需向客户/工程确认：振荡电路负阻裕量",
                    "",
                ],
            }
        )

        export_df = cm.build_bom_matched_export_df(upload_df, result_df)

        self.assertNotIn("待确认参数", export_df.columns)
        self.assertEqual(export_df.loc[0, "备注1"], "客户原备注")
        self.assertEqual(
            export_df.loc[0, "匹配备注"],
            "候选原备注；原型号资料缺少：负载电容(CL)；需向客户/工程确认：振荡电路负阻裕量",
        )
        self.assertEqual(export_df.loc[1, "备注1"], "")
        self.assertEqual(export_df.loc[1, "匹配说明"], "暂无跨品牌候选")
        self.assertEqual(export_df.loc[1, "匹配备注"], result_df.loc[1, "备注1"])

    def test_search_results_append_confirmation_to_existing_remark1_once(self):
        spec = {
            "型号": "FC2012AN",
            "器件类型": "晶体单元",
            "尺寸（inch）": "2012",
            "容值": "32.768",
            "容值单位": "kHz",
            "容值误差": "20",
            "负载电容（pF）": "12.5",
        }
        frame = pd.DataFrame(
            [
                {
                    "推荐等级": "部分参数匹配",
                    "品牌": "NDK",
                    "型号": "NX2012SA",
                    "备注1": "候选原备注",
                }
            ]
        )

        first = cm.add_match_confirmation_to_remark1(frame, spec)
        second = cm.add_match_confirmation_to_remark1(first, spec)

        self.assertNotIn("待确认参数", first.columns)
        self.assertTrue(first.loc[0, "备注1"].startswith("候选原备注；"))
        self.assertEqual(second.loc[0, "备注1"], first.loc[0, "备注1"])

    def test_workbook_export_reuses_existing_remark1_column(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["原型号", "备注1"])
        worksheet.append(["FC2012AN", "客户原备注"])
        source_df = pd.DataFrame({"原型号": ["FC2012AN"], "备注1": ["客户原备注"]})
        append_columns = [
            {"header": "匹配状态", "values": ["需确认"]},
            {"header": "备注1", "values": ["需向客户确认负载电容(CL)"]},
        ]

        cm.append_export_columns_to_worksheet(worksheet, source_df, append_columns)

        headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]
        self.assertEqual(headers.count("备注1"), 1)
        self.assertEqual(worksheet["B2"].value, "客户原备注；需向客户确认负载电容(CL)")
        self.assertEqual(worksheet["C1"].value, "匹配状态")

    def test_rtc_query_completeness_requires_interface_and_backup_details(self):
        incomplete = {
            "器件类型": "实时时钟模块",
            "封装代码": "SOP-14-208mil",
            "接口类型": "I²C",
            "工作温度": "-40~+85°C",
            "容值误差": "5",
        }
        complete = {
            **incomplete,
            "耐压（V）": "2.2~5.5",
            "计时电压（V）": "1.8~5.5",
            "备用电流（µA）": "0.8 Typ.",
        }

        self.assertFalse(cm.timing_query_has_complete_details(incomplete, "实时时钟模块"))
        self.assertTrue(cm.timing_query_has_complete_details(complete, "实时时钟模块"))

    def test_runtime_cache_preparation_keeps_multibrand_timing_rows(self):
        epson_frame = pd.DataFrame(
            [
                {
                    "品牌": "爱普生Epson",
                    "型号": "Q22FA12800007",
                    "系列": "FA-128",
                    "器件类型": "晶振",
                    "尺寸（inch）": "2016",
                    "容值": "38.4",
                    "容值单位": "MHz",
                }
            ]
        )
        companion_frame = pd.DataFrame(
            [
                {
                    "品牌": "Abracon",
                    "型号": "ABM11N-40.0000MHZ-8-D2X-T3",
                    "系列": "ABM11N",
                    "器件类型": "晶振",
                    "尺寸（inch）": "2016",
                    "容值": "40",
                    "容值单位": "MHz",
                }
            ]
        )
        with TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            epson_path = temp_root / "epson.csv"
            companion_path = temp_root / "multibrand.csv"
            companion_frame.to_csv(companion_path, index=False, encoding="utf-8-sig")
            with mock.patch.object(
                epson_sync,
                "MULTIBRAND_TIMING_SOURCE",
                companion_path,
            ):
                normalized, prepared, companion_rows = (
                    epson_sync.prepare_runtime_cache_frame(
                        epson_frame,
                        epson_path,
                    )
                )

        self.assertEqual(len(normalized), 1)
        self.assertEqual(companion_rows, 1)
        self.assertEqual(
            set(prepared["型号"].tolist()),
            {
                "Q22FA12800007",
                "ABM11N-40.0000MHZ-8-D2X-T3",
            },
        )

    def test_timing_query_parses_metric_package_and_ppm(self):
        spec = cm.parse_timing_spec_query("晶振 16MHz 3.2x2.5mm 10pF ±10ppm")

        self.assertIsNotNone(spec)
        self.assertEqual(spec["器件类型"], "晶振")
        self.assertEqual(spec["尺寸（inch）"], "3225")
        self.assertEqual(spec["容值"], "16")
        self.assertEqual(spec["容值单位"], "MHZ")
        self.assertEqual(spec["容值误差"], "10")
        self.assertEqual(spec["负载电容（pF）"], "10")

    def test_timing_query_parses_detailed_crystal_requirements(self):
        spec = cm.parse_timing_spec_query(
            "晶振 38.4MHz 2016 10pF ±15ppm "
            "-20~85℃ 温度特性±15ppm 老化±1ppm 基频"
        )

        self.assertEqual(spec["容值误差"], "15")
        self.assertEqual(spec["工作温度"], "-20~85℃")
        self.assertEqual(spec["频率温度特性（ppm）"], "±15ppm")
        self.assertEqual(spec["25℃老化（ppm）"], "±1ppm")
        self.assertEqual(spec["泛音阶次"], "FUNDAMENTAL")

    def test_timing_query_parses_32khz_turnover_and_parabolic_coefficient(self):
        spec = cm.parse_timing_spec_query(
            "晶振 32.768kHz 3215 7pF ±20ppm -40~85℃ "
            "老化±3ppm 拐点温度+25±5℃ 抛物线系数-0.04"
        )

        self.assertEqual(spec["拐点温度"], "+25℃ ±5℃")
        self.assertEqual(spec["抛物线系数（ppm/℃²）"], "-0.04ppm/℃²")

    def test_timing_voltage_range_accepts_supported_nominal_voltage(self):
        self.assertTrue(cm.timing_voltage_allows("1.62~3.63", "3.3"))
        self.assertTrue(cm.timing_voltage_allows("3.135~3.465", "3.3"))
        self.assertFalse(cm.timing_voltage_allows("1.71~1.89", "3.3"))

    def test_cross_brand_timing_match_prioritizes_epson(self):
        rows = pd.DataFrame(
            [
                {
                    "品牌": "Abracon",
                    "型号": "ASV-25.000MHZ",
                    "系列": "ASV",
                    "器件类型": "振荡器",
                    "尺寸（inch）": "3225",
                    "容值": "25",
                    "容值单位": "MHz",
                    "容值误差": "25",
                    "耐压（V）": "3.3",
                    "电源电压": "3.3",
                    "输出类型": "CMOS",
                    "工作温度": "-40~85℃",
                    "25℃老化（ppm）": "±3ppm",
                },
                {
                    "品牌": "爱普生Epson",
                    "型号": "X1G0051810011",
                    "系列": "SG-8101CG",
                    "器件类型": "振荡器",
                    "尺寸（inch）": "3225",
                    "容值": "25",
                    "容值单位": "MHz",
                    "容值误差": "25",
                    "耐压（V）": "1.62~3.63",
                    "电源电压": "1.62~3.63",
                    "输出类型": "CMOS",
                    "工作温度": "-40~85℃",
                    "25℃老化（ppm）": "±3ppm",
                },
            ]
        )
        normalized = cm.normalize_imported_component_dataframe(rows)
        prepared = cm.prepare_search_dataframe(normalized)
        spec = cm.parse_timing_spec_query(
            "振荡器 25MHz 3225 3.3V CMOS ±25ppm -40~85℃ 老化±3ppm"
        )

        with mock.patch.object(cm, "fetch_search_candidate_pairs", return_value=None):
            matched = cm.match_other_passive_spec(prepared, spec)

        self.assertEqual(
            matched["型号"].tolist(),
            ["X1G0051810011", "ASV-25.000MHZ"],
        )
        self.assertTrue(matched["推荐等级"].eq("完全匹配").all())

    def test_sparse_timing_query_is_not_labeled_complete(self):
        rows = pd.DataFrame(
            [
                {
                    "品牌": "爱普生Epson",
                    "型号": "Q22FA12800007",
                    "系列": "FA-128",
                    "器件类型": "晶振",
                    "尺寸（inch）": "2016",
                    "容值": "38.4",
                    "容值单位": "MHz",
                    "容值误差": "15",
                    "负载电容（pF）": "10",
                    "工作温度": "-20~85℃",
                    "频率温度特性（ppm）": "±15ppm",
                    "25℃老化（ppm）": "±1ppm",
                    "泛音阶次": "基频（Fundamental）",
                }
            ]
        )
        prepared = cm.prepare_search_dataframe(
            cm.normalize_imported_component_dataframe(rows)
        )
        spec = cm.parse_timing_spec_query("晶振 38.4MHz 2016 10pF ±15ppm")

        with mock.patch.object(cm, "fetch_search_candidate_pairs", return_value=None):
            matched = cm.match_other_passive_spec(prepared, spec)

        self.assertEqual(matched["推荐等级"].tolist(), ["部分参数匹配"])

    def test_detailed_crystal_match_rejects_known_parameter_conflicts(self):
        rows = pd.DataFrame(
            [
                {
                    "品牌": "爱普生Epson",
                    "型号": "Q22FA12800007",
                    "系列": "FA-128",
                    "器件类型": "晶振",
                    "尺寸（inch）": "2016",
                    "容值": "38.4",
                    "容值单位": "MHz",
                    "容值误差": "15",
                    "负载电容（pF）": "10",
                    "工作温度": "-20~85℃",
                    "频率温度特性（ppm）": "±15ppm",
                    "25℃老化（ppm）": "±1ppm",
                    "泛音阶次": "基频（Fundamental）",
                },
                {
                    "品牌": "Example",
                    "型号": "CONFLICT-38M4",
                    "系列": "CONFLICT",
                    "器件类型": "晶振",
                    "尺寸（inch）": "2016",
                    "容值": "38.4",
                    "容值单位": "MHz",
                    "容值误差": "15",
                    "负载电容（pF）": "10",
                    "工作温度": "-20~70℃",
                    "频率温度特性（ppm）": "±30ppm",
                    "25℃老化（ppm）": "±5ppm",
                    "泛音阶次": "三次泛音（3rd Overtone）",
                },
            ]
        )
        prepared = cm.prepare_search_dataframe(
            cm.normalize_imported_component_dataframe(rows)
        )
        spec = cm.parse_timing_spec_query(
            "晶振 38.4MHz 2016 10pF ±15ppm "
            "-20~85℃ 温度特性±15ppm 老化±1ppm 基频"
        )

        with mock.patch.object(cm, "fetch_search_candidate_pairs", return_value=None):
            matched = cm.match_other_passive_spec(prepared, spec)

        self.assertEqual(matched["型号"].tolist(), ["Q22FA12800007"])
        self.assertEqual(matched["推荐等级"].tolist(), ["完全匹配"])

    def test_candidate_loader_merges_database_and_search_sidecar_rows(self):
        database_rows = pd.DataFrame(
            [
                {
                    "品牌": "爱普生Epson",
                    "型号": "FC135R",
                    "器件类型": "晶振",
                    "_component_type": "晶振",
                }
            ]
        )
        sidecar_rows = pd.DataFrame(
            [
                {
                    "品牌": "爱普生Epson",
                    "型号": "Q13FC13500002",
                    "器件类型": "晶振",
                    "_component_type": "晶振",
                }
            ]
        )
        connection = mock.MagicMock()

        with (
            mock.patch.object(cm.os.path, "exists", return_value=True),
            mock.patch.object(cm, "is_public_mode", return_value=False),
            mock.patch.object(cm, "is_streamlit_cloud_runtime", return_value=False),
            mock.patch.object(cm.sqlite3, "connect", return_value=connection),
            mock.patch.object(cm.pd, "read_sql_query", return_value=database_rows),
            mock.patch.object(
                cm,
                "load_search_sidecar_rows_by_brand_model_pairs",
                return_value=sidecar_rows,
            ),
            mock.patch.object(cm, "prepare_search_dataframe", side_effect=lambda frame: frame),
        ):
            merged = cm.load_component_rows_by_brand_model_pairs(
                [
                    ("爱普生Epson", "FC135R"),
                    ("爱普生Epson", "Q13FC13500002"),
                ],
                preferred_component_type="晶振",
            )

        self.assertEqual(
            set(merged["型号"].tolist()),
            {"FC135R", "Q13FC13500002"},
        )

    def test_exact_model_lookup_uses_search_index_before_large_database(self):
        indexed_row = {
            "品牌": "爱普生Epson",
            "型号": "RX8025T-UC",
            "器件类型": "实时时钟模块",
            "_component_type": "实时时钟模块",
            "容值": "32.768",
            "容值单位": "kHz",
        }
        connection = mock.MagicMock()
        connection.execute.return_value.fetchall.return_value = [
            ("RX8025T-UC", "爱普生Epson", "RX8025T-UC")
        ]

        with (
            mock.patch.object(cm, "open_search_db_connection", return_value=connection),
            mock.patch.object(cm, "search_index_can_serve_queries", return_value=True),
            mock.patch.object(
                cm,
                "load_search_sidecar_rows_by_brand_model_pairs",
                return_value=cm.prepare_search_dataframe(
                    cm.normalize_imported_component_dataframe(pd.DataFrame([indexed_row]))
                ),
            ),
            mock.patch.object(
                cm,
                "load_component_rows_by_clean_models_from_database",
            ) as database_lookup,
        ):
            result = cm.load_component_rows_by_clean_model("RX8025T-UC")

        database_lookup.assert_not_called()
        self.assertEqual(result["型号"].tolist(), ["RX8025T-UC"])
        self.assertEqual(result["器件类型"].tolist(), ["实时时钟模块"])


if __name__ == "__main__":
    unittest.main()
