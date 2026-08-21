import base64
import hashlib
import logging
import json
import os
import runpy
import shutil
import sqlite3
import tempfile
import threading
import time
import unittest
import warnings
from io import BytesIO
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qs, urlsplit

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class UploadedBytes:
    def __init__(self, name, data):
        self.name = name
        self._data = data
        self.size = len(data)

    def getvalue(self):
        return self._data

    def read(self, *args):
        return self._data

    def seek(self, *args):
        return 0


def dataframe_to_xlsx_bytes(frame):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="报价")
    return output.getvalue()


def formatted_bom_xlsx_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "格式BOM"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "4F81BD"
    sheet.page_setup.orientation = "landscape"
    sheet.auto_filter.ref = "A1:C3"
    sheet.freeze_panes = None
    sheet.merge_cells("E1:F1")
    sheet["E1"] = "原表保留区"

    thin_blue = Side(style="thin", color="4F81BD")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="1F4E78")
    body_font = Font(name="微软雅黑", size=10, color="333333")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = [
        ["型号", "规格", "数量"],
        ["RC0402FR-071KL", "0402 1KΩ ±1% 1/16W", 12000],
        ["GRM155R71C104KA88D", "0402 X7R 100nF 16V", 8000],
    ]
    for row_idx, values in enumerate(rows, start=1):
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx > 1:
                cell.font = body_font
                cell.border = Border(bottom=thin_blue)
                cell.alignment = Alignment(vertical="center")
    sheet["C2"].number_format = '#,##0" PCS"'
    sheet["A2"].hyperlink = "https://example.com/RC0402FR-071KL"
    sheet["A2"].style = "Hyperlink"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 36
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 7
    sheet.column_dimensions["D"].hidden = True
    sheet.column_dimensions["E"].width = 20
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].hidden = True
    sheet.print_title_rows = "1:1"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def fojan_quote_xlsx_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "报价"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "电阻系列产品报价单"
    sheet.append(["TO:", "", "TEL:", "", "Date:2026/6/12", ""])
    sheet.merge_cells("D3:E3")
    sheet.append(["Series", "Type / Dimension", "Resistance Range", "New Unit Price/含税成本Kpcs", "", "Package"])
    sheet.append(["", "", "Ω (ohms)", "5%", "1%", ""])
    sheet["D5"] = 0.05
    sheet["E5"] = 0.01
    sheet["D5"].number_format = "0%"
    sheet["E5"].number_format = "0%"
    sheet.append(["FRC", "0603 1/10W", "0R,510R-10M", "2.60", "", "5000PCS"])
    sheet.append(["FRC", "0603 1/10W", "10R-470R", "2.80", "", "5000PCS"])
    sheet.append(["FRC", "0603 1/10W", "1R-9.9R", "3.60", "", "5000PCS"])
    sheet.append(["FRC", "0603 1/10W", "10R-1M", "", "3.10", "5000PCS"])
    sheet.append(["FRC", "0603 1/10W", "1R-9.9R/1M1-10M", "", "3.63 / 3.2", "5000PCS"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def fojan_multi_sheet_quote_xlsx_bytes():
    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(name, series, tolerance_headers, prices):
        sheet = workbook.create_sheet(name)
        sheet.append(["Series", "Type / Dimension", "Resistance Range", "New Unit Price", "", "Package"])
        sheet.append(["", "", "Ω (ohms)", *tolerance_headers, ""])
        sheet.append([series, "0603 1/10W", "10R-1M", *prices, "5000PCS"])

    add_sheet("FRC&FRL", "FRC", ["5%（J）", "1%（F）"], ["2.80", "3.10"])
    add_sheet("FRH", "FRH", ["0.5%（D）", "0.1%（B）"], ["7.90", "19.00"])
    add_sheet("FRQ", "FRQ", ["5%（J）", "1%（F）"], ["4.20", "5.40"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def fojan_alloy_quote_xlsx_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Alloy"
    sheet.append(["客户代码", "通用", "FRM&FPM合金电阻系列产品报价单", "", "", "", ""])
    sheet.append(["Date:2026/8/7", "", "", "", "", "", ""])
    sheet.append(["Series", "产品", "功率", "精度", "Resistance Range", "Unit Price", "Package"])
    sheet.append(["", "", "", "", "Ω (ohms)", "", ""])
    sheet.append(["FRM", "2512", "2W", 0.01, "1~4mR大电极", "124.2", "4000PCS"])
    sheet.append(["", "", "", 0.05, "1~4mR大电极", "71.3", "4000PCS"])
    sheet.append(["", "", "", 0.01, "1-100mR", "120.75", "4000PCS"])
    sheet.append(["", "", "", 0.02, "101~500mR", "112.7", "4000PCS"])
    sheet.append(["FPM", "2512", "3W", 0.01, "1~4mR大电极", "147.2", "4000PCS"])
    sheet.append(["", "", "", 0.05, "1~100mR", "78.2", "4000PCS"])
    sheet.append(["FRM", "2010", "1W~1.5W", 0.01, "2~100mR", "135.7", "4000PCS"])
    sheet.append(["FRM", "1206", "1W", 0.01, "1mR大电极", "112.7", "5000PCS"])
    sheet.append(["", "", "", 0.01, "1-100mR", "83.95", "5000PCS"])
    sheet.append(["FMH金属膜合金", "1206", "1W", 0.01, "120mR-910mR", "83.95", "5000PCS"])
    sheet.append(["FCM裸片合金", "2512", "3W~6W", 0.01, "0.2mR~5mR", "200.1", "1000PCS"])
    sheet.append(["FWP 塑封合金", "2725/2728", "4W", 0.01, "0.2~200mR", "300.1", "1000PCS"])
    sheet.append(["FWK裸片合金", "1216", "9W", 0.01, "0.3mR/0.5mR/1mR/2mR/3mR", "400.1/400.2/400.3/400.4/400.5", "1000PCS"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class SystemRegressionTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        logging.disable(logging.CRITICAL)
        warnings.filterwarnings("ignore", category=ResourceWarning)
        cls.base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        cls.temp_dir = tempfile.mkdtemp(prefix="component-matcher-regression-")
        cls.saved_env = {
            key: os.environ.get(key)
            for key in [
                "MEMBER_AUTH_DB_PATH",
                "COST_PRICE_DB_PATH",
                "NO_MATCH_REPORT_DB_PATH",
                "COMPONENT_MATCHER_BUILD_MODE",
                "COMPONENT_MATCHER_STARTUP_MAINTENANCE",
            ]
        }
        os.environ["MEMBER_AUTH_DB_PATH"] = os.path.join(cls.temp_dir, "member.sqlite")
        os.environ["COST_PRICE_DB_PATH"] = os.path.join(cls.temp_dir, "cost.sqlite")
        os.environ["NO_MATCH_REPORT_DB_PATH"] = os.path.join(cls.temp_dir, "reports.sqlite")
        os.environ["COMPONENT_MATCHER_BUILD_MODE"] = "1"
        os.environ["COMPONENT_MATCHER_STARTUP_MAINTENANCE"] = "0"
        loaded = runpy.run_path(
            os.path.join(cls.base_dir, "component_matcher.py"),
            run_name="component_matcher_regression",
        )
        # runpy returns a snapshot-like mapping. Function globals are the live
        # module namespace that tests must patch when isolating database paths.
        cls.app = loaded["clean_text"].__globals__
        cls.original_paths = {
            "DB_PATH": cls.app["DB_PATH"],
            "SEARCH_DB_PATH": cls.app["SEARCH_DB_PATH"],
            "NO_MATCH_REPORT_DB_PATH": cls.app["NO_MATCH_REPORT_DB_PATH"],
            "COST_PRICE_DB_PATH": cls.app["COST_PRICE_DB_PATH"],
        }

    @classmethod
    def tearDownClass(cls):
        for key, value in cls.saved_env.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value
        shutil.rmtree(cls.temp_dir, ignore_errors=True)

    def test_00_runtime_database_paths_are_isolated(self):
        temp_root = os.path.normcase(os.path.abspath(self.temp_dir))
        for key in ("MEMBER_AUTH_DB_PATH", "COST_PRICE_DB_PATH", "NO_MATCH_REPORT_DB_PATH"):
            database_path = os.path.normcase(os.path.abspath(self.app[key]))
            self.assertEqual(os.path.commonpath([temp_root, database_path]), temp_root, key)

    def test_00a_duplicate_search_rows_use_unique_report_button_keys(self):
        captured_keys = []
        original_button = self.app["st"].button
        self.app["st"].button = lambda *args, **kwargs: captured_keys.append(kwargs["key"])
        try:
            for line_index in (1, 2):
                self.app["render_no_match_report_button"](
                    query_text="FRC0402F3242TS",
                    mode="part_number",
                    reason="no alternate",
                    key_prefix="no_alt_report",
                    instance_key=line_index,
                )
        finally:
            self.app["st"].button = original_button

        self.assertEqual(len(captured_keys), 2)
        self.assertNotEqual(captured_keys[0], captured_keys[1])

    def test_00b_member_remote_state_survives_streamlit_runpy_reruns(self):
        import member_auth_runtime

        app = self.app
        self.assertIs(app["MEMBER_AUTH_REMOTE_LOCK"], member_auth_runtime.REMOTE_LOCK)
        self.assertIs(app["MEMBER_AUTH_REMOTE_REFRESH_LOCK"], member_auth_runtime.REFRESH_LOCK)
        self.assertIs(app["_MEMBER_AUTH_REMOTE_REFRESH_CACHE"], member_auth_runtime.REFRESH_CACHE)
        self.assertIs(app["_MEMBER_AUTH_REMOTE_FLUSH_STATE"], member_auth_runtime.FLUSH_STATE)
        self.assertIs(app["MEMBER_AUTH_SCHEMA_LOCK"], member_auth_runtime.SCHEMA_LOCK)
        self.assertIs(app["_MEMBER_AUTH_SCHEMA_READY_PATHS"], member_auth_runtime.SCHEMA_READY_PATHS)
        self.assertIs(app["_MEMBER_PASSWORD_CACHE"], member_auth_runtime.PASSWORD_CACHE)
        self.assertTrue(hasattr(member_auth_runtime, "APP_CODE_LOCK"))
        self.assertTrue(hasattr(member_auth_runtime, "APP_CODE_CACHE"))

        original_runtime_state = app["member_auth_runtime_state"]
        legacy_runtime_state = type(
            "LegacyMemberAuthRuntime",
            (),
            {
                "REMOTE_LOCK": original_runtime_state.REMOTE_LOCK,
                "REFRESH_LOCK": original_runtime_state.REFRESH_LOCK,
                "REFRESH_CACHE": original_runtime_state.REFRESH_CACHE,
                "FLUSH_CONDITION": original_runtime_state.FLUSH_CONDITION,
                "FLUSH_STATE": original_runtime_state.FLUSH_STATE,
            },
        )()
        try:
            app["member_auth_runtime_state"] = legacy_runtime_state
            app["ensure_member_auth_runtime_state_compatibility"]()
            self.assertTrue(hasattr(legacy_runtime_state, "SCHEMA_LOCK"))
            self.assertTrue(hasattr(legacy_runtime_state, "SCHEMA_READY_PATHS"))
            self.assertTrue(hasattr(legacy_runtime_state, "PASSWORD_CACHE_LOCK"))
            self.assertTrue(hasattr(legacy_runtime_state, "PASSWORD_CACHE_SECRET"))
            self.assertTrue(hasattr(legacy_runtime_state, "PASSWORD_CACHE"))
        finally:
            app["member_auth_runtime_state"] = original_runtime_state

    def test_01_exact_model_categories_and_library_rows(self):
        models = [
            "AC0402KRX7R9BB103",
            "GRM155R71C224KA12D",
            "BBGK00201209202Y00",
            "NCP15XH103F03RC",
        ]
        by_model = self.app["load_component_rows_by_exact_models_from_search_sidecar"](models)
        self.assertEqual(
            set(by_model[self.app["clean_model"](models[0])]["器件类型"].map(self.app["normalize_component_type"])),
            {"MLCC"},
        )
        self.assertEqual(
            set(by_model[self.app["clean_model"](models[1])]["器件类型"].map(self.app["normalize_component_type"])),
            {"MLCC"},
        )
        self.assertEqual(
            set(by_model[self.app["clean_model"](models[2])]["器件类型"].map(self.app["normalize_component_type"])),
            {"磁珠"},
        )
        self.assertEqual(
            set(by_model[self.app["clean_model"](models[3])]["器件类型"].map(self.app["normalize_component_type"])),
            {"热敏电阻"},
        )

        representative_models = [
            "FBF06FT-3R00N",
            "FAF02FVA1001QMH",
            "FPF05FTF1004NM",
            "FPS03FTE10R0NMD",
            "FMF06FTHR010-BH",
            "JAS103F344FB",
            "JFR103F344FB25025CPG",
            "JNR05S030L",
            "JVT10N180M",
            "JVZ10N180M",
        ]
        imported = self.app["load_component_rows_by_exact_models_from_search_sidecar"](representative_models)
        self.assertTrue(
            all(
                not imported[self.app["clean_model"](model)].empty
                for model in representative_models
            )
        )

        fojan_models = [
            "FRC1206P000TS", "FRC0603J100TS", "FRC0402J330TS", "FRC1206J201TS",
            "FRC0402J511TS", "FRC0402J102TS", "FRC0402J222TS", "FRC0402J472TS",
            "FRC0402J103TS", "FRC0805J103TS", "FRC0402J152TS", "FRC0402F1302TS",
            "FRC0402J513TS", "FRC0402P000TS", "FRC0603P000TS", "FRC0402F1003TS",
            "FRC0603J750TS", "FRC0603F8R20TS", "FRC0402F4701TS", "FRC0402J562TS",
            "FRC0402J303TS", "FRC0402J204TS", "FRC0402F2402TS", "FRC0603F3322TS",
            "FRC0603F5362TS", "FRC0603F2702TS", "FRC0603F4701TS", "FRC0402F2000TS",
            "FRC0402F49R9TS", "FRC0402F7502TS", "FRC0603F3R60TS", "FRC0603F1272TS",
            "FRC0603F2432TS", "FRC0402F3R30TS", "FRC0402F3922TS", "FRC0402F5113TS",
            "FRC0603J561TS", "FRC0603F1962TS", "FRC0603F1053TS",
        ]
        fojan_rows = self.app["load_component_rows_by_exact_models_from_search_sidecar"](fojan_models)
        self.assertTrue(
            all(not fojan_rows[self.app["clean_model"](model)].empty for model in fojan_models)
        )
        for model in fojan_models:
            row = fojan_rows[self.app["clean_model"](model)].iloc[0]
            expected_tolerance = "5" if "J" in model[7:9] or "P000" in model else "1"
            self.assertEqual(str(row["_tol"]), expected_tolerance, model)

        spaced_fojan_5_percent_models = [
            "FRC0603J100 TS", "FRC0402J330 TS", "FRC1206J201 TS", "FRC0402J511 TS",
            "FRC0402J102 TS", "FRC0402J222 TS", "FRC0402J472 TS", "FRC0402J103 TS",
            "FRC0805J103 TS", "FRC0402J152 TS", "FRC0402J513 TS", "FRC0603J750 TS",
            "FRC0402J562 TS", "FRC0402J303 TS", "FRC0402J204 TS", "FRC0603J561 TS",
        ]
        for model in spaced_fojan_5_percent_models:
            resolved = self.app["resolve_search_query_dataframe_and_spec"](model)
            self.assertEqual(resolved["mode"], "料号", model)
            self.assertNotEqual(resolved["resolution_path"], "model_token_prefix_lookup", model)
            rows = resolved["query_df"]
            self.assertTrue(
                rows["_model_clean"].astype(str).eq(self.app["clean_model"](model)).any(),
                model,
            )
            self.assertEqual(str(resolved["spec"].get("容值误差", "")), "5", model)

    def test_02_member_auth_approval_profile_and_search_logs(self):
        app = self.app
        app["ensure_configured_admin_member_account"]()
        admin, message = app["authenticate_member"]("TERRY46", "123456")
        self.assertEqual(message, "")
        self.assertEqual(admin["role"], "admin")
        self.assertTrue(admin["password_hash"].startswith("scrypt$"))
        self.assertNotIn("123456", admin["password_hash"])

        ok, message = app["create_member_account"](
            "CaseUser", "secret1", "Case User", "Old Co", "old@example.com", "100"
        )
        self.assertTrue(ok, message)
        pending, message = app["authenticate_member"]("caseuser", "secret1")
        self.assertIsNone(pending)
        self.assertIn("审核", message)
        duplicate_ok, _ = app["create_member_account"]("CASEUSER", "secret1")
        self.assertFalse(duplicate_ok)

        member = app["get_member_by_username"]("CASEuser")
        ok, message = app["approve_member_account_admin"](member["id"])
        self.assertTrue(ok, message)
        member, message = app["authenticate_member"]("CASEUSER", "secret1")
        self.assertIsNotNone(member, message)
        token = member["_session_token"]
        with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
            initial_expires_at = conn.execute(
                "SELECT expires_at_ts FROM member_sessions WHERE token=?", (token,)
            ).fetchone()[0]
        expected_ttl = app["MEMBER_AUTH_SESSION_TTL_SECONDS"]
        self.assertGreaterEqual(initial_expires_at, int(time.time()) + expected_ttl - 10)
        with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
            conn.execute(
                "UPDATE member_sessions SET expires_at_ts=? WHERE token=?",
                (int(time.time()) + 5, token),
            )
            conn.commit()
        self.assertIsNotNone(app["get_member_by_session_token"](token))
        with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
            expires_at = conn.execute(
                "SELECT expires_at_ts FROM member_sessions WHERE token=?", (token,)
            ).fetchone()[0]
        self.assertGreaterEqual(expires_at, int(time.time()) + expected_ttl - 10)

        ok, message = app["update_current_member_profile"](
            member["id"], "Case Renamed", "New Co", "new@example.com", "200", "客户A"
        )
        self.assertFalse(ok, message)
        self.assertIn("后台管理员", message)
        updated_member = app["get_member_by_id"](member["id"])
        self.assertEqual(updated_member["customer_name"], "")
        logs_before = app["list_member_profile_change_logs"](member["id"])
        self.assertEqual(len(logs_before), 0)

        ok, message = app["update_member_account_admin"](
            member["id"],
            username=member["username"],
            display_name="Case Renamed",
            company="New Co",
            customer_name="客户A",
            job_title=member.get("job_title", ""),
            email="new@example.com",
            phone="200",
            role=member["role"],
            status=member["status"],
            actor_username="regression-admin",
        )
        self.assertTrue(ok, message)
        updated_member = app["get_member_by_id"](member["id"])
        self.assertEqual(updated_member["customer_name"], "客户A")
        member_customers = app["list_member_sales_customers"](member["id"])
        self.assertEqual([row["customer_name"] for row in member_customers], ["客户A"])
        self.assertEqual(int(member_customers[0]["price_access_enabled"]), 1)
        logs_before = app["list_member_profile_change_logs"](member["id"])
        self.assertGreaterEqual(len(logs_before), 5)
        ok, message = app["change_current_member_password"](
            member["id"], "secret1", "secret2", "secret2"
        )
        self.assertTrue(ok, message)
        logs_after = app["list_member_profile_change_logs"](member["id"])
        self.assertEqual(len(logs_after), len(logs_before))
        self.assertTrue(all("password" not in str(row.get("field_name", "")).lower() for row in logs_after))
        member, message = app["authenticate_member"]("caseuser", "secret2")
        self.assertIsNotNone(member, message)

        app["record_member_search_logs"](
            member,
            ["0402 10K 1% 1/16W", "0402 10K ±1% 1/16W", "0805 X7R 100nF 10% 50V"],
            source="regression",
        )
        summary = app["list_member_search_log_summary"]("", "", "", 300)
        self.assertGreaterEqual(len(summary), 2)
        for period in ["daily", "weekly", "monthly"]:
            trend = app["build_member_search_trend_dataframe"](summary, period=period)
            self.assertFalse(trend.empty)
            self.assertLessEqual(int(trend["排名"].max()), 10)

    def test_02a_member_password_upgrade_and_configured_admin_fast_path(self):
        app = self.app
        app["ensure_configured_admin_member_account"]()

        original_verify = app["verify_member_password"]
        try:
            def unexpected_admin_hash_verify(*_args, **_kwargs):
                raise AssertionError("configured administrator login should not run password KDF")

            app["verify_member_password"] = unexpected_admin_hash_verify
            admin, message = app["authenticate_member"]("terry46", "123456")
            self.assertIsNotNone(admin, message)
        finally:
            app["verify_member_password"] = original_verify

        salt = b"legacy-member-salt"
        iterations = 12000
        digest = app["hashlib"].pbkdf2_hmac("sha256", b"legacy-secret", salt, iterations)
        legacy_hash = "pbkdf2_sha256${}${}${}".format(
            iterations,
            app["base64"].b64encode(salt).decode("ascii"),
            app["base64"].b64encode(digest).decode("ascii"),
        )
        now = app["current_timestamp_text"]()
        with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
            conn.execute("DELETE FROM members WHERE lower(username)=lower(?)", ("LegacyHashUser",))
            conn.execute(
                """
                INSERT INTO members (
                    username, password_hash, display_name, role, status, created_at, updated_at
                ) VALUES (?, ?, ?, 'member', 'active', ?, ?)
                """,
                ("LegacyHashUser", legacy_hash, "Legacy Hash User", now, now),
            )
            conn.commit()

        member, message = app["authenticate_member"]("legacyhashuser", "legacy-secret")
        self.assertIsNotNone(member, message)
        self.assertTrue(member["password_hash"].startswith("scrypt$"))
        self.assertTrue(
            app["member_password_cache_contains"]("legacy-secret", member["password_hash"])
        )
        with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
            stored_hash = conn.execute(
                "SELECT password_hash FROM members WHERE id=?", (int(member["id"]),)
            ).fetchone()[0]
        self.assertEqual(stored_hash, member["password_hash"])

    def test_02aa_legacy_member_database_adds_profile_columns_without_losing_accounts(self):
        app = self.app
        original_member_path = app["MEMBER_AUTH_DB_PATH"]
        legacy_path = os.path.join(self.temp_dir, "legacy-member-customer.sqlite")
        try:
            with sqlite3.connect(legacy_path) as conn:
                conn.executescript(
                    """
                    CREATE TABLE members (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        username TEXT NOT NULL UNIQUE,
                        password_hash TEXT NOT NULL,
                        display_name TEXT NOT NULL DEFAULT '',
                        company TEXT NOT NULL DEFAULT '',
                        email TEXT NOT NULL DEFAULT '',
                        phone TEXT NOT NULL DEFAULT '',
                        role TEXT NOT NULL DEFAULT 'member',
                        status TEXT NOT NULL DEFAULT 'active',
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL,
                        last_login_at TEXT NOT NULL DEFAULT ''
                    );
                    INSERT INTO members
                        (username, password_hash, display_name, created_at, updated_at)
                    VALUES ('legacy-customer', 'hash', 'Legacy Customer', '2026-01-01', '2026-01-01');
                    """
                )
            app["MEMBER_AUTH_DB_PATH"] = legacy_path
            # Reproduce a Streamlit hot deployment where the previous code left
            # an unversioned "schema ready" marker for the still-legacy database.
            app["_MEMBER_AUTH_SCHEMA_READY_PATHS"].add(os.path.abspath(legacy_path))
            app["ensure_member_auth_schema"]()
            with sqlite3.connect(legacy_path) as conn:
                columns = {row[1] for row in conn.execute("PRAGMA table_info(members)").fetchall()}
                row = conn.execute(
                    "SELECT username, display_name, customer_name, job_title FROM members"
                ).fetchone()
            self.assertIn("customer_name", columns)
            self.assertIn("job_title", columns)
            self.assertEqual(row, ("legacy-customer", "Legacy Customer", "", ""))
            with sqlite3.connect(legacy_path) as conn:
                customer_table = conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name='member_sales_customers'"
                ).fetchone()
            self.assertIsNotNone(customer_table)
            self.assertIn(
                (os.path.abspath(legacy_path), app["MEMBER_AUTH_SCHEMA_VERSION"]),
                app["_MEMBER_AUTH_SCHEMA_READY_PATHS"],
            )
        finally:
            app["MEMBER_AUTH_DB_PATH"] = original_member_path

    def test_02aaa_member_customer_lists_are_isolated_and_default_to_general_price(self):
        app = self.app
        member_ids = []
        for username in ["CustomerOwnerA", "CustomerOwnerB"]:
            ok, message = app["create_member_account"](
                username,
                "secret1",
                display_name=username,
                company="Test Co",
            )
            self.assertTrue(ok, message)
            member = app["get_member_by_username"](username)
            member_ids.append(int(member["id"]))
            ok, message = app["approve_member_account_admin"](member["id"])
            self.assertTrue(ok, message)

        ok, message, saved = app["save_member_sales_customer"](member_ids[0], "客户甲")
        self.assertFalse(ok)
        self.assertIn("全称", message)
        self.assertIsNone(saved)
        self.assertEqual(app["list_member_sales_customers"](member_ids[0]), [])

        ok, message, saved = app["save_member_sales_customer"](member_ids[0], "客户甲有限公司")
        self.assertTrue(ok, message)
        self.assertEqual(saved["customer_name"], "客户甲有限公司")
        self.assertEqual(int(saved["price_access_enabled"]), 0)
        self.assertEqual(len(app["list_member_sales_customers"](member_ids[0])), 1)
        self.assertEqual(app["list_member_sales_customers"](member_ids[1]), [])

        selected = app["select_member_sales_customer"](member_ids[0], "客户甲有限公司")
        self.assertIsNotNone(selected)
        self.assertIsNone(app["select_member_sales_customer"](member_ids[1], "客户甲有限公司"))

        customer_id = int(saved["id"])
        ok, message = app["set_member_sales_customer_price_access"](
            member_ids[0], customer_id, True
        )
        self.assertFalse(ok)
        self.assertIn("后台尚未维护", message)

        ok, message, _ = app["save_sales_customer"](
            "客户甲有限公司",
            "TEST-A-001",
            group_name="客户甲集团有限公司",
            updated_by="regression",
            sync_remote=False,
        )
        self.assertTrue(ok, message)
        ok, message = app["set_member_sales_customer_price_access"](
            member_ids[1], customer_id, True
        )
        self.assertFalse(ok)
        self.assertIn("没有这条客户记录", message)
        ok, message = app["set_member_sales_customer_price_access"](
            member_ids[0], customer_id, True
        )
        self.assertTrue(ok, message)
        authorized = app["list_member_sales_customers"](member_ids[0])[0]
        self.assertEqual(int(authorized["price_access_enabled"]), 1)

    def test_02aa1_admin_customer_selector_lists_all_member_customers_and_defaults_to_general(self):
        app = self.app
        customer_names = [
            "管理员查看深圳客户有限公司",
            "管理员查看上海客户有限公司",
        ]
        member_rows = []
        for username, customer_name in [
            ("AdminCostOwnerA", customer_names[0]),
            ("AdminCostOwnerB", customer_names[1]),
            ("AdminCostOwnerDisabled", "管理员隐藏客户有限公司"),
        ]:
            ok, message = app["create_member_account"](
                username,
                "secret1",
                display_name=username,
                company="Regression Co",
            )
            self.assertTrue(ok, message)
            member = app["get_member_by_username"](username)
            member_rows.append(member)
            ok, message = app["approve_member_account_admin"](member["id"])
            self.assertTrue(ok, message)
            ok, message, saved = app["save_member_sales_customer"](member["id"], customer_name)
            self.assertTrue(ok, message)
            self.assertEqual(saved["customer_name"], customer_name)

        disabled_member = member_rows[-1]
        ok, message = app["update_member_account_admin"](
            disabled_member["id"],
            username=disabled_member["username"],
            display_name=disabled_member["display_name"],
            company=disabled_member["company"],
            email=disabled_member["email"],
            phone=disabled_member["phone"],
            role=disabled_member["role"],
            status="disabled",
            actor_username="regression-admin",
        )
        self.assertTrue(ok, message)
        ok, message, _ = app["save_sales_customer"](
            customer_names[0],
            "ADM-COST-001",
            group_name="管理员客户集团有限公司",
            updated_by="regression",
            sync_remote=False,
        )
        self.assertTrue(ok, message)

        admin_customer_rows = app["list_member_sales_customers_for_admin"]()
        admin_customer_names = {row["customer_name"] for row in admin_customer_rows}
        self.assertIn(customer_names[0], admin_customer_names)
        self.assertIn(customer_names[1], admin_customer_names)
        self.assertNotIn("管理员隐藏客户有限公司", admin_customer_names)

        class FakeStreamlit:
            def __init__(self, selection=""):
                self.selection = selection
                self.session_state = {}
                self.selectbox_calls = []
                self.success_messages = []
                self.caption_messages = []
                self.info_messages = []

            def selectbox(self, label, options, key=None, **kwargs):
                options = list(options)
                self.selectbox_calls.append(
                    {
                        "label": label,
                        "options": options,
                        "key": key,
                        "help": kwargs.get("help", ""),
                    }
                )
                value = self.selection if self.selection in options else options[0]
                if key:
                    self.session_state[key] = value
                return value

            def success(self, value):
                self.success_messages.append(value)

            def caption(self, value):
                self.caption_messages.append(value)

            def info(self, value):
                self.info_messages.append(value)

        original_functions = {
            name: app[name]
            for name in [
                "st",
                "current_member",
                "current_member_is_admin",
            ]
        }
        try:
            app["current_member"] = lambda: {
                "id": 900001,
                "username": "regression-admin",
                "role": "admin",
                "status": "active",
            }
            app["current_member_is_admin"] = lambda: True

            fake_st = FakeStreamlit()
            app["st"] = fake_st
            customer_type, customer_name, ready = app["render_sales_cost_customer_selector"](
                key_prefix="admin_cost_regression"
            )
            self.assertEqual((customer_type, customer_name, ready), ("new", "", True))
            options = fake_st.selectbox_calls[0]["options"]
            self.assertEqual(options[0], "通用成本（不指定客户）")
            self.assertIn(customer_names[0], options)
            self.assertIn(customer_names[1], options)
            self.assertNotIn("管理员隐藏客户有限公司", options)
            self.assertIn("直接搜索料号", fake_st.caption_messages[-1])

            fake_selected_st = FakeStreamlit(selection=customer_names[0])
            app["st"] = fake_selected_st
            customer_type, customer_name, ready = app["render_sales_cost_customer_selector"](
                key_prefix="admin_cost_regression_selected"
            )
            self.assertEqual((customer_type, customer_name, ready), ("existing", customer_names[0], True))
            self.assertEqual(
                fake_selected_st.session_state[app["SALES_COST_CUSTOMER_TYPE_KEY"]],
                "existing",
            )
            self.assertEqual(
                fake_selected_st.session_state[app["SALES_COST_CUSTOMER_NAME_KEY"]],
                customer_names[0],
            )
            self.assertIn("客户价格", fake_selected_st.success_messages[-1])

            fake_unpriced_st = FakeStreamlit(selection=customer_names[1])
            app["st"] = fake_unpriced_st
            customer_type, customer_name, ready = app["render_sales_cost_customer_selector"](
                key_prefix="admin_cost_regression_unpriced"
            )
            self.assertEqual((customer_type, customer_name, ready), ("new", "", True))
            self.assertIn("通用价格", fake_unpriced_st.success_messages[-1])
        finally:
            app.update(original_functions)
            app["clear_cost_price_lookup_cache"]()

    def test_02aab_new_member_customer_requires_legal_company_full_name(self):
        app = self.app
        valid_names = [
            "深圳市示例科技有限公司",
            "示例股份有限公司",
            "Example Technologies Co., Ltd.",
            "Example Systems, Inc.",
            "Example Holdings LLC",
            "Example Electronics Pte. Ltd.",
            "Example Australia Pty Ltd",
            "Example Deutschland GmbH",
            "株式会社サンプル",
            "샘플 주식회사",
        ]
        for customer_name in valid_names:
            with self.subTest(customer_name=customer_name):
                ok, message = app["validate_customer_legal_full_name"](customer_name)
                self.assertTrue(ok, message)

        invalid_names = ["星际悦动", "Example", "Example Electronics", "客户甲"]
        for customer_name in invalid_names:
            with self.subTest(customer_name=customer_name):
                ok, message = app["validate_customer_legal_full_name"](customer_name)
                self.assertFalse(ok)
                self.assertIn("全称", message)

    def test_02ab_job_title_is_admin_managed_and_controls_cost_visibility(self):
        app = self.app
        ok, message = app["create_member_account"](
            "SalesTitleUser",
            "secret1",
            display_name="Sales Title User",
            company="Sales Co",
            email="sales@example.com",
            phone="300",
        )
        self.assertTrue(ok, message)
        member = app["get_member_by_username"]("salestitleuser")
        ok, message = app["approve_member_account_admin"](member["id"])
        self.assertTrue(ok, message)
        member = app["get_member_by_id"](member["id"])
        self.assertEqual(member.get("job_title", ""), "")
        self.assertEqual(app["member_cost_access_level"](member), "general")
        self.assertTrue(app["member_can_view_cost"](member))

        ok, message = app["update_member_account_admin"](
            member["id"],
            username=member["username"],
            display_name=member["display_name"],
            company=member["company"],
            customer_name=member.get("customer_name", ""),
            job_title="销售",
            email=member["email"],
            phone=member["phone"],
            role=member["role"],
            status=member["status"],
            actor_username="regression-admin",
        )
        self.assertTrue(ok, message)
        sales_member = app["get_member_by_id"](member["id"])
        self.assertEqual(sales_member["job_title"], "销售")
        self.assertTrue(app["member_can_view_cost"](sales_member))

        ok, message = app["update_current_member_profile"](
            member["id"],
            "Sales Renamed",
            "Sales Co 2",
            "sales2@example.com",
            "301",
        )
        self.assertTrue(ok, message)
        self.assertEqual(app["get_member_by_id"](member["id"])["job_title"], "销售")

        non_sales_member = dict(sales_member, job_title="工程")
        assistant_member = dict(sales_member, job_title=" 销售 助理 ")
        admin_member = dict(non_sales_member, role="admin")
        self.assertEqual(app["member_cost_access_level"](non_sales_member), "general")
        self.assertEqual(app["member_cost_access_level"](assistant_member), "sales")
        self.assertEqual(app["member_cost_access_level"](admin_member), "admin")
        self.assertTrue(app["member_can_view_cost"](non_sales_member))
        self.assertTrue(app["member_can_view_cost"](assistant_member))
        self.assertTrue(app["member_can_view_cost"](admin_member))

        visible_df = pd.DataFrame(
            [{"品牌": "FOJAN(富捷)", "型号": "FRC0402F1001TS", "成本": "1.70", "更新时间": "2026-08-12", "MOQ": "10000PCS"}]
        )
        restricted_df = app["apply_search_cost_visibility"](visible_df, can_view_cost=False)
        self.assertNotIn("成本", restricted_df.columns)
        self.assertNotIn("更新时间", restricted_df.columns)
        self.assertIn("MOQ", restricted_df.columns)
        self.assertIn("成本", app["apply_search_cost_visibility"](visible_df, can_view_cost=True).columns)

        summary = app["member_admin_summary_dataframe"]([app["get_member_by_id"](member["id"])])
        self.assertIn("职务", summary.columns)
        logs = app["list_member_profile_change_logs"](member["id"])
        self.assertTrue(any(row.get("field_name") == "job_title" for row in logs))

    def test_02ac_role_price_scopes_and_pm_brand_permissions(self):
        app = self.app
        customer_name = "权限测试客户有限公司"
        ok, message, _ = app["save_sales_customer"](
            customer_name,
            "AUTH-001",
            group_name="权限测试集团有限公司",
            updated_by="regression",
            sync_remote=False,
        )
        self.assertTrue(ok, message)

        members = {}
        for username, job_title in [
            ("PriceScopePm", "PM"),
            ("PriceScopeSales", "销售"),
            ("PriceScopeOther", "其他"),
        ]:
            ok, message = app["create_member_account"](
                username,
                "secret1",
                display_name=username,
                company="Price Scope Co",
            )
            self.assertTrue(ok, message)
            member = app["get_member_by_username"](username)
            ok, message = app["approve_member_account_admin"](member["id"])
            self.assertTrue(ok, message)
            member = app["get_member_by_id"](member["id"])
            ok, message = app["update_member_account_admin"](
                member["id"],
                username=member["username"],
                display_name=member["display_name"],
                company=member["company"],
                customer_name=member.get("customer_name", ""),
                job_title=job_title,
                pm_brands=["FOJAN(富捷)"] if job_title == "PM" else [],
                email=member["email"],
                phone=member["phone"],
                role=member["role"],
                status=member["status"],
                actor_username="regression-admin",
            )
            self.assertTrue(ok, message)
            members[job_title] = app["get_member_by_id"](member["id"])

        sales_member = members["销售"]
        ok, message, saved = app["save_member_sales_customer"](
            sales_member["id"], customer_name
        )
        self.assertTrue(ok, message)
        ok, message = app["set_member_sales_customer_price_access"](
            sales_member["id"], saved["id"], True
        )
        self.assertTrue(ok, message)

        self.assertEqual(app["list_member_pm_brands"](members["PM"]["id"]), ["FOJAN(富捷)"])
        self.assertEqual(
            app["authorize_cost_customer_context"](
                members["PM"], app["COST_CUSTOMER_TYPE_EXISTING"], customer_name
            ),
            (app["COST_CUSTOMER_TYPE_EXISTING"], customer_name),
        )
        self.assertEqual(
            app["authorize_cost_customer_context"](
                sales_member, app["COST_CUSTOMER_TYPE_EXISTING"], customer_name
            ),
            (app["COST_CUSTOMER_TYPE_EXISTING"], customer_name),
        )
        self.assertEqual(
            app["authorize_cost_customer_context"](
                members["其他"], app["COST_CUSTOMER_TYPE_EXISTING"], customer_name
            ),
            (app["COST_CUSTOMER_TYPE_NEW"], ""),
        )

        lookup = {
            "FOJAN-CUSTOMER": [
                {
                    "brand": "FOJAN(富捷)",
                    "customer_type": app["COST_CUSTOMER_TYPE_EXISTING"],
                    "_scope_rank": 0,
                }
            ],
            "YAGEO-CUSTOMER": [
                {
                    "brand": "国巨YAGEO",
                    "customer_type": app["COST_CUSTOMER_TYPE_EXISTING"],
                    "_scope_rank": 0,
                }
            ],
            "YAGEO-GENERAL": [
                {
                    "brand": "国巨YAGEO",
                    "customer_type": app["COST_CUSTOMER_TYPE_NEW"],
                    "_scope_rank": 25,
                }
            ],
            "__fojan_resistor_rules__": [{"brand": "FOJAN(富捷)"}],
        }
        pm_lookup = app["filter_cost_lookup_for_member"](lookup, members["PM"])
        self.assertIn("FOJAN-CUSTOMER", pm_lookup)
        self.assertNotIn("YAGEO-CUSTOMER", pm_lookup)
        self.assertIn("YAGEO-GENERAL", pm_lookup)
        self.assertIn("__fojan_resistor_rules__", pm_lookup)
        self.assertIs(app["filter_cost_lookup_for_member"](lookup, sales_member), lookup)

    def test_02b_member_login_returns_to_requesting_page(self):
        app = self.app
        original_functions = {
            name: app[name]
            for name in [
                "set_current_member",
                "is_member_page_requested",
                "is_bom_page_requested",
                "update_query_params",
                "st",
            ]
        }
        member_calls = []
        route_updates = []
        fake_st = type("FakeStreamlit", (), {"session_state": {}})()
        try:
            app["st"] = fake_st
            app["set_current_member"] = lambda member, query_updates=None: member_calls.append(
                (member, query_updates)
            )
            app["update_query_params"] = lambda **updates: route_updates.append(updates)
            app["is_bom_page_requested"] = lambda: False
            app["is_member_page_requested"] = lambda: True
            app["complete_member_login"]({"id": 7})
            self.assertEqual(
                member_calls,
                [({"id": 7}, {"member": "", "admin": "", "bom": ""})],
            )
            self.assertEqual(route_updates, [])

            app["is_member_page_requested"] = lambda: False
            app["complete_member_login"]({"id": 8})
            self.assertEqual(member_calls[-1], ({"id": 8}, None))
            self.assertEqual(len(route_updates), 0)

            fake_st.session_state[app["BOM_PENDING_UPLOAD_WAITING_LOGIN_KEY"]] = True
            app["is_bom_page_requested"] = lambda: True
            app["complete_member_login"]({"id": 9})
            self.assertEqual(member_calls[-1], ({"id": 9}, None))
            self.assertEqual(
                fake_st.session_state[app["BOM_POST_LOGIN_RESUME_STAGE_KEY"]],
                app["BOM_POST_LOGIN_STAGE_LOGIN_COMPLETE"],
            )
        finally:
            app.update(original_functions)

    def test_02ba_page_modes_are_mutually_exclusive(self):
        app = self.app
        original_get_query_param_value = app["get_query_param_value"]
        params = {}
        try:
            app["get_query_param_value"] = lambda name: params.get(name, "")

            params.update({"member": "1", "bom": "1"})
            self.assertEqual(app["requested_page_mode"](), "member")
            self.assertTrue(app["is_member_page_requested"]())
            self.assertFalse(app["is_bom_page_requested"]())

            params.update({"admin": "1"})
            self.assertEqual(app["requested_page_mode"](), "admin")
            self.assertTrue(app["is_no_match_admin_page_requested"]())
            self.assertFalse(app["is_member_page_requested"]())
            self.assertFalse(app["is_bom_page_requested"]())

            params.clear()
            params["bom"] = "1"
            self.assertEqual(app["requested_page_mode"](), "bom")
            self.assertTrue(app["is_bom_page_requested"]())
        finally:
            app["get_query_param_value"] = original_get_query_param_value

    def test_02baa_backend_entry_is_visible_only_to_admin_members(self):
        app = self.app

        class FakeStreamlit:
            def __init__(self):
                self.session_state = {
                    "_no_match_admin_authenticated": True,
                    "_member_auth_token": "admin-member-token",
                }
                self.markup = []

            def markdown(self, value, **kwargs):
                self.markup.append(value)

        fake_st = FakeStreamlit()
        original_functions = {
            name: app[name]
            for name in [
                "st",
                "current_member_is_admin",
                "is_no_match_admin_page_requested",
                "build_app_href",
                "get_query_param_value",
            ]
        }
        try:
            app["st"] = fake_st
            app["is_no_match_admin_page_requested"] = lambda: False
            app["build_app_href"] = lambda **updates: "?" + "&".join(
                f"{key}={value}" for key, value in updates.items()
            )
            app["get_query_param_value"] = lambda name: ""

            app["current_member_is_admin"] = lambda: False
            app["render_no_match_admin_entry_button"]()
            self.assertEqual(fake_st.markup, [])
            self.assertNotIn("_no_match_admin_authenticated", fake_st.session_state)

            app["current_member_is_admin"] = lambda: True
            app["render_no_match_admin_entry_button"]()
            self.assertEqual(len(fake_st.markup), 1)
            self.assertIn("admin-login-fixed", fake_st.markup[0])
            self.assertIn("进入后台", fake_st.markup[0])
            self.assertIn("admin=1", fake_st.markup[0])
            self.assertIn(
                f"{app['MEMBER_AUTH_QUERY_PARAM']}=admin-member-token",
                fake_st.markup[0],
            )
            self.assertNotIn("登入后台", fake_st.markup[0])
        finally:
            app.update(original_functions)

    def test_02bab_backend_access_uses_member_admin_role_only(self):
        app = self.app

        class FakeStreamlit:
            def __init__(self):
                self.session_state = {"_no_match_admin_authenticated": True}
                self.markup = []

            def markdown(self, value, **kwargs):
                self.markup.append(value)

        fake_st = FakeStreamlit()
        original_functions = {
            name: app[name]
            for name in [
                "st",
                "current_member_is_admin",
            ]
        }
        try:
            app["st"] = fake_st
            app["current_member_is_admin"] = lambda: False
            self.assertFalse(app["require_no_match_admin_login"]())
            self.assertNotIn("_no_match_admin_authenticated", fake_st.session_state)
            self.assertTrue(any("会员登录" in value for value in fake_st.markup))

            fake_st.markup.clear()
            app["current_member_is_admin"] = lambda: True
            self.assertTrue(app["require_no_match_admin_login"]())
            self.assertTrue(fake_st.session_state["_no_match_admin_authenticated"])
            self.assertEqual(fake_st.markup, [])
        finally:
            app.update(original_functions)

    def test_02bac_unauthorized_backend_route_keeps_member_login_entry(self):
        app = self.app

        class FakeStreamlit:
            def __init__(self):
                self.session_state = {}
                self.markup = []

            def markdown(self, value, **kwargs):
                self.markup.append(value)

        fake_st = FakeStreamlit()
        original_functions = {
            name: app[name]
            for name in [
                "st",
                "is_no_match_admin_page_requested",
                "current_member_is_admin",
                "current_member",
                "build_app_href",
            ]
        }
        try:
            app["st"] = fake_st
            app["is_no_match_admin_page_requested"] = lambda: True
            app["current_member_is_admin"] = lambda: False
            app["current_member"] = lambda: None
            app["build_app_href"] = lambda **updates: "?" + "&".join(
                f"{key}={value}" for key, value in updates.items()
            )
            app["render_member_entry_button"]()
            self.assertEqual(len(fake_st.markup), 1)
            self.assertIn("member-login-fixed", fake_st.markup[0])
            self.assertIn("会员登录", fake_st.markup[0])
            self.assertIn("member=1", fake_st.markup[0])

            fake_st.markup.clear()
            app["current_member_is_admin"] = lambda: True
            app["render_member_entry_button"]()
            self.assertEqual(fake_st.markup, [])
        finally:
            app.update(original_functions)

    def test_02bad_navigation_slots_compact_when_backend_entry_is_hidden(self):
        app = self.app

        class FakeStreamlit:
            def __init__(self):
                self.markup = []

            def markdown(self, value, **kwargs):
                self.markup.append(value)

        fake_st = FakeStreamlit()
        original_functions = {
            name: app[name]
            for name in [
                "st",
                "current_member_is_admin",
                "current_member",
                "is_member_page_requested",
                "is_bom_page_requested",
                "is_no_match_admin_page_requested",
                "build_app_href",
            ]
        }
        try:
            app["st"] = fake_st
            app["current_member"] = lambda: {"username": "ordinary-member"}
            app["is_member_page_requested"] = lambda: False
            app["is_bom_page_requested"] = lambda: False
            app["is_no_match_admin_page_requested"] = lambda: False
            app["build_app_href"] = lambda **updates: "?" + "&".join(
                f"{key}={value}" for key, value in updates.items()
            )

            app["current_member_is_admin"] = lambda: False
            app["render_member_entry_button"]()
            app["render_bom_entry_button"]()
            self.assertIn("member-login-fixed active nav-slot-first", fake_st.markup[0])
            self.assertIn("bom-entry-fixed nav-slot-second", fake_st.markup[1])

            fake_st.markup.clear()
            app["current_member_is_admin"] = lambda: True
            app["render_member_entry_button"]()
            app["render_bom_entry_button"]()
            self.assertIn('class="member-login-fixed active"', fake_st.markup[0])
            self.assertIn('class="bom-entry-fixed"', fake_st.markup[1])
            self.assertNotIn("nav-slot-", "".join(fake_st.markup))
        finally:
            app.update(original_functions)

    def test_02baf_member_and_bom_navigation_preserve_session_token(self):
        app = self.app

        class FakeStreamlit:
            def __init__(self):
                self.session_state = {"_member_auth_token": "admin-token"}
                self.markup = []

            def markdown(self, value, **kwargs):
                self.markup.append(value)

        fake_st = FakeStreamlit()
        captured_updates = []
        original_functions = {
            name: app[name]
            for name in [
                "st",
                "current_member_is_admin",
                "current_member",
                "is_member_page_requested",
                "is_bom_page_requested",
                "is_no_match_admin_page_requested",
                "get_query_param_value",
                "build_app_href",
            ]
        }
        try:
            app["st"] = fake_st
            app["current_member"] = lambda: {"username": "admin", "_session_token": "admin-token"}
            app["current_member_is_admin"] = lambda: True
            app["is_member_page_requested"] = lambda: False
            app["is_bom_page_requested"] = lambda: False
            app["is_no_match_admin_page_requested"] = lambda: False
            app["get_query_param_value"] = lambda name: ""
            app["build_app_href"] = lambda **updates: captured_updates.append(updates) or (
                "?" + "&".join(f"{key}={value}" for key, value in updates.items())
            )

            app["render_member_entry_button"]()
            app["render_bom_entry_button"]()
            self.assertEqual(captured_updates[0][app["MEMBER_AUTH_QUERY_PARAM"]], "admin-token")
            self.assertEqual(captured_updates[0]["member"], "1")
            self.assertEqual(captured_updates[1][app["MEMBER_AUTH_QUERY_PARAM"]], "admin-token")
            self.assertEqual(captured_updates[1]["bom"], "1")

            captured_updates.clear()
            fake_st.markup.clear()
            app["is_member_page_requested"] = lambda: True
            app["render_member_entry_button"]()
            self.assertEqual(captured_updates[0][app["MEMBER_AUTH_QUERY_PARAM"]], "admin-token")
            self.assertEqual(captured_updates[0]["member"], "0")

            captured_updates.clear()
            fake_st.markup.clear()
            fake_st.session_state["_member_auth_token"] = "ordinary-token"
            app["current_member"] = lambda: {"username": "ordinary", "_session_token": "ordinary-token"}
            app["current_member_is_admin"] = lambda: False
            app["is_member_page_requested"] = lambda: False
            app["render_member_entry_button"]()
            self.assertEqual(captured_updates[0][app["MEMBER_AUTH_QUERY_PARAM"]], "ordinary-token")
            self.assertEqual(captured_updates[0]["member"], "1")
        finally:
            app.update(original_functions)

    def test_02bb_member_logout_clears_ui_and_revokes_session(self):
        app = self.app
        app["ensure_configured_admin_member_account"]()
        member, message = app["authenticate_member"]("TERRY46", "123456")
        self.assertIsNotNone(member, message)
        token = member["_session_token"]
        fake_st = type(
            "FakeStreamlit",
            (),
            {
                "session_state": {
                    "_member_auth_token": token,
                    "_member_display_name": "Admin",
                    "_no_match_admin_authenticated": True,
                    app["MEMBER_PENDING_SEARCH_QUERY_KEY"]: "0402 10K",
                }
            },
        )()
        original_functions = {
            name: app[name]
            for name in [
                "st",
                "get_query_param_value",
                "update_query_params",
                "refresh_member_auth_remote_snapshot",
                "flush_member_auth_remote_snapshot",
            ]
        }
        route_updates = []
        remote_flushes = []
        try:
            app["st"] = fake_st
            app["get_query_param_value"] = (
                lambda name: token if name == app["MEMBER_AUTH_QUERY_PARAM"] else ""
            )
            app["update_query_params"] = lambda **updates: route_updates.append(updates)
            app["refresh_member_auth_remote_snapshot"] = lambda force=False: "current"
            app["flush_member_auth_remote_snapshot"] = lambda: remote_flushes.append(True) or True

            self.assertTrue(app["logout_member"]())
            self.assertNotIn("_member_auth_token", fake_st.session_state)
            self.assertNotIn("_member_display_name", fake_st.session_state)
            self.assertNotIn("_no_match_admin_authenticated", fake_st.session_state)
            self.assertTrue(fake_st.session_state["_member_auth_clear_browser_token"])
            self.assertEqual(
                route_updates,
                [
                    {
                        app["MEMBER_AUTH_QUERY_PARAM"]: "",
                        "member": "",
                        "admin": "",
                        "bom": "",
                    }
                ],
            )
            with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
                remaining = conn.execute(
                    "SELECT COUNT(*) FROM member_sessions WHERE token=?", (token,)
                ).fetchone()[0]
            self.assertEqual(remaining, 0)
            self.assertEqual(remote_flushes, [True])
        finally:
            app.update(original_functions)

    def test_02bc_admin_logout_signs_out_member_and_returns_to_search(self):
        app = self.app
        app["ensure_configured_admin_member_account"]()
        member, message = app["authenticate_member"]("TERRY46", "123456")
        self.assertIsNotNone(member, message)
        member_token = member["_session_token"]
        fake_st = type(
            "FakeStreamlit",
            (),
            {
                "session_state": {
                    "_member_auth_token": member_token,
                    "_member_display_name": "Admin",
                    "_no_match_admin_authenticated": True,
                }
            },
        )()
        original_functions = {
            name: app[name]
            for name in [
                "st",
                "get_query_param_value",
                "update_query_params",
                "refresh_member_auth_remote_snapshot",
                "flush_member_auth_remote_snapshot",
            ]
        }
        route_updates = []
        remote_flushes = []
        try:
            app["st"] = fake_st
            app["get_query_param_value"] = (
                lambda name: member_token if name == app["MEMBER_AUTH_QUERY_PARAM"] else ""
            )
            app["update_query_params"] = lambda **updates: route_updates.append(updates)
            app["refresh_member_auth_remote_snapshot"] = lambda force=False: "current"
            app["flush_member_auth_remote_snapshot"] = lambda: remote_flushes.append(True) or True

            self.assertTrue(app["logout_no_match_admin"]())

            self.assertNotIn("_no_match_admin_authenticated", fake_st.session_state)
            self.assertNotIn("_member_auth_token", fake_st.session_state)
            self.assertNotIn("_member_display_name", fake_st.session_state)
            self.assertTrue(fake_st.session_state["_member_auth_clear_browser_token"])
            self.assertEqual(fake_st.session_state["_member_auth_clear_browser_token_value"], member_token)
            self.assertTrue(fake_st.session_state[app["ADMIN_ROUTE_CLEAR_OUTER_SHELL_KEY"]])
            self.assertEqual(
                route_updates,
                [
                    {
                        app["MEMBER_AUTH_QUERY_PARAM"]: "",
                        "member": "",
                        "admin": "",
                        "bom": "",
                    }
                ],
            )
            with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
                remaining = conn.execute(
                    "SELECT COUNT(*) FROM member_sessions WHERE token=?", (member_token,)
                ).fetchone()[0]
            self.assertEqual(remaining, 0)
            self.assertEqual(remote_flushes, [True])
        finally:
            app.update(original_functions)

    def test_02c_pending_search_resumes_once_after_login(self):
        app = self.app
        original_st = app["st"]
        original_current_member = app["current_member"]
        fake_st = type("FakeStreamlit", (), {"session_state": {}})()
        try:
            app["st"] = fake_st
            app["current_member"] = lambda: None
            app["remember_pending_member_search"](
                " 0402 1% 10K ",
                app["SEARCH_BRAND_MODE_CUSTOM"],
                ["FOJAN(富捷)", "华新科Walsin"],
            )
            self.assertEqual(app["resumable_member_search_query"](), "")
            self.assertEqual(app["resumable_member_search_brand_settings"](), ("", []))

            app["current_member"] = lambda: {"id": 7}
            self.assertEqual(app["resumable_member_search_query"](), "0402 1% 10K")
            self.assertEqual(
                app["resumable_member_search_brand_settings"](),
                (app["SEARCH_BRAND_MODE_CUSTOM"], ["FOJAN(富捷)", "华新科Walsin"]),
            )
            app["clear_pending_member_search"]()
            self.assertEqual(app["resumable_member_search_query"](), "")
            self.assertEqual(app["resumable_member_search_brand_settings"](), ("", []))
        finally:
            app["st"] = original_st
            app["current_member"] = original_current_member

        resolve_bom_resume = app["resolve_bom_post_login_resume_action"]
        self.assertEqual(resolve_bom_resume(True, "", True, True), "announce_restore")
        self.assertEqual(
            resolve_bom_resume(False, app["BOM_POST_LOGIN_STAGE_LOGIN_COMPLETE"], True, True),
            "announce_restore",
        )
        self.assertEqual(
            resolve_bom_resume(False, app["BOM_POST_LOGIN_STAGE_UPLOAD_RESTORED"], True, True),
            "resume",
        )
        self.assertEqual(resolve_bom_resume(True, "", True, False), "missing_upload")
        self.assertEqual(resolve_bom_resume(True, "", False, True), "")
        auto_resume_ready = app["bom_post_login_auto_resume_ready"]
        restored_stage = app["BOM_POST_LOGIN_STAGE_UPLOAD_RESTORED"]
        self.assertFalse(auto_resume_ready(restored_stage, 101.0, now=100.0))
        self.assertTrue(auto_resume_ready(restored_stage, 101.0, now=101.0))
        self.assertTrue(auto_resume_ready(restored_stage, 999.0, now=100.0, manual_start=True))
        self.assertFalse(auto_resume_ready("login_complete", 0, now=100.0))

        original_resume_values = {
            name: app[name]
            for name in [
                "st",
                "current_member",
                "render_bom_progress_card",
                "render_bom_post_login_auto_resume_control",
            ]
        }
        transition_calls = []

        class StopAfterSuccessPage(Exception):
            pass

        fake_resume_st = type(
            "FakeResumeStreamlit",
            (),
            {
                "session_state": {
                    app["BOM_PENDING_UPLOAD_WAITING_LOGIN_KEY"]: True,
                    app["BOM_POST_LOGIN_RESUME_STAGE_KEY"]: app["BOM_POST_LOGIN_STAGE_LOGIN_COMPLETE"],
                },
                "empty": staticmethod(lambda: object()),
                "stop": staticmethod(lambda: (_ for _ in ()).throw(StopAfterSuccessPage())),
            },
        )()
        try:
            app["st"] = fake_resume_st
            app["current_member"] = lambda: {"id": 7}
            app["render_bom_progress_card"] = lambda *_args, **_kwargs: transition_calls.append("success")
            app["render_bom_post_login_auto_resume_control"] = lambda: transition_calls.append("auto_resume")
            with self.assertRaises(StopAfterSuccessPage):
                app["render_bom_post_login_resume_transition"](
                    type("RestoredUpload", (), {"name": "restored.xlsx"})()
                )
            self.assertEqual(transition_calls, ["success", "auto_resume"])
            self.assertEqual(
                fake_resume_st.session_state[app["BOM_POST_LOGIN_RESUME_STAGE_KEY"]],
                restored_stage,
            )
            self.assertGreater(
                float(fake_resume_st.session_state[app["BOM_POST_LOGIN_AUTO_RESUME_AT_KEY"]]),
                0,
            )
        finally:
            app.update(original_resume_values)

    def test_02d_compact_search_summary_and_read_only_runtime_snapshot(self):
        app = self.app
        progress_state = app["build_search_progress_state"](
            total_queries=3,
            completed_queries=3,
            stage_text="搜索已完成",
            elapsed_seconds=1.25,
            done=True,
            extra_chips=[{"label": "有结果", "value": "2", "tone": "success"}],
            summary_lines=["已返回可查看结果 2 条", "未找到匹配结果 1 条"],
        )
        summary_html = app["build_search_progress_summary_html"](progress_state)
        self.assertIn('class="search-progress-summary"', summary_html)
        self.assertIn("处理 3/3", summary_html)
        self.assertIn("有结果 2", summary_html)
        self.assertNotIn("bom-progress-track", summary_html)

        original_values = {
            name: app[name]
            for name in [
                "DB_PATH",
                "SEARCH_DB_PATH",
                "STREAMLIT_CLOUD_BUNDLE_MANIFEST_PATH",
                "MEMBER_AUTH_REMOTE_STATE_PATH",
                "RUNTIME_STORE_REMOTE_STATE_DIR",
            ]
        }
        original_release = os.environ.get("COMPONENT_MATCHER_RELEASE_STAMP")
        try:
            main_db = os.path.join(self.temp_dir, "runtime-main.sqlite")
            search_db = os.path.join(self.temp_dir, "runtime-search.sqlite")
            manifest_path = os.path.join(self.temp_dir, "runtime-manifest.json")
            with sqlite3.connect(main_db) as conn:
                conn.execute("CREATE TABLE components (id INTEGER PRIMARY KEY)")
                conn.executemany("INSERT INTO components (id) VALUES (?)", [(1,), (2,), (3,)])
            with sqlite3.connect(search_db) as conn:
                conn.execute("CREATE TABLE search_meta (meta_json TEXT NOT NULL)")
                conn.execute(
                    "INSERT INTO search_meta (meta_json) VALUES (?)",
                    (json.dumps({"table_row_counts": {"components_search_core": 321}}),),
                )
            with open(manifest_path, "w", encoding="utf-8") as handle:
                json.dump(
                    {
                        "build_epoch_ns": 1_700_000_000_000_000_000,
                        "members": [
                            {
                                "path": "cache/components_search.sqlite",
                                "size": 100,
                                "mtime_ns": 1,
                                "sha256": "abcdef1234567890",
                            }
                        ],
                    },
                    handle,
                )
            app["DB_PATH"] = main_db
            app["SEARCH_DB_PATH"] = search_db
            app["STREAMLIT_CLOUD_BUNDLE_MANIFEST_PATH"] = manifest_path
            app["MEMBER_AUTH_REMOTE_STATE_PATH"] = os.path.join(self.temp_dir, "runtime-member-state.json")
            app["RUNTIME_STORE_REMOTE_STATE_DIR"] = self.temp_dir
            os.environ["COMPONENT_MATCHER_RELEASE_STAMP"] = "2026-07-11T18:12:06+08:00"
            snapshot = app["build_runtime_status_snapshot"]()
            self.assertEqual(snapshot["component_rows"], 3)
            self.assertEqual(snapshot["search_rows"], 321)
            self.assertEqual(snapshot["database_version"], "abcdef123456")
            self.assertEqual(snapshot["release_stamp"], "2026-07-11T18:12:06+08:00")
        finally:
            app.update(original_values)
            if original_release is None:
                os.environ.pop("COMPONENT_MATCHER_RELEASE_STAMP", None)
            else:
                os.environ["COMPONENT_MATCHER_RELEASE_STAMP"] = original_release

    def test_02e_result_iframe_shrinks_to_actual_content(self):
        app = self.app
        estimate = app["estimate_result_table_iframe_height"]
        self.assertEqual(estimate(0), 150)
        self.assertEqual(estimate(8), 444)
        self.assertEqual(estimate(12), 444)
        self.assertEqual(estimate(100), 444)
        self.assertLessEqual(estimate(100), 460)

        match_estimate = app["estimate_match_card_iframe_height"]
        self.assertEqual(match_estimate(1, 1), 420)
        self.assertEqual(match_estimate(1, 100), 570)

        preview_estimate = app["estimate_bom_preview_iframe_height"]
        self.assertEqual(preview_estimate(0), 220)
        self.assertEqual(preview_estimate(5), 356)
        self.assertEqual(preview_estimate(100), 460)
        self.assertEqual(preview_estimate(100, compact=True), 280)

        iframe_html = app["build_result_table_iframe_html"](
            '<div class="result-section-card"><div class="result-table-wrap">'
            '<table class="result-table"><tbody><tr><td>row</td></tr></tbody></table>'
            "</div></div>"
        )
        self.assertIn("measureFrameContentHeight", iframe_html)
        self.assertIn("alignScrollableTableHeight", iframe_html)
        self.assertIn("visibleRowLimit = 8", iframe_html)
        self.assertIn("horizontalScrollbarReserve", iframe_html)
        self.assertIn("wrapper.offsetHeight - wrapper.clientHeight", iframe_html)
        self.assertIn("frameBottomReserve = 16", iframe_html)
        self.assertIn("overflow: hidden", iframe_html)
        self.assertIn("max-height: 440px", iframe_html)
        self.assertIn(".bom-preview-table-wrap", iframe_html)
        self.assertIn(".bom-preview-table-wrap-compact", iframe_html)
        self.assertNotIn("52vh", iframe_html)
        self.assertNotIn("document.documentElement.scrollHeight", iframe_html)

        preview_html = app["render_static_preview_table"](
            pd.DataFrame({"规格": [f"row-{index}" for index in range(20)]}),
            wrapper_class="bom-preview-table-wrap",
        )
        self.assertIn('class="bom-preview-table-wrap"', preview_html)

    def test_03_resistor_value_size_and_power_guards(self):
        app = self.app
        milliohm = app["parse_resistor_spec_query"]("1206 0.01R 1% 1/4W")
        megaohm = app["parse_resistor_spec_query"]("0402 1M 5% 1/16W")
        self.assertAlmostEqual(float(milliohm["_resistance_ohm"]), 0.01)
        self.assertAlmostEqual(float(megaohm["_resistance_ohm"]), 1_000_000.0)

        unlabeled_bom_specs = [
            ("0,50mW Resistor R_0201 1%", 0.0, "FRC0201F0000TS"),
            ("150,50mW Resistor R_0201 1%", 150.0, "FRC0201F1500TS"),
        ]
        for query, expected_ohm, expected_fojan_model in unlabeled_bom_specs:
            parsed = app["parse_resistor_spec_query"](query)
            self.assertIsNotNone(parsed, query)
            self.assertEqual(parsed["器件类型"], "贴片电阻", query)
            self.assertEqual(parsed["尺寸（inch）"], "0201", query)
            self.assertEqual(parsed["_power"], "1/20W", query)
            self.assertEqual(app["clean_tol_for_match"](parsed["容值误差"]), "1", query)
            self.assertEqual(parsed["_param_count"], 4, query)
            self.assertAlmostEqual(float(parsed["_resistance_ohm"]), expected_ohm, msg=query)
            self.assertEqual(app["build_fojan_resistor_model_from_spec"](parsed), expected_fojan_model, query)
            mode, detected = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            self.assertEqual(mode, "贴片电阻", query)
            self.assertAlmostEqual(float(detected["_resistance_ohm"]), expected_ohm, msg=query)
            resolved = app["resolve_search_query_dataframe_and_spec"](query)
            self.assertNotEqual(resolved["resolution_path"], "full_dataframe", query)
            self.assertFalse(resolved["query_df"].empty, query)
            matched = app["run_query_match"](resolved["query_df"], resolved["mode"], resolved["spec"])
            self.assertIn(
                expected_fojan_model,
                set(matched["型号"].astype(str).map(app["clean_model"])),
                query,
            )

        self.assertIsNone(
            app["find_leading_unlabeled_resistance_in_resistor_text"]("150,50mW Capacitor C_0201 1%")
        )
        self.assertIsNone(
            app["find_leading_unlabeled_resistance_in_resistor_text"]("0201 Resistor 50mW 1%")
        )

        complete_resistor_shorthand_specs = [
            ("5.6M +1% 0805 1/8W", 5_600_000.0, "0805", "1", "1/8W", "FRC0805F5604TS"),
            ("47R 1% :2512 2W", 47.0, "2512", "1", "2W", ""),
            ("62 1% 0603 1/10W", 62.0, "0603", "1", "1/10W", "FRC0603F62R0TS"),
            ("2M 1% 1/8W 0805", 2_000_000.0, "0805", "1", "1/8W", "FRC0805F2004TS"),
            ("3.9K 5% 0805 1/8W", 3_900.0, "0805", "5", "1/8W", "FRC0805J392 TS"),
        ]
        for query, expected_ohm, expected_size, expected_tol, expected_power, expected_model in complete_resistor_shorthand_specs:
            parsed = app["parse_resistor_spec_query"](query)
            self.assertIsNotNone(parsed, query)
            self.assertEqual(parsed["器件类型"], "贴片电阻", query)
            self.assertEqual(parsed["尺寸（inch）"], expected_size, query)
            self.assertEqual(app["clean_tol_for_match"](parsed["容值误差"]), expected_tol, query)
            self.assertEqual(parsed["_power"], expected_power, query)
            self.assertEqual(parsed["_param_count"], 4, query)
            self.assertAlmostEqual(float(parsed["_resistance_ohm"]), expected_ohm, msg=query)
            self.assertEqual(app["build_fojan_resistor_model_from_spec"](parsed), expected_model, query)
            mode, detected = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            self.assertEqual(mode, "贴片电阻", query)
            self.assertAlmostEqual(float(detected["_resistance_ohm"]), expected_ohm, msg=query)
            resolved = app["resolve_search_query_dataframe_and_spec"](query)
            self.assertNotEqual(resolved["resolution_path"], "full_dataframe", query)
            self.assertFalse(resolved["query_df"].empty, query)
            matched = app["run_query_match"](
                resolved["query_df"], resolved["mode"], resolved["spec"]
            )
            self.assertFalse(matched.empty, query)
            if expected_model:
                self.assertIn(
                    app["clean_model"](expected_model),
                    set(matched["型号"].astype(str).map(app["clean_model"])),
                    query,
                )
            if query == "47R 1% :2512 2W":
                self.assertTrue(
                    matched["功率"].astype(str).map(app["format_power_display"]).eq("2W").all(),
                    query,
                )

        self.assertIsNone(
            app["find_leading_plain_resistance_in_complete_resistor_spec"]("0603 1% 1/10W")
        )
        self.assertIsNone(
            app["find_leading_plain_resistance_in_complete_resistor_spec"]("62pF 1% 0603 1/10W")
        )

        image_resistor_specs = [
            ("1M,62.5mW Resistor R_0402 1%", 1_000_000.0, "0402", "1/16W"),
            ("4.02k,62.5mW Resistor R_0402 1%", 4_020.0, "0402", "1/16W"),
            ("4.7k,50mW Resistor R_0201 1%", 4_700.0, "0201", "1/20W"),
            ("4.8k,62.5mW Resistor R_0402 1%", 4_800.0, "0402", "1/16W"),
            ("10k,50mW Resistor R_0201 1%", 10_000.0, "0201", "1/20W"),
            ("10k,62.5mW Resistor R_0402 1%", 10_000.0, "0402", "1/16W"),
            ("12k,62.5mW Resistor R_0402 1%", 12_000.0, "0402", "1/16W"),
            ("15k,62.5mW Resistor R_0402 1%", 15_000.0, "0402", "1/16W"),
            ("18k,62.5mW Resistor R_0402 1%", 18_000.0, "0402", "1/16W"),
            ("20k,62.5mW Resistor R_0402 1%", 20_000.0, "0402", "1/16W"),
            ("30k,62.5mW Resistor R_0402 1%", 30_000.0, "0402", "1/16W"),
            ("60.4k,50mW Resistor R_0201 1%", 60_400.0, "0201", "1/20W"),
            ("60.4k,62.5mW Resistor R_0402 1%", 60_400.0, "0402", "1/16W"),
            ("100k,62.5mW Resistor R_0402 1%", 100_000.0, "0402", "1/16W"),
        ]
        for query, expected_ohm, expected_size, expected_power in image_resistor_specs:
            mode, parsed = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            self.assertEqual(mode, "贴片电阻", query)
            self.assertEqual(parsed["尺寸（inch）"], expected_size, query)
            self.assertEqual(parsed["_power"], expected_power, query)
            self.assertEqual(parsed["_param_count"], 4, query)
            self.assertAlmostEqual(float(parsed["_resistance_ohm"]), expected_ohm, msg=query)

        image_capacitor_specs = [
            ("1u,10V Capacitor C_0402 TCC0402X5R105K100AT", 1_000_000.0, "0402", "10"),
            ("2.2u,6.3V Capacitor C_0402 TCC0402X5R225M6R3AT", 2_200_000.0, "0402", "6.3"),
            ("10n,25V Capacitor C_0201 TCC0201X7R103K250ZT", 10_000.0, "0201", "25"),
            ("12p,50V Capacitor C_0201 TCC0201C0G120J500ZT", 12.0, "0201", "50"),
            ("22u,10V Capacitor C_0603 TCC0603X5R226M100CT", 22_000_000.0, "0603", "10"),
            ("100n,10V Capacitor C_0201 TCC0201X5R104K100ZT", 100_000.0, "0201", "10"),
            ("100n,16V Capacitor C_0402 0402B104K160CT", 100_000.0, "0402", "16"),
        ]
        for query, expected_pf, expected_size, expected_voltage in image_capacitor_specs:
            mode, parsed = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            self.assertEqual(mode, "规格", query)
            self.assertEqual(parsed["尺寸（inch）"], expected_size, query)
            self.assertEqual(app["clean_voltage"](parsed["耐压（V）"]), expected_voltage, query)
            self.assertGreaterEqual(parsed["_param_count"], 3, query)
            self.assertAlmostEqual(float(parsed["容值_pf"]), expected_pf, msg=query)

        thermistor_rows, thermistor_tokens, thermistor_match = app["load_component_rows_by_query_model_tokens"](
            "NCP03WF104F05RL Resistor 0201 无"
        )
        self.assertIn("NCP03WF104F05RL", thermistor_tokens)
        self.assertEqual(thermistor_match, "NCP03WF104F05RL")
        self.assertFalse(thermistor_rows.empty)
        self.assertEqual(set(thermistor_rows["器件类型"].map(app["normalize_component_type"])), {"热敏电阻"})

        slash_specs = [
            ("贴片\\1.24K\\±1%\\1/16W\\0402 ROHS", 1_240.0, "1"),
            ("贴片\\499R\\±1%\\1/16W\\0402 ROHS", 499.0, "1"),
            ("贴片\\499K\\±1%\\1/16W\\0402 ROHS", 499_000.0, "1"),
            ("贴片\\51R\\±5%\\1/16W\\0402 ROHS", 51.0, "5"),
        ]
        for query, expected_ohm, expected_tol in slash_specs:
            parsed = app["parse_resistor_spec_query"](query)
            self.assertIsNotNone(parsed, query)
            self.assertEqual(parsed["器件类型"], "贴片电阻", query)
            self.assertEqual(parsed["尺寸（inch）"], "0402", query)
            self.assertEqual(parsed["_power"], "1/16W", query)
            self.assertEqual(app["clean_tol_for_match"](parsed["容值误差"]), expected_tol, query)
            self.assertAlmostEqual(float(parsed["_resistance_ohm"]), expected_ohm, msg=query)
            mode, detected = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            self.assertEqual(mode, "贴片电阻", query)
            self.assertEqual(detected["器件类型"], "贴片电阻", query)
            self.assertIsNone(detected.get("容值_pf"), query)

        compact_1206_specs = [
            ("1206,3R,1%", 3.0, "1"),
            ("1206,3R.5%", 3.0, "5"),
            ("1206,3.3R,1%", 3.3, "1"),
            ("1206,3.3R,5%", 3.3, "5"),
            ("1206,4R,  1%", 4.0, "1"),
            ("1206,4R,  5%", 4.0, "5"),
        ]
        for query, expected_ohm, expected_tol in compact_1206_specs:
            parsed = app["parse_resistor_spec_query"](query)
            self.assertIsNotNone(parsed, query)
            self.assertEqual(parsed["尺寸（inch）"], "1206", query)
            self.assertEqual(app["clean_tol_for_match"](parsed["容值误差"]), expected_tol, query)
            self.assertAlmostEqual(float(parsed["_resistance_ohm"]), expected_ohm, msg=query)
            mode, detected = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            self.assertEqual(mode, "贴片电阻", query)
            resolved = app["resolve_search_query_dataframe_and_spec"](query)
            self.assertNotEqual(resolved["resolution_path"], "full_dataframe", query)
            self.assertFalse(resolved["query_df"].empty, query)

        self.assertEqual(
            app["normalize_resistor_value_tolerance_separator"]("1206,3R.5%"),
            "1206,3R,5%",
        )
        self.assertEqual(
            app["normalize_resistor_value_tolerance_separator"]("1206,3.3R,1%"),
            "1206,3.3R,1%",
        )

        for query in ["2010 100K士1%", "2010 100K土1%", "2010 100K士1％", "2010 100K±1%"]:
            parsed = app["parse_resistor_spec_query"](query)
            self.assertIsNotNone(parsed, query)
            self.assertEqual(parsed["尺寸（inch）"], "2010", query)
            self.assertEqual(parsed["_param_count"], 3, query)
            self.assertEqual(app["clean_tol_for_match"](parsed["容值误差"]), "1", query)
            self.assertAlmostEqual(float(parsed["_resistance_ohm"]), 100_000.0, msg=query)
            mode, detected = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            self.assertEqual(mode, "贴片电阻", query)
            self.assertEqual(detected["器件类型"], "贴片电阻", query)
        self.assertEqual(app["normalize_common_tolerance_symbol_typos"]("勇士1%"), "勇士1%")

        code_105 = app["parse_resistor_model_rule"]("FRC0402J105 TS", brand="FOJAN(富捷)")
        code_106 = app["parse_resistor_model_rule"]("FRC0402J106 TS", brand="FOJAN(富捷)")
        self.assertAlmostEqual(float(code_105["_resistance_ohm"]), 1_000_000.0)
        self.assertAlmostEqual(float(code_106["_resistance_ohm"]), 10_000_000.0)

        hkr_examples = [
            ("RCA031MFLF", "0603", 1_000_000.0, "1", "1/10W", "75", "5000PCS"),
            ("RCA0520KFLF", "0805", 20_000.0, "1", "1/8W", "150", "5000PCS"),
            ("RCA022R2JLF", "0402", 2.2, "5", "1/16W", "50", "10000PCS"),
        ]
        for model, size, resistance, tolerance, power, voltage, moq in hkr_examples:
            parsed_hkr = app["parse_resistor_model_rule"](model)
            self.assertIsNotNone(parsed_hkr, model)
            self.assertEqual(parsed_hkr["品牌"], "香港电阻HKR", model)
            self.assertEqual(parsed_hkr["系列"], "RCA", model)
            self.assertEqual(parsed_hkr["尺寸（inch）"], size, model)
            self.assertAlmostEqual(float(parsed_hkr["_resistance_ohm"]), resistance, msg=model)
            self.assertEqual(app["clean_tol_for_match"](parsed_hkr["容值误差"]), tolerance, model)
            self.assertEqual(parsed_hkr["功率"], power, model)
            self.assertEqual(parsed_hkr["耐压（V）"], voltage, model)
            self.assertEqual(parsed_hkr["MOQ"], moq, model)
            self.assertEqual(parsed_hkr["_model_rule_authority"], "hkr_rca_official_model", model)

        self.assertEqual(app["extract_requested_brand_from_query"]("香港电阻 RCA031MFLF"), "香港电阻HKR")
        resolved_hkr = app["resolve_search_query_dataframe_and_spec"]("RCA031MFLF")
        self.assertNotEqual(resolved_hkr["resolution_path"], "full_dataframe")
        self.assertFalse(resolved_hkr["query_df"].empty)
        resolved_hkr_spec = resolved_hkr["spec"]
        self.assertEqual(resolved_hkr_spec["品牌"], "香港电阻HKR")
        self.assertEqual(resolved_hkr_spec["尺寸（inch）"], "0603")
        self.assertAlmostEqual(float(resolved_hkr_spec["_resistance_ohm"]), 1_000_000.0)

        invalid = app["parse_resistor_spec_query"]("0420 10K 1% 1/16W")
        self.assertTrue(invalid.get("_unsupported_component"))
        self.assertEqual(invalid.get("_invalid_size_token"), "0420")
        resolved = app["resolve_search_query_dataframe_and_spec"]("0420 10K 1% 1/16W")
        self.assertEqual(resolved["resolution_path"], "unsupported_or_invalid_spec")
        self.assertTrue(resolved["query_df"].empty)

        candidates = pd.DataFrame(
            [
                {
                    "品牌": "A",
                    "型号": "R-1-16",
                    "器件类型": "贴片电阻",
                    "尺寸（inch）": "0603",
                    "材质（介质）": "",
                    "耐压（V）": "",
                    "容值_pf": None,
                    "容值": "10",
                    "容值单位": "KΩ",
                    "参数值": "10",
                    "参数单位": "KΩ",
                    "_resistance_ohm": 10000.0,
                    "容值误差": "1",
                    "功率": "1/16W",
                },
                {
                    "品牌": "B",
                    "型号": "R-1-10",
                    "器件类型": "贴片电阻",
                    "尺寸（inch）": "0603",
                    "材质（介质）": "",
                    "耐压（V）": "",
                    "容值_pf": None,
                    "容值": "10",
                    "容值单位": "KΩ",
                    "参数值": "10",
                    "参数单位": "KΩ",
                    "_resistance_ohm": 10000.0,
                    "容值误差": "1",
                    "功率": "1/10W",
                },
            ]
        )
        prepared = app["prepare_search_dataframe"](app["ensure_component_display_columns"](candidates))
        spec = app["parse_resistor_spec_query"]("0603 10K 1% 1/16W")
        original_fetch = app["fetch_search_candidate_pairs"]
        app["fetch_search_candidate_pairs"] = lambda _spec: None
        try:
            matched = app["run_query_match"](prepared, "规格", spec)
        finally:
            app["fetch_search_candidate_pairs"] = original_fetch
        self.assertEqual(set(matched["型号"]), {"R-1-16"})

        brand_candidates = pd.DataFrame(
            [
                {
                    "品牌": "FOJAN(富捷)",
                    "型号": "FRC0402F1002TS",
                    "器件类型": "贴片电阻",
                    "尺寸（inch）": "0402",
                    "材质（介质）": "",
                    "耐压（V）": "",
                    "容值_pf": None,
                    "_resistance_ohm": 10000.0,
                    "容值": "10",
                    "容值单位": "KΩ",
                    "容值误差": "1",
                    "功率": "1/16W",
                },
                {
                    "品牌": "华新科Walsin",
                    "型号": "WR04X1002FTL",
                    "器件类型": "贴片电阻",
                    "尺寸（inch）": "0402",
                    "材质（介质）": "",
                    "耐压（V）": "",
                    "容值_pf": None,
                    "_resistance_ohm": 10000.0,
                    "容值": "10",
                    "容值单位": "KΩ",
                    "容值误差": "1",
                    "功率": "1/16W",
                },
            ]
        )
        prepared_brand = app["prepare_search_dataframe"](app["ensure_component_display_columns"](brand_candidates))
        original_fetch = app["fetch_search_candidate_pairs"]
        app["fetch_search_candidate_pairs"] = lambda _spec: None
        try:
            no_brand = app["run_query_match"](
                prepared_brand,
                "贴片电阻",
                app["parse_resistor_spec_query"]("0402 1% 10K"),
            )
            self.assertEqual(set(no_brand["型号"]), {"FRC0402F1002TS", "WR04X1002FTL"})
            self.assertEqual(no_brand.iloc[0]["品牌"], "FOJAN(富捷)")
            self.assertEqual(app["brand_priority_value"]("FOJAN(富捷)", "贴片电阻"), 1)
            self.assertEqual(app["brand_priority_value"]("信昌PDC", "贴片电阻"), 2)
            self.assertEqual(app["brand_priority_value"]("华新科Walsin", "贴片电阻"), 3)
            self.assertEqual(app["brand_priority_value"]("厚声UNI-ROYAL", "贴片电阻"), 4)

            fojan_source_spec = app["parse_resistor_spec_query"]("0402 1% 10K")
            fojan_source_spec.update({"品牌": "FOJAN(富捷)", "型号": "FRC0402F1002TS"})
            fojan_source_matches = app["run_query_match"](prepared_brand, "料号", fojan_source_spec)
            self.assertEqual(fojan_source_matches.iloc[0]["品牌"], "FOJAN(富捷)")
            self.assertEqual(fojan_source_matches.iloc[0]["型号"], "FRC0402F1002TS")

            walsin_source_spec = app["parse_resistor_spec_query"]("0402 1% 10K")
            walsin_source_spec.update({"品牌": "华新科Walsin", "型号": "WR04X1002FTL"})
            walsin_source_matches = app["run_query_match"](prepared_brand, "料号", walsin_source_spec)
            self.assertEqual(set(walsin_source_matches["品牌"]), {"FOJAN(富捷)"})

            for query in ("富捷 0402 1% 10K", "0402 1% 10K 富捷", "FOJAN 0402 1% 10K"):
                mode, brand_spec = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
                self.assertEqual(mode, "贴片电阻", query)
                self.assertEqual(brand_spec["品牌"], "FOJAN(富捷)", query)
                self.assertFalse(brand_spec.get(app["BRAND_QUERY_FILTER_FLAG"], False), query)
                matched_brand = app["run_query_match"](prepared_brand, mode, brand_spec)
                self.assertEqual(
                    set(matched_brand["品牌"]),
                    {"FOJAN(富捷)", "华新科Walsin"},
                    query,
                )
                custom_brand_spec = app["apply_search_brand_scope_to_spec"](
                    brand_spec,
                    query,
                    app["SEARCH_BRAND_MODE_CUSTOM"],
                    ["FOJAN(富捷)"],
                )
                matched_custom_brand = app["run_query_match"](
                    prepared_brand, mode, custom_brand_spec
                )
                self.assertEqual(set(matched_custom_brand["品牌"]), {"FOJAN(富捷)"}, query)
                self.assertEqual(set(matched_custom_brand["型号"]), {"FRC0402F1002TS"}, query)

            fenghua_query = "10KΩ;75V;±1%;1/10W;0603;FENGHUA;RS-03K1002FT;无卤"
            fenghua_spec = app["parse_resistor_spec_query"]("0603 10K 1% 1/10W 75V 无卤")
            fenghua_spec.update({"品牌": "风华Fenghua", "型号": "RS-03K1002FT"})
            brand_filtered_part_spec = app["apply_query_brand_hint_to_spec"](fenghua_spec, fenghua_query)
            self.assertFalse(brand_filtered_part_spec.get(app["BRAND_QUERY_FILTER_FLAG"], False))
            part_lookup_spec = app["clear_query_brand_filter_for_part_lookup"](brand_filtered_part_spec)
            self.assertFalse(part_lookup_spec.get(app["BRAND_QUERY_FILTER_FLAG"], False))
            self.assertEqual(part_lookup_spec["品牌"], "风华Fenghua")

            embedded_part_spec = app["clear_query_brand_filter_for_embedded_part_metadata"](
                brand_filtered_part_spec,
                fenghua_query,
            )
            self.assertFalse(embedded_part_spec.get(app["BRAND_QUERY_FILTER_FLAG"], False))
            direct_brand_spec = app["apply_query_brand_hint_to_spec"](
                app["parse_resistor_spec_query"]("0603 10K 1%"),
                "FENGHUA 0603 10K 1%",
            )
            direct_brand_spec = app["clear_query_brand_filter_for_embedded_part_metadata"](
                direct_brand_spec,
                "FENGHUA 0603 10K 1%",
            )
            self.assertFalse(direct_brand_spec.get(app["BRAND_QUERY_FILTER_FLAG"], False))

            explicit_part_cases = [
                (
                    "1KΩ;50V;±1%;1/16W;0402;FENGHUA;RC-02K1001FT;无卤",
                    "RC-02K1001FT",
                    "FRC0402F1001TS",
                ),
                (
                    "10KΩ;/;±5%;1/16W;0402;FENGHUA;RC-02K103JT;无卤",
                    "RC-02K103JT",
                    "FRC0402J103 TS",
                ),
                (
                    "510Ω;±5%;1/16W;0402;FENGHUA;RC-02K511JT;无卤",
                    "RC-02K511JT",
                    "FRC0402J511 TS",
                ),
            ]
            for query, source_model, expected_fojan_model in explicit_part_cases:
                explicit_spec = app["parse_resistor_spec_query"](query)
                decoded_part_spec = dict(explicit_spec)
                decoded_part_spec.update(
                    {
                        "品牌": "风华Fenghua",
                        "型号": source_model,
                        "容值误差": "0.25",
                    }
                )
                merged_part_spec = app["merge_explicit_query_spec_into_part_spec"](
                    decoded_part_spec,
                    explicit_spec,
                )
                self.assertEqual(
                    app["build_fojan_resistor_model_from_spec"](merged_part_spec),
                    expected_fojan_model,
                    query,
                )

            low_ohm_candidate = app["finalize_search_candidate_frames"](
                [
                    pd.DataFrame(
                        [
                            {
                                "品牌": "FOJAN(富捷)",
                                "型号": "FRL1206FR050TS",
                                "器件类型": "厚膜电阻",
                                "尺寸（inch）": "1206",
                                "材质（介质）": "",
                                "耐压（V）": "",
                                "容值_pf": None,
                                "容值": "0.05",
                                "容值单位": "Ω",
                                "容值误差": "1",
                                "功率": "1/4W",
                                "_resistance_ohm": 0.05,
                                "特殊用途": "",
                            }
                        ]
                    )
                ]
            )
            self.assertEqual(low_ohm_candidate.iloc[0]["特殊用途"], "无卤")
            low_ohm_matches = app["run_query_match"](
                app["prepare_search_dataframe"](
                    app["ensure_component_display_columns"](low_ohm_candidate)
                ),
                "贴片电阻",
                app["parse_resistor_spec_query"]("1206 0.05Ω 1% 1/4W 无卤"),
            )
            self.assertEqual(low_ohm_matches["型号"].tolist(), ["FRL1206FR050TS"])

            duplicate_matches = app["deduplicate_component_matches"](
                pd.DataFrame(
                    [
                        {"品牌": "FOJAN(富捷)", "型号": "FRC0402F1001TS", "器件类型": "厚膜电阻"},
                        {"品牌": "FOJAN(富捷)", "型号": "FRC0402F1001RS", "器件类型": "贴片电阻"},
                        {"品牌": "FOJAN(富捷)", "型号": "FRC0402J563 TS", "器件类型": "厚膜电阻"},
                        {"品牌": "FOJAN(富捷)", "型号": "FRC0402J563RS", "器件类型": "贴片电阻"},
                    ]
                )
            )
            self.assertEqual(len(duplicate_matches), 2)
            self.assertEqual(
                duplicate_matches["型号"].tolist(),
                ["FRC0402F1001TS", "FRC0402J563 TS"],
            )

            cross_brand_candidates = pd.DataFrame(
                [
                    {
                        "品牌": "FOJAN(富捷)",
                        "型号": "FRC0603F1002TS",
                        "器件类型": "厚膜电阻",
                        "尺寸（inch）": "0603",
                        "材质（介质）": "",
                        "耐压（V）": "75",
                        "容值_pf": None,
                        "_resistance_ohm": 10000.0,
                        "容值": "10",
                        "容值单位": "KΩ",
                        "容值误差": "1",
                        "功率": "1/10W",
                        "特殊用途": "无卤",
                    },
                    {
                        "品牌": "风华Fenghua",
                        "型号": "RS-03K1002FT",
                        "器件类型": "厚膜电阻",
                        "尺寸（inch）": "0603",
                        "材质（介质）": "",
                        "耐压（V）": "75",
                        "容值_pf": None,
                        "_resistance_ohm": 10000.0,
                        "容值": "10",
                        "容值单位": "KΩ",
                        "容值误差": "1",
                        "功率": "1/10W",
                        "特殊用途": "无卤",
                    },
                ]
            )
            prepared_cross_brand = app["prepare_search_dataframe"](
                app["ensure_component_display_columns"](cross_brand_candidates)
            )
            cross_brand_matches = app["run_query_match"](prepared_cross_brand, "料号", part_lookup_spec)
            self.assertEqual(cross_brand_matches["型号"].tolist(), ["FRC0603F1002TS"])
            self.assertEqual(cross_brand_matches.iloc[0]["推荐等级"], "完全匹配")

            suffix_candidates = pd.DataFrame(
                [
                    {
                        "品牌": "FOJAN(富捷)",
                        "型号": model,
                        "器件类型": "厚膜电阻",
                        "尺寸（inch）": "0402",
                        "材质（介质）": "",
                        "耐压（V）": "50",
                        "容值_pf": None,
                        "_resistance_ohm": 1000.0,
                        "容值": "1",
                        "容值单位": "KΩ",
                        "容值误差": "1",
                        "功率": "1/16W",
                    }
                    for model in ["FRC0402F1001RS", "FRC0402F1001TS"]
                ]
            )
            prepared_suffix = app["prepare_search_dataframe"](
                app["ensure_component_display_columns"](suffix_candidates)
            )
            suffix_matches = app["apply_match_levels_and_sort"](
                prepared_suffix,
                app["parse_resistor_spec_query"]("0402 1K 1% 1/16W"),
            )
            self.assertEqual(
                suffix_matches["型号"].tolist(),
                ["FRC0402F1001TS", "FRC0402F1001RS"],
            )
            car_spec = app["parse_resistor_spec_query"]("0402 4.99Ω ±1% 1/16W 车规")
            car_spec["特殊用途"] = "车规"
            car_candidates = app["build_fojan_special_resistor_candidates_from_spec"](car_spec)
            car_matches = app["apply_match_levels_and_sort"](car_candidates, car_spec)
            self.assertTrue(car_matches.iloc[0]["型号"].startswith("FRQ0402F4R99"))
            self.assertEqual(app["special_use_specificity_rank"]("车规", "车规"), 0)
            self.assertGreater(app["special_use_specificity_rank"]("低阻/车规", "车规"), 0)

            reversed_car_query = car_matches.iloc[:2].iloc[::-1].copy()
            ordered_car_export = app["build_bom_export_candidate_frame"](
                car_matches.iloc[:2].copy(),
                query_df=reversed_car_query,
                spec=car_spec,
                mode="厚膜电阻",
            )
            self.assertEqual(
                ordered_car_export.iloc[0]["型号"],
                car_matches.iloc[0]["型号"],
            )

            source_brand_query = (
                "电阻-4.99R-±1%-1/16W-(-55~155℃)-车规-0402;"
                "YAGEO;AC0402FR-074R99L"
            )
            source_brand_mode, source_brand_spec = app["detect_query_mode_and_spec"](
                pd.DataFrame(),
                source_brand_query,
            )
            source_brand_spec = app["merge_query_text_hints_into_spec"](
                source_brand_spec,
                source_brand_query,
            )
            self.assertEqual(source_brand_spec["品牌"], "国巨YAGEO")
            self.assertTrue(app["fojan_brand_requested_or_unset"](source_brand_spec))
            source_brand_fojan = app["build_fojan_special_resistor_candidates_from_spec"](
                source_brand_spec
            )
            self.assertTrue(source_brand_fojan["型号"].str.startswith("FRQ0402F4R99").any())
            source_brand_matches = app["run_query_match"](
                source_brand_fojan,
                source_brand_mode,
                source_brand_spec,
            )
            self.assertTrue(source_brand_matches.iloc[0]["型号"].startswith("FRQ0402F4R99"))

            yageo_only_spec = app["apply_search_brand_scope_to_spec"](
                source_brand_spec,
                source_brand_query,
                brand_mode=app["SEARCH_BRAND_MODE_CUSTOM"],
                selected_brands=["国巨YAGEO"],
            )
            self.assertFalse(app["fojan_brand_requested_or_unset"](yageo_only_spec))
            self.assertTrue(
                app["build_fojan_special_resistor_candidates_from_spec"](
                    yageo_only_spec
                ).empty
            )

            missing_special_series = []
            for series_name, profile in app["FOJAN_SPECIAL_RESISTOR_CATALOG"].items():
                size_name, size_rule = next(iter(profile["sizes"].items()))
                minimum = float(size_rule.get("min_ohm", 0.0) or 0.0)
                maximum = float(size_rule.get("max_ohm", 1000.0) or 1000.0)
                resistance = min(max(1000.0, minimum), maximum)
                tolerance = next(iter(profile.get("tolerances", ("1",))))
                model = app["build_fojan_catalog_model"](
                    series_name,
                    profile,
                    size_name,
                    resistance,
                    tolerance,
                    suffix=next(iter(profile.get("suffixes", ("TS",)))),
                )
                if model == "":
                    missing_special_series.append(series_name)
            self.assertEqual(missing_special_series, [])
            self.assertEqual(
                app["normalize_fojan_frc_model_display"]("FRC0201F1003 TS", "FOJAN(富捷)"),
                "FRC0201F1003TS",
            )
            self.assertEqual(
                app["normalize_fojan_frc_model_display"]("FRC0402F1001RS", "FOJAN(富捷)"),
                "FRC0402F1001TS",
            )
            self.assertEqual(
                app["normalize_fojan_frc_model_display"]("FRC0402J102TS", "FOJAN(富捷)"),
                "FRC0402J102 TS",
            )
            self.assertEqual(
                app["normalize_fojan_frc_model_display"]("FRC0201J103TS", "FOJAN(富捷)"),
                "FRC0201J103 TS",
            )
            self.assertEqual(
                app["normalize_fojan_frc_model_display"]("FRC0402F1003 TS", "OtherBrand"),
                "FRC0402F1003 TS",
            )
            self.assertIsNone(app["parse_valid_fojan_resistor_model"]("FRC0402F103TS"))

            normalized_suffix = app["normalize_fojan_resistor_series_display_fields"](
                app["ensure_component_display_columns"](
                    pd.DataFrame(
                        [
                            {
                                "品牌": "FOJAN(富捷)",
                                "型号": "FRC0201F1003 TS",
                                "器件类型": "厚膜电阻",
                            },
                            {
                                "品牌": "FOJAN(富捷)",
                                "型号": "FRC0402F1001RS",
                                "器件类型": "厚膜电阻",
                            },
                            {
                                "品牌": "FOJAN(富捷)",
                                "型号": "FRC0402J102TS",
                                "器件类型": "厚膜电阻",
                            },
                            {
                                "品牌": "FOJAN(富捷)",
                                "型号": "FRC0201J103TS",
                                "器件类型": "",
                            },
                        ]
                    )
                )
            )
            self.assertEqual(
                normalized_suffix["型号"].tolist(),
                ["FRC0201F1003TS", "FRC0402F1001TS", "FRC0402J102 TS", "FRC0201J103 TS"],
            )
            rs_only_export_slots = app["build_bom_own_brand_export_slots"](
                prepared_suffix[prepared_suffix["型号"].eq("FRC0402F1001RS")].copy(),
                spec=app["parse_resistor_spec_query"]("0402 1K 1% 1/16W"),
                export_settings={
                    "mode": app["BOM_EXPORT_MODE_CUSTOM"],
                    "brands": ["富捷"],
                },
            )
            self.assertEqual(rs_only_export_slots["自有型号"], "FRC0402F1001TS")

            mode, rohm_spec = app["detect_query_mode_and_spec"](
                pd.DataFrame(), "贴片电阻 10K 0603 ±1% 0.25W ESR系列 ROHM"
            )
            self.assertEqual(mode, "厚膜电阻")
            self.assertEqual(rohm_spec["品牌"], "ROHM")
            self.assertFalse(rohm_spec.get(app["BRAND_QUERY_FILTER_FLAG"], False))
        finally:
            app["fetch_search_candidate_pairs"] = original_fetch

        fojan_no_power = app["parse_resistor_spec_query"]("0805 910R ±1%")
        self.assertEqual(app["build_fojan_resistor_model_from_spec"](fojan_no_power), "FRC0805F9100TS")
        for query, expected_model in (
            ("0201 10R ±5% 1/20W", "FRC0201J100 TS"),
            ("0201 33R ±5% 1/20W", "FRC0201J330 TS"),
            ("0201 1K ±5% 1/20W", "FRC0201J102 TS"),
            ("0201 10K ±5% 1/20W", "FRC0201J103 TS"),
            ("0201 1M ±5% 1/20W", "FRC0201J105 TS"),
        ):
            fojan_0201_5pct = app["parse_resistor_spec_query"](query)
            self.assertEqual(
                app["build_fojan_resistor_model_from_spec"](fojan_0201_5pct),
                expected_model,
            )
            fojan_0201_candidates = app["build_fojan_rule_candidate_from_spec"](fojan_0201_5pct)
            self.assertEqual(fojan_0201_candidates["型号"].tolist(), [expected_model])
            fojan_0201_slots = app["build_bom_own_brand_export_slots"](
                fojan_0201_candidates,
                spec=fojan_0201_5pct,
                export_settings={
                    "mode": app["BOM_EXPORT_MODE_CUSTOM"],
                    "brands": ["富捷"],
                },
            )
            self.assertEqual(fojan_0201_slots["自有型号"], expected_model)
        fojan_low_ohm_no_power = app["parse_resistor_spec_query"]("1206 10mΩ ±1%")
        self.assertEqual(app["build_fojan_resistor_model_from_spec"](fojan_low_ohm_no_power), "FRL1206FR010TS")
        fojan_wrong_power = app["parse_resistor_spec_query"]("1206 10mΩ ±1% 1W")
        self.assertEqual(app["build_fojan_resistor_model_from_spec"](fojan_wrong_power), "")

        real_fojan_row = pd.DataFrame(
            [
                {
                    "品牌": "FOJAN(富捷)",
                    "型号": "FRC0603J102 TS",
                    "器件类型": "厚膜电阻",
                    "_component_type": "厚膜电阻",
                    "系列": "FRC",
                    "系列说明": "普通厚膜贴片电阻",
                    "尺寸（inch）": "0603",
                    "容值": "1",
                    "容值单位": "KΩ",
                    "容值误差": "5",
                    "数据来源": "JLC-SMT官方元器件清单",
                }
            ]
        )
        fallback_fojan_row = pd.DataFrame(
            [
                {
                    "品牌": "FOJAN(富捷)",
                    "型号": "FRC0603J102TS",
                    "器件类型": "厚膜电阻",
                    "_component_type": "厚膜电阻",
                    "系列": "FRC",
                    "系列说明": "普通厚膜贴片电阻",
                    "尺寸（inch）": "0603",
                    "容值误差": "5",
                    "数据来源": "型号编码解析（成本按当前富捷系列规则）",
                }
            ]
        )
        merged_fojan = app["concat_component_frames"]([real_fojan_row, fallback_fojan_row])
        self.assertEqual(merged_fojan["型号"].tolist(), ["FRC0603J102 TS"])

        original_query = "100Ω;50V;±1%;1/16W;0402;RC0402FR-07100RL;无卤"
        exact_yageo = app["prepare_search_dataframe"](
            app["ensure_component_display_columns"](
                pd.DataFrame(
                    [
                        {
                            "品牌": "国巨YAGEO",
                            "型号": "RC0402FR-07100RL",
                            "器件类型": "厚膜电阻",
                            "系列": "RC",
                            "尺寸（inch）": "0402",
                            "材质（介质）": "",
                            "容值_pf": None,
                            "容值": "100",
                            "容值单位": "Ω",
                            "容值误差": "1",
                            "功率": "1/16W",
                            "耐压（V）": "50",
                            "特殊用途": "无卤",
                        }
                    ]
                )
            )
        )
        mode, original_spec = app["detect_query_mode_and_spec"](exact_yageo, original_query)
        original_spec = app["merge_query_text_hints_into_spec"](original_spec, original_query)
        self.assertEqual(app["normalize_special_use"](original_spec["特殊用途"]), "无卤")

        fojan_candidate = app["build_fojan_rule_candidate_from_spec"](original_spec)
        candidate_frame = app["finalize_search_candidate_frames"]([exact_yageo, fojan_candidate])
        fojan_rows = candidate_frame[
            candidate_frame["型号"].astype(str).map(app["clean_model"]).eq("FRC0402F1000TS")
        ]
        self.assertEqual(len(fojan_rows), 1)
        self.assertEqual(app["clean_voltage"](fojan_rows.iloc[0]["耐压（V）"]), "50")
        self.assertEqual(app["normalize_special_use"](fojan_rows.iloc[0]["特殊用途"]), "无卤")

        original_matches = app["run_query_match"](candidate_frame, mode, original_spec)
        self.assertIn(
            "FRC0402F1000TS",
            set(original_matches["型号"].astype(str).map(app["clean_model"])),
        )

        direct_query = "100Ω;50V;±1%;1/16W;0402;"
        direct_spec = app["parse_resistor_spec_query"](direct_query)
        direct_spec = app["merge_query_text_hints_into_spec"](direct_spec, direct_query)
        self.assertEqual(direct_spec["尺寸（inch）"], "0402")
        self.assertEqual(direct_spec["_power"], "1/16W")
        self.assertEqual(app["clean_voltage"](direct_spec["耐压（V）"]), "50")
        self.assertAlmostEqual(float(direct_spec["_resistance_ohm"]), 100.0)
        self.assertEqual(app["build_fojan_resistor_model_from_spec"](direct_spec), "FRC0402F1000TS")

        direct_candidates = app["finalize_search_candidate_frames"](
            [app["build_fojan_rule_candidate_from_spec"](direct_spec)]
        )
        direct_matches = app["run_query_match"](direct_candidates, "贴片电阻", direct_spec)
        self.assertEqual(
            set(direct_matches["型号"].astype(str).map(app["clean_model"])),
            {"FRC0402F1000TS"},
        )
        direct_spec_info = app["build_spec_info_df"](direct_spec)
        self.assertEqual(app["clean_text"](direct_spec_info.iloc[0]["系列"]), "")

        special_queries = {
            "0603 10K 1% 车规": "车规",
            "0603 10K 1% 抗硫化": "抗硫化",
            "1206 1M 1% 高耐压": "高压",
            "1206 10K 1% 0.5W 高功率": "高功率",
            "0603 10K 1% 抗浪涌": "抗浪涌",
            "2512 0.01R 1% 电流检测": "电流检测",
        }
        for query, expected_special in special_queries.items():
            parsed_special = app["parse_resistor_spec_query"](query)
            parsed_special = app["merge_query_text_hints_into_spec"](parsed_special, query)
            self.assertIn(
                expected_special,
                app["special_use_tokens"](parsed_special["特殊用途"]),
                query,
            )

        self.assertTrue(app["special_use_matches"]("车规 | 抗硫化", "车规/抗硫化"))
        self.assertFalse(app["special_use_matches"]("车规", "车规/抗硫化"))
        self.assertFalse(app["special_use_matches"]("抗硫化", "车规/抗硫化"))

        official_special_series = {
            "FPR0603F1002TS": ("FPR", {"车规", "高功率", "抗硫化"}),
            "FQV1206F1004TS": ("FQV", {"车规", "高压", "抗硫化"}),
            "FTH0603F1002TS": ("FTH", {"高精度", "低温漂"}),
            "FQA0603F1002TS": ("FQA", {"车规", "排阻"}),
            "FWPK2512FR001TS": ("FWPK", {"四端子", "高功率", "电流检测"}),
            "FUS2512FR001TS": ("FUS", {"车规", "高功率", "电流检测"}),
        }
        for model, (series, required_tags) in official_special_series.items():
            profile = app["lookup_official_resistor_series_profile_by_model"](
                model,
                "FOJAN(富捷)",
            )
            self.assertEqual(profile["系列"], series, model)
            actual_tags = {
                token.strip()
                for token in profile["特殊用途"].split("|")
                if token.strip()
            }
            self.assertTrue(required_tags.issubset(actual_tags), model)

        special_candidates = pd.DataFrame(
            [
                {
                    "品牌": "FOJAN(富捷)",
                    "型号": model,
                    "器件类型": "厚膜电阻",
                    "系列": series,
                    "尺寸（inch）": "0603",
                    "材质（介质）": "",
                    "耐压（V）": "",
                    "容值_pf": None,
                    "容值": "10",
                    "容值单位": "KΩ",
                    "容值误差": "1",
                    "功率": "1/10W",
                    "特殊用途": special_use,
                }
                for model, series, special_use in [
                    ("FRC0603F1002TS", "FRC", ""),
                    ("FRQ0603F1002TS", "FRQ", "车规"),
                    ("FRR0603F1002TS", "FRR", "抗硫化"),
                    ("FPR0603F1002TS", "FPR", "车规 | 抗硫化 | 高功率"),
                ]
            ]
        )
        prepared_special = app["prepare_search_dataframe"](
            app["ensure_component_display_columns"](special_candidates)
        )
        original_fetch = app["fetch_search_candidate_pairs"]
        app["fetch_search_candidate_pairs"] = lambda _spec: None
        try:
            combined_query = "0603 10K 1% 1/10W 车规 抗硫化"
            combined_spec = app["merge_query_text_hints_into_spec"](
                app["parse_resistor_spec_query"](combined_query),
                combined_query,
            )
            combined_matches = app["match_other_passive_spec"](
                prepared_special,
                combined_spec,
            )
        finally:
            app["fetch_search_candidate_pairs"] = original_fetch
        self.assertEqual(combined_matches["型号"].tolist(), ["FPR0603F1002TS"])

        official_matrix_cases = [
            (
                "0603 10K 1% 车规",
                10,
                {"FRQ0603F1002TS", "FRR0603F1002TS", "FPR0603F1002TS", "FQV0603F1002TS"},
                {"FRC0603F1002TS"},
            ),
            (
                "1206 10K 1% 1/2W 高功率",
                5,
                {"FRP1206F1002TS", "FPR1206F1002TS", "FPS1206F1002TS"},
                {"FRC1206F1002TS"},
            ),
            (
                "1206 100K 1% 500V 高压",
                2,
                {"FRV1206F1003TS", "FQV1206F1003TS"},
                {"FRC1206F1003TS"},
            ),
            (
                "0612 10K 1% 宽端子",
                2,
                {"FCW1206F1002TS", "FPW1206F1002TS"},
                set(),
            ),
            (
                "0805 0.33R 1% 车规 低阻 低TCR",
                2,
                {"FQL0805FR330TSL", "FQL0805FR330TSW"},
                set(),
            ),
            (
                "0612 1R 1% 1W 车规 低阻 宽端子",
                1,
                {"FQL091WF1R00TSR"},
                set(),
            ),
            (
                "064R 150R 5% 车规 排阻",
                2,
                {"FQA064RJ151TS", "FAR064RJ151TS"},
                set(),
            ),
            (
                "064R 150R 5% 车规 抗硫化 排阻",
                1,
                {"FAR064RJ151TS"},
                {"FQA064RJ151TS"},
            ),
        ]
        for query, minimum_count, required_models, forbidden_models in official_matrix_cases:
            spec = app["merge_query_text_hints_into_spec"](
                app["parse_resistor_spec_query"](query),
                query,
            )
            self.assertIsNotNone(spec, query)
            self.assertNotEqual(app["clean_text"](spec.get("特殊用途", "")), "", query)
            candidate = app["build_fojan_rule_candidate_from_spec"](spec)
            models = set(candidate["型号"].astype(str).map(app["clean_model"]))
            self.assertGreaterEqual(len(models), minimum_count, query)
            self.assertTrue(required_models.issubset(models), query)
            self.assertTrue(models.isdisjoint(forbidden_models), query)

        original_fetch = app["fetch_search_candidate_pairs"]
        try:
            app["fetch_search_candidate_pairs"] = lambda _spec: []
            end_to_end_query = "1206 10K 1% 1/2W 高功率"
            end_to_end_spec = app["merge_query_text_hints_into_spec"](
                app["parse_resistor_spec_query"](end_to_end_query),
                end_to_end_query,
            )
            end_to_end_frame = app["load_search_dataframe_for_query"](
                "贴片电阻",
                end_to_end_spec,
                query_text=end_to_end_query,
            )
            end_to_end_matches = app["run_query_match"](
                end_to_end_frame,
                "贴片电阻",
                end_to_end_spec,
            )
        finally:
            app["fetch_search_candidate_pairs"] = original_fetch
        self.assertIn(
            "FRP1206F1002TS",
            set(end_to_end_matches["型号"].astype(str).map(app["clean_model"])),
        )

        alloy_matrix_cases = {
            "1206 50mR 1% 0.5W 合金电阻": {"FMB1205FR050TM"},
            "2010 10mR 1% 1.5W 合金电阻": {"FRM2015FR010TM"},
            "2512 1mR 0.5% 3W 合金电阻": {"FPM253WDR001TM", "FPM253WDR001TML"},
        }
        for query, expected_models in alloy_matrix_cases.items():
            spec = app["parse_resistor_spec_query"](query)
            candidate = app["build_fojan_rule_candidate_from_spec"](spec)
            models = set(candidate["型号"].astype(str).map(app["clean_model"]))
            self.assertEqual(models, expected_models, query)

        official_fojan_series = {
            "FAR", "FCM", "FCN", "FCP", "FCR", "FCS", "FCW", "FHS", "FJR", "FMB",
            "FMH", "FMK", "FMS", "FNL", "FPL", "FPM", "FPR", "FPS", "FPW", "FQA",
            "FQL", "FQL-L", "FQP", "FQS", "FQT", "FQV", "FQW", "FRA", "FRB",
            "FRC", "FRC-P", "FRC-X", "FRD", "FRE", "FRG", "FRH", "FRH-X", "FRJ",
            "FRL", "FRL-L", "FRM", "FRN", "FRP", "FRP-X", "FRQ", "FRR", "FRS",
            "FRT", "FRV", "FRZ", "FSHM", "FSM", "FSP", "FSR", "FTH", "FTR",
            "FUP", "FUS", "FWK", "FWKP", "FWP", "FWPK",
        }
        covered_fojan_series = (
            set(app["FOJAN_SPECIAL_RESISTOR_CATALOG"].keys())
            | set(app["FOJAN_EXTENDED_ALLOY_MODEL_PROFILES"].keys())
            | {"FRC", "FMB", "FRM", "FPM"}
        )
        self.assertEqual(sorted(official_fojan_series - covered_fojan_series), [])

        official_fojan_samples = {
            "FQW1206F1001TS": ("FQW", 1000.0),
            "FRC1206F1001TSP": ("FRC-P", 1000.0),
            "FRH1206B1001TSX": ("FRH-X", 1000.0),
            "FSM25125WFR001TM": ("FSM", 0.001),
            "FMS252WFR010TM": ("FMS", 0.01),
            "FSHM28187WFR020TM": ("FSHM", 0.02),
            "FMK38205WFR005BK": ("FMK", 0.005),
            "FCN43125WFR001RK": ("FCN", 0.001),
            "FUS45mV200AFBKA": ("FUS", 0.000225),
            "FWKP27265WFR001RK": ("FWKP", 0.001),
            "FHS2025FR001BK": ("FHS", 0.001),
            "FJR39205WFR005BK": ("FJR", 0.005),
            "FSP25125WFR001TK": ("FSP", 0.001),
            "FCS851836WF0M050AP2S": ("FCS", 0.00005),
            "FSR2015JR005BCU": ("FSR", 0.005),
            "FCR11205WFR005BK": ("FCR", 0.005),
        }
        for model, (expected_series, expected_resistance) in official_fojan_samples.items():
            parsed = app["parse_fojan_catalog_resistor_model"](model, brand="FOJAN") or app[
                "parse_fojan_alloy_resistor_model"
            ](model, brand="FOJAN", component_type="合金电阻")
            self.assertIsNotNone(parsed, model)
            self.assertEqual(parsed["系列"], expected_series, model)
            self.assertAlmostEqual(float(parsed["_resistance_ohm"]), expected_resistance, msg=model)

        ordinary_spec = app["parse_resistor_spec_query"]("0603 10K 1% 1/10W")
        ordinary_candidate = app["build_fojan_rule_candidate_from_spec"](ordinary_spec)
        self.assertEqual(
            ordinary_candidate["型号"].astype(str).map(app["clean_model"]).tolist(),
            ["FRC0603F1002TS"],
        )

        high_ohmic_cases = (
            "1206 20M 5% 1/4W",
            "20MΩ;200V;±5%;1/4W;1206;FOJAN;FRG1206J206 TS;无卤",
        )
        for query in high_ohmic_cases:
            high_ohmic_spec = app["merge_query_text_hints_into_spec"](
                app["parse_resistor_spec_query"](query),
                query,
            )
            high_ohmic_candidate = app["build_fojan_rule_candidate_from_spec"](high_ohmic_spec)
            models = high_ohmic_candidate["型号"].astype(str).map(app["clean_model"]).tolist()
            self.assertEqual(models, ["FRG1206J206TS"], query)
            self.assertNotIn("FRC1206J206TS", models, query)

        boundary_spec = app["parse_resistor_spec_query"]("1206 10M 5% 1/4W")
        boundary_candidate = app["build_fojan_rule_candidate_from_spec"](boundary_spec)
        self.assertEqual(
            boundary_candidate["型号"].astype(str).map(app["clean_model"]).tolist(),
            ["FRC1206J106TS"],
        )

    def test_03a_walsin_array_maps_to_fojan_fra(self):
        app = self.app
        parsed_models = [
            app["parse_resistor_model_rule"](
                model,
                brand="华新科Walsin",
                component_type="厚膜电阻",
            )
            for model in ("WA04X680JTL", "WA04X680 JTL")
        ]
        self.assertTrue(all(item is not None for item in parsed_models))
        self.assertEqual(
            {app["clean_model"](item["型号"]) for item in parsed_models},
            {"WA04X680JTL"},
        )
        self.assertEqual({item["型号"] for item in parsed_models}, {"WA04X680 JTL"})
        walsin_array = parsed_models[0]
        self.assertIsNotNone(walsin_array)
        self.assertEqual(walsin_array["尺寸（inch）"], "044R")
        self.assertAlmostEqual(float(walsin_array["_resistance_ohm"]), 68.0)
        self.assertEqual(app["clean_tol_for_match"](walsin_array["容值误差"]), "5")
        self.assertEqual(walsin_array["_power"], "1/16W")
        self.assertIn("排阻", walsin_array["特殊用途"])

        walsin_array_candidates = app["build_fojan_rule_candidate_from_spec"](walsin_array)
        models = set(walsin_array_candidates["型号"].astype(str).map(app["clean_model"]))
        self.assertIn("FRA044RJ680TS", models)
        fra = walsin_array_candidates[
            walsin_array_candidates["型号"].astype(str).map(app["clean_model"]).eq("FRA044RJ680TS")
        ].iloc[0]
        self.assertEqual(fra["尺寸（inch）"], "044R")
        self.assertAlmostEqual(float(fra["_resistance_ohm"]), 68.0)
        self.assertEqual(app["clean_tol_for_match"](fra["容值误差"]), "5")
        self.assertEqual(fra["功率"], "1/16W")
        self.assertIn("排阻", app["clean_text"](fra["特殊用途"]))

        display_frame = pd.DataFrame(
            [{"品牌": "华新科Walsin", "型号": "WA04X680JTL", "器件类型": "厚膜电阻"}]
        )
        normalized = app["normalize_resistor_model_display_fields"](display_frame)
        self.assertEqual(normalized.iloc[0]["型号"], "WA04X680 JTL")

    def test_03aa_customer_kr_suffix_and_spaced_decimal_parse(self):
        app = self.app
        customer_kr_specs = [
            ("厚膜电阻, 0402, 11.3KR, ±1%, 1/16W, RoHS", 11_300.0, "FRC0402F1132TS"),
            ("厚膜电阻, 0402, 1KR, ±1%, 1/16W, RoHS", 1_000.0, "FRC0402F1001TS"),
            ("厚膜电阻, 0402, 2KR, ±1%, 1/16W, RoHS", 2_000.0, "FRC0402F2001TS"),
            ("厚膜电阻, 4. 7KR, ±1%, 0402, 1/16W, RoHS", 4_700.0, "FRC0402F4701TS"),
            ("厚膜电阻, 0402, 10KR, ±1%, 1/16W, RoHS", 10_000.0, "FRC0402F1002TS"),
            ("厚膜电阻, 0402, 20KR, ±1%, 1/16W, RoHS", 20_000.0, "FRC0402F2002TS"),
            ("厚膜电阻, 0402, 33KR, ±1%, 1/16W, RoHS", 33_000.0, "FRC0402F3302TS"),
            ("厚膜电阻, 0402, 100KR, ±1%, 1/16W, RoHS", 100_000.0, "FRC0402F1003TS"),
            ("厚膜电阻, 0603, 3. 9KR, ±1%, 1/10W, RoHS", 3_900.0, "FRC0603F3901TS"),
            ("厚膜电阻, 1206, 196KR, ±1%, 1/4W, RoHS", 196_000.0, "FRC1206F1963TS"),
            ("厚膜电阻, 0402, 22R, 1%, 1/16W, RoHS", 22.0, "FRC0402F22R0TS"),
        ]
        for query, expected_ohm, expected_model in customer_kr_specs:
            parsed = app["parse_resistor_spec_query"](query)
            self.assertIsNotNone(parsed, query)
            self.assertEqual(parsed["_param_count"], 4, query)
            self.assertAlmostEqual(float(parsed["_resistance_ohm"]), expected_ohm, msg=query)
            self.assertEqual(app["build_fojan_resistor_model_from_spec"](parsed), expected_model, query)

        self.assertEqual(
            app["normalize_resistor_value_tolerance_separator"]("CC0603KRX7R9BB103"),
            "CC0603KRX7R9BB103",
        )
        self.assertEqual(
            app["normalize_resistor_value_tolerance_separator"]("1206 50mR 1% 0.5W"),
            "1206 50mR 1% 0.5W",
        )
        milliohm = app["parse_resistor_spec_query"]("1206 50mR 1% 0.5W 合金电阻")
        self.assertAlmostEqual(float(milliohm["_resistance_ohm"]), 0.05)

    def test_03b_common_parts_keep_cross_brand_candidates_with_compliance_confirmation(self):
        app = self.app
        query = "39KΩ ±1% 0805 1/8W 0805W8F3902T5E 无卤 品牌:厚声/翔胜/华科/国巨"
        spec = app["parse_resistor_spec_query"]("0805 39K 1% 1/8W 无卤")
        spec.update({"品牌": "厚声UNI-ROYAL", "型号": "0805W8F3902T5E"})
        spec = app["merge_query_text_hints_into_spec"](spec, query)
        spec = app["clear_query_brand_filter_for_embedded_part_metadata"](spec, query)

        self.assertFalse(spec.get(app["BRAND_QUERY_FILTER_FLAG"], False))
        self.assertEqual(spec.get("品牌"), "厚声UNI-ROYAL")

        candidate_rows = []
        for brand, model, special_use in (
            ("厚声UNI-ROYAL", "0805W8F3902T5E", "无卤"),
            ("VO(翔胜)", "SCR0805F39K", ""),
            ("华新科Walsin", "WR08X3902FTL", "无卤"),
            ("国巨YAGEO", "RC0805FR-0739KL", ""),
            ("FOJAN(富捷)", "FRC0805F3902TS", "无卤"),
        ):
            candidate_rows.append(
                {
                    "品牌": brand,
                    "型号": model,
                    "器件类型": "厚膜电阻",
                    "尺寸（inch）": "0805",
                    "材质（介质）": "",
                    "耐压（V）": "",
                    "容值_pf": None,
                    "容值": "39",
                    "容值单位": "KΩ",
                    "容值误差": "1",
                    "功率": "1/8W",
                    "_resistance_ohm": 39000.0,
                    "特殊用途": special_use,
                }
            )
        prepared = app["prepare_search_dataframe"](
            app["ensure_component_display_columns"](pd.DataFrame(candidate_rows))
        )
        original_fetch = app["fetch_search_candidate_pairs"]
        app["fetch_search_candidate_pairs"] = lambda _spec: None
        try:
            matched = app["run_query_match"](prepared, "料号", spec)
        finally:
            app["fetch_search_candidate_pairs"] = original_fetch

        self.assertEqual(
            set(matched["品牌"]),
            {"VO(翔胜)", "华新科Walsin", "国巨YAGEO", "FOJAN(富捷)"},
        )
        levels = dict(zip(matched["型号"], matched["推荐等级"]))
        self.assertEqual(levels["WR08X3902FTL"], "完全匹配")
        self.assertEqual(levels["SCR0805F39K"], "需确认替代")
        self.assertEqual(levels["RC0805FR-0739KL"], "需确认替代")
        warnings_list = app["collect_recommendation_warnings"](
            matched.loc[matched["型号"].eq("RC0805FR-0739KL")].iloc[0],
            spec,
        )
        self.assertTrue(any("无卤" in warning for warning in warnings_list))

        custom_spec = app["apply_search_brand_scope_to_spec"](
            spec,
            query,
            app["SEARCH_BRAND_MODE_CUSTOM"],
            ["华新科Walsin", "国巨YAGEO"],
        )
        self.assertTrue(custom_spec.get(app["BRAND_QUERY_FILTER_FLAG"]))
        try:
            app["fetch_search_candidate_pairs"] = lambda _spec: None
            custom_matched = app["run_query_match"](prepared, "料号", custom_spec)
        finally:
            app["fetch_search_candidate_pairs"] = original_fetch
        self.assertEqual(set(custom_matched["品牌"]), {"华新科Walsin", "国巨YAGEO"})

        directive_query = "0805 39K 1% 1/8W 无卤 指定品牌:翔胜/国巨"
        directive_spec = app["merge_query_text_hints_into_spec"](
            app["parse_resistor_spec_query"]("0805 39K 1% 1/8W 无卤"),
            directive_query,
        )
        directive_aliases = set(app["requested_brand_aliases_from_spec"](directive_spec))
        self.assertIn("翔胜", directive_aliases)
        self.assertIn("国巨", directive_aliases)

        mlcc_query = "220pF ±10% 50V X7R 0603 0603B221K500CTSB 无卤品牌:华科/国巨/三环/风华"
        mlcc_spec = app["parse_spec_query"]("0603 X7R 220pF 10% 50V 无卤")
        mlcc_spec.update({"品牌": "华新科Walsin", "型号": "0603B221K500CTSB"})
        mlcc_spec = app["merge_query_text_hints_into_spec"](mlcc_spec, mlcc_query)
        mlcc_spec = app["clear_query_brand_filter_for_embedded_part_metadata"](mlcc_spec, mlcc_query)
        mlcc_rows = pd.DataFrame(
            [
                {
                    "品牌": "华新科Walsin",
                    "型号": "0603B221K500CTSB",
                    "器件类型": "MLCC",
                    "尺寸（inch）": "0603",
                    "材质（介质）": "X7R",
                    "容值_pf": 220.0,
                    "容值误差": "10",
                    "耐压（V）": "50",
                    "特殊用途": "无卤",
                },
                {
                    "品牌": "国巨YAGEO",
                    "型号": "CC0603KRX7R9BB221",
                    "器件类型": "MLCC",
                    "尺寸（inch）": "0603",
                    "材质（介质）": "X7R",
                    "容值_pf": 220.0,
                    "容值误差": "10",
                    "耐压（V）": "50",
                    "特殊用途": "",
                },
            ]
        )
        prepared_mlcc = app["prepare_search_dataframe"](
            app["ensure_component_display_columns"](mlcc_rows)
        )
        app["fetch_search_candidate_pairs"] = lambda _spec: None
        try:
            mlcc_matched = app["run_query_match"](prepared_mlcc, "料号", mlcc_spec)
        finally:
            app["fetch_search_candidate_pairs"] = original_fetch
        self.assertEqual(mlcc_matched["型号"].tolist(), ["CC0603KRX7R9BB221"])
        self.assertEqual(mlcc_matched.iloc[0]["推荐等级"], "需确认替代")
        self.assertTrue(
            any(
                "无卤" in warning
                for warning in app["collect_recommendation_warnings"](mlcc_matched.iloc[0], mlcc_spec)
            )
        )

    def test_04_no_match_resolution_persists_and_searches(self):
        app = self.app
        app["NO_MATCH_REPORT_DB_PATH"] = os.path.join(self.temp_dir, "reports.sqlite")
        app["DB_PATH"] = os.path.join(self.temp_dir, "components.sqlite")
        app["SEARCH_DB_PATH"] = os.path.join(self.temp_dir, "search.sqlite")
        spec = {
            "器件类型": "贴片电阻",
            "尺寸（inch）": "0603",
            "_resistance_ohm": 10000.0,
            "容值误差": "5",
            "功率": "1/10W",
            "规格摘要": "10KΩ ±5% 1/10W 0603",
        }
        query = "0603 10K ±5% 1/10W REGRESSION-MISSING"
        ok, message, report_id = app["submit_no_match_report"](
            query, mode="规格参数", spec=spec, reason="regression"
        )
        self.assertTrue(ok, message)
        self.assertTrue(
            app["resolve_no_match_report"](
                report_id,
                resolved_note="regression resolved",
                resolved_brand="FOJAN(富捷)",
                resolved_model="FRC0603J103 TS",
                resolved_component_type="贴片电阻",
            )
        )
        report = app["get_no_match_report_by_id"](report_id)
        self.assertEqual(report["library_status"], "已写入主库和搜索索引")
        with sqlite3.connect(app["DB_PATH"]) as conn:
            row = conn.execute(
                'SELECT 品牌, 型号 FROM components WHERE REPLACE(UPPER(型号), " ", "")=?',
                ("FRC0603J103TS",),
            ).fetchone()
        self.assertIsNotNone(row)
        for lookup in [query, "FRC0603J103 TS"]:
            resolved = app["resolve_no_match_report_as_query"](lookup)
            self.assertIsNotNone(resolved)
            self.assertFalse(resolved["query_df"].empty)

    def test_05_cost_list_updates_only_changed_cost_time(self):
        app = self.app
        app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "cost-test.sqlite")
        app["clear_cost_price_lookup_cache"]()
        first = pd.DataFrame(
            [
                {
                    "品牌": "FOJAN(富捷)",
                    "型号": "FRC0603J103 TS",
                    "规格参数": "0603 10K 5%",
                    "成本": "1.40",
                    "MOQ": "5000",
                    "L&T": "4W",
                },
                {
                    "品牌": "厚声UNI-ROYAL",
                    "型号": "0603WAJ0103T5E",
                    "规格参数": "0603 10K 5%",
                    "成本": "2.00",
                    "MOQ": "5000",
                    "L&T": "5W",
                },
            ]
        )
        app["current_timestamp_text"] = lambda: "2026-06-28 10:00:00"
        ok, message, _ = app["import_cost_price_list_from_upload"](
            UploadedBytes("cost1.xlsx", dataframe_to_xlsx_bytes(first)), "regression"
        )
        self.assertTrue(ok, message)
        second = first.copy()
        second.loc[0, "成本"] = "1.4"
        second.loc[1, "成本"] = "2.20"
        app["current_timestamp_text"] = lambda: "2026-06-29 11:00:00"
        ok, message, list_id = app["import_cost_price_list_from_upload"](
            UploadedBytes("cost2.xlsx", dataframe_to_xlsx_bytes(second)), "regression"
        )
        self.assertTrue(ok, message)
        by_model = {item["model"]: item for item in app["list_cost_price_items"](list_id, 10)}
        self.assertEqual(by_model["FRC0603J103 TS"]["cost_updated_at"], "2026-06-28 10:00:00")
        self.assertEqual(by_model["0603WAJ0103T5E"]["cost_updated_at"], "2026-06-29 11:00:00")
        entry = app["lookup_active_cost_price_for_row"](
            {"品牌": "FOJAN(富捷)", "型号": "FRC0603J103 TS"}
        )
        self.assertEqual(entry["cost"], "1.4")
        self.assertEqual(entry["moq"], "5000")
        self.assertEqual(entry["lead_time"], "4W")

    def test_05b_manual_cost_overrides_lists_and_can_be_disabled(self):
        app = self.app
        original_cost_path = app["COST_PRICE_DB_PATH"]
        original_timestamp = app["current_timestamp_text"]
        try:
            app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "manual-cost-test.sqlite")
            app["clear_cost_price_lookup_cache"]()
            app["current_timestamp_text"] = lambda: "2026-07-16 09:00:00"
            ok, message, manual_id = app["save_manual_cost_price_item"](
                brand="富捷",
                model="FRC0603F1002TS",
                cost="1.25",
                moq="6000PCS",
                lead_time="2W",
                spec_text="0603 10K 1% 1/10W",
                note="原厂单独询价",
                updated_by="regression",
            )
            self.assertTrue(ok, message)
            self.assertEqual(app["count_manual_cost_price_items"](True), 1)
            manual_only = app["lookup_active_cost_price_for_row"](
                {"品牌": "FOJAN(富捷)", "型号": "FRC0603F1002TS"}
            )
            self.assertEqual(manual_only["cost"], "1.25")
            self.assertEqual(manual_only["cost_source"], "单笔成本")

            list_frame = pd.DataFrame(
                [
                    {
                        "品牌": "FOJAN(富捷)",
                        "型号": "FRC0603F1002TS",
                        "规格参数": "0603 10K 1% 1/10W",
                        "成本": "2.00",
                        "MOQ": "5000PCS",
                        "L&T": "5W",
                    }
                ]
            )
            app["current_timestamp_text"] = lambda: "2026-07-16 10:00:00"
            ok, message, list_id = app["import_cost_price_list_from_upload"](
                UploadedBytes("manual-fallback.xlsx", dataframe_to_xlsx_bytes(list_frame)),
                "regression",
            )
            self.assertTrue(ok, message)
            self.assertIsNotNone(list_id)
            still_manual = app["lookup_active_cost_price_for_row"](
                {"品牌": "FOJAN(富捷)", "型号": "FRC0603F1002TS"}
            )
            self.assertEqual(still_manual["cost"], "1.25")

            app["current_timestamp_text"] = lambda: "2026-07-16 11:00:00"
            ok, message, updated_id = app["save_manual_cost_price_item"](
                brand="FOJAN",
                model="FRC0603F1002TS",
                cost="1.30",
                moq="7000PCS",
                lead_time="3W",
                spec_text="0603 10K 1% 1/10W",
                note="第二次询价",
                updated_by="regression",
                item_id=manual_id,
            )
            self.assertTrue(ok, message)
            self.assertEqual(updated_id, manual_id)
            self.assertEqual(len(app["list_manual_cost_price_items"](None, 10)), 1)
            enriched = app["enrich_component_cost_columns"](
                pd.DataFrame([{"品牌": "FOJAN(富捷)", "型号": "FRC0603F1002TS"}])
            ).iloc[0]
            self.assertEqual(enriched["成本"], "1.30")
            self.assertEqual(enriched["MOQ"], "7000PCS")
            self.assertEqual(enriched["MOQ来源"], "单笔成本")

            ok, message = app["set_manual_cost_price_item_active"](
                manual_id,
                False,
                updated_by="regression",
            )
            self.assertTrue(ok, message)
            fallback = app["lookup_active_cost_price_for_row"](
                {"品牌": "FOJAN(富捷)", "型号": "FRC0603F1002TS"}
            )
            self.assertEqual(app["normalize_cost_value_for_compare"](fallback["cost"]), "2")
            self.assertEqual(fallback["cost_source"], "当前启用成本清单")

            ok, message = app["set_manual_cost_price_item_active"](
                manual_id,
                True,
                updated_by="regression",
            )
            self.assertTrue(ok, message)
            restored_manual = app["lookup_active_cost_price_for_row"](
                {"品牌": "FOJAN(富捷)", "型号": "FRC0603F1002TS"}
            )
            self.assertEqual(restored_manual["cost"], "1.30")
            self.assertEqual(restored_manual["cost_updated_at"], "2026-07-16 11:00:00")
        finally:
            app["COST_PRICE_DB_PATH"] = original_cost_path
            app["current_timestamp_text"] = original_timestamp
            app["clear_cost_price_lookup_cache"]()

    def test_05c_cost_prices_are_isolated_by_customer(self):
        app = self.app
        original_cost_path = app["COST_PRICE_DB_PATH"]
        try:
            app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "customer-cost-test.sqlite")
            app["clear_cost_price_lookup_cache"]()

            def upload_cost(file_name, cost, customer_type="new", customer_name=""):
                frame = pd.DataFrame(
                    [{"brand": "FOJAN", "model": "FRC0603F1002TS", "cost": cost}]
                )
                return app["import_cost_price_list_from_upload"](
                    UploadedBytes(file_name, dataframe_to_xlsx_bytes(frame)),
                    "regression",
                    customer_type=customer_type,
                    customer_name=customer_name,
                )

            ok, message, base_id = upload_cost("base.xlsx", "1.00")
            self.assertTrue(ok, message)
            ok, message, customer_a_id = upload_cost("customer-a.xlsx", "1.20", "existing", "客户A")
            self.assertTrue(ok, message)
            ok, message, customer_b_id = upload_cost("customer-b.xlsx", "1.30", "existing", "客户B")
            self.assertTrue(ok, message)

            row = {"品牌": "FOJAN", "型号": "FRC0603F1002TS"}
            base = app["lookup_active_cost_price_for_row"](
                row,
                lookup=app["load_active_cost_price_lookup"]("new", ""),
            )
            customer_a = app["lookup_active_cost_price_for_row"](
                row,
                lookup=app["load_active_cost_price_lookup"]("existing", "客户A"),
            )
            customer_b = app["lookup_active_cost_price_for_row"](
                row,
                lookup=app["load_active_cost_price_lookup"]("existing", "客户B"),
            )
            self.assertEqual(app["normalize_cost_value_for_compare"](base["cost"]), "1")
            self.assertEqual(app["normalize_cost_value_for_compare"](customer_a["cost"]), "1.2")
            self.assertEqual(app["normalize_cost_value_for_compare"](customer_b["cost"]), "1.3")
            self.assertEqual(app["get_active_cost_price_list"]("new", "")["id"], base_id)
            self.assertEqual(app["get_active_cost_price_list"]("existing", "客户A")["id"], customer_a_id)
            self.assertEqual(app["get_active_cost_price_list"]("existing", "客户B")["id"], customer_b_id)

            ok, message, manual_id = app["save_manual_cost_price_item"](
                "FOJAN", "FRC0603F1002TS", "1.15",
                customer_type="existing", customer_name="客户A",
            )
            self.assertTrue(ok, message)
            self.assertIsNotNone(manual_id)
            customer_a_manual = app["lookup_active_cost_price_for_row"](
                row,
                lookup=app["load_active_cost_price_lookup"]("existing", "客户A"),
            )
            self.assertEqual(customer_a_manual["cost"], "1.15")
            unquoted_customer = app["lookup_active_cost_price_for_row"](
                row,
                lookup=app["load_active_cost_price_lookup"]("existing", "未报价客户"),
            )
            self.assertEqual(app["normalize_cost_value_for_compare"](unquoted_customer["cost"]), "1")
            self.assertEqual(set(app["list_existing_cost_customers"]()), {"客户A", "客户B"})
            self.assertEqual(
                app["resolve_member_customer_price_scope"](" 客户 A ", ["客户A", "客户B"]),
                ("existing", "客户A", True),
            )
            self.assertEqual(
                app["resolve_member_customer_price_scope"]("新客户C", ["客户A", "客户B"]),
                ("new", "", True),
            )
            self.assertEqual(
                app["resolve_member_customer_price_scope"]("", ["客户A", "客户B"]),
                ("", "", False),
            )

            signature_a = app["build_bom_workbook_run_signature"](
                UploadedBytes("same.xlsx", b"same"),
                {"Sheet1": {"model": "model"}},
                {"mode": app["BOM_EXPORT_MODE_AUTO"], "customer_type": "existing", "customer_name": "客户A"},
            )
            signature_b = app["build_bom_workbook_run_signature"](
                UploadedBytes("same.xlsx", b"same"),
                {"Sheet1": {"model": "model"}},
                {"mode": app["BOM_EXPORT_MODE_AUTO"], "customer_type": "existing", "customer_name": "客户B"},
            )
            self.assertNotEqual(signature_a, signature_b)
        finally:
            app["COST_PRICE_DB_PATH"] = original_cost_path
            app["clear_cost_price_lookup_cache"]()

    def test_05d_legacy_cost_database_migrates_without_losing_rows(self):
        app = self.app
        original_cost_path = app["COST_PRICE_DB_PATH"]
        try:
            legacy_path = os.path.join(self.temp_dir, "legacy-customer-cost.sqlite")
            app["COST_PRICE_DB_PATH"] = legacy_path
            with sqlite3.connect(legacy_path) as conn:
                conn.executescript(
                    """
                    CREATE TABLE cost_price_lists (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        file_name TEXT NOT NULL, file_sha256 TEXT NOT NULL DEFAULT '',
                        file_size INTEGER NOT NULL DEFAULT 0, uploaded_at TEXT NOT NULL,
                        uploaded_by TEXT NOT NULL DEFAULT '', row_count INTEGER NOT NULL DEFAULT 0,
                        active INTEGER NOT NULL DEFAULT 0, note TEXT NOT NULL DEFAULT ''
                    );
                    CREATE TABLE cost_price_manual_items (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        brand TEXT NOT NULL DEFAULT '', brand_key TEXT NOT NULL DEFAULT '',
                        model TEXT NOT NULL DEFAULT '', model_clean TEXT NOT NULL DEFAULT '',
                        spec_text TEXT NOT NULL DEFAULT '', cost TEXT NOT NULL DEFAULT '',
                        moq TEXT NOT NULL DEFAULT '', lead_time TEXT NOT NULL DEFAULT '',
                        note TEXT NOT NULL DEFAULT '', active INTEGER NOT NULL DEFAULT 1,
                        cost_updated_at TEXT NOT NULL DEFAULT '', created_at TEXT NOT NULL DEFAULT '',
                        created_by TEXT NOT NULL DEFAULT '', updated_at TEXT NOT NULL DEFAULT '',
                        updated_by TEXT NOT NULL DEFAULT ''
                    );
                    CREATE UNIQUE INDEX idx_cost_price_manual_brand_model
                    ON cost_price_manual_items(brand_key, model_clean);
                    INSERT INTO cost_price_lists
                        (file_name, uploaded_at, row_count, active)
                    VALUES ('legacy.xlsx', '2026-01-01 00:00:00', 10, 1);
                    INSERT INTO cost_price_manual_items
                        (brand, brand_key, model, model_clean, cost, active)
                    VALUES ('FOJAN', 'FOJAN', 'FRC0603F1002TS', 'FRC0603F1002TS', '1.00', 1);
                    """
                )
            app["init_cost_price_db"]()
            with sqlite3.connect(legacy_path) as conn:
                list_row = conn.execute(
                    "SELECT file_name, customer_type, customer_name, customer_key FROM cost_price_lists"
                ).fetchone()
                manual_row = conn.execute(
                    "SELECT model, cost, customer_type, customer_name, customer_key FROM cost_price_manual_items"
                ).fetchone()
                old_index = conn.execute(
                    "SELECT 1 FROM sqlite_master WHERE type='index' AND name='idx_cost_price_manual_brand_model'"
                ).fetchone()
            self.assertEqual(list_row, ("legacy.xlsx", "new", "", ""))
            self.assertEqual(manual_row, ("FRC0603F1002TS", "1.00", "new", "", ""))
            self.assertIsNone(old_index)
            ok, message, customer_record_id = app["save_manual_cost_price_item"](
                "FOJAN", "FRC0603F1002TS", "1.20",
                customer_type="existing", customer_name="客户A",
            )
            self.assertTrue(ok, message)
            self.assertIsNotNone(customer_record_id)
            self.assertEqual(len(app["list_manual_cost_price_items"](None, 10)), 2)
        finally:
            app["COST_PRICE_DB_PATH"] = original_cost_path
            app["clear_cost_price_lookup_cache"]()

    def test_05e_customer_group_price_pages_override_general_without_cross_group_leak(self):
        app = self.app
        original_cost_path = app["COST_PRICE_DB_PATH"]
        try:
            app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "customer-group-price-test.sqlite")
            app["clear_cost_price_lookup_cache"]()
            for name, code, group in [
                ("A集团深圳公司", "F0001", "A集团"),
                ("A集团东莞公司", "F0002", "A集团"),
                ("B集团公司", "B0001", "B集团"),
            ]:
                ok, message, _ = app["save_sales_customer"](
                    name, code, group_name=group, updated_by="regression", sync_remote=False
                )
                self.assertTrue(ok, message)

            workbook = Workbook()
            general = workbook.active
            general.title = "FRC通用"
            general["A1"] = "客户代码"
            general["B1"] = "通用"
            dedicated = workbook.create_sheet("A集团专价")
            dedicated["A1"] = "客户代码"
            dedicated["B1"] = "F0001"
            for sheet, price in [(general, "1.70"), (dedicated, "1.25")]:
                sheet.append([])
                sheet.append(["Series", "Type / Dimension", "Resistance Range", "New Unit Price", "", "Package"])
                sheet.append(["", "", "Ω (ohms)", "5%", "1%", ""])
                sheet.append(["FRC", "0603 1/10W", "10R-1M", "", price, "5000PCS"])
            output = BytesIO()
            workbook.save(output)
            upload = UploadedBytes("group-price.xlsx", output.getvalue())
            ok, message, _ = app["import_cost_price_list_from_upload"](upload, "regression")
            self.assertTrue(ok, message)

            row = {
                "品牌": "FOJAN(富捷)", "型号": "FRC0603F1002TS", "系列": "FRC",
                "尺寸（inch）": "0603", "阻值": 10, "阻值单位": "KΩ", "_resistance_ohm": 10000.0,
                "容值误差": "±1%", "功率": "1/10W",
            }
            a1 = app["lookup_active_cost_price_for_row"](
                row, app["load_active_cost_price_lookup"]("existing", "A集团深圳公司")
            )
            a2 = app["lookup_active_cost_price_for_row"](
                row, app["load_active_cost_price_lookup"]("existing", "A集团东莞公司")
            )
            b = app["lookup_active_cost_price_for_row"](
                row, app["load_active_cost_price_lookup"]("existing", "B集团公司")
            )
            self.assertEqual(app["normalize_cost_value_for_compare"](a1["cost"]), "1.25")
            self.assertEqual(app["normalize_cost_value_for_compare"](a2["cost"]), "1.25")
            self.assertEqual(app["normalize_cost_value_for_compare"](b["cost"]), "1.7")
            self.assertIn("集团共享价", a2["cost_source"])
            context = app["get_sales_customer_price_context"]("A集团深圳公司")
            self.assertEqual(set(context["group_code_keys"]), {"F0001", "F0002"})
            self.assertIn("A集团深圳公司", app["list_existing_cost_customers"]())
        finally:
            app["COST_PRICE_DB_PATH"] = original_cost_path
            app["clear_cost_price_lookup_cache"]()

    def test_06_bom_full_read_export_and_display_columns(self):
        app = self.app
        bom = pd.DataFrame(
            {
                "物料编号": [f"P{i:03d}" for i in range(85)],
                "规格": [f"0603 {i + 1}K 1% 1/10W" for i in range(85)],
                "需求数量": ["1000"] * 85,
            }
        )
        workbook = app["read_uploaded_bom_workbook"](
            UploadedBytes("85rows.xlsx", dataframe_to_xlsx_bytes(bom))
        )
        self.assertEqual(len(workbook["sheet_frames"][0]["df"]), 85)
        result = pd.DataFrame(
            [
                {
                    "自有品牌": "厚声UNI-ROYAL",
                    "自有型号": "0603WAF1002T5E",
                    "自有成本": "0.02",
                    "自有更新时间": "2026-06-28",
                    "自有MOQ": "5000",
                    "自有L&T": "4W",
                    "自有匹配说明": "厚声关键规格完全一致",
                    "自有匹配备注": "厚声包装需确认",
                    "自有品牌2": "FOJAN(富捷)",
                    "自有型号2": "FRC0603F1002 TS",
                    "自有成本2": "0.018",
                    "自有更新时间2": "2026-06-27",
                    "自有MOQ2": "5000",
                    "自有L&T2": "5W",
                    "自有匹配说明2": "富捷关键规格完全一致",
                    "自有匹配备注2": "富捷交期需确认",
                    "BOM行号": 1,
                    "BOM型号": "X",
                    "BOM规格": "0603 10K",
                    "状态": "可推荐",
                    "销售结论": "x",
                    "备选型号": "x",
                    "风险提示": "x",
                    "推荐理由": "x",
                    "解析说明": "x",
                    "客户回复型号": "x",
                    "可直接回复客户": "x",
                }
            ]
        )
        export = app["build_bom_matched_export_df"](bom, result)
        self.assertEqual(export.iloc[0]["匹配状态"], "可推荐")
        self.assertEqual(export.iloc[0]["匹配说明"], "厚声关键规格完全一致")
        self.assertEqual(export.iloc[0]["匹配备注"], "厚声包装需确认")
        self.assertEqual(export.iloc[0]["匹配说明2"], "富捷关键规格完全一致")
        self.assertEqual(export.iloc[0]["匹配备注2"], "富捷交期需确认")
        self.assertEqual(
            list(export.columns[-16:]),
            [
                "匹配品牌",
                "匹配型号",
                "匹配成本",
                "成本更新时间",
                "匹配MOQ",
                "匹配L&T",
                "匹配说明",
                "匹配备注",
                "匹配品牌2",
                "匹配型号2",
                "匹配成本2",
                "成本更新时间2",
                "匹配MOQ2",
                "匹配L&T2",
                "匹配说明2",
                "匹配备注2",
            ],
        )
        candidates = pd.DataFrame(
            [
                {
                    "品牌": "华新科Walsin",
                    "型号": "0402B103K500CT-A",
                    "器件类型": "MLCC",
                    "推荐等级": "完全匹配",
                    "成本": "",
                },
                {
                    "品牌": "华新科Walsin",
                    "型号": "0402B103K500CT-B",
                    "器件类型": "MLCC",
                    "推荐等级": "完全匹配",
                    "成本": "0.018",
                    "MOQ": "5000PCS",
                    "备注1": "华科候选备注",
                },
                {
                    "品牌": "信昌PDC",
                    "型号": "CC10B103K500A",
                    "器件类型": "MLCC",
                    "推荐等级": "完全匹配",
                    "成本": "0.020",
                    "备注1": "信昌候选备注",
                },
            ]
        )
        custom_settings = {"mode": app["BOM_EXPORT_MODE_CUSTOM"], "brands": ["华新科Walsin"]}
        custom_slots = app["build_bom_own_brand_export_slots"](
            candidates,
            spec={"器件类型": "MLCC"},
            export_settings=custom_settings,
        )
        self.assertEqual(custom_slots["自有品牌"], "华新科Walsin")
        self.assertEqual(custom_slots["自有型号"], "0402B103K500CT-B")
        self.assertEqual(custom_slots["自有成本"], "0.018")
        self.assertEqual(custom_slots["自有匹配说明"], "关键规格完全一致")
        self.assertEqual(custom_slots["自有品牌2"], "")

        auto_slots = app["build_bom_own_brand_export_slots"](
            candidates,
            spec={"器件类型": "MLCC"},
            export_settings={"mode": app["BOM_EXPORT_MODE_AUTO"]},
        )
        self.assertEqual(auto_slots["自有品牌"], "华科")
        self.assertEqual(auto_slots["自有品牌2"], "信昌")
        self.assertEqual(auto_slots["自有匹配备注"], "华科候选备注")
        self.assertEqual(auto_slots["自有匹配备注2"], "信昌候选备注")
        self.assertNotEqual(auto_slots["自有匹配说明"], "")
        self.assertNotEqual(auto_slots["自有匹配说明2"], "")
        self.assertIn("芯声微HRE", app["bom_export_brand_options"]())

        mapping = app["guess_bom_column_mapping"](
            pd.DataFrame(
                {
                    "品名": ["贴片电容"] * 4,
                    "规格": ["10nF;50V;±10%;0603;X7R"] * 4,
                    "国巨型号": ["CC0603KRX7R9BB103"] * 4,
                    "PDC料号": ["FN18X103K500PXG", "", "", ""],
                }
            )
        )
        self.assertEqual(mapping["model"], "国巨型号")
        self.assertEqual(mapping["spec"], "规格")

        blank_custom_slots = app["empty_bom_own_brand_export_slots"]()
        false_positive_row = {
            "状态": "可推荐",
            "解析状态": "解析成功",
            "匹配数量": 6,
            "首选推荐等级": "完全匹配",
            "推荐品牌": "国巨YAGEO",
            "推荐型号": "CC0603KRX7R9BB103",
            "推荐理由": "关键规格完全一致",
            "差异说明": "使用规格列解析，首选结果可推荐",
            "前5个其他品牌型号": "国巨YAGEO:CC0603KRX7R9BB103",
            "备注1": "国巨候选备注",
            "备注2": "通用候选备注",
            "品牌1": "国巨",
            "型号1": "CC0603KRX7R9BB103",
        }
        app["reconcile_bom_output_status"](
            false_positive_row,
            blank_custom_slots,
            export_settings={"mode": app["BOM_EXPORT_MODE_CUSTOM"], "brands": ["信昌PDC"]},
        )
        self.assertEqual(false_positive_row["状态"], "无匹配")
        self.assertEqual(false_positive_row["匹配数量"], 0)
        self.assertEqual(false_positive_row["推荐品牌"], "")
        self.assertEqual(false_positive_row["推荐型号"], "")
        self.assertEqual(false_positive_row["品牌1"], "")
        self.assertEqual(false_positive_row["型号1"], "")
        self.assertEqual(false_positive_row["备注1"], "")
        self.assertEqual(false_positive_row["备注2"], "")
        self.assertIn("指定品牌（信昌PDC）", false_positive_row["推荐理由"])
        self.assertIn("原厂无对应型号或当前数据库资料不足", false_positive_row["差异说明"])

        generic_slots = app["build_bom_own_brand_export_slots"](
            pd.DataFrame(
                [
                    {
                        "品牌": "村田Murata",
                        "型号": "LQH32PN100MN0L",
                        "器件类型": "功率电感",
                        "推荐等级": "完全匹配",
                        "成本": "0.5",
                    }
                ]
            ),
            spec={"器件类型": "功率电感"},
            export_settings={"mode": app["BOM_EXPORT_MODE_AUTO"]},
        )
        self.assertEqual(generic_slots["自有品牌"], "村田Murata")
        self.assertEqual(generic_slots["自有型号"], "LQH32PN100MN0L")

        signature_a = app["build_bom_workbook_run_signature"](
            UploadedBytes("same.xlsx", b"same"),
            {"Sheet1": {"model": "型号"}},
            export_settings=custom_settings,
        )
        signature_b = app["build_bom_workbook_run_signature"](
            UploadedBytes("same.xlsx", b"same"),
            {"Sheet1": {"model": "型号"}},
            export_settings={"mode": app["BOM_EXPORT_MODE_CUSTOM"], "brands": ["国巨YAGEO"]},
        )
        self.assertNotEqual(signature_a, signature_b)
        display = app["build_bom_display_df"](result)
        forbidden = {
            "销售结论",
            "备选型号",
            "风险提示",
            "推荐理由",
            "解析说明",
            "客户回复型号",
            "可直接回复客户",
        }
        self.assertTrue(forbidden.isdisjoint(set(display.columns)))

    def test_06b_bom_selected_brand_exports_active_cost(self):
        app = self.app
        self.assertFalse(
            app["should_start_bom_matching"]("old", "new", app["BOM_EXPORT_MODE_AUTO"])
        )
        self.assertTrue(
            app["should_start_bom_matching"](
                "old",
                "new",
                app["BOM_EXPORT_MODE_AUTO"],
                start_clicked=True,
            )
        )
        self.assertFalse(
            app["should_start_bom_matching"]("same", "same", app["BOM_EXPORT_MODE_AUTO"])
        )
        self.assertTrue(
            app["should_start_bom_matching"](
                "same",
                "same",
                app["BOM_EXPORT_MODE_AUTO"],
                start_clicked=True,
            )
        )
        self.assertFalse(
            app["should_start_bom_matching"]("old", "new", app["BOM_EXPORT_MODE_CUSTOM"])
        )
        self.assertTrue(
            app["should_start_bom_matching"](
                "old",
                "new",
                app["BOM_EXPORT_MODE_CUSTOM"],
                custom_start_clicked=True,
            )
        )
        self.assertFalse(app["bom_output_selection_ready"]("", []))
        self.assertTrue(
            app["bom_output_selection_ready"](app["BOM_EXPORT_MODE_AUTO"], [])
        )
        self.assertFalse(
            app["bom_output_selection_ready"](app["BOM_EXPORT_MODE_CUSTOM"], [])
        )
        self.assertTrue(
            app["bom_output_selection_ready"](
                app["BOM_EXPORT_MODE_CUSTOM"],
                ["信昌PDC"],
            )
        )
        self.assertFalse(
            app["should_start_bom_matching"]("same", "same", app["BOM_EXPORT_MODE_CUSTOM"])
        )
        self.assertTrue(
            app["should_start_bom_matching"](
                "same",
                "same",
                app["BOM_EXPORT_MODE_CUSTOM"],
                custom_start_clicked=True,
            )
        )
        original_cost_path = app["COST_PRICE_DB_PATH"]
        try:
            app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "bom-selected-brand-cost.sqlite")
            app["clear_cost_price_lookup_cache"]()
            cost_upload = UploadedBytes(
                "bom-cost.xlsx",
                dataframe_to_xlsx_bytes(
                    pd.DataFrame(
                        [
                            {
                                "品牌": "村田Murata",
                                "型号": "GRM155R71C224KA12D",
                                "成本": "0.123",
                                "MOQ": "10000PCS",
                                "L&T": "6W",
                            }
                        ]
                    )
                ),
            )
            ok, message, _ = app["import_cost_price_list_from_upload"](cost_upload, "regression")
            self.assertTrue(ok, message)
            result = app["bom_dataframe_from_upload"](
                None,
                pd.DataFrame([{"型号": "GRM155R71C224KA12D"}]),
                {"model": "型号", "spec": None, "name": None, "quantity": None},
                export_settings={
                    "mode": app["BOM_EXPORT_MODE_CUSTOM"],
                    "brands": ["村田Murata"],
                },
            )
            self.assertEqual(len(result), 1)
            self.assertEqual(result.iloc[0]["自有品牌"], "村田Murata", result.to_dict(orient="records"))
            self.assertEqual(result.iloc[0]["自有型号"], "GRM155R71C224KA12D")
            self.assertEqual(result.iloc[0]["自有成本"], "0.123")
            self.assertEqual(result.iloc[0]["自有MOQ"], "10000PCS")
            self.assertEqual(result.iloc[0]["自有L&T"], "6W")
            ok, message, _ = app["save_manual_cost_price_item"](
                brand="村田",
                model="GRM155R71C224KA12D",
                cost="0.111",
                moq="12000PCS",
                lead_time="2W",
                note="原厂单独报价",
                updated_by="regression",
            )
            self.assertTrue(ok, message)
            result = app["bom_dataframe_from_upload"](
                None,
                pd.DataFrame([{"型号": "GRM155R71C224KA12D"}]),
                {"model": "型号", "spec": None, "name": None, "quantity": None},
                export_settings={
                    "mode": app["BOM_EXPORT_MODE_CUSTOM"],
                    "brands": ["村田Murata"],
                },
            )
            self.assertEqual(result.iloc[0]["自有成本"], "0.111")
            self.assertEqual(result.iloc[0]["自有MOQ"], "12000PCS")
            self.assertEqual(result.iloc[0]["自有L&T"], "2W")
            source_df = pd.DataFrame([{"型号": "GRM155R71C224KA12D"}])
            source_upload = UploadedBytes("selected-brand.xlsx", dataframe_to_xlsx_bytes(source_df))
            source_workbook = app["read_uploaded_bom_workbook"](source_upload)
            export_bytes = app["bom_to_excel_bytes"](
                result,
                source_df,
                source_workbook=source_workbook,
                sheet_results=[
                    {
                        "sheet_name": source_workbook["sheet_frames"][0]["sheet_name"],
                        "source_df": source_df,
                        "result_df": result,
                    }
                ],
            )
            exported_workbook = load_workbook(BytesIO(export_bytes), data_only=False)
            exported_sheet = exported_workbook.active
            headers = [exported_sheet.cell(row=1, column=idx).value for idx in range(1, exported_sheet.max_column + 1)]
            values = {
                headers[idx - 1]: exported_sheet.cell(row=2, column=idx).value
                for idx in range(1, exported_sheet.max_column + 1)
            }
            self.assertEqual(values["匹配状态"], "可推荐")
            self.assertEqual(values["匹配品牌"], "村田Murata")
            self.assertEqual(values["匹配型号"], "GRM155R71C224KA12D")
            self.assertEqual(str(values["匹配成本"]), "0.111")
            first_brand_col = headers.index("匹配品牌")
            self.assertEqual(
                headers[first_brand_col : first_brand_col + 8],
                [
                    "匹配品牌",
                    "匹配型号",
                    "匹配成本",
                    "成本更新时间",
                    "匹配MOQ",
                    "匹配L&T",
                    "匹配说明",
                    "匹配备注",
                ],
            )
            self.assertNotEqual(values["匹配说明"], "")
            exported_workbook.close()

            restricted_export_bytes = app["bom_to_excel_bytes"](
                result,
                source_df,
                source_workbook=source_workbook,
                sheet_results=[
                    {
                        "sheet_name": source_workbook["sheet_frames"][0]["sheet_name"],
                        "source_df": source_df,
                        "result_df": result,
                    }
                ],
                include_cost=False,
            )
            restricted_workbook = load_workbook(BytesIO(restricted_export_bytes), data_only=False)
            restricted_sheet = restricted_workbook.active
            restricted_headers = [
                restricted_sheet.cell(row=1, column=idx).value
                for idx in range(1, restricted_sheet.max_column + 1)
            ]
            self.assertNotIn("匹配成本", restricted_headers)
            self.assertNotIn("成本更新时间", restricted_headers)
            self.assertIn("匹配型号", restricted_headers)
            self.assertIn("匹配MOQ", restricted_headers)
            restricted_workbook.close()
        finally:
            app["COST_PRICE_DB_PATH"] = original_cost_path
            app["clear_cost_price_lookup_cache"]()

    def test_06bb_bom_xlsx_export_preserves_original_format(self):
        app = self.app
        raw_bytes = formatted_bom_xlsx_bytes()
        source_df = pd.DataFrame(
            [
                {"型号": "RC0402FR-071KL", "规格": "0402 1KΩ ±1% 1/16W", "数量": "12000"},
                {"型号": "GRM155R71C104KA88D", "规格": "0402 X7R 100nF 16V", "数量": "8000"},
            ]
        )
        result_df = pd.DataFrame(
            [
                {
                    "状态": "可推荐",
                    "自有品牌": "富捷",
                    "自有型号": "FRC0402F1001TS",
                    "自有匹配说明": "关键规格完全一致",
                },
                {
                    "状态": "需确认",
                    "自有品牌": "信昌PDC",
                    "自有型号": "CC0402KRX7R7BB104",
                    "自有匹配说明": "需确认厚度",
                },
            ]
        )
        export_bytes = app["bom_to_excel_bytes"](
            result_df,
            source_df,
            source_workbook={
                "kind": "excel",
                "file_name": "格式BOM.xlsx",
                "file_bytes": raw_bytes,
            },
            sheet_results=[
                {
                    "sheet_name": "格式BOM",
                    "source_df": source_df,
                    "result_df": result_df,
                }
            ],
        )

        source_workbook = load_workbook(BytesIO(raw_bytes), data_only=False, rich_text=True)
        exported_workbook = load_workbook(BytesIO(export_bytes), data_only=False, rich_text=True)
        source_sheet = source_workbook["格式BOM"]
        exported_sheet = exported_workbook["格式BOM"]
        original_max_column = source_sheet.max_column

        self.assertGreater(exported_sheet.max_column, original_max_column)
        self.assertEqual(exported_sheet.cell(1, original_max_column + 1).value, "匹配状态")
        self.assertEqual(exported_sheet.cell(2, original_max_column + 1).value, "可推荐")
        self.assertEqual(exported_sheet.freeze_panes, source_sheet.freeze_panes)
        self.assertEqual(exported_sheet.auto_filter.ref, source_sheet.auto_filter.ref)
        self.assertEqual(
            {str(item) for item in exported_sheet.merged_cells.ranges},
            {str(item) for item in source_sheet.merged_cells.ranges},
        )
        self.assertEqual(exported_sheet.sheet_view.showGridLines, source_sheet.sheet_view.showGridLines)
        self.assertEqual(exported_sheet.page_setup.orientation, source_sheet.page_setup.orientation)
        self.assertEqual(exported_sheet.print_title_rows, source_sheet.print_title_rows)
        self.assertEqual(
            exported_sheet.sheet_properties.tabColor.rgb,
            source_sheet.sheet_properties.tabColor.rgb,
        )

        for column_name in ["A", "B", "C", "D", "E", "F"]:
            source_dimension = source_sheet.column_dimensions[column_name]
            exported_dimension = exported_sheet.column_dimensions[column_name]
            self.assertEqual(exported_dimension.width, source_dimension.width, column_name)
            self.assertEqual(exported_dimension.hidden, source_dimension.hidden, column_name)
            self.assertEqual(exported_dimension.outlineLevel, source_dimension.outlineLevel, column_name)
        for row_idx in [1, 2, 3]:
            source_dimension = source_sheet.row_dimensions[row_idx]
            exported_dimension = exported_sheet.row_dimensions[row_idx]
            self.assertEqual(exported_dimension.height, source_dimension.height, row_idx)
            self.assertEqual(exported_dimension.hidden, source_dimension.hidden, row_idx)
            for column_idx in range(1, original_max_column + 1):
                source_cell = source_sheet.cell(row=row_idx, column=column_idx)
                exported_cell = exported_sheet.cell(row=row_idx, column=column_idx)
                self.assertEqual(exported_cell.value, source_cell.value, source_cell.coordinate)
                self.assertEqual(exported_cell._style, source_cell._style, source_cell.coordinate)
                self.assertEqual(
                    getattr(exported_cell.hyperlink, "target", None),
                    getattr(source_cell.hyperlink, "target", None),
                    source_cell.coordinate,
                )

        source_workbook.close()
        exported_workbook.close()

        legacy_raw_bytes = b"legacy-biff-source-remains-unchanged"
        legacy_export_bytes = app["bom_to_excel_bytes"](
            result_df,
            source_df,
            source_workbook={
                "kind": "excel",
                "file_name": "旧版.xls",
                "file_bytes": legacy_raw_bytes,
            },
            sheet_results=[
                {
                    "sheet_name": "原分页",
                    "source_df": source_df,
                    "result_df": result_df,
                }
            ],
        )
        self.assertEqual(legacy_raw_bytes, b"legacy-biff-source-remains-unchanged")
        legacy_result_workbook = load_workbook(BytesIO(legacy_export_bytes), data_only=False)
        self.assertEqual(legacy_result_workbook.sheetnames, ["原分页"])
        legacy_result_sheet = legacy_result_workbook["原分页"]
        legacy_headers = [
            legacy_result_sheet.cell(row=1, column=column_idx).value
            for column_idx in range(1, legacy_result_sheet.max_column + 1)
        ]
        self.assertEqual(legacy_headers[:3], ["型号", "规格", "数量"])
        self.assertIn("匹配品牌", legacy_headers)
        self.assertIn("匹配型号", legacy_headers)
        self.assertEqual(
            legacy_result_sheet.cell(row=2, column=legacy_headers.index("匹配型号") + 1).value,
            "FRC0402F1001TS",
        )
        self.assertEqual(legacy_result_sheet.freeze_panes, "A2")
        legacy_result_workbook.close()

    def test_06c_bom_matching_reuses_bounded_cache_and_rich_candidates(self):
        app = self.app
        candidates = app["build_bom_query_candidates"](
            "GRM155R71C224KA12D",
            "0402 220nF 16V X7R 10%",
            "贴片电容",
            extra_values=["车规"],
        )
        sources = [item["source"] for item in candidates]
        self.assertEqual(sources[0], "型号列")
        self.assertLess(sources.index("型号列+规格列+品名列"), sources.index("规格列"))
        self.assertLess(sources.index("规格列+品名列+其他列"), sources.index("品名列"))

        cache = {}
        for index in range(300):
            app["store_bom_query_cache"](cache, f"Q{index}", {"index": index}, limit=256)
        self.assertEqual(len(cache), 256)
        self.assertNotIn("Q0", cache)
        self.assertEqual(app["bom_query_cache_key"](" 0402   10k  1% "), "0402 10K 1%")

        skipped = app["build_bom_upload_result_row"](
            None,
            0,
            {"型号": "MPN3", "规格": "Description", "品名": "项目"},
            {"model": "型号", "spec": "规格", "name": "品名", "quantity": None},
            query_cache={},
        )
        self.assertEqual(skipped["状态"], "已跳过")
        self.assertIn("重复表头", skipped["失败原因"])

        original_enrich_cost = app["enrich_component_cost_columns"]
        enriched_brands = []

        def capture_enrich_cost(frame):
            enriched_brands.extend(frame["品牌"].astype(str).tolist())
            return frame.copy()

        try:
            app["enrich_component_cost_columns"] = capture_enrich_cost
            app["build_bom_own_brand_export_slots"](
                pd.DataFrame(
                    [
                        {"品牌": "华新科Walsin", "型号": "0402B103K500CT", "器件类型": "MLCC"},
                        {"品牌": "信昌PDC", "型号": "FM05X103K500EGG", "器件类型": "MLCC"},
                        {"品牌": "村田Murata", "型号": "GRM155R71H103KA88D", "器件类型": "MLCC"},
                    ]
                ),
                spec={"器件类型": "MLCC"},
                export_settings={"mode": app["BOM_EXPORT_MODE_CUSTOM"], "brands": ["华新科Walsin"]},
            )
        finally:
            app["enrich_component_cost_columns"] = original_enrich_cost
        self.assertEqual(enriched_brands, ["华新科Walsin"])

        original_cache_signature = app["get_query_cache_signature"]
        try:
            app["get_query_cache_signature"] = lambda: self.fail(
                "BOM matching must not recalculate the interactive search cache signature"
            )
            direct_bom_result = app["evaluate_bom_candidate"](
                None,
                "0402 1K ±1% 1/16W",
                "规格列",
                0,
                query_cache={},
                export_settings={
                    "mode": app["BOM_EXPORT_MODE_CUSTOM"],
                    "brands": ["富捷"],
                },
            )
        finally:
            app["get_query_cache_signature"] = original_cache_signature
        self.assertEqual(direct_bom_result["status"], "可推荐")
        self.assertFalse(direct_bom_result["matched"].empty)

        original_bom_dataframe = app["bom_dataframe_from_upload"]
        seen_cache_ids = []

        def fake_bom_dataframe(_df, sheet_df, _mapping, **kwargs):
            seen_cache_ids.append(id(kwargs.get("query_cache")))
            return pd.DataFrame(index=sheet_df.index)

        try:
            app["bom_dataframe_from_upload"] = fake_bom_dataframe
            workbook = {
                "sheet_frames": [
                    {"sheet_name": "A", "df": pd.DataFrame([{"型号": "A1"}])},
                    {"sheet_name": "B", "df": pd.DataFrame([{"型号": "B1"}])},
                ]
            }
            app["build_bom_workbook_sheet_results"](
                workbook,
                sheet_mappings={
                    "A": {"model": "型号", "spec": None, "name": None, "quantity": None},
                    "B": {"model": "型号", "spec": None, "name": None, "quantity": None},
                },
            )
        finally:
            app["bom_dataframe_from_upload"] = original_bom_dataframe
        self.assertEqual(len(seen_cache_ids), 2)
        self.assertEqual(seen_cache_ids[0], seen_cache_ids[1])

    def test_06d_large_bom_matching_is_bounded_ordered_and_resumable(self):
        app = self.app
        original_builder = app["build_bom_upload_result_row"]
        active = 0
        max_active = 0
        calls = []
        state_lock = threading.Lock()

        def fake_builder(_df, row_index, record, _mapping, **_kwargs):
            nonlocal active, max_active
            with state_lock:
                active += 1
                max_active = max(max_active, active)
                calls.append(row_index)
            time.sleep(0.01)
            with state_lock:
                active -= 1
            return {
                "BOM行号": row_index + 2,
                "BOM型号": record["型号"],
                "状态": "可推荐",
                "首选推荐等级": "完全匹配",
            }

        checkpoints = []
        source = pd.DataFrame([{"型号": f"P{index:03d}"} for index in range(30)])
        try:
            app["build_bom_upload_result_row"] = fake_builder
            result = app["bom_dataframe_from_upload"](
                None,
                source,
                {"model": "型号", "spec": None, "name": None, "quantity": None},
                checkpoint_callback=lambda rows: checkpoints.append(rows),
                max_workers=4,
            )
            self.assertLessEqual(max_active, 4)
            self.assertEqual(result["BOM型号"].tolist(), source["型号"].tolist())
            self.assertTrue(checkpoints)
            self.assertEqual(len(checkpoints[-1]), len(source))

            calls.clear()
            resumed = app["bom_dataframe_from_upload"](
                None,
                source,
                {"model": "型号", "spec": None, "name": None, "quantity": None},
                resume_rows=checkpoints[-1],
                max_workers=4,
            )
            self.assertEqual(calls, [])
            self.assertEqual(resumed["BOM型号"].tolist(), source["型号"].tolist())

            calls.clear()
            duplicate_source = pd.DataFrame([
                {"型号": "RC0402FR-071KL", "数量": 100},
                {"型号": "RC0402FR-071KL", "数量": 200},
                {"型号": "RC0402FR-071KL", "数量": 300},
            ])
            deduped = app["bom_dataframe_from_upload"](
                None, duplicate_source,
                {"model": "型号", "spec": None, "name": None, "quantity": "数量"},
                max_workers=4,
            )
            self.assertEqual(len(calls), 1)
            self.assertEqual(deduped["BOM数量"].tolist(), ["100", "200", "300"])
            self.assertEqual(deduped.attrs["bom_metrics"]["unique_rows"], 1)
            self.assertEqual(deduped.attrs["bom_metrics"]["deduped_rows"], 2)
        finally:
            app["build_bom_upload_result_row"] = original_builder

    def test_06e_user_reported_golden_queries_remain_parseable(self):
        golden_path = os.path.join(self.base_dir, "tests", "golden_user_cases.json")
        with open(golden_path, "r", encoding="utf-8") as handle:
            cases = json.load(handle)
        for case in cases:
            with self.subTest(query=case["query"]):
                mode, spec = self.app["detect_query_mode_and_spec"](None, case["query"])
                self.assertNotIn(mode, {"无法识别", "暂不支持"})
                self.assertIsNotNone(spec)
                for field_name, expected_value in case.get("expected", {}).items():
                    actual_value = (spec or {}).get(field_name, "")
                    if isinstance(expected_value, (int, float)):
                        self.assertAlmostEqual(float(actual_value), float(expected_value))
                    else:
                        actual = self.app["clean_text"](actual_value)
                        self.assertEqual(actual.upper(), str(expected_value).upper())

    def test_07_member_database_remote_snapshot_survives_instance_reset(self):
        app = self.app
        snapshot = {"version": 0, "sha256": "", "payload_base64": "", "updated_at": ""}
        request_counts = {"get": 0, "put": 0}
        put_delay = {"seconds": 0.0}
        api_secret = "regression-secret"

        class SnapshotHandler(BaseHTTPRequestHandler):
            def log_message(self, *_args):
                return

            def _authorized(self):
                return self.headers.get("Authorization", "") == f"Bearer {api_secret}"

            def _send(self, status, payload):
                encoded = json.dumps(payload).encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(encoded)))
                self.end_headers()
                self.wfile.write(encoded)

            def do_GET(self):
                if not self._authorized():
                    self._send(401, {"error": "unauthorized"})
                    return
                request_counts["get"] += 1
                self._send(200, snapshot)

            def do_PUT(self):
                if not self._authorized():
                    self._send(401, {"error": "unauthorized"})
                    return
                request_counts["put"] += 1
                time.sleep(float(put_delay["seconds"]))
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length).decode("utf-8"))
                if int(body.get("expected_version") or 0) != int(snapshot["version"]):
                    self._send(409, {"error": "version_conflict", "version": snapshot["version"]})
                    return
                snapshot.update(
                    {
                        "version": snapshot["version"] + 1,
                        "sha256": body["sha256"],
                        "payload_base64": body["payload_base64"],
                        "updated_at": "2026-06-29T00:00:00Z",
                    }
                )
                self._send(200, {"ok": True, "version": snapshot["version"], "sha256": snapshot["sha256"]})

        server = ThreadingHTTPServer(("127.0.0.1", 0), SnapshotHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        saved_env = {key: os.environ.get(key) for key in ["MEMBER_AUTH_REMOTE_API_URL", "MEMBER_AUTH_REMOTE_API_SECRET", "MEMBER_AUTH_REMOTE_FORCE"]}
        original_state_path = app["MEMBER_AUTH_REMOTE_STATE_PATH"]
        try:
            os.environ["MEMBER_AUTH_REMOTE_API_URL"] = f"http://127.0.0.1:{server.server_port}/api/member-store/snapshot"
            os.environ["MEMBER_AUTH_REMOTE_API_SECRET"] = api_secret
            os.environ["MEMBER_AUTH_REMOTE_FORCE"] = "1"
            app["MEMBER_AUTH_REMOTE_STATE_PATH"] = os.path.join(self.temp_dir, "remote_state.json")
            ok, message = app["create_member_account"]("DurableUser", "secret1", "Durable User")
            self.assertTrue(ok, message)
            member = app["get_member_by_username"]("durableuser")
            app["approve_member_account_admin"](member["id"])
            self.assertGreaterEqual(snapshot["version"], 1)

            with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
                conn.execute("DELETE FROM members WHERE lower(username)=lower('DurableUser')")
                conn.commit()
            with sqlite3.connect(app["MEMBER_AUTH_DB_PATH"]) as conn:
                self.assertIsNone(
                    conn.execute(
                        "SELECT id FROM members WHERE lower(username)=lower('DurableUser')"
                    ).fetchone()
                )
            app["reset_member_auth_remote_refresh_cache"]()
            self.assertIsNotNone(app["get_member_by_username"]("DurableUser"))
            app["ensure_configured_admin_member_account"]()
            app["reset_member_auth_remote_refresh_cache"]()
            requests_before_login = dict(request_counts)
            app["initialize_member_auth_remote_storage"]()
            put_delay["seconds"] = 2.0
            login_started_at = time.perf_counter()
            restored, message = app["authenticate_member"]("DURABLEUSER", "secret1")
            login_elapsed = time.perf_counter() - login_started_at
            self.assertIsNotNone(restored, message)
            self.assertLess(login_elapsed, 1.5)
            self.assertEqual(request_counts["get"] - requests_before_login["get"], 1)
            self.assertTrue(app["wait_for_member_auth_remote_snapshot_flush"](timeout=5.0))
            put_delay["seconds"] = 0.0
            self.assertEqual(request_counts["put"] - requests_before_login["put"], 1)
            requests_after_login = dict(request_counts)
            for _ in range(3):
                self.assertIsNotNone(app["get_member_by_session_token"](restored["_session_token"]))
            self.assertEqual(request_counts, requests_after_login)

            remote_without_session_path = os.path.join(self.temp_dir, "remote-without-current-session.sqlite")
            with open(remote_without_session_path, "wb") as handle:
                handle.write(base64.b64decode(snapshot["payload_base64"]))
            with sqlite3.connect(remote_without_session_path) as conn:
                conn.execute(
                    "DELETE FROM member_sessions WHERE token=?",
                    (restored["_session_token"],),
                )
                conn.commit()
            with open(remote_without_session_path, "rb") as handle:
                stale_session_payload = handle.read()
            snapshot.update(
                {
                    "version": int(snapshot["version"]) + 1,
                    "sha256": hashlib.sha256(stale_session_payload).hexdigest(),
                    "payload_base64": base64.b64encode(stale_session_payload).decode("ascii"),
                    "updated_at": "2026-06-29T00:01:00Z",
                }
            )
            app["reset_member_auth_remote_refresh_cache"]()
            self.assertIsNotNone(
                app["get_member_by_session_token"](restored["_session_token"]),
                "a newer remote member snapshot must not erase an unexpired local login session",
            )
            self.assertTrue(app["wait_for_member_auth_remote_snapshot_flush"](timeout=5.0))

            stale_path = os.path.join(self.temp_dir, "member-stale.sqlite")
            shutil.copy2(app["MEMBER_AUTH_DB_PATH"], stale_path)
            ok, message = app["create_member_account"]("OtherInstanceUser", "secret2")
            self.assertTrue(ok, message)
            original_db_path = app["MEMBER_AUTH_DB_PATH"]
            app["MEMBER_AUTH_DB_PATH"] = stale_path
            try:
                usernames = {row["username"] for row in app["list_members_for_admin"]()}
                self.assertIn("OtherInstanceUser", usernames)
            finally:
                app["MEMBER_AUTH_DB_PATH"] = original_db_path
        finally:
            app["MEMBER_AUTH_REMOTE_STATE_PATH"] = original_state_path
            for key, value in saved_env.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value
            server.shutdown()
            server.server_close()

    def test_08_fojan_matrix_quote_imports_tolerance_price_rules(self):
        app = self.app
        app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "fojan-cost.sqlite")
        app["clear_cost_price_lookup_cache"]()
        upload = UploadedBytes("fojan-quote.xlsx", fojan_quote_xlsx_bytes())
        items, error = app["build_cost_price_items_from_workbook"](upload)
        self.assertEqual(error, "")
        self.assertEqual(len(items), 5)
        self.assertEqual({item["brand"] for item in items}, {"FOJAN(富捷)"})

        ok, message, _ = app["import_cost_price_list_from_upload"](upload, "regression")
        self.assertTrue(ok, message)
        lookup = app["load_active_cost_price_lookup"]()

        def price(model, resistance_ohm, tolerance):
            return app["lookup_active_cost_price_for_row"](
                {
                    "品牌": "FOJAN(富捷)",
                    "型号": model,
                    "器件类型": "贴片电阻",
                    "尺寸（inch）": "0603",
                    "功率": "1/10W",
                    "_resistance_ohm": resistance_ohm,
                    "容值误差": tolerance,
                },
                lookup=lookup,
            ).get("cost", "")

        self.assertEqual(price("FRC0603J100 TS", 10.0, "5"), "2.80")
        self.assertEqual(price("FRC0603J103 TS", 10000.0, "5"), "2.60")
        self.assertEqual(price("FRC0603F1002 TS", 10000.0, "1"), "3.10")
        self.assertEqual(price("FRC0603F8R20 TS", 8.2, "1"), "3.63")
        self.assertEqual(price("FRC0603F0000 TS", 0.0, "1"), "3.10")

        expected_zero_prices = {}
        for rule in app["load_resistor_series_pricing_rules"]():
            if rule.get("series") != "FRC":
                continue
            expected = app["select_resistor_segment_price"](
                rule.get("range", ""),
                rule.get("price_1", ""),
                10.0,
            )
            if expected:
                expected_zero_prices[rule["type_dimension_norm"]] = (expected, rule.get("package", ""))
        self.assertTrue(expected_zero_prices)
        for type_dimension, (expected_cost, expected_moq) in expected_zero_prices.items():
            size, power = type_dimension.split(" ", 1)
            zero_price = app["lookup_resistor_series_pricing"](
                {
                    "\u54c1\u724c": "FOJAN(\u5bcc\u6377)",
                    "\u578b\u53f7": f"FRC{size}F0000TS",
                    "\u5668\u4ef6\u7c7b\u578b": "\u539a\u819c\u7535\u963b",
                    "\u7cfb\u5217": "FRC",
                    "\u5c3a\u5bf8\uff08inch\uff09": size,
                    "\u529f\u7387": power,
                    "_res_ohm": 0.0,
                    "\u5bb9\u503c\u8bef\u5dee": "1",
                }
            )
            self.assertEqual(zero_price["\u6210\u672c"], expected_cost, type_dimension)
            self.assertEqual(zero_price["MOQ"], expected_moq, type_dimension)

        missing_range_model = app["build_rule_fallback_row_from_model"]("FRC0402F5233TS")
        self.assertEqual(len(missing_range_model), 1)
        fallback_row = missing_range_model.iloc[0]
        self.assertEqual(fallback_row["品牌"], "FOJAN(富捷)")
        self.assertEqual(fallback_row["系列"], "FRC")
        self.assertEqual(fallback_row["尺寸（inch）"], "0402")
        self.assertEqual(fallback_row["容值误差"], "1")
        self.assertAlmostEqual(float(fallback_row["_res_ohm"]), 523000.0)
        fallback_display = app["select_component_display_columns"](
            missing_range_model,
            fallback_row.to_dict(),
            prefix_columns=["品牌", "型号", "器件类别", "系列"],
        )
        self.assertEqual(fallback_display.iloc[0]["品牌"], "FOJAN(富捷)")
        self.assertEqual(fallback_display.iloc[0]["成本"], "1.7")
        self.assertEqual(fallback_display.iloc[0]["MOQ"], "10000PCS")

        for invalid_model in (
            "FRC0402F5243TS",
            "FRC0402F9993TS",
            "FRC0402F0003TS",
            "FRL0402F5233TS",
        ):
            self.assertTrue(app["build_rule_fallback_row_from_model"](invalid_model).empty, invalid_model)

        mode, spec = app["detect_query_mode_and_spec"](
            pd.DataFrame(),
            "0402 523K\u03a9 1% 1/16W \u539a\u819c\u7535\u963b",
        )
        self.assertEqual(mode, "\u539a\u819c\u7535\u963b")
        spec_rows = app["load_search_dataframe_for_query"](mode, spec)
        spec_fojan = spec_rows[
            spec_rows["\u54c1\u724c"].astype(str).str.contains("FOJAN|\u5bcc\u6377", case=False, regex=True)
        ]
        self.assertIn(
            "FRC0402F5233TS",
            set(spec_fojan["\u578b\u53f7"].map(app["clean_model"])),
        )
        spec_price = app["lookup_resistor_series_pricing"](spec_fojan.iloc[0].to_dict())
        self.assertEqual(spec_price["\u6210\u672c"], "1.7")

    def test_08a_fojan_multi_sheet_quote_imports_every_series_and_tolerance(self):
        app = self.app
        self.assertEqual(app["normalize_cost_price_tolerance_header"](1), "1")
        self.assertEqual(app["normalize_cost_price_tolerance_header"](0.01), "1")
        self.assertEqual(app["normalize_resistor_pricing_type_dimension"]("25121W"), "2512 1W")
        self.assertEqual(app["normalize_resistor_pricing_type_dimension"]("06031/10W"), "0603 1/10W")
        self.assertEqual(app["normalize_cost_price_tolerance_header"]("0.5%（D）"), "0.5")
        app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "fojan-multi-sheet-cost.sqlite")
        app["clear_cost_price_lookup_cache"]()
        upload = UploadedBytes("fojan-multi-sheet-quote.xlsx", fojan_multi_sheet_quote_xlsx_bytes())
        items, error = app["build_cost_price_items_from_workbook"](upload)
        self.assertEqual(error, "")
        self.assertEqual(len(items), 6)
        self.assertEqual({item["sheet_name"] for item in items}, {"FRC&FRL", "FRH", "FRQ"})
        rules = [json.loads(item["raw_json"]) for item in items]
        self.assertEqual({rule["series"] for rule in rules}, {"FRC", "FRH", "FRQ"})
        self.assertEqual({rule["tolerance"] for rule in rules}, {"5", "1", "0.5", "0.1"})

        ok, message, _ = app["import_cost_price_list_from_upload"](upload, "regression")
        self.assertTrue(ok, message)
        self.assertIn("覆盖 3 个分页", message)
        lookup = app["load_active_cost_price_lookup"]()

        def price(series, tolerance):
            return app["lookup_active_cost_price_for_row"](
                {
                    "品牌": "FOJAN(富捷)",
                    "型号": f"{series}0603F1002TS",
                    "器件类型": "厚膜电阻",
                    "系列": series,
                    "尺寸（inch）": "0603",
                    "功率": "1/10W",
                    "_resistance_ohm": 10000.0,
                    "容值误差": tolerance,
                },
                lookup=lookup,
            ).get("cost", "")

        self.assertEqual(price("FRC", "1"), "3.10")
        self.assertEqual(price("FRH", "0.5"), "7.90")
        self.assertEqual(price("FRH", "0.1"), "19.00")
        self.assertEqual(price("FRQ", "5"), "4.20")

    def test_08b_fojan_alloy_quote_imports_vertical_milliohm_rules(self):
        app = self.app
        original_cost_path = app["COST_PRICE_DB_PATH"]
        try:
            app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "fojan-alloy-cost.sqlite")
            app["clear_cost_price_lookup_cache"]()
            upload = UploadedBytes("fojan-alloy-quote.xlsx", fojan_alloy_quote_xlsx_bytes())
            items, error = app["build_cost_price_items_from_workbook"](upload)
            self.assertEqual(error, "")
            self.assertGreaterEqual(len(items), 10)
            self.assertEqual({item["sheet_name"] for item in items}, {"Alloy"})
            rules = [json.loads(item["raw_json"]) for item in items]
            self.assertIn("1mR-4mR", {rule["resistance_range"] for rule in rules})
            self.assertIn("101mR-500mR", {rule["resistance_range"] for rule in rules})
            self.assertEqual(
                {
                    rule["fojan_alloy_terminal"]
                    for rule in rules
                    if rule["series"] in {"FRM", "FPM"}
                },
                {"large", "standard"},
            )

            ok, message, _ = app["import_cost_price_list_from_upload"](upload, "regression")
            self.assertTrue(ok, message)
            self.assertIn("覆盖 1 个分页", message)
            lookup = app["load_active_cost_price_lookup"]()

            def price(model, series, size, power, resistance_ohm, tolerance):
                return app["lookup_active_cost_price_for_row"](
                    {
                        "品牌": "FOJAN(富捷)",
                        "型号": model,
                        "器件类型": "合金电阻",
                        "系列": series,
                        "尺寸（inch）": size,
                        "功率": power,
                        "_resistance_ohm": resistance_ohm,
                        "容值误差": tolerance,
                    },
                    lookup=lookup,
                )

            frm_large = price("FRM252WFR001TML", "FRM", "2512", "2W", 0.001, "1")
            self.assertEqual(frm_large["cost"], "124.2")
            self.assertEqual(frm_large["moq"], "4000PCS")

            frm_standard = price("FRM252WFR010TM", "FRM", "2512", "2W", 0.01, "1")
            self.assertEqual(frm_standard["cost"], "120.75")

            frm_high_range = price("FRM252WGR200TM", "FRM", "2512", "2W", 0.2, "2")
            self.assertEqual(frm_high_range["cost"], "112.7")

            fpm_large = price("FPM253WFR001TML", "FPM", "2512", "3W", 0.001, "1")
            self.assertEqual(fpm_large["cost"], "147.2")

            fpm_standard = price("FPM253WJR001TM", "FPM", "2512", "3W", 0.001, "5")
            self.assertEqual(fpm_standard["cost"], "78.2")

            frm_2010 = price("FRM2015FR010TM", "FRM", "2010", "1.5W", 0.01, "1")
            self.assertEqual(frm_2010["cost"], "135.7")

            frm_1206_large = price("FRM121WFR001TML", "FRM", "1206", "1W", 0.001, "1")
            self.assertEqual(frm_1206_large["cost"], "112.7")
            frm_1206_standard = price("FRM121WFR010TM", "FRM", "1206", "1W", 0.01, "1")
            self.assertEqual(frm_1206_standard["cost"], "83.95")

            fmh = price("FMH121WFR120TM", "FMH", "1206", "1W", 0.12, "1")
            self.assertEqual(fmh["cost"], "83.95")

            fcm_mid_power = price("FCM25125WF0M50TM", "FCM", "2512", "5W", 0.0005, "1")
            self.assertEqual(fcm_mid_power["cost"], "200.1")

            fwp = price("FWP27284WFR010TK", "FWP", "2728", "4W", 0.01, "1")
            self.assertEqual(fwp["cost"], "300.1")

            fwk = price("FWK12169WF0M50RK", "FWK", "1216", "9W", 0.0005, "1")
            self.assertEqual(fwk["cost"], "400.2")

            fwk_unsupported = price("FWK12169WFR003RK", "FWK", "1216", "9W", 0.003, "1")
            self.assertEqual(fwk_unsupported, {})
        finally:
            app["COST_PRICE_DB_PATH"] = original_cost_path
            app["clear_cost_price_lookup_cache"]()

    def test_09_pdc_series_descriptions_do_not_repeat_vendor_and_series(self):
        app = self.app
        profile = app["lookup_official_resistor_series_profile_by_model"](
            "FCF02FV-8062",
            "PSA(信昌电陶)",
        )
        self.assertEqual(profile.get("系列说明"), "通用厚膜贴片电阻/低阻电流检测贴片电阻")

        source = pd.DataFrame(
            [
                {"品牌": "PSA(信昌电陶)", "系列": "FCF", "系列说明": "PDC FCF 通用厚膜贴片电阻/低阻电流检测贴片电阻"},
                {"品牌": "PSA(信昌电陶)", "系列": "FCF", "系列说明": "PDC FCF-E 通用厚膜电流检测贴片电阻"},
                {"品牌": "PSA(信昌电陶)", "系列": "FWF", "系列说明": "PDC FWF 车规厚膜贴片电阻/抗硫化车规厚膜贴片电阻"},
                {"品牌": "华新科Walsin", "系列": "WR", "系列说明": "通用厚膜贴片电阻"},
            ]
        )
        formatted = app["format_display_df"](source)
        self.assertEqual(
            formatted["系列说明"].tolist(),
            [
                "通用厚膜贴片电阻/低阻电流检测贴片电阻",
                "通用厚膜电流检测贴片电阻",
                "车规厚膜贴片电阻/抗硫化车规厚膜贴片电阻",
                "通用厚膜贴片电阻",
            ],
        )

    def test_10_other_passive_specs_do_not_fall_back_to_wrong_models(self):
        app = self.app

        bead_query = "L2.0*W1.25*H0.9;2.0A;300Ω;±25%;CYBERMAX;CMBH2012S301NSP;SMD;0805;CAV"
        bead_mode, bead_spec = app["detect_query_mode_and_spec"](pd.DataFrame(), bead_query)
        self.assertEqual(bead_mode, "磁珠")
        self.assertEqual(bead_spec["器件类型"], "磁珠")
        self.assertEqual(bead_spec["容值"], "300")
        self.assertEqual(bead_spec["容值单位"], "Ω")

        inductor_query = "1000uH;±20%;180mA;6Ω max;L7.3*W7.3*H4.5;CYBERMAX;CMLH0704S102MTT;CAV"
        inductor_mode, inductor_spec = app["detect_query_mode_and_spec"](
            pd.DataFrame(), inductor_query
        )
        self.assertEqual(inductor_mode, "功率电感")
        self.assertEqual(inductor_spec["器件类型"], "功率电感")
        self.assertEqual(inductor_spec["容值"], "1000")
        self.assertEqual(inductor_spec["容值单位"], "UH")

        def match(rows, spec):
            frame = pd.DataFrame(rows)
            with sqlite3.connect(app["DB_PATH"]) as conn:
                source_columns = [row[1] for row in conn.execute('PRAGMA table_info("components")')]
            for column in source_columns:
                if column not in frame.columns:
                    frame[column] = ""
            prepared = app["prepare_search_dataframe"](frame)
            original_fetch = app["fetch_search_candidate_pairs"]
            app["fetch_search_candidate_pairs"] = lambda _spec: None
            try:
                return app["match_other_passive_spec"](prepared, spec)
            finally:
                app["fetch_search_candidate_pairs"] = original_fetch

        inductor_rows = [
            {
                "品牌": "Murata",
                "型号": "TEST-INDUCTOR-4R7",
                "器件类型": "功率电感",
                "容值": "4.7",
                "容值单位": "UH",
                "容值误差": "20",
            }
        ]
        self.assertTrue(
            match(
                inductor_rows,
                {"器件类型": "功率电感", "容值": "10", "容值单位": "UH", "容值误差": "20"},
            ).empty
        )
        self.assertEqual(
            match(
                inductor_rows,
                {"器件类型": "功率电感", "容值": "4.7", "容值单位": "UH", "容值误差": "20"},
            )["型号"].tolist(),
            ["TEST-INDUCTOR-4R7"],
        )

        varistor_rows = [
            {
                "品牌": "Littelfuse",
                "型号": "TEST-VARISTOR-470",
                "器件类型": "引线型压敏电阻",
                "耐压（V）": "470",
                "_varistor_voltage": "470",
                "_disc_size": "14D",
            }
        ]
        self.assertTrue(
            match(
                varistor_rows,
                {"器件类型": "引线型压敏电阻", "耐压（V）": "560", "_varistor_voltage": "560", "_disc_size": "14D"},
            ).empty
        )
        self.assertEqual(
            match(
                varistor_rows,
                {"器件类型": "引线型压敏电阻", "耐压（V）": "470", "_varistor_voltage": "470", "_disc_size": "14D"},
            )["型号"].tolist(),
            ["TEST-VARISTOR-470"],
        )

        mov_rows = [
            {
                "品牌": "Bourns",
                "型号": "MOV-14D471K",
                "器件类型": "引线型压敏电阻",
                "耐压（V）": "775",
                "压敏电压": "470",
                "直径（mm）": "14",
            },
            {
                "品牌": "Placeholder",
                "型号": "",
                "器件类型": "引线型压敏电阻",
                "压敏电压": "470",
                "直径（mm）": "14",
            },
        ]
        self.assertEqual(
            match(
                mov_rows,
                {"器件类型": "引线型压敏电阻", "耐压（V）": "470", "_varistor_voltage": "470", "_disc_size": "14D"},
            )["型号"].tolist(),
            ["MOV-14D471K"],
        )
        self.assertTrue(
            match(
                mov_rows,
                {"器件类型": "引线型压敏电阻", "耐压（V）": "775", "_varistor_voltage": "775", "_disc_size": "14D"},
            ).empty
        )

        common_mode_rows = [
            {
                "品牌": "Panasonic",
                "型号": "EXC14CE121U",
                "器件类型": "共模电感",
                "尺寸（inch）": "0302",
                "容值": "1.574",
                "容值单位": "NH",
                "电感值": "1.574",
                "电感单位": "NH",
                "共模阻抗": "120",
                "阻抗单位": "Ω",
            },
            {
                "品牌": "Panasonic",
                "型号": "EXC14CE900U",
                "器件类型": "共模电感",
                "尺寸（inch）": "0302",
                "共模阻抗": "90",
                "阻抗单位": "Ω",
            },
        ]
        self.assertEqual(
            match(
                common_mode_rows,
                {"器件类型": "共模电感", "容值": "120", "容值单位": "OHM"},
            )["型号"].tolist(),
            ["EXC14CE121U"],
        )
        parsed_common_mode = app["parse_inductor_spec_query"]("共模电感 0302 120OHM 100mA")
        self.assertEqual(parsed_common_mode["器件类型"], "共模电感")
        self.assertEqual(parsed_common_mode["尺寸（inch）"], "0302")
        self.assertEqual(parsed_common_mode["容值"], "120")
        self.assertEqual(parsed_common_mode["容值单位"], "Ω")
        self.assertEqual(parsed_common_mode["共模阻抗"], "120")
        self.assertEqual(parsed_common_mode["阻抗单位"], "Ω")
        self.assertEqual(
            match(common_mode_rows, parsed_common_mode)["型号"].tolist(),
            ["EXC14CE121U"],
        )

        crystal_rows = [
            {
                "品牌": "TXC",
                "型号": "TEST-CRYSTAL-16M",
                "器件类型": "晶振",
                "尺寸（inch）": "3225",
                "容值": "16",
                "容值单位": "MHZ",
                "容值误差": "20PPM",
                "负载电容（pF）": "12",
            }
        ]
        self.assertTrue(
            match(
                crystal_rows,
                {
                    "器件类型": "晶振",
                    "尺寸（inch）": "3225",
                    "容值": "16",
                    "容值单位": "MHZ",
                    "容值误差": "20PPM",
                    "负载电容（pF）": "8",
                },
            ).empty
        )
        self.assertEqual(
            match(
                crystal_rows,
                {
                    "器件类型": "晶振",
                    "尺寸（inch）": "3225",
                    "容值": "16",
                    "容值单位": "MHZ",
                    "容值误差": "20PPM",
                    "负载电容（pF）": "12",
                },
            )["型号"].tolist(),
            ["TEST-CRYSTAL-16M"],
        )

    def test_11_mlcc_special_use_terms_are_hard_constraints(self):
        app = self.app
        strict_queries = {
            "47nF 1210 630V 车规电容": "车规",
            "47nF 1210 630V 谐振电容": "谐振",
            "47nF 1210 630V 工业级电容": "工业",
            "47nF 1210 630V 软端电容": "软端子",
            "47nF 1210 630V 柔性端子电容": "软端子",
            "47nF 1210 630V FLEXITERM": "软端子",
            "47nF 1210 630V 车规软端电容": "车规/软端子",
            "47nF 1210 630V 次车规电容": "次车规",
            "47nF 1210 630V 高压电容": "高压",
            "47nF 1210 630V 中压电容": "中压",
            "47nF 1210 630V 抗弯电容": "抗弯",
            "47nF 1210 630V 安规电容": "安规",
            "47nF 1210 630V 高 Q 低损耗电容": "高Q",
            "47nF 1210 630V EMI 滤波电容": "EMI滤波",
        }
        for query_text, expected_class in strict_queries.items():
            with self.subTest(query=query_text):
                parsed = app["parse_spec_query"](query_text)
                self.assertEqual(parsed["特殊用途"], expected_class)
                self.assertTrue(app["mlcc_series_class_requires_filter"](expected_class))
                self.assertFalse(app["mlcc_series_class_matches"]("常规", expected_class))

        self.assertTrue(app["mlcc_series_class_matches"]("车规/软端子", "车规/软端子"))
        self.assertFalse(app["mlcc_series_class_matches"]("车规", "车规/软端子"))
        self.assertFalse(app["mlcc_series_class_matches"]("软端子", "车规/软端子"))

        query = "47nF 1210 630V 谐振电容"
        spec = app["parse_spec_query"](query)
        self.assertEqual(spec["特殊用途"], "谐振")
        self.assertEqual(app["infer_mlcc_series_class_from_spec"](spec), "谐振")
        spec_info = app["build_spec_info_df"](spec)
        self.assertIn("特殊用途", spec_info.columns)
        self.assertEqual(spec_info.iloc[0]["特殊用途"], "谐振")

        rows = pd.DataFrame(
            [
                {
                    "品牌": "ResonantBrand",
                    "型号": "RESONANT-1210-473-630V",
                    "器件类型": "MLCC",
                    "系列": "RZ",
                    "系列说明": "谐振 / Resonant MLCC",
                    "特殊用途": "谐振",
                    "尺寸（inch）": "1210",
                    "材质（介质）": "COG(NPO)",
                    "容值": "47",
                    "容值单位": "NF",
                    "耐压（V）": "630",
                    "_mlcc_series_class": "谐振",
                },
                {
                    "品牌": "GeneralBrand",
                    "型号": "GENERAL-X7R-1210-473-630V",
                    "器件类型": "MLCC",
                    "系列": "C",
                    "系列说明": "常规 / General-purpose MLCC",
                    "特殊用途": "",
                    "尺寸（inch）": "1210",
                    "材质（介质）": "X7R",
                    "容值": "47",
                    "容值单位": "NF",
                    "耐压（V）": "630",
                    "_mlcc_series_class": "常规",
                },
            ]
        )
        with sqlite3.connect(app["DB_PATH"]) as conn:
            source_columns = [row[1] for row in conn.execute('PRAGMA table_info("components")')]
        for column in source_columns:
            if column not in rows.columns:
                rows[column] = ""
        prepared = app["prepare_search_dataframe"](rows)
        original_fetch = app["fetch_search_candidate_pairs"]
        app["fetch_search_candidate_pairs"] = lambda _spec: None
        try:
            matched = app["match_by_partial_spec"](prepared, spec)
        finally:
            app["fetch_search_candidate_pairs"] = original_fetch
        self.assertEqual(matched["型号"].tolist(), ["RESONANT-1210-473-630V"])


    def test_12_runtime_databases_survive_instance_reset(self):
        app = self.app
        snapshots = {
            key: {"version": 0, "sha256": "", "payload_base64": "", "updated_at": ""}
            for key in ("cost-price", "no-match")
        }
        api_secret = "runtime-regression-secret"

        class RuntimeSnapshotHandler(BaseHTTPRequestHandler):
            def log_message(self, *_args):
                return

            def _store(self):
                return parse_qs(urlsplit(self.path).query).get("store", [""])[0]

            def _send(self, status, payload):
                encoded = json.dumps(payload).encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(encoded)))
                self.end_headers()
                self.wfile.write(encoded)

            def do_GET(self):
                if self.headers.get("Authorization", "") != f"Bearer {api_secret}":
                    self._send(401, {"error": "unauthorized"})
                    return
                store = self._store()
                self._send(200, {"store": store, **snapshots[store]})

            def do_PUT(self):
                if self.headers.get("Authorization", "") != f"Bearer {api_secret}":
                    self._send(401, {"error": "unauthorized"})
                    return
                store = self._store()
                length = int(self.headers.get("Content-Length", "0"))
                body = json.loads(self.rfile.read(length).decode("utf-8"))
                snapshot = snapshots[store]
                if int(body.get("expected_version") or 0) != int(snapshot["version"]):
                    self._send(409, {"error": "version_conflict", "version": snapshot["version"]})
                    return
                snapshot.update(
                    {
                        "version": snapshot["version"] + 1,
                        "sha256": body["sha256"],
                        "payload_base64": body["payload_base64"],
                        "updated_at": "2026-07-03T00:00:00Z",
                    }
                )
                self._send(200, {"ok": True, "store": store, "version": snapshot["version"], "sha256": snapshot["sha256"]})

        server = ThreadingHTTPServer(("127.0.0.1", 0), RuntimeSnapshotHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        saved_env = {
            key: os.environ.get(key)
            for key in ("RUNTIME_STORE_REMOTE_API_URL", "RUNTIME_STORE_REMOTE_API_SECRET", "RUNTIME_STORE_REMOTE_FORCE")
        }
        original_state_dir = app["RUNTIME_STORE_REMOTE_STATE_DIR"]
        original_cost_path = app["COST_PRICE_DB_PATH"]
        original_report_path = app["NO_MATCH_REPORT_DB_PATH"]
        try:
            app["RUNTIME_STORE_REMOTE_STATE_DIR"] = os.path.join(self.temp_dir, "runtime-state")
            app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "runtime-cost.sqlite")
            app["NO_MATCH_REPORT_DB_PATH"] = os.path.join(self.temp_dir, "runtime-reports.sqlite")
            app["reset_runtime_store_remote_refresh_cache"]()

            ok, message, _ = app["import_cost_price_list_from_upload"](
                UploadedBytes("runtime-cost.xlsx", fojan_quote_xlsx_bytes()),
                "regression",
            )
            self.assertTrue(ok, message)
            ok, message, manual_id = app["save_manual_cost_price_item"](
                brand="FOJAN(富捷)",
                model="FRC0603F1002TS",
                cost="1.23",
                moq="6000PCS",
                lead_time="2W",
                note="runtime snapshot regression",
                updated_by="regression",
            )
            self.assertTrue(ok, message)
            self.assertEqual(snapshots["cost-price"]["version"], 0)
            os.environ["RUNTIME_STORE_REMOTE_API_URL"] = f"http://127.0.0.1:{server.server_port}/api/runtime-store/snapshot"
            os.environ["RUNTIME_STORE_REMOTE_API_SECRET"] = api_secret
            os.environ["RUNTIME_STORE_REMOTE_FORCE"] = "1"
            app["reset_runtime_store_remote_refresh_cache"]("cost-price")
            self.assertIsNotNone(app["get_active_cost_price_list"]())
            self.assertGreater(snapshots["cost-price"]["version"], 0)
            app["COST_PRICE_DB_PATH"] = os.path.join(self.temp_dir, "runtime-cost-restored.sqlite")
            app["reset_runtime_store_remote_refresh_cache"]("cost-price")
            restored_cost = app["get_active_cost_price_list"]()
            self.assertIsNotNone(restored_cost)
            self.assertEqual(restored_cost["row_count"], 5)
            restored_manual = app["list_manual_cost_price_items"](active_only=True, limit=10)
            self.assertEqual([row["id"] for row in restored_manual], [manual_id])
            self.assertEqual(restored_manual[0]["cost"], "1.23")

            ok, message, report_id = app["submit_no_match_report"]("REMOTE-UNMATCHED-PART", reason="regression")
            self.assertTrue(ok, message)
            self.assertGreater(snapshots["no-match"]["version"], 0)
            app["NO_MATCH_REPORT_DB_PATH"] = os.path.join(self.temp_dir, "runtime-reports-restored.sqlite")
            app["reset_runtime_store_remote_refresh_cache"]("no-match")
            restored_reports = app["list_no_match_reports"]("all")
            self.assertEqual([row["id"] for row in restored_reports], [report_id])
            self.assertEqual(restored_reports[0]["query_text"], "REMOTE-UNMATCHED-PART")
        finally:
            app["RUNTIME_STORE_REMOTE_STATE_DIR"] = original_state_dir
            app["COST_PRICE_DB_PATH"] = original_cost_path
            app["NO_MATCH_REPORT_DB_PATH"] = original_report_path
            app["reset_runtime_store_remote_refresh_cache"]()
            for key, value in saved_env.items():
                if value is None:
                    os.environ.pop(key, None)
                else:
                    os.environ[key] = value
            server.shutdown()
            server.server_close()

    def test_13_manufacturer_packaging_moq_is_source_backed(self):
        lookup = self.app["lookup_manufacturer_packaging"]
        cases = [
            ({"品牌": "国巨YAGEO", "型号": "RC0603FR-0710KL", "系列": "RC", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "RC2010FK-0710KL", "系列": "RC", "尺寸（inch）": "2010"}, "4000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "RT0402BRD0733RL", "系列": "RT", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "RT0805BRA0710KL", "系列": "RT", "尺寸（inch）": "0805"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "RT2512BKB07100KL", "系列": "RT", "尺寸（inch）": "2512"}, "4000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "AA0603FR-071KL", "系列": "AA", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "AT0603BRC0710KL", "系列": "AT", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "RE1206BRE07100KL", "系列": "RE", "尺寸（inch）": "1206"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "PT2512DK-070R4L", "系列": "PT", "尺寸（inch）": "2512"}, "4000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "AR0805FR-07100KL", "系列": "AR", "尺寸（inch）": "0805"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "RL1206FR-070R011L", "系列": "RL", "尺寸（inch）": "1206"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "RP0603BRD07100KL", "系列": "RP", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "AC0603BRE0722KL", "系列": "AC", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "AC1020FK-07100KL", "系列": "AC", "尺寸（inch）": "1020"}, "4000PCS"),
            ({"品牌": "KOA", "型号": "RN73H1ETTP1000B25", "系列": "RN73H", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "KOA", "型号": "RS73F2BRTTD1000B", "系列": "RS73", "尺寸（inch）": "1206"}, "5000PCS"),
            ({"品牌": "KOA", "型号": "WK73R2HTTE1000F", "系列": "WK73R", "尺寸（inch）": "1020"}, "4000PCS"),
            ({"品牌": "威世Vishay", "型号": "CRCW0402100KJNED", "系列": "CRCW", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "威世Vishay", "型号": "CRCW06030000Z0EAHP", "系列": "CRCW", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "威世Vishay", "型号": "CRCW25120000Z0EG", "系列": "CRCW", "尺寸（inch）": "2512"}, "2000PCS"),
            ({"品牌": "威世Vishay", "型号": "CRCW0201100KFNEI", "系列": "CRCW", "尺寸（inch）": "0201"}, "20000PCS"),
            ({"品牌": "威世Vishay", "型号": "CRCW08050000ZSTA", "系列": "CRCW", "尺寸（inch）": "0805"}, "5000PCS"),
            ({"品牌": "威世Vishay", "型号": "CRCW25120000ZSTH", "系列": "CRCW", "尺寸（inch）": "2512"}, "4000PCS"),
            ({"品牌": "威世Vishay", "型号": "TNPW0402100KBEED", "系列": "TNPW", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "威世Vishay", "型号": "TNPW0603100KBETA", "系列": "TNPW", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "威世Vishay", "型号": "TNPW1206100KBECN", "系列": "TNPW", "尺寸（inch）": "1206"}, "1000PCS"),
            ({"品牌": "Panasonic", "型号": "ERJ6GEYJ103V", "系列": "ERJ", "尺寸（inch）": "0805"}, "5000PCS"),
            ({"品牌": "Panasonic", "型号": "ERA2AEB102X", "系列": "ERA-2A", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "威世Vishay", "型号": "NTCS0402E3103JL1T", "系列": "NTCS0402E", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "威世Vishay", "型号": "NTCS0603E3103FMT", "系列": "NTCS0603E", "尺寸（inch）": "0603"}, "4000PCS"),
            ({"品牌": "威世Vishay", "型号": "NTCS0805E3103FLT", "系列": "NTCS0805E", "尺寸（inch）": "0805"}, "4000PCS"),
            ({"品牌": "东电化TDK", "型号": "C1608C0G2E182J080AA", "系列": "C", "尺寸（inch）": "0603"}, "4000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "CC0402KRX7R9BB103", "系列": "CC", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "CC0201KRX5R7BB104", "系列": "CC", "尺寸（inch）": "0201"}, "15000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "CC0402KPX7R9BB103", "系列": "CC", "尺寸（inch）": "0402"}, "50000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "CC0603JRX7R9BB104", "系列": "CC", "尺寸（inch）": "0603"}, "4000PCS"),
            ({"品牌": "国巨YAGEO", "型号": "CC0603KPX7R7BB104", "系列": "CC", "尺寸（inch）": "0603"}, "15000PCS"),
            ({"品牌": "东电化TDK", "型号": "C0603X7S0J224K030BC", "系列": "C", "尺寸（inch）": "0201"}, "15000PCS"),
            ({"品牌": "东电化TDK", "型号": "C1005X7R1V224K050BC", "系列": "C", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "东电化TDK", "型号": "C2012C0G2W221J060AE", "系列": "C", "尺寸（inch）": "0805"}, "4000PCS"),
            ({"品牌": "东电化TDK", "型号": "C2012X5R1V226M125AC", "系列": "C", "尺寸（inch）": "0805"}, "2000PCS"),
            ({"品牌": "东电化TDK", "型号": "NTCG063JF103FTDS", "系列": "NTCG", "尺寸（inch）": "0201"}, "15000PCS"),
            ({"品牌": "东电化TDK", "型号": "NTCG103JF103FTDS", "系列": "NTCG", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "东电化TDK", "型号": "NTCG163JF103FTDS", "系列": "NTCG", "尺寸（inch）": "0603"}, "4000PCS"),
            ({"品牌": "村田Murata", "型号": "GRM155R71E472KA01D", "系列": "GRM", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "村田Murata", "型号": "GRM155R71E472KA01J", "系列": "GRM", "尺寸（inch）": "0402"}, "50000PCS"),
            ({"品牌": "村田Murata", "型号": "GRM188R11H104KA93D", "系列": "GRM", "尺寸（inch）": "0603"}, "4000PCS"),
            ({"品牌": "村田Murata", "型号": "GCM188R71H273KA55V", "系列": "GCM", "尺寸（inch）": "0603"}, "30000PCS"),
            ({"品牌": "村田Murata", "型号": "GCJ21BR71H104KA01L", "系列": "GCJ", "尺寸（inch）": "0805"}, "3000PCS"),
            ({"品牌": "村田Murata", "型号": "GCJ21BR71H104KA01K", "系列": "GCJ", "尺寸（inch）": "0805"}, "10000PCS"),
            ({"品牌": "三星Samsung", "型号": "CL02A102KP2NNNC", "系列": "CL", "尺寸（inch）": "01005", "高度（mm）": "0.20±0.02"}, "20000PCS"),
            ({"品牌": "三星Samsung", "型号": "CL03A102KA31INC", "系列": "CL", "尺寸（inch）": "0201", "高度（mm）": "0.30±0.03"}, "10000PCS"),
            ({"品牌": "三星Samsung", "型号": "CL05A104JO5NNNC", "系列": "CL", "尺寸（inch）": "0402", "高度（mm）": "0.50±0.05"}, "10000PCS"),
            ({"品牌": "三星Samsung", "型号": "CL10A104KA8NNNC", "系列": "CL", "尺寸（inch）": "0603", "高度（mm）": "0.80±0.10"}, "4000PCS"),
            ({"品牌": "三星Samsung", "型号": "CL21A105KACLNNC", "系列": "CL", "尺寸（inch）": "0805", "高度（mm）": "0.85±0.10"}, "4000PCS"),
            ({"品牌": "三星Samsung", "型号": "CL21A106KOQNNWC", "系列": "CL", "尺寸（inch）": "0805", "高度（mm）": "1.25±0.15"}, "2000PCS"),
            ({"品牌": "三星Samsung", "型号": "RC1005F100CS", "系列": "RC", "尺寸（inch）": "01005"}, "10000PCS"),
            ({"品牌": "三星Samsung", "型号": "RCS1608F100CS", "系列": "RCS", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "华新科Walsin", "型号": "WR04W1005FTL", "系列": "WR", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "华新科Walsin", "型号": "WR02W1005FAL", "系列": "WR", "尺寸（inch）": "01005"}, "15000PCS"),
            ({"品牌": "华新科Walsin", "型号": "WR18X40R2FTL", "系列": "WR", "尺寸（inch）": "1218"}, "3000PCS"),
            ({"品牌": "华新科Walsin", "型号": "0402B101J100CT", "系列": "常规", "尺寸（inch）": "0402", "高度（mm）": "0.50"}, "10000PCS"),
            ({"品牌": "华新科Walsin", "型号": "0805A106K250CT", "系列": "常规", "尺寸（inch）": "0805", "高度（mm）": "1.25"}, "3000PCS"),
            ({"品牌": "华新科Walsin", "型号": "1210B102J101CT", "系列": "常规", "尺寸（inch）": "1210", "高度（mm）": "2.50"}, "1000PCS"),
            ({"品牌": "华新科Walsin", "型号": "0402N0R1A500CT", "系列": "0402N", "尺寸（inch）": "0402", "高度（mm）": "0.50"}, "10000PCS"),
            ({"品牌": "三星Samsung", "型号": "RU1005FR020CS", "系列": "RU", "尺寸（inch）": "0402"}, "10000PCS"),
            ({"品牌": "三星Samsung", "型号": "RUK1608FR010CS", "系列": "RU", "尺寸（inch）": "0603"}, "5000PCS"),
            ({"品牌": "三星Samsung", "型号": "RUT2012FR100CS", "系列": "RU", "尺寸（inch）": "0805"}, "5000PCS"),
            ({"品牌": "三星Samsung", "型号": "RJ1220FR005CS", "系列": "RJ", "尺寸（inch）": "0508"}, "5000PCS"),
            ({"品牌": "三星Samsung", "型号": "RJ1220FR002CS", "系列": "RJ", "尺寸（inch）": "0508"}, "4000PCS"),
            ({"品牌": "华新科Walsin", "型号": "SH31B101K102CT", "系列": "SH", "尺寸（inch）": "1206", "高度（mm）": "1.60"}, "2000PCS"),
            ({"品牌": "华新科Walsin", "型号": "RF03N0R1A250CT", "系列": "RF", "尺寸（inch）": "0201", "高度（mm）": "0.30"}, "15000PCS"),
            ({"品牌": "华新科Walsin", "型号": "HH21N0R5B101CT", "系列": "HH", "尺寸（inch）": "0805", "高度（mm）": "1.25"}, "3000PCS"),
            ({"品牌": "华新科Walsin", "型号": "MT15N0R5B500CT", "系列": "MT", "尺寸（inch）": "0402", "高度（mm）": "0.50"}, "10000PCS"),
            ({"品牌": "FOJAN(富捷)", "型号": "FRM121WFR010TM", "系列": "FRM", "尺寸（inch）": "1206"}, "5000PCS"),
            ({"品牌": "FOJAN(富捷)", "型号": "FPM253WFR060TM", "系列": "FPM", "尺寸（inch）": "2512"}, "4000PCS"),
        ]
        for row, expected in cases:
            result = lookup(row)
            self.assertEqual(result.get("MOQ"), expected, row["型号"])
            self.assertIn("原厂标准包装数量", result.get("MOQ来源", ""), row["型号"])
            self.assertTrue(result.get("包装数量来源", "").startswith("https://"), row["型号"])

        self.assertEqual(
            lookup({"品牌": "国巨YAGEO", "型号": "RC0603FK-0710KL", "系列": "RC", "尺寸（inch）": "0603"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "国巨YAGEO", "型号": "RT2010BRD07100RL", "系列": "RT", "尺寸（inch）": "2010"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "KOA", "型号": "SLR1TTE1000D", "系列": "SLR", "尺寸（inch）": "2512"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "威世Vishay", "型号": "CRCW08050000Z0EB", "系列": "CRCW", "尺寸（inch）": "0805"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "威世Vishay", "型号": "TNPW1206100KBEEN", "系列": "TNPW", "尺寸（inch）": "1206"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "村田Murata", "型号": "LQW18AN20NG00#", "系列": "LQW18AN", "尺寸（inch）": "0603"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "国巨YAGEO", "型号": "CC0805MRX7R9BB104", "系列": "CC", "尺寸（inch）": "0805"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "三星Samsung", "型号": "CL10A105KA8NNND", "系列": "CL", "尺寸（inch）": "0603", "高度（mm）": "0.80±0.10"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "三星Samsung", "型号": "RC1608F100AS", "系列": "RC", "尺寸（inch）": "0603"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "华新科Walsin", "型号": "WR04W1005FBL", "系列": "WR", "尺寸（inch）": "0402"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "华新科Walsin", "型号": "1812B102J101CT", "系列": "常规", "尺寸（inch）": "1812", "高度（mm）": "3.20"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "华新科Walsin", "型号": "SH43B103K102CT", "系列": "SH", "尺寸（inch）": "1812", "高度（mm）": "3.20"}),
            {},
        )
        self.assertEqual(
            lookup({"品牌": "华新科Walsin", "型号": "MT15B101K500CT", "系列": "MT", "尺寸（inch）": "0402", "高度（mm）": "0.85"}),
            {},
        )
        original_cost_path = self.app["COST_PRICE_DB_PATH"]
        try:
            isolated_cost_path = os.path.join(self.temp_dir, "manufacturer-packaging-cost.sqlite")
            self.app["COST_PRICE_DB_PATH"] = isolated_cost_path
            self.app["clear_cost_price_lookup_cache"]()
            enriched = self.app["enrich_component_cost_columns"](
                pd.DataFrame([cases[0][0]])
            )
            self.assertEqual(enriched.iloc[0]["MOQ"], "5000PCS")
            self.assertIn("YAGEO", enriched.iloc[0]["MOQ来源"])

            cost_upload = UploadedBytes(
                "purchase-moq.xlsx",
                dataframe_to_xlsx_bytes(
                    pd.DataFrame(
                        [{"品牌": "国巨YAGEO", "型号": "RC0603FR-0710KL", "MOQ": "123PCS"}]
                    )
                ),
            )
            ok, message, _ = self.app["import_cost_price_list_from_upload"](cost_upload, "regression")
            self.assertTrue(ok, message)
            overridden = self.app["enrich_component_cost_columns"](pd.DataFrame([cases[0][0]]))
            self.assertEqual(overridden.iloc[0]["MOQ"], "123PCS")
            self.assertEqual(overridden.iloc[0]["MOQ来源"], "当前启用成本清单")
        finally:
            self.app["COST_PRICE_DB_PATH"] = original_cost_path
            self.app["clear_cost_price_lookup_cache"]()

    def test_14_fojan_alloy_resistor_rules_are_source_scoped(self):
        app = self.app

        parsed_frm = app["parse_resistor_model_rule"](
            "FRM121WFR010TM",
            brand="FOJAN(富捷)",
            component_type="合金电阻",
        )
        self.assertEqual(parsed_frm["器件类型"], "合金电阻")
        self.assertEqual(parsed_frm["系列"], "FRM")
        self.assertEqual(parsed_frm["尺寸（inch）"], "1206")
        self.assertEqual(parsed_frm["容值"], "10")
        self.assertEqual(parsed_frm["容值单位"], "mΩ")
        self.assertEqual(parsed_frm["容值误差"], "1")
        self.assertEqual(parsed_frm["功率"], "1W")

        parsed_fpm = app["parse_resistor_model_rule"](
            "FPM253WFR060TM",
            brand="FOJAN(富捷)",
            component_type="合金电阻",
        )
        self.assertEqual(parsed_fpm["器件类型"], "合金电阻")
        self.assertEqual(parsed_fpm["系列"], "FPM")
        self.assertEqual(parsed_fpm["尺寸（inch）"], "2512")
        self.assertEqual(parsed_fpm["容值"], "60")
        self.assertEqual(parsed_fpm["容值单位"], "mΩ")
        self.assertEqual(parsed_fpm["容值误差"], "1")
        self.assertEqual(parsed_fpm["功率"], "3W")

        parsed_fcm = app["parse_resistor_model_rule"](
            "FCM25125WF0M50TM",
            brand="FOJAN(富捷)",
            component_type="合金电阻",
        )
        self.assertEqual(parsed_fcm["系列"], "FCM")
        self.assertEqual(parsed_fcm["尺寸（inch）"], "2512")
        self.assertEqual(parsed_fcm["容值"], "0.5")
        self.assertEqual(parsed_fcm["容值单位"], "mΩ")
        self.assertEqual(parsed_fcm["功率"], "5W")

        parsed_fwp = app["parse_resistor_model_rule"](
            "FWP27284WFR010TK",
            brand="FOJAN(富捷)",
            component_type="合金电阻",
        )
        self.assertEqual(parsed_fwp["系列"], "FWP")
        self.assertEqual(parsed_fwp["尺寸（inch）"], "2728")
        self.assertEqual(parsed_fwp["容值"], "10")
        self.assertEqual(parsed_fwp["容值单位"], "mΩ")

        parsed_fwk = app["parse_resistor_model_rule"](
            "FWK12169WF0M50RK",
            brand="FOJAN(富捷)",
            component_type="合金电阻",
        )
        self.assertEqual(parsed_fwk["系列"], "FWK")
        self.assertEqual(parsed_fwk["尺寸（inch）"], "1216")
        self.assertEqual(parsed_fwk["容值"], "0.5")
        self.assertEqual(parsed_fwk["容值单位"], "mΩ")

        mode, spec = app["detect_query_mode_and_spec"](
            pd.DataFrame(),
            "合金电阻 电阻10毫欧 ±1% 1206",
        )
        self.assertEqual(mode, "合金电阻")
        rows = app["load_search_dataframe_for_query"](mode, spec)
        fojan_models = set(
            rows[rows["品牌"].astype(str).str.contains("FOJAN|富捷", case=False, regex=True)]["型号"].map(app["clean_model"])
        )
        self.assertIn("FRM121WFR010TM", fojan_models)
        self.assertNotIn("FRL1206FR010TS", fojan_models)
        display = app["select_component_display_columns"](
            rows[rows["型号"].map(app["clean_model"]).eq("FRM121WFR010TM")],
            spec,
            prefix_columns=["品牌", "型号", "系列"],
        )
        self.assertEqual(display.iloc[0]["MOQ"], "5000PCS")
        self.assertIn("FOJAN", display.iloc[0]["MOQ来源"])

        mode, spec = app["detect_query_mode_and_spec"](
            pd.DataFrame(),
            "富捷 贴片合金电阻 0.06R 2512 3W ±1%",
        )
        self.assertEqual(mode, "合金电阻")
        self.assertEqual(spec["品牌"], "FOJAN(富捷)")
        rows = app["load_search_dataframe_for_query"](mode, spec)
        fojan_rows = rows[
            rows["品牌"].astype(str).str.contains("FOJAN|富捷", case=False, regex=True)
        ]
        self.assertEqual(set(fojan_rows["型号"].map(app["clean_model"])), {"FPM253WFR060TM"})
        display = app["select_component_display_columns"](
            fojan_rows,
            spec,
            prefix_columns=["品牌", "型号", "系列"],
        )
        self.assertEqual(display.iloc[0]["MOQ"], "4000PCS")
        self.assertIn("FOJAN FPM", display.iloc[0]["MOQ来源"])

        for query in (
            "贴片合金电阻 0.3Ω ±1% 1206 1W",
        ):
            mode, spec = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            rows = app["load_search_dataframe_for_query"](mode, spec)
            if rows is None or rows.empty:
                continue
            fojan_rows = rows[rows["品牌"].astype(str).str.contains("FOJAN|富捷", case=False, regex=True)]
            self.assertTrue(fojan_rows.empty, query)

        query = "贴片合金电阻 2512 0.2R ±1%"
        mode, spec = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
        rows = app["load_search_dataframe_for_query"](mode, spec)
        fojan_models = set(
            rows[
                rows["品牌"].astype(str).str.contains("FOJAN|富捷", case=False, regex=True)
            ]["型号"].map(app["clean_model"])
        )
        self.assertIn("FRM252WFR200TM", fojan_models)

        mode, spec = app["detect_query_mode_and_spec"](
            pd.DataFrame(),
            "富捷 FCM 2512 0.5mR 5W ±1%",
        )
        rows = app["load_search_dataframe_for_query"](mode, spec)
        fojan_models = set(
            rows[
                rows["品牌"].astype(str).str.contains("FOJAN|富捷", case=False, regex=True)
            ]["型号"].map(app["clean_model"])
        )
        self.assertIn("FCM25125WF0M50TM", fojan_models)

        for query, expected_model in [
            ("富捷 FWP 2728 10mR 4W ±1%", "FWP27284WFR010TK"),
            ("富捷 FWK 1216 0.5mR 9W ±1%", "FWK12169WF0M50TK"),
            ("富捷 FMH 1206 120mR 1W ±1%", "FMH121WFR120TM"),
        ]:
            mode, spec = app["detect_query_mode_and_spec"](pd.DataFrame(), query)
            self.assertEqual(mode, "合金电阻", query)
            rows = app["load_search_dataframe_for_query"](mode, spec)
            fojan_models = set(
                rows[
                    rows["品牌"].astype(str).str.contains("FOJAN|富捷", case=False, regex=True)
                ]["型号"].map(app["clean_model"])
            )
            self.assertIn(expected_model, fojan_models, query)

    def test_15_joyin_ntc_b_tolerance_is_decoded_and_ranked(self):
        app = self.app
        app["DB_PATH"] = self.original_paths["DB_PATH"]
        app["SEARCH_DB_PATH"] = self.original_paths["SEARCH_DB_PATH"]
        expected_b_tolerances = {
            "JSNZ104F425FABXG": "1",
            "JSNZ104F425GABXG": "2",
            "JSNZ104F425HABXG": "3",
            "JSNZ104F425JABXG": "5",
        }
        candidate_rows = []
        for model, expected_b_tolerance in expected_b_tolerances.items():
            parsed = app["parse_joyin_ntc_common"](model, brand="JOYIN(久尹)")
            self.assertEqual(parsed["阻值误差"], "1", model)
            self.assertEqual(parsed["B值误差"], expected_b_tolerance, model)
            parsed["_res_ohm"] = parsed["_resistance_ohm"]
            candidate_rows.append(parsed)

        spec = {
            "品牌": "村田Murata",
            "型号": "NCP03WF104F05RL",
            "器件类型": "热敏电阻",
            "尺寸（inch）": "0201",
            "容值误差": "1",
            "阻值误差": "1",
            "_resistance_ohm": 100000.0,
            "B值": "4250",
            "B值条件": "25/50℃",
        }
        self.assertEqual(app["thermistor_b_tolerance_from_record"](spec), "1")

        source_rows = app["load_component_rows_by_clean_model"]("NCP03WF104F05RL")
        self.assertFalse(source_rows.empty)
        actual_spec = app["reverse_spec"](source_rows, "NCP03WF104F05RL")
        self.assertEqual(actual_spec["B值误差"], "1")

        ranked = app["apply_match_levels_and_sort"](pd.DataFrame(candidate_rows), spec)
        levels = dict(zip(ranked["型号"], ranked["推荐等级"]))
        self.assertEqual(levels["JSNZ104F425FABXG"], "完全匹配")
        for model in ("JSNZ104F425GABXG", "JSNZ104F425HABXG", "JSNZ104F425JABXG"):
            self.assertEqual(levels[model], "需确认替代")
        self.assertEqual(ranked.iloc[0]["型号"], "JSNZ104F425FABXG")

        display = app["select_component_display_columns"](
            pd.DataFrame(candidate_rows),
            spec,
            prefix_columns=["品牌", "型号"],
        )
        self.assertIn("B值误差", display.columns)
        formatted_display = app["format_display_df"](display)
        displayed = dict(zip(formatted_display["型号"], formatted_display["B值误差"]))
        self.assertEqual(displayed["JSNZ104F425FABXG"], "±1%")
        self.assertEqual(displayed["JSNZ104F425GABXG"], "±2%")

        actual_query_rows = app["load_search_dataframe_for_query"](
            "料号",
            actual_spec,
            "NCP03WF104F05RL",
            exact_part_rows=source_rows,
        )
        actual_matches = app["run_query_match"](actual_query_rows, "料号", actual_spec)
        if isinstance(actual_matches, pd.DataFrame) and not actual_matches.empty and "型号" in actual_matches.columns:
            actual_joyin = actual_matches[
                actual_matches["型号"].map(app["clean_model"]).isin(expected_b_tolerances)
            ]
            actual_levels = dict(zip(actual_joyin["型号"].map(app["clean_model"]), actual_joyin["推荐等级"]))
            self.assertEqual(actual_levels["JSNZ104F425FABXG"], "完全匹配")
            for model in ("JSNZ104F425GABXG", "JSNZ104F425HABXG", "JSNZ104F425JABXG"):
                self.assertEqual(actual_levels[model], "需确认替代")

        reported_query = "Thermistor NTC 10K OHM 240mW 1% 0402 SMD"
        mode, reported_spec = app["detect_query_mode_and_spec"](pd.DataFrame(), reported_query)
        self.assertEqual(mode, "热敏电阻")
        self.assertEqual(reported_spec["尺寸（inch）"], "0402")
        self.assertEqual(reported_spec["阻值@25C"], "10")
        self.assertEqual(reported_spec["阻值单位"], "KΩ")
        self.assertEqual(reported_spec["阻值误差"], "1")
        self.assertEqual(reported_spec["功率"], "240mW")
        self.assertEqual(reported_spec["B值"], "")
        self.assertEqual(reported_spec["B值误差"], "")
        self.assertEqual(reported_spec["B值条件"], "")

        joyin_0402_source = app["load_component_rows_by_clean_model"]("JSNA103F337FABXG")
        self.assertFalse(joyin_0402_source.empty)
        self.assertEqual(
            app["thermistor_max_power_text_from_record"](joyin_0402_source.iloc[0]),
            "170mW",
        )

        reported_rows = app["load_search_dataframe_for_query"](
            mode,
            reported_spec,
            reported_query,
        )
        reported_matches = app["run_query_match"](reported_rows, mode, reported_spec)
        reported_joyin = reported_matches[
            reported_matches["品牌"].astype(str).str.contains("久尹|JOYIN", case=False, regex=True)
        ]
        self.assertFalse(reported_joyin.empty)
        self.assertNotIn("完全匹配", set(reported_joyin["推荐等级"]))
        self.assertEqual(set(reported_joyin["_power"]), {"170mW"})
        reported_status, reported_reason = app["classify_recommendation_status"](
            reported_joyin.iloc[0],
            reported_spec,
        )
        self.assertEqual(reported_status, "参数冲突")
        self.assertIn("最大功率不足", reported_reason)

        reported_display = app["select_component_display_columns"](
            reported_joyin.head(1),
            reported_spec,
            prefix_columns=["品牌", "型号", "推荐等级"],
        )
        self.assertIn("功率", reported_display.columns)
        self.assertEqual(reported_display.iloc[0]["功率"], "170mW")

        complete_query = "Thermistor NTC 10K OHM 170mW 1% 0402 SMD B25/50=3370K ±1%"
        complete_mode, complete_spec = app["detect_query_mode_and_spec"](pd.DataFrame(), complete_query)
        self.assertEqual(complete_mode, "热敏电阻")
        self.assertEqual(complete_spec["B值"], "3370K")
        self.assertEqual(complete_spec["B值误差"], "1")
        self.assertEqual(complete_spec["B值条件"], "25/50℃")
        complete_rows = app["load_search_dataframe_for_query"](
            complete_mode,
            complete_spec,
            complete_query,
        )
        complete_matches = app["run_query_match"](complete_rows, complete_mode, complete_spec)
        complete_joyin = complete_matches[
            complete_matches["型号"].map(app["clean_model"]).isin(
                {
                    "JSNA103F337FABXG",
                    "JSNA103F337GABXG",
                    "JSNA103F337HABXG",
                    "JSNA103F337JABXG",
                }
            )
        ]
        complete_levels = dict(
            zip(complete_joyin["型号"].map(app["clean_model"]), complete_joyin["推荐等级"])
        )
        self.assertEqual(complete_levels["JSNA103F337FABXG"], "完全匹配")
        for model in ("JSNA103F337GABXG", "JSNA103F337HABXG", "JSNA103F337JABXG"):
            self.assertEqual(complete_levels[model], "需确认替代")


if __name__ == "__main__":
    unittest.main(verbosity=2)
