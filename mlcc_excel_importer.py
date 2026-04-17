import argparse
import json
import os
import re
import shutil
import sqlite3
import sys
import warnings
from collections import OrderedDict
from datetime import datetime
from pathlib import Path


BASE_DIR = Path(__file__).resolve().parent
VENV_SITE_PACKAGES = BASE_DIR / ".venv" / "Lib" / "site-packages"
if VENV_SITE_PACKAGES.exists():
    sys.path.insert(0, str(VENV_SITE_PACKAGES))

import pandas as pd
from openpyxl import load_workbook


warnings.filterwarnings(
    "ignore",
    message=r".*DataFrame concatenation with empty or all-NA entries.*",
    category=FutureWarning,
)


DEFAULT_MASTER = BASE_DIR / "Capacitor" / "MLCC.xlsx"
DEFAULT_INPUT_DIR = BASE_DIR.parent / "mlcc_incoming"
DEFAULT_TEMPLATE_DIR = BASE_DIR / "templates"
DEFAULT_BACKUP_DIR = BASE_DIR.parent / "mlcc_backups"
DEFAULT_DB_PATH = BASE_DIR / "components.db"
DEFAULT_SAMSUNG_CACHE = BASE_DIR / "cache" / "samsung_all_statuses_base.json"
SAMSUNG_DETAIL_URL = "https://product.samsungsem.com/mlcc/{part}.do"

STATUS_DISPLAY_COLUMN = "生产状态"
STATUS_HEADER_ALIASES = [STATUS_DISPLAY_COLUMN, "量产状态", "status", "备注1"]
LINK_HEADER_ALIASES = ["备注2", "remark2", "note2", "网址", "链接", "官网链接", "产品链接", "详情链接", "url", "link"]
STANDARD_COLUMNS = [
    "品牌", "型号", "系列", "尺寸（inch）", "材质（介质）", "容值", "容值单位",
    "容值误差", "耐压（V）", "特殊用途", "备注1", "备注2", "备注3",
]
SHEET_BRAND_MAP = OrderedDict([
    ("信昌PDC", ["信昌", "PDC"]),
    ("华新科Walsin", ["华新", "WALSIN"]),
    ("东电化TDK", ["TDK", "东电化"]),
    ("三星Samsung", ["SAMSUNG", "三星"]),
    ("太诱Taiyo", ["TAIYO", "太诱", "太阳诱电"]),
    ("村田Murata", ["MURATA", "村田"]),
    ("国巨YAGEO", ["YAGEO", "国巨"]),
    ("基美Kemet", ["KEMET", "基美"]),
    ("晶瓷Kyocera AVX", ["KYOCERA", "AVX", "晶瓷"]),
    ("三环CCTC", ["CCTC", "三环"]),
    ("三和SAMWHA", ["SAMWHA", "三和"]),
    ("宇阳EYANG", ["EYANG", "宇阳"]),
    ("达方DARFON", ["DARFON", "达方"]),
    ("微容VIIYONG", ["VIIYONG", "微容"]),
    ("风华Fenghua", ["FENGHUA", "风华"]),
    ("富捷FOJAN", ["FOJAN", "富捷"]),
    ("芯声微HRE", ["HRE", "芯声微"]),
])
HEADER_ALIASES = {
    "品牌": ["品牌", "厂牌", "品牌名称", "brand", "manufacturer", "mfr"],
    "型号": ["型号", "料号", "物料号", "产品型号", "规格型号", "订货号", "partnumber", "partno", "part#", "pn", "mpn", "model"],
    "系列": ["系列", "series"],
    "尺寸（inch）": ["尺寸", "封装", "尺寸inch", "case", "size", "inch", "casecode", "packagesize"],
    "材质（介质）": ["材质", "介质", "温特性", "dielectric", "material"],
    "容值": ["容值", "电容值", "capacitance", "value"],
    "容值单位": ["容值单位", "单位", "unit"],
    "容值误差": ["容值误差", "容值误差（%&pF）", "容值误差（%）", "容差", "误差", "误差%", "误差（%）", "误差（%&pF）", "容差%", "容差（%）", "tolerance", "tolerance(%)", "tol"],
    "耐压（V）": ["耐压", "电压", "额定电压", "voltage", "ratedvoltage", "wv"],
    "特殊用途": ["特殊用途", "用途", "application", "special"],
    "备注1": ["备注1", "备注", "remark1", "note1", "生产状态", "量产状态", "status"],
    "备注2": ["备注2", "remark2", "note2", "网址", "链接", "官网链接", "产品链接", "详情链接", "url", "link", "officialurl", "producturl"],
    "备注3": ["备注3", "remark3", "note3"],
}
OFFICIAL_TEMPLATE_COLUMNS = [
    "品牌", "型号", "系列", "尺寸（inch）", "材质（介质）", "容值", "容值单位",
    "容值误差", "耐压（V）", "特殊用途", "备注1", "备注2", "备注3",
]
OFFICIAL_TEMPLATE_SAMPLE = [{
    "品牌": "Samsung",
    "型号": "CL05A105KQ5NNNC",
    "系列": "MLCC",
    "尺寸（inch）": "0402",
    "材质（介质）": "X5R",
    "容值": "1",
    "容值单位": "uF",
    "容值误差": "±10%",
    "耐压（V）": "6.3V",
    "特殊用途": "",
    "备注1": "Mass Production",
    "备注2": "https://product.samsungsem.com/mlcc/CL05A105KQ5NNNC.do",
    "备注3": "",
}]


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_brand(value):
    return clean_text(value)


def clean_model(value):
    return clean_text(value).upper().replace(" ", "")


def clean_material(value):
    text = clean_text(value).upper()
    text = text.replace("（", "(").replace("）", ")").replace(" ", "")
    text = text.replace("C0G", "COG").replace("NP0", "NPO")
    return text


def clean_size(value):
    text = clean_text(value).replace(".0", "").replace(" ", "")
    pad_map = {
        "402": "0402",
        "603": "0603",
        "805": "0805",
        "201": "0201",
        "1005": "01005",
        "8004": "008004",
        "102": "0102",
        "15008": "015008",
        "008004": "008004",
        "0102": "0102",
        "015008": "015008",
    }
    if text in pad_map:
        return pad_map[text]
    return text.upper()


def normalize_tolerance_number(value):
    try:
        num = abs(float(clean_text(value)))
    except Exception:
        return None
    if 0 < num <= 1:
        num = num * 100
    if float(num).is_integer():
        num = int(num)
    return str(num)


def clean_tol_for_match(value):
    text = clean_text(value).upper()
    if text == "":
        return ""
    text = text.replace("（", "(").replace("）", ")").replace(" ", "")
    text = text.replace("％", "%").replace("﹪", "%")
    text = text.replace("＋", "+").replace("﹢", "+")
    text = text.replace("／", "/").replace("\\", "/")
    text = text.replace("卤", "+/-").replace("±", "+/-")
    for dash in ["－", "–", "—", "―", "−"]:
        text = text.replace(dash, "-")
    tol_letter_map = {
        "B": "0.1pF",
        "C": "0.25pF",
        "D": "0.5pF",
        "F": "1",
        "G": "2",
        "J": "5",
        "K": "10",
        "M": "20",
        "W": "0.05pF",
        "Z": "+80/-20",
    }
    if text in tol_letter_map:
        return tol_letter_map[text]
    text = text.replace("%", "").replace("+-", "+/-")
    if text.startswith("+/-"):
        normalized = normalize_tolerance_number(text[3:])
        return normalized if normalized is not None else text
    patterns = [
        (r"^\+?(\d+(?:\.\d+)?)/-(\d+(?:\.\d+)?)$", True),
        (r"^-(\d+(?:\.\d+)?)/\+?(\d+(?:\.\d+)?)$", False),
    ]
    for pattern, plus_first in patterns:
        match = re.fullmatch(pattern, text)
        if match:
            first = normalize_tolerance_number(match.group(1))
            second = normalize_tolerance_number(match.group(2))
            if first is None or second is None:
                return text
            if first == second:
                return first
            plus_val, minus_val = (first, second) if plus_first else (second, first)
            return f"+{plus_val}/-{minus_val}"
    normalized = normalize_tolerance_number(text)
    if normalized is not None:
        return normalized
    return text


def infer_tolerance_from_model(model, current_tolerance=""):
    if clean_text(current_tolerance) != "":
        return clean_tol_for_match(current_tolerance)

    text = clean_model(model).upper()
    if text == "":
        return ""

    tol_letter_map = {
        "B": "0.1pF",
        "C": "0.25pF",
        "D": "0.5pF",
        "F": "1",
        "G": "2",
        "J": "5",
        "K": "10",
        "M": "20",
        "W": "0.05pF",
        "Z": "+80/-20",
    }
    if text.startswith("CC"):
        search_text = text
        patterns = [
            r"^CC\d{4}([BCDFGJKMWZ])",
            r"\dR\d([BCDFGJKMWZ])\d",
            r"\d{3}([BCDFGJKMWZ])\d",
        ]
    elif text.startswith("C"):
        search_text = text[5:]
        patterns = [
            r"\d{3}([BCDFGJKMWZ])\d",
            r"\dR\d([BCDFGJKMWZ])\d",
        ]
    else:
        search_text = text
        patterns = [
            r"\d{3}([BCDFGJKMWZ])\d",
            r"\dR\d([BCDFGJKMWZ])\d",
        ]
    for pattern in patterns:
        match = re.search(pattern, search_text)
        if not match:
            continue
        return clean_tol_for_match(tol_letter_map.get(match.group(1), ""))
    return clean_tol_for_match(current_tolerance)


def fill_missing_tolerance_from_model(df):
    if df.empty or "型号" not in df.columns or "容值误差" not in df.columns:
        return df

    work = df.copy()
    work["容值误差"] = [
        infer_tolerance_from_model(model, tol)
        for model, tol in zip(work["型号"], work["容值误差"])
    ]
    return work


def clean_voltage(value):
    text = clean_text(value).upper().replace(" ", "").replace("V", "")
    if text == "":
        return ""
    try:
        num = float(text)
    except Exception:
        return text
    if num.is_integer():
        return str(int(num))
    return f"{num:.6f}".rstrip("0").rstrip(".")


def format_number(value):
    text = clean_text(value)
    if text == "":
        return ""
    try:
        num = float(text)
    except Exception:
        return text
    if num.is_integer():
        return str(int(num))
    return f"{num:.6f}".rstrip("0").rstrip(".")


def extract_url_value(value):
    text = clean_text(value)
    if text == "":
        return ""
    lower = text.lower()
    for prefix in ["official_url=", "链接=", "url=", "网址="]:
        if lower.startswith(prefix.lower()):
            text = clean_text(text.split("=", 1)[1] if "=" in text else "")
            break
    if not (text.startswith("https://") or text.startswith("http://")):
        return ""
    if "product.samsungsem.com/mlcc/" in text and not text.endswith(".do"):
        text = text + ".do"
    return text


def looks_like_official_status(value):
    text = clean_text(value)
    if text == "":
        return False
    upper = text.upper()
    return (
        "PRODUCTION" in upper or
        upper in {"NRND", "EOL"} or
        "量产" in text
    )


def normalize_samsung_size_code(size_code):
    token = clean_text(size_code)
    if "/" in token:
        token = token.split("/", 1)[0]
    return clean_size(token)


def extract_samsung_base_part(row):
    for column in ["备注2", "备注3"]:
        url = extract_url_value(row.get(column, ""))
        if "/mlcc/" in url:
            base = url.rsplit("/", 1)[-1]
            if base.endswith(".do"):
                base = base[:-3]
            return clean_model(base)
    return clean_model(row.get("型号", ""))


def load_samsung_cache_map(cache_path):
    path = Path(cache_path)
    if not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, dict):
        rows = raw.get("rows") or raw.get("data") or raw.get("result") or raw.get("list") or []
    elif isinstance(raw, list):
        rows = raw
    else:
        rows = []

    cache_map = {}
    for item in rows:
        base_part = clean_model(item.get("partNumber", ""))
        if base_part == "":
            continue
        cache_map[base_part] = {
            "尺寸（inch）": normalize_samsung_size_code(item.get("sizeCode", "")),
            "材质（介质）": clean_material(item.get("tcc", "")),
            "容值": format_number(item.get("capacitanceVal", "")),
            "容值单位": clean_text(item.get("capacitanceUnit", "")).upper(),
            "容值误差": clean_tol_for_match(item.get("tolerance", "")),
            "耐压（V）": clean_voltage(item.get("ratedVdc", "")),
            "备注1": clean_text(item.get("status", "")),
            "备注2": SAMSUNG_DETAIL_URL.format(part=base_part),
        }
    return cache_map


def repair_samsung_sheet(df, cache_map):
    work = ensure_standard_columns(df).copy()
    changed_rows = 0

    for idx, row in work.iterrows():
        brand = clean_brand(row.get("品牌", ""))
        if "三星" not in brand and "SAMSUNG" not in brand.upper():
            continue

        changed = False
        base_part = extract_samsung_base_part(row)
        official = cache_map.get(base_part, {})

        target_url = official.get("备注2", "") or extract_url_value(row.get("备注2", ""))
        if target_url == "" and base_part != "":
            target_url = SAMSUNG_DETAIL_URL.format(part=base_part)

        for column in ["尺寸（inch）", "材质（介质）", "容值", "容值单位", "容值误差", "耐压（V）"]:
            new_value = clean_text(official.get(column, ""))
            if new_value == "":
                continue
            current_value = clean_text(work.at[idx, column])
            normalized_current = clean_size(current_value) if column == "尺寸（inch）" else current_value
            normalized_new = clean_size(new_value) if column == "尺寸（inch）" else new_value
            if normalized_current != normalized_new:
                work.at[idx, column] = new_value
                changed = True

        status_value = clean_text(official.get("备注1", ""))
        if status_value == "" and looks_like_official_status(row.get("备注1", "")):
            status_value = clean_text(row.get("备注1", ""))
        if status_value != "" and clean_text(work.at[idx, "备注1"]) != status_value:
            work.at[idx, "备注1"] = status_value
            changed = True

        if target_url != "" and clean_text(work.at[idx, "备注2"]) != target_url:
            work.at[idx, "备注2"] = target_url
            changed = True

        if clean_text(work.at[idx, "备注3"]) != "":
            work.at[idx, "备注3"] = ""
            changed = True

        if changed:
            changed_rows += 1

    return work, changed_rows


def normalize_header_name(column_name):
    text = clean_text(column_name).lower()
    text = text.replace("（", "(").replace("）", ")")
    return re.sub(r"[\s\-_()/\\\[\]{}:：,.，;；#%％&]+", "", text)


def find_header_index(headers, aliases):
    normalized_aliases = {normalize_header_name(alias) for alias in aliases}
    for idx, header in enumerate(headers, start=1):
        if normalize_header_name(header) in normalized_aliases:
            return idx
    return None


def extract_sheet_status_metadata(workbook_path, sheet_name, row_count=None):
    metadata = {}
    try:
        wb = load_workbook(workbook_path, read_only=False, data_only=False)
    except Exception:
        return metadata

    try:
        if sheet_name not in wb.sheetnames:
            return metadata
        ws = wb[sheet_name]
        headers = [clean_text(ws.cell(1, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        status_col = find_header_index(headers, STATUS_HEADER_ALIASES)
        link_col = find_header_index(headers, LINK_HEADER_ALIASES)
        if status_col is None:
            return metadata

        max_row = ws.max_row
        if row_count is not None:
            max_row = min(max_row, row_count + 1)

        for excel_row in range(2, max_row + 1):
            status_cell = ws.cell(excel_row, status_col)
            status_text = clean_text(status_cell.value)
            target_url = ""
            if status_cell.hyperlink is not None:
                target_url = clean_text(
                    getattr(status_cell.hyperlink, "target", None)
                    or getattr(status_cell.hyperlink, "location", None)
                    or ""
                )
            if link_col is not None:
                target_url = extract_url_value(ws.cell(excel_row, link_col).value) or target_url

            if status_text == "" and target_url == "":
                continue
            metadata[excel_row - 2] = {
                "status": status_text,
                "url": extract_url_value(target_url),
            }
        return metadata
    finally:
        try:
            wb.close()
        except Exception:
            pass


def map_headers(df):
    rename_map = {}
    for column in df.columns:
        normalized = normalize_header_name(column)
        for standard_name, aliases in HEADER_ALIASES.items():
            normalized_aliases = [normalize_header_name(alias) for alias in aliases]
            if normalized in normalized_aliases:
                rename_map[column] = standard_name
                break
    mapped = df.rename(columns=rename_map)
    if not mapped.columns.duplicated().any():
        return mapped
    merged = pd.DataFrame(index=mapped.index)
    for column in dict.fromkeys(mapped.columns.tolist()):
        block = mapped.loc[:, mapped.columns == column]
        if isinstance(block, pd.Series):
            merged[column] = block
            continue
        if block.shape[1] == 1:
            merged[column] = block.iloc[:, 0]
            continue
        merged[column] = block.apply(
            lambda row: next((clean_text(value) for value in row if clean_text(value) != ""), ""),
            axis=1,
        )
    return merged


def parse_cap_value_unit(raw_value, raw_unit):
    value = clean_text(raw_value)
    unit = clean_text(raw_unit).upper().replace(" ", "")
    if value == "" and unit == "":
        return "", ""
    if value != "" and unit != "":
        return value, unit

    token = value.upper().replace(" ", "")
    token = token.replace("μF", "UF").replace("µF", "UF")
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(PF|NF|UF)", token)
    if match:
        return match.group(1), match.group(2)
    return value, unit


def cap_to_pf(value, unit):
    text = clean_text(value)
    normalized_unit = clean_text(unit).upper()
    if text == "":
        return None
    try:
        num = float(text)
    except Exception:
        return None

    if normalized_unit == "PF":
        factor = 1
    elif normalized_unit == "NF":
        factor = 1000
    elif normalized_unit == "UF":
        factor = 1000000
    else:
        return None
    return num * factor


def infer_brand_from_text(text):
    upper = clean_text(text).upper()
    for sheet_name, keys in SHEET_BRAND_MAP.items():
        if any(key.upper() in upper for key in keys):
            return sheet_name
    return ""


def canonical_sheet_name(brand_text, fallback_text=""):
    for candidate in [brand_text, fallback_text]:
        inferred = infer_brand_from_text(candidate)
        if inferred:
            return inferred
    cleaned_brand = clean_brand(brand_text)
    if cleaned_brand:
        return cleaned_brand
    fallback = clean_text(fallback_text)
    return fallback if fallback else "未分类"


def ensure_standard_columns(df):
    work = df.copy()
    for column in STANDARD_COLUMNS:
        if column not in work.columns:
            work[column] = ""
    return work


def standardize_df(df, inferred_brand, source_label):
    work = df.dropna(how="all").copy()
    if work.empty:
        return pd.DataFrame(columns=STANDARD_COLUMNS + ["目标Sheet"])

    work = map_headers(work)
    work = ensure_standard_columns(work)

    parsed = work.apply(lambda row: parse_cap_value_unit(row.get("容值", ""), row.get("容值单位", "")), axis=1)
    work["容值"] = [item[0] for item in parsed]
    work["容值单位"] = [item[1] for item in parsed]

    work["品牌"] = work["品牌"].apply(lambda x: clean_brand(x) or inferred_brand)
    work["型号"] = work["型号"].apply(clean_model)
    work["系列"] = work["系列"].apply(clean_text)
    work["尺寸（inch）"] = work["尺寸（inch）"].apply(clean_size)
    work["材质（介质）"] = work["材质（介质）"].apply(clean_material)
    work["容值"] = work["容值"].apply(clean_text)
    work["容值单位"] = work["容值单位"].apply(lambda x: clean_text(x).upper())
    work["容值误差"] = work["容值误差"].apply(clean_tol_for_match)
    work["耐压（V）"] = work["耐压（V）"].apply(clean_voltage)
    work["特殊用途"] = work["特殊用途"].apply(clean_text)
    work["备注1"] = work["备注1"].apply(clean_text)
    work["备注2"] = work["备注2"].apply(clean_text)
    work["备注3"] = work["备注3"].apply(clean_text)

    work = work[work["型号"] != ""].copy()
    if work.empty:
        return pd.DataFrame(columns=STANDARD_COLUMNS + ["目标Sheet"])

    work["目标Sheet"] = work.apply(
        lambda row: canonical_sheet_name(row.get("品牌", ""), inferred_brand or source_label),
        axis=1,
    )
    return work


def load_workbook_sheets(path):
    xls = pd.ExcelFile(path)
    sheets = OrderedDict()
    for sheet in xls.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet, dtype=str)
        metadata = extract_sheet_status_metadata(path, sheet, len(df))
        if metadata:
            if "备注1" not in df.columns:
                df["备注1"] = ""
            if "备注2" not in df.columns:
                df["备注2"] = ""
            for row_idx, info in metadata.items():
                if row_idx not in df.index:
                    continue
                if clean_text(df.at[row_idx, "备注1"]) == "" and clean_text(info.get("status", "")) != "":
                    df.at[row_idx, "备注1"] = clean_text(info.get("status", ""))
                if clean_text(df.at[row_idx, "备注2"]) == "" and clean_text(info.get("url", "")) != "":
                    df.at[row_idx, "备注2"] = clean_text(info.get("url", ""))

        if "备注1" not in df.columns:
            df["备注1"] = ""
        if "备注2" not in df.columns:
            df["备注2"] = ""
        if "品牌" in df.columns and "型号" in df.columns:
            for row_idx in df.index:
                brand = clean_brand(df.at[row_idx, "品牌"])
                if "三星" not in brand and "SAMSUNG" not in brand.upper():
                    continue
                status_text = clean_text(df.at[row_idx, "备注1"])
                if not looks_like_official_status(status_text):
                    continue
                if clean_text(df.at[row_idx, "备注2"]) != "":
                    continue
                part = clean_model(df.at[row_idx, "型号"])
                if part == "":
                    continue
                df.at[row_idx, "备注2"] = SAMSUNG_DETAIL_URL.format(part=part)
        sheets[sheet] = df
    return sheets


def analyze_master(master_path):
    print(f"主库文件: {master_path}")
    sheets = load_workbook_sheets(master_path)
    for sheet_name, df in sheets.items():
        work = ensure_standard_columns(df)
        model_count = int(work["型号"].astype(str).map(clean_model).ne("").sum())
        print(f"- {sheet_name}: 总行数 {len(df)}，有效型号行 {model_count}")


def create_official_templates(template_dir):
    template_dir.mkdir(parents=True, exist_ok=True)

    parts_template = template_dir / "official_confirmed_mlcc_template.csv"
    pd.DataFrame(OFFICIAL_TEMPLATE_SAMPLE, columns=OFFICIAL_TEMPLATE_COLUMNS).to_csv(parts_template, index=False, encoding="utf-8-sig")

    guide_path = template_dir / "official_confirmed_mlcc_template_README.md"
    guide_text = """# 官网确认料号录入模板

这个模板适用于你现在的拓库来源:
- 品牌官网料号查询系统可直接查到的料号
- 品牌官网筛选器可筛出的真实料号
- 品牌官方规格书中明确列出的真实料号

## 录入原则

1. 只有“官网能查到/能筛出来/规格书明确列出”的真实料号，才导入主库。
2. 规格书里只写“这个品牌能做到哪些规格”，但没有具体料号的，不要导入主库。
3. 建议把来源信息写进备注:
   - 备注1: `Mass Production / Pre Mass Production / NRND`
   - 备注2: `官方产品详情页链接`
   - 备注3: 可留空

## 列说明

- `品牌`: 建议写品牌名，例如 `Samsung`、`TDK`、`Murata`
- `型号`: 官方料号
- `系列`: 可写 `MLCC`、`常规`、`车规` 等
- `尺寸（inch）` / `材质（介质）` / `容值` / `容值单位` / `容值误差` / `耐压（V）`: 按官网结果填写
- `备注1`: 推荐写量产状态
- `备注2`: 推荐写官网产品详情页链接
- `备注3`: 当前版本建议留空

## 导入方式

先做分析，不写库:

```powershell
C:\\Users\\zjh\\AppData\\Local\\Programs\\Python\\Python314\\python.exe C:\\Users\\zjh\\Desktop\\data\\mlcc_excel_importer.py --input 你的文件.csv --dry-run
```

确认没问题后正式写入:

```powershell
C:\\Users\\zjh\\AppData\\Local\\Programs\\Python\\Python314\\python.exe C:\\Users\\zjh\\Desktop\\data\\mlcc_excel_importer.py --input 你的文件.csv
```
"""
    guide_path.write_text(guide_text, encoding="utf-8")

    print(f"\n已生成模板文件:")
    print(f"- {parts_template}")
    print(f"- {guide_path}")


def collect_input_files(input_paths, input_dir):
    files = []
    for item in input_paths:
        path = Path(item)
        if path.is_file():
            files.append(path)
    if input_dir and input_dir.exists():
        for path in sorted(input_dir.iterdir()):
            if path.is_file() and path.suffix.lower() in {".xlsx", ".xls", ".csv"}:
                files.append(path)
    deduped = []
    seen = set()
    for path in files:
        resolved = str(path.resolve())
        if resolved not in seen:
            deduped.append(path)
            seen.add(resolved)
    return deduped


def warn_if_input_inside_project(input_files):
    project_root = BASE_DIR.resolve()
    risky_files = []
    for path in input_files:
        try:
            resolved = path.resolve()
        except Exception:
            resolved = path
        if str(resolved).startswith(str(project_root)) and path.suffix.lower() in {".xlsx", ".xls"}:
            risky_files.append(resolved)
    if risky_files:
        print("\n警告: 这些原始 Excel 位于项目目录内，当前系统会递归扫描 data 下所有 xlsx，可能把它们误导入数据库。")
        for path in risky_files:
            print(f"- {path}")
        print("建议把原始品牌表放到项目目录外，例如默认目录 mlcc_incoming。")


def read_source_file(path):
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, dtype=str)
        return [(path.stem, df)]
    xls = pd.ExcelFile(path)
    return [(sheet, pd.read_excel(path, sheet_name=sheet, dtype=str)) for sheet in xls.sheet_names]


def make_row_key(row):
    brand_key = canonical_sheet_name(row.get("品牌", ""), row.get("目标Sheet", ""))
    return brand_key, clean_model(row.get("型号", ""))


def build_existing_key_map(master_sheets):
    key_map = {}
    for sheet_name, df in master_sheets.items():
        work = ensure_standard_columns(df)
        for row_idx, row in work.iterrows():
            model = clean_model(row.get("型号", ""))
            if model == "":
                continue
            key_map[(canonical_sheet_name(row.get("品牌", ""), sheet_name), model)] = {
                "sheet_name": sheet_name,
                "row_idx": row_idx,
            }
    return key_map


def merge_existing_and_incoming_row(existing_row, incoming_row):
    merged = existing_row.copy()
    changed = False
    for column in STANDARD_COLUMNS:
        if column in {"品牌", "型号"}:
            continue
        existing_value = clean_text(existing_row.get(column, ""))
        incoming_value = clean_text(incoming_row.get(column, ""))
        if incoming_value == "":
            continue
        if existing_value == "":
            merged[column] = incoming_value
            changed = True
            continue
        if column == "备注1" and looks_like_official_status(incoming_value) and existing_value != incoming_value:
            merged[column] = incoming_value
            changed = True
            continue
        if column == "备注2" and extract_url_value(incoming_value) != "" and existing_value != incoming_value:
            merged[column] = incoming_value
            changed = True
            continue
        if column == "备注3" and existing_value != incoming_value and incoming_value != "":
            merged[column] = incoming_value
            changed = True
    return merged, changed


def safe_concat_dataframes(frames, **kwargs):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r".*DataFrame concatenation with empty or all-NA entries.*",
            category=FutureWarning,
        )
        return pd.concat(frames, **kwargs)


def merge_into_master(master_sheets, import_rows):
    existing_keys = build_existing_key_map(master_sheets)
    sheet_order = list(master_sheets.keys())
    added_count = 0
    duplicate_count = 0
    updated_count = 0
    created_sheets = []

    for target_sheet, rows in import_rows.items():
        if target_sheet not in master_sheets:
            master_sheets[target_sheet] = pd.DataFrame(columns=STANDARD_COLUMNS)
            sheet_order.append(target_sheet)
            created_sheets.append(target_sheet)

        existing_df = ensure_standard_columns(master_sheets[target_sheet])
        new_rows = []
        for _, row in rows.iterrows():
            key = make_row_key(row)
            if key[1] == "":
                continue
            if key in existing_keys:
                duplicate_count += 1
                key_info = existing_keys[key]
                existing_sheet = key_info["sheet_name"]
                existing_row_idx = key_info["row_idx"]
                existing_sheet_df = ensure_standard_columns(master_sheets[existing_sheet])
                existing_row = existing_sheet_df.loc[existing_row_idx]
                merged_row, changed = merge_existing_and_incoming_row(existing_row, row)
                if changed:
                    for column in STANDARD_COLUMNS:
                        existing_sheet_df.at[existing_row_idx, column] = merged_row.get(column, "")
                    master_sheets[existing_sheet] = existing_sheet_df
                    updated_count += 1
                continue
            existing_keys[key] = target_sheet
            new_rows.append({column: row.get(column, "") for column in STANDARD_COLUMNS})
            added_count += 1

        if new_rows:
            master_sheets[target_sheet] = safe_concat_dataframes([existing_df, pd.DataFrame(new_rows)], ignore_index=True)
        else:
            master_sheets[target_sheet] = existing_df

    return OrderedDict((sheet, master_sheets[sheet]) for sheet in sheet_order), added_count, duplicate_count, updated_count, created_sheets


def backup_master(master_path):
    DEFAULT_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = DEFAULT_BACKUP_DIR / f"{master_path.stem}.backup_{timestamp}{master_path.suffix}"
    shutil.copy2(master_path, backup_path)
    return backup_path


def write_master(master_path, sheets):
    with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
        hyperlink_payload = []
        for sheet_name, df in sheets.items():
            out = ensure_standard_columns(df)
            out = out[STANDARD_COLUMNS].copy()
            display = out.copy()
            display.insert(display.columns.get_loc("备注1"), STATUS_DISPLAY_COLUMN, "")

            official_mask = display["备注1"].apply(looks_like_official_status)
            if official_mask.any():
                display.loc[official_mask, STATUS_DISPLAY_COLUMN] = display.loc[official_mask, "备注1"]
                display.loc[official_mask, "备注1"] = ""
                url_mask = display.loc[official_mask, "备注2"].apply(lambda x: extract_url_value(x) != "")
                if url_mask.any():
                    masked_index = display.loc[official_mask].index[url_mask]
                    display.loc[masked_index, "备注2"] = ""

            safe_sheet_name = sheet_name[:31]
            display.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            hyperlink_payload.append((safe_sheet_name, out[["备注1", "备注2"]].copy()))

        for safe_sheet_name, hyperlink_df in hyperlink_payload:
            ws = writer.book[safe_sheet_name]
            headers = [clean_text(cell.value) for cell in ws[1]]
            status_col = find_header_index(headers, [STATUS_DISPLAY_COLUMN])
            if status_col is None:
                continue
            for row_offset, (_, row) in enumerate(hyperlink_df.iterrows(), start=2):
                status_text = clean_text(row.get("备注1", ""))
                target_url = extract_url_value(row.get("备注2", ""))
                status_cell = ws.cell(row=row_offset, column=status_col)
                if looks_like_official_status(status_text):
                    status_cell.value = status_text
                    if target_url != "":
                        status_cell.hyperlink = target_url
                        status_cell.style = "Hyperlink"
                    else:
                        status_cell.hyperlink = None
                else:
                    status_cell.value = ""
                    status_cell.hyperlink = None


def refresh_sqlite_from_master(master_path, db_path):
    sheets = load_workbook_sheets(master_path)
    all_dfs = []
    for df in sheets.values():
        work = ensure_standard_columns(df)
        work = work[STANDARD_COLUMNS].copy()
        work["品牌"] = work["品牌"].apply(clean_text)
        work["型号"] = work["型号"].apply(clean_model)
        work["系列"] = work["系列"].apply(clean_text)
        work["尺寸（inch）"] = work["尺寸（inch）"].apply(clean_size)
        work["材质（介质）"] = work["材质（介质）"].apply(clean_material)
        work["容值"] = work["容值"].apply(clean_text)
        work["容值单位"] = work["容值单位"].apply(lambda x: clean_text(x).upper())
        work["容值误差"] = work["容值误差"].apply(clean_tol_for_match)
        work["耐压（V）"] = work["耐压（V）"].apply(clean_voltage)
        work["特殊用途"] = work["特殊用途"].apply(clean_text)
        work["备注1"] = work["备注1"].apply(clean_text)
        work["备注2"] = work["备注2"].apply(clean_text)
        work["备注3"] = work["备注3"].apply(clean_text)
        work["容值_pf"] = work.apply(lambda row: cap_to_pf(row["容值"], row["容值单位"]), axis=1)
        all_dfs.append(work)

    if not all_dfs:
        return 0

    df_all = safe_concat_dataframes(all_dfs, ignore_index=True).drop_duplicates()
    conn = sqlite3.connect(db_path)
    try:
        df_all.to_sql("components", conn, if_exists="replace", index=False)
    finally:
        conn.close()
    return len(df_all)


def repair_samsung_master(master_path, cache_path, db_path=None, dry_run=False):
    master_sheets = load_workbook_sheets(master_path)
    samsung_sheet_name = next((name for name in master_sheets if "三星" in str(name) or "Samsung" in str(name)), "")
    if samsung_sheet_name == "":
        return {
            "sheet_name": "",
            "changed_rows": 0,
            "backup_path": None,
            "db_row_count": 0,
        }

    cache_map = load_samsung_cache_map(cache_path)
    repaired_sheet, changed_rows = repair_samsung_sheet(master_sheets[samsung_sheet_name], cache_map)
    master_sheets[samsung_sheet_name] = repaired_sheet

    backup_path = None
    db_row_count = 0
    if not dry_run and changed_rows > 0:
        backup_path = backup_master(master_path)
        write_master(master_path, master_sheets)
        if db_path:
            db_row_count = refresh_sqlite_from_master(master_path, db_path)

    return {
        "sheet_name": samsung_sheet_name,
        "changed_rows": changed_rows,
        "backup_path": backup_path,
        "db_row_count": db_row_count,
    }


def main():
    parser = argparse.ArgumentParser(description="MLCC Excel 主库拓库导入工具")
    parser.add_argument("--master", default=str(DEFAULT_MASTER), help="主库 Excel 路径")
    parser.add_argument("--db", default=str(DEFAULT_DB_PATH), help="SQLite 数据库路径")
    parser.add_argument("--input", nargs="*", default=[], help="要导入的 Excel/CSV 文件路径")
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR), help="批量导入目录")
    parser.add_argument("--init-templates", action="store_true", help="生成官网确认料号录入模板")
    parser.add_argument("--template-dir", default=str(DEFAULT_TEMPLATE_DIR), help="模板输出目录")
    parser.add_argument("--repair-samsung", action="store_true", help="按三星官方缓存批量修复主库中的状态链接和关键规格字段")
    parser.add_argument("--samsung-cache", default=str(DEFAULT_SAMSUNG_CACHE), help="三星官方缓存 JSON 路径")
    parser.add_argument("--dry-run", action="store_true", help="只分析，不写回主库")
    args = parser.parse_args()

    master_path = Path(args.master)
    db_path = Path(args.db) if args.db else None
    input_dir = Path(args.input_dir) if args.input_dir else None
    template_dir = Path(args.template_dir)

    if args.init_templates:
        create_official_templates(template_dir)
        if not args.input:
            return

    if not master_path.exists():
        raise FileNotFoundError(f"主库不存在: {master_path}")

    if args.repair_samsung:
        repair_result = repair_samsung_master(
            master_path,
            Path(args.samsung_cache),
            db_path=db_path,
            dry_run=args.dry_run,
        )
        print("\n三星主库修复结果:")
        print(f"- 目标Sheet: {repair_result['sheet_name'] or '未找到三星Sheet'}")
        print(f"- 修复行数: {repair_result['changed_rows']}")
        if repair_result["backup_path"]:
            print(f"- 主库备份: {repair_result['backup_path']}")
        if repair_result["db_row_count"]:
            print(f"- 刷新后 SQLite 行数: {repair_result['db_row_count']}")
        if args.dry_run and repair_result["changed_rows"] > 0:
            print("- 当前为 dry-run，未写回主库。")
        if not args.input:
            return

    analyze_master(master_path)
    input_files = collect_input_files(args.input, input_dir)
    if not input_files:
        print("\n未发现要导入的原始文件。")
        print(f"你可以把品牌原始表放到: {input_dir}")
        print("或直接传入: --input brand_a.xlsx brand_b.xlsx")
        return

    print("\n准备导入这些文件:")
    for path in input_files:
        print(f"- {path}")
    warn_if_input_inside_project(input_files)

    import_rows = {}
    raw_row_count = 0
    standardized_count = 0

    for path in input_files:
        inferred_brand = infer_brand_from_text(path.stem)
        for sheet_name, df in read_source_file(path):
            raw_row_count += len(df)
            standardized = standardize_df(df, inferred_brand, sheet_name)
            standardized = fill_missing_tolerance_from_model(standardized)
            if standardized.empty:
                continue
            standardized_count += len(standardized)
            for target_sheet, group in standardized.groupby("目标Sheet", dropna=False):
                import_rows.setdefault(target_sheet, [])
                import_rows[target_sheet].append(group.copy())

    merged_rows = {sheet: safe_concat_dataframes(groups, ignore_index=True) for sheet, groups in import_rows.items()}
    master_sheets = load_workbook_sheets(master_path)
    merged_sheets, added_count, duplicate_count, updated_count, created_sheets = merge_into_master(master_sheets, merged_rows)

    print("\n导入分析结果:")
    print(f"- 原始总行数: {raw_row_count}")
    print(f"- 标准化后有效型号行: {standardized_count}")
    print(f"- 识别为重复并跳过: {duplicate_count}")
    print(f"- 预计新增入库: {added_count}")
    print(f"- 预计更新已有行: {updated_count}")
    if created_sheets:
        print(f"- 新增品牌Sheet: {', '.join(created_sheets)}")

    if args.dry_run:
        print("\n当前为 dry-run，未写回主库。")
        return

    backup_path = backup_master(master_path)
    write_master(master_path, merged_sheets)
    print(f"\n已写回主库: {master_path}")
    print(f"已生成备份: {backup_path}")
    if db_path:
        db_row_count = refresh_sqlite_from_master(master_path, db_path)
        print(f"已刷新 SQLite: {db_path}")
        print(f"SQLite 行数: {db_row_count}")


if __name__ == "__main__":
    main()
