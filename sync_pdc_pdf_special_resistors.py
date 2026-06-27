from __future__ import annotations

import argparse
import datetime as dt
import re
import shutil
import sqlite3
from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import component_matcher as cm
from sync_pdc_pdf_resistor_series import (
    E24_BASE,
    E96_BASE,
    low_ohm_code_mohm,
    resistor_code,
    resistance_values,
)
from sync_selected_cache_rows import stream_replace_prepared_rows


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "components.db"
PDF_DIR = Path.home() / "Desktop" / "被动产品线资料" / "信昌PDC" / "Resistor电阻"
RESISTOR_DIR = ROOT / "Resistor"
BRAND = "PSA(信昌电陶)"
SOURCE_TAG = "PDC special resistor official PDF generated"
CHECKED_AT = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

SIZE_DIMS = {
    "02": ("0402", "1.00", "0.50", "0.30"),
    "03": ("0603", "1.60", "0.80", "0.45"),
    "05": ("0805", "2.00", "1.25", "0.50"),
    "06": ("1206", "3.10", "1.60", "0.55"),
    "12": ("1210", "3.10", "2.60", "0.55"),
    "18": ("1218", "3.05", "4.60", "0.55"),
    "20": ("2010", "5.00", "2.50", "0.60"),
    "25": ("2512", "6.40", "3.20", "0.60"),
}

TOLERANCES = {
    "A": "±0.05%",
    "B": "±0.1%",
    "C": "±0.25%",
    "D": "±0.5%",
    "F": "±1%",
    "G": "±2%",
    "J": "±5%",
}


def clean_text(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"nan", "none"} else text


def decimal_text(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def display_resistance(value: Decimal) -> str:
    if value == 0:
        return "0Ω"
    if value < Decimal("0.001"):
        return f"{decimal_text(value * 1_000_000)}µΩ"
    if value < 1:
        return f"{decimal_text(value * 1000)}mΩ"
    if value >= 1_000_000:
        return f"{decimal_text(value / 1_000_000)}MΩ"
    if value >= 1_000:
        return f"{decimal_text(value / 1_000)}KΩ"
    return f"{decimal_text(value)}Ω"


def combined_e24_e96_values(minimum: Decimal, maximum: Decimal) -> list[Decimal]:
    return sorted(
        set(resistance_values(minimum, maximum, "E24"))
        | set(resistance_values(minimum, maximum, "E96"))
    )


def size_values(size_code: str) -> tuple[str, str, str, str]:
    return SIZE_DIMS[size_code]


def row(
    *,
    model: str,
    series: str,
    source_pdf: str,
    series_description: str,
    device_type: str,
    special_use: str,
    size_code: str,
    value_ohm: Decimal,
    tolerance: str,
    power: str,
    tcr: str,
    max_voltage: str = "",
    overload_voltage: str = "",
    note: str = "",
) -> dict[str, str]:
    inch, length, width, height = size_values(size_code)
    value_text = decimal_text(value_ohm)
    summary = " ".join(
        part
        for part in (
            series,
            inch,
            display_resistance(value_ohm),
            tolerance,
            power,
            f"TCR {tcr}" if tcr else "",
        )
        if part
    )
    return {
        "品牌": BRAND,
        "型号": model,
        "系列": series,
        "系列说明": series_description,
        "器件类型": device_type,
        "安装方式": "贴片",
        "封装代码": inch,
        "尺寸（inch）": inch,
        "尺寸（mm）": f"{length}x{width}x{height}",
        "长度（mm）": length,
        "宽度（mm）": width,
        "高度（mm）": height,
        "阻值@25C": value_text,
        "阻值单位": "Ω",
        "阻值误差": tolerance,
        "特殊用途": special_use,
        "规格摘要": summary,
        "备注1": f"{display_resistance(value_ohm)} {tolerance} {power} TCR {tcr}".strip(),
        "备注2": (
            f"最高工作电压 {max_voltage}; 最大过载电压 {overload_voltage}".strip()
        ),
        "备注3": note,
        "生产状态": "Active",
        "工作温度": "-55~155°C",
        "官网链接": str(PDF_DIR / source_pdf),
        "数据来源": f"{SOURCE_TAG}: {source_pdf}",
        "数据状态": "原厂PDF命名规则及规格范围生成",
        "校验时间": CHECKED_AT,
        "校验备注": note or f"依据 {source_pdf} 规格范围和命名规则生成",
        "_model_rule_authority": "pdc_official_pdf_generated",
        "_resistance_ohm": value_text,
    }


def faf_mh_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    sizes = {
        "02": (Decimal(40), Decimal(35_000), (("A", "0.0625W"),), "50V", "100V", "V"),
        "03": (
            Decimal(40),
            Decimal(130_000),
            (("B", "0.10W"), ("U", "0.15W")),
            "75V",
            "150V",
            "T",
        ),
        "05": (
            Decimal(10),
            Decimal(350_000),
            (("C", "0.125W"), ("S", "0.20W")),
            "100V",
            "200V",
            "T",
        ),
        "06": (
            Decimal(10),
            Decimal(1_000_000),
            (("D", "0.25W"), ("R", "0.40W")),
            "200V",
            "400V",
            "T",
        ),
    }
    for size_code, (minimum, maximum, powers, voltage, overload, packing) in sizes.items():
        for value in combined_e24_e96_values(minimum, maximum):
            for tolerance_code in ("A", "B", "C", "D", "F"):
                for watt_code, power in powers:
                    for tcr_code, tcr in (("V", "±10ppm/°C"), ("S", "±15ppm/°C"), ("Q", "±25ppm/°C"), ("P", "±50ppm/°C")):
                        model = (
                            f"FAF{size_code}{tolerance_code}{packing}{watt_code}"
                            f"{resistor_code(value, 4)}{tcr_code}MH"
                        )
                        rows.append(
                            row(
                                model=model,
                                series="FAF-MH",
                                source_pdf="FAF-MH.pdf",
                                series_description="信昌 FAF-MH AEC-Q200 抗硫化高精密薄膜电阻",
                                device_type="薄膜电阻",
                                special_use="高精密 | 薄膜 | AEC-Q200 | 抗硫化",
                                size_code=size_code,
                                value_ohm=value,
                                tolerance=TOLERANCES[tolerance_code],
                                power=power,
                                tcr=tcr,
                                max_voltage=voltage,
                                overload_voltage=overload,
                            )
                        )
    return rows


def fbf_10r_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    sizes = {
        "06": (("T",), (("-", "1/4W"), ("P", "1/2W")), "1.58V", "3.54V"),
        "20": (("P",), (("-", "3/4W"), ("P", "1W")), "2.74V", "6.12V"),
        "25": (("P",), (("-", "1W"), ("P", "2W")), "3.16V", "7.07V"),
    }
    for size_code, (packings, powers, voltage, overload) in sizes.items():
        for tolerance_code, e_series in (("F", "E96"), ("G", "E24"), ("J", "E24")):
            values = (
                combined_e24_e96_values(Decimal(1), Decimal(10))
                if tolerance_code == "F"
                else resistance_values(Decimal(1), Decimal(10), e_series)
            )
            for value in values:
                rcode = resistor_code(value, 4 if tolerance_code == "F" else 3)
                for packing in packings:
                    for power_code, power in powers:
                        for grade, grade_use in (
                            ("", "标准"),
                            ("M", "AEC-Q200"),
                            ("MD", "AEC-Q200 | 抗硫化"),
                        ):
                            model = (
                                f"FBF{size_code}{tolerance_code}{packing}"
                                f"{power_code}{rcode}N{grade}"
                            )
                            rows.append(
                                row(
                                    model=model,
                                    series="FBF-10R",
                                    source_pdf="FBF_10R.pdf",
                                    series_description="信昌 FBF ≥1Ω 金属浆料电流检测厚膜电阻",
                                    device_type="厚膜电阻",
                                    special_use=f"电流检测 | ≥1Ω | 金属浆料 | {grade_use}",
                                    size_code=size_code,
                                    value_ohm=value,
                                    tolerance=TOLERANCES[tolerance_code],
                                    power=power,
                                    tcr="±100ppm/°C",
                                    max_voltage=voltage,
                                    overload_voltage=overload,
                                )
                            )
    return rows


def mohm_e24_values(minimum: int, maximum: int) -> list[Decimal]:
    values: set[Decimal] = set()
    for exponent in range(0, 4):
        scale = Decimal(10) ** exponent
        for base in E24_BASE:
            value = base * scale
            if minimum <= value <= maximum:
                values.add(value)
    return sorted(values)


def fcf_e_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    sizes = {
        "03": ("1/8W", "1/4W", "D"),
        "05": ("1/4W", "1/2W", "F"),
        "06": ("1/3W", "3/4W", "G"),
        "12": ("2/3W", "3/4W", "G"),
        "20": ("3/4W", "1W", "H"),
        "25": ("1W", "2W", "J"),
    }
    for size_code, (normal_power, high_power, high_power_code) in sizes.items():
        packing = "P" if size_code in {"20", "25"} else "T"
        ranges = [(100, 910, "LE", "±200ppm/°C")]
        if size_code != "03":
            ranges.extend(
                [
                    (50, 99, "KE", "±400ppm/°C"),
                    (22, 49, "JE", "±600ppm/°C"),
                    (10, 20, "HE", "±1000ppm/°C"),
                ]
            )
        else:
            ranges.append((50, 99, "KE", "±400ppm/°C"))
        for minimum, maximum, control_code, tcr in ranges:
            for mohm in mohm_e24_values(minimum, maximum):
                value = mohm / Decimal(1000)
                rcode = low_ohm_code_mohm(mohm, width=3)
                for tolerance_code in ("F", "G", "J"):
                    normal_model = f"FCF{size_code}{tolerance_code}{packing}-{rcode}-E"
                    rows.append(
                        row(
                            model=normal_model,
                            series="FCF-E",
                            source_pdf="FCF-E.pdf",
                            series_description="信昌 FCF-E 通用厚膜低阻电流检测电阻",
                            device_type="厚膜电阻",
                            special_use="电流检测 | 低阻 | 厚膜 | 通用",
                            size_code=size_code,
                            value_ohm=value,
                            tolerance=TOLERANCES[tolerance_code],
                            power=normal_power,
                            tcr=tcr,
                        )
                    )
                    high_model = (
                        f"FCF{size_code}{tolerance_code}{packing}"
                        f"{high_power_code}{rcode}{control_code}"
                    )
                    rows.append(
                        row(
                            model=high_model,
                            series="FCF-E",
                            source_pdf="FCF-E.pdf",
                            series_description="信昌 FCF-E 高功率厚膜低阻电流检测电阻",
                            device_type="厚膜电阻",
                            special_use="电流检测 | 低阻 | 厚膜 | 高功率",
                            size_code=size_code,
                            value_ohm=value,
                            tolerance=TOLERANCES[tolerance_code],
                            power=high_power,
                            tcr=tcr,
                        )
                    )
    return rows


def fcf_green_rows(conn: sqlite3.Connection) -> list[dict[str, str]]:
    frame = pd.read_sql_query(
        """
        SELECT *
        FROM components
        WHERE [品牌] = ?
          AND [系列] = 'FCF'
          AND [尺寸（inch）] IN ('0402','0603','0805','1206','1218','2010','2512')
          AND [阻值误差] IN ('±1%','±5%')
        """,
        conn,
        params=(BRAND,),
    )
    rows: list[dict[str, str]] = []
    pattern = re.compile(r"^FCF(?:02|03|05|06|18|20|25)[FJ][TVWPQ]-[0-9R]+$")
    for item in frame.to_dict("records"):
        if not pattern.fullmatch(clean_text(item.get("型号"))):
            continue
        item["型号"] = f"{item['型号']}-G"
        item["系列"] = "FCF-G"
        item["系列说明"] = "信昌 FCF-G RoHS 豁免物质无铅绿色通用厚膜电阻"
        item["特殊用途"] = "通用厚膜 | RoHS Exemption Free | Pb≤100ppm | Green"
        item["数据来源"] = f"{SOURCE_TAG}: FCF-G.pdf"
        item["官网链接"] = str(PDF_DIR / "FCF-G.pdf")
        item["数据状态"] = "原厂PDF命名规则及规格范围生成"
        item["校验时间"] = CHECKED_AT
        item["校验备注"] = "由同尺寸/阻值的 FCF 规格按 FCF-G 原厂命名规则追加 -G"
        item["_model_rule_authority"] = "pdc_official_pdf_generated"
        rows.append(item)
    return rows


def triple_power_rows(prefix: str) -> list[dict[str, str]]:
    is_fps = prefix == "FPS"
    source_pdf = "FPS_Triple.pdf" if is_fps else "FPF_Triple.pdf"
    sizes = (
        {
            "03": ("E", "1/3W"),
            "05": ("F", "1/2W"),
            "06": ("G", "3/4W"),
        }
        if is_fps
        else {
            "03": ("E", "1/3W"),
            "05": ("F", "1/2W"),
            "06": ("G", "3/4W"),
            "12": ("G", "3/4W"),
            "20": ("I", "1.5W"),
            "25": ("K", "3W"),
        }
    )
    suffixes = (
        (("", "标准"), ("M", "AEC-Q200"), ("MB", "AEC-Q200 抗硫化50°C"), ("MD", "AEC-Q200 抗硫化90°C"))
        if is_fps
        else (("", "标准"), ("M", "AEC-Q200"))
    )
    rows: list[dict[str, str]] = []
    for size_code, (watt_code, power) in sizes.items():
        packing = "P" if size_code in {"20", "25"} else "T"
        for tolerance_code, e_series in (("F", "E96"), ("J", "E24")):
            values = (
                combined_e24_e96_values(Decimal(1), Decimal(1_000_000))
                if tolerance_code == "F"
                else resistance_values(Decimal(1), Decimal(1_000_000), e_series)
            )
            for value in values:
                rcode = resistor_code(value, 4 if tolerance_code == "F" else 3)
                if tolerance_code == "J":
                    rcode = f"{rcode}_"
                tcr = "±100ppm/°C" if value >= 10 else (
                    "±200ppm/°C" if size_code in {"03", "06"} else "±150ppm/°C"
                )
                tcr_code = "N" if "100" in tcr else ("Y" if "150" in tcr else "L")
                for suffix, suffix_use in suffixes:
                    model = (
                        f"{prefix}{size_code}{tolerance_code}{packing}{watt_code}"
                        f"{rcode}{tcr_code}{suffix}"
                    )
                    rows.append(
                        row(
                            model=model,
                            series=f"{prefix}-Triple",
                            source_pdf=source_pdf,
                            series_description=(
                                f"信昌 {prefix} 三倍额定功率"
                                + ("抗浪涌厚膜电阻" if is_fps else "厚膜电阻")
                            ),
                            device_type="厚膜电阻",
                            special_use=(
                                "三倍功率 | "
                                + ("抗浪涌 | " if is_fps else "")
                                + suffix_use
                            ),
                            size_code=size_code,
                            value_ohm=value,
                            tolerance=TOLERANCES[tolerance_code],
                            power=power,
                            tcr=tcr,
                            note=(
                                "FPF_Triple.pdf 料号示例将 0805 写作 FPF08，"
                                "规格表和尺寸表均定义为 FPF05；批量生成采用规格表 FPF05。"
                                if prefix == "FPF" and size_code == "05"
                                else ""
                            ),
                        )
                    )
    return rows


def fcf_array_rows() -> list[dict[str, str]]:
    configs = {
        "220": ("0402", "4P2R", "1/16W", ("J",), Decimal(10), Decimal(1_000_000), True),
        "240": ("0402", "8P4R Convex", "1/16W", ("F", "J"), Decimal(10), Decimal(1_000_000), True),
        "241": ("0402", "8P4R Concave", "1/16W", ("F", "J"), Decimal(3), Decimal(1_000_000), True),
        "320": ("0603", "4P2R Convex", "1/10W", ("F", "J"), Decimal(10), Decimal(1_000_000), True),
        "340": ("0603", "8P4R Convex", "1/10W", ("F", "J"), Decimal(10), Decimal(1_000_000), True),
        "341": ("0603", "8P4R Concave", "1/10W", ("F", "J"), Decimal(10), Decimal(1_000_000), True),
        "35R": ("0603", "10P8R", "1/16W", ("J",), Decimal(10), Decimal(100_000), False),
        "370": ("0603", "16P8R", "1/16W", ("F", "J"), Decimal(10), Decimal(100_000), True),
    }
    rows: list[dict[str, str]] = []
    size_code_by_inch = {"0402": "02", "0603": "03"}
    for type_code, (inch, circuit, power, tolerances, minimum, maximum, jumper) in configs.items():
        values = resistance_values(minimum, maximum, "E24")
        if jumper:
            values = [Decimal(0)] + values
        for tolerance_code in tolerances:
            for value in values:
                actual_tol = "J" if value == 0 else tolerance_code
                model = f"FCF{type_code}{actual_tol}T-{resistor_code(value, 3)}"
                rows.append(
                    row(
                        model=model,
                        series="FCF-Array",
                        source_pdf="FCF-Array.pdf",
                        series_description=f"信昌 FCF 阵列厚膜电阻网络（{circuit}）",
                        device_type="排阻（Resistor Network）",
                        special_use=f"电阻阵列 | {circuit} | 厚膜",
                        size_code=size_code_by_inch[inch],
                        value_ohm=value,
                        tolerance=TOLERANCES[actual_tol],
                        power=power,
                        tcr="±200ppm/°C" if value else "-300~+500ppm/°C",
                    )
                )
    return rows


def fmf_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    configs = [
        ("06", "F", "1/2W", "Low EMF", "BH", [1, 2], "±75ppm/°C"),
        ("06", "F", "1/2W", "Low EMF", "BH", [3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 20, 25], "±70ppm/°C"),
        ("06", "F", "1/2W", "Standard", "LH", [5, 10, 15, 18, 20, 25, 30], "±50ppm/°C"),
        ("06", "H", "1W", "Low EMF", "BH", [1, 2], "±75ppm/°C"),
        ("06", "H", "1W", "Low EMF", "BH", [3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 20, 25], "±70ppm/°C"),
        ("06", "H", "1W", "Standard", "LH", [5, 10, 15, 18, 20, 25, 30], "±50ppm/°C"),
        ("25", "H", "1W", "Low EMF", "BH", [1, 2, 2.5, 3, 4, 5, 10, 15, 20, 25], "±70ppm/°C"),
        ("25", "H", "1W", "Standard", "LH", [3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 18, 20, 22, 25, 30, 33, 35, 40, 50, 60, 70, 75, 80, 100], "±50ppm/°C"),
        ("25", "J", "2W", "Low EMF", "BH", [1, 2, 2.5, 3, 4, 5, 10, 15, 20, 25], "±70ppm/°C"),
        ("25", "J", "2W", "Standard", "LH", [3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 18, 20, 22, 25, 30, 33, 35, 40, 50, 60, 70, 75, 80, 100], "±50ppm/°C"),
        ("25", "K", "3W", "Low EMF", "BH", [0.5, 0.75, 1, 2, 2.5, 3, 4, 5, 6, 7, 8, 9, 10], "±70ppm/°C"),
        ("25", "K", "3W", "Low EMF", "BH", [20], "±50ppm/°C"),
        ("25", "K", "3W", "Standard", "LH", [5, 6, 8, 10], "±70ppm/°C"),
        ("25", "K", "3W", "Standard", "LH", [12, 14, 15, 16, 18, 20, 25, 30, 33, 35, 40, 50, 60, 75, 80, 100], "±50ppm/°C"),
    ]
    for size_code, watt_code, power, alloy, suffix, values, tcr in configs:
        pack = "T" if size_code == "06" else "P"
        tolerance_codes = ("F", "G", "J") if size_code == "06" else ("D", "F", "G", "J")
        for mohm_raw in values:
            mohm = Decimal(str(mohm_raw))
            value = mohm / Decimal(1000)
            rcode = low_ohm_code_mohm(mohm, width=3)
            use_x = size_code == "25" and mohm <= 3
            if use_x:
                rcode = f"{rcode}X"
            for tolerance_code in tolerance_codes:
                for actual_suffix, grade in (
                    (suffix, "通用"),
                    (f"{suffix}M", "AEC-Q200 | 抗硫化"),
                ):
                    model = (
                        f"FMF{size_code}{tolerance_code}{pack}{watt_code}"
                        f"{rcode}-{actual_suffix}"
                    )
                    rows.append(
                        row(
                            model=model,
                            series="FMF",
                            source_pdf="FMF.pdf",
                            series_description="信昌 FMF AEC-Q200 金属条高功率电流检测合金电阻",
                            device_type="合金电阻",
                            special_use=f"电流检测 | 金属条 | 低阻 | {alloy} | {grade}",
                            size_code=size_code,
                            value_ohm=value,
                            tolerance=TOLERANCES[tolerance_code],
                            power=power,
                            tcr=tcr,
                            note=(
                                "按 FMF.pdf Rating 与 Part Number 表生成；"
                                "BH=Low EMF，LH=Standard，后缀 M=AEC-Q200/抗硫化"
                            ),
                        )
                    )
    return rows


def build_rows() -> pd.DataFrame:
    with sqlite3.connect(DB_PATH, timeout=120) as conn:
        rows = (
            faf_mh_rows()
            + fbf_10r_rows()
            + fcf_e_rows()
            + fcf_green_rows(conn)
            + triple_power_rows("FPF")
            + triple_power_rows("FPS")
            + fcf_array_rows()
            + fmf_rows()
        )
    frame = pd.DataFrame(rows)
    return frame.drop_duplicates(
        subset=["品牌", "型号", "器件类型"], keep="last"
    ).reset_index(drop=True)


def db_columns(conn: sqlite3.Connection) -> list[str]:
    return [item[1] for item in conn.execute('PRAGMA table_info("components")')]


def backup_database() -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = DB_PATH.with_name(f"{DB_PATH.name}.pdc_special_resistors_{timestamp}.bak")
    shutil.copy2(DB_PATH, path)
    return path


def apply_database(frame: pd.DataFrame, dry_run: bool, no_backup: bool) -> Path | None:
    if dry_run:
        return None
    backup = None if no_backup else backup_database()
    with sqlite3.connect(DB_PATH, timeout=120) as conn:
        conn.execute("PRAGMA busy_timeout = 120000")
        columns = db_columns(conn)
        conn.execute("DELETE FROM components WHERE [数据来源] LIKE ?", (f"{SOURCE_TAG}:%",))
        records: list[dict[str, str]] = []
        for item in frame.to_dict("records"):
            record = {column: "" for column in columns}
            for column, value in item.items():
                if column in record:
                    record[column] = clean_text(value)
            records.append(record)
        models = frame["型号"].drop_duplicates().tolist()
        for offset in range(0, len(models), 600):
            chunk = models[offset : offset + 600]
            placeholders = ",".join("?" for _ in chunk)
            conn.execute(
                f"DELETE FROM components WHERE [品牌] = ? AND [型号] IN ({placeholders})",
                [BRAND, *chunk],
            )
        pd.DataFrame(records, columns=columns).to_sql(
            "components", conn, if_exists="append", index=False, chunksize=10, method="multi"
        )
        conn.commit()
    return backup


def workbook_row(item: dict[str, str], headers: list[str]) -> list[str]:
    mapped = dict(item)
    mapped["阻值"] = clean_text(item.get("阻值@25C"))
    mapped["功率"] = ""
    note1 = clean_text(item.get("备注1"))
    power_match = re.search(r"(?:\d+(?:\.\d+)?W|\d+/\d+W)", note1)
    if power_match:
        mapped["功率"] = power_match.group(0)
    mapped["TCR（ppm）"] = note1.split("TCR ", 1)[-1] if "TCR " in note1 else ""
    return [clean_text(mapped.get(header)) for header in headers]


def update_workbook(path: Path, frame: pd.DataFrame, dry_run: bool) -> int:
    workbook = load_workbook(path)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        headers = [clean_text(cell.value) for cell in sheet[1]]
        source_index = headers.index("数据来源") + 1
        matching_rows = [
            row_index
            for row_index in range(2, sheet.max_row + 1)
            if clean_text(sheet.cell(row_index, source_index).value).startswith(SOURCE_TAG)
        ]
        delete_ranges: list[tuple[int, int]] = []
        if matching_rows:
            start = previous = matching_rows[0]
            for row_index in matching_rows[1:]:
                if row_index == previous + 1:
                    previous = row_index
                    continue
                delete_ranges.append((start, previous))
                start = previous = row_index
            delete_ranges.append((start, previous))
        for start, end in reversed(delete_ranges):
            sheet.delete_rows(start, end - start + 1)
        for item in frame.to_dict("records"):
            sheet.append(workbook_row(item, headers))
        if not dry_run:
            workbook.save(path)
        return len(frame)
    finally:
        workbook.close()


def update_workbooks(frame: pd.DataFrame, dry_run: bool) -> dict[str, int]:
    targets = {
        "薄膜电阻": RESISTOR_DIR / "薄膜电阻.xlsx",
        "厚膜电阻": RESISTOR_DIR / "厚膜电阻.xlsx",
        "排阻（Resistor Network）": RESISTOR_DIR / "厚膜电阻.xlsx",
        "合金电阻": RESISTOR_DIR / "合金电阻.xlsx",
    }
    grouped: dict[Path, list[pd.DataFrame]] = {}
    for device_type, path in targets.items():
        subset = frame[frame["器件类型"] == device_type]
        if not subset.empty:
            grouped.setdefault(path, []).append(subset)
    counts: dict[str, int] = {}
    for path, parts in grouped.items():
        combined = pd.concat(parts, ignore_index=True)
        counts[path.name] = update_workbook(path, combined, dry_run)
    return counts


def refresh_search_sidecar_rows_bulk(prepared: pd.DataFrame) -> dict[str, int]:
    search_path = Path(cm.SEARCH_DB_PATH)
    if not search_path.exists():
        raise FileNotFoundError(search_path)

    frames = cm.build_search_sidecar_frames(prepared)
    if not frames:
        raise RuntimeError("no search sidecar frames generated")

    pairs = (
        prepared.loc[:, ["品牌", "型号"]]
        .drop_duplicates()
        .rename(columns={"品牌": "brand", "型号": "model"})
    )
    specs = cm.get_search_sidecar_table_specs()
    with sqlite3.connect(search_path, timeout=180) as conn:
        conn.execute("PRAGMA busy_timeout = 180000")
        conn.execute("DROP TABLE IF EXISTS main._pdc_refresh_models")
        conn.execute("DROP TABLE IF EXISTS temp._pdc_refresh_models")
        conn.execute(
            "CREATE TEMP TABLE _pdc_refresh_models "
            "(brand TEXT NOT NULL, model TEXT NOT NULL)"
        )
        conn.executemany(
            "INSERT INTO temp._pdc_refresh_models(brand, model) VALUES (?, ?)",
            pairs.itertuples(index=False, name=None),
        )
        conn.execute(
            "CREATE INDEX _pdc_refresh_models_key "
            "ON _pdc_refresh_models(brand, model)"
        )
        for table_name, spec in specs.items():
            columns = set(spec.get("columns", []))
            if not {"品牌", "型号"}.issubset(columns):
                continue
            try:
                conn.execute(
                    f'DELETE FROM "{table_name}" '
                    "WHERE EXISTS ("
                    "SELECT 1 FROM temp._pdc_refresh_models AS refresh "
                    f'WHERE refresh.brand = "{table_name}"."品牌" '
                    f'AND refresh.model = "{table_name}"."型号"'
                    ")"
                )
            except sqlite3.OperationalError:
                continue

        for table_name, sidecar_frame in frames.items():
            if sidecar_frame is None or sidecar_frame.empty:
                continue
            sidecar_frame.to_sql(
                table_name,
                conn,
                if_exists="append",
                index=False,
                method="multi",
                chunksize=cm.sqlite_bulk_insert_chunksize(sidecar_frame),
            )

        row_counts: dict[str, int] = {}
        for table_name in specs:
            try:
                row_counts[table_name] = int(
                    conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
                )
            except Exception:
                row_counts[table_name] = 0
        cm.write_search_index_meta(conn, row_counts)
        conn.commit()
    return row_counts


def refresh_cache(
    frame: pd.DataFrame,
    skip_cache: bool,
    skip_prepared_cache: bool = False,
) -> dict[str, int]:
    if skip_cache:
        return {}
    models = frame["型号"].drop_duplicates().tolist()
    parts: list[pd.DataFrame] = []
    with sqlite3.connect(DB_PATH, timeout=120) as conn:
        conn.execute("PRAGMA busy_timeout = 120000")
        for offset in range(0, len(models), 800):
            chunk = models[offset : offset + 800]
            placeholders = ",".join("?" for _ in chunk)
            parts.append(
                pd.read_sql_query(
                    f"SELECT * FROM components WHERE [品牌] = ? AND [型号] IN ({placeholders})",
                    conn,
                    params=[BRAND, *chunk],
                )
            )
    selected = pd.concat(parts, ignore_index=True)
    prepared = cm.prepare_search_dataframe(cm.deduplicate_component_rows(selected))
    print(f"cache_prepared_rows={len(prepared)}", flush=True)
    removed = inserted = 0
    if not skip_prepared_cache:
        removed, inserted = stream_replace_prepared_rows(prepared)
        print(f"cache_prepared_inserted={inserted}", flush=True)
    sidecar = refresh_search_sidecar_rows_bulk(prepared)
    return {
        "selected_db_rows": len(selected),
        "prepared_removed": int(removed),
        "prepared_inserted": int(inserted),
        "search_core_rows": int(sidecar.get(cm.COMPONENTS_SEARCH_CORE_TABLE, 0)),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sync special PDC resistor PDF series into database and workbooks."
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-db", action="store_true")
    parser.add_argument("--skip-workbooks", action="store_true")
    parser.add_argument("--skip-cache", action="store_true")
    parser.add_argument("--skip-prepared-cache", action="store_true")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()

    frame = build_rows()
    print(f"generated_rows={len(frame)}")
    print(frame.groupby(["系列", "器件类型"]).size().sort_values(ascending=False).to_string())

    backup = None
    if not args.skip_db:
        backup = apply_database(frame, args.dry_run, args.no_backup)
    if backup:
        print(f"backup_path={backup}")

    if not args.skip_workbooks:
        for name, count in update_workbooks(frame, args.dry_run).items():
            print(f"workbook_rows[{name}]={count}")

    if not args.dry_run:
        for key, value in refresh_cache(
            frame,
            args.skip_cache,
            args.skip_prepared_cache,
        ).items():
            print(f"{key}={value}")


if __name__ == "__main__":
    main()
