from __future__ import annotations

import argparse
import datetime as dt
import re
import shutil
import sqlite3
from pathlib import Path

import fitz
import pandas as pd
import pdfplumber
from openpyxl import load_workbook

import component_matcher as cm
from incremental_semiconductor_cache_update import refresh_search_sidecar_rows
from sync_selected_cache_rows import stream_replace_prepared_rows


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "components.db"
PDF_DIR = Path.home() / "Desktop" / "被动产品线资料" / "信昌PDC" / "Resistor电阻"
RESISTOR_DIR = ROOT / "Resistor"
NTC_WORKBOOK = RESISTOR_DIR / "热敏电阻_NTC.xlsx"
VDR_WORKBOOK = RESISTOR_DIR / "压敏电阻_VDR.xlsx"
BRAND = "JOYIN(久尹)"
SOURCE_TAG = "JOYIN official PDF sync 2026-06"
CHECKED_AT = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

R_TOLERANCE_CODES = {
    "10": ("K", "±10%"),
    "5": ("J", "±5%"),
    "3": ("H", "±3%"),
    "2": ("G", "±2%"),
    "1": ("F", "±1%"),
}
B_TOLERANCE_CODES = {
    "5": ("J", "±5%"),
    "3": ("H", "±3%"),
    "2": ("G", "±2%"),
    "1": ("F", "±1%"),
}

NTC_PROFILES = {
    "JAS": {
        "description": "久尹 JAS 端子头/耳环式 NTC 温度传感器，适用于家电、办公设备、汽车及安防测温",
        "use": "温度传感 | 端子头 | 引线/端子可定制",
        "mount": "引线/端子",
        "package": "端子头",
        "temperature": "-40~125°C",
    },
    "JAT": {
        "description": "久尹 JAT 环氧树脂包封径向引线 NTC 热敏电阻",
        "use": "温度检测 | 环氧包封 | 径向引线",
        "mount": "插件",
        "package": "径向引线",
        "temperature": "-40~125°C",
    },
    "JCR03": {
        "description": "久尹 JCR03 约 3mm 环氧树脂包封径向引线 NTC 热敏电阻",
        "use": "温度检测 | 3mm 本体 | 环氧包封 | 径向引线",
        "mount": "插件",
        "package": "3mm 径向引线",
        "temperature": "-40~125°C",
        "diameter": "3",
    },
    "JCR05": {
        "description": "久尹 JCR05 约 5mm 环氧树脂包封径向引线 NTC 热敏电阻",
        "use": "温度检测 | 5mm 本体 | 环氧包封 | 径向引线",
        "mount": "插件",
        "package": "5mm 径向引线",
        "temperature": "-40~125°C",
        "diameter": "5",
    },
    "JFR": {
        "description": "久尹 JFR 超小型快速响应引线 NTC 温度传感器",
        "use": "快速响应 | 超小型 | 引线测温 | 线长可选",
        "mount": "引线",
        "package": "快速响应探头",
        "temperature": "-40~125°C",
    },
    "JSR": {
        "description": "久尹 JSR 引线式 NTC 温度传感器，线材、长度及端子可按应用选择",
        "use": "温度传感 | 引线式 | 线材/长度/端子可定制",
        "mount": "引线",
        "package": "引线探头",
        "temperature": "-40~125°C",
    },
}

VARISTOR_PROFILES = {
    "JVT": {
        "description": "久尹 JVT 径向引线金属氧化物压敏电阻",
        "source_files": ("JVT.pdf", "JVT_Automotive.pdf"),
    },
    "JVZ": {
        "description": "久尹 JVZ 径向引线金属氧化物压敏电阻",
        "source_files": ("JVZ.pdf",),
    },
}

VARISTOR_FAMILY = {
    "N": ("标准型", "标准浪涌保护"),
    "S": ("高浪涌型", "高浪涌保护"),
    "U": ("超高浪涌型", "超高浪涌保护"),
}


def clean_text(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"nan", "none"} else text


def numeric_text(value: object) -> str:
    text = clean_text(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return match.group(0) if match else ""


def resistance_display(ohms: str) -> tuple[str, str, str]:
    value = float(ohms)
    if value >= 1_000_000:
        shown = f"{value / 1_000_000:g}"
        return shown, "MΩ", f"{shown}MΩ"
    if value >= 1_000:
        shown = f"{value / 1_000:g}"
        return shown, "KΩ", f"{shown}KΩ"
    shown = f"{value:g}"
    return shown, "Ω", f"{shown}Ω"


def tolerance_pairs(raw: str, mapping: dict[str, tuple[str, str]]) -> list[tuple[str, str]]:
    result: list[tuple[str, str]] = []
    for token in re.findall(r"\d+(?:\.\d+)?", raw):
        normalized = token.rstrip("0").rstrip(".") if "." in token else token
        if normalized in mapping:
            result.append(mapping[normalized])
    return result


def expand_ntc_template(template: str, r_code: str, b_code: str) -> str:
    match = re.fullmatch(
        r"(?P<prefix>(?:JAS|JAT|JCR|JFR|JSR)\d{3})X(?P<bvalue>\d{3})Y(?P<condition>[AB])",
        template,
    )
    if not match:
        raise ValueError(f"Unsupported NTC template: {template}")
    return (
        f"{match.group('prefix')}{r_code}{match.group('bvalue')}"
        f"{b_code}{match.group('condition')}"
    )


def ntc_table_rows(pdf_path: Path, series: str) -> list[dict[str, str]]:
    profile = NTC_PROFILES[series]
    prefix = "JCR" if series.startswith("JCR") else series
    source_note = f"{SOURCE_TAG}: {pdf_path.name}"
    rows: list[dict[str, str]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            for table in page.extract_tables() or []:
                for raw_row in table or []:
                    cells = [" ".join(clean_text(cell).split()) for cell in raw_row]
                    if not cells or not re.fullmatch(
                        rf"{prefix}\d{{3}}X\d{{3}}Y[AB]", cells[0]
                    ):
                        continue
                    if len(cells) < 8:
                        continue
                    template = cells[0]
                    r25_ohm = numeric_text(cells[1])
                    b_value = numeric_text(cells[3])
                    if not r25_ohm or not b_value:
                        continue
                    r_tolerances = tolerance_pairs(cells[2], R_TOLERANCE_CODES)
                    b_tolerances = tolerance_pairs(cells[4], B_TOLERANCE_CODES)
                    condition = "25/50°C" if template.endswith("A") else "25/85°C"
                    display_value, display_unit, resistance_label = resistance_display(r25_ohm)
                    official_mismatch = ""
                    encoded_resistance = template[len(prefix) : len(prefix) + 3]
                    expected_ohm = int(encoded_resistance[:2]) * (10 ** int(encoded_resistance[2]))
                    if abs(float(r25_ohm) - expected_ohm) > 0.01:
                        official_mismatch = (
                            f"原厂表格型号阻值码推算为 {expected_ohm:g}Ω，"
                            f"但表格阻值栏为 {float(r25_ohm):g}Ω；本库保留原厂表格阻值。"
                        )
                    for r_code, r_tolerance in r_tolerances:
                        for b_code, b_tolerance in b_tolerances:
                            model = expand_ntc_template(template, r_code, b_code)
                            notes = [
                                f"原厂电气模板 {template}",
                                f"PDF 第 {page_no} 页",
                                "型号为已展开电阻公差/B值公差的电气核心料号",
                            ]
                            if series in {"JAS", "JSR"}:
                                notes.append("线材、线长、端子等机械配置需按客户需求确认")
                            if official_mismatch:
                                notes.append(official_mismatch)
                            rows.append(
                                {
                                    "品牌": BRAND,
                                    "型号": model,
                                    "系列": series,
                                    "系列说明": profile["description"],
                                    "器件类型": "热敏电阻（NTC Thermistor）",
                                    "安装方式": profile["mount"],
                                    "封装代码": profile["package"],
                                    "尺寸（mm）": (
                                        f"{profile.get('diameter')}mm"
                                        if profile.get("diameter")
                                        else ""
                                    ),
                                    "直径（mm）": profile.get("diameter", ""),
                                    "阻值@25C": display_value,
                                    "阻值单位": display_unit,
                                    "阻值误差": r_tolerance,
                                    "B值": b_value,
                                    "B值条件": condition,
                                    "工作温度": profile["temperature"],
                                    "特殊用途": profile["use"],
                                    "规格摘要": (
                                        f"{series} {resistance_label} {r_tolerance} "
                                        f"B{condition}={b_value}K {b_tolerance}"
                                    ),
                                    "备注1": f"B值误差 {b_tolerance}",
                                    "备注2": (
                                        f"耗散系数 {clean_text(cells[5])}; "
                                        f"热时间常数 {clean_text(cells[6])}; "
                                        f"最大功率 {clean_text(cells[7])}"
                                    ),
                                    "备注3": "；".join(notes),
                                    "生产状态": "Active",
                                    "官网链接": str(pdf_path),
                                    "数据来源": source_note,
                                    "数据状态": "原厂PDF规格表",
                                    "校验时间": CHECKED_AT,
                                    "校验备注": "；".join(notes),
                                    "_model_rule_authority": "joyin_official_pdf_template",
                                    "_resistance_ohm": r25_ohm,
                                }
                            )
    return rows


def exact_jfr_rows(pdf_path: Path, source_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    with fitz.open(pdf_path) as doc:
        text = "\n".join(page.get_text("text") for page in doc)
    models = sorted(set(re.findall(r"JFR\d{3}[A-Z]\d{3}[A-Z][AB]\d{5}CP[A-Z]", text)))
    by_core = {row["型号"]: row for row in source_rows}
    rows: list[dict[str, str]] = []
    for model in models:
        core = model[:12]
        if core not in by_core:
            continue
        row = dict(by_core[core])
        row["型号"] = model
        row["系列"] = "JFR"
        row["封装代码"] = "快速响应探头"
        row["备注3"] = (
            f"原厂尺寸表完整料号；电气核心 {core}；"
            f"完整型号后段为引线长度/结构配置，来源 {pdf_path.name}"
        )
        row["校验备注"] = row["备注3"]
        row["_model_rule_authority"] = "joyin_official_pdf_exact_model"
        rows.append(row)
    return rows


def jnr_rows(pdf_path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            for table in page.extract_tables() or []:
                for raw_row in table or []:
                    cells = [" ".join(clean_text(cell).split()) for cell in raw_row]
                    if not cells or not re.fullmatch(r"JNR\d{2}S[0-9RA]+□", cells[0]):
                        continue
                    if len(cells) < 9:
                        continue
                    template = cells[0]
                    r25 = numeric_text(cells[1])
                    if not r25:
                        continue
                    body = re.match(r"JNR(?P<body>\d{2})", template).group("body")
                    for tolerance_code, tolerance in (("L", "±15%"), ("M", "±20%")):
                        model = template.replace("□", tolerance_code)
                        rows.append(
                            {
                                "品牌": BRAND,
                                "型号": model,
                                "系列": "JNR",
                                "系列说明": "久尹 JNR 浪涌抑制功率 NTC 热敏电阻（ICL）",
                                "器件类型": "功率热敏电阻（NTC Thermistor）",
                                "安装方式": "插件",
                                "封装代码": f"{int(body)}D",
                                "尺寸（mm）": f"{int(body)}mm",
                                "直径（mm）": str(int(body)),
                                "阻值@25C": r25,
                                "阻值单位": "Ω",
                                "阻值误差": tolerance,
                                "工作温度": clean_text(cells[8]).replace("～", "~"),
                                "特殊用途": "浪涌抑制 | ICL | 电源上电保护 | 功率NTC",
                                "规格摘要": (
                                    f"JNR {int(body)}D {r25}Ω {tolerance} "
                                    f"Imax={clean_text(cells[2])}A"
                                ),
                                "备注1": (
                                    f"Imax {clean_text(cells[2])}A; "
                                    f"RImax {clean_text(cells[3])}Ω"
                                ),
                                "备注2": (
                                    f"AC240V最大负载电容 {clean_text(cells[4])}uF; "
                                    f"Pmax {clean_text(cells[5])}W"
                                ),
                                "备注3": (
                                    f"耗散系数 {clean_text(cells[6])}; "
                                    f"热时间常数 {clean_text(cells[7])}; PDF 第 {page_no} 页"
                                ),
                                "生产状态": "Active",
                                "官网链接": str(pdf_path),
                                "数据来源": f"{SOURCE_TAG}: {pdf_path.name}",
                                "数据状态": "原厂PDF规格表",
                                "校验时间": CHECKED_AT,
                                "校验备注": (
                                    f"原厂模板 {template}；□ 按命名规则展开为 "
                                    "L=±15%、M=±20%；引脚/包装配置按订单确认"
                                ),
                                "_model_rule_authority": "joyin_official_pdf_template",
                                "_resistance_ohm": r25,
                            }
                        )
    return rows


def parse_varistor_records(pdf_path: Path, prefix: str) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    with fitz.open(pdf_path) as doc:
        lines = [clean_text(line) for page in doc for line in page.get_text("text").splitlines()]
    model_re = re.compile(rf"^{prefix}\s*(\d{{2}}[NSU])\s*(\d{{3}}[KLMNP])$")
    index = 0
    while index < len(lines):
        match = model_re.fullmatch(lines[index])
        if not match:
            index += 1
            continue
        values: list[str] = []
        cursor = index + 1
        while cursor < len(lines) and len(values) < 10:
            if model_re.fullmatch(lines[cursor]):
                break
            token = lines[cursor].replace("±", "")
            if re.fullmatch(r"\d+(?:\.\d+)?%?", token):
                values.append(lines[cursor])
            cursor += 1
        if len(values) == 10:
            records.append(
                {
                    "model": f"{prefix}{match.group(1)}{match.group(2)}",
                    "varistor_voltage": values[0],
                    "tolerance": values[1],
                    "ac_voltage": values[2],
                    "dc_voltage": values[3],
                    "clamp_voltage": values[4],
                    "clamp_current": values[5],
                    "surge_current": values[6],
                    "nominal_discharge": values[7],
                    "rated_wattage": values[8],
                    "energy": values[9],
                }
            )
        index = max(index + 1, cursor)
    unique: dict[str, dict[str, str]] = {}
    for record in records:
        unique[record["model"]] = record
    return list(unique.values())


def varistor_rows(pdf_path: Path, prefix: str) -> list[dict[str, str]]:
    profile = VARISTOR_PROFILES[prefix]
    rows: list[dict[str, str]] = []
    for record in parse_varistor_records(pdf_path, prefix):
        model = record["model"]
        match = re.fullmatch(rf"{prefix}(?P<body>\d{{2}})(?P<family>[NSU])\d{{3}}[KLMNP]", model)
        body = str(int(match.group("body")))
        family_name, family_use = VARISTOR_FAMILY[match.group("family")]
        series = f"{prefix}-{match.group('family')}"
        automotive = "可选 CX 汽车级版本" if prefix == "JVT" else ""
        rows.append(
            {
                "品牌": BRAND,
                "型号": model,
                "系列": series,
                "系列说明": f"{profile['description']}（{family_name}）",
                "器件类型": "引线型压敏电阻（MOV）",
                "安装方式": "插件",
                "封装代码": f"{body}D",
                "尺寸（mm）": f"{body}mm",
                "直径（mm）": body,
                "耐压（V）": record["varistor_voltage"],
                "容值误差": record["tolerance"],
                "特殊用途": f"{family_use} | 径向引线 | MOV" + (
                    " | 可选汽车级" if automotive else ""
                ),
                "规格摘要": (
                    f"{series} {body}D V1mA={record['varistor_voltage']}V "
                    f"{record['tolerance']} AC={record['ac_voltage']}V "
                    f"DC={record['dc_voltage']}V"
                ),
                "备注1": (
                    f"工作电压 AC {record['ac_voltage']}V / "
                    f"DC {record['dc_voltage']}V"
                ),
                "备注2": (
                    f"钳位电压 {record['clamp_voltage']}V @ "
                    f"{record['clamp_current']}A"
                ),
                "备注3": (
                    f"单次浪涌 {record['surge_current']}A; "
                    f"标称放电 {record['nominal_discharge']}kA; "
                    f"额定功率 {record['rated_wattage']}W; "
                    f"能量 {record['energy']}J"
                    + (f"; {automotive}" if automotive else "")
                ),
                "生产状态": "Active",
                "官网链接": str(pdf_path),
                "数据来源": f"{SOURCE_TAG}: {', '.join(profile['source_files'])}",
                "数据状态": "原厂PDF规格表",
                "校验时间": CHECKED_AT,
                "校验备注": (
                    "原厂电气特性表基础料号；引脚形式及包装后缀按订单确认"
                    + (f"；{automotive}" if automotive else "")
                ),
                "_model_rule_authority": "joyin_official_pdf_exact_model",
            }
        )
    return rows


def build_rows() -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for series, filename in {
        "JAS": "JAS.pdf",
        "JAT": "JAT.pdf",
        "JCR03": "JCR03.pdf",
        "JCR05": "JCR05.pdf",
        "JFR": "JFR.pdf",
        "JSR": "JSR.pdf",
    }.items():
        pdf_path = PDF_DIR / filename
        series_rows = ntc_table_rows(pdf_path, series)
        rows.extend(series_rows)
        if series == "JFR":
            rows.extend(exact_jfr_rows(pdf_path, series_rows))
    rows.extend(jnr_rows(PDF_DIR / "JNR.pdf"))
    rows.extend(varistor_rows(PDF_DIR / "JVT.pdf", "JVT"))
    rows.extend(varistor_rows(PDF_DIR / "JVZ.pdf", "JVZ"))
    frame = pd.DataFrame(rows)
    return frame.drop_duplicates(
        subset=["品牌", "型号", "器件类型"], keep="first"
    ).reset_index(drop=True)


def database_columns(conn: sqlite3.Connection) -> list[str]:
    return [row[1] for row in conn.execute('PRAGMA table_info("components")')]


def backup_database() -> Path:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = DB_PATH.with_name(f"{DB_PATH.name}.joyin_pdf_series_{timestamp}.bak")
    shutil.copy2(DB_PATH, backup)
    return backup


def apply_database(frame: pd.DataFrame, dry_run: bool, no_backup: bool) -> Path | None:
    if dry_run:
        return None
    backup = None if no_backup else backup_database()
    with sqlite3.connect(DB_PATH, timeout=120) as conn:
        conn.execute("PRAGMA busy_timeout = 120000")
        columns = database_columns(conn)
        conn.execute("DELETE FROM components WHERE [数据来源] LIKE ?", (f"{SOURCE_TAG}:%",))
        db_rows: list[dict[str, str]] = []
        for row in frame.to_dict("records"):
            merged = {column: "" for column in columns}
            for column, value in row.items():
                if column in merged:
                    merged[column] = clean_text(value)
            db_rows.append(merged)
        pd.DataFrame(db_rows, columns=columns).to_sql(
            "components", conn, if_exists="append", index=False, chunksize=20, method="multi"
        )
        conn.commit()
    return backup


def workbook_value(row: dict[str, str], column: str) -> str:
    if column == "压敏电压":
        return clean_text(row.get("耐压（V）"))
    if column == "工作电压AC":
        match = re.search(r"AC ([\d.]+)V", clean_text(row.get("备注1")))
        return match.group(1) if match else ""
    if column == "工作电压DC":
        match = re.search(r"DC ([\d.]+)V", clean_text(row.get("备注1")))
        return match.group(1) if match else ""
    if column == "钳位电压":
        match = re.search(r"钳位电压 ([\d.]+)V", clean_text(row.get("备注2")))
        return match.group(1) if match else ""
    if column == "峰值电流":
        match = re.search(r"单次浪涌 ([\d.]+)A", clean_text(row.get("备注3")))
        return match.group(1) if match else ""
    if column == "能量等级":
        match = re.search(r"能量 ([\d.]+)J", clean_text(row.get("备注3")))
        return match.group(1) if match else ""
    if column == "厚度（mm）":
        return clean_text(row.get("高度（mm）"))
    return clean_text(row.get(column))


def update_workbook(path: Path, frame: pd.DataFrame, dry_run: bool) -> int:
    workbook = load_workbook(path)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        headers = [clean_text(cell.value) for cell in sheet[1]]
        source_index = headers.index("数据来源") + 1
        for row_index in range(sheet.max_row, 1, -1):
            if clean_text(sheet.cell(row_index, source_index).value).startswith(SOURCE_TAG):
                sheet.delete_rows(row_index)
        for row in frame.to_dict("records"):
            sheet.append([workbook_value(row, header) for header in headers])
        if not dry_run:
            workbook.save(path)
        return len(frame)
    finally:
        workbook.close()


def update_workbooks(frame: pd.DataFrame, dry_run: bool) -> dict[str, int]:
    ntc = frame[frame["器件类型"].str.contains("NTC", na=False)]
    vdr = frame[frame["器件类型"].str.contains("压敏", na=False)]
    return {
        NTC_WORKBOOK.name: update_workbook(NTC_WORKBOOK, ntc, dry_run),
        VDR_WORKBOOK.name: update_workbook(VDR_WORKBOOK, vdr, dry_run),
    }


def refresh_cache(frame: pd.DataFrame, skip_cache: bool) -> dict[str, int]:
    if skip_cache:
        return {}
    models = frame["型号"].drop_duplicates().tolist()
    selected_parts: list[pd.DataFrame] = []
    with sqlite3.connect(DB_PATH, timeout=120) as conn:
        for offset in range(0, len(models), 800):
            chunk = models[offset : offset + 800]
            placeholders = ",".join("?" for _ in chunk)
            selected_parts.append(
                pd.read_sql_query(
                    f"SELECT * FROM components WHERE [型号] IN ({placeholders})",
                    conn,
                    params=chunk,
                )
            )
    selected = pd.concat(selected_parts, ignore_index=True)
    prepared = cm.prepare_search_dataframe(cm.deduplicate_component_rows(selected))
    removed, inserted = stream_replace_prepared_rows(prepared)
    sidecar = refresh_search_sidecar_rows(prepared)
    return {
        "selected_db_rows": len(selected),
        "prepared_removed": int(removed),
        "prepared_inserted": int(inserted),
        "search_core_rows": int(sidecar.get(cm.COMPONENTS_SEARCH_CORE_TABLE, 0)),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Sync JOYIN leaded NTC, power NTC, JVT and JVZ series from official PDFs."
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-db", action="store_true")
    parser.add_argument("--skip-workbooks", action="store_true")
    parser.add_argument("--skip-cache", action="store_true")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()

    frame = build_rows()
    print(f"generated_rows={len(frame)}")
    print(frame.groupby(["系列", "器件类型"]).size().to_string())

    backup = None
    if not args.skip_db:
        backup = apply_database(frame, args.dry_run, args.no_backup)
    if backup:
        print(f"backup_path={backup}")

    if not args.skip_workbooks:
        for name, count in update_workbooks(frame, args.dry_run).items():
            print(f"workbook_rows[{name}]={count}")

    if not args.dry_run and not args.skip_db:
        for key, value in refresh_cache(frame, args.skip_cache).items():
            print(f"{key}={value}")


if __name__ == "__main__":
    main()
