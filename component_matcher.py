import streamlit as st
import pandas as pd
import sqlite3
import os
import json
import glob
import hashlib
import hmac
import math
import logging
import threading
import time
import base64
import html
import sys
import streamlit.components.v1 as components
import re
import unicodedata
import zipfile
import shutil
from io import BytesIO
import urllib.parse
import urllib.request
import ssl
import warnings
import traceback
from copy import copy
try:
    import pyarrow as pa
    import pyarrow.parquet as pq
except Exception:
    pa = None
    pq = None
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from mlcc_excel_importer import map_headers as importer_map_headers, ensure_standard_columns as importer_ensure_standard_columns, STANDARD_COLUMNS as IMPORTER_STANDARD_COLUMNS
from resistor_series_rules import build_resistor_series_description, infer_resistor_series_profile


def quiet_nonessential_console_noise():
    warnings.filterwarnings("ignore", category=DeprecationWarning, module=r"pandas(\.|$)")
    warnings.filterwarnings("ignore", category=FutureWarning, module=r"pandas(\.|$)")
    warnings.filterwarnings(
        "ignore",
        message=r".*DataFrame concatenation with empty or all-NA entries.*",
        category=FutureWarning,
    )
    logging.getLogger("streamlit").setLevel(logging.ERROR)
    logging.getLogger("streamlit.runtime").setLevel(logging.ERROR)
    logging.getLogger("streamlit.elements").setLevel(logging.ERROR)


quiet_nonessential_console_noise()

BOM_MATCH_DEBUG = str(os.getenv("BOM_MATCH_DEBUG", "")).strip().lower() in {"1", "true", "yes", "on"}
COMPONENT_MATCHER_BUILD_MODE_ENV = "COMPONENT_MATCHER_BUILD_MODE"


def is_component_matcher_build_mode():
    return str(os.getenv(COMPONENT_MATCHER_BUILD_MODE_ENV, "")).strip().lower() in {"1", "true", "yes", "on"}


def bom_match_debug_log(*parts):
    if not BOM_MATCH_DEBUG:
        return
    print("[BOM-DEBUG]", *parts, flush=True)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "components.db")
DATA_FOLDER = BASE_DIR
MASTER_XLSX_PATH = os.path.join(BASE_DIR, "Capacitor", "MLCC.xlsx")
RESISTOR_LIBRARY_CACHE_PATH = os.path.join(BASE_DIR, "cache", "resistor_library_cache.csv")
PREPARED_CACHE_PATH = os.path.join(BASE_DIR, "cache", "components_prepared_v5.parquet")
PREPARED_CACHE_FALLBACK_PATH = os.path.join(BASE_DIR, "cache", "components_prepared_v5.pkl")
SAMSUNG_MLCC_STATUS_CACHE_PATH = os.path.join(BASE_DIR, "cache", "samsung_all_statuses_base.json")
SAMSUNG_MLCC_PACKAGE_CACHE_PATH = os.path.join(BASE_DIR, "cache", "samsung_package_cache.json")
MLCC_LCSC_DIMENSION_CACHE_PATH = os.path.join(BASE_DIR, "cache", "mlcc_lcsc_dimension_cache.json")
MLCC_OFFICIAL_DIMENSION_CACHE_PATH = os.path.join(BASE_DIR, "cache", "mlcc_official_dimension_cache.json")
PREPARED_CACHE_META_PATH = os.path.join(BASE_DIR, "cache", "components_prepared_v5_meta.json")
SOURCE_NORMALIZED_CACHE_DIR = os.path.join(BASE_DIR, "cache", "source_normalized")
SEARCH_DB_PATH = os.path.join(BASE_DIR, "cache", "components_search.sqlite")
STREAMLIT_CLOUD_BUNDLE_PATH = os.path.join(BASE_DIR, "streamlit_cloud_bundle.zip")
STREAMLIT_CLOUD_BUNDLE_PART_GLOB = os.path.join(BASE_DIR, "streamlit_cloud_bundle.zip.part*")
STREAMLIT_CLOUD_BUNDLE_MANIFEST_PATH = os.path.join(BASE_DIR, "streamlit_cloud_bundle.manifest.json")
STREAMLIT_CLOUD_BUNDLE_STATE_PATH = os.path.join(BASE_DIR, "cache", "streamlit_cloud_bundle_state.json")
MANUAL_CORRECTION_RULES_PATH = os.path.join(BASE_DIR, "cache", "manual_correction_rules.csv")
APP_ACCESS_CODE_ENV = "APP_ACCESS_CODE"
COMPONENT_MATCHER_PUBLIC_MODE_ENV = "COMPONENT_MATCHER_PUBLIC_MODE"
COMPONENTS_SEARCH_LEGACY_TABLE = "components_search"
SEARCH_META_TABLE = "search_meta"
COMPONENTS_SEARCH_CHUNK_ROWS = 50000
PREPARED_CACHE_VERSION = 7
SOURCE_NORMALIZED_CACHE_VERSION = 8
SEARCH_INDEX_SCHEMA_VERSION = 5
QUERY_RESULT_CACHE_VERSION = 12
MANUAL_CORRECTION_RULES_VERSION = 1
SEARCH_DB_FETCH_CHUNK = 300
LOGO_PATH = os.path.join(BASE_DIR, "logo.png")
REGRESSION_CASES_PATH = os.path.join(BASE_DIR, "regression_cases.csv")
PAGE_ICON = LOGO_PATH if os.path.exists(LOGO_PATH) else "📦"
PUBLIC_SEARCH_MAX_LINES = 50
PRIVATE_SEARCH_MAX_LINES = 200
PUBLIC_SEARCH_MAX_TOTAL_CHARS = 8000
PRIVATE_SEARCH_MAX_TOTAL_CHARS = 30000
PUBLIC_SEARCH_MAX_LINE_CHARS = 200
PRIVATE_SEARCH_MAX_LINE_CHARS = 400
PUBLIC_BOM_MAX_ROWS = 10000
PRIVATE_BOM_MAX_ROWS = 50000
PUBLIC_BOM_MAX_SHEETS = 10
PRIVATE_BOM_MAX_SHEETS = 25
STREAMLIT_CLOUD_BUNDLE_LOCK = threading.Lock()
CLOUD_SEARCH_ASSET_WARMUP_LOCK = threading.Lock()
CLOUD_SEARCH_ASSET_WARMUP_STARTED = False
STARTUP_TRACE_ENABLED = str(os.getenv("COMPONENT_MATCHER_STARTUP_TRACE", "")).strip().lower() in {"1", "true", "yes", "on"}
STARTUP_TRACE_PATH = os.path.join(BASE_DIR, "cache", "startup_trace.log")
# Public code nudge:
# This marker also participates in public query cache keys so stale session
# search results are invalidated when we ship a new public build or adjust
# matching/ranking behavior.
PUBLIC_CODE_STAMP = "2026-04-23T04:42:57+08:00"


def startup_trace(message):
    if not STARTUP_TRACE_ENABLED:
        return
    try:
        os.makedirs(os.path.dirname(STARTUP_TRACE_PATH), exist_ok=True)
        with open(STARTUP_TRACE_PATH, "a", encoding="utf-8") as handle:
            handle.write(f"{time.time():.6f} {message}\n")
    except Exception:
        pass


def bundle_member_path_for_local_path(path):
    rel_path = os.path.relpath(os.path.abspath(path), BASE_DIR)
    return rel_path.replace("\\", "/")


def _sqlite_table_row_count(path, table_name):
    if not os.path.exists(path):
        return 0
    try:
        conn = sqlite3.connect(path)
        try:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                (table_name,),
            )
            if cur.fetchone() is None:
                return 0
            cur.execute(f"SELECT COUNT(*) FROM {table_name}")
            return int(cur.fetchone()[0] or 0)
        finally:
            conn.close()
    except Exception:
        return 0


def bundle_target_is_valid(path):
    try:
        abs_path = os.path.abspath(path)
        if abs_path == os.path.abspath(DB_PATH):
            return database_has_component_rows()
        if abs_path == os.path.abspath(SEARCH_DB_PATH):
            return _sqlite_table_row_count(abs_path, COMPONENTS_SEARCH_CORE_TABLE) > 0
        if abs_path == os.path.abspath(PREPARED_CACHE_PATH):
            if pq is None:
                return os.path.exists(abs_path) and os.path.getsize(abs_path) > 0
            if not os.path.exists(abs_path) or os.path.getsize(abs_path) <= 0:
                return False
            try:
                parquet_file = pq.ParquetFile(abs_path)
                return int(parquet_file.metadata.num_rows or 0) > 0
            except Exception:
                return False
        return os.path.exists(path) and os.path.getsize(path) > 0
    except Exception:
        return False


def get_path_signature(path):
    abs_path = os.path.abspath(path)
    try:
        if not os.path.exists(abs_path):
            return {
                "path": abs_path,
                "exists": False,
            }
        stat = os.stat(abs_path)
        return {
            "path": abs_path,
            "exists": True,
            "mtime": round(stat.st_mtime, 6),
            "size": stat.st_size,
        }
    except Exception:
        return {
            "path": abs_path,
            "exists": False,
        }


def load_streamlit_cloud_bundle_manifest():
    if not os.path.exists(STREAMLIT_CLOUD_BUNDLE_MANIFEST_PATH):
        return {}
    try:
        with open(STREAMLIT_CLOUD_BUNDLE_MANIFEST_PATH, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except Exception:
        return {}
    members = data.get("members", []) if isinstance(data, dict) else []
    manifest = {}
    for entry in members:
        if not isinstance(entry, dict):
            continue
        member_path = clean_text(entry.get("path", ""))
        if member_path == "":
            continue
        manifest[member_path] = {
            "size": int(entry.get("size", 0) or 0),
            "mtime_ns": int(entry.get("mtime_ns", 0) or 0),
            "sha256": clean_text(entry.get("sha256", "")),
        }
    return manifest


def get_streamlit_cloud_bundle_signature():
    try:
        if os.path.exists(STREAMLIT_CLOUD_BUNDLE_MANIFEST_PATH):
            with open(STREAMLIT_CLOUD_BUNDLE_MANIFEST_PATH, "rb") as handle:
                return hashlib.sha256(handle.read()).hexdigest()
    except Exception:
        pass
    try:
        if os.path.exists(STREAMLIT_CLOUD_BUNDLE_PATH):
            stat = os.stat(STREAMLIT_CLOUD_BUNDLE_PATH)
            return f"{int(stat.st_mtime_ns)}:{int(stat.st_size)}"
    except Exception:
        pass
    part_paths = get_streamlit_cloud_bundle_part_paths()
    if part_paths:
        digest = hashlib.sha256()
        for part_path in part_paths:
            try:
                stat = os.stat(part_path)
            except Exception:
                continue
            digest.update(os.path.basename(part_path).encode("utf-8", "ignore"))
            digest.update(str(int(stat.st_mtime_ns)).encode("utf-8"))
            digest.update(str(int(stat.st_size)).encode("utf-8"))
        return f"parts:{digest.hexdigest()}"
    return ""


def get_streamlit_cloud_bundle_part_paths():
    part_paths = [
        path
        for path in glob.glob(STREAMLIT_CLOUD_BUNDLE_PART_GLOB)
        if os.path.isfile(path)
    ]

    def sort_key(path):
        base = os.path.basename(path)
        match = re.search(r"\.part(\d+)$", base)
        if match:
            return (0, int(match.group(1)), base)
        match = re.search(r"part[-_]?(\d+)", base)
        if match:
            return (0, int(match.group(1)), base)
        return (1, base)

    return sorted(part_paths, key=sort_key)


def replace_file_with_retry(source_path, target_path, attempts=8, initial_delay_sec=0.2):
    last_error = None
    for attempt in range(1, max(1, int(attempts)) + 1):
        try:
            os.replace(source_path, target_path)
            return True
        except OSError as exc:
            last_error = exc
            if attempt >= attempts:
                raise
            time.sleep(initial_delay_sec * attempt)
    if last_error is not None:
        raise last_error
    return False


def ensure_streamlit_cloud_bundle_archive_available():
    part_paths = get_streamlit_cloud_bundle_part_paths()
    if os.path.exists(STREAMLIT_CLOUD_BUNDLE_PATH):
        if not part_paths:
            return True
        current_signature = get_streamlit_cloud_bundle_signature()
        state = load_streamlit_cloud_bundle_state()
        state_signature = clean_text(state.get("bundle_signature", ""))
        if current_signature != "" and current_signature == state_signature:
            return True
        # If the deployment already has a bundle zip but the recorded bundle
        # version has changed, rebuild the archive from the freshly deployed
        # parts so the runtime cannot keep serving stale search data.
        try:
            os.remove(STREAMLIT_CLOUD_BUNDLE_PATH)
        except Exception:
            pass
    if not part_paths:
        return False

    temp_target_path = STREAMLIT_CLOUD_BUNDLE_PATH + ".part"
    if os.path.exists(temp_target_path):
        try:
            os.remove(temp_target_path)
        except Exception:
            pass

    try:
        with open(temp_target_path, "wb") as target_handle:
            for part_path in part_paths:
                with open(part_path, "rb") as source_handle:
                    shutil.copyfileobj(source_handle, target_handle, length=8 * 1024 * 1024)
        replace_file_with_retry(temp_target_path, STREAMLIT_CLOUD_BUNDLE_PATH)
        return os.path.exists(STREAMLIT_CLOUD_BUNDLE_PATH)
    except Exception as exc:
        logging.getLogger(__name__).warning("failed to rebuild cloud bundle archive from parts: %s", exc)
        return False
    finally:
        if os.path.exists(temp_target_path):
            try:
                os.remove(temp_target_path)
            except Exception:
                pass


def load_streamlit_cloud_bundle_state():
    if not os.path.exists(STREAMLIT_CLOUD_BUNDLE_STATE_PATH):
        return {}
    try:
        with open(STREAMLIT_CLOUD_BUNDLE_STATE_PATH, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_streamlit_cloud_bundle_state(signature, extracted_members=None):
    state = {
        "bundle_signature": clean_text(signature),
        "updated_at": time.time(),
        "members": sorted(
            {
                clean_text(member)
                for member in (extracted_members or [])
                if clean_text(member) != ""
            }
        ),
    }
    try:
        os.makedirs(os.path.dirname(STREAMLIT_CLOUD_BUNDLE_STATE_PATH), exist_ok=True)
        with open(STREAMLIT_CLOUD_BUNDLE_STATE_PATH, "w", encoding="utf-8") as handle:
            json.dump(state, handle, ensure_ascii=False, indent=2)
    except Exception:
        pass


def streamlit_cloud_bundle_refresh_needed(required_paths=None):
    if not (is_public_mode() or is_streamlit_cloud_runtime()):
        return False
    if not ensure_streamlit_cloud_bundle_archive_available():
        return False

    manifest = load_streamlit_cloud_bundle_manifest()
    signature = get_streamlit_cloud_bundle_signature()
    state = load_streamlit_cloud_bundle_state()
    state_signature = clean_text(state.get("bundle_signature", ""))
    target_paths = [
        os.path.abspath(path)
        for path in (
            required_paths
            or [DB_PATH, SEARCH_DB_PATH, PREPARED_CACHE_PATH]
        )
    ]
    for target_path in target_paths:
        member_name = bundle_member_path_for_local_path(target_path)
        expected = manifest.get(member_name)
        if expected:
            try:
                stat = os.stat(target_path)
                if int(stat.st_size) != int(expected.get("size", 0) or 0):
                    return True
            except Exception:
                return True
        elif not bundle_target_is_valid(target_path):
            return True
    return signature != "" and signature != state_signature


def ensure_streamlit_cloud_data_bundle(required_paths=None):
    target_paths = [
        os.path.abspath(path)
        for path in (
            required_paths
            or [DB_PATH, SEARCH_DB_PATH, PREPARED_CACHE_PATH]
        )
    ]
    if not ensure_streamlit_cloud_bundle_archive_available():
        return False
    refresh_needed = streamlit_cloud_bundle_refresh_needed(required_paths=target_paths)
    if not refresh_needed and all(bundle_target_is_valid(path) for path in target_paths):
        return True

    try:
        with STREAMLIT_CLOUD_BUNDLE_LOCK:
            extracted_members = []
            with zipfile.ZipFile(STREAMLIT_CLOUD_BUNDLE_PATH, "r") as archive:
                available_members = set(archive.namelist())
                for target_path in target_paths:
                    if not refresh_needed and bundle_target_is_valid(target_path):
                        continue
                    member_name = bundle_member_path_for_local_path(target_path)
                    if member_name not in available_members:
                        logging.getLogger(__name__).warning(
                            "cloud bundle missing member: %s",
                            member_name,
                        )
                        continue
                    target_dir = os.path.dirname(target_path)
                    if target_dir:
                        os.makedirs(target_dir, exist_ok=True)
                    temp_target_path = target_path + ".part"
                    if os.path.exists(temp_target_path):
                        try:
                            os.remove(temp_target_path)
                        except Exception:
                            pass
                    try:
                        with archive.open(member_name, "r") as source_handle, open(temp_target_path, "wb") as target_handle:
                            shutil.copyfileobj(source_handle, target_handle, length=1024 * 1024)
                        replace_file_with_retry(temp_target_path, target_path)
                        extracted_members.append(member_name)
                    finally:
                        if os.path.exists(temp_target_path):
                            try:
                                os.remove(temp_target_path)
                            except Exception:
                                pass
            save_streamlit_cloud_bundle_state(
                get_streamlit_cloud_bundle_signature(),
                extracted_members=extracted_members,
            )
    except Exception as exc:
        logging.getLogger(__name__).warning("failed to extract cloud bundle: %s", exc)
        return False

    return all(bundle_target_is_valid(path) for path in target_paths)


def search_sidecar_assets_available():
    return all(bundle_target_is_valid(path) for path in get_search_asset_bundle_paths())


def get_search_asset_bundle_paths():
    return [
        SEARCH_DB_PATH,
        MLCC_LCSC_DIMENSION_CACHE_PATH,
        os.path.join(BASE_DIR, "cache", "pdc_findchips_cache.json"),
    ]


def get_public_runtime_bundle_paths():
    return get_search_asset_bundle_paths()


def maybe_start_cloud_search_asset_warmup():
    global CLOUD_SEARCH_ASSET_WARMUP_STARTED
    startup_trace("cloud_warmup:enter")
    runtime_bundle_paths = get_public_runtime_bundle_paths()
    if not streamlit_cloud_bundle_refresh_needed(required_paths=runtime_bundle_paths) and (
        search_sidecar_assets_available() or prepared_cache_is_current()
    ):
        startup_trace("cloud_warmup:skip-ready")
        return
    if not ensure_streamlit_cloud_bundle_archive_available():
        startup_trace("cloud_warmup:no-bundle-archive")
        return
    with CLOUD_SEARCH_ASSET_WARMUP_LOCK:
        if CLOUD_SEARCH_ASSET_WARMUP_STARTED:
            startup_trace("cloud_warmup:already-started")
            return
        CLOUD_SEARCH_ASSET_WARMUP_STARTED = True
        startup_trace("cloud_warmup:started")

    def _warm():
        try:
            startup_trace("cloud_warmup:thread-start")
            ensure_streamlit_cloud_data_bundle(required_paths=get_search_asset_bundle_paths())
            startup_trace("cloud_warmup:thread-done")
        except Exception:
            startup_trace("cloud_warmup:thread-error")
            pass

    threading.Thread(target=_warm, daemon=True).start()
    startup_trace("cloud_warmup:thread-launched")


def ensure_component_data_ready(action_label=""):
    public_runtime = is_public_mode() or is_streamlit_cloud_runtime()
    full_bundle_paths = [DB_PATH, SEARCH_DB_PATH, PREPARED_CACHE_PATH]
    runtime_bundle_paths = get_public_runtime_bundle_paths() if public_runtime else full_bundle_paths
    bundle_refresh_needed = False
    if public_runtime:
        try:
            bundle_refresh_needed = streamlit_cloud_bundle_refresh_needed(required_paths=runtime_bundle_paths)
        except Exception:
            bundle_refresh_needed = False
    if public_runtime and bundle_refresh_needed and ensure_streamlit_cloud_data_bundle(required_paths=runtime_bundle_paths):
        clear_data_load_caches()
        return database_has_component_rows() or search_sidecar_assets_available() or prepared_cache_is_current()
    if prepared_cache_is_current() and not bundle_refresh_needed:
        return True
    if not bundle_refresh_needed and (
        database_has_component_rows() or search_sidecar_assets_available()
    ):
        return True

    action_text = clean_text(action_label)
    search_asset_paths = get_search_asset_bundle_paths()
    if bundle_refresh_needed and ensure_streamlit_cloud_data_bundle(required_paths=runtime_bundle_paths):
        clear_data_load_caches()
        return database_has_component_rows() or search_sidecar_assets_available() or prepared_cache_is_current()
    if public_runtime:
        if bundle_refresh_needed:
            return False
        return search_sidecar_assets_available() or prepared_cache_is_current()
    if action_text in {"搜索", "BOM 匹配"} and ensure_streamlit_cloud_data_bundle(required_paths=search_asset_paths):
        clear_data_load_caches()
        return database_has_component_rows() or search_sidecar_assets_available() or prepared_cache_is_current()

    if ensure_streamlit_cloud_data_bundle():
        clear_data_load_caches()
        return database_has_component_rows() or search_sidecar_assets_available() or prepared_cache_is_current()

    try:
        maybe_update_database(force=False)
    except Exception:
        return False
    return database_has_component_rows() or search_sidecar_assets_available()


def is_public_mode():
    return str(os.getenv(COMPONENT_MATCHER_PUBLIC_MODE_ENV, "")).strip().lower() in {"1", "true", "yes", "on"}


def is_streamlit_cloud_runtime():
    for env_name in ["STREAMLIT_SHARING_MODE", "IS_STREAMLIT_CLOUD", "STREAMLIT_RUNTIME"]:
        if clean_text(os.getenv(env_name, "")) != "":
            return True
    normalized_base_dir = os.path.abspath(BASE_DIR).replace("\\", "/")
    return normalized_base_dir == "/mount/src" or normalized_base_dir.startswith("/mount/src/")


def get_int_env(name, default_value):
    raw_value = os.getenv(name, None)
    if raw_value is None:
        return int(default_value)
    try:
        return int(str(raw_value).strip())
    except Exception:
        return int(default_value)


def get_runtime_security_limits():
    public_mode = is_public_mode()
    return {
        "max_search_lines": get_int_env("COMPONENT_MATCHER_MAX_SEARCH_LINES", PUBLIC_SEARCH_MAX_LINES if public_mode else PRIVATE_SEARCH_MAX_LINES),
        "max_search_total_chars": get_int_env("COMPONENT_MATCHER_MAX_SEARCH_TOTAL_CHARS", PUBLIC_SEARCH_MAX_TOTAL_CHARS if public_mode else PRIVATE_SEARCH_MAX_TOTAL_CHARS),
        "max_search_line_chars": get_int_env("COMPONENT_MATCHER_MAX_SEARCH_LINE_CHARS", PUBLIC_SEARCH_MAX_LINE_CHARS if public_mode else PRIVATE_SEARCH_MAX_LINE_CHARS),
        "max_bom_rows": get_int_env("COMPONENT_MATCHER_MAX_BOM_ROWS", PUBLIC_BOM_MAX_ROWS if public_mode else PRIVATE_BOM_MAX_ROWS),
        "max_bom_sheets": get_int_env("COMPONENT_MATCHER_MAX_BOM_SHEETS", PUBLIC_BOM_MAX_SHEETS if public_mode else PRIVATE_BOM_MAX_SHEETS),
    }


def get_configured_access_code():
    secret_value = ""
    try:
        if hasattr(st, "secrets"):
            secret_value = str(st.secrets.get("app_access_code", "")).strip()
            if secret_value == "":
                secret_value = str(st.secrets.get("APP_ACCESS_CODE", "")).strip()
    except Exception:
        secret_value = ""
    if secret_value != "":
        return secret_value
    return str(os.getenv(APP_ACCESS_CODE_ENV, "")).strip()


def get_query_param_value(name):
    try:
        value = st.query_params.get(name, "")
        if isinstance(value, list):
            value = value[0] if value else ""
        return str(value).strip()
    except Exception:
        try:
            params = st.experimental_get_query_params()
            value = params.get(name, [""])
            if isinstance(value, list):
                value = value[0] if value else ""
            return str(value).strip()
        except Exception:
            return ""


def should_render_inline_footer():
    embed_value = get_query_param_value("embed").lower()
    return embed_value not in {"1", "true", "yes", "on"}


def require_app_access():
    startup_trace("require_app_access:enter")
    access_code = get_configured_access_code()
    if access_code == "":
        startup_trace("require_app_access:open")
        return True

    if st.session_state.get("_app_access_granted", False):
        startup_trace("require_app_access:granted")
        return True

    st.markdown(
        """
        <div style="max-width: 560px; margin: 40px auto 0 auto; padding: 22px 22px 18px 22px; border: 1px solid rgba(59,130,246,0.18); border-radius: 18px; background: rgba(255,255,255,0.92); box-shadow: 0 12px 30px rgba(15,23,42,0.08);">
            <div style="font-size: 22px; font-weight: 800; color: #0f172a; margin-bottom: 8px;">访问验证</div>
            <div style="font-size: 14px; line-height: 1.7; color: #475569; margin-bottom: 14px;">当前站点已开启访问码保护。请输入访问码后继续使用搜索和 BOM 功能。</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    with st.form("app_access_form", clear_on_submit=False):
        entered_code = st.text_input("访问码", type="password", label_visibility="collapsed", placeholder="请输入访问码")
        submitted = st.form_submit_button("进入系统")

    if submitted:
        if entered_code and hmac.compare_digest(entered_code.strip(), access_code):
            st.session_state["_app_access_granted"] = True
            startup_trace("require_app_access:submit-ok")
            st.rerun()
        st.error("访问码不正确")
        startup_trace("require_app_access:submit-fail")
    st.stop()


COMPONENTS_SEARCH_CORE_TABLE = "components_search_core"
COMPONENTS_SEARCH_RESISTOR_TABLE = "components_search_resistor"
COMPONENTS_SEARCH_CAPACITOR_TABLE = "components_search_capacitor"
COMPONENTS_SEARCH_VALUE_TABLE = "components_search_value"
COMPONENTS_SEARCH_VARISTOR_TABLE = "components_search_varistor"
COMPONENTS_SEARCH_TABLE = COMPONENTS_SEARCH_CORE_TABLE
MODEL_REVERSE_LOOKUP_COLUMNS = [
    "_model_clean", "品牌", "型号", "系列", "器件类型", "尺寸（inch）", "尺寸（mm）", "材质（介质）", "容值",
    "容值单位", "容值误差", "耐压（V）", "长度（mm）", "宽度（mm）", "高度（mm）", "工作温度", "寿命（h）", "规格摘要", "特殊用途", "备注1", "备注2", "备注3", "安装方式", "封装代码",
    "_model_rule_authority",
]
MODEL_REVERSE_LOOKUP_CACHE = {}
MLCC_REFERENCE_LOOKUP_CACHE = {}
SAMSUNG_MLCC_DIMENSION_LOOKUP = None
SAMSUNG_MLCC_DIMENSION_LOOKUP_SIGNATURE = None
MLCC_LCSC_DIMENSION_CACHE = None
MLCC_LCSC_DIMENSION_CACHE_SIGNATURE = None
MLCC_OFFICIAL_DIMENSION_CACHE = None
MLCC_OFFICIAL_DIMENSION_CACHE_SIGNATURE = None
MLCC_LCSC_DIMENSION_CACHE_LOCK = threading.Lock()
MLCC_OFFICIAL_DIMENSION_CACHE_LOCK = threading.Lock()
PREPARED_SEARCH_REQUIRED_COLUMNS = [
    "_model_clean", "_size", "_mat", "_tol", "_volt", "_pf",
    "_tol_kind", "_tol_num", "_volt_num", "_component_type",
    "_res_ohm", "_power", "_power_watt", "_body_size", "_pitch", "_safety_class",
    "_varistor_voltage", "_disc_size", "_temp_low", "_temp_high", "_life_hours_num",
    "_mount_style", "_special_use_norm", "_mlcc_series_class", "_model_rule_authority",
]

st.set_page_config(
    page_title="富临通元器件匹配系统",
    page_icon=PAGE_ICON,
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
#MainMenu {visibility: hidden;}
header {visibility: hidden;}
footer {visibility: hidden;}
[data-testid="stToolbar"] {display: none !important;}
[data-testid="stDecoration"] {display: none !important;}
[data-testid="stStatusWidget"] {display: none !important;}
[data-testid="stHeader"] {display: none !important;}

.block-container {
    max-width: 1180px;
    padding-top: 0.8rem;
    padding-bottom: 1.6rem;
}
.main-title {
    text-align: center;
    font-size: 30px;
    font-weight: 700;
    margin-top: 8px;
    margin-bottom: 18px;
}
.sub-title {
    text-align: center;
    font-size: 15px;
    margin-bottom: 4px;
}
.sub-title-2 {
    text-align: center;
    font-size: 13px;
    margin-bottom: 20px;
}
.result-title {
    font-size: 22px;
    font-weight: 700;
    margin-top: 14px;
    margin-bottom: 6px;
}
.section-title {
    font-size: 20px;
    font-weight: 700;
    margin-top: 10px;
    margin-bottom: 4px;
}
.tool-panel {
    margin-top: 8px;
    padding: 16px 18px;
    border-radius: 18px;
    border: 1px solid rgba(37, 99, 235, 0.14);
    background: linear-gradient(180deg, rgba(255,255,255,0.98) 0%, rgba(245,248,255,0.98) 100%);
    box-shadow: 0 12px 24px rgba(15, 23, 42, 0.06);
}
.tool-panel-title {
    font-size: 18px;
    font-weight: 800;
    margin-bottom: 6px;
    color: #1f2937;
}
.tool-panel-note {
    font-size: 13px;
    color: #6b7280;
    margin-bottom: 12px;
    line-height: 1.55;
}
.interp-chip-row {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin: 8px 0 12px 0;
}
.interp-chip {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 6px 10px;
    border-radius: 999px;
    background: #eef2ff;
    color: #1f2937;
    font-size: 12px;
    font-weight: 700;
}
.interp-chip strong {
    color: #111827;
}
.interp-card {
    margin-top: 10px;
    padding: 14px 14px 10px 14px;
    border-radius: 16px;
    border: 1px solid #e5e7eb;
    background: #fff;
}
.interp-summary {
    font-size: 14px;
    line-height: 1.65;
    color: #334155;
    margin-bottom: 12px;
}
.interp-rule-note {
    font-size: 13px;
    color: #7c2d12;
    background: #fff7ed;
    border: 1px solid #fed7aa;
    border-radius: 12px;
    padding: 10px 12px;
    margin-top: 8px;
}
.bom-progress-card {
    margin: 10px 0 18px 0;
    padding: 16px 18px 14px 18px;
    border-radius: 18px;
    border: 1px solid rgba(51, 102, 255, 0.16);
    background: linear-gradient(180deg, rgba(255, 255, 255, 0.96) 0%, rgba(247, 250, 255, 0.96) 100%);
    box-shadow: 0 12px 28px rgba(31, 41, 55, 0.08);
}
.bom-progress-head {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    gap: 14px;
}
.bom-progress-title {
    font-size: 18px;
    font-weight: 800;
    color: #1f2937;
    line-height: 1.2;
}
.bom-progress-subtitle {
    margin-top: 5px;
    font-size: 13px;
    color: #6b7280;
    line-height: 1.45;
}
.bom-progress-pill {
    flex: 0 0 auto;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    min-width: 68px;
    padding: 8px 12px;
    border-radius: 999px;
    font-size: 13px;
    font-weight: 700;
    color: #1d4ed8;
    background: rgba(219, 234, 254, 0.9);
}
.bom-progress-pill.is-done {
    color: #166534;
    background: rgba(220, 252, 231, 0.95);
}
.bom-progress-track {
    position: relative;
    margin-top: 14px;
    height: 12px;
    border-radius: 999px;
    background: linear-gradient(180deg, rgba(229, 231, 235, 0.95) 0%, rgba(243, 244, 246, 0.95) 100%);
    overflow: hidden;
}
.bom-progress-fill {
    height: 100%;
    border-radius: inherit;
    background: linear-gradient(90deg, #5b8cff 0%, #39c7ff 55%, #6ee7b7 100%);
    background-size: 200% 100%;
    box-shadow: 0 0 18px rgba(91, 140, 255, 0.28);
    transition: width 180ms ease-out;
    animation: bom-progress-shimmer 2.8s linear infinite;
}
.bom-progress-fill.is-done {
    background: linear-gradient(90deg, #22c55e 0%, #34d399 55%, #86efac 100%);
    background-size: 200% 100%;
    box-shadow: 0 0 18px rgba(34, 197, 94, 0.28);
    animation: none;
}
.bom-progress-meta {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-top: 12px;
}
.bom-progress-chip {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 7px 10px;
    border-radius: 999px;
    background: rgba(249, 250, 251, 0.95);
    border: 1px solid rgba(229, 231, 235, 0.95);
    color: #374151;
    font-size: 12px;
    line-height: 1;
}
.bom-progress-chip strong {
    font-weight: 800;
}
.bom-progress-chip.success {
    background: rgba(236, 253, 245, 0.98);
    border-color: rgba(167, 243, 208, 0.95);
    color: #047857;
}
.bom-progress-chip.warn {
    background: rgba(255, 251, 235, 0.98);
    border-color: rgba(253, 230, 138, 0.95);
    color: #b45309;
}
.bom-progress-chip.fail {
    background: rgba(254, 242, 242, 0.98);
    border-color: rgba(254, 202, 202, 0.95);
    color: #b91c1c;
}
.bom-progress-current {
    margin-top: 12px;
    padding: 10px 12px;
    border-radius: 14px;
    background: rgba(248, 250, 252, 0.95);
    color: #334155;
    font-size: 12px;
    line-height: 1.45;
    border: 1px solid rgba(226, 232, 240, 0.95);
    word-break: break-word;
}
.bom-progress-current strong {
    color: #0f172a;
}
.bom-progress-summary {
    margin-top: 12px;
    padding: 12px 14px;
    border-radius: 14px;
    background: rgba(249, 250, 251, 0.95);
    border: 1px solid rgba(226, 232, 240, 0.95);
    color: #4b5563;
    font-size: 13px;
    line-height: 1.7;
    white-space: normal;
}
.bom-progress-summary-line + .bom-progress-summary-line {
    margin-top: 8px;
}
.bom-progress-summary strong {
    color: #111827;
}
.result-table-wrap {
    overflow: auto;
    max-height: min(560px, 52vh);
    margin-bottom: 0;
    position: relative;
}
.bom-result-table-wrap {
    max-height: min(560px, 52vh);
    overflow: auto;
    margin-bottom: 10px;
    position: relative;
}
.result-section-card {
    display: flex;
    flex-direction: column;
    gap: 10px;
    padding: 8px 8px 10px 8px;
    border: 1px solid rgba(191, 219, 254, 0.90);
    border-radius: 18px;
    background: linear-gradient(180deg, #ffffff 0%, #fbfdff 100%);
    box-shadow: 0 10px 24px rgba(15, 23, 42, 0.06);
    box-sizing: border-box;
}
.result-section-card .result-table-wrap,
.result-section-card .bom-result-table-wrap {
    max-height: min(560px, 52vh);
    margin-bottom: 0;
}
.bom-download-footer-outside {
    display: flex;
    justify-content: flex-end;
    align-items: center;
    gap: 10px;
    margin-top: 0;
    padding: 0 4px 0 4px;
}
.bom-download-btn {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    min-width: 190px;
    padding: 12px 18px;
    border-radius: 12px;
    border: 1px solid rgba(203, 213, 225, 0.98);
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
    color: #1f2937;
    font-size: 16px;
    font-weight: 700;
    text-decoration: none;
    box-shadow: 0 8px 18px rgba(31, 41, 55, 0.08);
}
.bom-download-btn:hover {
    border-color: rgba(148, 163, 184, 0.95);
    box-shadow: 0 12px 24px rgba(31, 41, 55, 0.12);
    transform: translateY(-1px);
    text-decoration: none;
}
.bom-download-btn:active {
    transform: translateY(0);
    box-shadow: 0 6px 12px rgba(31, 41, 55, 0.08);
}
.bom-download-hint {
    color: #64748b;
    font-size: 13px;
    line-height: 1.4;
}
.bom-preview-toggle-anchor {
    height: 0;
    padding: 0;
    line-height: 0;
}
.element-container:has(.bom-preview-toggle-anchor) {
    margin-top: -12px !important;
    margin-bottom: 8px !important;
}
.bom-preview-toggle-footer {
    margin-top: 0 !important;
    padding-top: 0 !important;
    padding-bottom: 0 !important;
}
.match-card-head {
    display: flex;
    align-items: center;
    flex-wrap: wrap;
    gap: 10px;
    margin-top: 0;
}
.match-card-head-spaced {
    margin-top: 4px;
}
.match-card-footer {
    height: 12px;
    margin-top: 4px;
    border-top: 1px solid rgba(191, 219, 254, 0.42);
    border-radius: 0 0 14px 14px;
    background: linear-gradient(180deg, rgba(248, 250, 255, 0.18) 0%, rgba(248, 250, 255, 0.72) 100%);
}
.app-footer-shell {
    margin-top: 20px;
    margin-bottom: 4px;
    display: block;
}
.app-footer {
    width: 100%;
    margin: 0;
    padding: 14px 0 10px;
    border: 0;
    background: transparent;
    box-shadow: none;
    text-align: center;
    color: #5f6b7a;
    font-size: 13px;
    line-height: 1.4;
    box-sizing: border-box;
    white-space: normal;
}
.match-card-query-pill {
    display: inline-flex;
    align-items: center;
    padding: 4px 10px;
    border-radius: 999px;
    background: rgba(59, 130, 246, 0.10);
    border: 1px solid rgba(59, 130, 246, 0.18);
    color: #1d4ed8;
    font-size: 14px;
    font-weight: 700;
    line-height: 1.2;
    word-break: break-all;
}
.result-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 14px;
    background: #ffffff;
}
.result-table th,
.result-table td {
    border: 1px solid #e6e6e6;
    padding: 8px 10px;
    text-align: center;
    vertical-align: middle;
    white-space: nowrap;
}
.result-table th {
    background: #f7f8fa;
    font-weight: 600;
    position: sticky;
    top: 0;
    z-index: 6;
    box-shadow: 0 1px 0 #e6e6e6;
}
.result-table tr.exact-match td {
    background: #fff59d;
    color: #111111;
}
.result-table tr.partial-match-row td {
    background: #fde2e1;
    color: #7f1d1d;
}
.result-table tr.substitute-row td {
    background: #dbeafe;
    color: #1d4ed8;
}
.result-table td.param-hit {
    color: #c62828;
    font-weight: 600;
}
.result-table td.model-list-cell,
.result-table th.model-list-cell,
.result-table td.detail-cell,
.result-table th.detail-cell {
    text-align: left;
    vertical-align: middle;
}
.result-table td.model-list-cell {
    white-space: nowrap;
}
.result-table td.model-list-cell summary {
    cursor: pointer;
    color: #1565c0;
    font-weight: 600;
    outline: none;
    white-space: nowrap;
}
.result-table td.model-list-cell summary::-webkit-details-marker {
    display: none;
}
.result-table td.model-list-cell .multi-model-cell {
    max-height: 96px;
    overflow-y: auto;
    overflow-x: auto;
    white-space: nowrap;
    word-break: normal;
    overflow-wrap: normal;
    line-height: 1.35;
}
.result-table tr.parse-fail-row td {
    background: #fde2e1;
}
.result-table tr.warn-row td {
    background: #fff4db;
}
.result-table a {
    color: #1565c0;
    text-decoration: none;
}
.result-table a:hover {
    text-decoration: underline;
}
@keyframes bom-progress-shimmer {
    0% { background-position: 0% 50%; }
    100% { background-position: 200% 50%; }
}
.query-title {
    font-size: 18px;
    font-weight: 700;
    margin-top: 6px;
    margin-bottom: 4px;
}
.query-inline-title {
    font-size: 18px;
    font-weight: 700;
    margin-top: 6px;
    margin-bottom: 0;
    color: #1f2937;
}
div[data-testid="stTextArea"] textarea {
    min-height: 90px !important;
    font-size: 16px !important;
}
div.stButton > button {
    display: block;
    margin: 0 auto;
    height: 48px;
    width: auto;
    min-width: 120px;
    padding: 0 18px;
    font-size: 18px;
    border-radius: 8px;
    white-space: nowrap;
}
</style>
""", unsafe_allow_html=True)

BOM_NONE_OPTION = "（不使用）"
BOM_ROLE_LABELS = {
    "model": "型号列",
    "spec": "规格列",
    "name": "品名列",
    "quantity": "数量列",
}
BOM_COLUMN_KEYWORDS = {
    "model": [
        ("manufacturerpartnumber", 120), ("partnumber", 110), ("partno", 110),
        ("mpn", 105), ("pn", 100), ("料号", 100), ("型号", 95), ("物料编码", 90),
        ("产品型号", 90), ("订货号", 85), ("规格型号", 80),
    ],
    "spec": [
        ("规格参数", 120), ("技术参数", 115), ("规格", 110), ("参数", 100),
        ("specification", 95), ("spec", 90), ("value", 70), ("参数描述", 70),
    ],
    "name": [
        ("物料名称", 120), ("品名", 110), ("名称", 100), ("描述", 90),
        ("description", 85), ("desc", 80), ("itemname", 80), ("name", 75),
    ],
    "quantity": [
        ("数量", 120), ("用量", 110), ("需求数量", 105), ("qty", 100),
        ("quantity", 100), ("pcs", 30), ("用数", 80),
    ],
}

BOM_COLUMN_NEGATIVE_KEYWORDS = {
    "quantity": [
        ("单价", 220), ("單價", 220), ("price", 220), ("金额", 220),
        ("rmb", 180), ("含税", 180), ("tax", 180), ("有效期", 160),
        ("lt", 120), ("leadtime", 120), ("交期", 120), ("moq", 80),
    ],
}

def image_to_base64(path):
    if not os.path.exists(path):
        return None
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

def clean_text(x):
    if x is None:
        return ""
    if isinstance(x, str):
        return x.strip()
    if not pd.api.types.is_scalar(x):
        try:
            values = list(x)
        except Exception:
            values = [x]
        parts = []
        for value in values:
            text = clean_text(value)
            if text != "":
                parts.append(text)
        if not parts:
            return ""
        if len(parts) == 1:
            return parts[0]
        return " ".join(parts).strip()
    if pd.isna(x):
        return ""
    return str(x).strip()

def display_text_width(text):
    text = clean_text(text)
    if text == "":
        return 0.0
    normalized = re.sub(r"\s+", " ", text)
    width = 0.0
    for ch in normalized:
        if ch == " ":
            width += 0.35
            continue
        if unicodedata.east_asian_width(ch) in {"W", "F"}:
            width += 2.0
        elif unicodedata.east_asian_width(ch) == "A":
            width += 1.1
        elif ch in ".,;:|/\\-()[]{}<>":
            width += 0.7
        else:
            width += 1.0
    return width

def clean_brand(x):
    return clean_text(x)

def clean_model(x):
    return clean_text(x).upper().replace(" ", "")

def extract_model_like_tokens(text):
    raw = clean_text(text).upper()
    if raw == "":
        return []
    tokens = []
    seen = set()
    whole = clean_model(raw)
    if (
        len(whole) >= 6
        and re.search(r"[A-Z]", whole)
        and re.search(r"\d", whole)
    ):
        tokens.append(whole)
        seen.add(whole)
    for token in re.findall(r"[A-Z0-9][A-Z0-9._/+\\-]{4,}", raw):
        compact = clean_model(token.strip(".,;:|()[]{}"))
        if compact == "" or compact in seen or len(compact) < 6:
            continue
        if not (re.search(r"[A-Z]", compact) and re.search(r"\d", compact)):
            continue
        tokens.append(compact)
        seen.add(compact)
    return tokens

def load_component_rows_by_query_model_tokens(query_text):
    tokens = extract_model_like_tokens(query_text)
    if not tokens:
        return pd.DataFrame(), [], ""
    for token in tokens:
        rows = load_component_rows_by_clean_model(token)
        if isinstance(rows, pd.DataFrame) and not rows.empty:
            return rows, tokens, token
    return pd.DataFrame(), tokens, ""

def clean_material(x):
    x = clean_text(x).upper()
    x = x.replace("（", "(").replace("）", ")").replace(" ", "")
    x = x.replace("C0G", "COG")
    x = x.replace("NP0", "NPO")
    return x

def clean_size(x):
    s = clean_text(x).replace(".0", "").replace(" ", "")
    pad_map = {
        "401": "0401",
        "402": "0402",
        "603": "0603",
        "805": "0805",
        "201": "0201",
        "1005": "01005",
        "8004": "008004",
        "102": "0102",
        "15008": "015008",
        "008004": "008004",
        "0401": "0401",
        "0102": "0102",
        "015008": "015008",
    }
    if s in pad_map:
        return pad_map[s]
    return s.upper()


SIZE_TOKEN_PATTERN = re.compile(r"(?<!\d)(008004|015008|01005|0102|0201|0401|0402|0603|0805|1206|1210|1808|1812|1825|2010|2220|2512|3225|4520|4532|5750)(?!\d)")
SPEC_EMBEDDED_MATERIALS = [
    ("COG(NPO)", "COG(NPO)"),
    ("C0G", "COG(NPO)"),
    ("COG", "COG(NPO)"),
    ("NP0", "COG(NPO)"),
    ("NPO", "COG(NPO)"),
    ("X5R", "X5R"),
    ("X7R", "X7R"),
    ("X7S", "X7S"),
    ("X7T", "X7T"),
    ("X6S", "X6S"),
    ("Y5V", "Y5V"),
]
RESISTOR_VALUE_PATTERN = re.compile(r"(?<![A-Z0-9])(\d+(?:\.\d+)?(?:mΩ|[RKM]|\s*Ω)|\d+[RKM]\d+)(?=(?:\+/-|[\s/|;,:()]|$))", flags=re.I)
RESISTOR_OHM_PATTERN = re.compile(r"(\d+(?:\.\d+)?mΩ|\d+(?:\.\d+)?(?:[RKM]\d+|[RKM]?)(?:\s*(?:OHMS?|Ω)))", flags=re.I)
RESISTOR_COMPACT_CONTEXT_PATTERN = re.compile(r"(?<![A-Z0-9])((?:R\d+(?:\.\d+)?|\d+(?:\.\d+)?[RKM](?:\d+)?))(?=(?:\+/-|[\s/|;,:()]|$))", flags=re.I)
TWO_DIM_SIZE_PATTERN = re.compile(r"(?:[DΦLW]?\s*)?(\d+(?:\.\d+)?)\s*[*X×]\s*(?:[HLDWΦ]?\s*)?(\d+(?:\.\d+)?)", flags=re.I)
THREE_DIM_SIZE_PATTERN = re.compile(r"(?:[DLWΦ]?\s*)?(\d+(?:\.\d+)?)\s*[*X×]\s*(?:[HLDWΦ]?\s*)?(\d+(?:\.\d+)?)\s*[*X×]\s*(?:[HLDWΦ]?\s*)?(\d+(?:\.\d+)?)", flags=re.I)
ELECTROLYTIC_SIZE_PATTERN = TWO_DIM_SIZE_PATTERN
PITCH_PATTERN = re.compile(r"(?:脚距|PITCH)\s*(\d+(?:\.\d+)?)\s*MM", flags=re.I)
VARISTOR_PITCH_PATTERN = re.compile(r"(?<![A-Z0-9])P\s*(\d+(?:\.\d+)?)(?=(?:\s*MM)?(?:[^A-Z0-9]|$))", flags=re.I)
MM_DIMENSION_PATTERN = re.compile(r"(?<![A-Z0-9])(\d+(?:\.\d+)?)\s*MM(?![A-Z0-9])", flags=re.I)
INDUCTANCE_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*(NH|UH|MH)\b", flags=re.I)
CURRENT_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*(MA|A)\b", flags=re.I)
FREQUENCY_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*(HZ|KHZ|MHZ)\b", flags=re.I)
VARISTOR_CODE_PATTERN = re.compile(r"(?<!\d)(\d{3})K(?![A-Z0-9])", flags=re.I)
DISC_SIZE_CODE_PATTERN = re.compile(r"(?<!\d)(\d{2})D(?!\d)", flags=re.I)
SMD_SIZE_CODE_PATTERN = re.compile(r"(?<![A-Z0-9])(0201|0401|0402|0603|0805|1206|1210|1812|2010|2512|3225)(?![A-Z0-9])", flags=re.I)
RESISTOR_TOLERANCE_CODE_MAP = {
    "B": "0.1",
    "C": "0.25",
    "D": "0.5",
    "F": "1",
    "G": "2",
    "J": "5",
    "K": "10",
    "M": "20",
}
RESISTOR_EIA96_VALUES = [
    100, 102, 105, 107, 110, 113, 115, 118, 121, 124, 127, 130,
    133, 137, 140, 143, 147, 150, 154, 158, 162, 165, 169, 174,
    178, 182, 187, 191, 196, 200, 205, 210, 215, 221, 226, 232,
    237, 243, 249, 255, 261, 267, 274, 280, 287, 294, 301, 309,
    316, 324, 332, 340, 348, 357, 365, 374, 383, 392, 402, 412,
    422, 432, 442, 453, 464, 475, 487, 499, 511, 523, 536, 549,
    562, 576, 590, 604, 619, 634, 649, 665, 681, 698, 715, 732,
    750, 768, 787, 806, 825, 845, 866, 887, 909, 931, 953, 976,
]
RESISTOR_EIA96_MULTIPLIERS = {
    "Z": 0.001,
    "Y": 0.01,
    "R": 0.01,
    "X": 0.1,
    "S": 0.1,
    "A": 1,
    "B": 10,
    "H": 10,
    "C": 100,
    "D": 1000,
    "E": 10000,
    "F": 100000,
}
TAI_RESISTOR_PREFIXES = ("RASS", "RMSV", "RMH", "RBA", "RMS", "RLS", "RB", "RM")
TAI_RESISTOR_SIZE_MAP = {
    "01": "01005",
    "02": "0201",
    "04": "0402",
    "06": "0603",
    "10": "0805",
    "12": "1206",
    "13": "1210",
    "20": "2010",
    "25": "2512",
}
TAI_RLS_OFFICIAL_PROFILE = {
    "系列": "RLS",
    "系列说明": "金属箔电流检测电阻",
    "器件类型": "合金电阻",
    "特殊用途": "车规 | 电流检测",
}
RESISTOR_MODEL_PREFIX_PATTERN = re.compile(
    r"^(AA|AC|AF|AR|ASG|AS|AT|RC|RT|WR|WF|MR|FCR|TRC|CR|TR|QR|CQ|NQ|LE|TC|MHR|PRF|NCP|NCU|RASS|RMSV|RMH|RBA|RMS|RLS|RB|RM|RTX|RTT|RAT|RLT)"
)
WALSIN_RESISTOR_SIZE_MAP = {
    "02": "01005",
    "03": "0201",
    "04": "0402",
    "06": "0603",
    "08": "0805",
    "12": "1206",
    "14": "1210",
    "18": "1812",
    "20": "2010",
    "25": "2512",
}
EVER_OHMS_RESISTOR_SIZE_MAP = {
    "02": "0201",
    "04": "0402",
    "06": "0603",
    "08": "0805",
    "12": "1206",
    "18": "1812",
    "20": "2010",
    "25": "2512",
}
UNIROYAL_RESISTOR_SIZE_MAP = {
    "01": "0201",
    "02": "0402",
    "03": "0603",
    "05": "0805",
    "06": "1206",
    "10": "1210",
    "12": "2010",
    "25": "2512",
}

RESISTOR_COMPONENT_TYPES = {"贴片电阻", "厚膜电阻", "薄膜电阻", "合金电阻", "碳膜电阻", "金属氧化膜电阻", "绕线电阻"}
VARISTOR_COMPONENT_TYPES = {"压敏电阻", "引线型压敏电阻", "贴片压敏电阻"}
SPECIAL_RESISTOR_COMPONENT_TYPES = {"热敏电阻"} | VARISTOR_COMPONENT_TYPES
CAPACITOR_COMPONENT_TYPES = {"MLCC", "薄膜电容", "钽电容", "铝电解电容", "引线型陶瓷电容"}
INDUCTOR_COMPONENT_TYPES = {"功率电感", "射频电感", "共模电感", "磁珠"}
POWER_INDUCTOR_COMPONENT_TYPES = {"功率电感", "射频电感"}
TIMING_COMPONENT_TYPES = {"晶振", "振荡器"}
ALL_RESISTOR_TYPES = RESISTOR_COMPONENT_TYPES | SPECIAL_RESISTOR_COMPONENT_TYPES
ALL_PASSIVE_COMPONENT_TYPES = CAPACITOR_COMPONENT_TYPES | ALL_RESISTOR_TYPES | INDUCTOR_COMPONENT_TYPES | TIMING_COMPONENT_TYPES
RESISTOR_SOURCE_EVIDENCE_TOKENS = ("来源:resistor", "source:resistor")
INDUCTOR_TOLERANCE_CODE_MAP = {"F": "1", "G": "2", "H": "3", "J": "5", "K": "10", "M": "20", "N": "30"}
COMPONENT_CATEGORY_DISPLAY_LABELS = {
    "MLCC": "陶瓷贴片电容（MLCC）",
    "薄膜电容": "薄膜电容（Film Capacitor）",
    "钽电容": "钽电容（Tantalum Capacitor）",
    "铝电解电容": "铝电解电容（Aluminum Electrolytic Capacitor）",
    "引线型陶瓷电容": "引线陶瓷电容（Lead Ceramic Capacitor）",
    "贴片电阻": "贴片电阻（Chip Resistor）",
    "厚膜电阻": "厚膜电阻（Thick Film Resistor）",
    "薄膜电阻": "薄膜电阻（Thin Film Resistor）",
    "合金电阻": "合金电阻（Metal Strip Resistor）",
    "碳膜电阻": "碳膜电阻（Carbon Film Resistor）",
    "金属氧化膜电阻": "金属氧化膜电阻（Metal Oxide Film Resistor）",
    "绕线电阻": "绕线电阻（Wirewound Resistor）",
    "热敏电阻": "热敏电阻（NTC Thermistor）",
    "压敏电阻": "压敏电阻（Varistor）",
    "引线型压敏电阻": "引线压敏电阻（Lead Varistor）",
    "贴片压敏电阻": "贴片压敏电阻（Chip Varistor）",
    "功率电感": "功率电感（Power Inductor）",
    "射频电感": "射频电感（RF Inductor）",
    "共模电感": "共模电感（Common Mode Choke）",
    "磁珠": "磁珠（Ferrite Bead）",
    "晶振": "石英晶体（Crystal Unit）",
    "振荡器": "晶体振荡器（Oscillator）",
}


def find_embedded_size(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    match = SIZE_TOKEN_PATTERN.search(upper)
    return clean_size(match.group(1)) if match else ""


def find_embedded_material(text):
    upper = clean_text(text).upper().replace(" ", "")
    if upper == "":
        return ""
    for token, normalized in SPEC_EMBEDDED_MATERIALS:
        if token in upper:
            return clean_material(normalized)
    return ""

def normalize_tolerance_number(value):
    try:
        num = abs(float(clean_text(value)))
    except:
        return None
    if float(num).is_integer():
        num = int(num)
    return str(num)

def normalize_pf_tolerance_number(value):
    try:
        num = abs(float(clean_text(value)))
    except:
        return None
    if float(num).is_integer():
        num = int(num)
    return str(num)

def normalize_tolerance_text(x):
    s = clean_text(x).upper()
    if s == "":
        return ""
    if s in {"NONE", "NAN", "NULL", "N/A", "NA", "-", "--"}:
        return ""

    s = s.replace("（", "(").replace("）", ")").replace(" ", "")
    s = s.replace("％", "%").replace("﹪", "%")
    s = s.replace("＋", "+").replace("﹢", "+")
    s = s.replace("／", "/").replace("\\", "/")
    s = s.replace("卤", "+/-").replace("±", "+/-")
    for dash in ["－", "–", "—", "―", "−"]:
        s = s.replace(dash, "-")

    tol_letter_map = {
        "B": "0.1pF",
        "C": "0.25pF",
        "D": "0.5pF",
        "F": "1",
        "G": "2",
        "J": "5",
        "K": "10",
        "M": "20",
        "W": "0.05pF",
        "Z": "+80/-20",
    }
    if s in tol_letter_map:
        return tol_letter_map[s]

    pf_pm_match = re.fullmatch(r"\+/-\s*(\d+(?:\.\d+)?)PF", s)
    if pf_pm_match:
        normalized_pf = normalize_pf_tolerance_number(pf_pm_match.group(1))
        return f"{normalized_pf}PF" if normalized_pf is not None else s

    pf_plain_match = re.fullmatch(r"(\d+(?:\.\d+)?)PF", s)
    if pf_plain_match:
        normalized_pf = normalize_pf_tolerance_number(pf_plain_match.group(1))
        return f"{normalized_pf}PF" if normalized_pf is not None else s

    s = s.replace("%", "")
    s = s.replace("+-", "+/-")

    if s.startswith("+/-"):
        normalized = normalize_tolerance_number(s[3:])
        return normalized if normalized is not None else s

    asymmetric_patterns = [
        (r"^\+?(\d+(?:\.\d+)?)/-(\d+(?:\.\d+)?)$", True),
        (r"^-(\d+(?:\.\d+)?)/\+?(\d+(?:\.\d+)?)$", False),
    ]
    for pattern, plus_first in asymmetric_patterns:
        m = re.fullmatch(pattern, s)
        if m:
            first = normalize_tolerance_number(m.group(1))
            second = normalize_tolerance_number(m.group(2))
            if first is None or second is None:
                return s
            if first == second:
                return first
            plus_val, minus_val = (first, second) if plus_first else (second, first)
            return f"+{plus_val}/-{minus_val}"

    normalized = normalize_tolerance_number(s)
    if normalized is not None:
        return normalized
    return s

def clean_tol_for_match(x):
    x = normalize_tolerance_text(x)
    if x == "":
        return ""
    return x

def clean_tol_for_display(x):
    x = normalize_tolerance_text(x)
    if x == "":
        return ""
    if re.fullmatch(r"\d+(?:\.\d+)?PF", x):
        return f"{x[:-2]}pF"
    if re.fullmatch(r"\+?\d+(?:\.\d+)?/-\d+(?:\.\d+)?", x):
        return f"{x}%"
    if re.fullmatch(r"\d+(?:\.\d+)?", x):
        return f"±{x}%"
    return x


def infer_tolerance_from_model(model):
    text = clean_model(model).upper()
    if text == "":
        return ""

    tol_letter_map = {
        "B": "0.1pF",
        "C": "0.25pF",
        "D": "0.5pF",
        "F": "1",
        "G": "2",
        "J": "5",
        "K": "10",
        "M": "20",
        "W": "0.05pF",
        "Z": "+80/-20",
    }
    if text.startswith("CC"):
        search_text = text
        patterns = [
            r"^CC\d{4}([BCDFGJKMWZ])",
            r"\dR\d([BCDFGJKMWZ])\d",
            r"\d{3}([BCDFGJKMWZ])\d",
        ]
    elif text.startswith("C"):
        search_text = text[5:]
        patterns = [
            r"\d{3}([BCDFGJKMWZ])\d",
            r"\dR\d([BCDFGJKMWZ])\d",
        ]
    else:
        search_text = text
        patterns = [
            r"\d{3}([BCDFGJKMWZ])\d",
            r"\dR\d([BCDFGJKMWZ])\d",
        ]

    for pattern in patterns:
        match = re.search(pattern, search_text)
        if not match:
            continue
        return clean_tol_for_match(tol_letter_map.get(match.group(1), ""))
    return ""


def tolerance_sort_key(value):
    tol = clean_tol_for_match(value)
    if tol == "":
        return ("empty", None)

    pf_match = re.fullmatch(r"(\d+(?:\.\d+)?)PF", tol)
    if pf_match:
        return ("pf", float(pf_match.group(1)))

    rank_map = {"1": 1, "2": 2, "5": 3, "10": 4, "20": 5, "+80/-20": 6}
    if tol in rank_map:
        return ("percent", rank_map[tol])

    return ("raw", tol)


def tolerance_equal(left, right):
    left_key = tolerance_sort_key(left)
    right_key = tolerance_sort_key(right)
    if left_key[0] == "empty" or right_key[0] == "empty":
        return False
    return left_key == right_key


def tolerance_allows(candidate, required):
    required_key = tolerance_sort_key(required)
    if required_key[0] == "empty":
        return True

    candidate_key = tolerance_sort_key(candidate)
    if candidate_key[0] == "empty":
        return False

    if candidate_key[0] != required_key[0]:
        return False

    if candidate_key[0] in {"percent", "pf"}:
        return candidate_key[1] <= required_key[1]

    return candidate_key[1] == required_key[1]


def tolerance_strictly_better(candidate, required):
    required_key = tolerance_sort_key(required)
    candidate_key = tolerance_sort_key(candidate)
    if candidate_key[0] == "empty" or candidate_key[0] != required_key[0]:
        return False
    if candidate_key[0] in {"percent", "pf"}:
        return candidate_key[1] < required_key[1]
    return False

def tolerance_equal_series(base, required):
    kind, value = tolerance_sort_key(required)
    if kind == "empty":
        return pd.Series(False, index=base.index)
    if kind == "raw":
        return base["_tol"].eq(clean_tol_for_match(required))
    return base["_tol_kind"].eq(kind) & base["_tol_num"].notna() & base["_tol_num"].eq(float(value))

def tolerance_allows_series(base, required):
    kind, value = tolerance_sort_key(required)
    if kind == "empty":
        return pd.Series(True, index=base.index)
    if kind == "raw":
        return base["_tol"].eq(clean_tol_for_match(required))
    return base["_tol_kind"].eq(kind) & base["_tol_num"].notna() & base["_tol_num"].le(float(value))

def parse_tolerance_token(token):
    t = clean_text(token).upper().replace(" ", "")
    if t == "":
        return ""
    leading_patterns = [
        r"^\+/-\s*\d+(?:\.\d+)?PF",
        r"^\+/-\s*\d+(?:\.\d+)?%",
        r"^\d+(?:\.\d+)?PF",
        r"^\d+(?:\.\d+)?%",
        r"^\+\d+(?:\.\d+)?%?/\-\d+(?:\.\d+)?%?",
        r"^\-\d+(?:\.\d+)?%?/\+\d+(?:\.\d+)?%?",
    ]
    for pattern in leading_patterns:
        match = re.match(pattern, t)
        if match:
            return clean_tol_for_match(match.group(0))
    tol_letter_map = {"F", "G", "J", "K", "M", "Z"}
    if t in tol_letter_map:
        return clean_tol_for_match(t)
    return ""

def clean_voltage(x):
    text = html.unescape(clean_text(x))
    if text == "":
        return ""

    placeholder_tokens = {"", "none", "nan", "null", "n/a", "na", "n.a.", "-", "--", "/", "\\"}

    def _format_voltage_number(value):
        try:
            num = float(value)
        except Exception:
            return clean_text(value)
        if abs(num - round(num)) < 1e-9:
            return str(int(round(num)))
        return f"{num:.6f}".rstrip("0").rstrip(".")

    compact_lower = re.sub(r"\s+", "", text).lower()
    if compact_lower in placeholder_tokens:
        return ""

    normalized = text.replace("−", "-").replace("–", "-").replace("—", "-")
    normalized = normalized.replace("～", "~").replace("至", "~")
    normalized = re.sub(r"(?i)\s*V(?:DC|AC)?\b", "", normalized)
    compact = re.sub(r"\s+", "", normalized)
    compact = re.sub(r"(?i)(?<=\d)TO(?=[+\-]?\d)", "~", compact)
    if compact.lower() in placeholder_tokens:
        return ""

    range_match = re.fullmatch(r"([+\-]?\d+(?:\.\d+)?)(?:~|-)([+\-]?\d+(?:\.\d+)?)", compact)
    if range_match:
        return f"{_format_voltage_number(range_match.group(1))}~{_format_voltage_number(range_match.group(2))}"

    if re.fullmatch(r"[+\-]?\d+(?:\.\d+)?", compact):
        return _format_voltage_number(compact)

    compact_lower = compact.lower()
    if compact_lower in placeholder_tokens:
        return ""
    return compact


def normalize_dimension_mm_value(value):
    text = html.unescape(clean_text(value))
    if text == "":
        return ""
    text = text.replace("\xa0", " ").replace("MM", "mm")
    text = re.sub(r"\s+", "", text)
    if text.lower().endswith("mm"):
        text = text[:-2]
    return text


def build_dimension_field_map(length="", width="", height=""):
    fields = {}
    for col, value in [("长度（mm）", length), ("宽度（mm）", width), ("高度（mm）", height)]:
        normalized = normalize_dimension_mm_value(value)
        if normalized != "":
            fields[col] = normalized
    return fields


def infer_mlcc_nominal_dimension_fields_from_size(size_hint):
    size_key = clean_size(size_hint)
    nominal_size = MLCC_NOMINAL_SIZE_MM_MAP.get(size_key)
    if not nominal_size:
        return {}
    fields = build_dimension_field_map(
        nominal_size[0],
        nominal_size[1],
        MLCC_NOMINAL_THICKNESS_MM_MAP.get(size_key, ""),
    )
    return fields


def merge_dimension_fields_into_record(record, fields, override_conflicts=False):
    merged = dict(record)
    for col in ["长度（mm）", "宽度（mm）", "高度（mm）"]:
        parsed_value = normalize_dimension_mm_value(fields.get(col, ""))
        current_value = normalize_dimension_mm_value(merged.get(col, ""))
        should_replace = override_conflicts or current_value == ""
        if (
            not should_replace
            and parsed_value != ""
            and "±" in parsed_value
            and "±" not in current_value
        ):
            should_replace = True
        if parsed_value != "" and should_replace:
            merged[col] = parsed_value
    return merged


def extract_dimension_fields_from_text(text):
    raw = html.unescape(clean_text(text))
    if raw == "":
        return {}
    compact = re.sub(r"\s+", "", raw)
    match = re.search(r"(\d+(?:\.\d+)?)\s*[x×]\s*(\d+(?:\.\d+)?)\s*[x×]\s*(\d+(?:\.\d+)?)\s*mm", compact, flags=re.I)
    if match:
        return build_dimension_field_map(match.group(1), match.group(2), match.group(3))
    match = re.search(r"(\d+(?:\.\d+)?)\s*[x×]\s*(\d+(?:\.\d+)?)\s*mm", compact, flags=re.I)
    if match:
        return build_dimension_field_map(match.group(1), match.group(2), "")
    return {}


def load_json_file_if_exists(path, default):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return default


def extract_url_candidates_from_text(text):
    raw = html.unescape(clean_text(text))
    if raw == "":
        return []
    candidates = []
    seen = set()
    for match in re.finditer(r"https?://[^\s<>'\"）】\]]+", raw, flags=re.I):
        candidate = match.group(0).rstrip(").,;:，。；】]}'\"")
        if candidate == "" or candidate in seen:
            continue
        seen.add(candidate)
        candidates.append(candidate)
    return candidates


def normalize_official_dimension_lookup_text(text):
    raw = html.unescape(clean_text(text))
    if raw == "":
        return ""
    normalized = raw.replace("\r", "\n").replace("\xa0", " ")
    normalized = re.sub(r"(?is)<(script|style)[^>]*>.*?</\1>", " ", normalized)
    normalized = re.sub(r"(?i)<br\s*/?>", "\n", normalized)
    normalized = re.sub(r"(?i)</(tr|p|div|li|section|article|table|thead|tbody|tfoot|h[1-6])>", "\n", normalized)
    normalized = re.sub(r"(?i)<(td|th)[^>]*>", " ", normalized)
    normalized = re.sub(r"(?i)<[^>]+>", " ", normalized)
    normalized = re.sub(r"[ \t]+", " ", normalized)
    normalized = re.sub(r"\n{2,}", "\n", normalized)
    return normalized.strip()


MLCC_DIMENSION_CELL_PATTERN = re.compile(r"\d+(?:\.\d+)?(?:\s*±\s*\d+(?:\.\d+)?)?")
MLCC_SIZE_SEARCH_TOKENS = {
    "01005": {"01005", "0100", "0402", "01R5", "C0402", "CC01005"},
    "0201": {"0201", "0603", "C0603", "CC0201"},
    "0204": {"0204"},
    "0402": {"0402", "1005", "C1005", "CC0402"},
    "0505": {"0505"},
    "0603": {"0603", "1608", "C1608", "CC0603"},
    "0612": {"0612"},
    "0805": {"0805", "2012", "C2012", "CC0805"},
    "1206": {"1206", "3216", "C3216", "CC1206"},
    "1210": {"1210", "3225", "C3225", "CC1210"},
    "2010": {"2010", "5025", "C5025", "CC2010"},
    "2512": {"2512", "6332", "C6332", "CC2512"},
    "3225": {"3225", "1210", "C3225", "CC3225"},
    "1808": {"1808", "4520", "C4520", "CC1808"},
    "1825": {"1825"},
    "1812": {"1812", "4532", "C4532", "CC1812"},
    "2211": {"2211"},
    "2220": {"2220", "5750", "C5750", "CC2220"},
    "2225": {"2225"},
    "4520": {"4520", "1808", "C4520", "CC4520"},
    "4532": {"4532", "1812", "C4532", "CC4532"},
    "5750": {"5750", "2220", "C5750", "CC5750"},
    "6332": {"6332", "2512", "C6332", "CC6332"},
}
MLCC_NOMINAL_SIZE_MM_MAP = {
    "01005": ("0.40", "0.20"),
    "0201": ("0.60", "0.30"),
    "0204": ("0.20", "0.40"),
    "0402": ("1.00", "0.50"),
    "0505": ("0.50", "0.50"),
    "0603": ("1.60", "0.80"),
    "0612": ("0.60", "1.20"),
    "0805": ("2.00", "1.25"),
    "1206": ("3.20", "1.60"),
    "1210": ("3.20", "2.50"),
    "2010": ("5.00", "2.50"),
    "2512": ("6.30", "3.20"),
    "3225": ("3.20", "2.50"),
    "1808": ("4.50", "2.00"),
    "1825": ("1.80", "2.50"),
    "1812": ("4.50", "3.20"),
    "2211": ("2.20", "1.10"),
    "4520": ("4.50", "2.00"),
    "4532": ("4.50", "3.20"),
    "2220": ("5.70", "5.00"),
    "2225": ("2.20", "2.50"),
    "5750": ("5.70", "5.00"),
    "6332": ("6.30", "3.20"),
}
MLCC_NOMINAL_THICKNESS_MM_MAP = {
    "01005": "0.20",
    "0201": "0.30",
    "0204": "0.20",
    "0402": "0.50",
    "0505": "0.50",
    "0603": "0.80",
    "0612": "0.60",
    "0805": "1.25",
    "1206": "1.60",
    "1210": "2.50",
    "1808": "2.00",
    "1825": "2.50",
    "1812": "3.20",
    "2010": "2.50",
    "2220": "5.00",
    "2211": "1.10",
    "2225": "2.50",
    "2512": "3.20",
    "3225": "2.50",
    "4520": "2.00",
    "4532": "3.20",
    "5750": "5.00",
    "6332": "3.20",
}
MURATA_SIZE_DIMENSION_MAP = {
    "02": ("0.40", "0.20"),
    "03": ("0.60", "0.30"),
    "05": ("0.50", "0.50"),
    "08": ("0.80", "0.80"),
    "0D": ("0.38", "0.38"),
    "0M": ("0.90", "0.60"),
    "15": ("1.00", "0.50"),
    "18": ("1.60", "0.80"),
    "1M": ("1.37", "1.00"),
    "21": ("2.00", "1.25"),
    "22": ("2.80", "2.80"),
    "31": ("3.20", "1.60"),
    "32": ("3.20", "2.50"),
    "42": ("4.50", "2.00"),
    "43": ("4.50", "3.20"),
    "52": ("5.70", "2.80"),
    "55": ("5.70", "5.00"),
}
MURATA_THICKNESS_CODE_MAP = {
    "2": "0.20",
    "3": "0.30",
    "5": "0.50",
    "6": "0.60",
    "7": "0.70",
    "8": "0.80",
    "9": "0.85",
    "A": "1.00",
    "B": "1.25",
    "C": "1.60",
    "D": "2.00",
    "E": "2.50",
    "F": "3.20",
    "M": "1.15",
    "N": "1.35",
    "Q": "1.50",
    "R": "1.80",
    "S": "2.80",
}
TDK_SIZE_DIMENSION_MAP = {
    "0402": ("0.40±0.02", "0.20±0.02"),
    "0603": ("0.60±0.03", "0.30±0.03"),
    "1005": ("1.00±0.05", "0.50±0.05"),
    "1608": ("1.60±0.10", "0.80±0.10"),
    "2012": ("2.00±0.20", "1.25±0.20"),
    "3216": ("3.20±0.20", "1.60±0.20"),
    "3225": ("3.20±0.40", "2.50±0.30"),
    "4532": ("4.50±0.40", "3.20±0.40"),
    "5750": ("5.70±0.40", "5.00±0.40"),
}
TDK_THICKNESS_CODE_MAP = {
    "020": "0.20",
    "030": "0.30",
    "050": "0.50",
    "060": "0.60",
    "080": "0.80",
    "085": "0.85",
    "115": "1.15",
    "125": "1.25",
    "130": "1.30",
    "160": "1.60",
    "200": "2.00",
    "230": "2.30",
    "250": "2.50",
    "280": "2.80",
    "320": "3.20",
}

FENGHUA_AM_SIZE_CODE_MAP = {
    "02": "0402",
    "03": "0603",
    "05": "0805",
    "06": "1206",
    "08": "1808",
    "10": "1210",
    "12": "1812",
    "20": "2220",
}
FENGHUA_AM_DIELECTRIC_CODE_MAP = {
    "B": "X7R",
    "C": "COG(NPO)",
}
FENGHUA_AM_TOLERANCE_CODE_MAP = {
    "F": "1",
    "G": "2",
    "J": "5",
    "K": "10",
    "M": "20",
}
FENGHUA_AM_VOLTAGE_CODE_MAP = {
    "160": "16",
    "250": "25",
    "500": "50",
    "631": "630",
    "101": "100",
    "102": "1000",
    "202": "2000",
}
FENGHUA_AM_DIMENSION_MAP = {
    ("02", "500"): ("1.00±0.05", "0.50±0.05", "0.50±0.05"),
    ("03", "160"): ("1.60±0.10", "0.80±0.10", "0.80±0.10"),
    ("03", "500"): ("1.60±0.10", "0.80±0.10", "0.80±0.10"),
    ("03", "101"): ("1.60±0.10", "0.80±0.10", "0.80±0.10"),
    ("05", "500"): ("2.00±0.20", "1.25±0.20", "1.25±0.20"),
    ("05", "101"): ("2.00±0.20", "1.25±0.20", "1.25±0.20"),
    ("10", "202"): ("3.20±0.30", "2.50±0.30", "2.00±0.30"),
}
FENGHUA_AM_MODEL_PATTERN = re.compile(
    r"^AM(?P<size>\d{2})(?P<dielectric>[A-Z])(?P<cap>(?:\d{3,4}|[0-9]R[0-9]+|R\d+))(?P<tol>[FGJKM])(?P<volt>\d{3})(?P<tail>[A-Z0-9]*)$"
)
FENGHUA_AM_SERIES_MEANING = {
    "AM": "汽车级 / AEC-Q200",
}


def build_mlcc_size_search_tokens(model="", size_hint=""):
    model_key = clean_model(model)
    size_key = clean_size(size_hint)
    tokens = set()
    if size_key:
        tokens.add(size_key)
        tokens.update(MLCC_SIZE_SEARCH_TOKENS.get(size_key, set()))
    if re.match(r"^C\d{4}", model_key):
        tokens.add(model_key[1:5])
        tokens.add("C" + model_key[1:5])
    if re.match(r"^C[CQ]\d{4}", model_key):
        tokens.add(model_key[2:6])
        tokens.add(model_key[:6])
    if re.match(r"^[CQ]C\d{4}", model_key):
        tokens.add(model_key[2:6])
        tokens.add(model_key[:6])
    if model_key.startswith("01R5"):
        tokens.add("01R5")
    if len(model_key) >= 4 and model_key[:4].isdigit():
        tokens.add(model_key[:4])
    if len(model_key) >= 2 and model_key[:2].isdigit():
        tokens.add(model_key[:2])
    return sorted((token for token in tokens if clean_text(token) != ""), key=lambda token: (-len(token), token))


def load_mlcc_lcsc_dimension_cache():
    global MLCC_LCSC_DIMENSION_CACHE, MLCC_LCSC_DIMENSION_CACHE_SIGNATURE
    signature = get_path_signature(MLCC_LCSC_DIMENSION_CACHE_PATH)
    if (
        MLCC_LCSC_DIMENSION_CACHE is not None
        and MLCC_LCSC_DIMENSION_CACHE_SIGNATURE == signature
    ):
        return MLCC_LCSC_DIMENSION_CACHE
    cached = load_json_file_if_exists(MLCC_LCSC_DIMENSION_CACHE_PATH, {})
    MLCC_LCSC_DIMENSION_CACHE = cached if isinstance(cached, dict) else {}
    MLCC_LCSC_DIMENSION_CACHE_SIGNATURE = signature
    return MLCC_LCSC_DIMENSION_CACHE


def save_mlcc_lcsc_dimension_cache():
    global MLCC_LCSC_DIMENSION_CACHE_SIGNATURE
    cache = load_mlcc_lcsc_dimension_cache()
    cache_dir = os.path.dirname(MLCC_LCSC_DIMENSION_CACHE_PATH)
    if cache_dir:
        os.makedirs(cache_dir, exist_ok=True)
    with open(MLCC_LCSC_DIMENSION_CACHE_PATH, "w", encoding="utf-8") as handle:
        json.dump(cache, handle, ensure_ascii=False, indent=2)
    MLCC_LCSC_DIMENSION_CACHE_SIGNATURE = get_path_signature(MLCC_LCSC_DIMENSION_CACHE_PATH)


def load_mlcc_official_dimension_cache():
    global MLCC_OFFICIAL_DIMENSION_CACHE, MLCC_OFFICIAL_DIMENSION_CACHE_SIGNATURE
    signature = get_path_signature(MLCC_OFFICIAL_DIMENSION_CACHE_PATH)
    if (
        MLCC_OFFICIAL_DIMENSION_CACHE is not None
        and MLCC_OFFICIAL_DIMENSION_CACHE_SIGNATURE == signature
    ):
        return MLCC_OFFICIAL_DIMENSION_CACHE
    cached = load_json_file_if_exists(MLCC_OFFICIAL_DIMENSION_CACHE_PATH, {})
    MLCC_OFFICIAL_DIMENSION_CACHE = cached if isinstance(cached, dict) else {}
    MLCC_OFFICIAL_DIMENSION_CACHE_SIGNATURE = signature
    return MLCC_OFFICIAL_DIMENSION_CACHE


def save_mlcc_official_dimension_cache():
    global MLCC_OFFICIAL_DIMENSION_CACHE_SIGNATURE
    cache = load_mlcc_official_dimension_cache()
    cache_dir = os.path.dirname(MLCC_OFFICIAL_DIMENSION_CACHE_PATH)
    if cache_dir:
        os.makedirs(cache_dir, exist_ok=True)
    with open(MLCC_OFFICIAL_DIMENSION_CACHE_PATH, "w", encoding="utf-8") as handle:
        json.dump(cache, handle, ensure_ascii=False, indent=2)
    MLCC_OFFICIAL_DIMENSION_CACHE_SIGNATURE = get_path_signature(MLCC_OFFICIAL_DIMENSION_CACHE_PATH)


def http_fetch_url_text(url, timeout=18):
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8",
        },
    )
    context = ssl._create_unverified_context()
    with urllib.request.urlopen(request, timeout=timeout, context=context) as response:
        payload = response.read()
        charset = getattr(response.headers, "get_content_charset", lambda: None)() or "utf-8"
    try:
        return payload.decode(charset, errors="replace")
    except Exception:
        return payload.decode("utf-8", errors="replace")


def http_fetch_url_bytes(url, timeout=24):
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8",
        },
    )
    context = ssl._create_unverified_context()
    with urllib.request.urlopen(request, timeout=timeout, context=context) as response:
        return response.read()


def extract_lcsc_product_code(url):
    raw = clean_text(url)
    if raw == "":
        return ""
    match = re.search(r"/product-detail(?:/[^/?#]+)?/(C\d+)\.html", raw, flags=re.I)
    if match:
        return clean_model(match.group(1))
    match = re.search(r"/datasheet/(C\d+)\.pdf", raw, flags=re.I)
    if match:
        return clean_model(match.group(1))
    match = re.search(r"\b(C\d{4,})\b", raw, flags=re.I)
    if match:
        return clean_model(match.group(1))
    return ""


def extract_lcsc_datasheet_pdf_url(page_text, product_code):
    if clean_text(page_text) == "" or clean_text(product_code) == "":
        return ""
    patterns = [
        rf"https:(?:\\\\u002F|/)(?:\\\\u002F|/)datasheet\.lcsc\.com(?:\\\\u002F|/)[^\"']+?\.pdf\?productCode={re.escape(product_code)}",
        rf"https:(?:\\\\u002F|/)(?:\\\\u002F|/)[^\"']+?\.pdf\?productCode={re.escape(product_code)}",
    ]
    for pattern in patterns:
        match = re.search(pattern, page_text, flags=re.I)
        if match:
            return html.unescape(match.group(0).replace("\\u002F", "/"))
    return ""


def extract_pdf_text_from_bytes(pdf_bytes, max_pages=18):
    if not pdf_bytes:
        return ""
    try:
        from pypdf import PdfReader
    except Exception:
        return ""
    try:
        reader = PdfReader(BytesIO(pdf_bytes))
    except Exception:
        return ""
    texts = []
    for page in reader.pages[:max_pages]:
        try:
            texts.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(texts)


def normalize_pdf_dimension_text(text):
    raw = clean_text(text)
    if raw == "":
        return ""
    normalized = html.unescape(raw).replace("\r", "\n").replace("\xa0", " ")
    normalized = normalized.replace("＋", "+").replace("－", "-").replace("−", "-")
    normalized = normalized.replace("± ", "±").replace(" ±", "±")
    normalized = re.sub(r"[ \t]+", " ", normalized)
    normalized = re.sub(r"\s*±\s*", "±", normalized)
    return normalized


def dimension_cell_leading_number(text):
    match = re.match(r"(\d+(?:\.\d+)?)", normalize_dimension_mm_value(text))
    if not match:
        return None
    try:
        return float(match.group(1))
    except Exception:
        return None


def get_mlcc_nominal_size_mm(size_hint):
    size_key = clean_size(size_hint)
    nominal_size = MLCC_NOMINAL_SIZE_MM_MAP.get(size_key)
    if not nominal_size:
        return None
    try:
        return float(nominal_size[0]), float(nominal_size[1])
    except Exception:
        return None


def score_mlcc_dimension_fields_against_size_hint(fields, size_hint):
    nominal_size = get_mlcc_nominal_size_mm(size_hint)
    if nominal_size is None or not isinstance(fields, dict):
        return None
    length = dimension_cell_leading_number(fields.get("长度（mm）", ""))
    width = dimension_cell_leading_number(fields.get("宽度（mm）", ""))
    if length is None or width is None:
        return None
    return abs(length - nominal_size[0]) + abs(width - nominal_size[1])


def mlcc_dimension_fields_match_size_hint(fields, size_hint):
    nominal_size = get_mlcc_nominal_size_mm(size_hint)
    if nominal_size is None or not isinstance(fields, dict):
        return True
    length = dimension_cell_leading_number(fields.get("长度（mm）", ""))
    width = dimension_cell_leading_number(fields.get("宽度（mm）", ""))
    if length is None or width is None:
        return True
    length_tolerance = max(nominal_size[0] * 0.25, 0.08)
    width_tolerance = max(nominal_size[1] * 0.25, 0.08)
    return abs(length - nominal_size[0]) <= length_tolerance and abs(width - nominal_size[1]) <= width_tolerance


def mlcc_dimension_fields_look_suspicious(fields):
    if not isinstance(fields, dict):
        return True
    length = normalize_dimension_mm_value(fields.get("长度（mm）", ""))
    width = normalize_dimension_mm_value(fields.get("宽度（mm）", ""))
    if length == "" or width == "":
        return False
    for value in [length, width]:
        if "±" in value or "." in value:
            return False
    return True


def extract_dimension_fields_from_size_table(text, size_candidates, size_hint=""):
    normalized_text = normalize_pdf_dimension_text(text)
    if normalized_text == "" or not size_candidates:
        return {}

    lines = [line.strip() for line in normalized_text.splitlines() if clean_text(line) != ""]
    line_entries = []
    for line in lines:
        normalized_line = re.sub(r"\s+", " ", line)
        line_entries.append((line, normalized_line))

    best_candidate = None
    best_key = None
    for token_rank, token in enumerate(size_candidates):
        token_pattern = re.compile(rf"(?<![A-Z0-9]){re.escape(token)}(?![A-Z0-9])", flags=re.I)
        for original_line, normalized_line in line_entries:
            if not token_pattern.search(normalized_line):
                continue
            tail = token_pattern.split(normalized_line, maxsplit=1)[-1]
            cells = [normalize_dimension_mm_value(cell) for cell in MLCC_DIMENSION_CELL_PATTERN.findall(tail)]
            cells = [cell for cell in cells if cell != ""]
            if len(cells) < 3:
                continue
            leading = [dimension_cell_leading_number(cell) for cell in cells[:3]]
            if any(value is None for value in leading):
                continue
            if any(value > 10 for value in leading[:3]):
                continue
            if sum(1 for cell in cells[:3] if "±" in cell) < 2:
                continue
            candidate_fields = build_dimension_field_map(cells[0], cells[1], cells[2])
            score = score_mlcc_dimension_fields_against_size_hint(candidate_fields, size_hint)
            if score is None:
                score = float(token_rank)
            fits_size_hint = mlcc_dimension_fields_match_size_hint(candidate_fields, size_hint)
            sort_key = (0 if fits_size_hint else 1, round(score, 6), token_rank)
            if best_candidate is None or sort_key < best_key:
                best_candidate = candidate_fields
                best_key = sort_key
    if best_candidate and mlcc_dimension_fields_match_size_hint(best_candidate, size_hint):
        return best_candidate
    if get_mlcc_nominal_size_mm(size_hint) is None:
        return best_candidate or {}
    return {}


def extract_dimension_fields_from_model_summary_line(text, model, size_hint=""):
    normalized_text = normalize_pdf_dimension_text(text)
    model_key = clean_model(model)
    size_key = clean_size(size_hint)
    if normalized_text == "" or model_key == "" or size_key == "":
        return {}
    nominal_size = MLCC_NOMINAL_SIZE_MM_MAP.get(size_key)
    if not nominal_size:
        return {}

    model_candidates = [model_key]
    for trim in [1, 2, 3]:
        if len(model_key) > trim:
            model_candidates.append(model_key[:-trim])

    for line in normalized_text.splitlines():
        normalized_line = re.sub(r"\s+", " ", line.strip())
        if normalized_line == "":
            continue
        compact_line = clean_model(normalized_line)
        if not any(candidate != "" and candidate in compact_line for candidate in model_candidates):
            continue
        direct_fields = extract_dimension_fields_from_text(normalized_line)
        if direct_fields:
            return direct_fields
        thickness_match = re.search(r"t\s*=\s*(\d+(?:\.\d+)?)\s*mm", normalized_line, flags=re.I)
        if thickness_match:
            return build_dimension_field_map(nominal_size[0], nominal_size[1], thickness_match.group(1))
    return {}


def decode_murata_dimension_fields_from_model(model):
    model_key = clean_model(model)
    prefixes = ("GRM", "GCM", "GCJ", "GJM", "GQM", "GRT", "GCG", "GCQ")
    prefix = next((item for item in prefixes if model_key.startswith(item)), None)
    if prefix is None or len(model_key) < len(prefix) + 3:
        return {}
    size_code = model_key[len(prefix):len(prefix) + 2]
    thickness_code = model_key[len(prefix) + 2:len(prefix) + 3]
    length_width = MURATA_SIZE_DIMENSION_MAP.get(size_code)
    thickness = MURATA_THICKNESS_CODE_MAP.get(thickness_code, "")
    if not length_width:
        return {}
    return build_dimension_field_map(length_width[0], length_width[1], thickness)


def decode_tdk_dimension_fields_from_model(model):
    model_key = clean_model(model)
    if not model_key.startswith("C") or len(model_key) < 17:
        return {}
    size_code = model_key[1:5]
    thickness_code = model_key[14:17]
    length_width = TDK_SIZE_DIMENSION_MAP.get(size_code)
    thickness = TDK_THICKNESS_CODE_MAP.get(thickness_code, "")
    if not length_width:
        return {}
    return build_dimension_field_map(length_width[0], length_width[1], thickness)


def decode_fenghua_am_dimension_fields_from_model(model):
    model_key = clean_model(model)
    match = FENGHUA_AM_MODEL_PATTERN.fullmatch(model_key)
    if match is None:
        return {}
    size_code = match.group("size")
    volt_code = match.group("volt")
    length_width_height = FENGHUA_AM_DIMENSION_MAP.get((size_code, volt_code))
    if length_width_height is None and size_code in {"02", "03", "05"}:
        generic_dims = {
            "02": ("1.00±0.05", "0.50±0.05", "0.50±0.05"),
            "03": ("1.60±0.10", "0.80±0.10", "0.80±0.10"),
            "05": ("2.00±0.20", "1.25±0.20", "1.25±0.20"),
        }
        length_width_height = generic_dims.get(size_code)
    if not length_width_height:
        return {}
    return build_dimension_field_map(length_width_height[0], length_width_height[1], length_width_height[2])


def lookup_mlcc_lcsc_dimension_fields(model, brand="", lcsc_url="", size_hint="", allow_online_lookup=True):
    product_code = extract_lcsc_product_code(lcsc_url)
    if product_code == "":
        return {}

    with MLCC_LCSC_DIMENSION_CACHE_LOCK:
        cache = load_mlcc_lcsc_dimension_cache()
        if product_code in cache and isinstance(cache.get(product_code), dict):
            cached_fields = cache.get(product_code, {}).copy()
            if (
                cached_fields
                and not mlcc_dimension_fields_look_suspicious(cached_fields)
                and mlcc_dimension_fields_match_size_hint(cached_fields, size_hint)
            ):
                return cached_fields

    if not allow_online_lookup:
        return {}

    try:
        datasheet_page_text = http_fetch_url_text(f"https://www.lcsc.com/datasheet/{product_code}.pdf")
        pdf_url = extract_lcsc_datasheet_pdf_url(datasheet_page_text, product_code)
        fields = {}
        if pdf_url != "":
            pdf_text = extract_pdf_text_from_bytes(http_fetch_url_bytes(pdf_url))
            fields = extract_dimension_fields_from_size_table(
                pdf_text,
                build_mlcc_size_search_tokens(model, size_hint=size_hint),
                size_hint=size_hint,
            )
            if not fields:
                fields = extract_dimension_fields_from_model_summary_line(
                    pdf_text,
                    model,
                    size_hint=size_hint,
                )
        fields = {col: value for col, value in fields.items() if normalize_dimension_mm_value(value) != ""}
        if fields and not mlcc_dimension_fields_match_size_hint(fields, size_hint):
            fields = {}
    except Exception:
        fields = {}

    with MLCC_LCSC_DIMENSION_CACHE_LOCK:
        cache = load_mlcc_lcsc_dimension_cache()
        cache[product_code] = fields.copy()
        try:
            save_mlcc_lcsc_dimension_cache()
        except Exception:
            pass
    return fields.copy()


def is_pdf_url(url):
    raw = clean_text(url).lower()
    if raw == "":
        return False
    return raw.endswith(".pdf") or ".pdf?" in raw or ".pdf#" in raw


def describe_mlcc_official_dimension_source(url):
    raw = clean_text(url)
    if raw == "":
        return "官方页面/规格书"
    lower = raw.lower()
    if "lcsc.com" in lower:
        return "LCSC规格书"
    if is_pdf_url(raw):
        return "官方规格书"
    return "官方页面/规格书"


def lookup_mlcc_official_dimension_fields(model, brand="", official_url="", size_hint="", allow_online_lookup=True):
    normalized_url = clean_text(official_url)
    if normalized_url == "":
        return {}

    lcsc_code = extract_lcsc_product_code(normalized_url)
    if lcsc_code != "":
        return lookup_mlcc_lcsc_dimension_fields(
            model,
            brand=brand,
            lcsc_url=normalized_url,
            size_hint=size_hint,
            allow_online_lookup=allow_online_lookup,
        )

    cache_key = f"url:{normalized_url}"
    with MLCC_OFFICIAL_DIMENSION_CACHE_LOCK:
        cache = load_mlcc_official_dimension_cache()
        if cache_key in cache and isinstance(cache.get(cache_key), dict):
            cached_fields = cache.get(cache_key, {}).copy()
            if (
                cached_fields
                and not mlcc_dimension_fields_look_suspicious(cached_fields)
                and mlcc_dimension_fields_match_size_hint(cached_fields, size_hint)
            ):
                return cached_fields

    if not allow_online_lookup:
        return {}

    fields = {}
    try:
        candidate_texts = []
        page_text = http_fetch_url_text(normalized_url)
        normalized_html = normalize_official_dimension_lookup_text(page_text)
        if normalized_html != "":
            candidate_texts.append(normalized_html)
        candidate_pdf_urls = []
        if is_pdf_url(normalized_url):
            candidate_pdf_urls.append(normalized_url)
        else:
            for candidate_url in extract_url_candidates_from_text(page_text):
                if is_pdf_url(candidate_url):
                    candidate_pdf_urls.append(candidate_url)
        seen_pdf_urls = set()
        for pdf_url in candidate_pdf_urls:
            if pdf_url in seen_pdf_urls:
                continue
            seen_pdf_urls.add(pdf_url)
            try:
                pdf_text = extract_pdf_text_from_bytes(http_fetch_url_bytes(pdf_url))
            except Exception:
                continue
            normalized_pdf_text = normalize_official_dimension_lookup_text(pdf_text)
            if normalized_pdf_text != "":
                candidate_texts.append(normalized_pdf_text)
        size_candidates = build_mlcc_size_search_tokens(model, size_hint=size_hint)
        for candidate_text in candidate_texts:
            fields = extract_dimension_fields_from_size_table(candidate_text, size_candidates, size_hint=size_hint)
            if not fields:
                fields = extract_dimension_fields_from_model_summary_line(candidate_text, model, size_hint=size_hint)
            if not fields:
                fields = extract_dimension_fields_from_text(candidate_text)
            if fields:
                break
        fields = {col: value for col, value in fields.items() if normalize_dimension_mm_value(value) != ""}
        if fields and not mlcc_dimension_fields_match_size_hint(fields, size_hint):
            fields = {}
    except Exception:
        fields = {}

    with MLCC_OFFICIAL_DIMENSION_CACHE_LOCK:
        cache = load_mlcc_official_dimension_cache()
        cache[cache_key] = fields.copy()
        try:
            save_mlcc_official_dimension_cache()
        except Exception:
            pass
    return fields.copy()


def load_samsung_mlcc_dimension_lookup():
    global SAMSUNG_MLCC_DIMENSION_LOOKUP, SAMSUNG_MLCC_DIMENSION_LOOKUP_SIGNATURE
    signature = (
        get_path_signature(SAMSUNG_MLCC_STATUS_CACHE_PATH),
        get_path_signature(SAMSUNG_MLCC_PACKAGE_CACHE_PATH),
    )
    if (
        SAMSUNG_MLCC_DIMENSION_LOOKUP is not None
        and SAMSUNG_MLCC_DIMENSION_LOOKUP_SIGNATURE == signature
    ):
        return SAMSUNG_MLCC_DIMENSION_LOOKUP

    lookup = {}
    base_lookup = {}
    status_rows = load_json_file_if_exists(SAMSUNG_MLCC_STATUS_CACHE_PATH, [])
    if isinstance(status_rows, list):
        for item in status_rows:
            if not isinstance(item, dict):
                continue
            base_model = clean_model(item.get("partNumber", ""))
            fields = build_dimension_field_map(item.get("lSize", ""), item.get("wSize", ""), item.get("tSize", ""))
            if base_model == "" or not fields:
                continue
            base_lookup[base_model] = fields
            lookup[base_model] = fields.copy()

    package_cache = load_json_file_if_exists(SAMSUNG_MLCC_PACKAGE_CACHE_PATH, {})
    if isinstance(package_cache, dict):
        for base_model, packaged_models in package_cache.items():
            base_key = clean_model(base_model)
            fields = base_lookup.get(base_key, {})
            if not fields:
                continue
            if not isinstance(packaged_models, list):
                packaged_models = [packaged_models]
            for packaged_model in packaged_models:
                packaged_key = clean_model(packaged_model)
                if packaged_key != "":
                    lookup[packaged_key] = fields.copy()

    SAMSUNG_MLCC_DIMENSION_LOOKUP = lookup
    SAMSUNG_MLCC_DIMENSION_LOOKUP_SIGNATURE = signature
    return SAMSUNG_MLCC_DIMENSION_LOOKUP


def lookup_samsung_mlcc_dimension_fields(model):
    model_key = clean_model(model)
    if model_key == "":
        return {}
    lookup = load_samsung_mlcc_dimension_lookup()
    candidates = [model_key]
    for trim in [1, 2, 3]:
        if len(model_key) > trim:
            candidates.append(model_key[:-trim])
    for candidate in candidates:
        if candidate in lookup:
            return lookup[candidate].copy()
    return {}


MLCC_DIMENSION_COLUMNS = ["长度（mm）", "宽度（mm）", "高度（mm）"]


def describe_mlcc_dimension_source(authority, brand=""):
    authority_text = clean_text(authority).lower()
    brand_text = clean_brand(brand)
    brand_upper = clean_text(brand).upper()
    if authority_text == "":
        return ""
    if "samsung" in authority_text:
        return "Samsung官方页面"
    if "murata" in authority_text:
        return "村田命名规则"
    if "tdk" in authority_text:
        return "TDK命名规则"
    if "fenghua" in authority_text:
        return "风华官方页面"
    if "walsin" in authority_text:
        return "华新科命名规则"
    if "yageo" in authority_text:
        return "国巨命名规则"
    if "kyocera" in authority_text or "avx" in authority_text:
        return "Kyocera AVX命名规则"
    if "generic_size_first" in authority_text:
        return "通用命名规则"
    if "hre" in authority_text:
        return "芯声微命名规则"
    if "pdc" in authority_text:
        return "信昌命名规则"
    if "jianghai" in authority_text:
        return "江海命名规则"
    if "三星" in brand_text or "SAMSUNG" in brand_upper:
        return "Samsung官方页面"
    if "村田" in brand_text or "MURATA" in brand_upper:
        return "村田命名规则"
    if "东电化" in brand_text or "TDK" in brand_upper:
        return "TDK命名规则"
    if "风华" in brand_text or "FENGHUA" in brand_upper:
        return "风华官方页面"
    if "华新科" in brand_text or "WALSIN" in brand_upper:
        return "华新科命名规则"
    if "国巨" in brand_text or "YAGEO" in brand_upper:
        return "国巨命名规则"
    if "晶瓷" in brand_text or "KYOCERA" in brand_upper or "AVX" in brand_upper:
        return "Kyocera AVX命名规则"
    if "芯声微" in brand_text or "HRE" in brand_upper:
        return "芯声微命名规则"
    if "信昌" in brand_text or "PDC" in brand_upper:
        return "信昌命名规则"
    if "江海" in brand_text or "JIANGHAI" in brand_upper:
        return "江海命名规则"
    return clean_text(authority)


def merge_mlcc_dimension_fields_with_source(fields, new_fields, source_labels, source_label):
    before = {col: normalize_dimension_mm_value(fields.get(col, "")) for col in MLCC_DIMENSION_COLUMNS}
    merged = merge_dimension_fields_into_record(fields, new_fields)
    after = {col: normalize_dimension_mm_value(merged.get(col, "")) for col in MLCC_DIMENSION_COLUMNS}
    if clean_text(source_label) != "" and any(before[col] != after[col] for col in MLCC_DIMENSION_COLUMNS):
        source_text = clean_text(source_label)
        if source_text not in source_labels:
            source_labels.append(source_text)
    return merged


def infer_mlcc_dimension_fields_and_source_from_record(record, allow_online_lookup=True):
    if not isinstance(record, dict):
        return {}, ""
    component_type = normalize_component_type(record.get("器件类型", ""))
    if component_type == "":
        component_type = normalize_component_type(infer_spec_component_type(record))
    if component_type != "MLCC":
        return {}, ""

    model = clean_model(record.get("型号", ""))
    if model == "":
        return {}, ""
    brand = clean_brand(record.get("品牌", ""))
    size_hint = clean_size(record.get("尺寸（inch）", ""))

    fields = {}
    source_labels = []
    parsed = parse_model_rule(model, brand=brand, component_type="MLCC")
    if isinstance(parsed, dict):
        parsed_source = describe_mlcc_dimension_source(parsed.get("_model_rule_authority", ""), brand)
        fields = merge_mlcc_dimension_fields_with_source(fields, parsed, source_labels, parsed_source)
        if size_hint == "":
            size_hint = clean_size(parsed.get("尺寸（inch）", ""))

    brand_upper = clean_text(brand).upper()
    if "SAMSUNG" in brand_upper or "三星" in brand:
        fields = merge_mlcc_dimension_fields_with_source(
            fields,
            lookup_samsung_mlcc_dimension_fields(model),
            source_labels,
            "Samsung官方页面",
        )
    elif "MURATA" in brand_upper or "村田" in brand:
        fields = merge_mlcc_dimension_fields_with_source(
            fields,
            decode_murata_dimension_fields_from_model(model),
            source_labels,
            "村田命名规则",
        )
    elif "TDK" in brand_upper or "东电化" in brand:
        fields = merge_mlcc_dimension_fields_with_source(
            fields,
            decode_tdk_dimension_fields_from_model(model),
            source_labels,
            "TDK命名规则",
        )

    needs_official_lookup = any(clean_text(fields.get(col, "")) == "" for col in MLCC_DIMENSION_COLUMNS)
    if not needs_official_lookup:
        needs_official_lookup = any(
            clean_text(fields.get(col, "")) != "" and "±" not in clean_text(fields.get(col, ""))
            for col in MLCC_DIMENSION_COLUMNS
        )
    if needs_official_lookup:
        official_url_candidates = []
        for url_col in ["官网链接", "备注3", "备注2"]:
            official_url_candidates.extend(extract_url_candidates_from_text(record.get(url_col, "")))
        for official_url in official_url_candidates:
            fields = merge_mlcc_dimension_fields_with_source(
                fields,
                lookup_mlcc_official_dimension_fields(
                    model,
                    brand=brand,
                    official_url=official_url,
                    size_hint=size_hint,
                    allow_online_lookup=allow_online_lookup,
                ),
                source_labels,
                describe_mlcc_official_dimension_source(official_url),
            )
            if all(clean_text(fields.get(col, "")) != "" for col in MLCC_DIMENSION_COLUMNS):
                break

    for note_col in ["备注3", "备注2"]:
        note_text = clean_text(record.get(note_col, ""))
        if note_text == "" or note_text.lower().startswith(("http://", "https://")):
            continue
        fields = merge_mlcc_dimension_fields_with_source(
            fields,
            extract_dimension_fields_from_text(note_text),
            source_labels,
            "备注文本",
        )

    if any(clean_text(fields.get(col, "")) == "" for col in MLCC_DIMENSION_COLUMNS):
        nominal_fields = infer_mlcc_nominal_dimension_fields_from_size(size_hint)
        fields = merge_mlcc_dimension_fields_with_source(
            fields,
            nominal_fields,
            source_labels,
            "尺寸码推断",
        )

    normalized_fields = {
        col: normalize_dimension_mm_value(fields.get(col, ""))
        for col in MLCC_DIMENSION_COLUMNS
        if normalize_dimension_mm_value(fields.get(col, "")) != ""
    }
    source_text = " / ".join(source_labels)
    if source_text == "" and normalized_fields:
        fallback_source = describe_mlcc_dimension_source(parsed.get("_model_rule_authority", ""), brand) if isinstance(parsed, dict) else ""
        source_text = clean_text(fallback_source)
    return normalized_fields, source_text

def voltage_display(x):
    x = clean_voltage(x)
    if x == "":
        return ""
    if re.search(r"[A-Za-z\u4e00-\u9fff]", x):
        return x
    return f"{x}V"

def decode_pdc_voltage_code(code):
    code = clean_text(code).upper().replace(" ", "")
    if code == "":
        return ""
    if code == "6R3":
        return "6.3"
    if len(code) == 3 and code.isdigit():
        try:
            base = int(code[:2])
            exponent = int(code[2])
            return clean_voltage(str(base * (10 ** exponent)))
        except:
            return ""
    return ""


PDC_MT_SIZE_CODE_MAP = {
    "03": "0201", "15": "0402", "18": "0603", "21": "0805",
    "31": "1206", "32": "1210", "42": "1808", "43": "1812",
    "46": "1825", "52": "2211", "55": "2220", "56": "2225",
}
PDC_MT_PACKAGE_CODE_MAP = {
    "B": "Bulk",
    "T": "Tray package",
    "E": 'Tape and 7" Reel, Embossed Tape',
    "P": 'Tape and 7" Reel, Paper Tape',
    "K": 'Tape and 10" Reel, Embossed Tape',
    "D": 'Tape and 10" Reel, Paper Tape',
    "L": 'Tape and 13" Reel, Embossed Tape',
    "G": 'Tape and 13" Reel, Paper Tape',
}
PDC_MT_THICKNESS_CODE_MAP = {
    "A": "0.60 ± 0.10",
    "B": "0.8 +0.15/-0.10",
    "C": "1.25 ± 0.10",
    "D": "1.40 ± 0.15",
    "E": "1.60 ± 0.20",
    "F": "2.00 ± 0.20",
    "G": "2.50 ± 0.30",
    "H": "2.80 ± 0.30",
    "I": "1.25 ± 0.20",
    "J": "1.15 ± 0.15",
    "K": "0.50 ± 0.20",
    "L": "0.30 ± 0.03",
    "M": "0.95 ± 0.10",
    "N": "0.50 ± 0.05",
    "O": "3.50 ± 0.20",
    "P": "1.60 +0.3/-0.10",
    "Q": "0.50 +0.02/-0.05",
    "R": "3.10 ± 0.30",
    "S": "0.80 ± 0.07",
    "T": "0.85 ± 0.10",
    "U": "0.50 ± 0.10",
    "V": "0.20 ± 0.02",
    "X": "0.80 ± 0.10",
    "Z": "0.25 ± 0.03",
}
PDC_MT_SPECIAL_CONTROL_MAP = {
    "E": "Soft Termination",
    "Z": "Anti-Arcing + Anti-Bending",
}
PDC_MT_CASE_DIMENSION_MAP = {
    "32": {"长度（mm）": "3.30±0.40", "宽度（mm）": "2.50±0.30"},
    "43": {"长度（mm）": "4.50±0.40", "宽度（mm）": "3.20±0.30"},
}
PDC_MLCC_SERIES_MEANING = {
    "FN": "常规 / General Purpose MLCC",
    "FS": "高容 / High Capacitance MLCC",
    "FM": "中压 / Medium Voltage MLCC",
    "FV": "高压 / High Voltage MLCC",
    "FP": "抗弯 / Anti-Bend General Purpose MLCC",
    "FK": "安规 / Safety Certified MLCC",
    "FH": "安规 / Safety Certified MLCC",
    "MT": "车规 / AEC-Q200",
    "MG": "次车规 / 无AEC-Q200",
    "MS": "车规 / 软端子",
}
PDC_MLCC_SERIES_CLASS = {
    "FN": "常规",
    "FS": "高容",
    "FM": "中压",
    "FV": "高压",
    "FP": "常规/抗弯",
    "FK": "安规",
    "FH": "安规",
    "MT": "车规",
    "MG": "次车规",
    "MS": "车规/软端子",
}
PDC_MLCC_SERIES_SPECIAL_USE = {
    "MT": "车规",
    "MG": "次车规",
    "MS": "车规",
}
PDC_MLCC_SERIES_PREFIXES = tuple(PDC_MLCC_SERIES_MEANING.keys())


def build_pdc_mt_dimension_fields(size_code, thickness_code):
    fields = {}
    size_dims = PDC_MT_CASE_DIMENSION_MAP.get(clean_text(size_code), {})
    if size_dims:
        fields.update(size_dims)
    thickness = clean_text(PDC_MT_THICKNESS_CODE_MAP.get(clean_text(thickness_code).upper(), ""))
    if thickness != "":
        fields["高度（mm）"] = thickness
    return fields


def pdc_mlcc_series_code_from_model(model):
    model_key = clean_model(model)
    if model_key == "":
        return ""
    for prefix in ("MS", "MG", "MT"):
        if model_key.startswith(prefix):
            return prefix
    return ""


def pdc_mlcc_series_meaning(series_code):
    return clean_text(PDC_MLCC_SERIES_MEANING.get(clean_text(series_code).upper(), ""))


def pdc_mlcc_series_class(series_code):
    return clean_text(PDC_MLCC_SERIES_CLASS.get(clean_text(series_code).upper(), ""))


def pdc_mlcc_series_profile(series_code):
    code = clean_text(series_code).upper()
    if code == "":
        return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    return {
        "系列": code,
        "系列说明": pdc_mlcc_series_meaning(code),
        "特殊用途": clean_text(PDC_MLCC_SERIES_SPECIAL_USE.get(code, "")),
        "_mlcc_series_class": pdc_mlcc_series_class(code),
    }


def pdc_general_mlcc_series_code_from_model(model):
    model_key = clean_model(model)
    if model_key == "":
        return ""
    for prefix in PDC_MLCC_SERIES_PREFIXES:
        if model_key.startswith(prefix):
            return prefix
    return ""


def pdc_mlcc_series_profile_from_model(model):
    model_key = clean_model(model)
    if model_key == "":
        return "", "", "", "", ""
    for prefix in ("MS", "MG", "MT"):
        match = re.fullmatch(
            rf"{prefix}(?P<size>\d{{2}})(?P<mat>[BCNXT])(?P<cap>(?:\d{{3,4}}|R\d+))(?P<tol>[FGJKMZ])(?P<volt>(?:6R3|\d{{3}}))(?P<rest>.*)",
            model_key,
        )
        if match is None:
            continue
        suffix = clean_text(match.group("rest"))
        package_code = suffix[0] if len(suffix) >= 1 else ""
        thickness_code = suffix[1] if len(suffix) >= 2 else ""
        special_code = suffix[2] if len(suffix) >= 3 else ""
        return prefix, pdc_mlcc_series_meaning(prefix), package_code, thickness_code, special_code
    return "", "", "", "", ""


def pdc_mlcc_special_control_meaning(code):
    return clean_text(PDC_MT_SPECIAL_CONTROL_MAP.get(clean_text(code).upper(), ""))

def cap_to_pf(value, unit):
    v = clean_text(value)
    u = clean_text(unit).upper().replace(" ", "")
    if v == "":
        return None
    try:
        num = float(v)
    except:
        return None
    if u == "PF":
        return num
    elif u == "NF":
        return num * 1000
    elif u == "UF":
        return num * 1000000
    return num

def pf_to_value_unit(pf):
    if pf is None:
        return "", ""
    if pf >= 1000000:
        value = pf / 1000000
        if float(value).is_integer():
            value = int(value)
        return str(value), "UF"
    elif pf >= 1000:
        value = pf / 1000
        if float(value).is_integer():
            value = int(value)
        return str(value), "NF"
    else:
        if float(pf).is_integer():
            pf = int(pf)
        return str(pf), "PF"


def ohm_to_value_unit(ohm_value):
    try:
        ohm = float(ohm_value)
    except:
        return "", ""
    if 0 < ohm < 1:
        value = ohm * 1000
        unit = "MILLIOHM"
    elif ohm >= 1000000:
        value = ohm / 1000000
        unit = "MOHM"
    elif ohm >= 1000:
        value = ohm / 1000
        unit = "KOHM"
    else:
        value = ohm
        unit = "OHM"
    return f"{value:.3f}".rstrip("0").rstrip("."), unit


def ohm_to_library_value_unit(ohm_value):
    value, unit = ohm_to_value_unit(ohm_value)
    display_unit_map = {
        "OHM": "Ω",
        "KOHM": "KΩ",
        "MOHM": "MΩ",
        "MILLIOHM": "mΩ",
    }
    return value, display_unit_map.get(unit, unit)


def parse_numeric_resistor_code(code):
    token = clean_text(code).upper()
    if token in {"0", "00", "000", "0000"}:
        return 0.0
    if len(token) == 3 and token.isdigit():
        return float(int(token[:2]) * (10 ** int(token[2])))
    if len(token) == 4 and token.isdigit():
        return float(int(token[:3]) * (10 ** int(token[3])))
    return None


def parse_eia96_resistor_code(code):
    token = clean_text(code).upper()
    if not re.fullmatch(r"\d{2}[A-Z]", token):
        return None
    base_index = int(token[:2])
    if not (1 <= base_index <= len(RESISTOR_EIA96_VALUES)):
        return None
    multiplier = RESISTOR_EIA96_MULTIPLIERS.get(token[2], None)
    if multiplier is None:
        return None
    return float(RESISTOR_EIA96_VALUES[base_index - 1] * multiplier)


def parse_resistor_value_code(code):
    token = clean_text(code).upper().replace("Ω", "").replace("OHMS", "").replace("OHM", "")
    if token == "":
        return None
    if token.isdigit():
        numeric = parse_numeric_resistor_code(token)
        if numeric is not None:
            return numeric
    direct = parse_resistance_token_to_ohm(token)
    if direct is not None:
        return direct
    return parse_eia96_resistor_code(token)


def match_tai_resistor_prefix(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    for prefix in TAI_RESISTOR_PREFIXES:
        if compact.startswith(prefix):
            return prefix
    return ""


def parse_tai_resistor_model(model, brand="", component_type=""):
    compact = clean_model(model)
    prefix = match_tai_resistor_prefix(compact)
    if prefix == "":
        return None
    rest = compact[len(prefix):]
    if len(rest) < 4:
        return None
    size_code = clean_text(rest[:2])
    size = clean_size(TAI_RESISTOR_SIZE_MAP.get(size_code, ""))
    tol_code = clean_text(rest[2:3]).upper()
    tol = clean_tol_for_match(RESISTOR_TOLERANCE_CODE_MAP.get(tol_code, ""))
    if size == "" or tol == "":
        return None
    packaging_code = clean_text(rest[3:4]).upper()
    tail = rest[4:] if len(rest) > 4 else ""
    resistance_ohm, value_code = extract_resistor_value_from_model(tail or rest, size_hint=size)
    if resistance_ohm is None:
        resistance_ohm, value_code = extract_resistor_value_from_model(compact, size_hint=size)
    if resistance_ohm is None:
        return None

    brand_hint = clean_brand(brand) or "大毅科技"
    fallback_component_type = "合金电阻" if prefix == "RLS" else (normalize_component_type(component_type) or "厚膜电阻")
    if prefix == "RLS":
        series_profile = dict(TAI_RLS_OFFICIAL_PROFILE)
    else:
        series_profile = infer_resistor_series_profile(
            compact,
            brand=brand_hint,
            component_type=fallback_component_type,
        )
    resolved_component_type = normalize_component_type(series_profile.get("器件类型", "")) or fallback_component_type
    resolved_special_use = clean_text(series_profile.get("特殊用途", ""))
    power_text = RESISTOR_POWER_BY_SIZE.get(size, "")

    return {
        "品牌": clean_brand(brand),
        "型号": compact,
        "器件类型": resolved_component_type,
        "系列": clean_text(series_profile.get("系列", "")) or prefix,
        "系列说明": clean_text(series_profile.get("系列说明", "")),
        "特殊用途": resolved_special_use,
        "尺寸（inch）": size,
        "容值误差": tol,
        "_resistance_ohm": resistance_ohm,
        "_power": power_text,
        "_model_rule_authority": "tai_resistor_model",
        "_value_code": value_code,
        "_packaging_code": packaging_code,
        "_param_count": sum([1 if size else 0, 1 if tol else 0, 1 if resistance_ohm is not None else 0]),
    }


def extract_resistor_value_from_model(model, size_hint=""):
    compact = clean_model(model)
    if compact == "":
        return None, ""
    size_hint = clean_size(size_hint)
    segments = [seg for seg in re.split(r"[-_/]", compact) if seg]
    if not segments:
        segments = [compact]

    best = (None, "", -1, -1)
    candidate_pattern = re.compile(r"R\d+(?:\d+)?|\d+[RKM]\d*|\d{4}|\d{3}|\d{2}[A-Z]")
    size_tokens = {size_hint} if size_hint else set()
    size_tokens |= {"01005", "0201", "0401", "0402", "0603", "0805", "1206", "1210", "1812", "2010", "2512", "3225"}

    for segment in segments:
        for match in candidate_pattern.finditer(segment):
            token = match.group(0)
            if token in size_tokens:
                continue
            ohm_value = parse_resistor_value_code(token)
            if ohm_value is None:
                continue
            score = 1
            if any(ch in token for ch in "RKM"):
                score += 4
            if token.isdigit():
                score += 2
            if match.end() == len(segment) or re.fullmatch(r"[A-Z0-9]{0,3}", segment[match.end():]):
                score += 1
            if score > best[2] or (score == best[2] and match.start() > best[3]):
                best = (ohm_value, token, score, match.start())
    return best[0], best[1]


def infer_resistor_size_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    tai_prefix = match_tai_resistor_prefix(compact)
    if tai_prefix != "":
        size_code = compact[len(tai_prefix):len(tai_prefix) + 2]
        mapped = clean_size(TAI_RESISTOR_SIZE_MAP.get(size_code, ""))
        if mapped != "":
            return mapped
    size = find_embedded_size(compact)
    if size != "":
        return size
    prefix_maps = [
        (("WR", "WF", "MR"), WALSIN_RESISTOR_SIZE_MAP),
        (("CR", "TR", "QR"), EVER_OHMS_RESISTOR_SIZE_MAP),
        (("CQ", "NQ", "LE", "TC"), UNIROYAL_RESISTOR_SIZE_MAP),
    ]
    for prefixes, size_map in prefix_maps:
        prefix = next((p for p in prefixes if compact.startswith(p)), None)
        if prefix is None:
            continue
        size_code = compact[len(prefix):len(prefix) + 2]
        mapped = clean_size(size_map.get(size_code, ""))
        if mapped != "":
            return mapped
    return ""


def infer_resistor_tolerance_from_model(model, size_hint=""):
    compact = clean_model(model)
    if compact == "":
        return ""
    tai_prefix = match_tai_resistor_prefix(compact)
    if tai_prefix != "":
        tol_idx = len(tai_prefix) + 2
        if tol_idx < len(compact):
            tol = clean_tol_for_match(RESISTOR_TOLERANCE_CODE_MAP.get(compact[tol_idx], ""))
            if tol != "":
                return tol
    if len(compact) >= 7 and compact[2:6].isdigit() and compact[6] in RESISTOR_TOLERANCE_CODE_MAP:
        return clean_tol_for_match(RESISTOR_TOLERANCE_CODE_MAP[compact[6]])
    value_ohm, value_code = extract_resistor_value_from_model(compact, size_hint=size_hint)
    if value_ohm is None or value_code == "":
        return ""
    segment = clean_model(compact.rsplit("-", 1)[0])
    if value_code in compact:
        idx = compact.find(value_code)
        if idx > 0:
            prev = compact[idx - 1]
            if prev in RESISTOR_TOLERANCE_CODE_MAP:
                return clean_tol_for_match(RESISTOR_TOLERANCE_CODE_MAP[prev])
    if segment and segment[-1] in RESISTOR_TOLERANCE_CODE_MAP:
        return clean_tol_for_match(RESISTOR_TOLERANCE_CODE_MAP[segment[-1]])
    return ""


def resistor_model_rule_candidate(model, brand="", component_type=""):
    compact = clean_model(model)
    if compact == "":
        return False
    normalized_type = normalize_component_type(component_type)
    if normalized_type != "" and normalized_type not in ALL_RESISTOR_TYPES:
        return False
    if normalized_type in ALL_RESISTOR_TYPES:
        return True
    if RESISTOR_MODEL_PREFIX_PATTERN.match(compact):
        return True
    upper_brand = clean_brand(brand).upper()
    return upper_brand != "" and any(token in upper_brand for token in ["YAGEO", "MURATA", "WALSIN", "UNI-ROYAL", "EVER OHMS", "RALEC"])


def parse_yageo_chip_resistor_model(model, brand="", component_type=""):
    compact = clean_model(model)
    if not re.match(r"^(AA|AC|AF|AR|AT|RC|RT)\d{4}[A-Z]", compact):
        return None
    series_profile = infer_resistor_series_profile(compact, brand=brand, component_type=normalize_component_type(component_type) or "厚膜电阻")
    resolved_component_type = normalize_component_type(series_profile.get("器件类型", "")) or normalize_component_type(component_type) or "厚膜电阻"
    resolved_special_use = clean_text(series_profile.get("特殊用途", ""))
    size = clean_size(compact[2:6])
    tol = clean_tol_for_match(RESISTOR_TOLERANCE_CODE_MAP.get(compact[6], ""))
    right_segment = compact.rsplit("-", 1)[-1] if "-" in compact else compact[7:]
    resistance_ohm, value_code = extract_resistor_value_from_model(right_segment[2:], size_hint=size)
    if resistance_ohm is None:
        resistance_ohm, value_code = extract_resistor_value_from_model(compact, size_hint=size)
    if size == "" and resistance_ohm is None and tol == "":
        return None
    return {
        "品牌": clean_brand(brand),
        "型号": compact,
        "器件类型": resolved_component_type,
        "系列": series_profile["系列"],
        "系列说明": series_profile["系列说明"],
        "特殊用途": resolved_special_use,
        "尺寸（inch）": size,
        "容值误差": tol,
        "_resistance_ohm": resistance_ohm,
        "_model_rule_authority": "yageo_chip_resistor_model",
        "_value_code": value_code,
        "_param_count": sum([1 if size else 0, 1 if tol else 0, 1 if resistance_ohm is not None else 0]),
    }


def parse_murata_mhr_resistor_model(model, brand="", component_type=""):
    compact = clean_model(model)
    match = re.fullmatch(r"MHR(\d{4})SA(\d{3})([BCDFGJKM])([A-Z0-9]{2,})", compact)
    if not match:
        return None
    series_profile = infer_resistor_series_profile(compact, brand=brand, component_type=normalize_component_type(component_type) or "厚膜电阻")
    resolved_component_type = normalize_component_type(series_profile.get("器件类型", "")) or normalize_component_type(component_type) or "厚膜电阻"
    resolved_special_use = clean_text(series_profile.get("特殊用途", ""))
    resistance_ohm = parse_resistor_value_code(match.group(2))
    tol = clean_tol_for_match(RESISTOR_TOLERANCE_CODE_MAP.get(match.group(3), ""))
    if resistance_ohm is None:
        return None
    return {
        "品牌": clean_brand(brand),
        "型号": compact,
        "器件类型": resolved_component_type,
        "系列": series_profile["系列"],
        "系列说明": series_profile["系列说明"],
        "特殊用途": resolved_special_use,
        "容值误差": tol,
        "_resistance_ohm": resistance_ohm,
        "_model_rule_authority": "murata_mhr_model",
        "_value_code": match.group(2),
        "_param_count": sum([1 if tol else 0, 1 if resistance_ohm is not None else 0]),
    }


def parse_generic_resistor_model(model, brand="", component_type=""):
    if not resistor_model_rule_candidate(model, brand=brand, component_type=component_type):
        return None
    compact = clean_model(model)
    series_profile = infer_resistor_series_profile(compact, brand=brand, component_type=normalize_component_type(component_type) or "厚膜电阻")
    resolved_component_type = normalize_component_type(series_profile.get("器件类型", "")) or normalize_component_type(component_type)
    resolved_special_use = clean_text(series_profile.get("特殊用途", ""))
    size = infer_resistor_size_from_model(compact)
    tol = infer_resistor_tolerance_from_model(compact, size_hint=size)
    resistance_ohm, value_code = extract_resistor_value_from_model(compact, size_hint=size)
    if resistance_ohm is None:
        return None
    return {
        "品牌": clean_brand(brand),
        "型号": compact,
        "器件类型": resolved_component_type,
        "系列": series_profile["系列"],
        "系列说明": series_profile["系列说明"],
        "特殊用途": resolved_special_use,
        "尺寸（inch）": size,
        "容值误差": tol,
        "_resistance_ohm": resistance_ohm,
        "_model_rule_authority": "generic_resistor_model",
        "_value_code": value_code,
        "_param_count": sum([1 if size else 0, 1 if tol else 0, 1 if resistance_ohm is not None else 0]),
    }


def parse_resistor_model_rule(model, brand="", component_type=""):
    for parser in (parse_murata_mhr_resistor_model, parse_yageo_chip_resistor_model, parse_tai_resistor_model, parse_generic_resistor_model):
        parsed = parser(model, brand=brand, component_type=component_type)
        if parsed is not None:
            return parsed
    return None


def merge_parsed_rule_into_record(record, parsed_rule, override_conflicts=False):
    if not parsed_rule:
        return dict(record)
    merged = dict(record)
    if "_model_rule_authority" not in merged:
        merged["_model_rule_authority"] = ""

    parsed_type = normalize_component_type(parsed_rule.get("器件类型", ""))
    if parsed_type == "":
        parsed_type = normalize_component_type(infer_spec_component_type(parsed_rule))
    current_type = normalize_component_type(merged.get("器件类型", ""))
    if parsed_type != "" and (override_conflicts or current_type == ""):
        merged["器件类型"] = parsed_type

    parsed_series = clean_text(parsed_rule.get("系列", ""))
    current_series = clean_text(merged.get("系列", ""))
    if parsed_series != "" and (override_conflicts or current_series == ""):
        merged["系列"] = parsed_series
    parsed_series_desc = clean_text(parsed_rule.get("系列说明", ""))
    current_series_desc = clean_text(merged.get("系列说明", ""))
    if parsed_series_desc != "" and (override_conflicts or current_series_desc == ""):
        merged["系列说明"] = parsed_series_desc

    for col, cleaner in [
        ("尺寸（inch）", clean_size),
        ("尺寸（mm）", clean_text),
        ("长度（mm）", clean_text),
        ("宽度（mm）", clean_text),
        ("高度（mm）", clean_text),
        ("材质（介质）", clean_material),
        ("容值误差", clean_tol_for_match),
        ("耐压（V）", clean_voltage),
        ("工作温度", normalize_working_temperature_text),
        ("寿命（h）", normalize_life_hours_value),
        ("安装方式", normalize_mounting_style),
        ("特殊用途", normalize_special_use),
        ("封装代码", clean_text),
        ("规格摘要", clean_text),
        ("容值", clean_text),
        ("容值单位", lambda x: clean_text(x).upper()),
        ("功率", format_power_display),
        ("脚距", clean_text),
        ("安规", clean_text),
        ("规格", clean_text),
        ("压敏电压", voltage_display),
    ]:
        parsed_value = cleaner(parsed_rule.get(col, ""))
        current_value = cleaner(merged.get(col, ""))
        if parsed_value != "" and (override_conflicts or current_value == ""):
            merged[col] = parsed_value

    parsed_pf = parsed_rule.get("容值_pf", None)
    if parsed_pf is not None and (override_conflicts or merged.get("容值_pf", None) is None):
        merged["容值_pf"] = parsed_pf
        value, unit = pf_to_value_unit(parsed_pf)
        merged["容值"] = value
        merged["容值单位"] = unit

    parsed_resistance = parsed_rule.get("_resistance_ohm", None)
    if parsed_resistance is not None:
        merged["_resistance_ohm"] = parsed_resistance
        value, unit = ohm_to_library_value_unit(parsed_resistance)
        merged["容值"] = value
        merged["容值单位"] = unit

    for col, cleaner in [
        ("_power", clean_text),
        ("_pitch", clean_text),
        ("_safety_class", clean_text),
        ("_body_size", clean_text),
        ("_disc_size", clean_text),
        ("_varistor_voltage", clean_voltage),
    ]:
        parsed_value = cleaner(parsed_rule.get(col, ""))
        current_value = cleaner(merged.get(col, ""))
        if parsed_value != "" and (override_conflicts or current_value == ""):
            merged[col] = parsed_value

    authority = clean_text(parsed_rule.get("_model_rule_authority", ""))
    if authority != "":
        merged["_model_rule_authority"] = authority
    return merged


def apply_model_rule_overrides_to_dataframe(df, override_conflicts=False):
    if df is None or df.empty or "型号" not in df.columns:
        return df
    work = df.copy()
    if "_model_rule_authority" not in work.columns:
        work["_model_rule_authority"] = ""
    for col in [
        "器件类型", "系列", "尺寸（inch）", "尺寸（mm）", "材质（介质）", "容值", "容值单位",
        "容值误差", "耐压（V）", "_model_rule_authority", "_resistance_ohm",
        "工作温度", "寿命（h）", "安装方式", "特殊用途", "封装代码", "规格摘要",
        "功率", "脚距", "安规", "压敏电压",
    ]:
        if col in work.columns and isinstance(work[col].dtype, pd.CategoricalDtype):
            work[col] = work[col].astype("object")

    model_series = work["型号"].astype(str).apply(clean_model)
    jianghai_mask = model_series.apply(lambda value: jianghai_series_code_from_model(value) != "")
    candidate_mask = model_series.str.match(RESISTOR_MODEL_PREFIX_PATTERN, na=False) | jianghai_mask
    candidate_idx = work[candidate_mask].index.tolist()
    if not candidate_idx:
        return work

    for idx in candidate_idx:
        if clean_text(work.at[idx, "_model_rule_authority"]) != "" and not clean_text(work.at[idx, "_model_rule_authority"]).startswith("manual"):
            continue
        row_model = work.at[idx, "型号"]
        row_brand = work.at[idx, "品牌"] if "品牌" in work.columns else ""
        row_brand_text = clean_brand(row_brand)
        row_is_jianghai = (
            "JIANGHAI" in clean_text(row_brand_text).upper()
            or "江海" in row_brand_text
            or jianghai_series_code_from_model(row_model) != ""
        )
        parsed_rule = parse_model_rule(
            row_model,
            brand=row_brand,
            component_type=work.at[idx, "器件类型"] if "器件类型" in work.columns else "",
        )
        if parsed_rule is None:
            continue
        merged = merge_parsed_rule_into_record(
            work.loc[idx].to_dict(),
            parsed_rule,
            override_conflicts=(override_conflicts and not row_is_jianghai),
        )
        for col, value in merged.items():
            work.at[idx, col] = value
    return work


SERIES_PLACEHOLDER_TOKENS = {"", "??", "？", "NONE", "NAN", "常规", "MLCC"}


def series_looks_missing(value):
    text = clean_text(value)
    if text == "":
        return True
    if text in SERIES_PLACEHOLDER_TOKENS:
        return True
    if text.upper() in SERIES_PLACEHOLDER_TOKENS:
        return True
    return bool(re.fullmatch(r"[?？]+", text))


def resistor_series_should_replace(current_value, canonical_value):
    current = clean_text(current_value)
    canonical = clean_text(canonical_value)
    if canonical == "":
        return False
    if series_looks_missing(current):
        return True
    if current == canonical:
        return False
    current_upper = current.upper()
    canonical_upper = canonical.upper()
    if canonical == "普通厚膜" and re.fullmatch(r"\d{4,5}[A-Z0-9]{2,}", current_upper):
        return True
    return current_upper.startswith(canonical_upper) and len(current_upper) > len(canonical_upper)


def resistor_series_desc_should_replace(current_value, canonical_value):
    current = clean_text(current_value)
    canonical = clean_text(canonical_value)
    if canonical == "":
        return False
    if current == "":
        return True
    if current == canonical:
        return False
    return current.endswith("电阻系列")


def mlcc_series_should_replace(current_value, canonical_value):
    current = clean_text(current_value)
    canonical = clean_text(canonical_value)
    if canonical == "":
        return False
    if series_looks_missing(current):
        return True
    if current == canonical:
        return False
    current_upper = current.upper()
    canonical_upper = canonical.upper()
    contaminated_patterns = [
        r"^C\d{4}$",
        r"^(?:CC|CQ|AC|AQ|AS|CS)\d{4}$",
        r"^(?:RF|HH|SH|RT|UF)\d{2}$",
        r"^CL\d{2}[A-Z]$",
        r"^(?:TMK|JMK|EMK|LMK|AMK|UMK|HMK|QMK|SMK|QVS|TVS)\d{2,4}$",
        r"^(?:MAAS|MBAS|MCAS|MCAST|MLAS|MMAS|MSAS)\d{2,3}$",
        r"^TCC\d{4}$",
        r"^(?:CGA|CSA)[A-Z0-9]{2,6}$",
        r"^(?:RHEL|RPER|ERB|RCE|RDE|RHE|RPE)[A-Z0-9]{2,}$",
    ]
    current_is_contaminated = any(re.fullmatch(pattern, current_upper) for pattern in contaminated_patterns)
    canonical_is_contaminated = any(re.fullmatch(pattern, canonical_upper) for pattern in contaminated_patterns)
    if current_is_contaminated and not canonical_is_contaminated:
        return True
    if current_upper in {"MS", "MBA"} and canonical_upper.startswith(current_upper):
        return True
    if current_upper == "AM" and canonical_upper == "AMK":
        return True
    return False


MLCC_GENERIC_SIZE_FIRST_EXCLUDED_BRAND_TOKENS = (
    "TDK",
    "东电化",
    "MURATA",
    "村田",
    "TAIYO",
    "太阳诱电",
    "YAGEO",
    "国巨",
    "KYOCERA",
    "AVX",
    "晶瓷",
    "PDC",
    "信昌",
    "WALSIN",
    "华新科",
    "FENGHUA",
    "风华",
    "JIANGHAI",
    "江海",
    "SAMSUNG",
    "三星",
    "CCTC",
    "三环",
)


def mlcc_generic_size_first_parser_allowed(brand):
    brand_text = clean_brand(brand)
    brand_upper = clean_text(brand_text).upper()
    if brand_text == "":
        return True
    if "HRE" in brand_upper or "芯声微" in brand_text:
        return True
    return not any(token in brand_upper or token in brand_text for token in MLCC_GENERIC_SIZE_FIRST_EXCLUDED_BRAND_TOKENS)


MURATA_SERIES_PREFIX_PATTERN = re.compile(
    r"^(AVR-M|AVRM|AVRL|AVRH|VAR|S\d{2}K|LS\d{2}K|CN\d{4}|SGN[A-Z0-9]{4}|"
    r"GRM|GCM|GCJ|GJM|GQM|GRT|GCG|GCQ|GRJ|GMA|GMD|GCH|GXT|GGM|GC3|GCD|GCE|GGD|"
    r"LLL|LLF|LLA|LLG|LLC|LLM|LLR|NFM|KCM|KRT|DK1|GA2|GA3|GR3|GR4|GR7|GJ4|KRM|KR3|KR9|"
    r"ZRA|ZRB|Z62|Z63|NTC|PRF|PTG|NXF|CEU|CGJ|CGB|CKG|CNA|CNC|CN0|"
    r"RHEL|RPER|ERB|RCE|RDE|RHE|RHS|RPE|DEH|DEJ|DHR|"
    r"CLLC|CLLE|CLLG|NCP\d{2}[A-Z]{2}|NCU\d{2}[A-Z]{2}|NCG\d{2}[A-Z]{2}|FTN\d{2}[A-Z]{2}|YNA)"
)

MURATA_SERIES_MEANING = {
    "GRM": "常规 / General-purpose MLCC",
    "GCM": "车规 / Automotive MLCC",
    "GRT": "车规 / Automotive MLCC",
    "GCJ": "Murata MLCC series",
    "GCG": "Murata MLCC series",
    "GCQ": "Murata MLCC series",
    "GJM": "高Q / High-Q MLCC",
    "GQM": "高Q / High-Q MLCC",
    "GRJ": "Murata MLCC series",
    "GMA": "Murata MLCC series",
    "GMD": "Murata MLCC series",
    "GCH": "Murata MLCC series",
    "GXT": "Murata MLCC series",
    "GGM": "Murata MLCC series",
    "GC3": "Murata MLCC series",
    "GCD": "Murata MLCC series",
    "GCE": "Murata MLCC series",
    "GGD": "Murata MLCC series",
    "LLL": "Murata MLCC series",
    "LLF": "Murata MLCC series",
    "LLA": "Murata MLCC series",
    "LLG": "Murata MLCC series",
    "LLC": "Murata MLCC series",
    "LLM": "常规低ESL / 10 terminals low ESL MLCC for General Purpose",
    "LLR": "常规低ESL控ESR / LW reversed controlled ESR low ESL MLCC for General Purpose",
    "NFM": "EMI滤波 / Murata EMI filter series",
    "KCM": "Murata MLCC series",
    "KRT": "Murata MLCC series",
    "DK1": "Murata MLCC series",
    "GA2": "Murata MLCC series",
    "GA3": "Murata MLCC series",
    "GR3": "Murata MLCC series",
    "GR4": "Murata MLCC series",
    "GR7": "Murata MLCC series",
    "GJ4": "Murata MLCC series",
    "KRM": "Murata MLCC series",
    "KR3": "Murata MLCC series",
    "KR9": "Murata MLCC series",
    "ZRA": "Murata MLCC series",
    "ZRB": "Murata MLCC series",
    "AVR-M": "Murata varistor series",
    "AVRM": "Murata varistor series",
    "AVRL": "Murata varistor series",
    "AVRH": "Murata varistor series",
    "VAR": "Murata varistor series",
    "S20K": "Murata varistor series",
    "LS40K": "Murata varistor series",
    "Z62": "Murata varistor series",
    "Z63": "Murata varistor series",
    "NTC": "Murata thermistor series",
    "PRF": "Murata thermistor series",
    "PTG": "Murata thermistor series",
    "NXF": "Murata thermistor series",
    "CEU": "Murata thermistor / sensor series",
    "CGJ": "Murata MLCC series",
    "CGB": "Murata MLCC series",
    "CKG": "Murata MLCC series",
    "CNA": "Murata MLCC series",
    "CNC": "Murata MLCC series",
    "CN0": "Murata capacitor / varistor series",
    "ERB": "Murata MLCC series",
    "RCE": "车规 / Automotive MLCC",
    "RDE": "常规 / General-purpose MLCC",
    "RHE": "车规 / Automotive MLCC",
    "RHS": "高温车规引线型 / High-temperature leaded automotive MLCC",
    "RPE": "Murata MLCC series",
    "DEH": "高压 / High Voltage (High Temperature Guaranteed, Low-dissipation Factor (Char. R, C))",
    "DEJ": "高压 / High Voltage (High Temperature Guaranteed, Low-dissipation Factor (Char. D))",
    "DHR": "超高压 / Ultrahigh Voltage",
    "RHEL": "Murata MLCC series",
    "RPER": "Murata MLCC series",
    "CLLC": "Murata capacitor series",
    "CLLE": "Murata capacitor series",
    "CLLG": "Murata capacitor series",
    "NCP15XQ": "Murata thermistor series",
    "NCP18WB": "Murata thermistor series",
    "NCU15WB": "Murata thermistor series",
    "NCU15WF": "Murata thermistor series",
    "NCU18WF": "Murata thermistor series",
    "YNA": "Murata thermistor series",
}


def murata_series_code_from_model(model):
    compact = clean_model(model)
    if not compact:
        return ""
    match = MURATA_SERIES_PREFIX_PATTERN.match(compact)
    if not match:
        return ""
    return match.group(1)


def murata_series_meaning(series_code):
    code = clean_text(series_code)
    if not code:
        return "Murata 官方系列代码"
    if code.startswith(("FTN", "NCP", "NCU", "NCG")):
        return "Murata NTC thermistor series"
    return MURATA_SERIES_MEANING.get(code, "Murata 官方系列代码")


MURATA_NTC_MODEL_PATTERN = re.compile(r"^(?P<series>(?:FTN|NCP|NCU|NCG)\d{2}[A-Z]{2})(?P<res>\d{3})(?P<tol>[BCDFGJKMZ])(?P<tail>.*)$")
MURATA_NTC_BODY_SIZE_MAP = {
    "02": "0402",
    "03": "0603",
    "15": "1005",
    "18": "1608",
    "21": "2012",
    "31": "3216",
    "32": "3225",
    "42": "4532",
    "43": "5750",
}
MURATA_NTC_INCH_SIZE_MAP = {
    "02": "01005",
    "03": "0201",
    "15": "0402",
    "18": "0603",
    "21": "0805",
    "31": "1206",
    "32": "1210",
    "42": "1812",
    "43": "2220",
}

def format_murata_ntc_size_label(size_code):
    body_code = clean_text(MURATA_NTC_BODY_SIZE_MAP.get(size_code, ""))
    inch_code = clean_text(MURATA_NTC_INCH_SIZE_MAP.get(size_code, ""))
    if body_code and inch_code:
        return f"{body_code}({inch_code})"
    return body_code or inch_code


def murata_ntc_series_code_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    match = MURATA_NTC_MODEL_PATTERN.match(compact)
    if not match:
        return ""
    return match.group("series")


def murata_ntc_series_profile_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return {
            "系列": "",
            "系列说明": "",
            "特殊用途": "",
            "尺寸（inch）": "",
            "尺寸（mm）": "",
            "长度（mm）": "",
            "宽度（mm）": "",
            "_ntc_resistance_code": "",
            "_ntc_tolerance_code": "",
            "_ntc_tail": "",
        }
    match = MURATA_NTC_MODEL_PATTERN.match(compact)
    if not match:
        return {
            "系列": "",
            "系列说明": "",
            "特殊用途": "",
            "尺寸（inch）": "",
            "尺寸（mm）": "",
            "长度（mm）": "",
            "宽度（mm）": "",
            "_ntc_resistance_code": "",
            "_ntc_tolerance_code": "",
            "_ntc_tail": "",
        }

    size_code = compact[3:5]
    size_inch = MURATA_NTC_INCH_SIZE_MAP.get(size_code, "")
    mm_dims = MURATA_SIZE_DIMENSION_MAP.get(size_code, ("", ""))
    length_mm, width_mm = mm_dims if len(mm_dims) == 2 else ("", "")
    size_mm = f"{length_mm}×{width_mm}" if length_mm and width_mm else ""
    return {
        "系列": match.group("series"),
        "系列说明": "Murata NTC thermistor series",
        "特殊用途": "",
        "尺寸（inch）": size_inch,
        "尺寸（mm）": size_mm,
        "长度（mm）": length_mm,
        "宽度（mm）": width_mm,
        "封装代码": size_inch,
        "_ntc_resistance_code": match.group("res"),
        "_ntc_tolerance_code": match.group("tol"),
        "_ntc_tail": match.group("tail"),
    }


def parse_murata_ntc_core(model, allow_partial=False):
    compact = clean_model(model)
    profile = murata_ntc_series_profile_from_model(compact)
    if clean_text(profile.get("系列", "")) == "":
        return None

    match = MURATA_NTC_MODEL_PATTERN.match(compact)
    if match is None:
        return None

    size_code = compact[3:5]
    size_label = format_murata_ntc_size_label(size_code)
    size_inch = clean_text(profile.get("尺寸（inch）", ""))
    size_mm = clean_text(profile.get("尺寸（mm）", ""))
    length_mm = clean_text(profile.get("长度（mm）", ""))
    width_mm = clean_text(profile.get("宽度（mm）", ""))
    resistance_ohm = parse_resistor_value_code(match.group("res"))
    if resistance_ohm is None:
        extracted_ohm, _ = extract_resistor_value_from_model(compact, size_hint=size_inch)
        resistance_ohm = extracted_ohm
    tol_code = clean_text(match.group("tol"))
    tol_value = {
        "B": "0.1",
        "C": "0.25",
        "D": "0.5",
        "E": "3",
        "F": "1",
        "G": "2",
        "J": "5",
        "K": "10",
        "M": "20",
        "Z": "+80/-20",
    }.get(tol_code, "")
    tol = clean_tol_for_match(f"±{tol_value}%") if tol_value and tol_value not in {"+80/-20"} else clean_text(tol_value)
    resistance_value, resistance_unit = ohm_to_library_value_unit(resistance_ohm) if resistance_ohm is not None else ("", "")
    summary = build_other_component_summary([
        f"{resistance_value}{resistance_unit}" if resistance_value and resistance_unit else "",
        clean_tol_for_display(tol) if clean_text(tol) != "" else "",
        size_label,
        size_mm,
    ])
    param_count = sum([
        1 if size_inch else 0,
        1 if resistance_ohm is not None else 0,
        1 if clean_text(tol) != "" else 0,
    ])
    if allow_partial and param_count < 3:
        return None
    return {
        "品牌": "村田Murata",
        "型号": compact,
        "器件类型": "热敏电阻",
        "系列": profile["系列"],
        "系列说明": profile["系列说明"],
        "安装方式": "贴片",
        "封装代码": size_inch,
        "尺寸（inch）": size_inch,
        "尺寸（mm）": size_mm,
        "长度（mm）": length_mm,
        "宽度（mm）": width_mm,
        "规格摘要": summary,
        "容值": resistance_value,
        "容值单位": resistance_unit,
        "容值误差": tol,
        "特殊用途": profile["特殊用途"],
        "_ntc_resistance_code": match.group("res"),
        "_ntc_tolerance_code": tol_code,
        "_ntc_tail": match.group("tail"),
        "_model_rule_authority": "murata_ntc_series",
        "_param_count": param_count,
        "_partial_part": True,
    }


def parse_murata_ntc_common(model):
    parsed = parse_murata_ntc_core(model, allow_partial=False)
    if parsed is None:
        return None
    parsed.pop("_param_count", None)
    parsed.pop("_partial_part", None)
    return parsed


MURATA_MLCC_SERIES_CLASS = {
    "GRM": "常规",
    "GCM": "车规",
    "GRT": "车规",
    "GJM": "高Q",
    "GQM": "高Q",
    "NFM": "EMI滤波",
    "RCE": "车规",
    "RDE": "常规",
    "RHE": "车规",
    "RHS": "车规",
    "LLM": "常规",
    "LLR": "常规",
    "DEH": "高压",
    "DEJ": "高压",
    "DHR": "高压",
}


def murata_series_profile(series_code):
    code = clean_text(series_code)
    if code == "":
        return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    class_text = clean_text(MURATA_MLCC_SERIES_CLASS.get(code, ""))
    special_use = "车规" if class_text == "车规" else ""
    return {
        "系列": code,
        "系列说明": murata_series_meaning(code),
        "特殊用途": special_use,
        "_mlcc_series_class": class_text,
    }


def fenghua_am_series_code_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    return "AM" if FENGHUA_AM_MODEL_PATTERN.fullmatch(compact) else ""


def fenghua_am_series_meaning(series_code):
    code = clean_text(series_code).upper()
    if code == "":
        return "风华 AM 汽车级 MLCC 系列"
    return FENGHUA_AM_SERIES_MEANING.get(code, "风华 AM 汽车级 MLCC 系列")


def fenghua_am_dimension_fields_from_model(model):
    return decode_fenghua_am_dimension_fields_from_model(model)


TDK_MLCC_SERIES_MEANING = {
    "C": "常规 / General-purpose MLCC",
    "CGA": "车规 / AEC-Q200",
}
TDK_MLCC_SERIES_CLASS = {
    "C": "常规",
    "CGA": "车规",
}
YAGEO_MLCC_SERIES_MEANING = {
    "CC": "常规 / General-purpose MLCC",
    "CQ": "高Q / High-Q MLCC",
    "AC": "车规 / Automotive MLCC",
    "AQ": "车规高Q / Automotive High-Q MLCC",
    "AS": "车规软端子 / Automotive Soft Termination MLCC",
    "CS": "软端子 / Soft Termination MLCC",
}
YAGEO_MLCC_SERIES_CLASS = {
    "CC": "常规",
    "CQ": "高Q",
    "AC": "车规",
    "AQ": "车规/高Q",
    "AS": "车规/软端子",
    "CS": "软端子",
}
WALSIN_MLCC_SERIES_MEANING = {
    "01R5": "超小型 / Ultra-small MLCC",
    "RF": "微波 / Ultra High Q / Ultra Low ESR MLCC",
    "HH": "高Q / Low ESR MLCC",
    "SH": "软端子 / Soft Termination MLCC",
    "RT": "车规高Q / AEC-Q200 Qualified Automotive High-Q MLCC",
    "UF": "微波窄容差 / Narrow Tolerance Ultra High Q MLCC",
}
WALSIN_MLCC_SERIES_CLASS = {
    "RF": "高Q",
    "HH": "高Q",
    "SH": "软端子",
    "RT": "车规/高Q",
    "UF": "高Q",
}
WALSIN_NUMERIC_MLCC_DIELECTRIC = {
    "A": "",
    "B": "X7R",
    "F": "Y5V",
    "N": "COG(NPO)",
    "S": "X6S",
    "T": "X7T",
    "X": "X5R",
}
TAIYO_MLCC_SERIES_MEANING = {
    "TMK": "Taiyo Yuden MLCC official series",
    "JMK": "Taiyo Yuden MLCC official series",
    "EMK": "Taiyo Yuden MLCC official series",
    "LMK": "Taiyo Yuden MLCC official series",
    "AMK": "Taiyo Yuden MLCC official series",
    "UMK": "Taiyo Yuden MLCC official series",
    "HMK": "Taiyo Yuden MLCC official series",
    "QMK": "Taiyo Yuden MLCC official series",
    "SMK": "Taiyo Yuden MLCC official series",
    "QVS": "Taiyo Yuden MLCC official series",
    "TVS": "Taiyo Yuden MLCC official series",
    "MSAY": "压电噪音抑制 / Piezoelectric Noise Suppression MLCC for General Electronic Equipment",
    "MSRL": "低电感焊线连接 / Low Inductance Wire-Bonding MLCC for General Electronic Equipment",
    "MSART": "高Q / High-Q MLCC for General Electronic Equipment",
    "MBARQ": "高Q / High-Q MLCC for Infrastructure & Industrial Equipment",
    "MCARQ": "车规高Q / Automotive High-Q MLCC",
    "MMARQ": "高Q / High-Q MLCC for Medical Devices Class C",
    "MAAS": "车规 / Automotive Powertrain & Safety MLCC",
    "MBAS": "工业通信 / Infrastructure & Industrial Equipment MLCC",
    "MCAS": "车规 / Automotive Body & Chassis / Infotainment MLCC",
    "MCAST": "车规 / Automotive Body & Chassis / Infotainment MLCC",
    "MLAS": "医疗 / Medical Devices Class A/B MLCC",
    "MMAS": "医疗 / Medical Devices Class C MLCC",
    "MSAS": "常规 / General Electronic Equipment MLCC",
    "UP": "Taiyo Yuden UP MLCC official series",
    "TP": "Taiyo Yuden TP MLCC official series",
    "JMR": "Taiyo Yuden JMR MLCC official series",
    "LMR": "Taiyo Yuden LMR MLCC official series",
    "HMR": "Taiyo Yuden HMR MLCC official series",
    "EMR": "Taiyo Yuden EMR MLCC official series",
    "MEASL": "Taiyo Yuden MEASL MLCC official series",
    "MEAST": "Taiyo Yuden MEAST MLCC official series",
    "MEASJ": "Taiyo Yuden MEASJ MLCC official series",
}
TAIYO_MLCC_SERIES_CLASS = {
    "MSART": "高Q/常规",
    "MBARQ": "高Q",
    "MCARQ": "车规/高Q",
    "MMARQ": "高Q",
    "MAAS": "车规",
    "MCAS": "车规",
    "MCAST": "车规",
    "MSAS": "常规",
    "UP": "常规",
    "TP": "常规",
    "JMR": "常规",
    "LMR": "常规",
    "HMR": "常规",
    "EMR": "常规",
}
GENERIC_MLCC_SERIES_MEANING = {
    "CL": "常规 / General-purpose MLCC",
    "CLR1": "常规 / General-purpose MLCC",
    "CC": "常规 / General-purpose MLCC",
    "CQ": "常规 / General-purpose MLCC",
    "AC": "车规 / Automotive MLCC",
    "TCC": "常规 / General-purpose MLCC",
    "LMK": "常规 / General-purpose MLCC",
    "TMK": "常规 / General-purpose MLCC",
    "JMK": "常规 / General-purpose MLCC",
    "EMK": "常规 / General-purpose MLCC",
    "AMK": "常规 / General-purpose MLCC",
    "MAAS": "常规 / General-purpose MLCC",
    "MSAS": "常规 / General-purpose MLCC",
    "MLAS": "常规 / General-purpose MLCC",
    "MCAST": "常规 / General-purpose MLCC",
    "MCAS": "常规 / General-purpose MLCC",
    "CSA": "常规 / General-purpose MLCC",
    "CGA": "车规 / AEC-Q200",
    "CAI": "车规 / AEC-Q200",
}
GENERIC_MLCC_SERIES_CLASS = {
    "CL": "常规",
    "CLR1": "常规",
    "CC": "常规",
    "CQ": "常规",
    "AC": "车规",
    "TCC": "常规",
    "LMK": "常规",
    "TMK": "常规",
    "JMK": "常规",
    "EMK": "常规",
    "AMK": "常规",
    "MAAS": "常规",
    "MSAS": "常规",
    "MLAS": "常规",
    "MCAST": "常规",
    "MCAS": "常规",
    "CSA": "常规",
    "CGA": "车规",
    "CAI": "车规",
}

HRE_MLCC_SERIES_MEANING = {
    "CAA": "车规 / Automotive MLCC",
    "CAI": "车规 / Automotive MLCC",
    "CIA": "工业 / Industrial MLCC",
    "CGA": "渠道常规 / General-purpose MLCC",
    "CSA": "常规 / General-purpose MLCC",
    "CSS": "常规 / Soft-termination MLCC",
    "CSO": "常规 / General-purpose MLCC",
}
HRE_MLCC_SERIES_CLASS = {
    "CAA": "车规",
    "CAI": "车规",
    "CIA": "工业",
    "CGA": "常规",
    "CSA": "常规",
    "CSS": "软端子",
    "CSO": "常规",
}
HRE_MLCC_SERIES_SPECIAL_USE = {
    "CAA": "车规",
    "CAI": "车规",
    "CIA": "工业",
    "CGA": "渠道专用",
    "CSS": "软端子",
}


MLCC_SERIES_CLASS_RULES = [
    ("次车规", [r"WITHOUT\s+AEC[- ]?Q200", r"无\s*AEC[- ]?Q200", r"次车规"]),
    ("车规", [r"\bAUTOMOTIVE\b", r"\bAEC[- ]?Q200\b", r"汽车级", r"(?<!次)车规"]),
    ("软端子", [r"SOFT\s*TERMINATION", r"软端子"]),
    ("工业", [r"INDUSTRIAL", r"工业"]),
    ("高容", [r"HIGH\s+CAPACITANCE", r"高容"]),
    ("高压", [r"HIGH\s+VOLTAGE", r"高压"]),
    ("中压", [r"MEDIUM\s+VOLTAGE", r"中压"]),
    ("抗弯", [r"ANTI[- ]?BEND", r"抗弯"]),
    ("安规", [r"SAFETY", r"安规", r"\bX1/Y2\b", r"\bX2\b"]),
    ("高Q", [r"HIGH[- ]?Q", r"高Q"]),
    ("EMI滤波", [r"EMI\s*FILTER", r"EMI", r"滤波"]),
    ("常规", [r"GENERAL\s*PURPOSE", r"一般共用", r"常规"]),
]
MLCC_STRICT_CLASS_TOKENS = {"车规", "次车规", "软端子", "工业", "高容", "高压", "中压", "抗弯", "安规", "高Q", "EMI滤波"}


def mlcc_series_class_tokens(value):
    text = clean_text(value)
    if text == "":
        return []
    tokens = []
    for label, patterns in MLCC_SERIES_CLASS_RULES:
        for pattern in patterns:
            if re.search(pattern, text, flags=re.I):
                tokens.append(label)
                break
    return list(dict.fromkeys(tokens))


def normalize_mlcc_series_class(value):
    tokens = mlcc_series_class_tokens(value)
    if not tokens:
        return ""
    return "/".join(tokens)


def generic_mlcc_series_code_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    if compact.startswith("CLR1"):
        return "CLR1"
    if compact.startswith(("CGA", "CSA")):
        return compact[:3]
    if compact.startswith("CAI"):
        return "CAI"
    if compact.startswith("CL"):
        return "CL"
    if compact.startswith(("CC", "CQ", "AC")):
        return compact[:2]
    for prefix in ("MEASL", "MEAST", "MEASJ", "LMR", "HMR", "EMR", "JMR", "LMK", "TMK", "JMK", "EMK", "AMK", "MAAS", "MSAS", "MLAS", "MCAST", "MCAS", "TCC", "UP", "TP"):
        if compact.startswith(prefix):
            return prefix
    return ""


def extract_leading_alpha_series_code(model, min_len=2, max_len=6):
    compact = clean_model(model)
    if compact == "":
        return ""
    match = re.match(rf"^([A-Z]{{{min_len},{max_len}}})(?=\d)", compact)
    if not match:
        return ""
    return clean_text(match.group(1)).upper()


def extract_size_prefixed_mlcc_series_code(model, size_len=4):
    compact = clean_model(model)
    if compact == "" or len(compact) <= size_len or not compact[:size_len].isdigit():
        return ""
    rest = compact[size_len:]
    for token_len in (3, 2, 1):
        if len(rest) <= token_len:
            continue
        token = clean_text(rest[:token_len]).upper()
        tail = clean_text(rest[token_len:]).upper()
        if not any(ch.isalpha() for ch in token):
            continue
        if re.match(r"^(?:\d{3,4}|R\d+)", tail):
            return f"{compact[:size_len]}{token}"
    return ""


def yageo_mlcc_series_code_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    for prefix in ("AQ", "AS", "CS", "CC", "CQ", "AC"):
        if compact.startswith(prefix):
            return prefix
    return ""


def yageo_mlcc_series_profile(series_code):
    code = clean_text(series_code).upper()
    if code == "":
        return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    class_text = clean_text(YAGEO_MLCC_SERIES_CLASS.get(code, ""))
    special_use = "车规" if "车规" in class_text else ""
    return {
        "系列": code,
        "系列说明": clean_text(YAGEO_MLCC_SERIES_MEANING.get(code, "")),
        "特殊用途": special_use,
        "_mlcc_series_class": class_text,
    }


def walsin_mlcc_series_code_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    if compact.startswith("01R5"):
        return "01R5"
    for prefix in ("RF", "HH", "SH", "RT", "UF"):
        if compact.startswith(prefix):
            return prefix
    numeric_family = extract_size_prefixed_mlcc_series_code(compact)
    if numeric_family != "":
        return numeric_family
    return ""


def walsin_mlcc_series_profile(series_code):
    code = clean_text(series_code).upper()
    if code == "":
        return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    class_text = clean_text(WALSIN_MLCC_SERIES_CLASS.get(code, ""))
    special_use = "车规" if "车规" in class_text else ""
    return {
        "系列": code,
        "系列说明": clean_text(WALSIN_MLCC_SERIES_MEANING.get(code, "")),
        "特殊用途": special_use,
        "_mlcc_series_class": class_text,
    }


def taiyo_mlcc_series_code_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    for prefix in (
        "MEASL", "MEAST", "MEASJ",
        "JMR", "LMR", "HMR", "EMR",
        "MSAY", "MSRL",
        "MSART", "MBARQ", "MCARQ", "MMARQ",
        "MCAST", "MAAS", "MBAS", "MCAS", "MLAS", "MMAS", "MSAS",
        "UP", "TP",
        "UMK", "HMK", "QMK", "SMK", "QVS", "TVS",
        "TMK", "JMK", "EMK", "LMK", "AMK",
    ):
        if compact.startswith(prefix):
            return prefix
    generic_prefix = extract_leading_alpha_series_code(compact, min_len=3, max_len=6)
    if generic_prefix != "":
        return generic_prefix
    return ""


def taiyo_mlcc_series_profile(series_code):
    code = clean_text(series_code).upper()
    if code == "":
        return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    class_text = clean_text(TAIYO_MLCC_SERIES_CLASS.get(code, ""))
    special_use = "车规" if "车规" in class_text else ""
    return {
        "系列": code,
        "系列说明": clean_text(TAIYO_MLCC_SERIES_MEANING.get(code, "")),
        "特殊用途": special_use,
        "_mlcc_series_class": class_text,
    }


def generic_mlcc_series_profile(series_code):
    code = clean_text(series_code).upper()
    if code == "":
        return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    class_text = clean_text(GENERIC_MLCC_SERIES_CLASS.get(code, ""))
    special_use = "车规" if "车规" in class_text else ""
    return {
        "系列": code,
        "系列说明": clean_text(GENERIC_MLCC_SERIES_MEANING.get(code, "")),
        "特殊用途": special_use,
        "_mlcc_series_class": class_text,
    }


def hre_mlcc_series_code_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return ""
    for prefix in ("CAA", "CAI", "CIA", "CGA", "CSA", "CSS", "CSO"):
        if compact.startswith(prefix):
            return prefix
    return ""


def hre_mlcc_series_profile(series_code):
    code = clean_text(series_code).upper()
    if code == "":
        return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    class_text = clean_text(HRE_MLCC_SERIES_CLASS.get(code, ""))
    special_use = clean_text(HRE_MLCC_SERIES_SPECIAL_USE.get(code, ""))
    return {
        "系列": code,
        "系列说明": clean_text(HRE_MLCC_SERIES_MEANING.get(code, "")),
        "特殊用途": special_use,
        "_mlcc_series_class": class_text,
    }


def hre_mlcc_series_profile_from_model(model):
    return hre_mlcc_series_profile(hre_mlcc_series_code_from_model(model))


def expand_hre_cga_rows_to_csa(df):
    if df is None or df.empty or "型号" not in df.columns:
        return df
    work = df.copy()
    model_clean = work["型号"].astype("string").fillna("").apply(clean_model)
    if "品牌" in work.columns:
        brand_clean = work["品牌"].astype("string").fillna("").apply(clean_brand)
    else:
        brand_clean = pd.Series([""] * len(work), index=work.index, dtype="string")
    if "器件类型" in work.columns:
        component_type_clean = work["器件类型"].astype("string").fillna("").apply(normalize_component_type)
    else:
        component_type_clean = pd.Series([""] * len(work), index=work.index, dtype="string")
    hre_mask = (
        brand_clean.str.contains(r"HRE|芯声微", case=False, regex=True, na=False)
        & model_clean.str.startswith("CGA", na=False)
        & (component_type_clean.eq("") | component_type_clean.eq("MLCC"))
    )
    if not hre_mask.any():
        return work
    clones = work.loc[hre_mask].copy()
    clones["型号"] = clones["型号"].astype("string").fillna("").apply(clean_model).str.replace(r"^CGA", "CSA", regex=True)
    clones["系列"] = "CSA"
    clones["系列说明"] = "常规 / General-purpose MLCC"
    if "特殊用途" in clones.columns:
        clones["特殊用途"] = ""
    if "_mlcc_series_class" in clones.columns:
        clones["_mlcc_series_class"] = "常规"
    if "_model_rule_authority" in clones.columns:
        clones["_model_rule_authority"] = "hre_cga_to_csa_clone"
    combined = pd.concat([work, clones], ignore_index=True, sort=False)
    return combined


def generic_mlcc_series_profile_from_model(model, brand=""):
    brand_text = clean_brand(brand)
    brand_upper = clean_text(brand_text).upper()
    if "HRE" in brand_upper or "芯声微" in brand_text:
        return hre_mlcc_series_profile_from_model(model)
    return generic_mlcc_series_profile(generic_mlcc_series_code_from_model(model))


def tdk_mlcc_series_profile_from_model(model):
    compact = clean_model(model)
    if compact == "":
        return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    if compact.startswith("CGA"):
        return {
            "系列": "CGA",
            "系列说明": TDK_MLCC_SERIES_MEANING["CGA"],
            "特殊用途": "车规",
            "_mlcc_series_class": TDK_MLCC_SERIES_CLASS["CGA"],
        }
    if re.match(r"^C\d{4}[A-Z0-9].*", compact):
        return {
            "系列": "C",
            "系列说明": TDK_MLCC_SERIES_MEANING["C"],
            "特殊用途": "",
            "_mlcc_series_class": TDK_MLCC_SERIES_CLASS["C"],
        }
    return {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}


def resolve_mlcc_series_profile(brand="", model="", series="", series_desc="", special_use=""):
    brand_text = clean_brand(brand)
    brand_upper = clean_text(brand_text).upper()
    model_key = clean_model(model)
    series_key = clean_text(series).upper()
    is_pdc_brand = "PDC" in brand_upper or "信昌" in brand_text
    is_yageo_brand = "YAGEO" in brand_upper or "国巨" in brand_text
    is_walsin_brand = "WALSIN" in brand_upper or "华新科" in brand_text
    is_taiyo_brand = "TAIYO" in brand_upper or "太诱" in brand_text or "太阳诱电" in brand_text
    result = {
        "系列": clean_text(series),
        "系列说明": clean_text(series_desc),
        "特殊用途": clean_text(special_use),
        "_mlcc_series_class": normalize_mlcc_series_class(special_use),
    }

    profile = {"系列": "", "系列说明": "", "特殊用途": "", "_mlcc_series_class": ""}
    if model_key.startswith("AM") or "FENGHUA" in brand_upper or "风华" in brand_text:
        profile = {
            "系列": "AM",
            "系列说明": fenghua_am_series_meaning("AM"),
            "特殊用途": "车规",
            "_mlcc_series_class": "车规",
        }
    else:
        pdc_series_code = pdc_general_mlcc_series_code_from_model(model_key) if is_pdc_brand or brand_text == "" else ""
        murata_series_code = murata_series_code_from_model(model_key)
        yageo_series_code = yageo_mlcc_series_code_from_model(model_key if is_yageo_brand else "")
        walsin_series_code = walsin_mlcc_series_code_from_model(model_key if is_walsin_brand else "")
        taiyo_series_code = taiyo_mlcc_series_code_from_model(model_key if is_taiyo_brand else "")
        if pdc_series_code != "":
            profile = pdc_mlcc_series_profile(pdc_series_code)
        elif murata_series_code != "":
            profile = murata_series_profile(murata_series_code)
        elif yageo_series_code != "":
            profile = yageo_mlcc_series_profile(yageo_series_code)
        elif walsin_series_code != "":
            profile = walsin_mlcc_series_profile(walsin_series_code)
        elif taiyo_series_code != "":
            profile = taiyo_mlcc_series_profile(taiyo_series_code)
        elif "HRE" in brand_upper or "芯声微" in brand_text:
            profile = parse_generic_size_first_mlcc(model_key, brand=brand_text)
            if profile is None or clean_text(profile.get("系列", "")) == "":
                profile = generic_mlcc_series_profile_from_model(model_key, brand=brand_text)
        elif model_key.startswith("CGA") or re.match(r"^C\d{4}[A-Z0-9].*", model_key):
            profile = tdk_mlcc_series_profile_from_model(model_key)
        elif series_key in PDC_MLCC_SERIES_MEANING:
            profile = pdc_mlcc_series_profile(series_key)
        elif series_key in MURATA_SERIES_MEANING:
            profile = murata_series_profile(series_key)
        elif is_yageo_brand and series_key in YAGEO_MLCC_SERIES_MEANING:
            profile = yageo_mlcc_series_profile(series_key)
        elif is_walsin_brand and series_key in WALSIN_MLCC_SERIES_MEANING:
            profile = walsin_mlcc_series_profile(series_key)
        elif is_taiyo_brand and series_key in TAIYO_MLCC_SERIES_MEANING:
            profile = taiyo_mlcc_series_profile(series_key)
        elif series_key == "C" or series_key == "CGA" or series_key.startswith("CGA") or re.match(r"^C\d{4}$", series_key):
            profile = tdk_mlcc_series_profile_from_model(series_key)
        else:
            generic_profile = generic_mlcc_series_profile_from_model(model_key or series_key, brand=brand_text)
            if clean_text(generic_profile.get("系列", "")) == "":
                generic_profile = parse_generic_size_first_mlcc(model_key or series_key, brand=brand_text)
            if generic_profile is not None:
                profile = generic_profile

    profile_series = clean_text(profile.get("系列", ""))
    if mlcc_series_should_replace(result["系列"], profile_series):
        result["系列"] = profile_series
    if result["系列说明"] == "" and clean_text(profile.get("系列说明", "")) != "":
        result["系列说明"] = clean_text(profile.get("系列说明", ""))
    if result["特殊用途"] == "" and clean_text(profile.get("特殊用途", "")) != "":
        result["特殊用途"] = clean_text(profile.get("特殊用途", ""))
    profile_class_text = clean_text(profile.get("_mlcc_series_class", ""))
    current_class_text = normalize_mlcc_series_class(
        "/".join(
            part for part in [
                clean_text(result["系列说明"]),
                clean_text(result["特殊用途"]),
            ] if clean_text(part) != ""
        )
    )
    if (
        profile_series != ""
        and clean_text(result["系列"]) == profile_series
        and profile_class_text != ""
        and current_class_text != ""
        and current_class_text != profile_class_text
    ):
        profile_desc_text = clean_text(profile.get("系列说明", ""))
        if profile_desc_text != "":
            result["系列说明"] = profile_desc_text
        result["特殊用途"] = clean_text(profile.get("特殊用途", ""))

    if "HRE" in brand_upper or "芯声微" in brand_text:
        if profile_series != "":
            result["系列"] = profile_series
        profile_desc_text = clean_text(profile.get("系列说明", ""))
        if profile_desc_text != "":
            result["系列说明"] = profile_desc_text
        result["特殊用途"] = clean_text(profile.get("特殊用途", ""))
        result["_mlcc_series_class"] = clean_text(profile.get("_mlcc_series_class", ""))

    result["_mlcc_series_class"] = normalize_mlcc_series_class(
        "/".join(
            part for part in [
                clean_text(profile.get("_mlcc_series_class", "")),
                result["系列说明"],
                result["特殊用途"],
            ] if clean_text(part) != ""
        )
    )
    return result


def infer_mlcc_series_class_from_spec(spec):
    if spec is None:
        return ""
    profile = resolve_mlcc_series_profile(
        brand=spec.get("品牌", ""),
        model=spec.get("型号", ""),
        series=spec.get("系列", ""),
        series_desc=spec.get("系列说明", ""),
        special_use=spec.get("特殊用途", ""),
    )
    return clean_text(profile.get("_mlcc_series_class", ""))


def mlcc_series_class_requires_filter(value):
    tokens = set(mlcc_series_class_tokens(value))
    return bool(tokens & MLCC_STRICT_CLASS_TOKENS) or "常规" in tokens


def mlcc_series_class_matches(candidate, target):
    target_tokens = set(mlcc_series_class_tokens(target))
    if not target_tokens:
        return True
    candidate_tokens = set(mlcc_series_class_tokens(candidate))
    if target_tokens <= {"常规"}:
        return not bool(candidate_tokens & MLCC_STRICT_CLASS_TOKENS)
    if not candidate_tokens:
        return False
    if "车规" in target_tokens and "车规" not in candidate_tokens:
        return False
    if "次车规" in target_tokens and not ({"次车规", "车规"} & candidate_tokens):
        return False
    for token in MLCC_STRICT_CLASS_TOKENS - {"车规", "次车规"}:
        if token in target_tokens and token not in candidate_tokens:
            return False
    return True


def mlcc_series_class_sort_rank(candidate, target):
    target_text = normalize_mlcc_series_class(target)
    target_tokens = set(mlcc_series_class_tokens(target_text))
    if not target_tokens:
        return 1
    candidate_text = normalize_mlcc_series_class(candidate)
    candidate_tokens = set(mlcc_series_class_tokens(candidate_text))
    if target_tokens <= {"常规"}:
        if candidate_tokens <= {"常规"}:
            return 0 if candidate_text == "常规" else 1
        return 3
    if candidate_text != "" and candidate_text == target_text:
        return 0
    if mlcc_series_class_matches(candidate_text, target_text):
        return 1
    if candidate_text == "":
        return 2
    return 3


NICHICON_SERIES_PREFIX_PATTERN = re.compile(r"^(UV[A-Z0-9])")
CHEMI_CON_SERIES_PREFIX_PATTERN = re.compile(r"^E([A-Z0-9]{3})")
NICHICON_GENERIC_SERIES_PREFIX_PATTERN = re.compile(r"^([A-Z]{3})(?=[A-Z0-9])")


def nichicon_series_code_from_model(model):
    compact = clean_model(model)
    if not compact:
        return ""
    match = NICHICON_SERIES_PREFIX_PATTERN.match(compact)
    if not match:
        match = NICHICON_GENERIC_SERIES_PREFIX_PATTERN.match(compact)
        if not match:
            return ""
    return match.group(1)


def chemi_con_series_code_from_model(model):
    compact = clean_model(model)
    if not compact:
        return ""
    match = CHEMI_CON_SERIES_PREFIX_PATTERN.match(compact)
    if not match:
        return ""
    return match.group(1)


JIANGHAI_SERIES_MEANING = {
    "GUP": "江海 CD137U GUP 系列",
    "WUP": "江海 CD137U WUP 系列",
    "VPR": "江海 CD137S VPR 系列",
    "GPR": "江海 CD137S GPR 系列",
    "JBK": "江海 CD263 JBK 系列",
    "ABZ": "江海 CD293 ABZ 系列",
    "CBZ": "江海 CD293 CBZ 系列",
    "KBZ": "江海 CD293 KBZ 系列",
    "VBZ": "江海 CD293 VBZ 系列",
    "WBZ": "江海 CD293 WBZ 系列",
    "XBZ": "江海 CD293 XBZ 系列",
    "DQH": "江海 CD29H DQH 系列",
    "WQH": "江海 CD29H WQH 系列",
    "EQH": "江海 CD29H EQH 系列",
    "GQH": "江海 CD29H GQH 系列",
    "VNF": "江海 CD29NF VNF 系列",
    "GNF": "江海 CD29NF GNF 系列",
    "WNF": "江海 CD29NF WNF 系列",
    "EVA": "江海 CDA220 / PHVA EVA 系列",
    "GVA": "江海 CDA220 / PHVA GVA 系列",
    "JVA": "江海 CDA220 / PHVA JVA 系列",
    "EVZ": "江海 CDC220 EVZ 系列",
    "GVZ": "江海 CDC220 GVZ 系列",
    "JVZ": "江海 CDC220 JVZ 系列",
    "ELA": "江海 PHLA ELA 系列",
    "VLA": "江海 PHLA VLA 系列",
    "HLA": "江海 PHLA HLA 系列",
    "JLA": "江海 PHLA JLA 系列",
    "KLA": "江海 PHLA KLA 系列",
    "VLY": "江海 CD284L LY 系列",
    "HLY": "江海 CD284L LY 系列",
    "JLY": "江海 CD284L LY 系列",
    "KLY": "江海 CD284L LY 系列",
    "VVA": "江海 PHVA VVA 系列",
    "HVA": "江海 PHVA HVA 系列",
    "KVA": "江海 PHVA KVA 系列",
    "AVM": "江海 PCV1 AVM 系列",
    "CVM": "江海 PCV1 CVM 系列",
    "EVM": "江海 PCV1 EVM 系列",
    "EVF": "江海 PCV1 EVF 系列",
    "JVM": "江海 PCV1 JVM 系列",
    "EEQ": "江海 ECR1 EEQ 系列",
    "UP": "江海 CD137U UP 系列",
    "PR": "江海 CD137S PR 系列",
    "BK": "江海 CD263 BK 系列",
    "ELB": "江海 PHLB ELB 系列",
    "VLB": "江海 PHLB VLB 系列",
    "HLB": "江海 PHLB HLB 系列",
    "JLB": "江海 PHLB JLB 系列",
    "KLB": "江海 PHLB KLB 系列",
    "DPA": "江海 HPA DPA 系列",
    "EPA": "江海 HPA EPA 系列",
    "GPA": "江海 HPA GPA 系列",
    "JPA": "江海 HPA JPA 系列",
    "KPA": "江海 HPA KPA 系列",
    "APA": "江海 HPA APA 系列",
    "CPA": "江海 HPA CPA 系列",
    "VT1": "江海欧洲 CD VT1 系列",
    "VTD": "江海欧洲 CD VTD 系列",
    "VZ2": "江海欧洲 CD VZ2 系列",
    "VZL": "江海欧洲 CD VZL 系列",
    "VZS": "江海欧洲 CD VZS 系列",
}


def jianghai_series_code_from_model(model):
    compact = clean_model(model)
    if not compact:
        return ""
    patterns = [
        r"^ECG2([A-Z]{3})",
        r"^ECR\d([A-Z]{3})",
        r"^PCR\d([A-Z]{3})",
        r"^PCRA\d([A-Z]{3})",
        r"^PCV\d([A-Z]{3})",
        r"^ECS1([A-Z]{3})",
        r"^ECS2([A-Z]{3})",
        r"^ECA1([A-Z]{3})",
        r"^ECC1([A-Z]{3})",
        r"^PHR1([A-Z]{3})",
        r"^PHV1([A-Z]{3})",
        r"^ECG(?:[0-9A-Z]{2})(UP|PR)",
        r"^ECR(?:[0-9A-Z]{2})(BK)",
        r"^PCR(?:[0-9A-Z]{2})([A-Z]{3})",
        r"^PCV(?:[0-9A-Z]{2})([A-Z]{3})",
        r"^PHR1([A-Z]LB)",
        r"^PCP\d([A-Z]PA)",
        r"^ECV(?:0J|1A|1C|1E|1V|1H|1J|1K|2A)(VT1|VTD|VZ2|VZL|VZS)",
    ]
    for pattern in patterns:
        match = re.match(pattern, compact)
        if match:
            return match.group(1)
    return ""


def jianghai_series_meaning(series_code, model=""):
    code = clean_text(series_code).upper()
    if not code:
        return "Jianghai 官方系列代码"
    if code in JIANGHAI_SERIES_MEANING:
        return JIANGHAI_SERIES_MEANING.get(code, "Jianghai 官方系列代码")
    profile = jianghai_series_profile(code, model)
    family = clean_text(profile.get("family", ""))
    if family != "":
        return f"江海 {family} {code} 系列"
    return "Jianghai 官方系列代码"


JIANGHAI_HPA_VOLTAGE_MAP = {
    "DPA": "2",
    "EPA": "2.5",
    "GPA": "4",
    "JPA": "6.3",
    "KPA": "8",
    "APA": "10",
    "CPA": "16",
}

JIANGHAI_PHR_VOLTAGE_MAP = {
    "E": "25",
    "V": "35",
    "H": "50",
    "J": "63",
    "K": "80",
}

JIANGHAI_PHV_VOLTAGE_MAP = {
    "V": "16",
    "E": "25",
    "G": "35",
    "H": "50",
    "J": "63",
    "K": "80",
}

JIANGHAI_PHR_FAMILY_PROFILES = {
    "LA": {"family": "PHLA", "安装方式": "插件", "封装代码": "HYBRID", "工作温度": "-55~105℃", "寿命（h）": "5000", "特殊用途": "车规/混合聚合物"},
    "LB": {"family": "PHLB", "安装方式": "插件", "封装代码": "HYBRID", "工作温度": "-55~125℃", "寿命（h）": "4000", "特殊用途": "车规/混合聚合物"},
    "LD": {"family": "PHLD", "安装方式": "插件", "封装代码": "HYBRID", "工作温度": "-55~125℃", "寿命（h）": "4000", "特殊用途": "车规/混合聚合物"},
    "LE": {"family": "PHLE", "安装方式": "插件", "封装代码": "HYBRID", "工作温度": "-55~105℃", "寿命（h）": "10000", "特殊用途": "车规/混合聚合物"},
    "LF": {"family": "PHLF", "安装方式": "插件", "封装代码": "HYBRID", "工作温度": "-55~135℃", "寿命（h）": "4000", "特殊用途": "车规/混合聚合物"},
}

JIANGHAI_PHV_FAMILY_PROFILES = {
    "VA": {"family": "PHVA", "安装方式": "贴片", "封装代码": "HYBRID", "工作温度": "-55~105℃", "寿命（h）": "5000", "特殊用途": "车规/混合聚合物"},
    "VB": {"family": "PHVB", "安装方式": "贴片", "封装代码": "HYBRID", "工作温度": "-55~125℃", "寿命（h）": "4000", "特殊用途": "车规/混合聚合物"},
    "VD": {"family": "PHVD", "安装方式": "贴片", "封装代码": "HYBRID", "工作温度": "-55~125℃", "寿命（h）": "4000", "特殊用途": "车规/混合聚合物"},
    "VE": {"family": "PHVE", "安装方式": "贴片", "封装代码": "HYBRID", "工作温度": "-55~105℃", "寿命（h）": "10000", "特殊用途": "车规/混合聚合物"},
    "VF": {"family": "PHVF", "安装方式": "贴片", "封装代码": "HYBRID", "工作温度": "-55~135℃", "寿命（h）": "4000", "特殊用途": "车规/混合聚合物"},
}

JIANGHAI_GENERIC_SERIES_PROFILES = {
    "PA": {"family": "HPA", "安装方式": "贴片", "封装代码": "POLYMER", "工作温度": "105℃", "寿命（h）": "2000", "特殊用途": "消费/导电聚合物"},
    "UP": {"family": "CD137U", "安装方式": "螺栓式", "封装代码": "SCREW", "工作温度": "-40~105℃", "寿命（h）": "5000", "特殊用途": "工业/高纹波"},
    "PR": {"family": "CD137S", "安装方式": "螺栓式", "封装代码": "SCREW", "工作温度": "-40~105℃", "寿命（h）": "5000", "特殊用途": "工业/高纹波"},
    "BK": {"family": "CD263", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "BZ": {"family": "CD293", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "QH": {"family": "CD29H", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "NF": {"family": "CD29NF", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "-25~105℃", "寿命（h）": "2000", "特殊用途": "工业/耐腐蚀"},
    "VLY": {"family": "CD284L LY", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "-40~105℃", "寿命（h）": "", "特殊用途": "低ESR/长寿命"},
    "HLY": {"family": "CD284L LY", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "-40~105℃", "寿命（h）": "", "特殊用途": "低ESR/长寿命"},
    "JLY": {"family": "CD284L LY", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "-40~105℃", "寿命（h）": "", "特殊用途": "低ESR/长寿命"},
    "KLY": {"family": "CD284L LY", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "-40~105℃", "寿命（h）": "", "特殊用途": "低ESR/长寿命"},
    "EEQ": {"family": "ECR1 EEQ", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "105℃", "寿命（h）": "", "特殊用途": "低ESR"},
    "AVM": {"family": "PCV1 AVM", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "", "特殊用途": "混合物"},
    "CVM": {"family": "PCV1 CVM", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "", "特殊用途": "混合物"},
    "EVM": {"family": "PCV1 EVM", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "", "特殊用途": "混合物"},
    "EVF": {"family": "PCV1 EVF", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "3000", "特殊用途": "混合物"},
    "JVM": {"family": "PCV1 JVM", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "", "特殊用途": "混合物"},
}

JIANGHAI_EUROPE_SMD_VOLTAGE_MAP = {
    "0J": "6.3",
    "1A": "10",
    "1C": "16",
    "1E": "25",
    "1V": "35",
    "1H": "50",
    "1J": "63",
    "1K": "80",
    "2A": "100",
}

JIANGHAI_EUROPE_SMD_TOLERANCE_MAP = {
    "J": "5",
    "K": "10",
    "M": "20",
    "Q": "+30/-10",
}

JIANGHAI_EUROPE_SMD_SIZE_CODE_MAP = {
    "0405": "4*5.4mm",
    "0505": "5*5.4mm",
    "0506": "5*6mm",
    "0605": "6.3*5.4mm",
    "0606": "6.3*6mm",
    "0607": "6.3*7.7mm",
    "0806": "8*6.2mm",
    "0810": "8*10.2mm",
    "1010": "10*10.2mm",
    "1012": "10*12.5mm",
}

JIANGHAI_EUROPE_SMD_SERIES_PROFILES = {
    "VT1": {"family": "CD VT1", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "1000", "特殊用途": "低ESR/标准型"},
    "VTD": {"family": "CD VTD", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "2000", "特殊用途": "低ESR"},
    "VZ2": {"family": "CD VZ2", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "2000", "特殊用途": "高纹波/低阻抗"},
    "VZL": {"family": "CD VZL", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "5000", "特殊用途": "长寿命/低ESR"},
    "VZS": {"family": "CD VZS", "安装方式": "贴片", "封装代码": "SMD", "工作温度": "-55~105℃", "寿命（h）": "2000", "特殊用途": "小型化/低ESR"},
}

JIANGHAI_SCREW_VOLTAGE_MAP = {
    "GUP": "400",
    "WUP": "450",
    "UP": "400~450",
    "VPR": "350",
    "GPR": "400",
    "WPR": "450",
    "HPR": "500",
    "PR": "350~500",
}

JIANGHAI_LY_VOLTAGE_MAP = {
    "VLY": "35",
    "HLY": "50",
    "JLY": "63",
    "KLY": "80",
}

JIANGHAI_PCV_VOLTAGE_MAP = {
    "AVM": "10",
    "CVM": "16",
    "EVM": "25",
    "EVF": "25",
    "JVM": "6.3",
}

JIANGHAI_ECR1_RADIAL_SERIES_PROFILES = {
    "ALL": {"family": "CD281 / CD281L", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "AXX": {"family": "CD282", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "JYL": {"family": "CD282L", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "AEQ": {"family": "CD282X", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "JXY": {"family": "CD284", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "KLY": {"family": "CD284L", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "-40~105℃", "寿命（h）": "", "特殊用途": "低ESR/长寿命"},
    "EQL": {"family": "CD28L", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
}

JIANGHAI_ECR2_RADIAL_SERIES_PROFILES = {
    "XDE": {"family": "CD261L", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "GKH": {"family": "CD264", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "CFK": {"family": "CD266", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "WZR": {"family": "CD26HL", "安装方式": "插件", "封装代码": "RADIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
}

JIANGHAI_ECS_SNAPIN_SERIES_PROFILES = {
    "CBW": {"family": "CD294", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "FBC": {"family": "CD295", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "WKC": {"family": "CD296", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "WQK": {"family": "CD296Q", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "EBB": {"family": "CD297", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "GPG": {"family": "CD299", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "GQT": {"family": "CD29CT", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "GQR": {"family": "CD29HE", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
    "EQL": {"family": "CD29L", "安装方式": "插件", "封装代码": "SNAP-IN", "工作温度": "", "寿命（h）": "", "特殊用途": "工业"},
}

JIANGHAI_PHR_SIZE_CODE_MAP = {
    "BAB": "8*11.5mm",
    "C10": "10*10mm",
    "CAC": "10*12.5mm",
    "C16": "10*16mm",
    "C20": "10*20mm",
}

JIANGHAI_PHV_SIZE_CODE_MAP = {
    "F80": "6.3*6.6mm",
    "B10": "8*8.3mm",
    "C10": "10*10.3mm",
    "C12": "10*12.2mm",
    "C16": "10*16mm",
}

JIANGHAI_HPA_SIZE_CODE_MAP = {
    "V": "7.3*4.3*1.9mm",
    "E": "7.3*4.3*1.1mm",
    "B": "7.3*4.3*1.4mm",
    "D": "7.3*4.3*2.8mm",
}

JIANGHAI_CD137U_DIAMETER_MAP = {
    "C": "51",
    "D": "64",
    "E": "77",
    "F": "90",
}

JIANGHAI_SNAPIN_DIAMETER_MAP = {
    "LA": "35",
    "LB": "30",
}


def jianghai_series_profile(series_code, model=""):
    code = clean_text(series_code).upper()
    model_clean = clean_model(model)
    if code == "":
        return {}
    if model_clean.startswith("ECV") and code in JIANGHAI_EUROPE_SMD_SERIES_PROFILES:
        profile = dict(JIANGHAI_EUROPE_SMD_SERIES_PROFILES.get(code, {}))
        if len(model_clean) >= 5:
            profile["voltage"] = JIANGHAI_EUROPE_SMD_VOLTAGE_MAP.get(model_clean[3:5], "")
        return profile
    if model_clean.startswith("PHR1") and len(code) == 3 and code[1:] in JIANGHAI_PHR_FAMILY_PROFILES:
        profile = dict(JIANGHAI_PHR_FAMILY_PROFILES[code[1:]])
        profile["voltage"] = JIANGHAI_PHR_VOLTAGE_MAP.get(code[0], "")
        return profile
    if model_clean.startswith("PHV1") and len(code) == 3 and code[1:] in JIANGHAI_PHV_FAMILY_PROFILES:
        profile = dict(JIANGHAI_PHV_FAMILY_PROFILES[code[1:]])
        profile["voltage"] = JIANGHAI_PHV_VOLTAGE_MAP.get(code[0], "")
        return profile
    if model_clean.startswith("PCV") and code in JIANGHAI_PCV_VOLTAGE_MAP:
        profile = dict(JIANGHAI_GENERIC_SERIES_PROFILES.get(code, {}))
        profile["voltage"] = JIANGHAI_PCV_VOLTAGE_MAP.get(code, "")
        return profile
    if model_clean.startswith("ECR1") and code in JIANGHAI_ECR1_RADIAL_SERIES_PROFILES:
        profile = dict(JIANGHAI_ECR1_RADIAL_SERIES_PROFILES.get(code, {}))
        profile["voltage"] = jianghai_voltage_from_cd263_model(model_clean)
        return profile
    if model_clean.startswith("ECR2") and code in JIANGHAI_ECR2_RADIAL_SERIES_PROFILES:
        profile = dict(JIANGHAI_ECR2_RADIAL_SERIES_PROFILES.get(code, {}))
        profile["voltage"] = jianghai_voltage_from_cd263_model(model_clean)
        return profile
    if model_clean.startswith(("ECS1", "ECS2")) and code in JIANGHAI_ECS_SNAPIN_SERIES_PROFILES:
        profile = dict(JIANGHAI_ECS_SNAPIN_SERIES_PROFILES.get(code, {}))
        profile["voltage"] = jianghai_voltage_from_snapin_model(model_clean)
        return profile
    if model_clean.startswith(("ECR1", "ECR2")) and code in JIANGHAI_LY_VOLTAGE_MAP:
        profile = dict(JIANGHAI_GENERIC_SERIES_PROFILES.get(code, {}))
        profile["voltage"] = JIANGHAI_LY_VOLTAGE_MAP.get(code, "")
        return profile
    if model_clean.startswith("ECR1") and code == "EEQ":
        profile = dict(JIANGHAI_GENERIC_SERIES_PROFILES.get("EEQ", {}))
        profile["voltage"] = "25"
        return profile
    if model_clean.startswith("PCP") and code.endswith("PA"):
        profile = dict(JIANGHAI_GENERIC_SERIES_PROFILES["PA"])
        profile["voltage"] = JIANGHAI_HPA_VOLTAGE_MAP.get(code, "")
        return profile
    if model_clean.startswith("ECG2") and code.endswith("UP"):
        profile = dict(JIANGHAI_GENERIC_SERIES_PROFILES["UP"])
        profile["voltage"] = JIANGHAI_SCREW_VOLTAGE_MAP.get(code, "")
        return profile
    if model_clean.startswith("ECG2") and code.endswith("PR"):
        profile = dict(JIANGHAI_GENERIC_SERIES_PROFILES["PR"])
        profile["voltage"] = JIANGHAI_SCREW_VOLTAGE_MAP.get(code, "")
        return profile
    if model_clean.startswith("ECR") and code.endswith("BK"):
        return dict(JIANGHAI_GENERIC_SERIES_PROFILES["BK"])
    if model_clean.startswith(("ECS1", "ECS2")) and code.endswith("BZ"):
        return dict(JIANGHAI_GENERIC_SERIES_PROFILES["BZ"])
    if model_clean.startswith(("ECS1", "ECS2")) and code.endswith("QH"):
        return dict(JIANGHAI_GENERIC_SERIES_PROFILES["QH"])
    if model_clean.startswith(("ECS1", "ECS2")) and code.endswith("NF"):
        return dict(JIANGHAI_GENERIC_SERIES_PROFILES["NF"])
    if model_clean.startswith("ECA1") and code.endswith("VA"):
        return {"family": "CDA220", "安装方式": "插件", "封装代码": "AXIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "轴向"}
    if model_clean.startswith("ECC1") and code.endswith("VZ"):
        return {"family": "CDC220", "安装方式": "插件", "封装代码": "AXIAL", "工作温度": "", "寿命（h）": "", "特殊用途": "轴向"}
    return {}


def jianghai_capacitance_from_model(model):
    compact = clean_model(model)
    match = re.search(r"(?<!\d)(\d{2,4})(?=[MQW])", compact)
    if not match:
        return ""
    code = match.group(1)
    if len(code) == 2:
        return code
    if len(code) == 3:
        return str(int(code[:2]) * (10 ** int(code[2])))
    if len(code) == 4:
        return str(int(code[:3]) * (10 ** int(code[3])))
    return code


def jianghai_dimension_from_tenth_token(token):
    digits = re.sub(r"[^0-9]", "", clean_text(token))
    if digits == "":
        return ""
    try:
        value = int(digits)
    except Exception:
        return ""
    if len(digits) == 3 and value >= 100:
        return _format_compact_number(value / 10.0)
    return str(value)


def jianghai_dimension_from_integer_token(token):
    digits = re.sub(r"[^0-9]", "", clean_text(token))
    if digits == "":
        return ""
    try:
        return str(int(digits))
    except Exception:
        return ""


def jianghai_size_from_six_digit_code(code):
    digits = re.sub(r"[^0-9]", "", clean_text(code))
    if len(digits) != 6:
        return ""
    diameter_code = digits[:3]
    height_code = digits[3:]
    diameter = jianghai_dimension_from_tenth_token(diameter_code)
    height_map = {
        "011": "11.5",
        "012": "12.5",
        "013": "13.5",
        "014": "14.5",
        "015": "15.5",
        "115": "11.5",
        "125": "12.5",
        "135": "13.5",
        "145": "14.5",
        "155": "15.5",
        "016": "16",
        "020": "20",
        "025": "25",
        "030": "30",
        "031": "31.5",
        "035": "35",
        "040": "40",
        "045": "45",
        "050": "50",
    }
    height = height_map.get(height_code, jianghai_dimension_from_integer_token(height_code))
    if diameter == "" or height == "":
        return ""
    return jianghai_size_text(diameter, height)


def jianghai_size_text(diameter, height, width=""):
    dia = clean_text(diameter)
    hgt = clean_text(height)
    wid = clean_text(width)
    if dia == "" or hgt == "":
        return ""
    if wid != "":
        return f"{dia}*{wid}*{hgt}mm"
    return f"{dia}*{hgt}mm"


def jianghai_size_from_model_rule(model, series_code=""):
    compact = clean_model(model)
    code = clean_text(series_code).upper()
    if compact == "":
        return ""
    if compact.startswith("PHR1"):
        match = re.search(r"M([A-Z0-9]{3})", compact)
        if match:
            return JIANGHAI_PHR_SIZE_CODE_MAP.get(match.group(1), "")
    if compact.startswith("PHV1"):
        match = re.search(r"M([A-Z0-9]{3})", compact)
        if match:
            return JIANGHAI_PHV_SIZE_CODE_MAP.get(match.group(1), "")
    if compact.startswith("PCP"):
        suffix = compact[-1]
        return JIANGHAI_HPA_SIZE_CODE_MAP.get(suffix, "")
    if compact.startswith("PCV"):
        pcv_size_map = {
            "MF60": "6.3*5.7mm",
            "MF08": "6.3*8mm",
            "MB10": "8*10mm",
            "MC10": "10*10mm",
        }
        for token, size_text in pcv_size_map.items():
            if token in compact:
                return size_text
    if compact.startswith("ECG2") and code.endswith("UP"):
        match = re.search(r"M([A-Z])(\d{3})$", compact)
        if match:
            diameter = JIANGHAI_CD137U_DIAMETER_MAP.get(match.group(1), "")
            height = jianghai_dimension_from_integer_token(match.group(2))
            return jianghai_size_text(diameter, height)
    if compact.startswith(("ECR1", "ECR2")):
        match = re.search(r"(\d{3})(\d{3})(?:[A-Z])?$", compact)
        if match:
            size_text = jianghai_size_from_six_digit_code("".join(match.groups()))
            if size_text != "":
                return size_text
    if compact.startswith(("ECS1", "ECS2")):
        match = re.search(r"M([A-Z]{2})(\d{3})(\d{3})V?$", compact)
        if match:
            diameter = JIANGHAI_SNAPIN_DIAMETER_MAP.get(match.group(1), "")
            height = jianghai_dimension_from_integer_token(match.group(3))
            if diameter != "" and height != "":
                return jianghai_size_text(diameter, height)
        match = re.search(r"M(\d{3})(\d{3})V?$", compact)
        if match:
            size_text = jianghai_size_from_six_digit_code("".join(match.groups()))
            if size_text != "":
                return size_text
    if compact.startswith(("ECA1", "ECC1")):
        tail_match = re.search(r"(\d{3})(\d{3})$", compact)
        if tail_match:
            diameter = jianghai_dimension_from_tenth_token(tail_match.group(1))
            height = jianghai_dimension_from_integer_token(tail_match.group(2))
            return jianghai_size_text(diameter, height)
    return ""


JIANGHAI_EUROPE_SMD_DATASHEET_URL = "https://jianghai-europe.com/wp-content/uploads/JE25_ECap_Catalogue.pdf"


def jianghai_europe_cap_code_to_uf(code):
    token = clean_model(code)
    if token == "":
        return ""
    if "R" in token:
        match = re.fullmatch(r"(\d*)R(\d+)", token)
        if not match:
            return ""
        left = match.group(1) or "0"
        right = match.group(2)
        try:
            return _format_compact_number(float(f"{left}.{right}"))
        except Exception:
            return ""
    if not re.fullmatch(r"\d{3}", token):
        return ""
    try:
        base = int(token[:2])
        exponent = int(token[2])
        return _format_compact_number(base * (10 ** exponent))
    except Exception:
        return ""


def parse_jianghai_europe_smd_model(model):
    compact = clean_model(model)
    if compact == "":
        return None
    match = re.fullmatch(
        r"ECV(?P<volt>0J|1A|1C|1E|1V|1H|1J|1K|2A)(?P<series>VT1|VTD|VZ2|VZL|VZS)(?P<cap>(?:\d{3}|[0-9]*R[0-9]+))(?P<tol>[A-Z])(?P<size>\d{4})(?P<rest>[A-Z0-9]*)",
        compact,
    )
    if match is None:
        return None
    series_code = match.group("series")
    series_profile = dict(JIANGHAI_EUROPE_SMD_SERIES_PROFILES.get(series_code, {}))
    voltage = clean_voltage(JIANGHAI_EUROPE_SMD_VOLTAGE_MAP.get(match.group("volt"), ""))
    capacitance_uf = jianghai_europe_cap_code_to_uf(match.group("cap"))
    if capacitance_uf == "":
        return None
    capacitance_pf = cap_to_pf(capacitance_uf, "UF")
    tolerance = clean_tol_for_match(JIANGHAI_EUROPE_SMD_TOLERANCE_MAP.get(match.group("tol"), ""))
    body_size = clean_text(JIANGHAI_EUROPE_SMD_SIZE_CODE_MAP.get(match.group("size"), ""))
    work_temp = normalize_working_temperature_text(series_profile.get("工作温度", ""))
    life_hours = normalize_life_hours_value(series_profile.get("寿命（h）", ""))
    mounting_style = normalize_mounting_style(series_profile.get("安装方式", ""), series_profile.get("封装代码", ""))
    special_use = normalize_special_use(series_profile.get("特殊用途", ""))
    summary = build_other_component_summary(
        [
            f"{capacitance_uf}UF",
            clean_tol_for_display(tolerance),
            voltage_display(voltage),
            work_temp,
            format_life_hours_display(life_hours),
            body_size,
            mounting_style,
            special_use,
        ]
    )
    param_count = sum(
        [
            1 if capacitance_pf is not None else 0,
            1 if tolerance != "" else 0,
            1 if voltage != "" else 0,
            1 if body_size != "" else 0,
            1 if work_temp != "" else 0,
            1 if life_hours != "" else 0,
        ]
    )
    if param_count < 3:
        return None
    return {
        "品牌": "江海Jianghai",
        "型号": compact,
        "系列": series_code,
        "系列说明": jianghai_series_meaning(series_code, compact),
        "器件类型": "铝电解电容",
        "容值": clean_text(capacitance_uf),
        "容值单位": "UF",
        "容值_pf": capacitance_pf,
        "容值误差": tolerance,
        "耐压（V）": voltage,
        "尺寸（mm）": body_size,
        "工作温度": work_temp,
        "寿命（h）": life_hours,
        "安装方式": mounting_style,
        "封装代码": clean_text(series_profile.get("封装代码", "")),
        "特殊用途": special_use,
        "规格摘要": summary,
        "备注2": JIANGHAI_EUROPE_SMD_DATASHEET_URL,
        "_body_size": body_size,
        "_core_param_count": param_count,
        "_param_count": param_count,
        "_model_rule_authority": "jianghai_europe_smd_series",
    }


def jianghai_voltage_from_cd263_model(model):
    compact = clean_model(model)
    match = re.search(r"M(\d{3})\d{3}$", compact)
    if not match:
        return ""
    return clean_voltage(match.group(1))


def jianghai_voltage_from_snapin_model(model):
    compact = clean_model(model)
    match = re.search(r"M(?:[A-Z]{2})?(\d{3})(\d{3})V?$", compact)
    if not match:
        return ""
    return clean_voltage(match.group(1))


def jianghai_voltage_from_model_rule(model, series_code=""):
    compact = clean_model(model)
    code = clean_text(series_code).upper()
    if compact == "":
        return ""
    profile = jianghai_series_profile(code, compact)
    profile_voltage = clean_voltage(profile.get("voltage", ""))
    if profile_voltage != "":
        return profile_voltage
    if compact.startswith("ECR"):
        return jianghai_voltage_from_cd263_model(compact)
    if compact.startswith(("ECS1", "ECS2")):
        return jianghai_voltage_from_snapin_model(compact)
    return ""


def parse_jianghai_aluminum_model(model):
    compact = clean_model(model)
    if compact == "":
        return None
    europe_smd_parsed = parse_jianghai_europe_smd_model(compact)
    if europe_smd_parsed is not None:
        return europe_smd_parsed
    series_code = jianghai_series_code_from_model(compact)
    if series_code == "":
        return None
    profile = jianghai_series_profile(series_code, compact)
    cap_uf = jianghai_capacitance_from_model(compact)
    cap_pf = cap_to_pf(cap_uf, "UF") if cap_uf != "" else None
    voltage = jianghai_voltage_from_model_rule(compact, series_code)
    body_size = jianghai_size_from_model_rule(compact, series_code)
    work_temp = normalize_working_temperature_text(profile.get("工作温度", ""))
    life_hours = normalize_life_hours_value(profile.get("寿命（h）", ""))
    mounting_style = normalize_mounting_style(profile.get("安装方式", ""), profile.get("封装代码", ""))
    special_use = normalize_special_use(profile.get("特殊用途", ""))
    summary = build_other_component_summary([
        f"{cap_uf}UF" if cap_uf != "" else "",
        "±20%",
        voltage_display(voltage) if voltage != "" else "",
        work_temp,
        format_life_hours_display(life_hours),
        body_size,
        mounting_style,
        special_use,
    ])
    param_count = sum([
        1 if cap_pf is not None else 0,
        1 if voltage != "" else 0,
        1 if body_size != "" else 0,
        1 if work_temp != "" else 0,
        1 if life_hours != "" else 0,
    ])
    if param_count < 2:
        return None
    return {
        "品牌": "江海Jianghai",
        "型号": compact,
        "系列": series_code,
        "系列说明": jianghai_series_meaning(series_code, compact),
        "器件类型": "铝电解电容",
        "容值_pf": cap_pf,
        "容值误差": "20",
        "耐压（V）": voltage,
        "尺寸（mm）": body_size,
        "工作温度": work_temp,
        "寿命（h）": life_hours,
        "安装方式": mounting_style,
        "封装代码": clean_text(profile.get("封装代码", "")),
        "特殊用途": special_use,
        "规格摘要": summary,
        "_body_size": body_size,
        "_core_param_count": param_count,
        "_param_count": param_count,
        "_model_rule_authority": "jianghai_series_model",
    }


def build_jianghai_seed_row(
    model,
    series,
    capacitance_uf,
    voltage,
    size_mm,
    mounting_style,
    package_code,
    work_temp="",
    life_hours="",
    special_use="",
    esr="",
    data_source="用户样本",
    data_status="待复核",
    official_url="",
):
    def split_jianghai_size_text(value):
        text = normalize_dimension_mm_value(value)
        if text == "":
            return "", ""
        text = text.replace("×", "*").replace("x", "*").replace("X", "*")
        parts = [clean_text(part) for part in re.split(r"\*", text) if clean_text(part) != ""]
        if len(parts) >= 3:
            return parts[0], parts[-1]
        if len(parts) == 2:
            return parts[0], parts[1]
        return "", ""

    size_text = normalize_dimension_mm_value(size_mm).replace("*", "×").replace("x", "×").replace("X", "×")
    size_d, size_h = split_jianghai_size_text(size_text)
    cap_text = clean_text(capacitance_uf)
    cap_pf = cap_to_pf(cap_text, "UF") if cap_text != "" else None
    voltage_text = voltage_display(voltage) if clean_text(voltage) != "" else ""
    summary_parts = [
        f"{cap_text}uF" if cap_text != "" else "",
        "±20%",
        voltage_text,
        normalize_working_temperature_text(work_temp),
        size_text,
        normalize_mounting_style(mounting_style, package_code),
        normalize_special_use(special_use),
    ]
    summary = build_other_component_summary(summary_parts)
    return {
        "品牌": "江海Jianghai",
        "型号": clean_model(model),
        "系列": clean_text(series).upper(),
        "器件类型": "铝电解电容",
        "容值_pf": cap_pf,
        "容值": cap_text,
        "容值单位": "UF",
        "容值误差": "±20%",
        "耐压（V）": voltage_text,
        "材质（介质）": "",
        "尺寸（inch）": "",
        "尺寸（mm）": size_text,
        "直径（mm）": size_d,
        "高度（mm）": size_h,
        "长度（mm）": "",
        "宽度（mm）": "",
        "安装方式": clean_text(mounting_style),
        "封装代码": clean_text(package_code),
        "工作温度": normalize_working_temperature_text(work_temp),
        "寿命（h）": normalize_life_hours_value(life_hours),
        "特殊用途": normalize_special_use(special_use),
        "ESR": clean_text(esr),
        "规格摘要": summary,
        "数据来源": clean_text(data_source),
        "数据状态": clean_text(data_status),
        "官网链接": clean_text(official_url),
        "校验时间": time.strftime("%Y-%m-%d"),
        "校验备注": clean_text(data_status or data_source),
        "备注1": "",
        "备注2": "",
        "备注3": "",
        "_model_rule_authority": "jianghai_seed",
    }


def load_jianghai_seed_rows():
    return [
        build_jianghai_seed_row(
            "ECR1VLY152MLL125030E",
            "VLY",
            "1500",
            "35V",
            "12.5×30",
            "插件",
            "RADIAL",
            work_temp="-40~105℃",
            special_use="低ESR/长寿命",
            esr="25mΩ",
            data_source="用户样本 / Jianghai CD284L LY series PDF",
            data_status="待官方复核",
            official_url="https://www.jlc-smt.com/lcsc/detail/C233069.html",
        ),
        build_jianghai_seed_row(
            "ECR1EEQ681MLL100020E",
            "EEQ",
            "680",
            "25V",
            "10×20",
            "插件",
            "RADIAL",
            work_temp="105℃",
            special_use="低ESR",
            esr="62mΩ",
            data_source="用户样本",
            data_status="待官方复核",
        ),
        build_jianghai_seed_row(
            "PCV1EVF221MB10FVTSWP",
            "EVF",
            "220",
            "25V",
            "8×10",
            "贴片",
            "SMD",
            work_temp="-55~105℃",
            life_hours="3000",
            special_use="混合物",
            esr="27mΩ",
            data_source="用户样本",
            data_status="待官方复核",
        ),
    ]


SUNLORD_MCL_OFFICIAL_PDF_URL = "https://www.sunlordinc.com/uploads/files/20221205/MCL%20Series%20of%20Multilayer%20Chip%20Ferrite%20Inductor.pdf"


def build_sunlord_mcl_seed_row(
    model,
    series,
    size_inch,
    size_mm,
    length_mm,
    width_mm,
    height_mm,
    inductance_uh,
    tolerance_pct,
    rated_current_a,
    dcr_ohm,
    srf_mhz,
):
    inductance_text = clean_text(inductance_uh)
    tolerance_text = clean_text(tolerance_pct)
    current_text = clean_text(rated_current_a)
    dcr_text = clean_text(dcr_ohm)
    srf_text = clean_text(srf_mhz)
    summary = build_other_component_summary([
        f"{inductance_text}uH" if inductance_text else "",
        f"±{tolerance_text}%" if tolerance_text else "",
        clean_text(size_inch),
        f"Ir {current_text}A" if current_text else "",
        f"DCR {dcr_text}Ω" if dcr_text else "",
        f"SRF {srf_text}MHz" if srf_text else "",
        "SMT",
    ])
    return {
        "品牌": "Sunlord(顺络)",
        "型号": clean_model(model),
        "系列": clean_text(series).upper(),
        "系列说明": "Sunlord MCL multilayer chip ferrite inductor",
        "器件类型": "功率电感",
        "容值": inductance_text,
        "容值单位": "UH",
        "容值误差": tolerance_text,
        "电感值": inductance_text,
        "电感单位": "UH",
        "电感误差": tolerance_text,
        "耐压（V）": "",
        "材质（介质）": "",
        "尺寸（inch）": clean_text(size_inch),
        "尺寸（mm）": clean_text(size_mm),
        "长度（mm）": clean_text(length_mm),
        "宽度（mm）": clean_text(width_mm),
        "高度（mm）": clean_text(height_mm),
        "安装方式": "SMT",
        "封装代码": clean_text(series).upper(),
        "特殊用途": "Choke",
        "额定电流": current_text,
        "DCR": dcr_text,
        "规格摘要": summary,
        "生产状态": "Active",
        "数据来源": "Sunlord official PDF",
        "数据状态": "官方网页抽取",
        "官网链接": SUNLORD_MCL_OFFICIAL_PDF_URL,
        "校验时间": time.strftime("%Y-%m-%d"),
        "校验备注": "official pdf table",
        "备注1": f"SRF {srf_text}MHz" if srf_text else "",
        "备注2": "",
        "备注3": "",
        "_model_rule_authority": "sunlord_mcl_seed",
    }


def load_sunlord_mcl_seed_rows():
    rows = []
    base_rows = [
        ("MCL1608", "0603", "1.6×0.8×0.8 mm", "1.6", "0.8", "0.8", "MCL1608SR10", "0.1", "0.70", "0.182/0.140", "240"),
        ("MCL1608", "0603", "1.6×0.8×0.8 mm", "1.6", "0.8", "0.8", "MCL1608SR22", "0.22", "0.55", "0.351/0.270", "150"),
        ("MCL1608", "0603", "1.6×0.8×0.8 mm", "1.6", "0.8", "0.8", "MCL1608SR47", "0.47", "0.40", "0.546/0.420", "105"),
        ("MCL1608", "0603", "1.6×0.8×0.8 mm", "1.6", "0.8", "0.8", "MCL1608S1R0", "1.0", "0.19", "0.260/0.200", "75"),
        ("MCL1608", "0603", "1.6×0.8×0.8 mm", "1.6", "0.8", "0.8", "MCL1608S2R2", "2.2", "0.14", "0.520/0.400", "50"),
        ("MCL1608", "0603", "1.6×0.8×0.8 mm", "1.6", "0.8", "0.8", "MCL1608S4R7", "4.7", "0.10", "0.780/0.600", "35"),
        ("MCL1608", "0603", "1.6×0.8×0.8 mm", "1.6", "0.8", "0.8", "MCL1608S100", "10", "0.05", "1.170/0.900", "20"),
        ("MCL2012", "0805", "2.0×1.25×0.85 mm", "2.0", "1.25", "0.85", "MCL2012SR10", "0.1", "1.00", "0.091/0.070", "235"),
        ("MCL2012", "0805", "2.0×1.25×0.85 mm", "2.0", "1.25", "0.85", "MCL2012SR22", "0.22", "0.80", "0.169/0.130", "170"),
        ("MCL2012", "0805", "2.0×1.25×0.85 mm", "2.0", "1.25", "0.85", "MCL2012SR47", "0.47", "0.55", "0.234/0.180", "125"),
        ("MCL2012", "0805", "2.0×1.25×0.85 mm", "2.0", "1.25", "0.85", "MCL2012S1R0", "1.0", "0.30", "0.260/0.200", "75"),
        ("MCL2012", "0805", "2.0×1.25×0.85 mm", "2.0", "1.25", "0.85", "MCL2012S2R2", "2.2", "0.22", "0.364/0.280", "50"),
        ("MCL2012", "0805", "2.0×1.25×0.85 mm", "2.0", "1.25", "0.85", "MCL2012S4R7", "4.7", "0.18", "0.390/0.300", "25"),
        ("MCL2012", "0805", "2.0×1.25×0.85 mm", "2.0", "1.25", "0.85", "MCL2012S100", "10", "0.06", "0.650/0.500", "15"),
        ("MCL2012", "0805", "2.0×1.25×0.85 mm", "2.0", "1.25", "0.85", "MCL2012H100", "10", "0.10", "0.650/0.500", "20"),
    ]
    for series, size_inch, size_mm, length_mm, width_mm, height_mm, base_model, inductance_uh, current_a, dcr_ohm, srf_mhz in base_rows:
        for tol_code, tol_pct in (("M", "20"), ("N", "30")):
            rows.append(
                build_sunlord_mcl_seed_row(
                    model=f"{base_model}{tol_code}T",
                    series=series,
                    size_inch=size_inch,
                    size_mm=size_mm,
                    length_mm=length_mm,
                    width_mm=width_mm,
                    height_mm=height_mm,
                    inductance_uh=inductance_uh,
                    tolerance_pct=tol_pct,
                    rated_current_a=current_a,
                    dcr_ohm=dcr_ohm,
                    srf_mhz=srf_mhz,
                )
            )
    return rows


def fill_missing_series_from_model(df):
    if df is None or df.empty or "型号" not in df.columns:
        return df
    work = df.copy()
    if "系列" not in work.columns:
        work["系列"] = ""
    for col in ["系列", "型号", "品牌", "器件类型"]:
        if col in work.columns and isinstance(work[col].dtype, pd.CategoricalDtype):
            work[col] = work[col].astype("object")

    series_text = work["系列"].astype("string").fillna("").apply(clean_text)
    missing_mask = series_text.eq("") | series_text.str.fullmatch(r"[?？]+", na=False) | series_text.isin(SERIES_PLACEHOLDER_TOKENS) | series_text.str.lower().isin({"none", "nan"})

    model_clean = work["型号"].astype("string").fillna("").apply(clean_model)
    brand_clean = work["品牌"].astype("string").fillna("").apply(clean_brand) if "品牌" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")

    if "系列说明" not in work.columns:
        work["系列说明"] = ""
    if "特殊用途" not in work.columns:
        work["特殊用途"] = ""
    fenghua_model_mask = model_clean.str.match(r"^AM\d{2}[A-Z].*", na=False)
    fenghua_brand_mask = brand_clean.str.contains(r"FENGHUA|风华", case=False, regex=True, na=False)
    fenghua_mask = fenghua_model_mask | (fenghua_brand_mask & fenghua_model_mask)
    if fenghua_mask.any():
        fenghua_series_blank = work["系列"].astype("string").fillna("").apply(clean_text).eq("")
        if fenghua_series_blank.any():
            work.loc[fenghua_mask & fenghua_series_blank, "系列"] = "AM"
        fenghua_desc_blank = work["系列说明"].astype("string").fillna("").apply(clean_text).eq("")
        if fenghua_desc_blank.any():
            work.loc[fenghua_mask & fenghua_desc_blank, "系列说明"] = fenghua_am_series_meaning("AM")

    pdc_brand_mask = brand_clean.str.contains(r"PDC|信昌", case=False, regex=True, na=False)
    pdc_series = model_clean.apply(pdc_general_mlcc_series_code_from_model)
    pdc_mask = pdc_brand_mask & pdc_series.ne("")
    if pdc_mask.any():
        work.loc[pdc_mask, "系列"] = pdc_series[pdc_mask]
        if "系列说明" not in work.columns:
            work["系列说明"] = ""
        pdc_series_desc = pdc_series.apply(pdc_mlcc_series_meaning)
        pdc_desc_blank = work["系列说明"].astype("string").fillna("").apply(clean_text).eq("")
        pdc_desc_mask = pdc_mask & pdc_desc_blank & ~pdc_series.isin(["MT", "MG", "MS"])
        if pdc_desc_mask.any():
            work.loc[pdc_desc_mask, "系列说明"] = pdc_series_desc[pdc_desc_mask]

    mlcc_type_mask = (
        work["器件类型"].astype(str).apply(normalize_component_type).eq("MLCC")
        if "器件类型" in work.columns
        else pd.Series([False] * len(work), index=work.index)
    )
    if mlcc_type_mask.any():
        mlcc_placeholder_mask = mlcc_type_mask & work["系列"].astype("string").fillna("").apply(series_looks_missing)
        if mlcc_placeholder_mask.any():
            work.loc[mlcc_placeholder_mask, "系列"] = ""
        mlcc_profile_df = work.loc[mlcc_type_mask, ["品牌", "型号", "系列", "系列说明", "特殊用途"]].apply(
            lambda row: resolve_mlcc_series_profile(
                brand=row.get("品牌", ""),
                model=row.get("型号", ""),
                series=row.get("系列", ""),
                series_desc=row.get("系列说明", ""),
                special_use=row.get("特殊用途", ""),
            ),
            axis=1,
            result_type="expand",
        )
        current_series = work.loc[mlcc_type_mask, "系列"].astype("string").fillna("").apply(clean_text)
        profile_series = mlcc_profile_df["系列"].astype("string").fillna("").apply(clean_text)
        series_replace_mask = current_series.combine(profile_series, mlcc_series_should_replace)
        if series_replace_mask.any():
            replace_idx = series_replace_mask[series_replace_mask].index
            work.loc[replace_idx, "系列"] = profile_series.loc[replace_idx]
        canonical_profile_df = work.loc[mlcc_type_mask, ["品牌", "型号", "系列"]].apply(
            lambda row: resolve_mlcc_series_profile(
                brand=row.get("品牌", ""),
                model=row.get("型号", ""),
                series=row.get("系列", ""),
                series_desc="",
                special_use="",
            ),
            axis=1,
            result_type="expand",
        )
        hre_mask = mlcc_type_mask & brand_clean.str.contains(r"HRE|芯声微", case=False, regex=True, na=False)
        if hre_mask.any():
            hre_idx = hre_mask[hre_mask].index
            work.loc[hre_idx, "系列"] = canonical_profile_df.loc[hre_idx, "系列"].astype("string").fillna("").apply(clean_text)
            work.loc[hre_idx, "系列说明"] = canonical_profile_df.loc[hre_idx, "系列说明"].astype("string").fillna("").apply(clean_text)
            work.loc[hre_idx, "特殊用途"] = canonical_profile_df.loc[hre_idx, "特殊用途"].astype("string").fillna("").apply(clean_text)
            if "_mlcc_series_class" in canonical_profile_df.columns:
                work.loc[hre_idx, "_mlcc_series_class"] = canonical_profile_df.loc[hre_idx, "_mlcc_series_class"].astype("string").fillna("").apply(clean_text)
        if series_replace_mask.any():
            replace_idx = series_replace_mask[series_replace_mask].index
            work.loc[replace_idx, "系列说明"] = canonical_profile_df.loc[replace_idx, "系列说明"].astype("string").fillna("").apply(clean_text)
            work.loc[replace_idx, "特殊用途"] = canonical_profile_df.loc[replace_idx, "特殊用途"].astype("string").fillna("").apply(clean_text)
        for col in ["系列说明", "特殊用途"]:
            current_col = work.loc[mlcc_type_mask, col].astype("string").fillna("").apply(clean_text)
            profile_col = canonical_profile_df[col].astype("string").fillna("").apply(clean_text)
            blank_mask = current_col.eq("") & profile_col.ne("")
            if blank_mask.any():
                replace_idx = blank_mask[blank_mask].index
                work.loc[replace_idx, col] = profile_col.loc[replace_idx]
        current_series_after = work.loc[mlcc_type_mask, "系列"].astype("string").fillna("").apply(clean_text)
        profile_class = canonical_profile_df["_mlcc_series_class"].astype("string").fillna("").apply(clean_text)
        current_desc_after = work.loc[mlcc_type_mask, "系列说明"].astype("string").fillna("").apply(clean_text)
        current_use_after = work.loc[mlcc_type_mask, "特殊用途"].astype("string").fillna("").apply(clean_text)
        current_class_after = pd.Series(
            [
                normalize_mlcc_series_class("/".join(part for part in [desc, use] if clean_text(part) != ""))
                for desc, use in zip(current_desc_after.tolist(), current_use_after.tolist())
            ],
            index=current_desc_after.index,
            dtype="string",
        ).fillna("").apply(clean_text)
        class_refresh_mask = (
            current_series_after.eq(profile_series)
            & profile_series.ne("")
            & profile_class.ne("")
            & current_class_after.ne(profile_class)
        )
        if class_refresh_mask.any():
            refresh_idx = class_refresh_mask[class_refresh_mask].index
            work.loc[refresh_idx, "系列说明"] = canonical_profile_df.loc[refresh_idx, "系列说明"].astype("string").fillna("").apply(clean_text)
            work.loc[refresh_idx, "特殊用途"] = canonical_profile_df.loc[refresh_idx, "特殊用途"].astype("string").fillna("").apply(clean_text)
            current_series_after = work.loc[mlcc_type_mask, "系列"].astype("string").fillna("").apply(clean_text)
        filled_mlcc_mask = mlcc_type_mask & ~current_series_after.apply(series_looks_missing)
        if filled_mlcc_mask.any():
            missing_mask.loc[filled_mlcc_mask] = False

    resistor_type_mask = (
        work["器件类型"].astype(str).apply(normalize_component_type).isin(RESISTOR_COMPONENT_TYPES)
        if "器件类型" in work.columns
        else pd.Series([False] * len(work), index=work.index)
    )
    if resistor_type_mask.any():
        if "特殊用途" not in work.columns:
            work["特殊用途"] = ""

        def resolve_resistor_profile(row):
            profile = infer_resistor_series_profile(
                row.get("型号", ""),
                brand=row.get("品牌", ""),
                component_type=row.get("器件类型", ""),
                special_use=row.get("特殊用途", ""),
            )
            inferred_series = clean_text(profile.get("系列", ""))
            inferred_desc = clean_text(profile.get("系列说明", "")) or build_resistor_series_description(
                brand=row.get("品牌", ""),
                series=inferred_series,
                component_type=row.get("器件类型", ""),
                special_use=row.get("特殊用途", ""),
            )
            inferred_special_use = clean_text(profile.get("特殊用途", ""))
            return pd.Series(
                {
                    "系列": inferred_series,
                    "系列说明": inferred_desc,
                    "特殊用途": inferred_special_use,
                }
            )

        resistor_profile_df = work.loc[
            resistor_type_mask,
            ["品牌", "型号", "器件类型", "特殊用途", "系列", "系列说明"],
        ].apply(resolve_resistor_profile, axis=1, result_type="expand")
        if not resistor_profile_df.empty:
            resistor_profile_df["系列"] = resistor_profile_df["系列"].astype("string").fillna("").apply(clean_text)
            resistor_profile_df["系列说明"] = resistor_profile_df["系列说明"].astype("string").fillna("").apply(clean_text)
            resistor_profile_df["特殊用途"] = resistor_profile_df["特殊用途"].astype("string").fillna("").apply(clean_text)

            current_series = work.loc[resistor_type_mask, "系列"].astype("string").fillna("").apply(clean_text)
            current_desc = work.loc[resistor_type_mask, "系列说明"].astype("string").fillna("").apply(clean_text)
            current_special = work.loc[resistor_type_mask, "特殊用途"].astype("string").fillna("").apply(clean_text)
            profile_series = resistor_profile_df["系列"]
            profile_desc = resistor_profile_df["系列说明"]
            profile_special = resistor_profile_df["特殊用途"]

            series_replace_mask = current_series.combine(profile_series, resistor_series_should_replace)
            if series_replace_mask.any():
                replace_idx = series_replace_mask[series_replace_mask].index
                work.loc[replace_idx, "系列"] = profile_series.loc[replace_idx]
                missing_mask.loc[replace_idx] = False

            desc_refresh_mask = current_desc.combine(profile_desc, resistor_series_desc_should_replace)
            desc_fill_mask = series_replace_mask | desc_refresh_mask
            if desc_fill_mask.any():
                replace_idx = desc_fill_mask[desc_fill_mask].index
                work.loc[replace_idx, "系列说明"] = profile_desc.loc[replace_idx]

            special_fill_mask = (
                series_replace_mask
                | (current_special.eq("") & profile_special.ne(""))
            ) & profile_special.ne("")
            if special_fill_mask.any():
                replace_idx = special_fill_mask[special_fill_mask].index
                work.loc[replace_idx, "特殊用途"] = profile_special.loc[replace_idx]

            current_series_after = work.loc[resistor_type_mask, "系列"].astype("string").fillna("").apply(clean_text)
            filled_resistor_mask = resistor_type_mask & ~current_series_after.apply(series_looks_missing)
            if filled_resistor_mask.any():
                missing_mask.loc[filled_resistor_mask] = False

    series_desc_blank = work["系列说明"].astype("string").fillna("").apply(clean_text).eq("")
    special_use_blank = work["特殊用途"].astype("string").fillna("").apply(clean_text).eq("")
    component_type_series = (
        work["器件类型"].astype("string").fillna("").apply(normalize_component_type)
        if "器件类型" in work.columns
        else pd.Series([""] * len(work), index=work.index, dtype="string")
    )
    series_profile_candidate_mask = (
        model_clean.ne("")
        & component_type_series.isin(ALL_PASSIVE_COMPONENT_TYPES)
        & (
            missing_mask
            | series_desc_blank
            | (
                special_use_blank
                & component_type_series.isin({"热敏电阻", "铝电解电容"} | VARISTOR_COMPONENT_TYPES)
            )
        )
    )
    if series_profile_candidate_mask.any():
        profile_df = work.loc[
            series_profile_candidate_mask,
            ["品牌", "型号", "器件类型", "系列", "系列说明", "特殊用途"],
        ].apply(
            lambda row: resolve_official_series_profile(
                brand=row.get("品牌", ""),
                model=row.get("型号", ""),
                component_type=row.get("器件类型", ""),
                series=row.get("系列", ""),
                series_desc=row.get("系列说明", ""),
                special_use=row.get("特殊用途", ""),
            ),
            axis=1,
            result_type="expand",
        )
        if not profile_df.empty:
            profile_df = profile_df.reindex(columns=["系列", "系列说明", "特殊用途", "器件类型"]).fillna("")
            current_series = work.loc[series_profile_candidate_mask, "系列"].astype("string").fillna("").apply(clean_text)
            current_desc = work.loc[series_profile_candidate_mask, "系列说明"].astype("string").fillna("").apply(clean_text)
            current_special = work.loc[series_profile_candidate_mask, "特殊用途"].astype("string").fillna("").apply(clean_text)
            current_type = component_type_series.loc[series_profile_candidate_mask].astype("string").fillna("").apply(normalize_component_type)

            profile_series = profile_df["系列"].astype("string").fillna("").apply(clean_text)
            profile_desc = profile_df["系列说明"].astype("string").fillna("").apply(clean_text)
            profile_special = profile_df["特殊用途"].astype("string").fillna("").apply(clean_text)
            profile_type = profile_df["器件类型"].astype("string").fillna("").apply(normalize_component_type)

            series_replace_mask = pd.Series(
                [
                    passive_series_should_replace(current, canonical, model=model_clean.at[idx])
                    for idx, current, canonical in zip(current_series.index, current_series.tolist(), profile_series.tolist())
                ],
                index=current_series.index,
                dtype="bool",
            )
            if series_replace_mask.any():
                replace_idx = series_replace_mask[series_replace_mask].index
                work.loc[replace_idx, "系列"] = profile_series.loc[replace_idx]
                missing_mask.loc[replace_idx] = False

            desc_fill_mask = (current_desc.eq("") | series_replace_mask) & profile_desc.ne("")
            if desc_fill_mask.any():
                replace_idx = desc_fill_mask[desc_fill_mask].index
                work.loc[replace_idx, "系列说明"] = profile_desc.loc[replace_idx]

            special_fill_mask = (current_special.eq("") | series_replace_mask) & profile_special.ne("")
            if special_fill_mask.any():
                replace_idx = special_fill_mask[special_fill_mask].index
                work.loc[replace_idx, "特殊用途"] = profile_special.loc[replace_idx]

            if "器件类型" in work.columns:
                type_fill_mask = current_type.eq("") & profile_type.ne("")
                if type_fill_mask.any():
                    replace_idx = type_fill_mask[type_fill_mask].index
                    work.loc[replace_idx, "器件类型"] = profile_type.loc[replace_idx]

    if not missing_mask.any():
        return work

    def assign_series(pattern, extra_mask=None):
        nonlocal missing_mask
        candidate_mask = missing_mask & model_clean.str.match(pattern, na=False)
        if extra_mask is not None:
            candidate_mask &= extra_mask
        if not candidate_mask.any():
            return
        extracted = model_clean[candidate_mask].str.extract(pattern, expand=False).fillna("").astype("string").str.strip()
        valid_mask = extracted.ne("")
        if not valid_mask.any():
            return
        idx = extracted[valid_mask].index
        work.loc[idx, "系列"] = extracted.loc[idx]
        missing_mask.loc[idx] = False

    murata_brand_mask = brand_clean.str.contains(r"MURATA|村田|东电化", case=False, regex=True, na=False)
    murata_series = model_clean.apply(murata_series_code_from_model)
    murata_mask = murata_brand_mask & murata_series.ne("")
    if murata_mask.any():
        work.loc[murata_mask, "系列"] = murata_series[murata_mask]
        missing_mask.loc[murata_mask] = False

    nichicon_brand_mask = brand_clean.str.contains(r"NICHICON|尼吉康", case=False, regex=True, na=False)
    nichicon_series = model_clean.apply(nichicon_series_code_from_model)
    nichicon_mask = nichicon_brand_mask & nichicon_series.ne("")
    if nichicon_mask.any():
        work.loc[nichicon_mask, "系列"] = nichicon_series[nichicon_mask]
        missing_mask.loc[nichicon_mask] = False

    chemi_con_brand_mask = brand_clean.str.contains(r"CHEMI-CON|贵弥功|日本贵弥功|NIPPON", case=False, regex=True, na=False)
    chemi_con_series = model_clean.apply(chemi_con_series_code_from_model)
    chemi_con_mask = chemi_con_brand_mask & chemi_con_series.ne("")
    if chemi_con_mask.any():
        work.loc[chemi_con_mask, "系列"] = chemi_con_series[chemi_con_mask]
        missing_mask.loc[chemi_con_mask] = False

    jianghai_brand_mask = brand_clean.str.contains(r"JIANGHAI|江海|南通江海", case=False, regex=True, na=False)
    jianghai_series = model_clean.apply(jianghai_series_code_from_model)
    jianghai_mask = jianghai_brand_mask & jianghai_series.ne("")
    if jianghai_mask.any():
        work.loc[jianghai_mask, "系列"] = jianghai_series[jianghai_mask]
        jianghai_desc_blank = work["系列说明"].astype("string").fillna("").apply(clean_text).eq("")
        jianghai_desc_mask = jianghai_mask & jianghai_desc_blank
        if jianghai_desc_mask.any():
            jianghai_desc = jianghai_series[jianghai_desc_mask].combine(
                model_clean[jianghai_desc_mask],
                lambda series, model: jianghai_series_meaning(series, model),
            )
            work.loc[jianghai_desc_mask, "系列说明"] = jianghai_desc
        missing_mask.loc[jianghai_mask] = False

    fenghua_mask = fenghua_model_mask | (fenghua_brand_mask & fenghua_model_mask)
    if fenghua_mask.any():
        work.loc[fenghua_mask, "系列"] = "AM"
        fenghua_desc_mask = fenghua_mask & work["系列说明"].astype("string").fillna("").apply(clean_text).eq("")
        if fenghua_desc_mask.any():
            work.loc[fenghua_desc_mask, "系列说明"] = fenghua_am_series_meaning("AM")
        missing_mask.loc[fenghua_mask] = False

    pdc_series_profile = model_clean.apply(pdc_mlcc_series_profile_from_model)
    pdc_series_code = pdc_series_profile.apply(lambda item: item[0] if isinstance(item, tuple) and len(item) >= 1 else "")
    pdc_series_desc = pdc_series_profile.apply(lambda item: item[1] if isinstance(item, tuple) and len(item) >= 2 else "").astype("string").fillna("")
    pdc_special_desc = pdc_series_profile.apply(lambda item: pdc_mlcc_special_control_meaning(item[4]) if isinstance(item, tuple) and len(item) >= 5 else "").astype("string").fillna("")
    pdc_series_desc = pdc_series_desc.mask(
        pdc_special_desc.ne("") & pdc_series_desc.eq(""),
        pdc_special_desc,
    )
    pdc_series_desc = pdc_series_desc.mask(
        pdc_special_desc.ne("") & pdc_series_desc.ne(""),
        pdc_series_desc + " / " + pdc_special_desc,
    )
    pdc_mask = pdc_brand_mask & pdc_series_code.ne("")
    if pdc_mask.any():
        work.loc[pdc_mask, "系列"] = pdc_series_code[pdc_mask]
        if "系列说明" not in work.columns:
            work["系列说明"] = ""
        pdc_desc_mask = pdc_mask & work["系列说明"].astype("string").fillna("").apply(clean_text).eq("")
        if pdc_desc_mask.any():
            work.loc[pdc_desc_mask, "系列说明"] = pdc_series_desc[pdc_desc_mask]
        missing_mask.loc[pdc_mask] = False

    walsin_brand_mask = brand_clean.str.contains(r"WALSIN|华新科", case=False, regex=True, na=False)
    assign_series(r"^(01R5)", extra_mask=walsin_brand_mask)
    assign_series(r"^((?:RF|HH|SH|RT|UF))(?:\d{2})", extra_mask=walsin_brand_mask)

    # MLCC / ceramic families with well-defined official naming prefixes.
    assign_series(r"^((?:GRM|GCM|GCJ|GJM|GQM|GRT|GCG|GCQ|GRJ|GMA|GMD|GCH|GXT|GGM|GC3|GCD|GCE|GGD|LLL|LLF|LLA|LLG|LLC|NFM|KCM|KRT|DK1|GA2|GA3|GR3|GR4|GR7|GJ4|KRM|KR3|KR9|ZRA|ZRB|NTC|PRF|PTG|NXF|CEU|CGJ|CGB|CKG|CNA|CNC|CN0|YNA|RHEL|RPER|ERB|RCE|RDE|RHE|RPE))")
    assign_series(r"^(AVR-M)")
    assign_series(r"^(AVRM)")
    assign_series(r"^(AVRL)")
    assign_series(r"^(VAR)")
    assign_series(r"^(S\d{2}K)")
    assign_series(r"^(LS\d{2}K)")
    assign_series(r"^(SGN[A-Z0-9]{4})")
    assign_series(r"^(CN\d{4})")
    assign_series(r"^(CL)")
    assign_series(r"^((?:CC|CQ|AC|AQ|AS|CS))")
    assign_series(r"^((?:UMK|HMK|QMK|SMK|QVS|TVS|TMK|JMK|EMK|LMK|AMK))")
    assign_series(r"^((?:MSAY|MSRL|MSART|MBARQ|MCARQ|MMARQ|MAAS|MBAS|MCAS|MCAST|MLAS|MMAS|MSAS))")
    assign_series(r"^(TCC)")
    assign_series(r"^(CGA)")
    assign_series(r"^(CSA)")
    assign_series(r"^((?:MT|FP|FS|FN|FM|FV|FK|FH))(?:\d{2})")
    assign_series(r"^(ECR\d[A-Z]{3})")
    assign_series(r"^(ECA\d[A-Z]{3})")
    assign_series(r"^(EEE\d[A-Z]{3})")
    assign_series(r"^(EEU\d[A-Z]{3})")
    assign_series(r"^(ECL\d[A-Z]{3})")
    assign_series(r"^(EKX\d[A-Z]{3})")
    assign_series(r"^(EKY\d[A-Z]{3})")
    assign_series(r"^(ECP\d[A-Z]{3})")
    assign_series(r"^(FCS\d[A-Z]{3})")
    assign_series(r"^(FCR\d[A-Z]{3})")
    assign_series(r"^(FCP\d[A-Z]{3})")

    # Resistors and thermistors / varistors.
    assign_series(r"^(ERJ-[A-Z0-9]{3,4})")
    assign_series(r"^(ERJ\d{2}[A-Z]{2})")
    assign_series(r"^(ERJ[A-Z]{2}\d)")
    assign_series(r"^(ERJ\d[A-Z]{2})")
    assign_series(r"^(ERA\d{2}[A-Z])")
    assign_series(r"^(ERA\d[A-Z])")
    assign_series(r"^((?:RN73|RK73|WK73)(?:[A-Z]\d[A-Z]|\d[A-Z]))")
    assign_series(r"^(RNCF\d{4})")
    assign_series(r"^(RNCP\d{4})")
    assign_series(r"^(RMCF\d{4})")
    assign_series(r"^(RNCS\d{4})")
    assign_series(r"^(RTAN\d{4})")
    assign_series(r"^(RMEF\d{4})")
    assign_series(r"^(HMC\d{4})")
    assign_series(r"^(HVC\d{4})")
    assign_series(r"^(HVCB\d{4})")
    assign_series(r"^(CRM\d{4})")
    assign_series(r"^(FCR\d{4})")
    assign_series(r"^(SM\d{4})")
    assign_series(r"^(HCJ\d{4})")
    assign_series(r"^(MLFA\d)")
    assign_series(r"^(RPC\d{4})")
    assign_series(r"^(RMC\d/\d{2})")
    assign_series(r"^(RF73\d[A-Z])")
    assign_series(r"^(RK73[A-Z]{2}\d[A-Z])")
    assign_series(r"^(CPF\d{4})")
    assign_series(r"^(CRGH\d{4})")
    assign_series(r"^(CRGP\d{4})")
    assign_series(r"^(RP73PF\d[A-Z])")
    assign_series(r"^((?:OMM|MMA|MMU)\d{4})")
    assign_series(r"^((?:AA|AC|AF|AR|AT|RC|RT|RL|RS|RK|RP|RA)\d{4})")
    assign_series(r"^(ERJP\d{2}[A-Z])")
    assign_series(r"^(ERA-\d[A-Z])")
    assign_series(r"^(ERJ\d{2}[A-Z])")
    assign_series(r"^((?:WW)\d{2}WR)")
    assign_series(r"^(RN(?:55|60|65|70|75))")
    assign_series(r"^(RNC\d{2,4})")
    assign_series(r"^(RMC\d{2,4})")
    assign_series(r"^(RNR\d{1,2})")
    assign_series(r"^(RNX\d)")
    assign_series(r"^(RGC\d)")
    assign_series(r"^(RCP\d?)")
    assign_series(r"^(RG\d)")
    assign_series(r"^((?:CRCW|TNPW|SMM|PAT)\d{4})")
    assign_series(r"^(M55342)")
    assign_series(r"^(D55342)")
    assign_series(r"^(MBB)")
    assign_series(r"^(MBA)")
    assign_series(r"^(MCT)")
    assign_series(r"^(PHP)")
    assign_series(r"^((?:RNCF|RMCF|RNCS|RTAN|HMC|HVCB|CRM|RMEF)\d{4})")
    assign_series(r"^((?:CR|CRM)\d{4})")
    assign_series(r"^((?:MCR|LTR|PMR|KTR|ESR|SDR|SFR|UCR)\d{2})")
    assign_series(r"^((?:WR|WF|MR)\d{2}[A-Z])")
    assign_series(r"^((?:RS|RC)-\d{2}[A-Z])")
    assign_series(r"^(MHR\d{4})")
    assign_series(r"^(PRG\d{2})")
    assign_series(r"^(MF72)")
    assign_series(r"^(CL-\d{2})")
    assign_series(r"^(B\d{5,6})")

    # Small inductive / timing families that still follow a readable prefix rule.
    assign_series(r"^((?:LQH|LQW|LQM|LQG)\d{2,3})")

    return work


MANUAL_CORRECTION_RULE_COLUMNS = [
    "启用",
    "优先级",
    "匹配方式",
    "品牌",
    "型号匹配",
    "器件类型",
    "尺寸（inch）",
    "材质（介质）",
    "参数值",
    "参数单位",
    "参数误差",
    "耐压（V）",
    "压敏电压",
    "功率",
    "脚距",
    "安规",
    "备注",
]

MANUAL_CORRECTION_RULE_MATCH_MODES = ["精确", "前缀", "包含", "正则"]


def manual_correction_rules_default_df():
    return pd.DataFrame([
        {
            "启用": False,
            "优先级": 100,
            "匹配方式": "精确",
            "品牌": "信昌PDC",
            "型号匹配": "MT43X104K152EHE",
            "器件类型": "MLCC",
            "尺寸（inch）": "1812",
            "材质（介质）": "X7R",
            "参数值": "100",
            "参数单位": "NF",
            "参数误差": "10%",
            "耐压（V）": "1500V",
            "压敏电压": "",
            "功率": "",
            "脚距": "",
            "安规": "",
            "备注": "示例：按信昌原厂命名规则解析为 1812 / X7R / 100nF / 1500V",
        }
    ], columns=MANUAL_CORRECTION_RULE_COLUMNS)


def get_manual_correction_rules_signature():
    if not os.path.exists(MANUAL_CORRECTION_RULES_PATH):
        return {
            "manual_rules_missing": True,
            "manual_rules_version": MANUAL_CORRECTION_RULES_VERSION,
        }
    stat = os.stat(MANUAL_CORRECTION_RULES_PATH)
    return {
        "manual_rules_path": MANUAL_CORRECTION_RULES_PATH,
        "manual_rules_mtime": round(stat.st_mtime, 6),
        "manual_rules_size": stat.st_size,
        "manual_rules_version": MANUAL_CORRECTION_RULES_VERSION,
    }


@st.cache_data(ttl=60)
def _load_manual_correction_rules_cached(cache_signature):
    if not os.path.exists(MANUAL_CORRECTION_RULES_PATH):
        return manual_correction_rules_default_df()
    try:
        df = pd.read_csv(MANUAL_CORRECTION_RULES_PATH, encoding="utf-8-sig")
    except Exception:
        return manual_correction_rules_default_df()
    return normalize_manual_correction_rules_dataframe(df)


def load_manual_correction_rules():
    return _load_manual_correction_rules_cached(json.dumps(get_manual_correction_rules_signature(), sort_keys=True, ensure_ascii=True))


def manual_rule_has_effect(row):
    if row is None:
        return False
    if clean_text(row.get("型号匹配", "")) == "":
        return False
    effect_columns = [
        "器件类型", "尺寸（inch）", "材质（介质）", "参数值", "参数单位", "参数误差",
        "耐压（V）", "压敏电压", "功率", "脚距", "安规",
    ]
    return any(clean_text(row.get(col, "")) != "" for col in effect_columns)


def save_manual_correction_rules(df):
    normalized = normalize_manual_correction_rules_dataframe(df)
    if normalized is None:
        return pd.DataFrame()
    save_rows = normalized[normalized.apply(manual_rule_has_effect, axis=1)].copy()
    if save_rows.empty:
        save_rows = normalized.iloc[0:0].copy()
    rules_dir = os.path.dirname(MANUAL_CORRECTION_RULES_PATH)
    if rules_dir:
        os.makedirs(rules_dir, exist_ok=True)
    save_rows.to_csv(MANUAL_CORRECTION_RULES_PATH, index=False, encoding="utf-8-sig")
    return save_rows


def normalize_manual_correction_rules_dataframe(df):
    if df is None or df.empty:
        return manual_correction_rules_default_df()
    work = df.copy()
    for col in MANUAL_CORRECTION_RULE_COLUMNS:
        if col not in work.columns:
            work[col] = ""
    work = work[MANUAL_CORRECTION_RULE_COLUMNS].copy()
    work["启用"] = work["启用"].apply(lambda value: bool(value) if str(value).strip() != "" else False)
    work["优先级"] = pd.to_numeric(work["优先级"], errors="coerce").fillna(100).astype(int)
    work["匹配方式"] = work["匹配方式"].apply(normalize_manual_rule_match_mode)
    work["品牌"] = work["品牌"].apply(clean_brand)
    work["型号匹配"] = work["型号匹配"].apply(clean_model)
    work["器件类型"] = work["器件类型"].apply(normalize_component_type)
    work["尺寸（inch）"] = work["尺寸（inch）"].apply(clean_size)
    work["材质（介质）"] = work["材质（介质）"].apply(clean_material)
    work["参数值"] = work["参数值"].apply(clean_text)
    work["参数单位"] = work["参数单位"].apply(lambda value: clean_text(value).upper())
    work["参数误差"] = work["参数误差"].apply(clean_tol_for_match)
    work["耐压（V）"] = work["耐压（V）"].apply(clean_voltage)
    work["压敏电压"] = work["压敏电压"].apply(clean_voltage)
    work["功率"] = work["功率"].apply(format_power_display)
    work["脚距"] = work["脚距"].apply(clean_text)
    work["安规"] = work["安规"].apply(clean_text)
    work["备注"] = work["备注"].apply(clean_text)
    work = work.drop_duplicates(subset=["品牌", "型号匹配", "匹配方式", "器件类型"], keep="first").reset_index(drop=True)
    return work


def normalize_manual_rule_match_mode(value):
    text = clean_text(value)
    if text == "":
        return "精确"
    mapping = {
        "exact": "精确",
        "精准": "精确",
        "精确": "精确",
        "prefix": "前缀",
        "前缀": "前缀",
        "contains": "包含",
        "包含": "包含",
        "regex": "正则",
        "regexp": "正则",
        "正则": "正则",
    }
    return mapping.get(text.lower(), mapping.get(text, "精确"))


def has_active_manual_correction_rules():
    rules = load_manual_correction_rules()
    if rules is None or rules.empty:
        return False
    try:
        return bool(rules["启用"].astype(bool).any())
    except Exception:
        return False


def manual_rule_matches_text(model_text, brand_text, rule_row):
    rule_model = clean_text(rule_row.get("型号匹配", ""))
    if rule_model == "":
        return False
    rule_brand = clean_brand(rule_row.get("品牌", ""))
    model_clean = clean_model(model_text)
    brand_clean = clean_brand(brand_text)
    match_mode = normalize_manual_rule_match_mode(rule_row.get("匹配方式", ""))
    if rule_brand:
        brand_ok = (
            rule_brand in brand_clean
            or brand_clean in rule_brand
            or rule_brand.upper() in brand_clean.upper()
            or brand_clean.upper() in rule_brand.upper()
        )
        if not brand_ok:
            return False
    if match_mode == "正则":
        try:
            return re.search(rule_model, clean_text(model_text), flags=re.I) is not None or re.search(rule_model, model_clean, flags=re.I) is not None
        except Exception:
            return False
    rule_model_clean = clean_model(rule_model)
    if match_mode == "前缀":
        return model_clean.startswith(rule_model_clean)
    if match_mode == "包含":
        return rule_model_clean in model_clean
    return model_clean == rule_model_clean


def manual_rule_row_to_parsed_rule(rule_row):
    component_type = normalize_component_type(rule_row.get("器件类型", ""))
    parsed = {
        "器件类型": component_type,
        "_model_rule_authority": f"manual_rule:{normalize_manual_rule_match_mode(rule_row.get('匹配方式', ''))}",
    }
    for col in ["尺寸（inch）", "材质（介质）", "容值误差", "耐压（V）"]:
        value = clean_text(rule_row.get(col, ""))
        if value != "":
            parsed[col] = clean_size(value) if col == "尺寸（inch）" else clean_material(value) if col == "材质（介质）" else clean_tol_for_match(value) if col == "容值误差" else clean_voltage(value)
    power = clean_text(rule_row.get("功率", ""))
    if power != "":
        parsed["功率"] = format_power_display(power)
        parsed["_power"] = power
    pitch = clean_text(rule_row.get("脚距", ""))
    if pitch != "":
        parsed["脚距"] = pitch
        parsed["_pitch"] = pitch
    safety = clean_text(rule_row.get("安规", ""))
    if safety != "":
        parsed["安规"] = safety
        parsed["_safety_class"] = safety
    varistor_voltage = clean_voltage(rule_row.get("压敏电压", ""))
    if varistor_voltage != "":
        parsed["压敏电压"] = voltage_display(varistor_voltage)
        parsed["_varistor_voltage"] = varistor_voltage
        parsed["耐压（V）"] = voltage_display(varistor_voltage)
    param_value = clean_text(rule_row.get("参数值", ""))
    param_unit = clean_text(rule_row.get("参数单位", "")).upper()
    if component_type in RESISTOR_COMPONENT_TYPES or component_type == "热敏电阻":
        resistance_text = "".join([param_value, param_unit])
        parsed_resistance = find_resistance_in_text(resistance_text) if resistance_text != "" else None
        if parsed_resistance is None and param_value != "":
            parsed_resistance = parse_resistor_value_code(param_value)
        if parsed_resistance is not None:
            parsed["_resistance_ohm"] = parsed_resistance
            value, unit = ohm_to_library_value_unit(parsed_resistance)
            parsed["容值"] = value
            parsed["容值单位"] = unit
    elif component_type in CAPACITOR_COMPONENT_TYPES:
        pf = cap_to_pf(param_value, param_unit)
        if pf is not None:
            parsed["容值_pf"] = pf
            value, unit = pf_to_value_unit(pf)
            parsed["容值"] = value
            parsed["容值单位"] = unit
    elif component_type in INDUCTOR_COMPONENT_TYPES or component_type in TIMING_COMPONENT_TYPES:
        if param_value != "":
            parsed["容值"] = param_value
        if param_unit != "":
            parsed["容值单位"] = param_unit
    elif component_type in VARISTOR_COMPONENT_TYPES:
        if varistor_voltage == "" and param_value != "":
            varistor_voltage = clean_voltage(param_value if param_unit in {"", "V"} else f"{param_value}{param_unit}")
            if varistor_voltage != "":
                parsed["压敏电压"] = voltage_display(varistor_voltage)
                parsed["_varistor_voltage"] = varistor_voltage
                parsed["耐压（V）"] = voltage_display(varistor_voltage)
    else:
        if param_value != "":
            parsed["容值"] = param_value
        if param_unit != "":
            parsed["容值单位"] = param_unit
    return parsed


def find_matching_manual_correction_rule(model, brand="", component_type=""):
    rules = load_manual_correction_rules()
    if rules is None or rules.empty:
        return None
    model_text = clean_text(model)
    brand_text = clean_text(brand)
    active_rules = rules[rules["启用"].astype(bool)].copy()
    if active_rules.empty:
        return None
    active_rules = active_rules.sort_values(by=["优先级"], ascending=False, kind="stable")
    for _, rule_row in active_rules.iterrows():
        if manual_rule_matches_text(model_text, brand_text, rule_row):
            return rule_row.to_dict()
    return None


def apply_manual_correction_rules_to_dataframe(df, override_conflicts=True):
    if df is None or df.empty or "型号" not in df.columns:
        return df
    rules = load_manual_correction_rules()
    if rules is None or rules.empty:
        return df
    active_rules = rules[rules["启用"].astype(bool)].copy()
    if active_rules.empty:
        return df
    work = df.copy()
    if "_model_rule_authority" not in work.columns:
        work["_model_rule_authority"] = ""
    brand_series = work["品牌"].astype(str).apply(clean_brand) if "品牌" in work.columns else pd.Series([""] * len(work), index=work.index)
    model_series = work["型号"].astype(str).apply(clean_model)
    applied_mask = pd.Series(False, index=work.index)
    active_rules = active_rules.sort_values(by=["优先级"], ascending=False, kind="stable")
    for _, rule_row in active_rules.iterrows():
        rule_model = clean_text(rule_row.get("型号匹配", ""))
        if rule_model == "":
            continue
        match_mode = normalize_manual_rule_match_mode(rule_row.get("匹配方式", ""))
        rule_brand = clean_brand(rule_row.get("品牌", ""))
        if match_mode == "正则":
            try:
                model_mask = model_series.str.contains(rule_model, regex=True, na=False, flags=re.I)
            except Exception:
                model_mask = pd.Series(False, index=work.index)
        else:
            rule_model_clean = clean_model(rule_model)
            if match_mode == "前缀":
                model_mask = model_series.str.startswith(rule_model_clean, na=False)
            elif match_mode == "包含":
                model_mask = model_series.str.contains(re.escape(rule_model_clean), na=False)
            else:
                model_mask = model_series.eq(rule_model_clean)
        if rule_brand != "":
            brand_mask = brand_series.apply(
                lambda value: (
                    rule_brand in value
                    or value in rule_brand
                    or rule_brand.upper() in value.upper()
                    or value.upper() in rule_brand.upper()
                )
            )
            model_mask = model_mask & brand_mask
        if applied_mask.any():
            model_mask = model_mask & (~applied_mask)
        hit_idx = work.index[model_mask].tolist()
        if not hit_idx:
            continue
        parsed_rule = manual_rule_row_to_parsed_rule(rule_row)
        for idx in hit_idx:
            merged = merge_parsed_rule_into_record(
                work.loc[idx].to_dict(),
                parsed_rule,
                override_conflicts=override_conflicts,
            )
            for col, value in merged.items():
                work.at[idx, col] = value
        applied_mask.loc[hit_idx] = True
    return work


def build_manual_rule_match_summary(rule_row):
    if rule_row is None:
        return ""
    pieces = []
    brand = clean_text(rule_row.get("品牌", ""))
    model = clean_text(rule_row.get("型号匹配", ""))
    mode = normalize_manual_rule_match_mode(rule_row.get("匹配方式", ""))
    if brand:
        pieces.append(f"品牌={brand}")
    if model:
        pieces.append(f"{mode}匹配={model}")
    ctype = clean_text(rule_row.get("器件类型", ""))
    if ctype:
        pieces.append(f"器件类型={ctype}")
    return "，".join(pieces)


def build_pdc_mt_rule_breakdown(model, parsed=None):
    model_clean = clean_model(model)
    mt_match = re.fullmatch(
        r"(?P<prefix>MT|MG|MS)(?P<size>\d{2})(?P<mat>[BCNXT])(?P<cap>(?:\d{3,4}|R\d+))(?P<tol>[FGJKMZ])(?P<volt>(?:6R3|\d{3}))(?P<rest>.*)",
        model_clean,
    )
    if mt_match is None:
        return pd.DataFrame([
            {"段位": "型号", "原始片段": model_clean, "含义": "未能按 PDC MT 结构继续拆分", "结果": ""}
        ])
    material_map = {"B": "X5R", "X": "X7R", "T": "X7T", "N": "COG(NPO)"}
    tol_map = {"F": "1%", "G": "2%", "J": "5%", "K": "10%", "M": "20%", "Z": "-20%~+80%"}
    suffix = clean_text(mt_match.group("rest"))
    package_code = suffix[0] if len(suffix) >= 1 else ""
    thickness_code = suffix[1] if len(suffix) >= 2 else ""
    special_code = suffix[2] if len(suffix) >= 3 else ""
    series_code = clean_text(mt_match.group("prefix"))
    series_desc = clean_text(parsed.get("系列说明", "")) if parsed else ""
    if series_desc == "":
        series_desc = pdc_mlcc_series_meaning(series_code)
    special_meaning = pdc_mlcc_special_control_meaning(special_code)
    if series_desc == "" and special_meaning != "":
        series_desc = special_meaning
    rows = [
        {"段位": "系列前缀", "原始片段": series_code, "含义": "信昌 PDC 系列代号", "结果": series_code},
        {"段位": "系列说明", "原始片段": series_desc, "含义": "系列类别 / 产品定位", "结果": series_desc},
        {"段位": "尺寸码", "原始片段": mt_match.group("size"), "含义": PDC_MT_SIZE_CODE_MAP.get(mt_match.group("size"), "未知尺寸"), "结果": clean_size(parsed.get("尺寸（inch）", "")) if parsed else ""},
        {"段位": "介质码", "原始片段": mt_match.group("mat"), "含义": material_map.get(mt_match.group("mat"), "未知介质"), "结果": clean_material(parsed.get("材质（介质）", "")) if parsed else ""},
        {"段位": "容值码", "原始片段": mt_match.group("cap"), "含义": f"{mt_match.group('cap')} -> {pf_to_value_unit(murata_cap_code_to_pf(mt_match.group('cap')))[0]}{pf_to_value_unit(murata_cap_code_to_pf(mt_match.group('cap')))[1]}", "结果": ""},
        {"段位": "容差码", "原始片段": mt_match.group("tol"), "含义": tol_map.get(mt_match.group("tol"), "未知容差"), "结果": clean_tol_for_match(parsed.get("容值误差", "")) if parsed else ""},
        {"段位": "耐压码", "原始片段": mt_match.group("volt"), "含义": f"{decode_pdc_voltage_code(mt_match.group('volt'))}V", "结果": voltage_display(parsed.get("耐压（V）", "")) if parsed else ""},
    ]
    if package_code:
        rows.append({"段位": "包装码", "原始片段": package_code, "含义": PDC_MT_PACKAGE_CODE_MAP.get(package_code, "未知包装"), "结果": package_code})
    if thickness_code:
        rows.append({
            "段位": "厚度码",
            "原始片段": thickness_code,
            "含义": PDC_MT_THICKNESS_CODE_MAP.get(thickness_code, "未知厚度"),
            "结果": clean_text(parsed.get("高度（mm）", "")) if parsed else "",
        })
    if special_code:
        rows.append({"段位": "特性码", "原始片段": special_code, "含义": special_meaning or "未知特性", "结果": special_code})
    if suffix and len(suffix) > 3:
        rows.append({"段位": "后缀余量", "原始片段": suffix[3:], "含义": "额外控制码", "结果": suffix[3:]})
    return pd.DataFrame(rows)


def build_murata_ntc_rule_breakdown(model, parsed=None):
    model_clean = clean_model(model)
    series_code = murata_ntc_series_code_from_model(model_clean)
    if not series_code:
        return pd.DataFrame([
            {"段位": "型号", "原始片段": model_clean, "含义": "Murata NTC 官方系列代码未能从型号前缀识别", "结果": ""}
        ])

    match = MURATA_NTC_MODEL_PATTERN.fullmatch(model_clean)
    rows = [{
        "段位": "系列前缀",
        "原始片段": series_code,
        "含义": murata_series_meaning(series_code),
        "结果": series_code,
    }]
    if match is not None:
        size_code = match.group("series")[3:5]
        size_label = format_murata_ntc_size_label(size_code)
        resistance_code = match.group("res")
        tol_code = match.group("tol")
        tail = match.group("tail")
        resistance_value = clean_text(parsed.get("容值", "")) if parsed is not None else ""
        resistance_unit = clean_text(parsed.get("容值单位", "")) if parsed is not None else ""
        if (resistance_value == "" or resistance_unit == "") and parsed is not None:
            extracted_ohm = parse_resistor_value_code(resistance_code)
            if extracted_ohm is None:
                extracted_ohm, _ = extract_resistor_value_from_model(model_clean, size_hint=clean_size(parsed.get("尺寸（inch）", "")))
            if extracted_ohm is not None:
                resistance_value, resistance_unit = ohm_to_library_value_unit(extracted_ohm)
        rows.extend([
            {
                "段位": "尺寸码",
                "原始片段": size_code,
                "含义": size_label or "未知尺寸",
                "结果": clean_size(parsed.get("尺寸（inch）", "")) if parsed is not None else "",
            },
            {
                "段位": "电阻码",
                "原始片段": resistance_code,
                "含义": "R25 额定阻值编码",
                "结果": f"{resistance_value}{resistance_unit}" if resistance_value and resistance_unit else "",
            },
            {
                "段位": "容差码",
                "原始片段": tol_code,
                "含义": clean_tol_for_match(parsed.get("容值误差", "")) if parsed is not None else clean_tol_for_match(tol_code),
                "结果": clean_tol_for_match(parsed.get("容值误差", "")) if parsed is not None else clean_tol_for_match(tol_code),
            },
        ])
        if clean_text(tail) != "":
            rows.append({
                "段位": "后缀",
                "原始片段": tail,
                "含义": "包装 / 版本 / 额外控制码",
                "结果": clean_text(tail),
            })
    else:
        rows.append({
            "段位": "后续编码",
            "原始片段": model_clean[len(series_code):],
            "含义": "Murata NTC 系列内部的阻值 / 容差 / 版本编码",
            "结果": "",
        })
    return pd.DataFrame(rows)


def build_murata_rule_breakdown(model, parsed=None):
    model_clean = clean_model(model)
    if murata_ntc_series_code_from_model(model_clean):
        return build_murata_ntc_rule_breakdown(model_clean, parsed=parsed)
    series_code = murata_series_code_from_model(model_clean)
    if not series_code:
        return pd.DataFrame([
            {"段位": "型号", "原始片段": model_clean, "含义": "Murata 官方系列代码未能从型号前缀识别", "结果": ""}
        ])
    match = re.fullmatch(r"(?P<prefix>GRM|GCM|GCJ|GJM|GQM|GRT|GCG|GCQ)(?P<size>\d{2})(?P<thickness>.)(?P<mat>[A-Z0-9]{2})(?P<volt>[0-9A-Z]{2})(?P<cap>(?:\d{3,4}|R\d+))(?P<tol>[BCDFGJKMZ]).*", model_clean)
    size_map = {"02": "01005", "03": "0201", "15": "0402", "18": "0603", "21": "0805", "31": "1206", "32": "1210", "42": "1808", "43": "1812", "55": "2220"}
    material_map = {"B1": "COG(NPO)", "C1": "COG(NPO)", "5C": "COG(NPO)", "R6": "X5R", "R7": "X7R", "R9": "X8R", "C6": "X5S", "C7": "X7S", "C8": "X6S", "D6": "X5T", "D7": "X7T", "D8": "X6T", "E7": "X7U", "L8": "X8L", "M8": "X8M", "N8": "X8N", "U2": "U2J", "7U": "U2J", "Z7": "X7R"}
    rows = [{
        "段位": "系列前缀",
        "原始片段": series_code,
        "含义": murata_series_meaning(series_code),
        "结果": series_code,
    }]
    if match is not None:
        rows.extend([
            {"段位": "尺寸码", "原始片段": match.group("size"), "含义": size_map.get(match.group("size"), "未知尺寸"), "结果": clean_size(parsed.get("尺寸（inch）", "")) if parsed else ""},
            {"段位": "厚度码", "原始片段": match.group("thickness"), "含义": "厚度/系列内部结构码", "结果": clean_text(match.group("thickness"))},
            {"段位": "介质码", "原始片段": match.group("mat"), "含义": material_map.get(match.group("mat"), "未知介质"), "结果": clean_material(parsed.get("材质（介质）", "")) if parsed else ""},
            {"段位": "耐压码", "原始片段": match.group("volt"), "含义": clean_voltage(parsed.get("耐压（V）", "")) + "V" if parsed else "", "结果": voltage_display(parsed.get("耐压（V）", "")) if parsed else ""},
            {"段位": "容值码", "原始片段": match.group("cap"), "含义": f"{murata_cap_code_to_pf(match.group('cap'))}pF", "结果": ""},
            {"段位": "容差码", "原始片段": match.group("tol"), "含义": clean_tol_for_match(match.group("tol")), "结果": clean_tol_for_match(parsed.get("容值误差", "")) if parsed else ""},
        ])
    else:
        rows.append({
            "段位": "后续编码",
            "原始片段": model_clean[len(series_code):],
            "含义": "Murata 系列内部的尺寸 / 介质 / 耐压 / 版本编码",
            "结果": ""
        })
    return pd.DataFrame(rows)


def build_resistor_rule_breakdown(model, parsed=None):
    model_clean = clean_model(model)
    rows = [{"段位": "型号", "原始片段": model_clean, "含义": "电阻命名通常以尺寸、阻值、容差、功率等字段组成", "结果": ""}]
    if parsed is not None:
        series = clean_text(parsed.get("系列", ""))
        series_desc = clean_text(parsed.get("系列说明", ""))
        if series == "":
            series_profile = infer_resistor_series_profile(model_clean, brand=parsed.get("品牌", ""), component_type=parsed.get("器件类型", ""))
            series = series_profile["系列"]
            if series_desc == "":
                series_desc = series_profile["系列说明"]
        if series_desc == "":
            series_desc = build_series_description_fallback(
                parsed.get("器件类型", ""),
                brand=parsed.get("品牌", ""),
                series=series,
                special_use=parsed.get("特殊用途", ""),
            )
        rows.extend([
            {"段位": "系列前缀", "原始片段": series, "含义": "电阻系列代码 / 命名前缀", "结果": series},
            {"段位": "系列说明", "原始片段": series_desc, "含义": "系列类别 / 产品定位", "结果": series_desc},
            {"段位": "尺寸（inch）", "原始片段": clean_size(parsed.get("尺寸（inch）", "")), "含义": "封装尺寸", "结果": clean_size(parsed.get("尺寸（inch）", ""))},
            {"段位": "阻值", "原始片段": clean_text(parsed.get("容值", "")) + clean_text(parsed.get("容值单位", "")), "含义": "按系列命名规则反推的阻值", "结果": f"{clean_text(parsed.get('容值', ''))}{clean_text(parsed.get('容值单位', ''))}"},
            {"段位": "容差", "原始片段": clean_tol_for_match(parsed.get("容值误差", "")), "含义": "阻值容差", "结果": clean_tol_for_match(parsed.get("容值误差", ""))},
            {"段位": "功率", "原始片段": clean_text(parsed.get("_power", "")), "含义": "额定功率", "结果": format_power_display(parsed.get("_power", ""))},
        ])
    return pd.DataFrame(rows)


def build_generic_model_rule_breakdown(model, parsed=None):
    model_clean = clean_model(model)
    rows = [{"段位": "型号", "原始片段": model_clean, "含义": "当前未命中更细的系列拆分规则，显示已识别字段", "结果": ""}]
    if parsed is not None:
        for label, key in [
            ("品牌", "品牌"),
            ("器件类型", "器件类型"),
            ("尺寸", "尺寸（inch）"),
            ("材质（介质）", "材质（介质）"),
            ("参数值", "容值"),
            ("参数单位", "容值单位"),
            ("参数误差", "容值误差"),
            ("耐压（V）", "耐压（V）"),
            ("功率", "_power"),
            ("脚距", "_pitch"),
            ("安规", "_safety_class"),
        ]:
            value = parsed.get(key, "")
            if clean_text(value) != "":
                rows.append({"段位": label, "原始片段": clean_text(value), "含义": label, "结果": clean_text(value)})
    return pd.DataFrame(rows)


def build_model_naming_interpretation(model, brand="", component_type=""):
    model_text = clean_model(model)
    brand_text = clean_brand(brand)
    if model_text == "":
        return {
            "model": "",
            "brand": brand_text,
            "component_type": "",
            "authority": "",
            "summary": "请输入型号后再解释。",
            "parsed": None,
            "breakdown": pd.DataFrame(),
            "manual_rule": None,
        }
    manual_rule = find_matching_manual_correction_rule(model_text, brand_text, component_type)
    parsed = parse_model_rule(model_text, brand=brand_text, component_type=component_type)
    if parsed is None:
        return {
            "model": model_text,
            "brand": brand_text,
            "component_type": normalize_component_type(component_type),
            "authority": "",
            "summary": "当前型号未命中专用命名规则，请检查品牌、型号是否输入完整。",
            "parsed": None,
            "breakdown": pd.DataFrame(),
            "manual_rule": manual_rule,
        }
    resolved_type = normalize_component_type(parsed.get("器件类型", "")) or infer_spec_component_type(parsed)
    authority = clean_text(parsed.get("_model_rule_authority", ""))
    family = authority
    pdc_series_code, pdc_series_desc, _, _, _ = pdc_mlcc_series_profile_from_model(model_text)
    if pdc_series_code != "":
        family = f"pdc_{pdc_series_code.lower()}_series"
    elif murata_ntc_series_code_from_model(model_text):
        family = "murata_ntc_series"
    elif murata_series_code_from_model(model_text):
        family = "murata_core_series"
    elif resolved_type in RESISTOR_COMPONENT_TYPES or resolved_type == "热敏电阻":
        family = "resistor"
    elif resolved_type in INDUCTOR_COMPONENT_TYPES:
        family = "inductor"
    elif resolved_type in TIMING_COMPONENT_TYPES:
        family = "timing"
    elif resolved_type in VARISTOR_COMPONENT_TYPES:
        family = "varistor"
    if authority.startswith("manual_rule"):
        summary = "命中人工修正规则表，系统会优先采用这里的覆盖值。"
    elif family.startswith("pdc_"):
        series_label = clean_text(parsed.get("系列", "")) or pdc_series_code
        series_desc = clean_text(parsed.get("系列说明", "")) or pdc_series_desc
        if series_label != "" and series_desc != "":
            summary = f"按信昌 PDC {series_label} 系列（{series_desc}）原厂料号结构拆分。"
        elif series_label != "":
            summary = f"按信昌 PDC {series_label} 系列原厂料号结构拆分。"
        else:
            summary = "按信昌 PDC 系列原厂料号结构拆分。"
    elif family == "murata_ntc_series":
        summary = "按村田 Murata NTC 官方热敏电阻系列规则拆分。"
    elif family == "murata_core_series":
        summary = "按村田 Murata 官方系列代码拆分。"
    elif family == "fenghua_am_series":
        series_label = clean_text(parsed.get("系列", "")) or "AM"
        series_desc = clean_text(parsed.get("系列说明", "")) or fenghua_am_series_meaning(series_label)
        summary = f"按风华 {series_label} 系列（{series_desc}）官方命名规则拆分。"
    elif family == "resistor":
        series_label = clean_text(parsed.get("系列", ""))
        series_desc = clean_text(parsed.get("系列说明", ""))
        if series_label != "" and series_desc != "":
            summary = f"按电阻 {series_label} 系列（{series_desc}）命名规则反推阻值、容差与封装。"
        elif series_label != "":
            summary = f"按电阻 {series_label} 系列命名规则反推阻值、容差与封装。"
        else:
            summary = "按电阻系列命名规则反推阻值、容差与封装。"
    else:
        summary = f"已按 {resolved_type or '已知'} 命名规则解析。"
    if family.startswith("pdc_"):
        breakdown = build_pdc_mt_rule_breakdown(model_text, parsed=parsed)
    elif family == "murata_ntc_series":
        breakdown = build_murata_ntc_rule_breakdown(model_text, parsed=parsed)
    elif family == "murata_core_series":
        breakdown = build_murata_rule_breakdown(model_text, parsed=parsed)
    elif family == "fenghua_am_series":
        parsed_size = clean_size(parsed.get("尺寸（inch）", ""))
        parsed_material = clean_material(parsed.get("材质（介质）", ""))
        parsed_pf = parsed.get("容值_pf", None)
        cap_value, cap_unit = pf_to_value_unit(parsed_pf) if parsed_pf is not None else ("", "")
        parsed_tol = clean_tol_for_match(parsed.get("容值误差", ""))
        parsed_volt = voltage_display(parsed.get("耐压（V）", ""))
        tail = FENGHUA_AM_MODEL_PATTERN.fullmatch(model_text)
        rows = [
            {"段位": "系列前缀", "原始片段": "AM", "含义": "风华 AM 汽车级 MLCC 系列", "结果": clean_text(parsed.get("系列", "")) or "AM"},
        ]
        if tail is not None:
            rows.extend([
                {"段位": "尺寸码", "原始片段": tail.group("size"), "含义": FENGHUA_AM_SIZE_CODE_MAP.get(tail.group("size"), "未知尺寸"), "结果": parsed_size},
                {"段位": "介质码", "原始片段": tail.group("dielectric"), "含义": FENGHUA_AM_DIELECTRIC_CODE_MAP.get(tail.group("dielectric"), "未知介质"), "结果": parsed_material},
                {"段位": "容值码", "原始片段": tail.group("cap"), "含义": f"{tail.group('cap')} -> {clean_text(cap_value)}{clean_text(cap_unit).upper()}", "结果": f"{clean_text(cap_value)}{clean_text(cap_unit).upper()}"},
                {"段位": "容差码", "原始片段": tail.group("tol"), "含义": f"±{FENGHUA_AM_TOLERANCE_CODE_MAP.get(tail.group('tol'), tail.group('tol'))}%", "结果": parsed_tol},
                {"段位": "电压码", "原始片段": tail.group("volt"), "含义": f"{FENGHUA_AM_VOLTAGE_CODE_MAP.get(tail.group('volt'), tail.group('volt'))}V", "结果": parsed_volt},
            ])
            if clean_text(tail.group("tail")) != "":
                rows.append({"段位": "尾码", "原始片段": tail.group("tail"), "含义": "订购尾码 / 包装变体", "结果": clean_text(tail.group("tail"))})
        breakdown = pd.DataFrame(rows)
    elif family == "resistor":
        breakdown = build_resistor_rule_breakdown(model_text, parsed=parsed)
    elif family == "inductor":
        breakdown = pd.DataFrame([
            {"段位": "系列/类型", "原始片段": model_text, "含义": "电感/磁珠系列按品牌命名规则拆分", "结果": resolved_type},
            {"段位": "尺寸（inch）", "原始片段": clean_size(parsed.get("尺寸（inch）", "")), "含义": "封装尺寸", "结果": clean_size(parsed.get("尺寸（inch）", ""))},
            {"段位": "感量", "原始片段": clean_text(parsed.get("容值", "")), "含义": "电感值", "结果": f"{clean_text(parsed.get('容值', ''))}{clean_text(parsed.get('容值单位', '')).upper()}"},
            {"段位": "误差", "原始片段": clean_tol_for_match(parsed.get("容值误差", "")), "含义": "电感公差", "结果": clean_tol_for_match(parsed.get("容值误差", ""))},
        ])
    elif family == "timing":
        breakdown = pd.DataFrame([
            {"段位": "系列/类型", "原始片段": model_text, "含义": "晶振/振荡器系列按品牌命名规则拆分", "结果": resolved_type},
            {"段位": "尺寸（inch）", "原始片段": clean_size(parsed.get("尺寸（inch）", "")), "含义": "封装尺寸", "结果": clean_size(parsed.get("尺寸（inch）", ""))},
            {"段位": "频率", "原始片段": clean_text(parsed.get("容值", "")), "含义": "输出频率/基频", "结果": f"{clean_text(parsed.get('容值', ''))}{clean_text(parsed.get('容值单位', '')).upper()}"},
            {"段位": "频差", "原始片段": clean_tol_for_match(parsed.get("容值误差", "")), "含义": "频差容差", "结果": clean_tol_for_match(parsed.get("容值误差", ""))},
        ])
    elif family == "varistor":
        breakdown = pd.DataFrame([
            {"段位": "系列/类型", "原始片段": model_text, "含义": "压敏电阻系列按品牌命名规则拆分", "结果": resolved_type},
            {"段位": "压敏电压", "原始片段": voltage_display(parsed.get("_varistor_voltage", parsed.get("耐压（V）", ""))), "含义": "压敏动作电压/额定电压", "结果": voltage_display(parsed.get("_varistor_voltage", parsed.get("耐压（V）", "")))},
            {"段位": "尺寸/脚距", "原始片段": clean_text(parsed.get("规格", "")) or clean_text(parsed.get("_pitch", "")), "含义": "系列尺寸或脚距", "结果": clean_text(parsed.get("规格", "")) or clean_text(parsed.get("_pitch", ""))},
        ])
    else:
        breakdown = build_generic_model_rule_breakdown(model_text, parsed=parsed)
    if manual_rule is not None and not breakdown.empty:
        note = build_manual_rule_match_summary(manual_rule)
        if note:
            breakdown = safe_concat_dataframes([
                pd.DataFrame([{
                    "段位": "人工修正规则",
                    "原始片段": clean_text(manual_rule.get("型号匹配", "")),
                    "含义": note,
                    "结果": clean_text(manual_rule.get("备注", "")),
                }]),
                breakdown,
            ], ignore_index=True)
    return {
        "model": model_text,
        "brand": brand_text,
        "component_type": resolved_type,
        "authority": authority,
        "summary": summary,
        "parsed": parsed,
        "breakdown": breakdown,
        "manual_rule": manual_rule,
    }


def parse_power_to_watts(power_text):
    raw = clean_text(power_text)
    if raw == "":
        return None
    compact = raw.replace("毫瓦", "mW").replace("瓦", "W").replace(" ", "")
    fraction_match = re.fullmatch(r"(\d+)/(\d+)W", compact, flags=re.I)
    if fraction_match:
        denominator = float(fraction_match.group(2))
        if denominator == 0:
            return None
        return float(fraction_match.group(1)) / denominator
    mw_match = re.fullmatch(r"(\d+(?:\.\d+)?)MW", compact, flags=re.I)
    if mw_match:
        return float(mw_match.group(1)) / 1000.0
    watt_match = re.fullmatch(r"(\d+(?:\.\d+)?)W", compact, flags=re.I)
    if watt_match:
        return float(watt_match.group(1))
    return None


def format_power_display(power_text):
    watts = parse_power_to_watts(power_text)
    if watts is None:
        return clean_text(power_text)
    common = [
        (1 / 16, "1/16W"),
        (1 / 10, "1/10W"),
        (1 / 8, "1/8W"),
        (1 / 4, "1/4W"),
        (1 / 2, "1/2W"),
        (1.0, "1W"),
        (2.0, "2W"),
        (3.0, "3W"),
        (5.0, "5W"),
    ]
    for target, label in common:
        if abs(watts - target) < 1e-9:
            return label
    if 0 < watts < 1:
        value = watts * 1000.0
        return f"{value:.3f}".rstrip("0").rstrip(".") + "mW"
    return f"{watts:.3f}".rstrip("0").rstrip(".") + "W"

def eia_code_to_pf(code):
    code = clean_text(code)
    if not code.isdigit():
        return None
    if len(code) == 3:
        base = int(code[:2]); mul = int(code[2]); return base * (10 ** mul)
    if len(code) == 4:
        base = int(code[:3]); mul = int(code[3]); return base * (10 ** mul)
    return None


def murata_cap_code_to_pf(code):
    code = clean_model(code)
    if code == "":
        return None

    pf = eia_code_to_pf(code)
    if pf is not None:
        return pf

    decimal_pf = re.fullmatch(r"(\d*)R(\d+)", code)
    if decimal_pf:
        left = decimal_pf.group(1) or "0"
        right = decimal_pf.group(2)
        return float(f"{left}.{right}")

    return None

def parse_samsung_cl(model):
    model = clean_model(model)
    if not model.startswith("CL") or len(model) < 10:
        return None
    size_map = {"02":"01005","03":"0201","05":"0402","10":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"A":"X5R","B":"X7R","C":"COG(NPO)","U":"X7S","Y":"X7T","Z":"X7R"}
    tol_map = {"F":"1","G":"2","J":"5","K":"10","M":"20","Z":"+80/-20"}
    voltage_map = {"R":"4","Q":"6.3","P":"10","O":"16","A":"25","L":"35","B":"50","C":"100","D":"200","E":"250","F":"500","G":"630","H":"1000"}
    try:
        size = size_map.get(model[2:4], "")
        material = clean_material(material_map.get(model[4], ""))
        cap_pf = murata_cap_code_to_pf(model[5:8])
        tol = clean_tol_for_match(tol_map.get(model[8], ""))
        volt = clean_voltage(voltage_map.get(model[9], ""))
        if size == "" or material == "" or cap_pf is None or tol == "" or volt == "":
            return None
        result = {
            "品牌":"Samsung",
            "型号":model,
            "尺寸（inch）":size,
            "材质（介质）":material,
            "容值_pf":cap_pf,
            "容值误差":tol,
            "耐压（V）":volt,
            "_model_rule_authority": "samsung_cl",
        }
        result.update(lookup_samsung_mlcc_dimension_fields(model))
        return result
    except:
        return None


def parse_samsung_clr1(model):
    model = clean_model(model)
    if not model.startswith("CLR1"):
        return None

    tol_map = {
        "A": "0.05PF",
        "B": "0.1pF",
        "C": "0.25pF",
        "D": "0.5pF",
        "F": "1",
        "G": "2",
        "J": "5",
        "K": "10",
        "M": "20",
    }
    voltage_map = {"A": "25", "O": "16"}
    match = re.fullmatch(r"CLR1C(?P<cap>(?:\d{3}|\dR\d|R\d+))(?P<tol>[A-Z])(?P<volt>[A-Z])(?P<rest>.*)", model)
    if not match:
        return None
    cap_pf = murata_cap_code_to_pf(match.group("cap"))
    if cap_pf is None:
        return None
    result = {
        "品牌": "Samsung",
        "型号": model,
        "尺寸（inch）": "0201",
        "材质（介质）": "COG(NPO)",
        "容值_pf": cap_pf,
        "容值误差": clean_tol_for_match(tol_map.get(match.group("tol"), "")),
        "耐压（V）": clean_voltage(voltage_map.get(match.group("volt"), "")),
        "_model_rule_authority": "samsung_clr1",
    }
    result.update(lookup_samsung_mlcc_dimension_fields(model))
    return result

def parse_pdc_fp(model):
    model = clean_model(model)
    if not model.startswith("FP") or len(model) < 12:
        return None
    size_map = {"31":"1206","32":"1210","42":"1808","43":"1812","46":"1825","55":"2220"}
    material_map = {"X":"X7R","T":"X7T","N":"COG(NPO)"}
    tol_map = {"J":"5","K":"10","M":"20"}
    voltage_map = {"100":"10","101":"100","102":"1000","200":"20","201":"200","202":"2000","250":"25","251":"250","252":"2500","300":"30","301":"300","302":"3000","450":"45","451":"450","500":"50","501":"500","630":"63","631":"630"}
    try:
        series_profile = pdc_mlcc_series_profile("FP")
        return {
            "品牌":"信昌PDC",
            "型号":model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）":size_map.get(model[2:4], ""),
            "材质（介质）":clean_material(material_map.get(model[4], "")),
            "容值_pf":eia_code_to_pf(model[5:8]),
            "容值误差":clean_tol_for_match(tol_map.get(model[8], "")),
            "耐压（V）":clean_voltage(voltage_map.get(model[9:12], "")),
            "_model_rule_authority": "pdc_fp_series",
        }
    except:
        return None

def parse_pdc_fs_voltage(model):
    model = clean_model(model)
    voltage_map_1 = {
        "R": "4", "Q": "6.3", "P": "10", "O": "16",
        "A": "25", "L": "35", "B": "50", "C": "100",
        "D": "200", "E": "250", "F": "500", "G": "630", "H": "1000"
    }
    if len(model) >= 12:
        code3 = model[9:12]
        decoded = decode_pdc_voltage_code(code3)
        if decoded != "":
            return decoded
    if len(model) >= 10:
        code1 = model[9]
        if code1 in voltage_map_1:
            return clean_voltage(voltage_map_1[code1])
    return ""

def parse_pdc_fs(model):
    model = clean_model(model)
    if not model.startswith("FS") or len(model) < 11:
        return None
    size_map = {"15":"0402","18":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"B":"X5R","X":"X7R","T":"X7T","N":"COG(NPO)"}
    tol_map = {"J":"5","K":"10","M":"20"}
    try:
        series_profile = pdc_mlcc_series_profile("FS")
        return {
            "品牌":"信昌PDC",
            "型号":model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）":size_map.get(model[2:4], ""),
            "材质（介质）":clean_material(material_map.get(model[4], "")),
            "容值_pf":eia_code_to_pf(model[5:8]),
            "容值误差":clean_tol_for_match(tol_map.get(model[8], "")),
            "耐压（V）":parse_pdc_fs_voltage(model),
            "_model_rule_authority": "pdc_fs_series",
        }
    except:
        return None

def parse_pdc_fn(model):
    model = clean_model(model)
    if not model.startswith("FN") or len(model) < 12:
        return None
    size_map = {"15":"0402","18":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"B":"X5R","X":"X7R","T":"X7T","N":"COG(NPO)"}
    tol_map = {"J":"5","K":"10","M":"20"}
    try:
        series_profile = pdc_mlcc_series_profile("FN")
        return {
            "品牌":"信昌PDC",
            "型号":model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）":size_map.get(model[2:4], ""),
            "材质（介质）":clean_material(material_map.get(model[4], "")),
            "容值_pf":eia_code_to_pf(model[5:8]),
            "容值误差":clean_tol_for_match(tol_map.get(model[8], "")),
            "耐压（V）":clean_voltage(parse_pdc_fs_voltage(model)),
            "_model_rule_authority": "pdc_fn_series",
        }
    except:
        return None

def parse_pdc_fm(model):
    model = clean_model(model)
    if not model.startswith("FM") or len(model) < 12:
        return None
    size_map = {"15":"0402","18":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"X":"X7R","N":"COG(NPO)","Y":"Y5V"}
    tol_map = {"J":"5","K":"10","M":"20"}
    try:
        series_profile = pdc_mlcc_series_profile("FM")
        return {
            "品牌":"信昌PDC",
            "型号":model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）":size_map.get(model[2:4], ""),
            "材质（介质）":clean_material(material_map.get(model[4], "")),
            "容值_pf":eia_code_to_pf(model[5:8]),
            "容值误差":clean_tol_for_match(tol_map.get(model[8], "")),
            "耐压（V）":clean_voltage(parse_pdc_fs_voltage(model)),
            "_model_rule_authority": "pdc_fm_series",
        }
    except:
        return None

def parse_pdc_fv(model):
    model = clean_model(model)
    if not model.startswith("FV") or len(model) < 12:
        return None
    size_map = {"15":"0402","18":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"X":"X7R","N":"COG(NPO)"}
    tol_map = {"J":"5","K":"10","M":"20"}
    try:
        series_profile = pdc_mlcc_series_profile("FV")
        return {
            "品牌":"信昌PDC",
            "型号":model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）":size_map.get(model[2:4], ""),
            "材质（介质）":clean_material(material_map.get(model[4], "")),
            "容值_pf":eia_code_to_pf(model[5:8]),
            "容值误差":clean_tol_for_match(tol_map.get(model[8], "")),
            "耐压（V）":clean_voltage(parse_pdc_fs_voltage(model)),
            "_model_rule_authority": "pdc_fv_series",
        }
    except:
        return None


def parse_pdc_fk(model):
    model = clean_model(model)
    if not model.startswith("FK") or len(model) < 12:
        return None
    size_map = {"08": "1808", "12": "1812", "20": "2220", "21": "2211"}
    material_map = {"N": "COG(NPO)", "X": "X7R"}
    tol_map = {"J": "5", "K": "10"}
    voltage_map = {"502": "250"}
    try:
        series_profile = pdc_mlcc_series_profile("FK")
        return {
            "品牌": "信昌PDC",
            "型号": model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）": size_map.get(model[2:4], ""),
            "材质（介质）": clean_material(material_map.get(model[4], "")),
            "容值_pf": eia_code_to_pf(model[5:8]),
            "容值误差": clean_tol_for_match(tol_map.get(model[8], "")),
            "耐压（V）": clean_voltage(voltage_map.get(model[9:12], "")),
            "_model_rule_authority": "pdc_fk_series",
        }
    except:
        return None


def parse_pdc_fh(model):
    model = clean_model(model)
    if not model.startswith("FH") or len(model) < 12:
        return None
    size_map = {"08": "1808", "12": "1812", "20": "2220"}
    material_map = {"N": "COG(NPO)", "X": "X7R"}
    tol_map = {"J": "5", "K": "10"}
    voltage_map = {"302": "250"}
    try:
        series_profile = pdc_mlcc_series_profile("FH")
        return {
            "品牌": "信昌PDC",
            "型号": model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）": size_map.get(model[2:4], ""),
            "材质（介质）": clean_material(material_map.get(model[4], "")),
            "容值_pf": eia_code_to_pf(model[5:8]),
            "容值误差": clean_tol_for_match(tol_map.get(model[8], "")),
            "耐压（V）": clean_voltage(voltage_map.get(model[9:12], "")),
            "_model_rule_authority": "pdc_fh_series",
        }
    except:
        return None


def parse_pdc_mlcc_core(model, family_prefix="MT", allow_partial=False):
    model = clean_model(model)
    family_prefix = clean_text(family_prefix).upper()
    if family_prefix not in {"MT", "MG", "MS"}:
        return None
    mt_match = re.fullmatch(
        rf"{family_prefix}(?P<size>\d{{2}})(?P<mat>[BCNXT])(?P<cap>(?:\d{{3,4}}|R\d+))(?P<tol>[FGJKMZ])(?P<volt>(?:6R3|\d{{3}}))(?P<rest>.*)",
        model,
    )
    if mt_match is None:
        return None

    material_map = {"B": "X5R", "X": "X7R", "T": "X7T", "N": "COG(NPO)"}
    tol_map = {"F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}

    try:
        size_code = mt_match.group("size")
        size = clean_size(PDC_MT_SIZE_CODE_MAP.get(size_code, ""))
        mat = clean_material(material_map.get(mt_match.group("mat"), ""))
        cap_pf = murata_cap_code_to_pf(mt_match.group("cap"))
        tol = clean_tol_for_match(tol_map.get(mt_match.group("tol"), ""))
        volt = decode_pdc_voltage_code(mt_match.group("volt"))
        suffix = clean_text(mt_match.group("rest"))
        package_code = suffix[0] if len(suffix) >= 1 else ""
        thickness_code = suffix[1] if len(suffix) >= 2 else ""
        special_code = suffix[2] if len(suffix) >= 3 else ""
        special_meaning = pdc_mlcc_special_control_meaning(special_code)
        series_code, series_meaning, _, _, _ = pdc_mlcc_series_profile_from_model(model)
        if series_code == "":
            series_code = family_prefix
        if series_meaning == "":
            series_meaning = pdc_mlcc_series_meaning(series_code)
        series_profile = pdc_mlcc_series_profile(series_code)
        if special_meaning != "":
            series_desc = f"{series_meaning} / {special_meaning}" if series_meaning else special_meaning
        else:
            series_desc = series_meaning
        special_use = clean_text(series_profile.get("特殊用途", ""))
        if special_meaning != "":
            special_use = f"{special_use}/{special_meaning}" if special_use else special_meaning

        param_count = sum([
            1 if size else 0,
            1 if mat else 0,
            1 if cap_pf is not None else 0,
            1 if tol else 0,
            1 if volt else 0,
        ])
        if param_count < 3:
            return None

        cap_value, cap_unit = pf_to_value_unit(cap_pf)
        result = {
            "品牌": "信昌PDC",
            "型号": model,
            "器件类型": "MLCC",
            "系列": series_code,
            "系列说明": series_desc,
            "尺寸（inch）": size,
            "材质（介质）": mat,
            "容值_pf": cap_pf,
            "容值": cap_value,
            "容值单位": cap_unit,
            "容值误差": tol,
            "耐压（V）": clean_voltage(volt),
            "安装方式": "贴片",
            "特殊用途": special_use,
            "_mlcc_series_class": normalize_mlcc_series_class(
                "/".join(
                    part for part in [
                        clean_text(series_profile.get("_mlcc_series_class", "")),
                        series_desc,
                        special_use,
                    ]
                    if clean_text(part) != ""
                )
            ),
            "_pdc_package_code": package_code,
            "_pdc_thickness_code": thickness_code,
            "_pdc_control_code": special_code,
        }
        result.update(build_pdc_mt_dimension_fields(size_code, thickness_code))
        if allow_partial:
            result["_param_count"] = param_count
            result["_partial_part"] = True
        else:
            result["_model_rule_authority"] = f"pdc_{series_code.lower()}_series" if series_code else "pdc_mlcc_series"
        return result
    except:
        return None


def parse_pdc_mt_core(model, allow_partial=False):
    return parse_pdc_mlcc_core(model, family_prefix="MT", allow_partial=allow_partial)


def parse_pdc_mg_core(model, allow_partial=False):
    return parse_pdc_mlcc_core(model, family_prefix="MG", allow_partial=allow_partial)


def parse_pdc_ms_core(model, allow_partial=False):
    return parse_pdc_mlcc_core(model, family_prefix="MS", allow_partial=allow_partial)



def parse_murata_core(model, allow_partial=False):
    model = clean_model(model)
    prefixes = ["GRM", "GCM", "GCJ", "GJM", "GQM", "GRT", "GCG", "GCQ"]
    prefix = next((p for p in prefixes if model.startswith(p)), None)
    if prefix is None:
        return None

    # Murata official order:
    # series + 2-char size + 1-char thickness + 2-char dielectric + 2-char voltage + 3-char capacitance + 1-char tolerance ...
    size_map = {
        "02": "01005", "03": "0201", "15": "0402", "18": "0603",
        "21": "0805", "31": "1206", "32": "1210", "42": "1808",
        "43": "1812", "55": "2220"
    }
    material_map = {
        "B1": "COG(NPO)", "C1": "COG(NPO)", "5C": "COG(NPO)",
        "R6": "X5R", "R7": "X7R", "R9": "X8R",
        "C6": "X5S", "C7": "X7S", "C8": "X6S",
        "D6": "X5T", "D7": "X7T", "D8": "X6T",
        "E7": "X7U", "L8": "X8L", "M8": "X8M", "N8": "X8N",
        "U2": "U2J", "7U": "U2J", "Z7": "X7R"
    }
    voltage_map = {
        "0E": "2.5", "0G": "4", "0J": "6.3", "1A": "10",
        "1C": "16", "1E": "25", "1H": "50", "2A": "100",
        "2D": "200", "2E": "250", "2J": "630", "2K": "1000",
        "YA": "35"
    }
    tol_map = {
        "B": "0.1pF", "C": "0.25pF", "D": "0.5pF", "W": "0.05pF",
        "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"
    }

    p = len(prefix)
    if len(model) < p + 11:
        return None

    try:
        size_code = model[p:p+2]
        mat_code = model[p+3:p+5]
        volt_code = model[p+5:p+7]
        cap_code = model[p+7:p+10]
        tol_code = model[p+10]
        series_profile = murata_series_profile(prefix)

        size = clean_size(size_map.get(size_code, ""))
        mat = clean_material(material_map.get(mat_code, ""))
        pf = murata_cap_code_to_pf(cap_code)
        tol = clean_tol_for_match(tol_map.get(tol_code, ""))
        volt = clean_voltage(voltage_map.get(volt_code, ""))
        dimension_fields = decode_murata_dimension_fields_from_model(model)
        param_count = sum([
            1 if size else 0,
            1 if mat else 0,
            1 if pf is not None else 0,
            1 if tol else 0,
            1 if volt else 0,
        ])
        if allow_partial and param_count < 3:
            return None

        return {
            "品牌": "村田Murata",
            "型号": model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）": size,
            "材质（介质）": mat,
            "容值_pf": pf,
            "容值误差": tol,
            "耐压（V）": volt,
            "_model_rule_authority": "murata_core_series",
            "_param_count": param_count,
            "_partial_part": True,
            **dimension_fields,
        }
    except:
        return None


def parse_murata_common(model):
    parsed = parse_murata_core(model, allow_partial=False)
    if parsed is None:
        return None
    parsed.pop("_param_count", None)
    parsed.pop("_partial_part", None)
    return parsed


def parse_tdk_c_series(model):
    model = clean_model(model)
    if not model.startswith("C") or len(model) < 14:
        return None

    size_map = {
        "0402": "0402", "0603": "0603", "0805": "0805", "1206": "1206",
        "1210": "1210", "1808": "1808", "1812": "1812", "2220": "2220",
        "1005": "0402", "1608": "0603", "2012": "0805", "3216": "1206",
        "3225": "1210", "4520": "1808", "4532": "1812", "5750": "2220"
    }
    material_map = {
        "C0G": "COG(NPO)", "COG": "COG(NPO)", "NP0": "COG(NPO)", "NPO": "COG(NPO)",
        "X5R": "X5R", "X7R": "X7R", "X7S": "X7S", "X6S": "X6S", "X7T": "X7T",
        "X8R": "X8R", "X8L": "X8L"
    }
    voltage_map = {
        "0E": "2.5", "0G": "4", "0J": "6.3", "1A": "10", "1C": "16",
        "1E": "25", "1H": "50", "2A": "100", "2D": "200", "2E": "250", "2J": "630"
    }
    tol_map = {
        "C": "0.25pF", "D": "0.5pF", "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"
    }

    try:
        size_code = model[1:5]
        mat_code = model[5:8]
        volt_code = model[8:10]
        cap_code = model[10:13]
        tol_code = model[13]
        dimension_fields = decode_tdk_dimension_fields_from_model(model)
        series_profile = tdk_mlcc_series_profile_from_model(model)

        return {
            "品牌": "TDK",
            "型号": model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）": size_map.get(size_code, ""),
            "材质（介质）": clean_material(material_map.get(mat_code, "")),
            "容值_pf": murata_cap_code_to_pf(cap_code),
            "容值误差": clean_tol_for_match(tol_map.get(tol_code, "")),
            "耐压（V）": clean_voltage(voltage_map.get(volt_code, "")),
            "_model_rule_authority": "tdk_c_series",
            **dimension_fields,
        }
    except:
        return None


def parse_tdk_cga_series(model):
    model = clean_model(model)
    if not model.startswith("CGA") or len(model) < 12:
        return None

    size_map = {
        "1": "0201", "2": "0402", "3": "0603", "4": "0805",
        "5": "1206", "6": "1210", "7": "1808", "8": "1812",
        "9": "2220", "D": "7563",
    }
    material_map = {
        "C0G": "COG(NPO)", "COG": "COG(NPO)", "NP0": "COG(NPO)", "NPO": "COG(NPO)",
        "X5R": "X5R", "X7R": "X7R", "X7S": "X7S", "X7T": "X7T",
        "X6S": "X6S", "X8R": "X8R", "X8L": "X8L",
    }
    voltage_map = {
        "0E": "2.5", "0G": "4", "0J": "6.3", "1A": "10", "1B": "16",
        "1C": "25", "1D": "50", "1E": "100", "1H": "50", "2A": "100",
        "2D": "200", "2E": "250", "2J": "630"
    }
    tol_map = {
        "C": "0.25PF", "D": "0.5PF", "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"
    }

    match = re.fullmatch(
        r"CGA([1-9D])([A-Z])([0-3])?(C0G|COG|NP0|NPO|X5R|X7R|X7S|X7T|X6S|X8R|X8L)([0-3][A-Z])(\d{3,4}|R\d+)([BCDFGJKMZ])(.*)",
        model,
    )
    if not match:
        return None

    try:
        series_profile = tdk_mlcc_series_profile_from_model(model)
        return {
            "品牌": "TDK",
            "型号": model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）": size_map.get(match.group(1), ""),
            "材质（介质）": clean_material(material_map.get(match.group(4), "")),
            "容值_pf": murata_cap_code_to_pf(match.group(6)),
            "容值误差": clean_tol_for_match(tol_map.get(match.group(7), "")),
            "耐压（V）": clean_voltage(voltage_map.get(match.group(5), "")),
            "_model_rule_authority": "tdk_cga_series",
            **decode_tdk_dimension_fields_from_model(model),
        }
    except:
        return None


def parse_walsin_common(model, brand=""):
    model = clean_model(model)
    brand_text = clean_brand(brand)
    brand_upper = clean_text(brand_text).upper()
    if "WALSIN" not in brand_upper and "华新科" not in brand_text:
        return None
    if len(model) < 11:
        return None

    size_map = {
        "0201": "0201", "0402": "0402", "0603": "0603", "0805": "0805",
        "1206": "1206", "1210": "1210", "1808": "1808", "1812": "1812", "2220": "2220",
        "03": "0201", "05": "0402", "06": "0603", "09": "0805", "11": "0505", "12": "1206",
        "15": "0402", "18": "0603", "21": "0805", "31": "1206", "32": "1210",
        "42": "1808", "43": "1812", "56": "2225",
    }
    material_map = {
        "B": "X7R", "C": "COG(NPO)", "F": "Y5V", "N": "COG(NPO)", "S": "X6S", "T": "X7T", "X": "X5R"
    }
    tol_map = {
        "A": "0.25PF", "B": "0.1PF", "C": "0.25PF", "D": "0.5PF",
        "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20",
    }

    special_01005_match = re.fullmatch(
        r"01R5(?P<mat>[BCNXT])(?P<cap>(?:\d{3,4}|R\d+))(?P<tol>[FGJKMZ])(?P<volt>(?:6R3|\d{3}))(?P<rest>.*)",
        model,
    )
    if special_01005_match:
        volt = special_01005_match.group("volt")
        vmap = {"6R3": "6.3", "100": "10", "160": "16", "250": "25", "500": "50", "630": "63"}
        series_profile = walsin_mlcc_series_profile("01R5")
        return {
            "品牌": "华新科Walsin",
            "型号": model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）": "01005",
            "材质（介质）": clean_material(material_map.get(special_01005_match.group("mat"), "")),
            "容值_pf": murata_cap_code_to_pf(special_01005_match.group("cap")),
            "容值误差": clean_tol_for_match(tol_map.get(special_01005_match.group("tol"), "")),
            "耐压（V）": clean_voltage(vmap.get(volt, volt)),
            "_model_rule_authority": "walsin_01005_series",
        }

    alpha_match = re.fullmatch(
        r"(?P<prefix>[A-Z]{2})(?P<size>\d{2})(?P<mat>[BCNXT])(?P<cap>(?:\d{3,4}|R\d+|\dR\d+))(?P<tol>[ABCDEFGJKMZ])(?P<volt>(?:6R3|\d{3}))(?P<rest>.*)",
        model,
    )
    if alpha_match:
        volt = alpha_match.group("volt")
        vmap = {"6R3": "6.3", "100": "10", "160": "16", "250": "25", "500": "50", "630": "63"}
        series_profile = walsin_mlcc_series_profile(alpha_match.group("prefix"))
        return {
            "品牌": "华新科Walsin",
            "型号": model,
            "器件类型": "MLCC",
            "系列": series_profile["系列"],
            "系列说明": series_profile["系列说明"],
            "特殊用途": series_profile["特殊用途"],
            "_mlcc_series_class": series_profile["_mlcc_series_class"],
            "尺寸（inch）": size_map.get(alpha_match.group("size"), clean_size(alpha_match.group("size"))),
            "材质（介质）": clean_material(material_map.get(alpha_match.group("mat"), "")),
            "容值_pf": murata_cap_code_to_pf(alpha_match.group("cap")),
            "容值误差": clean_tol_for_match(tol_map.get(alpha_match.group("tol"), "")),
            "耐压（V）": clean_voltage(vmap.get(volt, volt)),
            "_model_rule_authority": "walsin_alpha_prefix",
        }

    size_code = model[:4]
    if not size_code.isdigit():
        return None

    try:
        mat_code = model[4]
        cap_code = model[5:8]
        tol_code = model[8]
        rest = model[9:]
        volt = ""
        if rest.startswith("6R3"):
            volt = "6.3"
        elif len(rest) >= 3 and rest[:3].isdigit():
            vcode = rest[:3]
            vmap = {"100": "10", "160": "16", "250": "25", "500": "50", "630": "63"}
            volt = vmap.get(vcode, vcode)

        return {
            "品牌": "华新科Walsin",
            "型号": model,
            "器件类型": "MLCC",
            "系列": f"{size_code}{mat_code}",
            "系列说明": f"华新科Walsin {size_code}{mat_code} {clean_material(material_map.get(mat_code, '')) or 'MLCC'}系列",
            "尺寸（inch）": size_map.get(size_code, ""),
            "材质（介质）": clean_material(material_map.get(mat_code, "")),
            "容值_pf": murata_cap_code_to_pf(cap_code),
            "容值误差": clean_tol_for_match(tol_map.get(tol_code, "")),
            "耐压（V）": clean_voltage(volt),
            "_model_rule_authority": "walsin_numeric_series",
        }
    except:
        return None


def parse_fenghua_am_series(model):
    model = clean_model(model)
    match = FENGHUA_AM_MODEL_PATTERN.fullmatch(model)
    if match is None:
        return None

    size_code = match.group("size")
    dielectric_code = match.group("dielectric")
    cap_code = match.group("cap")
    tol_code = match.group("tol")
    volt_code = match.group("volt")

    size = clean_size(FENGHUA_AM_SIZE_CODE_MAP.get(size_code, ""))
    material = clean_material(FENGHUA_AM_DIELECTRIC_CODE_MAP.get(dielectric_code, ""))
    cap_pf = murata_cap_code_to_pf(cap_code)
    cap_value, cap_unit = pf_to_value_unit(cap_pf) if cap_pf is not None else ("", "")
    tol = clean_tol_for_match(FENGHUA_AM_TOLERANCE_CODE_MAP.get(tol_code, ""))
    volt = clean_voltage(FENGHUA_AM_VOLTAGE_CODE_MAP.get(volt_code, volt_code))
    dimension_fields = fenghua_am_dimension_fields_from_model(model)
    param_count = sum([
        1 if size != "" else 0,
        1 if material != "" else 0,
        1 if cap_pf is not None else 0,
        1 if tol != "" else 0,
        1 if volt != "" else 0,
    ])

    result = {
        "品牌": "风华Fenghua",
        "型号": model,
        "器件类型": "MLCC",
        "系列": "AM",
        "系列说明": fenghua_am_series_meaning("AM"),
        "尺寸（inch）": size,
        "材质（介质）": material,
        "容值_pf": cap_pf,
        "容值": clean_text(cap_value),
        "容值单位": clean_text(cap_unit).upper(),
        "容值误差": tol,
        "耐压（V）": volt,
        "安装方式": "贴片",
        "_model_rule_authority": "fenghua_am_series",
        "_param_count": param_count,
    }
    if dimension_fields:
        result.update(dimension_fields)
    return result



def parse_cap_token_to_pf(token):
    t = clean_text(token).lower().replace(" ", "")
    t = t.replace("μf", "uf")
    if t == "":
        return None

    m = re.match(r"^(\d+(?:\.\d+)?)(pf|nf|uf)$", t)
    if m:
        val = float(m.group(1))
        unit = m.group(2).upper()
        return cap_to_pf(val, unit)

    if re.fullmatch(r"\d{3,4}", t):
        return eia_code_to_pf(t)

    return None




def looks_like_spec_query(line):
    raw = clean_text(line).upper()
    if raw == "":
        return False
    if re.search(r"[\s,/\\|;:%]", raw):
        return True
    spec_keywords = [
        "X5R", "X7R", "X7S", "X7T", "Y5V", "COG", "NPO", "C0G", "NP0", "X6S",
        "UF", "NF", "PF", "6R3", "10%", "5%", "20%", "630V", "6.3V",
        "OHM", "Ω", "电阻", "电解", "脚距", "PITCH"
    ]
    if any(k in raw for k in spec_keywords):
        return True
    if any(size_code in raw for size_code in {"008004", "01005", "015008"}):
        return True
    if raw[:4].isdigit() and raw[:4] in {"0100", "0102", "0201", "0401", "0402", "0603", "0805", "1206", "1210", "1808", "1812", "1825", "2220"}:
        return True
    if re.fullmatch(r"\d{3,4}[FGJKMZ]?", raw):
        return True
    return False


def count_core_params(spec):
    if spec is None:
        return 0
    explicit_count = spec.get("_core_param_count")
    if explicit_count is not None:
        try:
            return int(explicit_count)
        except:
            pass
    return sum([
        1 if clean_text(spec.get("尺寸（inch）", "")) != "" else 0,
        1 if clean_text(spec.get("材质（介质）", "")) != "" else 0,
        1 if spec.get("容值_pf", None) is not None else 0,
        1 if clean_text(spec.get("容值误差", "")) != "" else 0,
        1 if clean_text(spec.get("耐压（V）", "")) != "" else 0,
    ])


def count_query_params(spec):
    if spec is None:
        return 0
    explicit_count = spec.get("_param_count")
    if explicit_count is not None:
        try:
            return int(explicit_count)
        except:
            pass
    return count_core_params(spec)


def other_passive_min_required_params(spec):
    if spec is None:
        return 2
    component_type = normalize_component_type(spec.get("器件类型", ""))
    if component_type in (RESISTOR_COMPONENT_TYPES | {"热敏电阻"}):
        return 1 if spec.get("_resistance_ohm") is not None else 2
    if component_type in VARISTOR_COMPONENT_TYPES:
        varistor_voltage = clean_voltage(spec.get("_varistor_voltage", "")) or clean_voltage(spec.get("耐压（V）", ""))
        disc_size = clean_text(spec.get("_disc_size", ""))
        return 1 if (varistor_voltage != "" or disc_size != "") else 2
    if component_type in {"铝电解电容", "薄膜电容", "钽电容"}:
        return 2
    return 2


OTHER_PASSIVE_TYPES = ALL_PASSIVE_COMPONENT_TYPES - {"MLCC"}


def is_other_passive_mode(mode):
    return clean_text(mode) in OTHER_PASSIVE_TYPES or clean_text(mode) == "其他器件"


def normalized_other_passive_mode(spec, fallback="其他器件"):
    component_type = normalize_component_type(spec.get("器件类型", "")) if spec else ""
    if component_type in OTHER_PASSIVE_TYPES:
        return component_type
    fallback_text = clean_text(fallback)
    if fallback_text in OTHER_PASSIVE_TYPES:
        return fallback_text
    return fallback_text or "其他器件"


def looks_like_compact_part_query(line):
    raw = clean_model(line)
    if raw == "":
        return False
    if re.search(r"[\s,/\\|;:%]", clean_text(line).upper()):
        return False
    if not (re.search(r"[A-Z]", raw) and re.search(r"\d", raw)):
        return False

    # Short industrial/inductor model names such as ICM2020-A or ILHB-1206
    # are valid compact part queries, but they used to miss the older length-only
    # heuristic and fall all the way back to the full dataframe path.
    if len(raw) >= 8 and (
        re.fullmatch(r"[A-Z]{2,8}\d{3,6}[A-Z0-9-]*", raw)
        or re.fullmatch(r"[A-Z]{2,8}-\d{3,6}[A-Z0-9-]*", raw)
    ):
        return True

    known_prefixes = (
        "CL", "AM", "FP", "FS", "FN", "FM", "FV", "FK", "FH",
        "GRM", "GCM", "GCJ", "GJM", "GQM",
        "TCC", "TMK", "JMK", "EMK", "LMK", "AMK", "CC",
        "ECG2", "ECR", "ECS1", "ECS2", "ECA1", "ECC1", "ECV", "PHR1", "PHV1", "PCP",
    )
    if raw.startswith(known_prefixes):
        return True
    if raw.startswith("C") and len(raw) >= 14:
        return True
    if len(raw) >= 11 and raw[:4].isdigit():
        return True
    alpha_count = len(re.findall(r"[A-Z]", raw))
    digit_count = len(re.findall(r"\d", raw))
    if len(raw) >= 12 and alpha_count >= 3 and digit_count >= 3:
        return True
    return False

def parse_cap_token_to_pf(token):
    t = clean_text(token).lower().replace(" ", "")
    t = t.replace("μf", "uf").replace("µf", "uf")
    if t == "":
        return None

    # 1uf / 100nf / 47pf / 0.1uf
    m = re.match(r"^(\d+(?:\.\d+)?)(pf|nf|uf)$", t)
    if m:
        val = float(m.group(1))
        unit = m.group(2).upper()
        return cap_to_pf(val, unit)

    # 105 / 104 / 474 / 1005
    if re.fullmatch(r"\d{3,4}", t):
        return eia_code_to_pf(t)

    # 102P / 224P 常见 BOM 写法：视为 EIA 容值码
    if re.fullmatch(r"\d{3,4}p", t):
        return eia_code_to_pf(t[:-1])

    # 22P / 4N7 / 1U0 这类紧凑写法
    compact_suffix_match = re.fullmatch(r"(\d+(?:\.\d+)?)([PNU])", t.upper())
    if compact_suffix_match:
        num = float(compact_suffix_match.group(1))
        unit = {"P": "PF", "N": "NF", "U": "UF"}[compact_suffix_match.group(2)]
        return cap_to_pf(num, unit)

    # 1U0 / 4N7 / 2P2 style (optional support)
    m2 = re.match(r"^(\d+)([UPN])(\d+)$", t.upper())
    if m2:
        a, u, b = m2.groups()
        num = float(f"{a}.{b}")
        unit = {"P": "PF", "N": "NF", "U": "UF"}[u]
        return cap_to_pf(num, unit)

    m3 = re.match(r"^(\d*)R(\d+)$", t.upper())
    if m3:
        a, b = m3.groups()
        num = float(f"{a or '0'}.{b}")
        return cap_to_pf(num, "PF")

    return None

def parse_spec_query(line):
    raw = clean_text(line)
    if raw == "":
        return None

    s = raw.upper()
    s = s.replace("μF", "UF").replace("µF", "UF")
    s = s.replace("％", "%").replace("﹪", "%")
    s = s.replace("＋", "+").replace("﹢", "+")
    s = s.replace("／", "/").replace("\\", "/")
    s = s.replace("±", "+/-").replace("卤", "+/-")
    s = s.replace("_", " ")

    # 把连写的 “容值/电压 + 容差” 拆开，避免 6.3V±20% / 1PF±0.1PF 这类格式被误判
    s = re.sub(r"(\d+(?:\.\d+)?(?:PF|NF|UF|V))(\+/-)", r"\1 \2", s)
    s = re.sub(r"(\+/-\s*\d+(?:\.\d+)?%?)(?=\d+(?:\.\d+)?V\b)", r"\1 ", s)
    s = re.sub(r"(\d+R\d+)(\+/-)", r"\1 \2", s)
    s = re.sub(r"(\d{3,4}P)(\+/-)", r"\1 \2", s)

    s = s.replace("+/-", " PLUSMINUS ")
    for old in ["，", ",", "/", "|", ";", "；", "(", ")", "（", "）", "：", ":", "OR", "或"]:
        s = s.replace(old, " ")
    s = s.replace("PLUSMINUS", "+/-")
    s = re.sub(r"(\+/-)\s*(\d+(?:\.\d+)?(?:PF|%))", r"\1\2", s)

    # split combined patterns like 105K / 474K / 104J / 6R3V
    s = re.sub(r"(\d{3,4})([FGJKMZ])\b", r"\1 \2", s)
    s = re.sub(r"(\d+R\d+)(V)\b", r"\1 \2", s)
    s = re.sub(r"\s+", " ", s).strip()

    tokens = s.split(" ")

    size = find_embedded_size(raw)
    material = find_embedded_material(raw)
    cap_pf = None
    tol = ""
    volt = ""

    size_candidates = {"008004", "01005", "0102", "015008", "0201", "0401", "0402", "0603", "0805", "1206", "1210", "1808", "1812", "1825", "2220"}
    material_candidates = {"X5R", "X7R", "X7S", "X7T", "Y5V", "COG", "NPO", "COG(NPO)", "C0G", "NP0", "X6S"}
    tol_letter_map = {"F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}

    compact_raw = raw.upper().replace(" ", "")
    if tol == "":
        pf_tol_match = re.search(r"(\+/-\d+(?:\.\d+)?PF)", compact_raw)
        if pf_tol_match:
            tol = clean_tol_for_match(pf_tol_match.group(1))
    if tol == "":
        percent_tol_match = re.search(r"(\+/-\d+(?:\.\d+)?%)", compact_raw)
        if percent_tol_match:
            tol = clean_tol_for_match(percent_tol_match.group(1))
    if tol == "":
        asym_tol_match = re.search(r"(\+\d+(?:\.\d+)?%?/-\d+(?:\.\d+)?%?)", compact_raw)
        if asym_tol_match:
            tol = clean_tol_for_match(asym_tol_match.group(1))

    for token in tokens:
        t = token.strip().upper()
        if t == "":
            continue

        # size
        if size == "" and t in size_candidates:
            size = t
            continue
        if size == "":
            embedded_size = find_embedded_size(t)
            if embedded_size != "":
                size = embedded_size

        if t in size_candidates:
            continue

        # material
        if material == "" and t in material_candidates:
            material = clean_material(t)
            continue
        if material == "":
            embedded_material = find_embedded_material(t)
            if embedded_material != "":
                material = embedded_material

        # tolerance
        if tol == "":
            parsed_tol = parse_tolerance_token(t)
            if parsed_tol != "":
                tol = parsed_tol
                continue
            if t in tol_letter_map:
                tol = clean_tol_for_match(tol_letter_map[t])
                continue

        # voltage
        if volt == "" and re.fullmatch(r"\d+(?:\.\d+)?V", t):
            volt = clean_voltage(t)
            continue
        if volt == "" and re.fullmatch(r"\d+R\d+", t):
            volt = clean_voltage(t.replace("R", "."))
            continue
        if volt == "" and re.fullmatch(r"\d+(?:\.\d+)?", t) and "." in t:
            num = float(t)
            if 0 < num <= 1000:
                volt = clean_voltage(t)
                continue
        # common bare voltages when clearly voltage-like
        if volt == "" and t in {"4", "6.3", "10", "16", "25", "35", "50", "100", "200", "250", "500", "630", "1000"}:
            volt = clean_voltage(t)
            continue

        # capacitance
        pf = parse_cap_token_to_pf(t)
        if cap_pf is None and pf is not None:
            cap_pf = pf
            continue

    if cap_pf is None:
        explicit_cap_match = re.search(r"(?<!\d)(\d+(?:\.\d+)?)(PF|NF|UF)(?![A-Z])", compact_raw)
        if explicit_cap_match:
            cap_pf = cap_to_pf(explicit_cap_match.group(1), explicit_cap_match.group(2))
    if cap_pf is None:
        compact_code_match = re.search(r"(?<!\d)(\d{3,4})P(?![A-Z0-9])", compact_raw)
        if compact_code_match:
            cap_pf = eia_code_to_pf(compact_code_match.group(1))

    param_count = sum([
        1 if size else 0,
        1 if material else 0,
        1 if cap_pf is not None else 0,
        1 if tol else 0,
        1 if volt else 0
    ])

    if param_count == 0:
        return None

    return {
        "品牌": "",
        "型号": raw,
        "尺寸（inch）": size,
        "材质（介质）": material,
        "容值_pf": cap_pf,
        "容值误差": tol,
        "耐压（V）": clean_voltage(volt),
        "_param_count": param_count
    }

def format_resistance_display(ohm_value):
    try:
        ohm = float(ohm_value)
    except:
        return ""
    if 0 < ohm < 1:
        value = ohm * 1000
        text = f"{value:.3f}".rstrip("0").rstrip(".")
        return f"{text}mΩ"
    if ohm >= 1000000:
        value = ohm / 1000000
        text = f"{value:.3f}".rstrip("0").rstrip(".")
        return f"{text}MΩ"
    if ohm >= 1000:
        value = ohm / 1000
        text = f"{value:.3f}".rstrip("0").rstrip(".")
        return f"{text}KΩ"
    text = f"{ohm:.3f}".rstrip("0").rstrip(".")
    return f"{text}Ω"

def parse_resistance_token_to_ohm(token):
    raw = clean_text(token).replace(" ", "")
    raw = raw.replace("OHMS", "Ω").replace("OHM", "Ω").replace("ohms", "Ω").replace("ohm", "Ω")
    milliohm_match = re.fullmatch(r"(\d+(?:\.\d+)?)mΩ", raw, flags=re.I)
    if milliohm_match:
        return float(milliohm_match.group(1)) / 1000.0
    t = raw.upper()
    if t.endswith("Ω"):
        bare = t[:-1]
        if re.fullmatch(r"\d+(?:\.\d+)?(?:[RKM])", bare) or re.fullmatch(r"\d+[RKM]\d+", bare):
            t = bare
        elif re.fullmatch(r"\d+(?:\.\d+)?", bare):
            t = bare + "R"
        else:
            t = bare
    if t == "":
        return None
    leading_r_match = re.fullmatch(r"R(\d+(?:\.\d+)?)", t)
    if leading_r_match:
        digits = leading_r_match.group(1)
        if "." in digits:
            return float(f"0{digits}")
        return float(digits) / (10 ** len(digits))
    direct_match = re.fullmatch(r"(\d+(?:\.\d+)?)([RKM])", t)
    if direct_match:
        value = float(direct_match.group(1))
        unit = direct_match.group(2)
        multiplier = {"R": 1, "K": 1000, "M": 1000000}[unit]
        return value * multiplier
    compact_match = re.fullmatch(r"(\d+)([RKM])(\d+)", t)
    if compact_match:
        whole, unit, frac = compact_match.groups()
        value = float(f"{whole}.{frac}")
        multiplier = {"R": 1, "K": 1000, "M": 1000000}[unit]
        return value * multiplier
    plain_match = re.fullmatch(r"\d+(?:\.\d+)?", t)
    if plain_match:
        return float(t)
    return None

def find_resistance_in_text(text):
    upper = clean_text(text).upper().replace("OHMS", "Ω").replace("OHM", "Ω")
    upper = upper.replace("±", "+/-").replace("卤", "+/-")
    ohm_match = RESISTOR_OHM_PATTERN.search(upper)
    if ohm_match:
        return parse_resistance_token_to_ohm(ohm_match.group(1))
    resistor_context = detect_component_type_hint(upper) in (RESISTOR_COMPONENT_TYPES | {"热敏电阻"}) or looks_like_resistor_context(upper)
    if resistor_context:
        compact_match = RESISTOR_COMPACT_CONTEXT_PATTERN.search(upper)
        if compact_match:
            return parse_resistance_token_to_ohm(compact_match.group(1))
    match = RESISTOR_VALUE_PATTERN.search(upper)
    if match:
        return parse_resistance_token_to_ohm(match.group(1))
    return None

def find_inductance_in_text(text):
    upper = clean_text(text).upper().replace("μ", "U").replace("µ", "U").replace("Μ", "U")
    match = INDUCTANCE_PATTERN.search(upper)
    if not match:
        return ""
    return f"{match.group(1)}{match.group(2).upper()}"


def text_contains_inductance_range(text):
    upper = clean_text(text).upper().replace("μ", "U").replace("µ", "U").replace("Μ", "U")
    if upper == "":
        return False
    if "INDUCTANCE RANGE" in upper:
        return True
    return bool(
        re.search(
            r"\d+(?:\.\d+)?\s*(?:-|~|–|TO)\s*\d+(?:\.\d+)?\s*(NH|UH|MH|H)\b",
            upper,
            flags=re.I,
        )
    )

def find_current_in_text(text):
    upper = clean_text(text).upper()
    matches = CURRENT_PATTERN.findall(upper)
    if not matches:
        return ""
    for value, unit in matches:
        if unit.upper() == "A" or float(value) >= 1:
            return f"{value}{unit.upper()}"
    value, unit = matches[0]
    return f"{value}{unit.upper()}"


def text_contains_current_range(text):
    upper = clean_text(text).upper()
    if upper == "":
        return False
    if "CURRENT RANGE" in upper:
        return True
    return bool(
        re.search(
            r"\d+(?:\.\d+)?\s*(?:-|~|–|TO)\s*\d+(?:\.\d+)?\s*(UA|MA|A)\b",
            upper,
            flags=re.I,
        )
    )


def find_labeled_current_in_text(text, labels):
    upper = clean_text(text).upper().replace("DC MAX.", "DC").replace("DC MAX", "DC")
    if upper == "":
        return ""
    if isinstance(labels, str):
        labels = [labels]
    for label in labels or []:
        label_text = clean_text(label).upper()
        if label_text == "":
            continue
        pattern = re.compile(rf"{re.escape(label_text)}[^0-9]{{0,40}}(\d+(?:\.\d+)?)\s*(UA|MA|A)\b", flags=re.I)
        match = pattern.search(upper)
        if match:
            return f"{match.group(1)}{match.group(2).upper()}"
    return ""


def find_frequency_in_text(text):
    upper = clean_text(text).upper()
    match = FREQUENCY_PATTERN.search(upper)
    if not match:
        return ""
    return f"{match.group(1)}{match.group(2).upper()}"


def find_frequency_tolerance_in_text(text):
    upper = clean_text(text).upper().replace("±", "+/-")
    if upper == "":
        return ""
    match = re.search(r"(\+/-)\s*(\d+(?:\.\d+)?)\s*(PPM|%)", upper, flags=re.I)
    if match:
        return f"{match.group(1)}{match.group(2)}{match.group(3).upper()}"
    match = re.search(r"(\d+(?:\.\d+)?)\s*(PPM)\b", upper, flags=re.I)
    if match:
        return f"{match.group(1)}{match.group(2).upper()}"
    return ""


def _find_labeled_resistive_measurement(text, labels):
    raw = clean_text(text).replace("OHMS", "OHM").replace("ohms", "ohm")
    if raw == "":
        return ""
    if isinstance(labels, str):
        labels = [labels]
    for label in labels or []:
        label_text = clean_text(label)
        if label_text == "":
            continue
        pattern = re.compile(rf"{re.escape(label_text)}[^0-9]{{0,48}}(\d+(?:\.\d+)?)\s*(mΩ|MΩ|KΩ|MOHM|KOHM|OHM|Ω)", flags=re.I)
        match = pattern.search(raw)
        if match:
            return combine_value_and_unit(match.group(1), normalize_library_value_unit(match.group(2)), separator="")
    return ""


def find_dcr_in_text(text):
    return _find_labeled_resistive_measurement(text, ["DCR", "DC RESISTANCE"])


def find_esr_in_text(text):
    return _find_labeled_resistive_measurement(text, ["ESR"])


def text_contains_impedance_range(text):
    upper = clean_text(text).upper()
    if upper == "":
        return False
    if "IMPEDANCE RANGE" in upper:
        return True
    return bool(
        re.search(
            r"\d+(?:,\d{3})*(?:\.\d+)?\s*(OHM|Ω)\s*(?:-|~|–|TO)\s*\d+(?:,\d{3})*(?:\.\d+)?\s*(OHM|Ω)?",
            upper,
            flags=re.I,
        )
    )


def find_impedance_at_100mhz_in_text(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    match = re.search(r"100\s*MHZ[^0-9]{0,24}(\d+(?:\.\d+)?)\s*(OHM|Ω)", upper, flags=re.I)
    if match:
        return combine_value_and_unit(match.group(1), normalize_library_value_unit(match.group(2)), separator="")
    match = re.search(r"IMPEDANCE[^0-9]{0,24}(\d+(?:\.\d+)?)\s*(OHM|Ω)", upper, flags=re.I)
    if match:
        return combine_value_and_unit(match.group(1), normalize_library_value_unit(match.group(2)), separator="")
    return ""


def find_b_value_in_text(text):
    upper = clean_text(text).upper().replace("β", "B")
    if upper == "":
        return ""
    match = re.search(r"(?:B(?:ETA)?(?:25/50|25/85)?|B值)[^0-9]{0,12}(\d{3,5}(?:\.\d+)?)\s*K?", upper, flags=re.I)
    if not match:
        return ""
    return normalize_b_value_text(match.group(1))


def find_b_value_condition_in_text(text):
    upper = clean_text(text).upper().replace("β", "B")
    if upper == "":
        return ""
    match = re.search(r"(25/50|25/85|50/85|25℃/50℃|25℃/85℃)", upper, flags=re.I)
    if not match:
        return ""
    return clean_text(match.group(1)).replace("℃", "℃")


def find_circuit_count_in_text(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    match = re.search(r"(\d+)\s*(?:LINES?|CIRCUITS?|回路)", upper, flags=re.I)
    if not match:
        return ""
    return clean_text(match.group(1))


def find_shielding_type_in_text(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    if "UNSHIELDED" in upper or "NON-SHIELDED" in upper or "非屏蔽" in upper:
        return "非屏蔽"
    if "SHIELDED" in upper or "屏蔽" in upper:
        return "屏蔽"
    return ""


def find_varistor_voltage_in_text(text):
    explicit = parse_voltage_from_text(text)
    if explicit != "":
        return explicit
    upper = clean_text(text).upper().replace(" ", "")
    match = VARISTOR_CODE_PATTERN.search(upper)
    if not match:
        return ""
    code = match.group(1)
    try:
        base = int(code[:2])
        multiplier = int(code[2])
        voltage = base * (10 ** multiplier)
        return clean_voltage(str(voltage))
    except Exception:
        return ""

def find_disc_size_code(text):
    upper = clean_text(text).upper().replace(" ", "")
    match = DISC_SIZE_CODE_PATTERN.search(upper)
    if not match:
        return ""
    return f"{match.group(1)}D"

def find_tolerance_in_text(text):
    upper = clean_text(text).upper().replace("±", "+/-").replace("卤", "+/-")
    if upper == "":
        return ""
    patterns = [
        r"(\+/-\s*\d+(?:\.\d+)?\s*%)",
        r"(?<![\d.])(\d+(?:\.\d+)?\s*%)",
        r"(\+\s*\d+(?:\.\d+)?\s*%?\s*/-\s*\d+(?:\.\d+)?\s*%?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, upper)
        if match:
            return clean_tol_for_match(match.group(1))
    compact = upper.replace(" ", "")
    for pattern in [r"(\+/-\d+(?:\.\d+)?%)", r"(?<![A-Z0-9.])(\d+(?:\.\d+)?%)", r"(\+\d+(?:\.\d+)?%?/-\d+(?:\.\d+)?%?)"]:
        match = re.search(pattern, compact)
        if match:
            return clean_tol_for_match(match.group(1))
    return ""

def find_power_in_text(text):
    raw = clean_text(text)
    if raw == "":
        return ""
    normalized = raw.replace("毫瓦", "mW").replace("瓦", "W")
    for pattern in [
        r"(?<![A-Z0-9.])(\d+\s*/\s*\d+\s*W)(?![A-Z0-9])",
        r"(?<![A-Z0-9.])(\d+(?:\.\d+)?\s*mW)(?![A-Z0-9])",
        r"(?<![A-Z0-9.])(\d+(?:\.\d+)?\s*W)(?![A-Z0-9])",
    ]:
        match = re.search(pattern, normalized, flags=re.I)
        if match:
            return format_power_display(match.group(1))
    return ""

def parse_voltage_from_text(text):
    upper = clean_text(text).upper()
    upper = upper.replace("±", "+/-").replace("卤", "+/-")
    compact = upper.replace(" ", "")
    # Support tightly packed values like 100UF25V or 0.1UF275V.
    compact_match = re.search(r"(?:PF|NF|UF)(\d+(?:\.\d+)?)V", compact)
    if compact_match:
        return clean_voltage(compact_match.group(1))
    voltage_match = re.search(r"(\d+(?:\.\d+)?)\s*V", upper)
    if voltage_match:
        return clean_voltage(voltage_match.group(1))
    return ""

def normalize_component_keyword_compact(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    compact = upper.replace(" ", "")
    return re.sub(r"[^A-Z0-9\u4E00-\u9FFFΩ]+", "", compact)


def alias_token_matches(upper_text, compact_text, token):
    token_upper = clean_text(token).upper()
    if token_upper == "":
        return False
    token_compact = re.sub(r"[^A-Z0-9\u4E00-\u9FFFΩ]+", "", token_upper.replace(" ", ""))
    if token_compact == "":
        return False
    if re.fullmatch(r"[A-Z0-9]+", token_compact) and len(token_compact) <= 4:
        boundary_patterns = [
            rf"(?<![A-Z0-9]){re.escape(token_compact)}(?![A-Z0-9])",
            rf"(?<![A-Z0-9]){re.escape(token_compact)}(?=\d)",
            rf"^{re.escape(token_compact)}(?=\d)",
        ]
        return any(re.search(pattern, upper_text) is not None for pattern in boundary_patterns)
    if re.search(r"[^A-Z0-9Ω]", token_upper) or " " in token_upper:
        return token_upper in upper_text
    return token_compact in compact_text


COMPONENT_ALIAS_TOKENS = {
    "MLCC": {
        "compact": [
            "MLCC", "CCAP", "CCER", "CERCAP", "CERAMICCAPACITOR",
            "MONOLITHICCERAMIC", "MONOCERAMIC", "CHIPCERAMIC",
            "CERAMICCHIP", "MULTILAYERCERAMICCHIPCAPACITOR", "CERAMIC",
        ],
        "upper": ["C.CAP", "C CAP", "CER CAP", "CERAMIC CAP", "陶瓷", "瓷片", "片容", "独石电容", "独石电容器"],
    },
    "薄膜电容": {
        "compact": [
            "PPCAP", "FILMCAP", "FILMCAPACITOR", "POLYPROPYLENE",
            "POLYESTER", "METALLIZEDFILM", "SUPPRESSIONCAPACITOR",
            "SAFETYCAPACITOR", "CBB", "MKP", "MKT", "PPS", "PEN", "PETCAP",
        ],
        "upper": ["PP CAP", "FILM CAP", "POLYPROPYLENE CAP", "POLYESTER CAP", "薄膜", "安规", "安规电容", "安規", "安規電容", "薄膜电容", "薄膜電容器"],
    },
    "钽电容": {
        "compact": ["TANTALUM", "TANTCAP", "TANT", "TANCAP"],
        "upper": ["TANTALUM CAP", "TANTALUM", "钽电容", "钽电容器"],
    },
    "铝电解电容": {
        "compact": [
            "ECAP", "PEECAP", "ALCAP", "ALECAP", "ELECCAP",
            "ELECTROLYTICCAPACITOR", "ALUMINUMELECTROLYTIC",
            "POLYMERECAP", "POLYMERALUMINUM",
        ],
        "upper": [
            "E CAP", "PE E CAP", "AL CAP", "AL E CAP", "ELECTROLYTIC",
            "ALUMINUM ELECTROLYTIC", "铝电解", "铝电解电容", "铝电解电容器",
            "电解", "电解电容", "电解电容器", "编带式",
            "鋁電解", "鋁電解電容", "鋁電解電容器", "電解電容", "電解電容器",
        ],
    },
    "压敏电阻": {
        "compact": ["VARISTOR", "MOV", "VDR", "ZNR", "METALOXIDEVARISTOR", "压敏电阻", "压敏"],
        "upper": ["VARISTOR", "MOV", "VDR", "ZNR", "压敏电阻", "压敏", "壓敏電阻", "壓敏", "氧化锌压敏", "氧化鋅壓敏"],
    },
    "引线型压敏电阻": {
        "compact": ["LEADEDVARISTOR", "RADIALVARISTOR", "引线型压敏电阻", "引线压敏电阻", "插件压敏电阻", "直插压敏电阻"],
        "upper": ["LEADED VARISTOR", "RADIAL VARISTOR", "引线型压敏电阻", "引线压敏电阻", "插件压敏电阻", "直插压敏电阻", "引线压敏", "插件压敏", "直插压敏"],
    },
    "贴片压敏电阻": {
        "compact": ["SMDVARISTOR", "CHIPVARISTOR", "MLV", "贴片压敏电阻", "贴片压敏"],
        "upper": ["SMD VARISTOR", "CHIP VARISTOR", "贴片压敏电阻", "贴片压敏", "貼片壓敏電阻", "貼片壓敏", "MLV"],
    },
    "热敏电阻": {
        "compact": ["THERMISTOR", "NTC", "PTC", "热敏电阻", "热敏"],
        "upper": ["THERMISTOR", "NTC", "PTC", "热敏电阻", "热敏", "熱敏電阻", "熱敏"],
    },
    "合金电阻": {
        "compact": ["ALLOY", "METALSTRIP", "SHUNT", "CURRENTSENSE", "LOWOHMIC"],
        "upper": ["合金", "分流电阻", "采样电阻", "取样电阻", "电流检测电阻", "电流采样电阻", "分流電阻", "採樣電阻", "取樣電阻", "電流檢測電阻", "電流採樣電阻"],
    },
    "碳膜电阻": {
        "compact": ["CARBONFILMRESISTOR", "CARBONRESISTOR", "CARBONFILM", "CARBONCOMPOSITION"],
        "upper": ["CARBON RESISTOR", "CARBON FILM", "CARBON FILM RESISTOR", "碳膜电阻", "碳質電阻"],
    },
    "金属氧化膜电阻": {
        "compact": ["METALOXIDEFILMRESISTOR", "METALOXIDEFILM", "METALOXIDE"],
        "upper": ["METAL OXIDE", "METAL OXIDE FILM", "METAL OXIDE FILM RESISTOR", "金属氧化膜电阻", "金屬氧化膜電阻"],
    },
    "绕线电阻": {
        "compact": ["WIREWOUNDRESISTOR", "WIREWOUND"],
        "upper": ["WIREWOUND", "WIREWOUND RESISTOR", "绕线电阻", "繞線電阻"],
    },
    "薄膜电阻": {
        "compact": ["THINFILMRESISTOR", "THINFILM", "METALFILMRESISTOR", "METALFILM"],
        "upper": ["THIN FILM", "THIN-FILM", "METAL FILM", "METAL FILM RESISTOR", "薄膜电阻", "金属膜电阻", "金屬膜電阻"],
    },
    "厚膜电阻": {
        "compact": ["THICKFILMRESISTOR", "THICKFILM", "CHIPRESISTOR"],
        "upper": ["THICK FILM", "THICK-FILM", "厚膜电阻", "貼片電阻", "贴片电阻", "CHIP RESISTOR"],
    },
    "贴片电阻": {
        "compact": ["RESISTOR", "OHM"],
        "upper": ["RESISTOR", "电阻", "電阻", "OHM", "Ω"],
    },
    "功率电感": {
        "compact": ["POWERINDUCTOR", "INDUCTOR", "CHOKE", "WIREWOUNDINDUCTOR", "MOLDEDINDUCTOR"],
        "upper": ["POWER INDUCTOR", "功率电感", "功率电感器", "电感", "電感", "CHOKE"],
    },
    "射频电感": {
        "compact": [
            "RFINDUCTOR",
            "AIRCORE",
            "AIRCOIL",
            "WIREWOUNDINDUCTOR",
            "CHIPINDUCTOR",
            "MULTILAYERCHIPINDUCTOR",
            "RFCHOKE",
        ],
        "upper": [
            "RF INDUCTOR",
            "AIR CORE",
            "AIR COIL",
            "WIREWOUND INDUCTOR",
            "CHIP INDUCTOR",
            "MULTILAYER CHIP INDUCTOR",
            "RF CHOKE",
        ],
    },
    "共模电感": {
        "compact": ["COMMONMODE", "COMMONMODECHOKE", "CMCHOKE", "CMC"],
        "upper": ["COMMON MODE", "COMMON-MODE", "COMMON MODE CHOKE", "共模电感", "共模扼流圈", "共模扼流圈"],
    },
    "磁珠": {
        "compact": ["FERRITEBEAD", "CHIPBEAD", "BEAD"],
        "upper": ["FERRITE BEAD", "CHIP BEAD", "磁珠"],
    },
    "晶振": {
        "compact": ["CRYSTAL", "XTAL", "QUARTZ"],
        "upper": ["CRYSTAL", "XTAL", "QUARTZ", "晶振", "晶體", "石英晶体", "石英晶體", "石英晶振"],
    },
    "振荡器": {
        "compact": ["OSCILLATOR", "VCXO", "TCXO", "OCXO", "XO"],
        "upper": ["OSCILLATOR", "振荡器", "振盪器", "有源晶振"],
    },
}


def matches_component_alias(text, component_type):
    upper = clean_text(text).upper()
    compact = normalize_component_keyword_compact(text)
    if upper == "":
        return False
    token_group = COMPONENT_ALIAS_TOKENS.get(component_type, {})
    compact_tokens = token_group.get("compact", [])
    upper_tokens = token_group.get("upper", [])
    if any(alias_token_matches(upper, compact, token) for token in compact_tokens):
        return True
    return any(alias_token_matches(upper, compact, token) for token in upper_tokens)

def looks_like_mlcc_context(text):
    upper = clean_text(text).upper()
    compact = normalize_component_keyword_compact(text)
    if upper == "":
        return False
    if matches_component_alias(text, "MLCC"):
        return True
    if any(token in compact for token, _ in SPEC_EMBEDDED_MATERIALS) and (
        any(unit in compact for unit in ["PF", "NF", "UF"]) or find_embedded_size(upper) != ""
    ):
        return True
    return False

def looks_like_film_capacitor_context(text):
    upper = clean_text(text).upper()
    if upper == "":
        return False
    if matches_component_alias(text, "薄膜电容"):
        return True
    if re.search(r"(?<![A-Z0-9])(X1|X2|Y1|Y2)(?![A-Z0-9])", upper):
        return True
    return False

def looks_like_tantalum_context(text):
    return matches_component_alias(text, "钽电容")

def looks_like_leaded_ceramic_context(text):
    upper = clean_text(text).upper()
    compact = normalize_component_keyword_compact(text)
    if upper == "":
        return False
    if "Y5P" not in compact and "Y5P" not in upper:
        return False
    if re.search(r"D\d+\s*\*\s*L\d+", upper) is None and re.search(r"(LEADED|DISC|RADIAL|圆片|圆瓷|引线|瓷片|陶瓷圆片)", upper) is None:
        return False
    return True

def looks_like_varistor_context(text):
    return (
        matches_component_alias(text, "压敏电阻")
        or matches_component_alias(text, "引线型压敏电阻")
        or matches_component_alias(text, "贴片压敏电阻")
    )

def looks_like_smd_varistor_context(text):
    if not looks_like_varistor_context(text):
        return False
    upper = clean_text(text).upper()
    if upper == "":
        return False
    if matches_component_alias(text, "贴片压敏电阻"):
        return True
    if any(token in upper for token in ["SMD", "SMT", "CHIP", "贴片", "貼片", "MLV"]):
        return True
    return SMD_SIZE_CODE_PATTERN.search(upper) is not None

def looks_like_leaded_varistor_context(text):
    if not looks_like_varistor_context(text):
        return False
    if looks_like_smd_varistor_context(text):
        return False
    upper = clean_text(text).upper()
    compact = normalize_component_keyword_compact(text)
    if upper == "":
        return False
    if matches_component_alias(text, "引线型压敏电阻"):
        return True
    if find_disc_size_code(text) != "":
        return True
    if extract_pitch_from_text(text) != "":
        return True
    if re.search(r"(LEADED|RADIAL|引线|插件|直插|脚距)", upper):
        return True
    return MM_DIMENSION_PATTERN.search(upper) is not None and "SMD" not in compact

def detect_resistor_subtype_hint(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    if looks_like_varistor_context(text):
        if looks_like_smd_varistor_context(text):
            return "贴片压敏电阻"
        if looks_like_leaded_varistor_context(text):
            return "引线型压敏电阻"
        return "压敏电阻"
    if matches_component_alias(text, "热敏电阻"):
        return "热敏电阻"
    if matches_component_alias(text, "合金电阻"):
        return "合金电阻"
    if matches_component_alias(text, "碳膜电阻"):
        return "碳膜电阻"
    if matches_component_alias(text, "薄膜电阻"):
        return "薄膜电阻"
    if matches_component_alias(text, "厚膜电阻"):
        return "厚膜电阻"
    compact = normalize_component_keyword_compact(text)
    if matches_component_alias(text, "贴片电阻") or upper.startswith("RES"):
        return "贴片电阻"
    return ""

def detect_inductor_subtype_hint(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    if matches_component_alias(text, "射频电感"):
        return "射频电感"
    if matches_component_alias(text, "共模电感"):
        return "共模电感"
    if matches_component_alias(text, "磁珠"):
        return "磁珠"
    if matches_component_alias(text, "功率电感"):
        return "功率电感"
    return ""

def detect_timing_subtype_hint(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    if matches_component_alias(text, "振荡器"):
        return "振荡器"
    if matches_component_alias(text, "晶振"):
        return "晶振"
    return ""

def find_film_material(text):
    upper = clean_text(text).upper().replace(" ", "")
    material_tokens = [
        ("MKP", "MKP"),
        ("MKT", "MKT"),
        ("CBB", "CBB"),
        ("PPS", "PPS"),
        ("PEN", "PEN"),
        ("PP", "PP"),
        ("PET", "PET"),
    ]
    for token, label in material_tokens:
        if token in upper:
            return label
    return ""

def find_safety_class(text):
    upper = clean_text(text).upper()
    match = re.search(r"(?<![A-Z0-9])(X1|X2|Y1|Y2)(?![A-Z0-9])", upper)
    if match:
        return match.group(1)
    return ""

def detect_component_type_hint(text):
    upper = clean_text(text).upper()
    if upper == "":
        return ""
    timing_hint = detect_timing_subtype_hint(text)
    if timing_hint != "":
        return timing_hint
    inductor_hint = detect_inductor_subtype_hint(text)
    if inductor_hint != "":
        return inductor_hint
    resistor_hint = detect_resistor_subtype_hint(text)
    if resistor_hint in (VARISTOR_COMPONENT_TYPES | {"热敏电阻"}):
        return resistor_hint
    if looks_like_film_capacitor_context(text):
        return "薄膜电容"
    if looks_like_tantalum_context(text):
        return "钽电容"
    if looks_like_leaded_ceramic_context(text):
        return "引线型陶瓷电容"
    if looks_like_electrolytic_context(text):
        return "铝电解电容"
    if resistor_hint != "":
        return resistor_hint
    if looks_like_mlcc_context(text):
        return "MLCC"
    return ""

def normalize_component_type(value):
    text = clean_text(value)
    if text == "":
        return ""
    if text in ALL_PASSIVE_COMPONENT_TYPES:
        return text
    hint = detect_component_type_hint(text)
    if hint != "":
        return hint
    return text


def resolve_component_category_type(value):
    if isinstance(value, pd.Series):
        value = value.to_dict()
    if isinstance(value, dict):
        direct_type = normalize_component_type(value.get("器件类型", ""))
        if direct_type != "":
            return direct_type
        inferred_row_type = infer_db_component_type(value)
        if inferred_row_type != "":
            return inferred_row_type
        inferred_spec_type = infer_spec_component_type(value)
        return inferred_spec_type if inferred_spec_type in ALL_PASSIVE_COMPONENT_TYPES else ""
    return normalize_component_type(value)


def format_component_category_display(value):
    component_type = resolve_component_category_type(value)
    if component_type == "":
        return ""
    return COMPONENT_CATEGORY_DISPLAY_LABELS.get(component_type, component_type)


def format_component_size_display_value(value):
    text = clean_text(value)
    if text == "":
        return ""
    normalized = text.lower()
    if "mm" in normalized or "×" in text or "*" in text or re.search(r"\d\s*[xX]\s*\d", text):
        return text
    return clean_size(text)


def resistor_source_evidence_present(value):
    note = clean_text(value).lower()
    if note == "":
        return False
    return any(token in note for token in RESISTOR_SOURCE_EVIDENCE_TOKENS)


def infer_strong_component_type_override(row, direct="", text=""):
    direct_type = normalize_component_type(direct)
    if direct_type in ALL_RESISTOR_TYPES or direct_type in INDUCTOR_COMPONENT_TYPES or direct_type in TIMING_COMPONENT_TYPES:
        return ""
    summary = clean_text(row.get("规格摘要", ""))
    notes = clean_text(row.get("校验备注", ""))
    if summary == "" and notes == "":
        return ""
    explicit_resistor_summary = (
        detect_resistor_subtype_hint(summary) != ""
        or (
            looks_like_resistor_context(summary)
            and not looks_like_varistor_context(summary)
            and not looks_like_thermistor_context(summary)
        )
    )
    if not resistor_source_evidence_present(notes) and not explicit_resistor_summary:
        return ""
    evidence_text = " ".join(
        part for part in [
            clean_text(text),
            summary,
            notes,
            clean_text(row.get("品牌", "")),
            clean_text(row.get("型号", "")),
            clean_text(row.get("系列", "")),
            clean_text(row.get("备注1", "")),
            clean_text(row.get("备注2", "")),
            clean_text(row.get("备注3", "")),
        ]
        if clean_text(part) != ""
    )
    resistor_hint = detect_resistor_subtype_hint(evidence_text)
    if resistor_hint != "":
        return resistor_hint
    if looks_like_resistor_context(evidence_text):
        return "贴片电阻"
    return ""

def infer_db_component_type(row):
    direct = normalize_component_type(row.get("器件类型", ""))
    raw_unit = clean_text(row.get("容值单位", "")).upper()
    text = " ".join([
        clean_text(row.get("器件类型", "")),
        clean_text(row.get("品牌", "")),
        clean_text(row.get("型号", "")),
        clean_text(row.get("系列", "")),
        clean_text(row.get("安装方式", "")),
        clean_text(row.get("封装代码", "")),
        clean_text(row.get("尺寸（mm）", "")),
        clean_text(row.get("材质（介质）", "")),
        clean_text(row.get("规格摘要", "")),
        clean_text(row.get("特殊用途", "")),
        clean_text(row.get("备注1", "")),
        clean_text(row.get("备注2", "")),
        clean_text(row.get("备注3", "")),
    ])
    if direct in VARISTOR_COMPONENT_TYPES or looks_like_varistor_context(text):
        if direct == "贴片压敏电阻" or looks_like_smd_varistor_context(text):
            return "贴片压敏电阻"
        if direct == "引线型压敏电阻" or looks_like_leaded_varistor_context(text):
            return "引线型压敏电阻"
        if direct in VARISTOR_COMPONENT_TYPES:
            return direct
        return "压敏电阻"
    if looks_like_thermistor_context(text):
        return "热敏电阻"
    if direct == "MLCC" and looks_like_leaded_ceramic_context(text):
        return "引线型陶瓷电容"
    strong_override = infer_strong_component_type_override(row, direct=direct, text=text)
    if strong_override != "":
        return strong_override
    if direct in INDUCTOR_COMPONENT_TYPES or raw_unit in {"NH", "UH", "MH"}:
        return direct if direct in INDUCTOR_COMPONENT_TYPES else (detect_inductor_subtype_hint(text) or "功率电感")
    if direct in TIMING_COMPONENT_TYPES or raw_unit in {"HZ", "KHZ", "MHZ"}:
        return direct if direct in TIMING_COMPONENT_TYPES else (detect_timing_subtype_hint(text) or "晶振")
    if direct != "":
        return direct
    hint = detect_component_type_hint(text)
    if hint != "":
        return hint
    if looks_like_film_capacitor_context(text):
        return "薄膜电容"
    # Some vendor tables only expose electrolytic / resistor / thermistor clues
    # in package text, pitch, or ohm-related fields, so give these a second pass
    # before falling back to MLCC.
    if looks_like_leaded_ceramic_context(text):
        return "引线型陶瓷电容"
    if looks_like_electrolytic_context(text):
        return "铝电解电容"
    if looks_like_thermistor_context(text):
        return "热敏电阻"
    if looks_like_resistor_context(text):
        return "贴片电阻"
    if clean_material(row.get("材质（介质）", "")) != "" or pd.notna(row.get("容值_pf", None)):
        return "MLCC"
    if clean_text(row.get("容值单位", "")).upper() in {"PF", "NF", "UF"}:
        return "MLCC"
    return ""

def infer_spec_component_type(spec):
    if spec is None:
        return ""
    direct = normalize_component_type(spec.get("器件类型", ""))
    raw_unit = clean_text(spec.get("容值单位", "")).upper()
    spec_text = " ".join([
        clean_text(spec.get("器件类型", "")),
        clean_text(spec.get("型号", "")),
        clean_text(spec.get("规格摘要", "")),
        clean_text(spec.get("安装方式", "")),
        clean_text(spec.get("封装代码", "")),
        clean_text(spec.get("尺寸（mm）", "")),
        clean_text(spec.get("材质（介质）", "")),
        clean_text(spec.get("_body_size", "")),
        clean_text(spec.get("_pitch", "")),
        clean_text(spec.get("_safety_class", "")),
    ])
    if direct in VARISTOR_COMPONENT_TYPES or looks_like_varistor_context(spec_text):
        if direct == "贴片压敏电阻" or looks_like_smd_varistor_context(spec_text):
            return "贴片压敏电阻"
        if direct == "引线型压敏电阻" or looks_like_leaded_varistor_context(spec_text):
            return "引线型压敏电阻"
        if direct in VARISTOR_COMPONENT_TYPES:
            return direct
        return "压敏电阻"
    if direct == "MLCC" and looks_like_leaded_ceramic_context(spec_text):
        return "引线型陶瓷电容"
    strong_override = infer_strong_component_type_override(spec, direct=direct, text=spec_text)
    if strong_override != "":
        return strong_override
    if direct in INDUCTOR_COMPONENT_TYPES or raw_unit in {"NH", "UH", "MH"}:
        return direct if direct in INDUCTOR_COMPONENT_TYPES else (detect_inductor_subtype_hint(spec_text) or "功率电感")
    if direct in TIMING_COMPONENT_TYPES or raw_unit in {"HZ", "KHZ", "MHZ"}:
        return direct if direct in TIMING_COMPONENT_TYPES else (detect_timing_subtype_hint(spec_text) or "晶振")
    if direct != "":
        return direct
    hint = detect_component_type_hint(spec_text)
    if hint != "":
        return hint
    if looks_like_film_capacitor_context(spec_text):
        return "薄膜电容"
    if looks_like_leaded_ceramic_context(spec_text):
        return "引线型陶瓷电容"
    if looks_like_electrolytic_context(spec_text):
        return "铝电解电容"
    if looks_like_thermistor_context(spec_text):
        return "热敏电阻"
    if looks_like_resistor_context(spec_text):
        return "贴片电阻"
    if spec.get("容值_pf", None) is not None or clean_material(spec.get("材质（介质）", "")) != "":
        return "MLCC"
    return ""

def looks_like_thermistor_context(text):
    raw = clean_text(text)
    if raw == "":
        return False
    upper = raw.upper()
    compact = normalize_component_keyword_compact(raw)
    if matches_component_alias(raw, "热敏电阻"):
        return True
    if upper.startswith("NTC") or compact.startswith("NTC"):
        return True
    if any(token in raw for token in ["过流保护", "正极电阻", "自恢复保险丝", "可恢复保险丝", "保险丝热敏"]):
        return True
    if any(token in upper for token in ["PTC THERMISTOR", "OVERCURRENT PROTECTION", "RESETTABLE FUSE", "RESETTABLE", "POLYFUSE", "PPTC", "POSISTOR"]):
        return True
    if "MURATA" in upper and re.search(r"\bPRG\d", upper):
        return True
    if "MURATA" in upper and re.search(r"PRG\d", compact):
        return True
    if re.search(r"(?:FTN|NCP|NCU|NCG)\d{2}[A-Z]{2}\d{3}[BCDFGJKMZ]", compact):
        return True
    return False


def reverse_lookup_row_priority(row):
    model_text = clean_model(row.get("_model_clean", row.get("型号", "")))
    row_type = normalize_component_type(row.get("_component_type", "")) or infer_db_component_type(row) or normalize_component_type(row.get("器件类型", ""))
    data_source = clean_text(row.get("数据来源", ""))
    data_status = clean_text(row.get("数据状态", ""))
    series = clean_text(row.get("系列", ""))
    summary = clean_text(row.get("规格摘要", ""))
    score = 0

    if "官方PDF" in data_status:
        score += 1000
    elif "官方" in data_status:
        score += 400

    if "Part Number List" in data_source or "官方" in data_source:
        score += 200

    if looks_like_thermistor_context(model_text):
        if row_type == "热敏电阻":
            score += 300
        elif row_type in RESISTOR_COMPONENT_TYPES:
            score -= 300

    if looks_like_varistor_context(model_text):
        if row_type in VARISTOR_COMPONENT_TYPES:
            score += 300
        elif row_type == "贴片电阻":
            score -= 300

    if series != "":
        score += 10
    if summary != "":
        score += 5
    if clean_text(row.get("容值", "")) != "" or clean_text(row.get("阻值@25C", "")) != "":
        score += 5
    return score


def prioritize_component_rows_for_lookup(df):
    if df is None or df.empty:
        return df
    if "品牌" not in df.columns or "型号" not in df.columns:
        return df

    work = df.copy()
    work["_lookup_order"] = range(len(work))
    try:
        work["_lookup_priority"] = work.apply(reverse_lookup_row_priority, axis=1)
    except Exception:
        work["_lookup_priority"] = 0

    work = work.sort_values(
        by=["品牌", "型号", "_lookup_priority", "_lookup_order"],
        ascending=[True, True, False, True],
        kind="mergesort",
    )
    return work.drop(columns=["_lookup_priority", "_lookup_order"], errors="ignore")


def looks_like_resistor_context(text):
    hint = detect_component_type_hint(text)
    if hint in (VARISTOR_COMPONENT_TYPES | {"热敏电阻", "铝电解电容", "薄膜电容", "钽电容", "功率电感", "共模电感", "磁珠", "晶振", "振荡器", "MLCC"}):
        return False
    if hint in RESISTOR_COMPONENT_TYPES:
        return True
    upper = clean_text(text).upper().replace("±", "+/-").replace("卤", "+/-")
    compact = normalize_component_keyword_compact(text).replace("±", "+/-").replace("卤", "+/-")
    if any(token in upper for token in ["电阻", "RESISTOR", "OHM", "Ω"]) or upper.startswith("RES"):
        return True
    if any(token in compact for token in ["UF", "NF", "PF"]):
        return False
    has_compact_resistance = RESISTOR_VALUE_PATTERN.search(upper) is not None or RESISTOR_COMPACT_CONTEXT_PATTERN.search(upper) is not None
    return has_compact_resistance and "%" in upper

def build_other_component_summary(parts):
    return " ".join([clean_text(part) for part in parts if clean_text(part) != ""])


def parse_inductor_spec_query(line):
    raw = clean_text(line)
    if raw == "":
        return None

    upper = raw.upper().replace("μ", "U").replace("µ", "U").replace("Μ", "U")
    component_type = detect_inductor_subtype_hint(raw)
    inductance = find_inductance_in_text(upper)
    if component_type == "" and inductance == "":
        return None
    if component_type == "" and any(
        checker(raw)
        for checker in (
            looks_like_mlcc_context,
            looks_like_film_capacitor_context,
            looks_like_electrolytic_context,
            looks_like_resistor_context,
            looks_like_varistor_context,
            looks_like_thermistor_context,
        )
    ):
        return None
    if component_type == "":
        component_type = "功率电感"

    size = find_embedded_size(raw)
    tol = find_tolerance_in_text(raw)
    tokens = [token.strip().upper() for token in re.split(r"[\s,;/|()（）]+", upper) if token.strip()]
    if tol == "":
        for token in tokens:
            if token in INDUCTOR_TOLERANCE_CODE_MAP:
                tol = clean_tol_for_match(INDUCTOR_TOLERANCE_CODE_MAP[token])
                break
    current = find_current_in_text(upper)

    value = ""
    unit = ""
    inductance_match = re.fullmatch(r"(\d+(?:\.\d+)?)(NH|UH|MH)", inductance, flags=re.I)
    if inductance_match:
        value = clean_text(inductance_match.group(1))
        unit = clean_text(inductance_match.group(2)).upper()

    param_count = sum([
        1 if size else 0,
        1 if value and unit else 0,
        1 if tol else 0,
        1 if current else 0,
    ])
    if param_count == 0:
        return None

    summary = build_other_component_summary([
        f"{value}{unit}" if value and unit else "",
        clean_tol_for_display(tol) if tol else "",
        current,
        size,
    ])
    return {
        "品牌": "",
        "型号": raw,
        "尺寸（inch）": size,
        "材质（介质）": "",
        "容值": value,
        "容值单位": unit,
        "容值_pf": None,
        "容值误差": tol,
        "耐压（V）": "",
        "器件类型": component_type,
        "规格摘要": summary,
        "_current": current,
        "_core_param_count": param_count,
        "_param_count": param_count,
    }


def parse_resistor_spec_query(line):
    raw = clean_text(line)
    if raw == "" or not looks_like_resistor_context(raw):
        return None

    normalized = raw.upper()
    normalized = normalized.replace("μF", "UF").replace("µF", "UF")
    normalized = normalized.replace("％", "%").replace("﹪", "%")
    normalized = normalized.replace("＋", "+").replace("﹢", "+")
    normalized = normalized.replace("／", "/").replace("\\", "/")
    normalized = normalized.replace("±", "+/-").replace("卤", "+/-")
    normalized = normalized.replace("+/-", " PLUSMINUS ")
    for old in ["，", ",", "/", "|", ";", "；", "(", ")", "（", "）", "：", ":"]:
        normalized = normalized.replace(old, " ")
    normalized = normalized.replace("PLUSMINUS", "+/-")
    tokens = [token.strip().upper() for token in normalized.split(" ") if token.strip()]

    component_type = detect_resistor_subtype_hint(raw) or "贴片电阻"
    if component_type in SPECIAL_RESISTOR_COMPONENT_TYPES:
        return None
    size = find_embedded_size(raw)
    tol = find_tolerance_in_text(raw)
    resistance_ohm = find_resistance_in_text(raw)
    power = find_power_in_text(raw)
    if size == "":
        for token in tokens:
            embedded_size = find_embedded_size(token)
            if embedded_size != "":
                size = embedded_size
                break
    if tol == "":
        for token in tokens:
            if token in RESISTOR_TOLERANCE_CODE_MAP:
                tol = clean_tol_for_match(RESISTOR_TOLERANCE_CODE_MAP[token])
                break

    param_count = sum([
        1 if size else 0,
        1 if resistance_ohm is not None else 0,
        1 if tol else 0,
        1 if power else 0,
    ])
    if param_count == 0:
        return None

    summary = build_other_component_summary([
        format_resistance_display(resistance_ohm) if resistance_ohm is not None else "",
        clean_tol_for_display(tol) if tol else "",
        power,
        size,
    ])
    return {
        "品牌": "",
        "型号": raw,
        "尺寸（inch）": size,
        "材质（介质）": "",
        "容值_pf": None,
        "容值误差": tol,
        "耐压（V）": "",
        "器件类型": component_type,
        "规格摘要": summary,
        "_resistance_ohm": resistance_ohm,
        "_power": power,
        "_core_param_count": param_count,
        "_param_count": param_count,
    }

def looks_like_electrolytic_context(text):
    if looks_like_film_capacitor_context(text):
        return False
    upper = clean_text(text).upper()
    if upper == "":
        return False
    compact = normalize_component_keyword_compact(text)
    if matches_component_alias(text, "铝电解电容"):
        return True
    if "铝电容" in upper or "鋁電容" in upper:
        return True
    electrolytic_tokens = [
        "电解",
        "電解",
        "ELECTROLYTIC",
        "ALUMINUM ELECTROLYTIC",
        "ALUMINIUM ELECTROLYTIC",
        "E CAP",
        "E-CAP",
        "ECAP",
        "PE E CAP",
        "PE E-CAP",
        "PEECAP",
    ]
    capacitor_tokens = ["电容", "電容", "CAP", "CAPACITOR"]
    if any(token in upper for token in electrolytic_tokens) or any(token.replace(" ", "") in compact for token in electrolytic_tokens):
        if any(token in upper for token in capacitor_tokens) or any(token in compact for token in capacitor_tokens):
            return True
    if ("电容" in upper or "電容" in upper or "CAP" in compact or "CAPACITOR" in compact) and (
        "电解" in upper or "電解" in upper or "ELECTROLYTIC" in upper or "ELECTROLYTIC" in compact
    ):
        return True
    return "UF" in compact and (extract_body_size_from_text(upper) != "" or extract_pitch_from_text(upper) != "")


def parse_varistor_spec_query(line):
    raw = clean_text(line)
    if raw == "" or not looks_like_varistor_context(raw):
        return None

    component_type = detect_resistor_subtype_hint(raw) or "压敏电阻"
    tol = find_tolerance_in_text(raw)
    varistor_voltage = clean_voltage(find_varistor_voltage_in_text(raw))
    disc_size = clean_text(find_disc_size_code(raw))
    pitch = extract_pitch_from_text(raw)
    power = find_power_in_text(raw)

    param_count = sum([
        1 if varistor_voltage else 0,
        1 if tol else 0,
        1 if disc_size else 0,
        1 if pitch else 0,
        1 if power else 0,
    ])
    if param_count == 0:
        return None

    summary = build_other_component_summary([
        voltage_display(varistor_voltage) if varistor_voltage else "",
        clean_tol_for_display(tol) if tol else "",
        disc_size,
        pitch,
        power,
    ])
    return {
        "品牌": "",
        "型号": raw,
        "尺寸（inch）": "",
        "材质（介质）": "",
        "容值_pf": None,
        "容值误差": tol,
        "耐压（V）": varistor_voltage,
        "器件类型": component_type,
        "规格摘要": summary,
        "_varistor_voltage": varistor_voltage,
        "_disc_size": disc_size,
        "_pitch": pitch,
        "_power": power,
        "_core_param_count": param_count,
        "_param_count": param_count,
    }

def parse_thermistor_spec_query(line):
    raw = clean_text(line)
    if raw == "" or not looks_like_thermistor_context(raw):
        return None

    size = find_embedded_size(raw)
    tol = find_tolerance_in_text(raw)
    resistance_ohm = find_resistance_in_text(raw)

    param_count = sum([
        1 if size else 0,
        1 if resistance_ohm is not None else 0,
        1 if tol else 0,
    ])
    if param_count == 0:
        return None

    summary = build_other_component_summary([
        format_resistance_display(resistance_ohm) if resistance_ohm is not None else "",
        clean_tol_for_display(tol) if tol else "",
        size,
    ])
    return {
        "品牌": "",
        "型号": raw,
        "尺寸（inch）": size,
        "材质（介质）": "",
        "容值_pf": None,
        "容值误差": tol,
        "耐压（V）": "",
        "器件类型": "热敏电阻",
        "规格摘要": summary,
        "_resistance_ohm": resistance_ohm,
        "_core_param_count": param_count,
        "_param_count": param_count,
    }

def parse_film_capacitor_spec_query(line):
    raw = clean_text(line)
    if raw == "" or not looks_like_film_capacitor_context(raw):
        return None

    base_spec = parse_spec_query(raw) or {}
    cap_pf = base_spec.get("容值_pf", None)
    tol = clean_tol_for_match(base_spec.get("容值误差", ""))
    volt = clean_voltage(base_spec.get("耐压（V）", ""))
    body_size = extract_body_size_from_text(raw)
    pitch = extract_pitch_from_text(raw)
    material = find_film_material(raw)
    safety_class = find_safety_class(raw)

    param_count = sum([
        1 if cap_pf is not None else 0,
        1 if tol else 0,
        1 if volt else 0,
        1 if body_size else 0,
        1 if pitch else 0,
        1 if material else 0,
        1 if safety_class else 0,
    ])
    if param_count == 0:
        return None

    value, unit = pf_to_value_unit(cap_pf)
    summary = build_other_component_summary([
        material,
        f"{value}{unit}" if value != "" and unit != "" else "",
        clean_tol_for_display(tol) if tol else "",
        voltage_display(volt) if volt else "",
        safety_class,
        body_size,
        pitch,
    ])
    return {
        "品牌": "",
        "型号": raw,
        "尺寸（inch）": "",
        "材质（介质）": material,
        "容值_pf": cap_pf,
        "容值误差": tol,
        "耐压（V）": volt,
        "器件类型": "薄膜电容",
        "规格摘要": summary,
        "_body_size": body_size,
        "_pitch": pitch,
        "_safety_class": safety_class,
        "_core_param_count": param_count,
        "_param_count": param_count,
    }

def parse_electrolytic_spec_query(line):
    raw = clean_text(line)
    if raw == "" or not looks_like_electrolytic_context(raw):
        return None

    compact = raw.upper().replace("±", "+/-").replace("卤", "+/-")
    compact = compact.replace("μF", "UF").replace("µF", "UF")

    base_spec = parse_spec_query(raw) or {}
    cap_pf = base_spec.get("容值_pf", None)
    tol = clean_tol_for_match(base_spec.get("容值误差", ""))
    volt = clean_voltage(base_spec.get("耐压（V）", ""))
    body_size = extract_body_size_from_text(compact)
    pitch = extract_pitch_from_text(compact)
    work_temp = normalize_working_temperature_text(base_spec.get("工作温度", "")) or extract_working_temperature_from_text(compact)
    life_hours = normalize_life_hours_value(base_spec.get("寿命（h）", "")) or parse_life_hours_from_text(compact)
    mounting_style = normalize_mounting_style(base_spec.get("安装方式", "") or extract_mounting_style_from_text(compact))
    special_use = normalize_special_use(base_spec.get("特殊用途", "") or extract_special_use_from_text(compact))
    if tol == "":
        tol = find_tolerance_in_text(compact)
    if volt == "":
        volt = parse_voltage_from_text(compact)
    if volt == "":
        compact_match = re.search(r"(?:PF|NF|UF)(\d+(?:\.\d+)?)V", compact.replace(" ", ""))
        if compact_match:
            volt = clean_voltage(compact_match.group(1))

    param_count = sum([
        1 if cap_pf is not None else 0,
        1 if tol else 0,
        1 if volt else 0,
        1 if body_size else 0,
        1 if pitch else 0,
        1 if work_temp else 0,
        1 if life_hours else 0,
        1 if mounting_style else 0,
        1 if special_use else 0,
    ])
    if param_count == 0:
        return None

    value, unit = pf_to_value_unit(cap_pf)
    summary = build_other_component_summary([
        f"{value}{unit}" if value != "" and unit != "" else "",
        clean_tol_for_display(tol) if tol else "",
        voltage_display(volt) if volt else "",
        work_temp,
        format_life_hours_display(life_hours),
        body_size,
        mounting_style,
        special_use,
        pitch,
    ])
    return {
        "品牌": "",
        "型号": raw,
        "尺寸（inch）": "",
        "材质（介质）": "",
        "容值_pf": cap_pf,
        "容值误差": tol,
        "耐压（V）": volt,
        "工作温度": work_temp,
        "寿命（h）": life_hours,
        "安装方式": mounting_style,
        "特殊用途": special_use,
        "器件类型": "铝电解电容",
        "规格摘要": summary,
        "_body_size": body_size,
        "_pitch": pitch,
        "_core_param_count": param_count,
        "_param_count": param_count,
    }

def parse_other_passive_query(line):
    for parser in [
        parse_inductor_spec_query,
        parse_film_capacitor_spec_query,
        parse_electrolytic_spec_query,
        parse_varistor_spec_query,
        parse_thermistor_spec_query,
        parse_resistor_spec_query,
    ]:
        parsed = parser(line)
        if parsed is not None:
            return parsed
    return None


def build_component_detail_lines(
    component_type,
    size="",
    material="",
    value="",
    unit="",
    tol="",
    volt="",
    resistance_ohm=None,
    power="",
    body_size="",
    pitch="",
    safety_class="",
    varistor_voltage="",
    disc_size="",
    work_temp="",
    life_hours="",
    mounting_style="",
    special_use="",
    production_status="",
    package_code="",
    diameter_mm="",
    height_mm="",
    polarity="",
    esr="",
    ripple_current="",
    resistance_25c="",
    resistance_unit="",
    resistance_tol="",
    b_value="",
    b_value_condition="",
    common_mode_impedance="",
    impedance_unit="",
    rated_current="",
    dcr="",
    circuit_count="",
    inductance_value="",
    inductance_unit="",
    inductance_tol="",
    saturation_current="",
    shielding="",
    impedance_100mhz="",
    load_capacitance="",
    drive_level="",
    output_type="",
    duty_cycle="",
):
    lines = []
    component_type = normalize_component_type(component_type)
    size = clean_size(size)
    material = clean_material(material)
    value = clean_text(value)
    unit = clean_text(unit).upper()
    tol = clean_tol_for_display(tol) if clean_text(tol) != "" else ""
    volt = voltage_display(volt) if clean_text(volt) != "" else ""
    power = format_power_display(power)
    body_size = clean_text(body_size)
    pitch = clean_text(pitch)
    safety_class = clean_text(safety_class)
    varistor_voltage = voltage_display(varistor_voltage) if clean_text(varistor_voltage) != "" else ""
    disc_size = clean_text(disc_size)
    work_temp = normalize_working_temperature_text(work_temp)
    life_hours = normalize_life_hours_value(life_hours)
    mounting_style = normalize_mounting_style(mounting_style)
    special_use = normalize_special_use(special_use)
    production_status = clean_text(production_status)
    package_code = clean_text(package_code)
    diameter_mm = normalize_dimension_mm_value(diameter_mm)
    height_mm = normalize_dimension_mm_value(height_mm)
    polarity = clean_text(polarity)
    esr = clean_text(esr)
    ripple_current = format_current_display(ripple_current)
    resistance_25c = clean_text(resistance_25c)
    resistance_unit = normalize_library_value_unit(resistance_unit)
    resistance_tol = clean_tol_for_display(resistance_tol) if clean_text(resistance_tol) != "" else ""
    b_value = normalize_b_value_text(b_value)
    b_value_condition = clean_text(b_value_condition)
    common_mode_impedance = format_impedance_display(common_mode_impedance, impedance_unit)
    rated_current = format_current_display(rated_current)
    dcr = clean_text(dcr)
    circuit_count = clean_text(circuit_count)
    inductance_value = clean_text(inductance_value)
    inductance_unit = clean_text(inductance_unit)
    inductance_tol = clean_tol_for_display(inductance_tol) if clean_text(inductance_tol) != "" else ""
    saturation_current = format_current_display(saturation_current)
    shielding = clean_text(shielding)
    impedance_100mhz = format_impedance_display(impedance_100mhz, impedance_unit)
    load_capacitance = clean_text(load_capacitance)
    drive_level = clean_text(drive_level)
    output_type = clean_text(output_type)
    duty_cycle = clean_text(duty_cycle)
    pitch = format_pitch_display(pitch)

    if component_type == "MLCC":
        if size != "":
            lines.append(f"尺寸: {size}")
        if material != "":
            lines.append(f"介质: {material}")
        if value != "" and unit != "":
            lines.append(f"容值: {value} {unit}")
        if tol != "":
            lines.append(f"误差: {tol}")
        if volt != "":
            lines.append(f"耐压: {volt}")
        if work_temp != "":
            lines.append(f"温度: {work_temp}")
        if package_code != "":
            lines.append(f"封装: {package_code}")
        return lines

    if component_type in RESISTOR_COMPONENT_TYPES:
        if size != "":
            lines.append(f"尺寸: {size}")
        if resistance_ohm is not None:
            lines.append(f"阻值: {format_resistance_display(resistance_ohm)}")
        if tol != "":
            lines.append(f"误差: {tol}")
        if power != "":
            lines.append(f"功率: {power}")
        if volt != "":
            lines.append(f"耐压: {volt}")
        if work_temp != "":
            lines.append(f"温度: {work_temp}")
        if package_code != "":
            lines.append(f"封装: {package_code}")
        return lines

    if component_type == "热敏电阻":
        if size != "":
            lines.append(f"尺寸: {size}")
        if resistance_25c != "":
            lines.append(f"R25: {combine_value_and_unit(resistance_25c, resistance_unit)}")
        elif resistance_ohm is not None:
            lines.append(f"阻值: {format_resistance_display(resistance_ohm)}")
        if resistance_tol != "":
            lines.append(f"误差: {resistance_tol}")
        elif tol != "":
            lines.append(f"误差: {tol}")
        if b_value != "":
            b_text = f"B值: {b_value}"
            if b_value_condition != "":
                b_text = f"{b_text} ({b_value_condition})"
            lines.append(b_text)
        if work_temp != "":
            lines.append(f"温度: {work_temp}")
        if mounting_style != "":
            lines.append(f"安装: {mounting_style}")
        return lines

    if component_type in VARISTOR_COMPONENT_TYPES:
        if varistor_voltage != "":
            lines.append(f"压敏电压: {varistor_voltage}")
        if tol != "":
            lines.append(f"误差: {tol}")
        if disc_size != "":
            lines.append(f"规格: {disc_size}")
        if pitch != "":
            lines.append(f"{pitch}")
        if power != "":
            lines.append(f"功率: {power}")
        if package_code != "":
            lines.append(f"封装: {package_code}")
        return lines

    if component_type == "铝电解电容":
        if value != "" and unit != "":
            lines.append(f"容值: {value} {unit}")
        if tol != "":
            lines.append(f"误差: {tol}")
        if volt != "":
            lines.append(f"耐压: {volt}")
        if work_temp != "":
            lines.append(f"温度: {work_temp}")
        if life_hours != "":
            lines.append(f"寿命: {format_life_hours_display(life_hours)}")
        if body_size != "":
            lines.append(f"尺寸: {body_size}")
        elif diameter_mm != "" or height_mm != "":
            body_text = " x ".join(part for part in [diameter_mm, height_mm] if part != "")
            if body_text != "":
                lines.append(f"尺寸: {body_text}")
        if diameter_mm != "":
            lines.append(f"直径: {diameter_mm}mm")
        if height_mm != "":
            lines.append(f"高度: {height_mm}mm")
        if polarity != "":
            lines.append(f"极性: {polarity}")
        if esr != "":
            lines.append(f"ESR: {esr}")
        if ripple_current != "":
            lines.append(f"纹波电流: {ripple_current}")
        if mounting_style != "":
            lines.append(f"安装: {mounting_style}")
        if special_use != "":
            lines.append(f"用途: {special_use}")
        if pitch != "":
            lines.append(f"{pitch}")
        return lines

    if component_type == "钽电容":
        if size != "":
            lines.append(f"尺寸: {size}")
        if value != "" and unit != "":
            lines.append(f"容值: {value} {unit}")
        if tol != "":
            lines.append(f"误差: {tol}")
        if volt != "":
            lines.append(f"耐压: {volt}")
        if package_code != "":
            lines.append(f"封装: {package_code}")
        return lines

    if component_type == "薄膜电容":
        if material != "":
            lines.append(f"介质: {material}")
        if value != "" and unit != "":
            lines.append(f"容值: {value} {unit}")
        if tol != "":
            lines.append(f"误差: {tol}")
        if volt != "":
            lines.append(f"耐压: {volt}")
        if safety_class != "":
            lines.append(f"安规: {safety_class}")
        if body_size != "":
            lines.append(f"尺寸: {body_size}")
        if work_temp != "":
            lines.append(f"温度: {work_temp}")
        if mounting_style != "":
            lines.append(f"安装: {mounting_style}")
        if pitch != "":
            lines.append(f"{pitch}")
        return lines

    if component_type == "引线型陶瓷电容":
        if material != "":
            lines.append(f"介质: {material}")
        if value != "" and unit != "":
            lines.append(f"容值: {value} {unit}")
        if tol != "":
            lines.append(f"误差: {tol}")
        if volt != "":
            lines.append(f"耐压: {volt}")
        if body_size != "":
            lines.append(f"尺寸: {body_size}")
        if work_temp != "":
            lines.append(f"温度: {work_temp}")
        if mounting_style != "":
            lines.append(f"安装: {mounting_style}")
        if pitch != "":
            lines.append(f"{pitch}")
        return lines

    if component_type in INDUCTOR_COMPONENT_TYPES:
        if size != "":
            lines.append(f"尺寸: {size}")
        if component_type == "磁珠":
            bead_impedance = impedance_100mhz or combine_value_and_unit(value, unit)
            if bead_impedance != "":
                lines.append(f"阻抗@100MHz: {bead_impedance}")
        elif component_type == "共模电感":
            if common_mode_impedance != "":
                lines.append(f"共模阻抗: {common_mode_impedance}")
            elif value != "" and unit != "":
                lines.append(f"阻抗: {combine_value_and_unit(value, unit)}")
            if circuit_count != "":
                lines.append(f"回路数: {circuit_count}")
        else:
            induct_text = combine_value_and_unit(inductance_value or value, inductance_unit or unit)
            if induct_text != "":
                lines.append(f"电感值: {induct_text}")
            if inductance_tol != "":
                lines.append(f"误差: {inductance_tol}")
            elif tol != "":
                lines.append(f"误差: {tol}")
        if rated_current != "":
            lines.append(f"额定电流: {rated_current}")
        if saturation_current != "":
            lines.append(f"饱和电流: {saturation_current}")
        if dcr != "":
            lines.append(f"DCR: {dcr}")
        if shielding != "":
            lines.append(f"屏蔽: {shielding}")
        if work_temp != "":
            lines.append(f"温度: {work_temp}")
        if mounting_style != "":
            lines.append(f"安装: {mounting_style}")
        return lines

    if component_type in TIMING_COMPONENT_TYPES:
        if size != "":
            lines.append(f"尺寸: {size}")
        if value != "" and unit != "":
            label = "输出频率" if component_type == "振荡器" else "频率"
            lines.append(f"{label}: {value} {unit}")
        if tol != "":
            lines.append(f"频差: {tol}")
        if volt != "":
            lines.append(f"工作电压: {volt}")
        if work_temp != "":
            lines.append(f"温度: {work_temp}")
        if package_code != "":
            lines.append(f"封装: {package_code}")
        if component_type == "晶振":
            if load_capacitance != "":
                lines.append(f"负载电容: {load_capacitance}")
            if esr != "":
                lines.append(f"ESR: {esr}")
            if drive_level != "":
                lines.append(f"驱动电平: {drive_level}")
        else:
            if output_type != "":
                lines.append(f"输出: {output_type}")
            if duty_cycle != "":
                lines.append(f"占空比: {duty_cycle}")
        if production_status != "":
            lines.append(f"状态: {production_status}")
        return lines

    if size != "":
        lines.append(f"尺寸: {size}")
    if material != "":
        lines.append(f"材质: {material}")
    if value != "" and unit != "":
        lines.append(f"参数: {value} {unit}")
    if tol != "":
        lines.append(f"误差: {tol}")
    if volt != "":
        lines.append(f"耐压: {volt}")
    return lines


def format_component_detail_inline(lines):
    lines = [clean_text(line) for line in (lines or []) if clean_text(line) != ""]
    return " | ".join(lines)


def build_component_spec_detail_from_spec(spec):
    if spec is None:
        return ""
    component_type = infer_spec_component_type(spec)
    value, unit = spec_display_value_unit(spec)
    lines = build_component_detail_lines(
        component_type,
        size=spec.get("尺寸（inch）", ""),
        material=spec.get("材质（介质）", ""),
        value=value,
        unit=unit,
        tol=spec.get("容值误差", ""),
        volt=spec.get("耐压（V）", ""),
        resistance_ohm=spec.get("_resistance_ohm", None),
        power=spec.get("_power", ""),
        body_size=spec.get("_body_size", ""),
        pitch=spec.get("_pitch", ""),
        safety_class=spec.get("_safety_class", ""),
        varistor_voltage=spec.get("_varistor_voltage", spec.get("耐压（V）", "")),
        disc_size=spec.get("_disc_size", ""),
        work_temp=spec.get("工作温度", ""),
        life_hours=spec.get("寿命（h）", ""),
        mounting_style=spec.get("安装方式", ""),
        special_use=spec.get("特殊用途", ""),
        production_status=spec.get("生产状态", ""),
        package_code=spec.get("封装代码", ""),
        diameter_mm=spec.get("直径（mm）", ""),
        height_mm=spec.get("高度（mm）", ""),
        polarity=spec.get("极性", ""),
        esr=spec.get("ESR", ""),
        ripple_current=spec.get("纹波电流", ""),
        resistance_25c=spec.get("阻值@25C", ""),
        resistance_unit=spec.get("阻值单位", ""),
        resistance_tol=spec.get("阻值误差", ""),
        b_value=spec.get("B值", ""),
        b_value_condition=spec.get("B值条件", ""),
        common_mode_impedance=spec.get("共模阻抗", ""),
        impedance_unit=spec.get("阻抗单位", ""),
        rated_current=spec.get("额定电流", ""),
        dcr=spec.get("DCR", ""),
        circuit_count=spec.get("回路数", ""),
        inductance_value=spec.get("电感值", ""),
        inductance_unit=spec.get("电感单位", ""),
        inductance_tol=spec.get("电感误差", ""),
        saturation_current=spec.get("饱和电流", ""),
        shielding=spec.get("屏蔽类型", ""),
        impedance_100mhz=spec.get("阻抗@100MHz", ""),
        load_capacitance=spec.get("负载电容（pF）", ""),
        drive_level=spec.get("驱动电平", ""),
        output_type=spec.get("输出类型", ""),
        duty_cycle=spec.get("占空比", ""),
    )
    return format_component_detail_inline(lines)


def spec_display_value_unit(spec):
    if spec is None:
        return "", ""
    component_type = infer_spec_component_type(spec)
    if component_type in ALL_RESISTOR_TYPES:
        return ohm_to_value_unit(spec.get("_resistance_ohm", None))
    if component_type in INDUCTOR_COMPONENT_TYPES:
        raw_value = clean_text(spec.get("电感值", "")) or clean_text(spec.get("容值", ""))
        raw_unit = clean_text(spec.get("电感单位", "")) or clean_text(spec.get("容值单位", ""))
        if raw_value != "" and raw_unit != "":
            return normalize_numeric_range_text(raw_value), raw_unit.upper()
        return parse_inductance_value_unit(find_inductance_in_text(build_component_context_text(spec)))
    if component_type in TIMING_COMPONENT_TYPES:
        raw_value = clean_text(spec.get("输出频率", "")) or clean_text(spec.get("频率", "")) or clean_text(spec.get("容值", ""))
        raw_unit = clean_text(spec.get("频率单位", "")) or clean_text(spec.get("容值单位", ""))
        if raw_value != "" and raw_unit != "":
            return normalize_numeric_range_text(raw_value), raw_unit.upper()
        return parse_frequency_value_unit(find_frequency_in_text(build_component_context_text(spec)))
    cap_pf = spec.get("容值_pf", None)
    if cap_pf is not None:
        return pf_to_value_unit(cap_pf)
    return normalize_numeric_range_text(clean_text(spec.get("容值", ""))), clean_text(spec.get("容值单位", "")).upper()


def resolve_component_display_type(spec_or_type):
    if isinstance(spec_or_type, dict):
        return normalize_component_type(resolve_component_category_type(spec_or_type))
    if isinstance(spec_or_type, pd.Series):
        return normalize_component_type(resolve_component_category_type(spec_or_type.to_dict()))
    return normalize_component_type(spec_or_type)


def get_component_display_schema(spec_or_type):
    component_type = resolve_component_display_type(spec_or_type)
    if component_type == "MLCC":
        return [
            ("系列", "系列"),
            ("系列说明", "系列说明"),
            ("尺寸（inch）", "尺寸"),
            ("材质（介质）", "材质（介质）"),
            ("容值", "容值"),
            ("容值单位", "容值单位"),
            ("容值误差", "容值误差"),
            ("耐压（V）", "耐压（V）"),
            ("长度（mm）", "长度(mm)"),
            ("宽度（mm）", "宽度(mm)"),
            ("高度（mm）", "厚度(mm)"),
            ("尺寸来源", "尺寸来源"),
        ]
    if component_type in RESISTOR_COMPONENT_TYPES:
        return [
            ("系列", "系列"),
            ("系列说明", "系列说明"),
            ("尺寸（inch）", "尺寸"),
            ("容值", "阻值"),
            ("容值单位", "阻值单位"),
            ("容值误差", "误差"),
            ("功率", "功率"),
            ("耐压（V）", "最高工作电压"),
            ("工作温度", "工作温度"),
            ("安装方式", "安装方式"),
            ("生产状态", "生产状态"),
        ]
    if component_type == "热敏电阻":
        return [
            ("系列", "系列"),
            ("系列说明", "系列说明"),
            ("尺寸（inch）", "尺寸"),
            ("阻值@25C", "R25"),
            ("阻值单位", "阻值单位"),
            ("阻值误差", "误差"),
            ("B值", "B值"),
            ("B值条件", "B值条件"),
            ("工作温度", "工作温度"),
            ("安装方式", "安装方式"),
            ("生产状态", "生产状态"),
        ]
    if component_type in INDUCTOR_COMPONENT_TYPES:
        if component_type == "磁珠":
            return [
                ("系列", "系列"),
                ("系列说明", "系列说明"),
                ("尺寸（inch）", "尺寸"),
                ("阻抗@100MHz", "阻抗@100MHz"),
                ("阻抗单位", "阻抗单位"),
                ("额定电流", "额定电流"),
                ("DCR", "DCR"),
                ("工作温度", "工作温度"),
                ("安装方式", "安装方式"),
                ("生产状态", "生产状态"),
            ]
        if component_type == "共模电感":
            return [
                ("系列", "系列"),
                ("系列说明", "系列说明"),
                ("尺寸（inch）", "尺寸"),
                ("电感值", "电感值"),
                ("电感单位", "电感单位"),
                ("共模阻抗", "共模阻抗"),
                ("阻抗单位", "阻抗单位"),
                ("额定电流", "额定电流"),
                ("DCR", "DCR"),
                ("回路数", "回路数"),
                ("工作温度", "工作温度"),
                ("安装方式", "安装方式"),
                ("生产状态", "生产状态"),
            ]
        return [
            ("系列", "系列"),
            ("系列说明", "系列说明"),
            ("尺寸（inch）", "尺寸"),
            ("电感值", "电感值"),
            ("电感单位", "电感单位"),
            ("电感误差", "误差"),
            ("额定电流", "额定电流"),
            ("饱和电流", "饱和电流"),
            ("DCR", "DCR"),
            ("屏蔽类型", "屏蔽类型"),
            ("工作温度", "工作温度"),
            ("安装方式", "安装方式"),
            ("生产状态", "生产状态"),
        ]
    if component_type == "晶振":
        return [
            ("系列", "系列"),
            ("系列说明", "系列说明"),
            ("尺寸（inch）", "尺寸"),
            ("封装代码", "封装代码"),
            ("容值", "频率"),
            ("容值单位", "频率单位"),
            ("容值误差", "频差"),
            ("工作温度", "工作温度"),
            ("负载电容（pF）", "负载电容(pF)"),
            ("ESR", "ESR"),
            ("驱动电平", "驱动电平"),
            ("安装方式", "安装方式"),
            ("生产状态", "生产状态"),
        ]
    if component_type == "振荡器":
        return [
            ("系列", "系列"),
            ("系列说明", "系列说明"),
            ("尺寸（inch）", "尺寸"),
            ("封装代码", "封装代码"),
            ("容值", "输出频率"),
            ("容值单位", "频率单位"),
            ("容值误差", "频差"),
            ("耐压（V）", "工作电压（V）"),
            ("工作温度", "工作温度"),
            ("输出类型", "输出类型"),
            ("占空比", "占空比"),
            ("安装方式", "安装方式"),
            ("生产状态", "生产状态"),
        ]
    if component_type in VARISTOR_COMPONENT_TYPES:
        if component_type == "贴片压敏电阻":
            return [
                ("系列", "系列"),
                ("尺寸（inch）", "尺寸"),
                ("压敏电压", "压敏电压"),
                ("容值误差", "误差"),
                ("功率", "功率"),
                ("安装方式", "安装方式"),
                ("生产状态", "生产状态"),
            ]
        return [
            ("系列", "系列"),
            ("压敏电压", "压敏电压"),
            ("容值误差", "误差"),
            ("规格", "规格"),
            ("脚距", "脚距"),
            ("功率", "功率"),
            ("安装方式", "安装方式"),
            ("生产状态", "生产状态"),
        ]
    if component_type == "铝电解电容":
        return [
            ("系列", "系列"),
            ("系列说明", "系列说明"),
            ("容值", "容值"),
            ("容值单位", "容值单位"),
            ("容值误差", "容值误差"),
            ("耐压（V）", "耐压（V）"),
            ("工作温度", "工作温度"),
            ("寿命（h）", "寿命(h)"),
            ("直径（mm）", "直径(mm)"),
            ("高度（mm）", "高度(mm)"),
            ("尺寸(mm)", "外形(mm)"),
            ("极性", "极性"),
            ("ESR", "ESR"),
            ("纹波电流", "纹波电流"),
            ("安装方式", "安装方式"),
            ("特殊用途", "特殊用途"),
            ("脚距", "脚距"),
            ("生产状态", "生产状态"),
        ]
    if component_type == "钽电容":
        return [
            ("系列", "系列"),
            ("尺寸（inch）", "尺寸"),
            ("容值", "容值"),
            ("容值单位", "容值单位"),
            ("容值误差", "容值误差"),
            ("耐压（V）", "耐压（V）"),
            ("生产状态", "生产状态"),
        ]
    if component_type == "薄膜电容":
        return [
            ("系列", "系列"),
            ("系列说明", "系列说明"),
            ("材质（介质）", "材质（介质）"),
            ("容值", "容值"),
            ("容值单位", "容值单位"),
            ("容值误差", "容值误差"),
            ("耐压（V）", "耐压（V）"),
            ("安规", "安规"),
            ("尺寸(mm)", "尺寸(mm)"),
            ("脚距", "脚距"),
            ("工作温度", "工作温度"),
            ("安装方式", "安装方式"),
            ("生产状态", "生产状态"),
        ]
    if component_type == "引线型陶瓷电容":
        return [
            ("系列", "系列"),
            ("材质（介质）", "材质（介质）"),
            ("容值", "容值"),
            ("容值单位", "容值单位"),
            ("容值误差", "容值误差"),
            ("耐压（V）", "耐压（V）"),
            ("尺寸(mm)", "尺寸(mm)"),
            ("脚距", "脚距"),
            ("工作温度", "工作温度"),
            ("安装方式", "安装方式"),
            ("生产状态", "生产状态"),
        ]
    return [
        ("系列", "系列"),
        ("尺寸（inch）", "尺寸"),
        ("材质（介质）", "材质（介质）"),
        ("容值", "参数值"),
        ("容值单位", "参数单位"),
        ("容值误差", "误差"),
        ("耐压（V）", "额定电压（V）"),
    ]


def get_component_header_labels(spec_or_type):
    labels = {source: label for source, label in get_component_display_schema(spec_or_type)}
    labels["前5个其他品牌型号"] = "其他品牌型号"
    labels["推荐品牌1"] = "推荐品牌1"
    labels["推荐型号1"] = "推荐型号1"
    labels["推荐品牌2"] = "推荐品牌2"
    labels["推荐型号2"] = "推荐型号2"
    labels["推荐品牌3"] = "推荐品牌3"
    labels["推荐型号3"] = "推荐型号3"
    return labels


def infer_mlcc_dimension_fields_from_record(record, allow_online_lookup=True):
    fields, _ = infer_mlcc_dimension_fields_and_source_from_record(record, allow_online_lookup=allow_online_lookup)
    return fields


def enrich_mlcc_dimension_fields_in_record(record, allow_online_lookup=False):
    enriched = dict(record) if isinstance(record, dict) else {}
    fields, source_text = infer_mlcc_dimension_fields_and_source_from_record(enriched, allow_online_lookup=allow_online_lookup)
    if fields:
        enriched = merge_dimension_fields_into_record(enriched, fields)
    if source_text != "" and clean_text(enriched.get("尺寸来源", "")) == "":
        enriched["尺寸来源"] = source_text
    return enriched


def enrich_mlcc_dimension_fields_in_dataframe(df, spec_or_type=None, allow_online_lookup=False):
    if df is None or df.empty:
        return df
    component_type = resolve_component_display_type(spec_or_type)
    if component_type != "MLCC":
        if "器件类型" not in df.columns:
            return df
        type_mask = df["器件类型"].astype(str).apply(normalize_component_type).eq("MLCC")
        if not type_mask.any():
            return df
    out = df.copy()
    for dim_col in ["长度（mm）", "宽度（mm）", "高度（mm）"]:
        if dim_col not in out.columns:
            out[dim_col] = ""
    if "尺寸来源" not in out.columns:
        out["尺寸来源"] = ""
    for idx in out.index.tolist():
        if (
            all(clean_text(out.at[idx, col]) != "" for col in ["长度（mm）", "宽度（mm）", "高度（mm）"] if col in out.columns)
            and clean_text(out.at[idx, "尺寸来源"]) != ""
        ):
            continue
        record = out.loc[idx].to_dict()
        fields, source_text = infer_mlcc_dimension_fields_and_source_from_record(record, allow_online_lookup=allow_online_lookup)
        if fields:
            for col, value in fields.items():
                if clean_text(out.at[idx, col]) == "":
                    out.at[idx, col] = value
        if source_text != "" and clean_text(out.at[idx, "尺寸来源"]) == "":
            out.at[idx, "尺寸来源"] = source_text
    return out


def build_component_display_row(spec, allow_online_lookup=False):
    spec = enrich_mlcc_dimension_fields_in_record(spec, allow_online_lookup=allow_online_lookup)
    component_type = normalize_component_type(spec.get("器件类型", ""))
    if component_type == "MLCC" and any(
        clean_text(spec.get(col, "")) == ""
        for col in ["长度（mm）", "宽度（mm）", "高度（mm）"]
    ):
        official_url_candidates = []
        for url_col in ["官网链接", "备注3", "备注2"]:
            official_url_candidates.extend(extract_url_candidates_from_text(spec.get(url_col, "")))
        if official_url_candidates:
            spec = enrich_mlcc_dimension_fields_in_record(spec, allow_online_lookup=True)
    display_fallbacks = infer_display_fallback_fields_from_record(spec)
    if display_fallbacks:
        merged_spec = dict(spec)
        for col, value in display_fallbacks.items():
            if clean_text(merged_spec.get(col, "")) == "" and clean_text(value) != "":
                merged_spec[col] = value
        spec = merged_spec
    value, unit = spec_display_value_unit(spec)
    diameter_mm = normalize_dimension_mm_value(spec.get("直径（mm）", ""))
    height_mm = normalize_dimension_mm_value(spec.get("高度（mm）", ""))
    body_size = clean_text(spec.get("_body_size", spec.get("尺寸（mm）", "")))
    if body_size == "" and (diameter_mm != "" or height_mm != ""):
        body_size = " x ".join(part for part in [diameter_mm, height_mm] if part != "")
    disc_size = clean_text(spec.get("_disc_size", ""))
    series_code = clean_text(spec.get("系列", ""))
    series_desc = clean_text(spec.get("系列说明", ""))
    special_use = clean_text(spec.get("特殊用途", ""))
    component_type = resolve_component_display_type(spec)
    series_profile = resolve_official_series_profile(
        brand=spec.get("品牌", ""),
        model=spec.get("型号", ""),
        component_type=component_type,
        series=series_code,
        series_desc=series_desc,
        special_use=special_use,
    )
    profile_type = normalize_component_type(series_profile.get("器件类型", ""))
    if profile_type != "":
        component_type = profile_type
    profile_series = clean_text(series_profile.get("系列", ""))
    if profile_series != "":
        series_code = profile_series
    if series_desc == "":
        series_desc = clean_text(series_profile.get("系列说明", ""))
    if special_use == "":
        special_use = clean_text(series_profile.get("特殊用途", ""))
    if series_desc == "" and series_code != "":
        series_desc = build_series_description_fallback(
            component_type,
            brand=spec.get("品牌", ""),
            series=series_code,
            special_use=special_use,
        )
    pitch_display = format_pitch_display(spec.get("_pitch", "") or spec.get("脚距", "") or spec.get("脚距（mm）", ""))
    size_display = clean_size(spec.get("尺寸（inch）", ""))
    if (
        component_type == "共模电感"
        and normalize_mounting_style(spec.get("安装方式", ""), spec.get("封装代码", "")) == "THT"
        and body_size != ""
    ):
        size_display = body_size
    power_source = clean_text(spec.get("_power", "")) or clean_text(spec.get("功率", ""))
    return {
        "器件类别": format_component_category_display(component_type),
        "系列": series_code,
        "系列说明": series_desc,
        "尺寸（inch）": format_component_size_display_value(size_display),
        "封装代码": clean_text(spec.get("封装代码", "")),
        "材质（介质）": clean_material(spec.get("材质（介质）", "")),
        "容值": value,
        "容值单位": clean_text(unit).upper(),
        "容值误差": clean_tol_for_display(spec.get("容值误差", "")),
        "耐压（V）": voltage_display(spec.get("耐压（V）", "")),
        "长度（mm）": normalize_dimension_mm_value(spec.get("长度（mm）", "")),
        "宽度（mm）": normalize_dimension_mm_value(spec.get("宽度（mm）", "")),
        "高度（mm）": height_mm,
        "尺寸来源": clean_text(spec.get("尺寸来源", "")),
        "工作温度": normalize_working_temperature_text(spec.get("工作温度", "")),
        "寿命（h）": format_life_hours_display(spec.get("寿命（h）", "")),
        "功率": format_power_display(power_source),
        "安装方式": normalize_mounting_style(spec.get("安装方式", ""), spec.get("封装代码", "")),
        "特殊用途": normalize_special_use(special_use),
        "脚距": pitch_display,
        "安规": clean_text(spec.get("_safety_class", "")),
        "规格": disc_size or body_size,
        "压敏电压": voltage_display(spec.get("_varistor_voltage", spec.get("耐压（V）", ""))),
        "尺寸(mm)": body_size,
        "生产状态": clean_text(spec.get("生产状态", "")),
        "直径（mm）": diameter_mm,
        "极性": clean_text(spec.get("极性", "")),
        "ESR": clean_text(spec.get("ESR", "")),
        "纹波电流": format_current_display(spec.get("纹波电流", "")),
        "阻值@25C": clean_text(spec.get("阻值@25C", "")),
        "阻值单位": normalize_library_value_unit(spec.get("阻值单位", "")),
        "阻值误差": clean_tol_for_display(spec.get("阻值误差", "")),
        "B值": normalize_b_value_text(spec.get("B值", "")),
        "B值条件": clean_text(spec.get("B值条件", "")),
        "共模阻抗": clean_text(spec.get("共模阻抗", "")),
        "阻抗单位": normalize_library_value_unit(spec.get("阻抗单位", "")),
        "额定电流": format_current_display(spec.get("额定电流", "")),
        "DCR": clean_text(spec.get("DCR", "")),
        "回路数": clean_text(spec.get("回路数", "")),
        "电感值": clean_text(spec.get("电感值", "")),
        "电感单位": clean_text(spec.get("电感单位", "")).upper(),
        "电感误差": clean_tol_for_display(spec.get("电感误差", "")),
        "饱和电流": format_current_display(spec.get("饱和电流", "")),
        "屏蔽类型": clean_text(spec.get("屏蔽类型", "")),
        "阻抗@100MHz": clean_text(spec.get("阻抗@100MHz", "")),
        "负载电容（pF）": clean_text(spec.get("负载电容（pF）", "")),
        "驱动电平": clean_text(spec.get("驱动电平", "")),
        "输出类型": clean_text(spec.get("输出类型", "")),
        "占空比": clean_text(spec.get("占空比", "")),
    }


def build_component_context_text(record):
    fields = [
        "器件类型", "品牌", "型号", "系列", "系列说明", "安装方式", "封装代码",
        "尺寸（inch）", "尺寸（mm）", "材质（介质）", "规格摘要", "容值", "容值单位", "容值误差",
        "工作温度", "寿命（h）", "特殊用途", "生产状态",
        "备注1", "备注2", "备注3", "官网链接", "数据来源", "校验备注",
        "阻值@25C", "阻值单位", "阻值误差", "B值", "B值条件",
        "共模阻抗", "阻抗单位", "额定电流", "DCR", "回路数",
        "电感值", "电感单位", "电感误差", "饱和电流", "屏蔽类型", "阻抗@100MHz",
        "直径（mm）", "高度（mm）", "极性", "ESR", "纹波电流",
        "输出频率", "频率", "频率单位", "频差（ppm）", "电源电压",
        "负载电容（pF）", "驱动电平", "输出类型", "占空比",
    ]
    return " ".join(clean_text(record.get(field, "")) for field in fields if clean_text(record.get(field, "")) != "")


def normalize_official_source_text(text):
    raw = html.unescape(clean_text(text))
    if raw == "":
        return ""
    normalized = raw.replace("\xa0", " ")
    normalized = normalized.replace("µ", "u").replace("μ", "u").replace("Μ", "U").replace("Ω", "Ω")
    normalized = normalized.replace('">', " ").replace("“", '"').replace("”", '"')
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def extract_inductor_series_family_from_text(text):
    upper = normalize_official_source_text(text).upper()
    if upper == "":
        return ""
    patterns = [
        r"\b([A-Z]{2,8}(?:-[A-Z0-9]+)?)\s+SERIES\b",
        r"\bPRODUCT\s+SERIES[:\s]+([A-Z]{2,8}(?:-[A-Z0-9]+)?)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, upper, flags=re.I)
        if match:
            return clean_text(match.group(1)).upper()
    return ""


def infer_official_inductor_fields_from_row(row):
    if row is None:
        return {}
    component_type = normalize_component_type(row.get("器件类型", "")) or infer_db_component_type(row)
    if component_type not in INDUCTOR_COMPONENT_TYPES:
        return {}

    text_fields = [
        "规格摘要", "备注1", "备注2", "备注3", "特殊用途", "官网链接",
        "数据来源", "校验备注", "系列", "型号", "尺寸（mm）",
    ]
    row_text = " | ".join(
        normalize_official_source_text(row.get(field, ""))
        for field in text_fields
        if normalize_official_source_text(row.get(field, "")) != ""
    )
    if row_text == "":
        row_text = build_component_context_text(row)

    result = {}
    brand_text = clean_brand(row.get("品牌", ""))
    model_text = clean_text(row.get("型号", ""))
    series_text = clean_text(row.get("系列", ""))
    special_text = normalize_official_source_text(row.get("特殊用途", ""))
    source_text = normalize_official_source_text(row.get("数据来源", ""))
    combined_series_text = " | ".join(part for part in [special_text, source_text, row_text] if part != "")

    inferred_series = extract_inductor_series_family_from_text(combined_series_text)
    if inferred_series != "" and (series_text == "" or series_text.upper() == model_text.upper()):
        result["系列"] = inferred_series

    if component_type in POWER_INDUCTOR_COMPONENT_TYPES:
        has_inductance_range = text_contains_inductance_range(row_text)
        has_current_range = text_contains_current_range(row_text)
        if (clean_text(row.get("电感值", "")) == "" or clean_text(row.get("电感单位", "")) == "") and not has_inductance_range:
            inductance_value, inductance_unit = parse_inductance_value_unit(find_inductance_in_text(row_text))
            if clean_text(row.get("电感值", "")) == "" and inductance_value != "":
                result["电感值"] = inductance_value
            if clean_text(row.get("电感单位", "")) == "" and inductance_unit != "":
                result["电感单位"] = inductance_unit
        if clean_text(row.get("额定电流", "")) == "" and not has_current_range:
            rated_current = find_labeled_current_in_text(row_text, ["RATED CURRENT", "HEATING CURRENT", "IRMS", "IR", "IDC", "CURRENT RANGE", "CURRENT"])
            if rated_current != "":
                result["额定电流"] = rated_current
        if clean_text(row.get("饱和电流", "")) == "":
            saturation_current = find_labeled_current_in_text(row_text, ["SATURATION CURRENT", "ISAT", "SAT CURRENT"])
            if saturation_current != "":
                result["饱和电流"] = saturation_current
        if clean_text(row.get("DCR", "")) == "":
            dcr = find_dcr_in_text(row_text)
            if dcr != "":
                result["DCR"] = dcr
        if clean_text(row.get("屏蔽类型", "")) == "":
            shielding = find_shielding_type_in_text(row_text)
            if shielding != "":
                result["屏蔽类型"] = shielding
    elif component_type == "共模电感":
        has_inductance_range = text_contains_inductance_range(row_text)
        has_current_range = text_contains_current_range(row_text)
        has_impedance_range = text_contains_impedance_range(row_text)
        if (clean_text(row.get("电感值", "")) == "" or clean_text(row.get("电感单位", "")) == "") and not has_inductance_range:
            inductance_value, inductance_unit = parse_inductance_value_unit(find_inductance_in_text(row_text))
            if clean_text(row.get("电感值", "")) == "" and inductance_value != "":
                result["电感值"] = inductance_value
            if clean_text(row.get("电感单位", "")) == "" and inductance_unit != "":
                result["电感单位"] = inductance_unit
        if clean_text(row.get("共模阻抗", "")) == "" and not has_impedance_range:
            impedance = find_impedance_at_100mhz_in_text(row_text)
            value, unit = parse_impedance_value_unit(impedance)
            if value != "":
                result["共模阻抗"] = value
            if clean_text(row.get("阻抗单位", "")) == "" and unit != "":
                result["阻抗单位"] = unit
        if clean_text(row.get("额定电流", "")) == "" and not has_current_range:
            rated_current = find_labeled_current_in_text(row_text, ["RATED CURRENT", "IR", "CURRENT"])
            if rated_current != "":
                result["额定电流"] = rated_current
        if clean_text(row.get("DCR", "")) == "":
            dcr = find_dcr_in_text(row_text)
            if dcr != "":
                result["DCR"] = dcr
    elif component_type == "磁珠":
        has_impedance_range = text_contains_impedance_range(row_text)
        if clean_text(row.get("阻抗@100MHz", "")) == "" and not has_impedance_range:
            impedance = find_impedance_at_100mhz_in_text(row_text)
            value, unit = parse_impedance_value_unit(impedance)
            if value != "":
                result["阻抗@100MHz"] = value
            if clean_text(row.get("阻抗单位", "")) == "" and unit != "":
                result["阻抗单位"] = unit
        if clean_text(row.get("额定电流", "")) == "":
            rated_current = find_labeled_current_in_text(row_text, ["RATED CURRENT", "IR", "CURRENT"])
            if rated_current != "":
                result["额定电流"] = rated_current
        if clean_text(row.get("DCR", "")) == "":
            dcr = find_dcr_in_text(row_text)
            if dcr != "":
                result["DCR"] = dcr

    if clean_text(row.get("工作温度", "")) == "":
        working_temp = extract_working_temperature_from_text(row_text)
        if working_temp != "":
            result["工作温度"] = working_temp
    if clean_text(row.get("安装方式", "")) == "":
        mounting_style = extract_mounting_style_from_text(row_text)
        if mounting_style != "":
            result["安装方式"] = mounting_style
    if clean_text(row.get("生产状态", "")) == "":
        status_match = re.search(r"\b(ACTIVE|NEW|DISCONTINUED|INACTIVE|PRODUCTION)\b", row_text, flags=re.I)
        if status_match:
            result["生产状态"] = status_match.group(1).capitalize()

    display_series = clean_text(result.get("系列", "")) or series_text
    current_series_desc = clean_text(row.get("系列说明", ""))
    should_refresh_series_desc = current_series_desc == ""
    if (
        not should_refresh_series_desc
        and display_series != ""
        and model_text != ""
        and display_series.upper() != model_text.upper()
        and re.search(re.escape(model_text), current_series_desc, flags=re.I)
    ):
        should_refresh_series_desc = True
    if should_refresh_series_desc and display_series != "":
        if component_type in POWER_INDUCTOR_COMPONENT_TYPES:
            if "RF CHOKE" in row_text.upper():
                result["系列说明"] = f"{brand_text} {display_series} 射频电感系列".strip()
            elif "SHIELDED" in row_text.upper():
                result["系列说明"] = f"{brand_text} {display_series} 屏蔽功率电感系列".strip()
            elif component_type == "射频电感":
                result["系列说明"] = f"{brand_text} {display_series} 射频电感系列".strip()
            else:
                result["系列说明"] = f"{brand_text} {display_series} 功率电感系列".strip()
        elif component_type == "共模电感":
            if "HIGH VOLTAGE" in row_text.upper():
                result["系列说明"] = f"{brand_text} {display_series} 高压共模电感系列".strip()
            elif "FILTER" in row_text.upper():
                result["系列说明"] = f"{brand_text} {display_series} 共模滤波器系列".strip()
            else:
                result["系列说明"] = f"{brand_text} {display_series} 共模电感系列".strip()
        elif component_type == "磁珠":
            result["系列说明"] = f"{brand_text} {display_series} 磁珠系列".strip()
    return result


def infer_component_display_fallbacks_from_row(row):
    if row is None:
        return {}
    component_type = normalize_component_type(row.get("器件类型", "")) or infer_db_component_type(row)
    row_text = build_component_context_text(row)
    series_profile = resolve_official_series_profile(
        brand=row.get("品牌", ""),
        model=row.get("型号", ""),
        component_type=component_type,
        series=row.get("系列", ""),
        series_desc=row.get("系列说明", ""),
        special_use=row.get("特殊用途", ""),
    )
    result = {
        key: clean_text(series_profile.get(key, ""))
        for key in ["系列", "系列说明", "特殊用途"]
        if clean_text(series_profile.get(key, "")) != ""
    }
    if normalize_component_type(series_profile.get("器件类型", "")) != "":
        component_type = normalize_component_type(series_profile.get("器件类型", ""))
    result.update(infer_official_inductor_fields_from_row(row))

    series_code = clean_text(row.get("系列", ""))
    if clean_text(result.get("系列", "")) != "":
        series_code = clean_text(result.get("系列", ""))
    if clean_text(row.get("系列说明", "")) == "" and series_code != "":
        fallback_desc = build_series_description_fallback(
            component_type,
            brand=row.get("品牌", ""),
            series=series_code,
            special_use=row.get("特殊用途", ""),
        )
        if fallback_desc != "":
            result["系列说明"] = fallback_desc
    if clean_text(row.get("工作温度", "")) == "":
        working_temp = extract_working_temperature_from_text(row_text)
        if working_temp != "":
            result["工作温度"] = working_temp
    if clean_text(row.get("寿命（h）", "")) == "":
        life_hours = parse_life_hours_from_text(row_text)
        if life_hours != "":
            result["寿命（h）"] = format_life_hours_display(life_hours)
    if clean_text(row.get("安装方式", "")) == "":
        mounting_style = extract_mounting_style_from_text(row_text)
        if mounting_style != "":
            result["安装方式"] = mounting_style

    display_fallbacks = infer_display_fallback_fields_from_record(row)
    for col, value in display_fallbacks.items():
        if col not in result and clean_text(value) != "":
            result[col] = value

    if component_type == "铝电解电容":
        diameter_mm, height_mm, body_text = extract_electrolytic_dimensions_from_text(row_text)
        if clean_text(row.get("直径（mm）", "")) == "" and diameter_mm != "":
            result["直径（mm）"] = diameter_mm
        if clean_text(row.get("高度（mm）", "")) == "" and height_mm != "":
            result["高度（mm）"] = height_mm
        if clean_text(row.get("尺寸(mm)", "")) == "" and body_text != "":
            result["尺寸(mm)"] = body_text
        if clean_text(row.get("极性", "")) == "":
            result["极性"] = "有极性"
        if clean_text(row.get("ESR", "")) == "":
            esr = find_esr_in_text(row_text)
            if esr != "":
                result["ESR"] = esr
        if clean_text(row.get("纹波电流", "")) == "":
            ripple = find_labeled_current_in_text(row_text, ["RIPPLE CURRENT", "RIPPLE"])
            if ripple != "":
                result["纹波电流"] = ripple

    if component_type == "热敏电阻":
        if clean_text(row.get("阻值@25C", "")) == "" or clean_text(row.get("阻值单位", "")) == "":
            resistance_ohm = find_resistance_in_text(row_text)
            if resistance_ohm is not None:
                resistance_value, resistance_unit = ohm_to_library_value_unit(resistance_ohm)
                if clean_text(row.get("阻值@25C", "")) == "" and resistance_value != "":
                    result["阻值@25C"] = resistance_value
                if clean_text(row.get("阻值单位", "")) == "" and resistance_unit != "":
                    result["阻值单位"] = resistance_unit
        if clean_text(row.get("阻值误差", "")) == "":
            tol = clean_tol_for_display(row.get("容值误差", "")) or clean_tol_for_display(find_tolerance_in_text(row_text))
            if tol != "":
                result["阻值误差"] = tol
        if clean_text(row.get("B值", "")) == "":
            b_value = find_b_value_in_text(row_text)
            if b_value != "":
                result["B值"] = b_value
        if clean_text(row.get("B值条件", "")) == "":
            b_condition = find_b_value_condition_in_text(row_text)
            if b_condition != "":
                result["B值条件"] = b_condition

    if component_type in INDUCTOR_COMPONENT_TYPES:
        if component_type in POWER_INDUCTOR_COMPONENT_TYPES:
            has_inductance_range = text_contains_inductance_range(row_text)
            has_current_range = text_contains_current_range(row_text)
            if (clean_text(row.get("电感值", "")) == "" or clean_text(row.get("电感单位", "")) == "") and not has_inductance_range:
                inductance_value, inductance_unit = parse_inductance_value_unit(find_inductance_in_text(row_text))
                if clean_text(row.get("电感值", "")) == "" and inductance_value != "":
                    result["电感值"] = inductance_value
                if clean_text(row.get("电感单位", "")) == "" and inductance_unit != "":
                    result["电感单位"] = inductance_unit
            if clean_text(row.get("电感误差", "")) == "":
                tol = clean_tol_for_display(find_tolerance_in_text(row_text))
                if tol != "":
                    result["电感误差"] = tol
            if clean_text(row.get("额定电流", "")) == "" and not has_current_range:
                rated_current = find_labeled_current_in_text(row_text, ["RATED CURRENT", "HEATING CURRENT", "IRMS", "IDC", "CURRENT"])
                if rated_current != "":
                    result["额定电流"] = rated_current
            if clean_text(row.get("饱和电流", "")) == "":
                saturation_current = find_labeled_current_in_text(row_text, ["SATURATION CURRENT", "ISAT", "SAT CURRENT"])
                if saturation_current != "":
                    result["饱和电流"] = saturation_current
            if clean_text(row.get("DCR", "")) == "":
                dcr = find_dcr_in_text(row_text)
                if dcr != "":
                    result["DCR"] = dcr
            if clean_text(row.get("屏蔽类型", "")) == "":
                shielding = find_shielding_type_in_text(row_text)
                if shielding != "":
                    result["屏蔽类型"] = shielding
        elif component_type == "共模电感":
            has_inductance_range = text_contains_inductance_range(row_text)
            has_current_range = text_contains_current_range(row_text)
            has_impedance_range = text_contains_impedance_range(row_text)
            if (clean_text(row.get("电感值", "")) == "" or clean_text(row.get("电感单位", "")) == "") and not has_inductance_range:
                inductance_value, inductance_unit = parse_inductance_value_unit(find_inductance_in_text(row_text))
                if clean_text(row.get("电感值", "")) == "" and inductance_value != "":
                    result["电感值"] = inductance_value
                if clean_text(row.get("电感单位", "")) == "" and inductance_unit != "":
                    result["电感单位"] = inductance_unit
            if clean_text(row.get("共模阻抗", "")) == "" and not has_impedance_range:
                impedance = find_impedance_at_100mhz_in_text(row_text)
                value, unit = parse_impedance_value_unit(impedance)
                if value != "":
                    result["共模阻抗"] = value
                if clean_text(row.get("阻抗单位", "")) == "" and unit != "":
                    result["阻抗单位"] = unit
            if clean_text(row.get("额定电流", "")) == "" and not has_current_range:
                rated_current = find_labeled_current_in_text(row_text, ["RATED CURRENT", "IR", "CURRENT"])
                if rated_current != "":
                    result["额定电流"] = rated_current
            if clean_text(row.get("DCR", "")) == "":
                dcr = find_dcr_in_text(row_text)
                if dcr != "":
                    result["DCR"] = dcr
            if clean_text(row.get("回路数", "")) == "":
                circuit_count = find_circuit_count_in_text(row_text)
                if circuit_count != "":
                    result["回路数"] = circuit_count
        elif component_type == "磁珠":
            has_impedance_range = text_contains_impedance_range(row_text)
            if clean_text(row.get("阻抗@100MHz", "")) == "" and not has_impedance_range:
                impedance = find_impedance_at_100mhz_in_text(row_text)
                value, unit = parse_impedance_value_unit(impedance)
                if value != "":
                    result["阻抗@100MHz"] = value
                if clean_text(row.get("阻抗单位", "")) == "" and unit != "":
                    result["阻抗单位"] = unit
            if clean_text(row.get("额定电流", "")) == "":
                rated_current = find_labeled_current_in_text(row_text, ["RATED CURRENT", "CURRENT"])
                if rated_current != "":
                    result["额定电流"] = rated_current
            if clean_text(row.get("DCR", "")) == "":
                dcr = find_dcr_in_text(row_text)
                if dcr != "":
                    result["DCR"] = dcr

    if component_type in TIMING_COMPONENT_TYPES:
        if clean_text(row.get("容值", "")) == "" or clean_text(row.get("容值单位", "")) == "":
            frequency_value, frequency_unit = parse_frequency_value_unit(find_frequency_in_text(row_text))
            if clean_text(row.get("容值", "")) == "" and frequency_value != "":
                result["容值"] = frequency_value
            if clean_text(row.get("容值单位", "")) == "" and frequency_unit != "":
                result["容值单位"] = frequency_unit
        if clean_text(row.get("容值误差", "")) == "":
            frequency_tol = clean_text(row.get("频差（ppm）", "")) or find_frequency_tolerance_in_text(row_text)
            if frequency_tol != "":
                result["容值误差"] = frequency_tol
        if component_type == "晶振":
            if clean_text(row.get("ESR", "")) == "":
                esr = find_esr_in_text(row_text)
                if esr != "":
                    result["ESR"] = esr
        else:
            if clean_text(row.get("耐压（V）", "")) == "" and clean_text(row.get("电源电压", "")) != "":
                result["耐压（V）"] = clean_text(row.get("电源电压", ""))

    return result


def ensure_component_display_columns(df):
    if df is None:
        return pd.DataFrame()
    out = df.copy()
    for col in out.columns:
        if isinstance(out[col].dtype, pd.CategoricalDtype):
            out[col] = out[col].astype("object")
    row_count = len(out.index)

    def blank_series():
        return pd.Series([""] * row_count, index=out.index, dtype="object")

    if "功率" not in out.columns:
        out["功率"] = out["_power"].apply(format_power_display) if "_power" in out.columns else blank_series()
    if "脚距" not in out.columns:
        if "_pitch" in out.columns:
            out["脚距"] = out["_pitch"].astype(str).apply(format_pitch_display)
        elif "脚距（mm）" in out.columns:
            out["脚距"] = out["脚距（mm）"].astype(str).apply(format_pitch_display)
        else:
            out["脚距"] = blank_series()
    if "安规" not in out.columns:
        out["安规"] = out["_safety_class"].astype(str).apply(clean_text) if "_safety_class" in out.columns else blank_series()
    for col in [
        "封装代码", "生产状态", "直径（mm）", "极性", "ESR", "纹波电流",
        "阻值@25C", "阻值单位", "阻值误差", "B值", "B值条件",
        "共模阻抗", "阻抗单位", "额定电流", "DCR", "回路数",
        "电感值", "电感单位", "电感误差", "饱和电流", "屏蔽类型", "阻抗@100MHz",
        "负载电容（pF）", "驱动电平", "输出类型", "占空比",
    ]:
        if col not in out.columns:
            out[col] = blank_series()
        out[col] = out[col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
    if "系列" not in out.columns:
        out["系列"] = blank_series()
    out["系列"] = out["系列"].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
    if "系列说明" not in out.columns:
        out["系列说明"] = blank_series()
    out["系列说明"] = out["系列说明"].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
    if "器件类别" not in out.columns:
        out["器件类别"] = blank_series()
    out["器件类别"] = out["器件类别"].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
    if "特殊用途" not in out.columns:
        out["特殊用途"] = blank_series()
    if "尺寸（inch）" not in out.columns:
        out["尺寸（inch）"] = blank_series()
    out["尺寸（inch）"] = out["尺寸（inch）"].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
    size_fallback = blank_series()
    for fallback_col in ["尺寸（mm）", "尺寸(mm)", "_body_size"]:
        if fallback_col in out.columns:
            size_fallback = out[fallback_col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
            break
    fill_size_mask = out["尺寸（inch）"].eq("") & size_fallback.ne("")
    if fill_size_mask.any():
        out.loc[fill_size_mask, "尺寸（inch）"] = size_fallback.loc[fill_size_mask]
    if "器件类型" in out.columns and "安装方式" in out.columns:
        tht_common_mode_mask = (
            out["器件类型"].astype(str).apply(normalize_component_type).eq("共模电感")
            & out["安装方式"].astype(str).apply(normalize_mounting_style).eq("THT")
            & size_fallback.ne("")
        )
        if tht_common_mode_mask.any():
            out.loc[tht_common_mode_mask, "尺寸（inch）"] = size_fallback.loc[tht_common_mode_mask]
    out["尺寸（inch）"] = out["尺寸（inch）"].astype(str).replace("nan", "").replace("None", "").apply(format_component_size_display_value)
    timing_value_sources = [col for col in ["输出频率", "频率"] if col in out.columns]
    for source_col in timing_value_sources:
        blank_value_mask = out["容值"].astype(str).apply(clean_text).eq("")
        if blank_value_mask.any():
            out.loc[blank_value_mask, "容值"] = out.loc[blank_value_mask, source_col].astype(str).apply(clean_text)
    if "频率单位" in out.columns:
        blank_unit_mask = out["容值单位"].astype(str).apply(clean_text).eq("")
        if blank_unit_mask.any():
            out.loc[blank_unit_mask, "容值单位"] = out.loc[blank_unit_mask, "频率单位"].astype(str).apply(clean_text)
    if "频差（ppm）" in out.columns:
        blank_tol_mask = out["容值误差"].astype(str).apply(clean_text).eq("")
        if blank_tol_mask.any():
            out.loc[blank_tol_mask, "容值误差"] = out.loc[blank_tol_mask, "频差（ppm）"].astype(str).apply(clean_text)
    if "电源电压" in out.columns:
        blank_volt_mask = out["耐压（V）"].astype(str).apply(clean_text).eq("")
        if blank_volt_mask.any():
            out.loc[blank_volt_mask, "耐压（V）"] = out.loc[blank_volt_mask, "电源电压"].astype(str).apply(clean_text)
    if "型号" in out.columns:
        mlcc_type_mask = (
            out["器件类型"].astype(str).apply(normalize_component_type).eq("MLCC")
            if "器件类型" in out.columns
            else pd.Series(True, index=out.index)
        )
        if mlcc_type_mask.any():
            mlcc_profiles = out.loc[mlcc_type_mask].apply(
                lambda row: resolve_mlcc_series_profile(
                    brand=row.get("品牌", ""),
                    model=row.get("型号", ""),
                    series=row.get("系列", ""),
                    series_desc=row.get("系列说明", ""),
                    special_use=row.get("特殊用途", ""),
                ),
                axis=1,
            )
            mlcc_profile_df = pd.DataFrame(list(mlcc_profiles), index=mlcc_profiles.index)
            series_values = mlcc_profile_df["系列"].astype(str).apply(clean_text)
            series_valid_idx = series_values[series_values.ne("")].index
            if len(series_valid_idx) > 0:
                out.loc[series_valid_idx, "系列"] = series_values.loc[series_valid_idx]
            for col in ["系列说明", "特殊用途"]:
                blank_mask = out.loc[mlcc_type_mask, col].astype(str).apply(clean_text).eq("")
                if blank_mask.any():
                    blank_idx = blank_mask[blank_mask].index
                    fill_values = mlcc_profile_df.loc[blank_idx, col].astype(str).apply(clean_text)
                    valid_idx = fill_values[fill_values.ne("")].index
                    if len(valid_idx) > 0:
                        out.loc[valid_idx, col] = fill_values.loc[valid_idx]
    non_mlcc_series_mask = out["系列说明"].astype(str).apply(clean_text).eq("") & out["系列"].astype(str).apply(clean_text).ne("")
    if non_mlcc_series_mask.any():
        filled_desc = out.loc[non_mlcc_series_mask].apply(
            lambda row: build_series_description_fallback(
                row.get("器件类型", ""),
                brand=row.get("品牌", ""),
                series=row.get("系列", ""),
                special_use=row.get("特殊用途", ""),
            ),
            axis=1,
        )
        valid_idx = filled_desc[filled_desc.astype(str).apply(clean_text).ne("")].index
        if len(valid_idx) > 0:
            out.loc[valid_idx, "系列说明"] = filled_desc.loc[valid_idx]
    if "规格" not in out.columns:
        if "_disc_size" in out.columns:
            disc_size = out["_disc_size"].astype(str).apply(clean_text)
        else:
            disc_size = blank_series()
        if "_body_size" in out.columns:
            body_size = out["_body_size"].astype(str).apply(clean_text)
        elif "尺寸（mm）" in out.columns:
            body_size = out["尺寸（mm）"].astype(str).apply(clean_text)
        else:
            body_size = blank_series()
        out["规格"] = disc_size.where(disc_size.ne(""), body_size)
    if "工作温度" not in out.columns:
        out["工作温度"] = blank_series()
    out["工作温度"] = out["工作温度"].astype(str).apply(normalize_working_temperature_text)
    if "寿命（h）" not in out.columns:
        out["寿命（h）"] = blank_series()
    out["寿命（h）"] = out["寿命（h）"].astype(str).apply(format_life_hours_display)
    out["功率"] = out["功率"].astype(str).apply(format_power_display)
    if "安装方式" not in out.columns:
        out["安装方式"] = blank_series()
    if "封装代码" in out.columns:
        out["安装方式"] = [
            normalize_mounting_style(value, package_code)
            for value, package_code in zip(out["安装方式"].astype(str).tolist(), out["封装代码"].astype(str).tolist())
        ]
    else:
        out["安装方式"] = out["安装方式"].astype(str).apply(normalize_mounting_style)
    out["特殊用途"] = out["特殊用途"].astype(str).apply(normalize_special_use)
    if "品牌" in out.columns and "型号" in out.columns and "器件类型" in out.columns:
        bourns_cm1309_mask = (
            out["品牌"].astype(str).apply(clean_brand).str.upper().eq("BOURNS")
            & out["器件类型"].astype(str).apply(normalize_component_type).eq("共模电感")
            & out["型号"].astype(str).apply(clean_model).str.match(r"^CM1309\dCT-\d+$", na=False)
        )
        if bourns_cm1309_mask.any():
            out.loc[bourns_cm1309_mask, "系列"] = "CM1309"
            out.loc[bourns_cm1309_mask, "系列说明"] = "Bourns CM1309 共模电感系列"
            out.loc[bourns_cm1309_mask, "安装方式"] = "THT"
            if "封装代码" in out.columns:
                out.loc[bourns_cm1309_mask, "封装代码"] = "CM1309"
    if "压敏电压" not in out.columns:
        if "_varistor_voltage" in out.columns:
            out["压敏电压"] = out["_varistor_voltage"].apply(voltage_display)
        else:
            out["压敏电压"] = blank_series()
    if "尺寸(mm)" not in out.columns:
        if "_body_size" in out.columns:
            out["尺寸(mm)"] = out["_body_size"].astype(str).apply(clean_text)
        elif "尺寸（mm）" in out.columns:
            out["尺寸(mm)"] = out["尺寸（mm）"].astype(str).apply(clean_text)
        else:
            out["尺寸(mm)"] = blank_series()
    if "直径（mm）" in out.columns:
        out["直径（mm）"] = out["直径（mm）"].astype(str).apply(normalize_dimension_mm_value)
    for dim_col in ["长度（mm）", "宽度（mm）", "高度（mm）"]:
        if dim_col not in out.columns:
            out[dim_col] = blank_series()
        out[dim_col] = out[dim_col].astype(str).replace("nan", "").replace("None", "").apply(normalize_dimension_mm_value)
    out["脚距"] = out["脚距"].astype(str).apply(format_pitch_display)
    out["容值单位"] = out["容值单位"].astype(str).apply(normalize_library_value_unit)
    out["容值误差"] = out["容值误差"].astype(str).apply(clean_tol_for_display)
    out["耐压（V）"] = out["耐压（V）"].astype(str).apply(voltage_display)
    out["阻值单位"] = out["阻值单位"].astype(str).apply(normalize_library_value_unit)
    out["阻值误差"] = out["阻值误差"].astype(str).apply(clean_tol_for_display)
    out["B值"] = out["B值"].astype(str).apply(normalize_b_value_text)
    out["阻抗单位"] = out["阻抗单位"].astype(str).apply(normalize_library_value_unit)
    out["电感单位"] = out["电感单位"].astype(str).apply(lambda value: clean_text(value).upper())
    out["电感误差"] = out["电感误差"].astype(str).apply(clean_tol_for_display)
    for current_col in ["纹波电流", "额定电流", "饱和电流"]:
        out[current_col] = out[current_col].astype(str).apply(format_current_display)
    if "尺寸来源" not in out.columns:
        out["尺寸来源"] = blank_series()
    if row_count > 0 and row_count <= 5000:
        fallback_series = out.apply(lambda row: infer_component_display_fallbacks_from_row(row), axis=1)
        fallback_df = pd.DataFrame(list(fallback_series), index=out.index)
        for col in fallback_df.columns:
            if col not in out.columns:
                out[col] = blank_series()
            fill_values = fallback_df[col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
            blank_mask = out[col].astype(str).replace("nan", "").replace("None", "").apply(clean_text).eq("")
            valid_mask = fill_values.ne("")
            target_idx = fill_values[blank_mask & valid_mask].index
            if len(target_idx) > 0:
                out.loc[target_idx, col] = fill_values.loc[target_idx]
        inductor_mask = out["器件类型"].astype(str).apply(normalize_component_type).isin(INDUCTOR_COMPONENT_TYPES) if "器件类型" in out.columns else pd.Series(False, index=out.index)
        if inductor_mask.any():
            official_series = out.loc[inductor_mask].apply(lambda row: infer_official_inductor_fields_from_row(row), axis=1)
            official_df = pd.DataFrame(list(official_series), index=out.loc[inductor_mask].index)
            for col in official_df.columns:
                if col not in out.columns:
                    out[col] = blank_series()
                fill_values = official_df[col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
                if col == "系列":
                    current_series = out.loc[inductor_mask, col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
                    current_models = out.loc[inductor_mask, "型号"].astype(str).replace("nan", "").replace("None", "").apply(clean_text) if "型号" in out.columns else pd.Series("", index=current_series.index)
                    replace_mask = current_series.eq("") | current_series.eq(current_models)
                    target_idx = fill_values[replace_mask & fill_values.ne("")].index
                elif col == "系列说明":
                    current_desc = out.loc[inductor_mask, col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
                    current_models = out.loc[inductor_mask, "型号"].astype(str).replace("nan", "").replace("None", "").apply(clean_text) if "型号" in out.columns else pd.Series("", index=current_desc.index)
                    replace_mask = current_desc.eq("")
                    model_tokens = [re.escape(model) for model in current_models[current_models.ne("")].unique()]
                    if model_tokens:
                        replace_mask = replace_mask | current_desc.str.contains("|".join(model_tokens), case=False, na=False)
                    target_idx = fill_values[replace_mask & fill_values.ne("")].index
                else:
                    blank_mask = out.loc[inductor_mask, col].astype(str).replace("nan", "").replace("None", "").apply(clean_text).eq("")
                    target_idx = fill_values[blank_mask & fill_values.ne("")].index
                if len(target_idx) > 0:
                    out.loc[target_idx, col] = fill_values.loc[target_idx]
    if row_count > 0:
        out["器件类别"] = out.apply(lambda row: format_component_category_display(row), axis=1)
    return out


def fill_component_display_blanks(df, spec):
    if df is None or df.empty or spec is None:
        return df
    out = df.copy()
    fallback = build_component_display_row(spec)
    for col, value in fallback.items():
        if col not in out.columns:
            continue
        blank_mask = out[col].astype(str).apply(clean_text).eq("")
        if blank_mask.any() and clean_text(value) != "":
            out.loc[blank_mask, col] = value
    return out


def select_component_display_columns(df, spec_or_type, prefix_columns=None, suffix_columns=None, allow_online_lookup=False):
    if df is None:
        return pd.DataFrame()
    out = ensure_component_display_columns(df)
    out = enrich_mlcc_dimension_fields_in_dataframe(
        out,
        spec_or_type=spec_or_type,
        allow_online_lookup=allow_online_lookup,
    )
    schema_columns = [source for source, _ in get_component_display_schema(spec_or_type)]
    selected = []
    for col in (prefix_columns or []):
        if col in out.columns and col not in selected:
            selected.append(col)
    for col in schema_columns:
        if col in out.columns and col not in selected:
            selected.append(col)
    for col in (suffix_columns or []):
        if col in out.columns and col not in selected:
            selected.append(col)
    return out[selected].copy() if selected else out


def build_component_column_config(columns, spec_or_type=None):
    header_labels = get_component_header_labels(spec_or_type)
    width_map = {
        "推荐等级": "small",
        "品牌": "small",
        "型号": "medium",
        "推荐品牌": "small",
        "推荐型号": "medium",
        "推荐品牌1": "small",
        "推荐型号1": "medium",
        "推荐品牌2": "small",
        "推荐型号2": "medium",
        "推荐品牌3": "small",
        "推荐型号3": "medium",
        "品牌1": "small",
        "型号1": "medium",
        "品牌2": "small",
        "型号2": "medium",
        "品牌3": "small",
        "型号3": "medium",
        "器件类别": "medium",
        "系列": "small",
        "系列说明": "medium",
        "信昌料号": "large",
        "华科料号": "large",
        "前5个其他品牌型号": "large",
        "其他品牌型号": "large",
        "规格参数明细": "large",
        "匹配参数明细": "large",
        "长度（mm）": "small",
        "宽度（mm）": "small",
        "高度（mm）": "small",
        "封装代码": "small",
        "生产状态": "small",
        "直径（mm）": "small",
        "极性": "small",
        "ESR": "small",
        "纹波电流": "small",
        "阻值@25C": "small",
        "阻值单位": "small",
        "阻值误差": "small",
        "B值": "small",
        "B值条件": "small",
        "共模阻抗": "small",
        "阻抗单位": "small",
        "额定电流": "small",
        "DCR": "small",
        "回路数": "small",
        "电感值": "small",
        "电感单位": "small",
        "电感误差": "small",
        "饱和电流": "small",
        "屏蔽类型": "small",
        "阻抗@100MHz": "small",
        "负载电容（pF）": "small",
        "驱动电平": "small",
        "输出类型": "small",
        "占空比": "small",
        "尺寸来源": "medium",
        "备注1": "medium",
        "备注2": "medium",
        "备注3": "medium",
    }
    config = {}
    for col in columns:
        label = header_labels.get(col, col)
        width = width_map.get(col)
        if width:
            config[col] = st.column_config.TextColumn(label, width=width)
        else:
            config[col] = st.column_config.TextColumn(label)
    return config


def build_component_section_title(spec, suffix):
    component_label = format_component_category_display(spec)
    if component_label == "":
        component_label = resolve_component_display_type(spec)
    return f"{component_label}{suffix}" if component_label != "" else suffix


def build_component_summary_from_spec(spec):
    if spec is None:
        return ""
    component_type = infer_spec_component_type(spec)
    value, unit = spec_display_value_unit(spec)
    if component_type == "MLCC":
        return build_other_component_summary([
            clean_size(spec.get("尺寸（inch）", "")),
            clean_material(spec.get("材质（介质）", "")),
            f"{value}{unit}" if value != "" and unit != "" else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            voltage_display(spec.get("耐压（V）", "")),
        ])
    if component_type in RESISTOR_COMPONENT_TYPES:
        return build_other_component_summary([
            format_resistance_display(spec.get("_resistance_ohm")) if spec.get("_resistance_ohm") is not None else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            clean_text(spec.get("_power", "")),
            clean_size(spec.get("尺寸（inch）", "")),
        ])
    if component_type == "热敏电阻":
        return build_other_component_summary([
            format_resistance_display(spec.get("_resistance_ohm")) if spec.get("_resistance_ohm") is not None else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            clean_size(spec.get("尺寸（inch）", "")),
        ])
    if component_type in VARISTOR_COMPONENT_TYPES:
        return build_other_component_summary([
            voltage_display(spec.get("_varistor_voltage", spec.get("耐压（V）", ""))),
            clean_tol_for_display(spec.get("容值误差", "")),
            clean_text(spec.get("_disc_size", "")),
            clean_text(spec.get("_pitch", "")),
            clean_text(spec.get("_power", "")),
        ])
    if component_type == "铝电解电容":
        return build_other_component_summary([
            f"{value}{unit}" if value != "" and unit != "" else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            voltage_display(spec.get("耐压（V）", "")),
            normalize_working_temperature_text(spec.get("工作温度", "")),
            format_life_hours_display(spec.get("寿命（h）", "")),
            clean_text(spec.get("_body_size", "")),
            normalize_mounting_style(spec.get("安装方式", "")),
            normalize_special_use(spec.get("特殊用途", "")),
            clean_text(spec.get("_pitch", "")),
        ])
    if component_type == "钽电容":
        return build_other_component_summary([
            clean_size(spec.get("尺寸（inch）", "")),
            f"{value}{unit}" if value != "" and unit != "" else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            voltage_display(spec.get("耐压（V）", "")),
        ])
    if component_type == "薄膜电容":
        return build_other_component_summary([
            clean_material(spec.get("材质（介质）", "")),
            f"{value}{unit}" if value != "" and unit != "" else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            voltage_display(spec.get("耐压（V）", "")),
            clean_text(spec.get("_safety_class", "")),
            clean_text(spec.get("_body_size", "")),
            clean_text(spec.get("_pitch", "")),
        ])
    if component_type == "引线型陶瓷电容":
        return build_other_component_summary([
            clean_material(spec.get("材质（介质）", "")),
            f"{value}{unit}" if value != "" and unit != "" else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            voltage_display(spec.get("耐压（V）", "")),
            clean_text(spec.get("_body_size", "")),
            clean_text(spec.get("_pitch", "")),
        ])
    if component_type in INDUCTOR_COMPONENT_TYPES:
        summary = build_other_component_summary([
            f"{value}{unit}" if value != "" and unit != "" else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            format_current_display(spec.get("额定电流", "")),
            clean_text(spec.get("DCR", "")),
            clean_text(spec.get("尺寸（mm）", "")) or clean_size(spec.get("尺寸（inch）", "")),
        ])
        return summary or clean_text(spec.get("规格摘要", ""))
    if component_type in TIMING_COMPONENT_TYPES:
        return build_other_component_summary([
            f"{value}{unit}" if value != "" and unit != "" else "",
            clean_tol_for_display(spec.get("容值误差", "")),
            voltage_display(spec.get("耐压（V）", "")),
            clean_size(spec.get("尺寸（inch）", "")),
        ])
    return clean_text(spec.get("规格摘要", ""))


def extract_body_size_from_text(text):
    upper = clean_text(text).upper()
    match3 = THREE_DIM_SIZE_PATTERN.search(upper)
    if match3:
        return f"{match3.group(1)}*{match3.group(2)}*{match3.group(3)}mm"
    match = ELECTROLYTIC_SIZE_PATTERN.search(upper)
    if not match:
        return ""
    return f"{match.group(1)}*{match.group(2)}mm"


def extract_pitch_from_text(text):
    upper = clean_text(text).upper()
    match = PITCH_PATTERN.search(upper)
    if match:
        return f"脚距{match.group(1)}mm"
    if looks_like_varistor_context(text):
        match = VARISTOR_PITCH_PATTERN.search(upper)
        if match:
            return f"脚距{match.group(1)}mm"
    return ""


def _format_compact_number(value):
    try:
        num = float(value)
    except Exception:
        return clean_text(value)
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    return f"{num:.6f}".rstrip("0").rstrip(".")


def normalize_numeric_range_text(value):
    text = clean_text(value)
    if text == "":
        return ""
    normalized = text.replace("−", "-").replace("–", "-").replace("—", "-")
    normalized = normalized.replace("～", "~").replace("至", "~")
    compact = re.sub(r"\s+", "", normalized)
    compact = re.sub(r"(?i)(?<=\d)TO(?=[+\-]?\d)", "~", compact)
    range_match = re.fullmatch(r"([+\-]?\d+(?:\.\d+)?)(?:~|-)([+\-]?\d+(?:\.\d+)?)", compact)
    if range_match:
        return f"{_format_compact_number(range_match.group(1))}~{_format_compact_number(range_match.group(2))}"
    if re.fullmatch(r"[+\-]?\d+(?:\.\d+)?", compact):
        return _format_compact_number(compact)
    return text


def working_temperature_bounds(value):
    text = clean_text(value)
    if text == "":
        return None, None
    upper = text.upper()
    upper = upper.replace("−", "-").replace("–", "-").replace("—", "-")
    upper = upper.replace("～", "~").replace("至", "~").replace("TO", "~")
    upper = upper.replace("℃", "C").replace("°C", "C")
    compact = re.sub(r"\s+", "", upper)

    range_match = re.search(
        r"([+\-]?\d+(?:\.\d+)?)\s*C?\s*(?:~|-)\s*([+\-]?\d+(?:\.\d+)?)(?:C)?",
        compact,
    )
    if range_match:
        try:
            return float(range_match.group(1)), float(range_match.group(2))
        except Exception:
            return None, None

    single_match = re.search(r"([+\-]?\d+(?:\.\d+)?)(?:C)$", compact)
    if single_match:
        try:
            single = float(single_match.group(1))
            return None, single
        except Exception:
            return None, None

    if re.fullmatch(r"[+\-]?\d+(?:\.\d+)?", compact):
        try:
            single = float(compact)
            return None, single
        except Exception:
            return None, None
    return None, None


def normalize_working_temperature_text(value):
    low, high = working_temperature_bounds(value)
    if low is None and high is None:
        return ""
    if low is None:
        return f"{_format_compact_number(high)}℃"
    return f"{_format_compact_number(low)}~{_format_compact_number(high)}℃"


def extract_working_temperature_from_text(text):
    raw = clean_text(text)
    if raw == "":
        return ""
    upper = raw.upper().replace("−", "-").replace("–", "-").replace("—", "-")
    upper = upper.replace("～", "~").replace("至", "~").replace("TO", "~").replace("℃", "C").replace("°C", "C")
    patterns = [
        r"(?:WORK(?:ING)?|CATEGORY|OPERATING)\s*TEMPERATURE(?:\s*RANGE)?[^0-9+\-]{0,24}([+\-]?\d+(?:\.\d+)?\s*(?:~|-)\s*[+\-]?\d+(?:\.\d+)?)",
        r"(?:工作温度|温度范围)[^0-9+\-]{0,8}([+\-]?\d+(?:\.\d+)?\s*(?:~|-)\s*[+\-]?\d+(?:\.\d+)?)",
        r"([+\-]?\d+(?:\.\d+)?\s*(?:~|-)\s*[+\-]?\d+(?:\.\d+)?)\s*C(?=$|[^0-9A-Z])",
        r"([+\-]?\d+(?:\.\d+)?)\s*C(?=$|[^0-9A-Z])",
    ]
    for pattern in patterns:
        match = re.search(pattern, upper, flags=re.I)
        if not match:
            continue
        normalized = normalize_working_temperature_text(match.group(1))
        if normalized != "":
            return normalized
    bare_negative_range = re.search(r"(-\d+(?:\.\d+)?)\s*(?:~|-)\s*(\+?\d+(?:\.\d+)?)(?=$|[^0-9A-Z])", upper, flags=re.I)
    if bare_negative_range:
        normalized = normalize_working_temperature_text(f"{bare_negative_range.group(1)}~{bare_negative_range.group(2)}℃")
        if normalized != "":
            return normalized
    return ""


def life_hours_to_number(value):
    text = clean_text(value).upper().replace(",", "")
    if text == "":
        return None
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        try:
            return float(text)
        except Exception:
            return None
    if not re.search(r"(?:H|HR|HRS|HOUR)", text):
        return None
    match = re.search(r"(\d{3,6}(?:\.\d+)?)", text)
    if not match:
        return None
    try:
        return float(match.group(1))
    except Exception:
        return None


def normalize_life_hours_value(value):
    number = life_hours_to_number(value)
    if number is None:
        return ""
    return _format_compact_number(number)


def parse_life_hours_from_text(text):
    raw = clean_text(text)
    if raw == "":
        return ""
    upper = raw.upper().replace(",", "").replace("小时", "H").replace("小時", "H")
    match = re.search(r"(\d{3,6}(?:\.\d+)?)\s*(?:H|HR|HRS|HOUR|HOURS)(?=$|[^0-9A-Z])", upper)
    if not match:
        return ""
    return normalize_life_hours_value(match.group(1))


def format_life_hours_display(value):
    normalized = normalize_life_hours_value(value)
    return f"{normalized}h" if normalized != "" else ""


def normalize_b_value_text(value):
    text = clean_text(value).upper().replace(" ", "")
    if text == "":
        return ""
    if text.endswith("KK"):
        text = text[:-1]
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(K)?", text)
    if match:
        return f"{match.group(1)}K"
    return clean_text(value)


def combine_value_and_unit(value, unit="", separator=" "):
    value_text = clean_text(value)
    unit_text = clean_text(unit)
    if value_text == "":
        return ""
    if unit_text == "":
        return value_text
    if value_text.upper().endswith(unit_text.upper()):
        return value_text
    joiner = "" if unit_text in {"Ω", "KΩ", "MΩ", "mΩ", "A", "mA", "uA", "nH", "uH", "mH", "H", "Hz", "kHz", "MHz", "GHz"} else separator
    return f"{value_text}{joiner}{unit_text}"


def format_current_display(value):
    text = clean_text(value).replace(" ", "")
    if text == "":
        return ""
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(UA|MA|A)", text, flags=re.I)
    if match:
        return f"{match.group(1)}{match.group(2).upper()}"
    return clean_text(value)


def format_pitch_display(value):
    text = clean_text(value)
    if text == "":
        return ""
    if "脚距" in text or "PITCH" in text.upper():
        return text
    normalized = normalize_dimension_mm_value(text)
    if normalized == "":
        return text
    return f"脚距{normalized}mm"


def format_impedance_display(value, unit=""):
    value_text = clean_text(value)
    unit_text = normalize_library_value_unit(unit)
    if value_text == "":
        return ""
    upper = value_text.upper()
    if any(token in upper for token in ["Ω", "OHM"]):
        return value_text.replace("OHMS", "Ω").replace("OHM", "Ω")
    if unit_text == "":
        unit_text = "Ω"
    return combine_value_and_unit(value_text, unit_text, separator=" ")


def parse_inductance_value_unit(text):
    normalized = clean_text(text).upper().replace("Μ", "U").replace("ΜH", "UH").replace("μ", "U").replace("µ", "U").replace(" ", "")
    if normalized == "":
        return "", ""
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(NH|UH|MH|H)", normalized, flags=re.I)
    if not match:
        return "", ""
    return clean_text(match.group(1)), clean_text(match.group(2)).upper()


def parse_frequency_value_unit(text):
    normalized = clean_text(text).upper().replace(" ", "")
    if normalized == "":
        return "", ""
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(HZ|KHZ|MHZ|GHZ)", normalized, flags=re.I)
    if not match:
        return "", ""
    return clean_text(match.group(1)), clean_text(match.group(2)).upper()


def parse_impedance_value_unit(text):
    normalized = clean_text(text).replace(" ", "")
    if normalized == "":
        return "", ""
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(MΩ|KΩ|Ω|mΩ|MOHM|KOHM|OHM)", normalized, flags=re.I)
    if not match:
        return "", ""
    return clean_text(match.group(1)), normalize_library_value_unit(match.group(2))


def extract_electrolytic_dimensions_from_text(text):
    upper = clean_text(text).upper().replace("Φ", "").replace("Ø", "")
    if upper == "":
        return "", "", ""
    match = ELECTROLYTIC_SIZE_PATTERN.search(upper)
    if not match:
        return "", "", ""
    diameter = normalize_dimension_mm_value(match.group(1))
    height = normalize_dimension_mm_value(match.group(2))
    body_text = " x ".join(part for part in [diameter, height] if part != "")
    return diameter, height, body_text


def build_series_description_fallback(component_type, brand="", series="", special_use=""):
    component_type = normalize_component_type(component_type)
    series_text = clean_text(series)
    if series_text == "":
        return ""
    special_text = normalize_special_use(special_use)
    brand_text = clean_brand(brand)
    prefix = f"{brand_text} " if brand_text != "" else ""
    if component_type == "铝电解电容":
        return f"{prefix}{series_text} 铝电解电容系列".strip()
    if component_type == "薄膜电容":
        return f"{prefix}{series_text} 薄膜电容系列".strip()
    if component_type == "钽电容":
        return f"{prefix}{series_text} 钽电容系列".strip()
    if component_type == "引线型陶瓷电容":
        return f"{prefix}{series_text} 引线型陶瓷电容系列".strip()
    if component_type in RESISTOR_COMPONENT_TYPES:
        if special_text != "":
            return f"{prefix}{series_text} {special_text} {component_type}系列".strip()
        return f"{prefix}{series_text} {component_type}系列".strip()
    if component_type == "热敏电阻":
        if special_text != "":
            return f"{prefix}{series_text} {special_text} 热敏电阻系列".strip()
        return f"{prefix}{series_text} 热敏电阻系列".strip()
    if component_type in VARISTOR_COMPONENT_TYPES:
        if special_text != "":
            return f"{prefix}{series_text} {special_text} 压敏电阻系列".strip()
        return f"{prefix}{series_text} 压敏电阻系列".strip()
    if component_type in INDUCTOR_COMPONENT_TYPES:
        suffix = "磁珠系列" if component_type == "磁珠" else f"{component_type}系列"
        return f"{prefix}{series_text} {suffix}".strip()
    if component_type in TIMING_COMPONENT_TYPES:
        return f"{prefix}{series_text} {component_type}系列".strip()
    if special_text != "":
        return f"{series_text} / {special_text}"
    return f"{prefix}{series_text} 系列".strip()


THERMISTOR_SERIES_RULES = [
    (re.compile(r"^(NTCLE\d{3}[A-Z]\d)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(B\d{5}[A-Z])", flags=re.I), "NTC thermistor series", ""),
    (re.compile(r"^(WBP-TR)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(WMF\d+)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(TS\d+)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(CT\d+)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(HCT)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(MF73T)", flags=re.I), "NTC inrush current limiter series", "浪涌抑制"),
    (re.compile(r"^(MF72)", flags=re.I), "NTC inrush current limiter series", "浪涌抑制"),
    (re.compile(r"^(MF58[A-Z]?)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(MF52[A-Z]?)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(MF25)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(MF11)", flags=re.I), "NTC thermistor series", "测温"),
    (re.compile(r"^(SCL\d+(?:\.\d+)?D?)", flags=re.I), "NTC inrush current limiter series", "浪涌抑制"),
    (re.compile(r"^(SL\d{2})", flags=re.I), "NTC inrush current limiter series", "浪涌抑制"),
    (re.compile(r"^(CL-\d{2}|CL\d{2})", flags=re.I), "NTC inrush current limiter series", "浪涌抑制"),
    (re.compile(r"^(NTC\d+D)", flags=re.I), "NTC inrush current limiter series", "浪涌抑制"),
    (re.compile(r"^(\d+(?:\.\d+)?D)", flags=re.I), "NTC inrush current limiter series", "浪涌抑制"),
    (re.compile(r"^(\d+D)", flags=re.I), "NTC inrush current limiter series", "浪涌抑制"),
]
VARISTOR_SERIES_RULES = [
    (re.compile(r"^(K\d{4}ESDA)", flags=re.I), lambda m: m.group(1).upper(), "贴片压敏电阻系列"),
    (re.compile(r"^(KESD\d{4})", flags=re.I), lambda m: m.group(1).upper(), "贴片压敏电阻系列"),
    (re.compile(r"^(CNR-\d+[A-Z])", flags=re.I), lambda m: m.group(1).upper(), "压敏电阻系列"),
    (re.compile(r"^(MGUG|VCUG)", flags=re.I), lambda m: m.group(1).upper(), "贴片压敏电阻系列"),
    (re.compile(r"^(\d{4})(?=-\d+(?:\.\d+)?V)", flags=re.I), lambda m: m.group(1).upper(), "贴片压敏电阻系列"),
    (re.compile(r"^(FNR)", flags=re.I), lambda m: "FNR", "氧化锌压敏电阻系列"),
    (re.compile(r"^(?:[A-Z]+-)?(\d+(?:\.\d+)?D)(?=-|\d|[A-Z]|$)", flags=re.I), lambda m: clean_text(m.group(1)).upper(), "氧化锌压敏电阻系列"),
    (re.compile(r"^\d{3}(KD)(\d{2})(?:[A-Z]|-|$)", flags=re.I), lambda m: f"{m.group(1).upper()}{m.group(2)}", "金属氧化物压敏电阻系列"),
    (re.compile(r"^(STE)(\d{2}D)", flags=re.I), lambda m: m.group(2).upper(), "氧化锌压敏电阻系列"),
    (re.compile(r"^(B722)(\d{2})([A-Z])", flags=re.I), lambda m: f"{m.group(3).upper()}{m.group(2)}K", "TDK/EPCOS SIOV 压敏电阻系列"),
    (re.compile(r"^(ERZ[A-Z0-9]{1,2})", flags=re.I), lambda m: m.group(1).upper(), "Panasonic 压敏电阻系列"),
    (re.compile(r"^(MLV\d{4})", flags=re.I), lambda m: m.group(1).upper(), "多层贴片压敏电阻系列"),
    (re.compile(r"^(V\d{4})", flags=re.I), lambda m: m.group(1).upper(), "贴片压敏电阻系列"),
]
TIMING_SERIES_PREFIX_MEANING = {
    "TG": "温补晶体振荡器 / TCXO",
    "SG": "石英晶体振荡器 / SPXO",
    "VG": "压控晶体振荡器 / VCXO",
    "FC": "石英晶体单元 / Crystal Unit",
    "FA": "石英晶体单元 / Crystal Unit",
}
PANASONIC_ALUMINUM_SERIES_PATTERNS = [
    re.compile(r"^(ECR\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(ECA\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(EEE\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(EEU\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(ECL\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(EKX\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(EKY\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(ECP\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(FCS\d[A-Z]{3})", flags=re.I),
    re.compile(r"^(FCR\d[A-Z]{3})", flags=re.I),
]


def passive_series_should_replace(current_value, canonical_value, model=""):
    current = clean_text(current_value)
    canonical = clean_text(canonical_value)
    model_text = clean_model(model)
    if canonical == "":
        return False
    if series_looks_missing(current):
        return True
    if current == canonical:
        return False
    current_upper = current.upper()
    canonical_upper = canonical.upper()
    if model_text != "" and current_upper == model_text.upper():
        return True
    if current_upper.startswith(canonical_upper) and len(current_upper) > len(canonical_upper):
        return True
    return False


def merge_series_profile_values(result, profile, model=""):
    if not isinstance(profile, dict):
        return dict(result)
    merged = dict(result)
    profile_series = clean_text(profile.get("系列", ""))
    series_replaced = False
    if passive_series_should_replace(merged.get("系列", ""), profile_series, model=model):
        merged["系列"] = profile_series
        series_replaced = True
    for col in ("系列说明", "特殊用途"):
        profile_value = clean_text(profile.get(col, ""))
        if profile_value == "":
            continue
        if clean_text(merged.get(col, "")) == "" or series_replaced:
            merged[col] = profile_value
    profile_type = normalize_component_type(profile.get("器件类型", ""))
    if normalize_component_type(merged.get("器件类型", "")) == "" and profile_type != "":
        merged["器件类型"] = profile_type
    return merged


def resolve_aluminum_electrolytic_series_profile(brand="", model="", series="", series_desc="", special_use=""):
    brand_text = clean_brand(brand)
    brand_upper = brand_text.upper()
    model_text = clean_model(model)
    result = {
        "系列": clean_text(series),
        "系列说明": clean_text(series_desc),
        "特殊用途": normalize_special_use(special_use),
        "器件类型": "铝电解电容",
    }
    if model_text == "":
        return result
    if "JIANGHAI" in brand_upper or "江海" in brand_text or jianghai_series_code_from_model(model_text) != "":
        series_code = jianghai_series_code_from_model(model_text)
        if series_code != "":
            profile = jianghai_series_profile(series_code, model_text)
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": jianghai_series_meaning(series_code, model_text),
                    "特殊用途": clean_text(profile.get("特殊用途", "")),
                    "器件类型": "铝电解电容",
                },
                model=model_text,
            )
    if "NICHICON" in brand_upper or "尼吉康" in brand_text:
        series_code = nichicon_series_code_from_model(model_text)
        if series_code != "":
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": f"Nichicon {series_code} 铝电解电容系列",
                    "器件类型": "铝电解电容",
                },
                model=model_text,
            )
    if any(token in brand_upper for token in ("CHEMI-CON", "CHEMICON")) or "贵弥功" in brand_text:
        series_code = chemi_con_series_code_from_model(model_text)
        if series_code != "":
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": f"Chemi-Con {series_code} 铝电解电容系列",
                    "器件类型": "铝电解电容",
                },
                model=model_text,
            )
    if "PANASONIC" in brand_upper or "松下" in brand_text:
        for pattern in PANASONIC_ALUMINUM_SERIES_PATTERNS:
            match = pattern.match(model_text)
            if match:
                series_code = clean_text(match.group(1)).upper()
                return merge_series_profile_values(
                    result,
                    {
                        "系列": series_code,
                        "系列说明": f"Panasonic {series_code} 铝电解电容系列",
                        "器件类型": "铝电解电容",
                    },
                    model=model_text,
                )
    if clean_text(result.get("系列", "")) != "" and clean_text(result.get("系列说明", "")) == "":
        result["系列说明"] = build_series_description_fallback("铝电解电容", brand=brand_text, series=result["系列"], special_use=result.get("特殊用途", ""))
    return result


def resolve_thermistor_series_profile(brand="", model="", series="", series_desc="", special_use=""):
    brand_text = clean_brand(brand)
    model_text = clean_model(model)
    result = {
        "系列": clean_text(series),
        "系列说明": clean_text(series_desc),
        "特殊用途": normalize_special_use(special_use),
        "器件类型": "热敏电阻",
    }
    if model_text == "":
        return result
    murata_profile = murata_ntc_series_profile_from_model(model_text)
    murata_series = clean_text(murata_profile.get("系列", ""))
    if murata_series != "":
        return merge_series_profile_values(
            result,
            {
                "系列": murata_series,
                "系列说明": clean_text(murata_profile.get("系列说明", "")),
                "特殊用途": clean_text(murata_profile.get("特殊用途", "")),
                "器件类型": "热敏电阻",
            },
            model=model_text,
        )
    for pattern, desc_suffix, inferred_use in THERMISTOR_SERIES_RULES:
        match = pattern.match(model_text)
        if not match:
            continue
        series_code = clean_text(match.group(1)).upper()
        series_brand = "TDK/EPCOS" if series_code.startswith("B") else brand_text
        return merge_series_profile_values(
            result,
            {
                "系列": series_code,
                "系列说明": f"{series_brand} {series_code} {desc_suffix}".strip(),
                "特殊用途": normalize_special_use("/".join(part for part in [result.get("特殊用途", ""), inferred_use] if clean_text(part) != "")),
                "器件类型": "热敏电阻",
            },
            model=model_text,
        )
    if "FENGHUA" in brand_text.upper() or "风华" in brand_text:
        fenghua_match = re.match(r"^(CMF[A-Z])", model_text)
        if fenghua_match:
            series_code = clean_text(fenghua_match.group(1)).upper()
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": f"{brand_text} {series_code} NTC thermistor series",
                    "器件类型": "热敏电阻",
                },
                model=model_text,
            )
    if clean_text(result.get("系列", "")) != "" and clean_text(result.get("系列说明", "")) == "":
        result["系列说明"] = build_series_description_fallback("热敏电阻", brand=brand_text, series=result["系列"], special_use=result.get("特殊用途", ""))
    return result


def resolve_varistor_series_profile(brand="", model="", component_type="", series="", series_desc="", special_use=""):
    brand_text = clean_brand(brand)
    model_text = clean_model(model)
    normalized_type = normalize_component_type(component_type) or "压敏电阻"
    result = {
        "系列": clean_text(series),
        "系列说明": clean_text(series_desc),
        "特殊用途": normalize_special_use(special_use),
        "器件类型": normalized_type,
    }
    if model_text == "":
        return result
    for pattern, code_builder, desc_suffix in VARISTOR_SERIES_RULES:
        match = pattern.match(model_text)
        if not match:
            continue
        return merge_series_profile_values(
            result,
            {
                "系列": clean_text(code_builder(match)),
                "系列说明": f"{brand_text} {clean_text(code_builder(match))} {desc_suffix}".strip(),
                "器件类型": normalized_type,
            },
            model=model_text,
        )
    if "LITTELFUSE" in brand_text.upper():
        littelfuse_match = re.match(r"^(V(?:\d{2}P|8Z[A-Z]))", model_text)
        if littelfuse_match:
            series_code = clean_text(littelfuse_match.group(1)).upper()
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": f"{brand_text} {series_code} multilayer varistor series",
                    "器件类型": normalized_type,
                },
                model=model_text,
            )
    if "BOURNS" in brand_text.upper():
        bourns_match = re.match(r"^(CG\d{4}ML[A-Z])", model_text)
        if bourns_match:
            series_code = clean_text(bourns_match.group(1)).upper()
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": f"{brand_text} {series_code} chip varistor series",
                    "器件类型": normalized_type,
                },
                model=model_text,
            )
    if clean_text(result.get("系列", "")) != "" and clean_text(result.get("系列说明", "")) == "":
        result["系列说明"] = build_series_description_fallback(normalized_type, brand=brand_text, series=result["系列"], special_use=result.get("特殊用途", ""))
    return result


def resolve_timing_series_profile(brand="", model="", component_type="", series="", series_desc="", special_use=""):
    brand_text = clean_brand(brand)
    model_text = clean_model(model)
    normalized_type = normalize_component_type(component_type)
    result = {
        "系列": clean_text(series),
        "系列说明": clean_text(series_desc),
        "特殊用途": normalize_special_use(special_use),
        "器件类型": normalized_type,
    }
    if model_text == "":
        return result
    series_code = clean_text(series)
    if series_code == "":
        match = re.match(r"^([A-Z]{2}\d{3,6})", model_text)
        if match:
            series_code = clean_text(match.group(1)).upper()
    if series_code == "":
        return result
    series_desc_text = clean_text(series_desc)
    if series_desc_text == "":
        prefix_match = re.match(r"^[A-Z]{2}", series_code)
        prefix_key = clean_text(prefix_match.group(0)).upper() if prefix_match else ""
        family_desc = TIMING_SERIES_PREFIX_MEANING.get(prefix_key, "")
        if family_desc != "":
            series_desc_text = f"{brand_text} {series_code} {family_desc}".strip()
        elif normalized_type in TIMING_COMPONENT_TYPES:
            series_desc_text = build_series_description_fallback(normalized_type, brand=brand_text, series=series_code, special_use=result.get("特殊用途", ""))
    return merge_series_profile_values(
        result,
        {
            "系列": series_code,
            "系列说明": series_desc_text,
            "器件类型": normalized_type,
        },
        model=model_text,
    )


def resolve_generic_mlcc_brand_series_profile(brand="", model="", series="", series_desc="", special_use=""):
    brand_text = clean_brand(brand)
    brand_upper = brand_text.upper()
    model_text = clean_model(model)
    result = {
        "系列": clean_text(series),
        "系列说明": clean_text(series_desc),
        "特殊用途": normalize_special_use(special_use),
        "器件类型": "MLCC",
        "_mlcc_series_class": "",
    }
    if model_text == "":
        return result
    if "HRE" in brand_upper or "芯声微" in brand_text:
        parsed = parse_generic_size_first_mlcc(model_text, brand=brand_text)
        if parsed is not None:
            series_code = clean_text(parsed.get("系列", ""))
            if series_code != "":
                return merge_series_profile_values(result, parsed, model=model_text)
    if "TAIYO" in brand_upper or "太诱" in brand_text or "太阳诱电" in brand_text:
        series_code = taiyo_mlcc_series_code_from_model(model_text)
        if series_code != "":
            profile = taiyo_mlcc_series_profile(series_code)
            if clean_text(profile.get("系列说明", "")) == "":
                profile["系列说明"] = f"Taiyo Yuden {series_code} MLCC official series"
            return merge_series_profile_values(result, profile, model=model_text)
    if "WALSIN" in brand_upper or "华新科" in brand_text:
        series_code = walsin_mlcc_series_code_from_model(model_text) or extract_size_prefixed_mlcc_series_code(model_text)
        if series_code != "":
            profile = walsin_mlcc_series_profile(series_code)
            if clean_text(profile.get("系列说明", "")) == "":
                dielectric = clean_text(WALSIN_NUMERIC_MLCC_DIELECTRIC.get(series_code[-1], ""))
                if dielectric != "":
                    profile["系列说明"] = f"{brand_text} {series_code} {dielectric} MLCC系列"
                else:
                    profile["系列说明"] = f"{brand_text} {series_code} MLCC系列"
            return merge_series_profile_values(result, profile, model=model_text)
    if "KYOCERA" in brand_upper or "AVX" in brand_upper or "晶瓷" in brand_text:
        match = re.match(r"^([A-Z]{2,4})(?=\d)", model_text)
        series_code = clean_text(match.group(1)).upper() if match else extract_size_prefixed_mlcc_series_code(model_text)
        if series_code == "":
            fallback_patterns = (
                r"^(M39014)",
                r"^(\d{3}[A-Z])",
                r"^([A-Z]{5,8})(?=[0-9/])",
            )
            for pattern in fallback_patterns:
                fallback_match = re.match(pattern, model_text)
                if fallback_match:
                    series_code = clean_text(fallback_match.group(1)).upper()
                    break
        if series_code:
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": f"Kyocera AVX {series_code} MLCC official series",
                    "器件类型": "MLCC",
                },
                model=model_text,
            )
    if "MURATA" in brand_upper or "村田" in brand_text:
        series_code = murata_series_code_from_model(model_text)
        if series_code == "":
            murata_generic_match = re.match(r"^(DEA|DE[12]|DEBB|DEBE|DEBF|DECB|DECE|DEC|DEF|DESD|KCA|KC|LLD|WBM|RHDL|EVA|GCB|GJ)", model_text)
            series_code = clean_text(murata_generic_match.group(1)).upper() if murata_generic_match else extract_leading_alpha_series_code(model_text, min_len=2, max_len=4)
        if series_code != "":
            series_desc_text = murata_series_meaning(series_code)
            if series_desc_text in {"Murata 官方系列代码", ""}:
                series_desc_text = f"Murata {series_code} capacitor official series"
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": series_desc_text,
                    "器件类型": "MLCC",
                },
                model=model_text,
            )
    if "HRE" in brand_upper or "芯声微" in brand_text:
        parsed = parse_generic_size_first_mlcc(model_text, brand=brand_text)
        if parsed is not None:
            series_code = clean_text(parsed.get("系列", ""))
            if series_code != "":
                return merge_series_profile_values(
                    result,
                    {
                        "系列": series_code,
                        "系列说明": clean_text(parsed.get("系列说明", "")) or f"{brand_text} {series_code} MLCC系列",
                        "特殊用途": clean_text(parsed.get("特殊用途", "")),
                        "器件类型": "MLCC",
                    },
                    model=model_text,
                )
    if "VIIYONG" in brand_upper or "微容" in brand_text:
        match = re.match(r"^V[0-9R]+[A-Z](?P<size>\d{4})(?P<family>C0G|COG|HQC|NP0|NPO|X5R|X6S|X7R|X7S|X8R|Y5V)", model_text)
        if match:
            series_code = f"{clean_text(match.group('size'))}{clean_text(match.group('family')).upper()}"
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": f"{brand_text} {series_code} MLCC系列",
                    "器件类型": "MLCC",
                },
                model=model_text,
            )
    if "FOJAN" in brand_upper or "富捷" in brand_text:
        series_code = extract_leading_alpha_series_code(model_text, min_len=2, max_len=6)
        if series_code != "":
            return merge_series_profile_values(
                result,
                {
                    "系列": series_code,
                    "系列说明": f"{brand_text} {series_code} MLCC系列",
                    "器件类型": "MLCC",
                },
                model=model_text,
            )
    return result


def resolve_official_series_profile(brand="", model="", component_type="", series="", series_desc="", special_use=""):
    brand_text = clean_brand(brand)
    model_text = clean_model(model)
    normalized_type = normalize_component_type(component_type)
    result = {
        "系列": clean_text(series),
        "系列说明": clean_text(series_desc),
        "特殊用途": normalize_special_use(special_use),
        "器件类型": normalized_type,
    }
    parsed = None
    if model_text != "":
        try:
            parsed = parse_model_rule(model_text, brand=brand_text, component_type=normalized_type)
        except Exception:
            parsed = None
    if isinstance(parsed, dict):
        result = merge_series_profile_values(
            result,
            {
                "系列": clean_text(parsed.get("系列", "")),
                "系列说明": clean_text(parsed.get("系列说明", "")),
                "特殊用途": clean_text(parsed.get("特殊用途", "")),
                "器件类型": normalize_component_type(parsed.get("器件类型", "")),
            },
            model=model_text,
        )
    resolved_type = normalize_component_type(result.get("器件类型", "")) or normalized_type
    if resolved_type == "MLCC":
        result = merge_series_profile_values(
            result,
            resolve_mlcc_series_profile(
                brand=brand_text,
                model=model_text,
                series=result.get("系列", ""),
                series_desc=result.get("系列说明", ""),
                special_use=result.get("特殊用途", ""),
            ),
            model=model_text,
        )
        result = merge_series_profile_values(
            result,
            resolve_generic_mlcc_brand_series_profile(
                brand=brand_text,
                model=model_text,
                series=result.get("系列", ""),
                series_desc=result.get("系列说明", ""),
                special_use=result.get("特殊用途", ""),
            ),
            model=model_text,
        )
    elif resolved_type in RESISTOR_COMPONENT_TYPES:
        result = merge_series_profile_values(
            result,
            infer_resistor_series_profile(
                model_text,
                brand=brand_text,
                component_type=resolved_type,
                special_use=result.get("特殊用途", ""),
            ),
            model=model_text,
        )
    elif resolved_type == "热敏电阻":
        result = merge_series_profile_values(
            result,
            resolve_thermistor_series_profile(
                brand=brand_text,
                model=model_text,
                series=result.get("系列", ""),
                series_desc=result.get("系列说明", ""),
                special_use=result.get("特殊用途", ""),
            ),
            model=model_text,
        )
    elif resolved_type == "铝电解电容":
        result = merge_series_profile_values(
            result,
            resolve_aluminum_electrolytic_series_profile(
                brand=brand_text,
                model=model_text,
                series=result.get("系列", ""),
                series_desc=result.get("系列说明", ""),
                special_use=result.get("特殊用途", ""),
            ),
            model=model_text,
        )
    elif resolved_type in VARISTOR_COMPONENT_TYPES:
        result = merge_series_profile_values(
            result,
            resolve_varistor_series_profile(
                brand=brand_text,
                model=model_text,
                component_type=resolved_type,
                series=result.get("系列", ""),
                series_desc=result.get("系列说明", ""),
                special_use=result.get("特殊用途", ""),
            ),
            model=model_text,
        )
    elif resolved_type in TIMING_COMPONENT_TYPES:
        result = merge_series_profile_values(
            result,
            resolve_timing_series_profile(
                brand=brand_text,
                model=model_text,
                component_type=resolved_type,
                series=result.get("系列", ""),
                series_desc=result.get("系列说明", ""),
                special_use=result.get("特殊用途", ""),
            ),
            model=model_text,
        )
    if clean_text(result.get("系列说明", "")) == "" and clean_text(result.get("系列", "")) != "":
        fallback_desc = build_series_description_fallback(
            normalize_component_type(result.get("器件类型", "")) or resolved_type,
            brand=brand_text,
            series=result.get("系列", ""),
            special_use=result.get("特殊用途", ""),
        )
        if fallback_desc != "":
            result["系列说明"] = fallback_desc
    return result


def normalize_mounting_style(value, package_code=""):
    text = " ".join([clean_text(value), clean_text(package_code)]).upper()
    if text == "":
        return ""
    if any(token in text for token in ["贴片", "SMD", "SMT", "CHIP"]):
        return "贴片"
    if any(token in text for token in ["螺栓", "SCREW"]):
        return "螺栓式"
    if any(token in text for token in ["AXIAL", "CROWN", "轴向"]):
        return "轴向"
    if any(token in text for token in ["SNAP-IN", "LUG", "RADIAL", "插件", "引线", "LEADED", "牛角"]):
        return "插件"
    return clean_text(value)


def extract_mounting_style_from_text(text):
    raw = clean_text(text)
    if raw == "":
        return ""
    upper = raw.upper()
    if any(token in upper for token in ["贴片", "SMD", "SMT", "CHIP"]):
        return "贴片"
    if any(token in upper for token in ["螺栓", "SCREW"]):
        return "螺栓式"
    if any(token in upper for token in ["AXIAL", "CROWN", "轴向"]):
        return "轴向"
    if any(token in upper for token in ["SNAP-IN", "LUG", "RADIAL", "插件", "引线", "LEADED", "牛角", "THROUGH HOLE", "THRU HOLE"]):
        return "插件"
    return ""


SPECIAL_USE_RULES = [
    ("车规", [r"AEC[- ]?Q200", r"AUTOMOTIVE", r"车规"]),
    ("抗硫化", [r"抗硫", r"ANTI[- ]?SULFUR", r"SULFUR"]),
    ("消费", [r"CONSUMER", r"NOTEBOOK", r"CPU", r"消费"]),
    ("工业", [r"INDUSTRIAL", r"工业"]),
    ("耐腐蚀", [r"CORROSION", r"耐腐蚀"]),
    ("高纹波", [r"HIGH\s*RIPPLE", r"高纹波"]),
    ("混合聚合物", [r"HYBRID", r"混合"]),
    ("导电聚合物", [r"POLYMER", r"固态", r"导电高分子"]),
]


def special_use_tokens(value):
    text = clean_text(value)
    if text == "":
        return []
    upper = text.upper()
    tokens = []
    for label, patterns in SPECIAL_USE_RULES:
        for pattern in patterns:
            if re.search(pattern, upper, flags=re.I):
                tokens.append(label)
                break
    if tokens:
        return list(dict.fromkeys(tokens))
    compact = clean_text(value)
    return [compact] if compact != "" else []


def normalize_special_use(value):
    tokens = special_use_tokens(value)
    if not tokens:
        return ""
    return "/".join(tokens)


def extract_special_use_from_text(text):
    raw = clean_text(text)
    if raw == "":
        return ""
    upper = raw.upper()
    tokens = []
    for label, patterns in SPECIAL_USE_RULES:
        for pattern in patterns:
            if re.search(pattern, upper, flags=re.I):
                tokens.append(label)
                break
    if not tokens:
        return ""
    return "/".join(dict.fromkeys(tokens))


def working_temperature_covers(candidate, target):
    target_norm = normalize_working_temperature_text(target)
    if target_norm == "":
        return True
    candidate_norm = normalize_working_temperature_text(candidate)
    if candidate_norm == "":
        return False
    if candidate_norm == target_norm:
        return True
    cand_low, cand_high = working_temperature_bounds(candidate_norm)
    target_low, target_high = working_temperature_bounds(target_norm)
    if target_high is not None:
        if cand_high is None or cand_high < target_high:
            return False
    if target_low is not None:
        if cand_low is None or cand_low > target_low:
            return False
    return True


def life_hours_covers(candidate, target):
    target_num = life_hours_to_number(target)
    if target_num is None:
        return True
    candidate_num = life_hours_to_number(candidate)
    if candidate_num is None:
        return False
    return candidate_num >= target_num


def special_use_matches(candidate, target):
    target_tokens = set(special_use_tokens(target))
    if not target_tokens:
        return True
    candidate_tokens = set(special_use_tokens(candidate))
    if not candidate_tokens:
        return False
    return bool(candidate_tokens & target_tokens)


def build_component_spec_detail_from_row(row, component_type_hint=""):
    if row is None:
        return ""
    component_type = normalize_component_type(component_type_hint) or infer_db_component_type(row)
    row_text = " ".join([
        clean_text(row.get("器件类型", "")),
        clean_text(row.get("品牌", "")),
        clean_text(row.get("型号", "")),
        clean_text(row.get("系列", "")),
        clean_text(row.get("安装方式", "")),
        clean_text(row.get("封装代码", "")),
        clean_text(row.get("尺寸（inch）", "")),
        clean_text(row.get("尺寸（mm）", "")),
        clean_text(row.get("材质（介质）", "")),
        clean_text(row.get("规格摘要", "")),
        clean_text(row.get("容值", "")),
        clean_text(row.get("容值单位", "")),
        clean_text(row.get("工作温度", "")),
        clean_text(row.get("寿命（h）", "")),
        clean_text(row.get("特殊用途", "")),
        clean_text(row.get("备注1", "")),
        clean_text(row.get("备注2", "")),
        clean_text(row.get("备注3", "")),
    ])
    lines = build_component_detail_lines(
        component_type,
        size=row.get("尺寸（inch）", ""),
        material=row.get("材质（介质）", ""),
        value=row.get("容值", ""),
        unit=row.get("容值单位", ""),
        tol=row.get("容值误差", ""),
        volt=row.get("耐压（V）", ""),
        resistance_ohm=find_resistance_in_text(row_text) if component_type in (RESISTOR_COMPONENT_TYPES | {"热敏电阻"}) else None,
        power=find_power_in_text(row_text) if component_type in (RESISTOR_COMPONENT_TYPES | VARISTOR_COMPONENT_TYPES) else "",
        body_size=clean_text(row.get("尺寸（mm）", "")) or (extract_body_size_from_text(row_text) if component_type in {"铝电解电容", "薄膜电容", "引线型陶瓷电容"} else ""),
        pitch=row.get("脚距（mm）", "") or (extract_pitch_from_text(row_text) if component_type in ({"铝电解电容", "薄膜电容", "引线型陶瓷电容"} | VARISTOR_COMPONENT_TYPES) else ""),
        safety_class=find_safety_class(row_text) if component_type == "薄膜电容" else "",
        varistor_voltage=find_varistor_voltage_in_text(row_text) if component_type in VARISTOR_COMPONENT_TYPES else "",
        disc_size=find_disc_size_code(row_text) if component_type in VARISTOR_COMPONENT_TYPES else "",
        work_temp=row.get("工作温度", "") or extract_working_temperature_from_text(row_text),
        life_hours=row.get("寿命（h）", "") or parse_life_hours_from_text(row_text),
        mounting_style=row.get("安装方式", "") or extract_mounting_style_from_text(row_text),
        special_use=row.get("特殊用途", "") or extract_special_use_from_text(row_text),
        production_status=row.get("生产状态", ""),
        package_code=row.get("封装代码", ""),
        diameter_mm=row.get("直径（mm）", ""),
        height_mm=row.get("高度（mm）", ""),
        polarity=row.get("极性", ""),
        esr=row.get("ESR", "") or find_esr_in_text(row_text),
        ripple_current=row.get("纹波电流", "") or find_labeled_current_in_text(row_text, ["RIPPLE CURRENT", "RIPPLE"]),
        resistance_25c=row.get("阻值@25C", ""),
        resistance_unit=row.get("阻值单位", ""),
        resistance_tol=row.get("阻值误差", ""),
        b_value=row.get("B值", "") or find_b_value_in_text(row_text),
        b_value_condition=row.get("B值条件", "") or find_b_value_condition_in_text(row_text),
        common_mode_impedance=row.get("共模阻抗", ""),
        impedance_unit=row.get("阻抗单位", ""),
        rated_current=row.get("额定电流", "") or find_labeled_current_in_text(row_text, ["RATED CURRENT", "HEATING CURRENT", "IRMS", "IDC", "CURRENT"]),
        dcr=row.get("DCR", "") or find_dcr_in_text(row_text),
        circuit_count=row.get("回路数", "") or find_circuit_count_in_text(row_text),
        inductance_value=row.get("电感值", ""),
        inductance_unit=row.get("电感单位", ""),
        inductance_tol=row.get("电感误差", ""),
        saturation_current=row.get("饱和电流", "") or find_labeled_current_in_text(row_text, ["SATURATION CURRENT", "ISAT", "SAT CURRENT"]),
        shielding=row.get("屏蔽类型", "") or find_shielding_type_in_text(row_text),
        impedance_100mhz=row.get("阻抗@100MHz", "") or find_impedance_at_100mhz_in_text(row_text),
        load_capacitance=row.get("负载电容（pF）", ""),
        drive_level=row.get("驱动电平", ""),
        output_type=row.get("输出类型", ""),
        duty_cycle=row.get("占空比", ""),
    )
    return format_component_detail_inline(lines)

def match_by_partial_spec(df, spec):
    if spec is None or df.empty:
        return pd.DataFrame()

    base = scope_search_dataframe(df, spec)
    if base.empty:
        return pd.DataFrame()
    provided_size = clean_size(spec.get("尺寸（inch）", ""))
    provided_mat = clean_material(spec.get("材质（介质）", ""))
    provided_tol = clean_tol_for_match(spec.get("容值误差", ""))
    provided_volt = clean_voltage(spec.get("耐压（V）", ""))
    provided_pf = spec.get("容值_pf", None)
    target_type = infer_spec_component_type(spec)

    # MLCC 的核心参数是尺寸和容值，缺任意一个都不进入部分匹配。
    if target_type == "MLCC" and (provided_size == "" or provided_pf is None):
        return pd.DataFrame()

    provided_count = sum([
        1 if provided_size != "" else 0,
        1 if provided_mat != "" else 0,
        1 if provided_pf is not None else 0,
        1 if provided_tol != "" else 0,
        1 if provided_volt != "" else 0,
    ])
    if provided_count == 0:
        return pd.DataFrame()

    match_size = pd.Series(False, index=base.index)
    match_mat = pd.Series(False, index=base.index)
    match_pf = pd.Series(False, index=base.index)
    match_tol = pd.Series(False, index=base.index)
    match_volt = pd.Series(False, index=base.index)

    if provided_size != "":
        match_size = base["_size"].eq(provided_size)
    if provided_mat != "":
        match_mat = base["_mat"].eq(provided_mat)
    if provided_pf is not None:
        target_pf = float(provided_pf)
        match_pf = base["_pf"].notna() & ((base["_pf"] - target_pf).abs() < 1e-6)
    if provided_tol != "":
        match_tol = tolerance_equal_series(base, provided_tol)
    if provided_volt != "":
        match_volt = base["_volt"].eq(provided_volt)

    # 核心参数一旦输入，就必须严格一致，避免完整规格查询时混出其它尺寸/容值/材质。
    core_mask = pd.Series(True, index=base.index)
    if provided_size != "":
        core_mask &= match_size
    if provided_mat != "":
        core_mask &= match_mat
    if provided_pf is not None:
        core_mask &= match_pf
    if not core_mask.any():
        return pd.DataFrame()

    def volt_num(value):
        try:
            return float(value)
        except:
            return None

    # 次要参数走“可替代”约束：
    # 容差允许更严不允许更松，电压允许更高不允许更低。
    if provided_tol != "":
        tol_ok = tolerance_allows_series(base, provided_tol)
    else:
        tol_ok = pd.Series(True, index=base.index)

    if provided_volt != "":
        spec_volt_num = volt_num(provided_volt)
        volt_ok = base["_volt_num"].notna() & (spec_volt_num is not None) & base["_volt_num"].ge(spec_volt_num)
    else:
        volt_ok = pd.Series(True, index=base.index)

    final_mask = core_mask & tol_ok & volt_ok
    if not final_mask.any():
        return pd.DataFrame()

    work = base[final_mask].copy()
    work["_match_size"] = match_size[final_mask].astype(bool)
    work["_match_mat"] = match_mat[final_mask].astype(bool)
    work["_match_pf"] = match_pf[final_mask].astype(bool)
    work["_match_tol"] = match_tol[final_mask].astype(bool)
    work["_match_volt"] = match_volt[final_mask].astype(bool)
    work["_tol_ok"] = tol_ok[final_mask].astype(bool)
    work["_volt_ok"] = volt_ok[final_mask].astype(bool)
    work["_matched_param_count"] = (
        work["_match_size"].astype(int)
        + work["_match_mat"].astype(int)
        + work["_match_pf"].astype(int)
        + work["_match_tol"].astype(int)
        + work["_match_volt"].astype(int)
    )
    work["_provided_param_count"] = provided_count

    out = work.copy()
    out = exclude_same_brand(out, spec.get("品牌", ""))
    out = apply_match_levels_and_sort(out, spec)

    drop_cols = [
        c for c in [
            "_size", "_mat", "_tol", "_volt", "_pf",
            "_tol_kind", "_tol_num", "_volt_num", "_component_type", "_res_ohm",
            "_match_size", "_match_mat", "_match_pf", "_match_tol", "_match_volt",
            "_tol_ok", "_volt_ok",
            "_matched_param_count", "_provided_param_count",
        ] if c in out.columns
    ]
    return out.drop(columns=drop_cols)

def build_spec_info_df(spec):
    if spec is None:
        return pd.DataFrame()
    row = pd.DataFrame([build_component_display_row(spec)])
    row = select_component_display_columns(row, spec)
    return format_display_df(row)



MLCC_REFERENCE_ALIAS_MAP = {
    "信昌料号": ("信昌", "PDC"),
    "华科料号": ("华新科", "WALSIN", "华科"),
}
MLCC_REFERENCE_EXCLUDE_ALIASES = tuple({alias for aliases in MLCC_REFERENCE_ALIAS_MAP.values() for alias in aliases if clean_text(alias) != ""})
BOM_PREFERRED_BRAND_SLOTS = 3
BOM_PREFERRED_BRAND_ALIAS_GROUPS = (
    ("信昌", ("信昌", "PDC")),
    ("华新科", ("华新科", "WALSIN", "华科")),
    ("芯声微", ("芯声微", "HRE")),
    ("厚声", ("厚声", "UNI-ROYAL", "UNIROYAL")),
    ("江海", ("江海", "JIANGHAI", "NANTONG JIANGHAI")),
    ("爱普生", ("爱普生", "EPSON")),
)
BOM_PREFERRED_BRAND_EXCLUDE_ALIASES = tuple(
    {
        alias
        for _, aliases in BOM_PREFERRED_BRAND_ALIAS_GROUPS
        for alias in aliases
        if clean_text(alias) != ""
    }
)
BOM_PREFERRED_SLOT_COLUMNS = tuple(
    col
    for idx in range(1, BOM_PREFERRED_BRAND_SLOTS + 1)
    for col in (f"品牌{idx}", f"型号{idx}")
)
BOM_PREFERRED_DISPLAY_COLUMN_MAP = (
    ("推荐品牌", "推荐品牌"),
    ("推荐型号", "推荐型号"),
    ("品牌1", "推荐品牌1"),
    ("型号1", "推荐型号1"),
    ("品牌2", "推荐品牌2"),
    ("型号2", "推荐型号2"),
    ("品牌3", "推荐品牌3"),
    ("型号3", "推荐型号3"),
)


def brand_alias_matches(brand, aliases):
    brand_text = clean_brand(brand).upper()
    if brand_text == "":
        return False
    for alias in aliases:
        alias_text = clean_text(alias).upper()
        if alias_text != "" and alias_text in brand_text:
            return True
    return False


def empty_bom_preferred_brand_slots(limit=BOM_PREFERRED_BRAND_SLOTS):
    slots = {}
    for idx in range(1, limit + 1):
        slots[f"品牌{idx}"] = ""
        slots[f"型号{idx}"] = ""
    return slots


def match_bom_preferred_brand_label(brand):
    for label, aliases in BOM_PREFERRED_BRAND_ALIAS_GROUPS:
        if brand_alias_matches(brand, aliases):
            return label
    return ""


def normalize_brand_model_pair_set(pairs):
    normalized = set()
    for item in pairs or []:
        if not isinstance(item, (list, tuple)) or len(item) < 2:
            continue
        brand = clean_brand(item[0])
        model = clean_text(item[1])
        if brand != "" and model != "":
            normalized.add((brand, model))
        label = match_bom_preferred_brand_label(brand)
        if label != "" and model != "":
            normalized.add((label, model))
    return normalized


def build_bom_preferred_brand_slots(frame, limit=BOM_PREFERRED_BRAND_SLOTS, exclude_pairs=None):
    slots = empty_bom_preferred_brand_slots(limit=limit)
    if frame is None or frame.empty:
        return slots
    if "品牌" not in frame.columns or "型号" not in frame.columns:
        return slots
    excluded = normalize_brand_model_pair_set(exclude_pairs)
    items = []
    seen = set()
    for brand_value, model_value in zip(frame["品牌"].astype(str).tolist(), frame["型号"].astype(str).tolist()):
        label = match_bom_preferred_brand_label(brand_value)
        model = clean_text(model_value)
        if label == "" or model == "":
            continue
        key = (label, model)
        if key in excluded or key in seen:
            continue
        seen.add(key)
        items.append(key)
        if len(items) >= limit:
            break
    for idx, (brand, model) in enumerate(items, start=1):
        slots[f"品牌{idx}"] = brand
        slots[f"型号{idx}"] = model
    return slots


def choose_bom_display_match(frame):
    if frame is None or frame.empty:
        return None
    if "品牌" not in frame.columns:
        return None
    for _, row in frame.iterrows():
        if match_bom_preferred_brand_label(row.get("品牌", "")) != "":
            return row
    return None


def collect_brand_models_in_frame(frame, aliases):
    if frame is None or frame.empty:
        return ""
    if "品牌" not in frame.columns or "型号" not in frame.columns:
        return ""
    models = []
    seen = set()
    for brand_value, model_value in zip(frame["品牌"].astype(str).tolist(), frame["型号"].astype(str).tolist()):
        if not brand_alias_matches(brand_value, aliases):
            continue
        model = clean_text(model_value)
        if model == "" or model in seen:
            continue
        seen.add(model)
        models.append(model)
    return " | ".join(models)


def empty_mlcc_reference_result():
    return {key: "" for key in MLCC_REFERENCE_ALIAS_MAP}


def lookup_brand_models_for_spec_map(df, spec):
    refs = empty_mlcc_reference_result()
    if df is None or df.empty or spec is None:
        return refs
    if infer_spec_component_type(spec) != "MLCC":
        return refs

    try:
        cache_key = (get_query_cache_signature(), serialize_spec_for_cache(spec))
    except Exception:
        cache_key = None
    if cache_key is not None:
        cached = MLCC_REFERENCE_LOOKUP_CACHE.get(cache_key)
        if cached is not None:
            return cached.copy()

    base = scope_search_dataframe(df, spec)
    if base.empty or "品牌" not in base.columns or "型号" not in base.columns:
        return refs

    spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
    spec_volt = clean_voltage(spec.get("耐压（V）", ""))
    if spec_tol != "":
        base = base[tolerance_allows_series(base, spec_tol)]
        if base.empty:
            return refs
    if spec_volt != "":
        try:
            spec_volt_num = float(spec_volt)
            base = base[base["_volt_num"].notna() & base["_volt_num"].ge(spec_volt_num)]
        except Exception:
            base = base[base["_volt"].eq(spec_volt)]
        if base.empty:
            return refs

    base = apply_match_levels_and_sort(base, spec)
    if base.empty:
        return refs

    for field_name, aliases in MLCC_REFERENCE_ALIAS_MAP.items():
        refs[field_name] = collect_brand_models_in_frame(base, aliases)

    if cache_key is not None:
        MLCC_REFERENCE_LOOKUP_CACHE[cache_key] = refs.copy()
    return refs


def lookup_brand_models_for_spec(df, spec, aliases):
    refs = lookup_brand_models_for_spec_map(df, spec)
    for field_name, field_aliases in MLCC_REFERENCE_ALIAS_MAP.items():
        if tuple(field_aliases) == tuple(aliases):
            return refs.get(field_name, "")
    return ""


def format_other_brand_models(frame, limit=None, exclude_aliases=None, exclude_pairs=None):
    if frame is None or frame.empty:
        return ""
    if "品牌" not in frame.columns or "型号" not in frame.columns:
        return ""
    alias_filter = tuple(exclude_aliases or MLCC_REFERENCE_EXCLUDE_ALIASES)
    excluded_pairs = normalize_brand_model_pair_set(exclude_pairs)
    items = []
    seen = set()
    for brand_value, model_value in zip(frame["品牌"].astype(str).tolist(), frame["型号"].astype(str).tolist()):
        brand = clean_brand(brand_value)
        if alias_filter and brand_alias_matches(brand, alias_filter):
            continue
        model = clean_text(model_value)
        if model == "":
            continue
        key = (brand, model)
        if key in excluded_pairs:
            continue
        if key in seen:
            continue
        seen.add(key)
        items.append(f"{brand}:{model}" if brand != "" else model)
        if limit is not None and len(items) >= limit:
            break
    return " | ".join(items)


def resolve_mlcc_brand_references(df, spec, matched=None, current_model=""):
    refs = empty_mlcc_reference_result()
    if infer_spec_component_type(spec) != "MLCC":
        return refs

    source_brand = clean_brand(spec.get("品牌", "")) if spec is not None else ""
    source_model = clean_text(current_model) or (clean_text(spec.get("型号", "")) if spec is not None else "")
    fallback_refs = None

    for field_name, aliases in MLCC_REFERENCE_ALIAS_MAP.items():
        value = ""
        if matched is not None and not matched.empty:
            value = collect_brand_models_in_frame(matched, aliases)
        if value == "" and df is not None and not df.empty:
            if fallback_refs is None:
                fallback_refs = lookup_brand_models_for_spec_map(df, spec)
            value = fallback_refs.get(field_name, "")
        if value == "" and brand_alias_matches(source_brand, aliases) and source_model != "":
            value = source_model
        refs[field_name] = value

    return refs


def move_columns_after(df, anchor_col, move_cols):
    if df is None or df.empty:
        return df
    existing_moves = [col for col in move_cols if col in df.columns]
    if not existing_moves or anchor_col not in df.columns:
        return df
    columns = [col for col in df.columns if col not in existing_moves]
    insert_at = columns.index(anchor_col) + 1
    for offset, col in enumerate(existing_moves):
        columns.insert(insert_at + offset, col)
    return df[columns]


def move_reference_model_columns_after_rank(df):
    if df is None or df.empty:
        return df
    move_cols = [col for col in ["信昌料号", "华科料号"] if col in df.columns]
    if not move_cols:
        return df
    for anchor in ["首选推荐等级", "推荐等级"]:
        if anchor in df.columns:
            return move_columns_after(df, anchor, move_cols)
    return df


def build_bom_component_distribution_text(result_df):
    if result_df is None or result_df.empty or "器件类型" not in result_df.columns:
        return ""

    type_series = result_df["器件类型"].apply(clean_text)
    type_series = type_series[type_series != ""]
    if type_series.empty:
        return ""

    counts = type_series.value_counts()
    parts = [f"{name} {int(count)} 行" for name, count in counts.items()]
    return "器件类型分布：" + " | ".join(parts)


def compact_bom_detail_text(value):
    text = clean_text(value)
    if text == "":
        return ""
    lines = [clean_text(line) for line in re.split(r"[\r\n]+", text) if clean_text(line) != ""]
    if not lines:
        return text

    def compact_detail_token(token):
        token = clean_text(token)
        if token == "":
            return ""
        token = token.replace("：", ":")
        if ":" not in token:
            return token
        label, raw = token.split(":", 1)
        label = clean_text(label)
        raw = clean_text(raw)
        if raw == "":
            return ""
        raw = raw.replace(" ", "")
        if label in {"容值", "阻值", "压敏电压", "耐压", "功率"}:
            return raw
        if label in {"尺寸", "介质", "材质", "安规", "规格", "脚距"}:
            return raw
        return raw if raw != "" else token

    compacted = [compact_detail_token(line) for line in lines]
    compacted = [item for item in compacted if item != ""]
    if not compacted:
        return text
    joined = " | ".join(compacted)
    joined = re.sub(r"\s+", " ", joined).strip()
    return joined


def build_bom_parse_summary(row):
    status = clean_text(row.get("解析状态", ""))
    source = clean_text(row.get("解析来源", ""))
    failure_reason = clean_text(row.get("失败原因", ""))
    difference_note = clean_text(row.get("差异说明", ""))
    parts = []
    if status != "":
        parts.append(f"状态: {status}")
    if source != "":
        parts.append(f"来源: {source}")
    if failure_reason != "":
        parts.append(f"原因: {failure_reason}")
    if difference_note != "":
        parts.append(f"说明: {difference_note}")
    return " | ".join(parts)


def build_bom_display_df(result_df):
    if result_df is None or result_df.empty:
        return pd.DataFrame()

    display_df = result_df.copy()
    for col in BOM_PREFERRED_SLOT_COLUMNS:
        if col not in display_df.columns:
            display_df[col] = ""
    for source_col, display_col in BOM_PREFERRED_DISPLAY_COLUMN_MAP:
        if source_col in display_df.columns:
            display_df[display_col] = display_df[source_col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
        else:
            display_df[display_col] = ""
    if "前5个其他品牌型号" in display_df.columns and "其他品牌型号" not in display_df.columns:
        display_df["其他品牌型号"] = display_df["前5个其他品牌型号"]
    display_df["解析说明"] = display_df.apply(build_bom_parse_summary, axis=1)

    preferred_columns = [
        "BOM行号",
        "BOM型号",
        "BOM规格",
        "BOM品名",
        "器件类型",
        "规格参数明细",
        "BOM数量",
        "匹配参数明细",
        "首选推荐等级",
        "推荐品牌",
        "推荐型号",
        "推荐品牌1",
        "推荐型号1",
        "推荐品牌2",
        "推荐型号2",
        "推荐品牌3",
        "推荐型号3",
        "其他品牌型号",
        "状态",
        "解析说明",
    ]
    existing_columns = [col for col in preferred_columns if col in display_df.columns]
    if not existing_columns:
        return display_df
    return display_df[existing_columns]


def parse_taiyo_common(model):
    model = clean_model(model)
    # 常见太阳诱电格式示例：TMK105BJ105KV-F / JMK105BJ105KV-F
    prefixes = ["TMK", "JMK", "EMK", "LMK", "AMK"]
    prefix = next((p for p in prefixes if model.startswith(p)), None)
    if prefix is None or len(model) < 11:
        return None

    size_map = {
        "1005": "0402", "1608": "0603", "2012": "0805", "3216": "1206",
        "3225": "1210", "4520": "1808", "4532": "1812", "5750": "2220"
    }
    material_map = {
        "C": "COG(NPO)", "B": "X5R", "E": "X6S", "F": "X7R", "R": "X7R"
    }
    tol_map = {"F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}
    voltage_map = {
        "0E": "2.5", "0G": "4", "0J": "6.3", "1A": "10", "1C": "16",
        "1E": "25", "1H": "50", "2A": "100", "2D": "200", "2E": "250", "2J": "630",
        "V": "35", "KV": "10"
    }
    series_profile = taiyo_mlcc_series_profile(prefix)

    try:
        size_code = model[3:7]
        mat_code = model[7]
        cap_code = model[8:11]

        tol = ""
        volt = ""

        # 常见：...J105KV-F / ...J105KV
        rest = model[11:]
        if len(rest) >= 1 and rest[0] in tol_map:
            tol = tol_map.get(rest[0], "")
            rest2 = rest[1:]
        else:
            rest2 = rest

        if rest2.startswith("6R3"):
            volt = "6.3"
        elif rest2.startswith("KV"):
            volt = "10"
        elif len(rest2) >= 2 and rest2[:2] in voltage_map:
            volt = voltage_map.get(rest2[:2], "")
        elif len(rest2) >= 1 and rest2[:1] in voltage_map:
            volt = voltage_map.get(rest2[:1], "")

        return {
            "品牌": "太阳诱电Taiyo",
            "型号": model,
            "系列": clean_text(series_profile.get("系列", "")),
            "系列说明": clean_text(series_profile.get("系列说明", "")),
            "特殊用途": clean_text(series_profile.get("特殊用途", "")),
            "_mlcc_series_class": clean_text(series_profile.get("_mlcc_series_class", "")),
            "尺寸（inch）": size_map.get(size_code, ""),
            "材质（介质）": clean_material(material_map.get(mat_code, "")),
            "容值_pf": murata_cap_code_to_pf(cap_code),
            "容值误差": clean_tol_for_match(tol),
            "耐压（V）": clean_voltage(volt),
            "_model_rule_authority": "taiyo_old_series",
        }
    except:
        return None


def parse_taiyo_new_common(model):
    model = clean_model(model)
    prefixes = ("MAAS", "MBAS", "MCAS", "MCAST", "MLAS", "MMAS", "MSAS")
    if not model.startswith(prefixes):
        return None

    size_map = {
        "021": "0201",
        "31": "1206",
        "32": "1210",
        "63": "0603",
        "105": "0402",
        "168": "0603",
    }
    material_map = {
        "LAB": "X7R",
        "LBC": "X6S",
        "SCG": "COG(NPO)",
        "CC6": "X6S",
        "MAB": "X5R",
        "SL8": "X8L",
    }
    tol_map = {"A": "0.05PF", "B": "0.1pF", "C": "0.25pF", "D": "0.5pF", "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}
    voltage_map = {"F": "4", "P": "6.3", "T": "10", "W": "25", "N": "16", "H": "50"}
    series_code = taiyo_mlcc_series_code_from_model(model)
    series_profile = taiyo_mlcc_series_profile(series_code)

    match = re.fullmatch(
        r"(?P<prefix>M[A-Z]{4})(?P<size>\d{2,3})(?P<mat>[A-Z0-9]{3})(?P<body>.*)",
        model,
    )
    if not match:
        return None

    body = match.group("body")
    cap_match = re.fullmatch(r"(?:(?P<cond>\d)(?P<cap>\d{3})|(?P<cap2>R\d+|\d{3,4}))(?P<tol>[BCDFGJKMZ])(?P<volt>[A-Z])(?P<rest>.*)", body)
    if not cap_match:
        return None

    cap_token = cap_match.group("cap") or cap_match.group("cap2")
    if cap_token is None:
        return None
    cap_pf = murata_cap_code_to_pf(cap_token)
    if cap_pf is None:
        return None

    size_code = match.group("size")
    size = size_map.get(size_code, "")
    if size == "" and size_code.isdigit():
        if len(size_code) == 2:
            size = {"31": "1206", "32": "1210"}.get(size_code, "")

    return {
        "品牌": "太阳诱电Taiyo",
        "型号": model,
        "系列": clean_text(series_profile.get("系列", "")),
        "系列说明": clean_text(series_profile.get("系列说明", "")),
        "特殊用途": clean_text(series_profile.get("特殊用途", "")),
        "_mlcc_series_class": clean_text(series_profile.get("_mlcc_series_class", "")),
        "尺寸（inch）": size,
        "材质（介质）": clean_material(material_map.get(match.group("mat"), "")),
        "容值_pf": cap_pf,
        "容值误差": clean_tol_for_match(tol_map.get(cap_match.group("tol"), "")),
        "耐压（V）": clean_voltage(voltage_map.get(cap_match.group("volt"), "")),
        "_model_rule_authority": "taiyo_new_series",
    }


def parse_generic_size_first_mlcc(model, brand=""):
    model = clean_model(model)
    match = re.fullmatch(
        r"(?P<prefix>[A-Z]{1,3})(?P<size>\d{4})(?P<mat>C0G|COG|NP0|NPO|X5R|X7R|X7S|X7T|X6S|X8R|X8L|Y5V)(?P<cap>(?:\d{3,4}|R\d+))(?P<tol>[BCDFGJKMZ])(?P<volt>(?:6R3|0J|0G|0E|1A|1B|1C|1D|1E|1H|2A|2D|2E|2J|250|500|630|\d{3}))(?P<rest>.*)",
        model,
    )
    if not match:
        return None

    size_map = {
        "01005": "01005", "0201": "0201", "0401": "0401", "0402": "0402",
        "0603": "0603", "0805": "0805", "1206": "1206", "1210": "1210",
        "1808": "1808", "1812": "1812", "2010": "2010", "2220": "2220",
    }
    voltage_map = {
        "6R3": "6.3", "0J": "6.3", "0G": "4", "0E": "2.5", "1A": "10",
        "1B": "16", "1C": "25", "1D": "50", "1E": "100", "1H": "50",
        "2A": "100", "2D": "200", "2E": "250", "2J": "630",
        "250": "25", "500": "50", "630": "63",
    }
    material_map = {
        "C0G": "COG(NPO)", "COG": "COG(NPO)", "NP0": "COG(NPO)", "NPO": "COG(NPO)",
        "X5R": "X5R", "X7R": "X7R", "X7S": "X7S", "X7T": "X7T", "X6S": "X6S",
        "X8R": "X8R", "X8L": "X8L", "Y5V": "Y5V"
    }
    tol_map = {"A": "0.05PF", "B": "0.1pF", "C": "0.25pF", "D": "0.5pF", "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}

    brand_text = clean_brand(brand)
    if "HRE" in brand_text or "芯声微" in brand_text:
        display_brand = "芯声微HRE"
        authority = "hre_generic_size_first"
    else:
        display_brand = clean_brand(brand) or "未知"
        authority = "generic_size_first_mlcc"
    series_profile = generic_mlcc_series_profile_from_model(model, brand=brand_text)
    series_code = clean_text(series_profile.get("系列", "")) or clean_text(match.group("prefix"))

    return {
        "品牌": display_brand,
        "型号": model,
        "系列": series_code,
        "系列说明": clean_text(series_profile.get("系列说明", "")),
        "特殊用途": clean_text(series_profile.get("特殊用途", "")),
        "_mlcc_series_class": clean_text(series_profile.get("_mlcc_series_class", "")),
        "尺寸（inch）": size_map.get(match.group("size"), clean_size(match.group("size"))),
        "材质（介质）": clean_material(material_map.get(match.group("mat"), match.group("mat"))),
        "容值_pf": murata_cap_code_to_pf(match.group("cap")),
        "容值误差": clean_tol_for_match(tol_map.get(match.group("tol"), "")),
        "耐压（V）": clean_voltage(voltage_map.get(match.group("volt"), "")),
        "_model_rule_authority": authority,
    }


def parse_kyocera_avx_common(model):
    model = clean_model(model)
    match = re.fullmatch(
        r"(?P<size>\d{4})(?P<volt>[0-9AZYD])(?P<mat>[A-Z])(?P<cap>(?:\d{3,4}|R\d+))(?P<tol>[BCDFGJKMZ])(?P<rest>.*)",
        model,
    )
    if not match:
        return None

    size_map = {
        "01005": "01005", "0201": "0201", "0402": "0402", "0603": "0603",
        "0805": "0805", "1206": "1206", "1210": "1210", "1808": "1808",
        "1812": "1812", "2220": "2220",
    }
    voltage_map = {
        "4": "4", "6": "6.3", "Z": "10", "Y": "16", "1": "25", "3": "25", "D": "35", "5": "50", "8": "100"
    }
    material_map = {
        "C": "X7R", "D": "X5R", "Z": "COG(NPO)", "A": "X8R", "B": "X6S"
    }
    tol_map = {"A": "0.05PF", "B": "0.1pF", "C": "0.25pF", "D": "0.5pF", "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}

    return {
        "品牌": "晶瓷Kyocera AVX",
        "型号": model,
        "系列": extract_size_prefixed_mlcc_series_code(model),
        "系列说明": f"Kyocera AVX {extract_size_prefixed_mlcc_series_code(model)} MLCC official series" if extract_size_prefixed_mlcc_series_code(model) else "",
        "尺寸（inch）": size_map.get(match.group("size"), clean_size(match.group("size"))),
        "材质（介质）": clean_material(material_map.get(match.group("mat"), "")),
        "容值_pf": murata_cap_code_to_pf(match.group("cap")),
        "容值误差": clean_tol_for_match(tol_map.get(match.group("tol"), "")),
        "耐压（V）": clean_voltage(voltage_map.get(match.group("volt"), "")),
        "_model_rule_authority": "kyocera_avx_common",
    }


def parse_yageo_common(model):
    model = clean_model(model)
    # 常见国巨：CC0402KRX5R5BB105 / CC0805KKX7R9BB104
    prefix = yageo_mlcc_series_code_from_model(model)
    if prefix == "" or len(model) < 16:
        return None

    size_map = {
        "0100": "01005", "0201": "0201", "0402": "0402", "0603": "0603",
        "0805": "0805", "1206": "1206", "1210": "1210", "1808": "1808",
        "1812": "1812", "2220": "2220"
    }
    material_map = {
        "X5R": "X5R", "X7R": "X7R", "C0G": "COG(NPO)", "NP0": "COG(NPO)", "NPO": "COG(NPO)", "Y5V": "Y5V"
    }
    tol_map = {"A": "0.05PF", "B": "0.1pF", "C": "0.25pF", "D": "0.5pF", "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}
    voltage_map = {
        "4": "4", "5": "6.3", "6": "10", "7": "16", "8": "25", "9": "50",
        "A": "100", "B": "200", "C": "250", "D": "500", "E": "630"
    }
    series_profile = yageo_mlcc_series_profile(prefix)

    try:
        size_code = model[2:6]
        tol_code = model[6]
        mat_code = model[8:11]
        volt_code = model[11]
        cap_code = model[14:17]

        return {
            "品牌": "国巨YAGEO",
            "型号": model,
            "系列": clean_text(series_profile.get("系列", "")),
            "系列说明": clean_text(series_profile.get("系列说明", "")),
            "特殊用途": clean_text(series_profile.get("特殊用途", "")),
            "_mlcc_series_class": clean_text(series_profile.get("_mlcc_series_class", "")),
            "尺寸（inch）": size_map.get(size_code, ""),
            "材质（介质）": clean_material(material_map.get(mat_code, "")),
            "容值_pf": murata_cap_code_to_pf(cap_code),
            "容值误差": clean_tol_for_match(tol_map.get(tol_code, "")),
            "耐压（V）": clean_voltage(voltage_map.get(volt_code, "")),
            "_model_rule_authority": "yageo_cc_cq",
        }
    except:
        return None


def parse_cctc_common(model):
    model = clean_model(model)
    # 常见三环：TCC0402X5R105K6R3AT
    if not model.startswith("TCC") or len(model) < 14:
        return None

    size_map = {
        "0100": "01005", "0201": "0201", "0402": "0402", "0603": "0603",
        "0805": "0805", "1206": "1206", "1210": "1210", "1808": "1808",
        "1812": "1812", "2220": "2220"
    }
    material_map = {"X5R": "X5R", "X7R": "X7R", "C0G": "COG(NPO)", "NP0": "COG(NPO)", "X7T": "X7T"}
    tol_map = {"F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}

    try:
        size_code = model[3:7]
        mat_code = model[7:10]
        cap_code = model[10:13]
        tol_code = model[13]
        rest = model[14:]

        volt = ""
        if rest.startswith("6R3"):
            volt = "6.3"
        elif len(rest) >= 3 and rest[:3].isdigit():
            vmap = {"100": "10", "160": "16", "250": "25", "500": "50", "630": "630"}
            volt = vmap.get(rest[:3], rest[:3])

        return {
            "品牌": "三环CCTC",
            "型号": model,
            "尺寸（inch）": size_map.get(size_code, ""),
            "材质（介质）": clean_material(material_map.get(mat_code, "")),
            "容值_pf": eia_code_to_pf(cap_code),
            "容值误差": clean_tol_for_match(tol_map.get(tol_code, "")),
            "耐压（V）": clean_voltage(volt)
        }
    except:
        return None


def parse_samsung_cl_partial(model):
    model = clean_model(model)
    if not model.startswith("CL") or len(model) < 5:
        return None
    size_map = {"02":"01005","03":"0201","05":"0402","10":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"A":"X5R","B":"X7R","C":"COG(NPO)","U":"X7S","Y":"X7T","Z":"X7R"}
    tol_map = {"F":"1","G":"2","J":"5","K":"10","M":"20","Z":"+80/-20"}
    voltage_map = {"R":"4","Q":"6.3","P":"10","O":"16","A":"25","L":"35","B":"50","C":"100","D":"200","E":"250","F":"500","G":"630","H":"1000"}
    size = size_map.get(model[2:4], "") if len(model) >= 4 else ""
    mat = material_map.get(model[4], "") if len(model) >= 5 else ""
    pf = eia_code_to_pf(model[5:8]) if len(model) >= 8 else None
    tol = tol_map.get(model[8], "") if len(model) >= 9 else ""
    volt = voltage_map.get(model[9], "") if len(model) >= 10 else ""
    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if pf is not None else 0, 1 if tol else 0, 1 if volt else 0])
    if param_count < 3:
        return None
    return {"品牌": "Samsung","型号": model,"尺寸（inch）": clean_size(size),"材质（介质）": clean_material(mat),"容值_pf": pf,"容值误差": clean_tol_for_match(tol),"耐压（V）": clean_voltage(volt),"_param_count": param_count,"_partial_part": True}

def parse_pdc_fp_partial(model):
    model = clean_model(model)
    if not model.startswith("FP") or len(model) < 5:
        return None
    size_map = {"31":"1206","32":"1210","42":"1808","43":"1812","46":"1825","55":"2220"}
    material_map = {"X":"X7R","T":"X7T","N":"COG(NPO)"}
    tol_map = {"J":"5","K":"10","M":"20"}
    voltage_map = {"100":"10","101":"100","102":"1000","200":"20","201":"200","202":"2000","250":"25","251":"250","252":"2500","300":"30","301":"300","302":"3000","450":"45","451":"450","500":"50","501":"500","630":"63","631":"630"}
    size = size_map.get(model[2:4], "") if len(model) >= 4 else ""
    mat = material_map.get(model[4], "") if len(model) >= 5 else ""
    pf = eia_code_to_pf(model[5:8]) if len(model) >= 8 else None
    tol = tol_map.get(model[8], "") if len(model) >= 9 else ""
    volt = voltage_map.get(model[9:12], "") if len(model) >= 12 else ""
    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if pf is not None else 0, 1 if tol else 0, 1 if volt else 0])
    if param_count < 3:
        return None
    return {"品牌": "信昌PDC","型号": model,"尺寸（inch）": clean_size(size),"材质（介质）": clean_material(mat),"容值_pf": pf,"容值误差": clean_tol_for_match(tol),"耐压（V）": clean_voltage(volt),"_param_count": param_count,"_partial_part": True}

def parse_pdc_fs_partial(model):
    model = clean_model(model)
    if not model.startswith("FS") or len(model) < 5:
        return None
    size_map = {"15":"0402","18":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"B":"X5R","X":"X7R","T":"X7T","N":"COG(NPO)"}
    tol_map = {"J":"5","K":"10","M":"20"}
    size = size_map.get(model[2:4], "") if len(model) >= 4 else ""
    mat = material_map.get(model[4], "") if len(model) >= 5 else ""
    pf = eia_code_to_pf(model[5:8]) if len(model) >= 8 else None
    tol = tol_map.get(model[8], "") if len(model) >= 9 else ""
    volt = parse_pdc_fs_voltage(model)
    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if pf is not None else 0, 1 if tol else 0, 1 if volt else 0])
    if param_count < 3:
        return None
    return {"品牌": "信昌PDC","型号": model,"尺寸（inch）": clean_size(size),"材质（介质）": clean_material(mat),"容值_pf": pf,"容值误差": clean_tol_for_match(tol),"耐压（V）": clean_voltage(volt),"_param_count": param_count,"_partial_part": True}

def parse_pdc_fn_partial(model):
    model = clean_model(model)
    if not model.startswith("FN") or len(model) < 5:
        return None
    size_map = {"15":"0402","18":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"B":"X5R","X":"X7R","T":"X7T","N":"COG(NPO)"}
    tol_map = {"J":"5","K":"10","M":"20"}
    size = size_map.get(model[2:4], "") if len(model) >= 4 else ""
    mat = material_map.get(model[4], "") if len(model) >= 5 else ""
    pf = eia_code_to_pf(model[5:8]) if len(model) >= 8 else None
    tol = tol_map.get(model[8], "") if len(model) >= 9 else ""
    volt = parse_pdc_fs_voltage(model)
    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if pf is not None else 0, 1 if tol else 0, 1 if volt else 0])
    if param_count < 3:
        return None
    return {"品牌": "信昌PDC","型号": model,"尺寸（inch）": clean_size(size),"材质（介质）": clean_material(mat),"容值_pf": pf,"容值误差": clean_tol_for_match(tol),"耐压（V）": clean_voltage(volt),"_param_count": param_count,"_partial_part": True}

def parse_pdc_fm_partial(model):
    model = clean_model(model)
    if not model.startswith("FM") or len(model) < 5:
        return None
    size_map = {"15":"0402","18":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"X":"X7R","N":"COG(NPO)","Y":"Y5V"}
    tol_map = {"J":"5","K":"10","M":"20"}
    size = size_map.get(model[2:4], "") if len(model) >= 4 else ""
    mat = material_map.get(model[4], "") if len(model) >= 5 else ""
    pf = eia_code_to_pf(model[5:8]) if len(model) >= 8 else None
    tol = tol_map.get(model[8], "") if len(model) >= 9 else ""
    volt = parse_pdc_fs_voltage(model)
    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if pf is not None else 0, 1 if tol else 0, 1 if volt else 0])
    if param_count < 3:
        return None
    return {"品牌": "信昌PDC","型号": model,"尺寸（inch）": clean_size(size),"材质（介质）": clean_material(mat),"容值_pf": pf,"容值误差": clean_tol_for_match(tol),"耐压（V）": clean_voltage(volt),"_param_count": param_count,"_partial_part": True}

def parse_pdc_fv_partial(model):
    model = clean_model(model)
    if not model.startswith("FV") or len(model) < 5:
        return None
    size_map = {"15":"0402","18":"0603","21":"0805","31":"1206","32":"1210","42":"1808","43":"1812","55":"2220"}
    material_map = {"X":"X7R","N":"COG(NPO)"}
    tol_map = {"J":"5","K":"10","M":"20"}
    size = size_map.get(model[2:4], "") if len(model) >= 4 else ""
    mat = material_map.get(model[4], "") if len(model) >= 5 else ""
    pf = eia_code_to_pf(model[5:8]) if len(model) >= 8 else None
    tol = tol_map.get(model[8], "") if len(model) >= 9 else ""
    volt = parse_pdc_fs_voltage(model)
    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if pf is not None else 0, 1 if tol else 0, 1 if volt else 0])
    if param_count < 3:
        return None
    return {"品牌": "信昌PDC","型号": model,"尺寸（inch）": clean_size(size),"材质（介质）": clean_material(mat),"容值_pf": pf,"容值误差": clean_tol_for_match(tol),"耐压（V）": clean_voltage(volt),"_param_count": param_count,"_partial_part": True}


def parse_pdc_fk_partial(model):
    model = clean_model(model)
    if not model.startswith("FK") or len(model) < 5:
        return None
    size_map = {"08": "1808", "12": "1812", "20": "2220", "21": "2211"}
    material_map = {"N": "COG(NPO)", "X": "X7R"}
    tol_map = {"J": "5", "K": "10"}
    voltage_map = {"502": "250"}
    size = size_map.get(model[2:4], "") if len(model) >= 4 else ""
    mat = material_map.get(model[4], "") if len(model) >= 5 else ""
    pf = eia_code_to_pf(model[5:8]) if len(model) >= 8 else None
    tol = tol_map.get(model[8], "") if len(model) >= 9 else ""
    volt = voltage_map.get(model[9:12], "") if len(model) >= 12 else ""
    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if pf is not None else 0, 1 if tol else 0, 1 if volt else 0])
    if param_count < 3:
        return None
    return {"品牌": "信昌PDC","型号": model,"尺寸（inch）": clean_size(size),"材质（介质）": clean_material(mat),"容值_pf": pf,"容值误差": clean_tol_for_match(tol),"耐压（V）": clean_voltage(volt),"_param_count": param_count,"_partial_part": True}


def parse_pdc_fh_partial(model):
    model = clean_model(model)
    if not model.startswith("FH") or len(model) < 5:
        return None
    size_map = {"08": "1808", "12": "1812", "20": "2220"}
    material_map = {"N": "COG(NPO)", "X": "X7R"}
    tol_map = {"J": "5", "K": "10"}
    voltage_map = {"302": "250"}
    size = size_map.get(model[2:4], "") if len(model) >= 4 else ""
    mat = material_map.get(model[4], "") if len(model) >= 5 else ""
    pf = eia_code_to_pf(model[5:8]) if len(model) >= 8 else None
    tol = tol_map.get(model[8], "") if len(model) >= 9 else ""
    volt = voltage_map.get(model[9:12], "") if len(model) >= 12 else ""
    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if pf is not None else 0, 1 if tol else 0, 1 if volt else 0])
    if param_count < 3:
        return None
    return {"品牌": "信昌PDC","型号": model,"尺寸（inch）": clean_size(size),"材质（介质）": clean_material(mat),"容值_pf": pf,"容值误差": clean_tol_for_match(tol),"耐压（V）": clean_voltage(volt),"_param_count": param_count,"_partial_part": True}


def parse_pdc_mt_partial(model):
    return parse_pdc_mt_core(model, allow_partial=True)


def parse_murata_partial(model):
    return parse_murata_core(model, allow_partial=True)

def parse_tdk_partial(model):
    model = clean_model(model)
    if not model.startswith("C"):
        return None

    size_map = {
        "1005": "0402", "1608": "0603", "2012": "0805", "3216": "1206",
        "3225": "1210", "4520": "1808", "4532": "1812", "5750": "2220"
    }
    material_map = {"C0G": "COG(NPO)", "X5R": "X5R", "X7R": "X7R", "X7S": "X7S", "X6S": "X6S"}
    voltage_map = {
        "0E": "2.5", "0G": "4", "0J": "6.3", "1A": "10", "1C": "16",
        "1E": "25", "1H": "50", "2A": "100", "2D": "200", "2E": "250", "2J": "630"
    }
    tol_map = {"C": "0.25pF", "D": "0.5pF", "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}

    size = size_map.get(model[1:5], "") if len(model) >= 5 else ""
    mat = material_map.get(model[5:8], "") if len(model) >= 8 else ""
    volt = voltage_map.get(model[8:10], "") if len(model) >= 10 else ""
    pf = eia_code_to_pf(model[10:13]) if len(model) >= 13 else None
    tol = tol_map.get(model[13], "") if len(model) >= 14 else ""

    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if volt else 0, 1 if pf is not None else 0, 1 if tol else 0])
    if param_count < 3:
        return None

    return {
        "品牌": "TDK",
        "型号": model,
        "尺寸（inch）": clean_size(size),
        "材质（介质）": clean_material(mat),
        "容值_pf": pf,
        "容值误差": clean_tol_for_match(tol),
        "耐压（V）": clean_voltage(volt),
        "_param_count": param_count,
        "_partial_part": True
    }

def parse_yageo_partial(model):
    model = clean_model(model)
    if not model.startswith("CC"):
        return None

    size_map = {
        "0100": "01005", "0201": "0201", "0402": "0402", "0603": "0603",
        "0805": "0805", "1206": "1206", "1210": "1210", "1808": "1808",
        "1812": "1812", "2220": "2220"
    }
    material_map = {"X5R": "X5R", "X7R": "X7R", "C0G": "COG(NPO)", "NP0": "COG(NPO)", "Y5V": "Y5V"}
    tol_map = {"B": "0.1pF", "C": "0.25pF", "D": "0.5pF", "F": "1", "G": "2", "J": "5", "K": "10", "M": "20", "Z": "+80/-20"}
    voltage_map = {"4": "4", "5": "6.3", "6": "10", "7": "16", "8": "25", "9": "50", "A": "100", "B": "200", "C": "250", "D": "500", "E": "630"}

    size = size_map.get(model[2:6], "") if len(model) >= 6 else ""
    tol = tol_map.get(model[6], "") if len(model) >= 7 else ""
    mat = material_map.get(model[8:11], "") if len(model) >= 11 else ""
    volt = voltage_map.get(model[11], "") if len(model) >= 12 else ""
    pf = eia_code_to_pf(model[14:17]) if len(model) >= 17 else None

    param_count = sum([1 if size else 0, 1 if mat else 0, 1 if volt else 0, 1 if pf is not None else 0, 1 if tol else 0])
    if param_count < 3:
        return None

    return {
        "品牌": "国巨Yageo",
        "型号": model,
        "尺寸（inch）": clean_size(size),
        "材质（介质）": clean_material(mat),
        "容值_pf": pf,
        "容值误差": clean_tol_for_match(tol),
        "耐压（V）": clean_voltage(volt),
        "_param_count": param_count,
        "_partial_part": True
    }

def reverse_spec_partial(model):
    m = clean_model(model)
    if m.startswith("CL"):
        return parse_samsung_cl_partial(m)
    if m.startswith("FP"):
        return parse_pdc_fp_partial(m)
    if m.startswith("FS"):
        return parse_pdc_fs_partial(m)
    if m.startswith("FN"):
        return parse_pdc_fn_partial(m)
    if m.startswith("FM"):
        return parse_pdc_fm_partial(m)
    if m.startswith("FV"):
        return parse_pdc_fv_partial(m)
    if m.startswith("FK"):
        return parse_pdc_fk_partial(m)
    if m.startswith("FH"):
        return parse_pdc_fh_partial(m)
    if m.startswith("MT"):
        return parse_pdc_mt_partial(m)
    if murata_ntc_series_code_from_model(m):
        parsed = parse_murata_ntc_core(m, allow_partial=True)
        if parsed is not None:
            return parsed
    if m.startswith(("GRM", "GCM", "GCJ", "GJM", "GQM")):
        return parse_murata_partial(m)
    if m.startswith("CC"):
        return parse_yageo_partial(m)
    if m.startswith("C"):
        return parse_tdk_partial(m)
    parsed = parse_model_rule(m, brand="", component_type="")
    if isinstance(parsed, dict) and parsed:
        return parsed
    return None

def normalize_header(col):
    text = clean_text(col).replace("（", "(").replace("）", ")")

    def compact_key(value):
        return re.sub(r"[\s\-_()/\\\[\]{}:：.。,，;；#%％&]+", "", clean_text(value)).lower()

    candidate_keys = [compact_key(text)]
    candidate_keys.extend(compact_key(x) for x in re.findall(r"\(([^()]*)\)", text))
    candidate_keys = [x for x in dict.fromkeys(candidate_keys) if x]

    alias_map = {
        "品牌": ["品牌", "厂牌", "品牌名称", "brand", "manufacturer", "mfr"],
        "型号": ["型号", "料号", "物料号", "产品型号", "规格型号", "订货号", "partnumber", "partno", "part#", "pn", "mpn", "model"],
        "系列": ["系列", "series"],
        "尺寸（inch）": ["尺寸", "封装", "尺寸inch", "case", "size", "inch", "casecode", "packagesize"],
        "材质（介质）": ["材质", "介质", "温特性", "dielectric", "material"],
        "容值": ["容值", "电容值", "capacitance", "value"],
        "容值单位": ["容值单位", "单位", "unit"],
        "容值误差": [
            "容值误差", "容值误差（%&pF）", "容值误差（%）", "容差", "误差", "误差%", "误差（%）", "误差（%&pF）",
            "容差%", "容差（%）", "tolerance", "tolerance(%)", "tol"
        ],
        "耐压（V）": ["耐压", "电压", "额定电压", "voltage", "ratedvoltage", "wv"],
        "长度（mm）": ["长度", "长度(mm)", "长", "长(mm)", "length", "bodylength"],
        "宽度（mm）": ["宽度", "宽度(mm)", "宽", "宽(mm)", "width", "bodywidth"],
        "高度（mm）": ["高度", "高度(mm)", "高", "高(mm)", "厚度", "厚度(mm)", "厚", "thickness", "height", "bodyheight", "bodythickness"],
        "特殊用途": ["特殊用途", "用途", "application", "special"],
        "备注1": ["备注1", "备注", "remark1", "note1", "生产状态", "量产状态", "status"],
        "备注2": ["备注2", "remark2", "note2", "网址", "链接", "官网链接", "产品链接", "详情链接", "url", "link", "officialurl", "producturl"],
        "备注3": ["备注3", "remark3", "note3"],
    }

    for standard_name, aliases in alias_map.items():
        alias_keys = [compact_key(alias) for alias in aliases]
        for candidate in candidate_keys:
            if any(candidate == alias or candidate.startswith(alias) for alias in alias_keys):
                return standard_name
    return None


def merge_duplicate_columns(df):
    if not getattr(df.columns, "duplicated", None) or not df.columns.duplicated().any():
        return df
    merged = pd.DataFrame(index=df.index)
    for col in dict.fromkeys(df.columns.tolist()):
        block = df.loc[:, df.columns == col]
        if isinstance(block, pd.Series):
            merged[col] = block
            continue
        if block.shape[1] == 1:
            merged[col] = block.iloc[:, 0]
            continue
        merged[col] = block.apply(
            lambda row: next((clean_text(v) for v in row if clean_text(v) != ""), ""),
            axis=1,
        )
    return merged


def parse_model_rule(model, brand="", component_type=""):
    m = clean_model(model)
    if m == "":
        return None
    brand_text = clean_brand(brand)
    brand_upper = clean_text(brand_text).upper()
    if m.startswith("CLR1"):
        parsed = parse_samsung_clr1(m)
        if parsed is not None:
            return parsed
    if "SAMSUNG" in brand_upper or "三星" in brand_text:
        if m.startswith("CL"):
            return parse_samsung_cl(m)
    if "MURATA" in brand_upper or "村田" in brand_text:
        if murata_ntc_series_code_from_model(m):
            parsed = parse_murata_ntc_common(m)
            if parsed is not None:
                return parsed
        if m.startswith(("GRM", "GCM", "GCJ", "GJM", "GQM", "GRT", "GCG", "GCQ")):
            return parse_murata_common(m)
    if murata_ntc_series_code_from_model(m):
        parsed = parse_murata_ntc_common(m)
        if parsed is not None:
            return parsed
    if "TDK" in brand_upper or "东电化" in brand_text:
        if m.startswith("CGA"):
            parsed = parse_tdk_cga_series(m)
            if parsed is not None:
                return parsed
        if m.startswith("C") and len(m) >= 14:
            return parse_tdk_c_series(m)
    if m.startswith(("MAAS", "MSAS", "MLAS", "MCAST", "MCAS")):
        parsed = parse_taiyo_new_common(m)
        if parsed is not None:
            return parsed
    if "TAIYO" in brand_upper or "太阳诱电" in brand_text:
        parsed = parse_taiyo_new_common(m)
        if parsed is not None:
            return parsed
        if m.startswith(("TMK", "JMK", "EMK", "LMK", "AMK")):
            return parse_taiyo_common(m)
    if "YAGEO" in brand_upper or "国巨" in brand_text:
        if m.startswith(("CC", "CQ")):
            return parse_yageo_common(m)
    if "KYOCERA" in brand_upper or "AVX" in brand_upper or "晶瓷" in brand_text:
        parsed = parse_kyocera_avx_common(m)
        if parsed is not None:
            return parsed
    if re.fullmatch(r"\d{4}[0-9AZYD][A-Z]\d{3,4}[BCDFGJKMZ].*", m):
        parsed = parse_kyocera_avx_common(m)
        if parsed is not None:
            return parsed
    if mlcc_generic_size_first_parser_allowed(brand_text):
        parsed_generic_size_first = parse_generic_size_first_mlcc(m, brand=brand_text)
        if parsed_generic_size_first is not None:
            return parsed_generic_size_first
    if FENGHUA_AM_MODEL_PATTERN.fullmatch(m):
        parsed = parse_fenghua_am_series(m)
        if parsed is not None:
            return parsed
    pdc_series_code = pdc_mlcc_series_code_from_model(m)
    if pdc_series_code == "MG":
        parsed = parse_pdc_mg_core(m, allow_partial=False)
        if parsed is not None:
            return parsed
    if pdc_series_code == "MS":
        parsed = parse_pdc_ms_core(m, allow_partial=False)
        if parsed is not None:
            return parsed
    if m.startswith("MT"):
        parsed = parse_pdc_mt_core(m, allow_partial=False)
        if parsed is not None:
            return parsed
    if "WALSIN" in brand_upper or "华新科" in brand_text:
        parsed = parse_pdc_mt_core(m, allow_partial=False) if m.startswith("MT") else None
        if parsed is not None:
            return parsed
        parsed = parse_walsin_common(m, brand=brand_text)
        if parsed is not None:
            return parsed
    if "JIANGHAI" in brand_upper or "江海" in brand_text or jianghai_series_code_from_model(m) != "":
        parsed = parse_jianghai_aluminum_model(m)
        if parsed is not None:
            return parsed
    if m.startswith("CL"):
        return parse_samsung_cl(m)
    if m.startswith("FP"):
        return parse_pdc_fp(m)
    if m.startswith("FS"):
        return parse_pdc_fs(m)
    if m.startswith("FN"):
        return parse_pdc_fn(m)
    if m.startswith("FM"):
        return parse_pdc_fm(m)
    if m.startswith("FV"):
        return parse_pdc_fv(m)
    if m.startswith("FK"):
        return parse_pdc_fk(m)
    if m.startswith("FH"):
        return parse_pdc_fh(m)
    if m.startswith(("GRM", "GCM", "GCJ", "GJM", "GQM", "GRT", "GCG", "GCQ")):
        return parse_murata_common(m)
    if m.startswith("TCC"):
        return parse_cctc_common(m)
    if m.startswith(("TMK", "JMK", "EMK", "LMK", "AMK")):
        return parse_taiyo_common(m)
    if m.startswith(("CC", "CQ")):
        return parse_yageo_common(m)
    if m.startswith(("CGA", "CSA", "CTA", "CBA")) and len(m) >= 7 and m[3:7].isdigit():
        parsed = parse_generic_size_first_mlcc(m, brand=brand_text)
        if parsed is not None:
            return parsed
    if m.startswith("CGA") and len(m) >= 6 and m[3].isdigit() and not m[3:7].isdigit():
        parsed = parse_tdk_cga_series(m)
        if parsed is not None:
            return parsed
    if m.startswith("C") and len(m) >= 14:
        return parse_tdk_c_series(m)
    parsed_resistor = parse_resistor_model_rule(m, brand=brand, component_type=component_type)
    if parsed_resistor is not None:
        return parsed_resistor
    if len(m) >= 11 and (m[:4].isdigit() or re.fullmatch(r"[A-Z]{2}\d{2}.*", m)):
        return parse_walsin_common(m, brand=brand_text)
    return None


def fill_missing_spec_fields_from_model(df):
    df = fill_missing_series_from_model(df)
    key_cols = ["器件类型", "尺寸（inch）", "材质（介质）", "容值误差", "耐压（V）", "容值_pf"]
    for col in ["器件类型", "尺寸（inch）", "材质（介质）", "容值", "容值单位", "容值误差", "耐压（V）", "长度（mm）", "宽度（mm）", "高度（mm）", "系列", "_model_rule_authority"]:
        if col in df.columns and isinstance(df[col].dtype, pd.CategoricalDtype):
            df[col] = df[col].astype("object")
    missing_mask = pd.Series(False, index=df.index)
    for col in key_cols:
        if col == "容值_pf":
            missing_mask |= pd.to_numeric(df[col], errors="coerce").isna()
        else:
            missing_mask |= df[col].astype(str).apply(clean_text).eq("")
    if "_model_rule_authority" in df.columns:
        missing_mask |= df["_model_rule_authority"].astype(str).apply(clean_text).eq("")
    candidate_idx = df[missing_mask].index.tolist()
    if not candidate_idx:
        return df

    for idx in candidate_idx:
        parsed = parse_model_rule(
            df.at[idx, "型号"],
            brand=df.at[idx, "品牌"] if "品牌" in df.columns else "",
            component_type=df.at[idx, "器件类型"] if "器件类型" in df.columns else "",
        )
        if clean_text(df.at[idx, "容值误差"]) == "":
            inferred_tol = infer_tolerance_from_model(df.at[idx, "型号"])
            if inferred_tol:
                df.at[idx, "容值误差"] = inferred_tol
        if not parsed:
            continue
        parsed_type = normalize_component_type(parsed.get("器件类型", ""))
        if parsed_type == "":
            parsed_type = normalize_component_type(infer_spec_component_type(parsed))
        if "器件类型" in df.columns and clean_text(df.at[idx, "器件类型"]) == "" and parsed_type != "":
            df.at[idx, "器件类型"] = parsed_type
        if "系列" in df.columns and clean_text(df.at[idx, "系列"]) == "":
            parsed_series = clean_text(parsed.get("系列", ""))
            if parsed_series != "":
                df.at[idx, "系列"] = parsed_series
        if "系列说明" in df.columns and clean_text(df.at[idx, "系列说明"]) == "":
            parsed_series_desc = clean_text(parsed.get("系列说明", ""))
            if parsed_series_desc != "":
                df.at[idx, "系列说明"] = parsed_series_desc
        if "特殊用途" in df.columns and clean_text(df.at[idx, "特殊用途"]) == "":
            parsed_special_use = clean_text(parsed.get("特殊用途", ""))
            if parsed_special_use != "":
                df.at[idx, "特殊用途"] = parsed_special_use
        if clean_text(df.at[idx, "尺寸（inch）"]) == "":
            df.at[idx, "尺寸（inch）"] = parsed.get("尺寸（inch）", "")
        if clean_text(df.at[idx, "材质（介质）"]) == "":
            df.at[idx, "材质（介质）"] = parsed.get("材质（介质）", "")
        if clean_text(df.at[idx, "容值误差"]) == "":
            df.at[idx, "容值误差"] = parsed.get("容值误差", "")
        if clean_text(df.at[idx, "耐压（V）"]) == "":
            df.at[idx, "耐压（V）"] = parsed.get("耐压（V）", "")
        for dim_col in ["长度（mm）", "宽度（mm）", "高度（mm）"]:
            if dim_col in df.columns and clean_text(df.at[idx, dim_col]) == "":
                df.at[idx, dim_col] = parsed.get(dim_col, "")
        if "_model_rule_authority" in df.columns:
            parsed_authority = clean_text(parsed.get("_model_rule_authority", ""))
            if parsed_authority != "" and clean_text(df.at[idx, "_model_rule_authority"]) == "":
                df.at[idx, "_model_rule_authority"] = parsed_authority
        current_pf = pd.to_numeric(pd.Series([df.at[idx, "容值_pf"]]), errors="coerce").iloc[0]
        parsed_pf = parsed.get("容值_pf", None)
        if pd.isna(current_pf) and parsed_pf is not None:
            value, unit = pf_to_value_unit(parsed_pf)
            if clean_text(df.at[idx, "容值"]) == "" and value != "":
                df.at[idx, "容值"] = value
            if clean_text(df.at[idx, "容值单位"]) == "" and unit != "":
                df.at[idx, "容值单位"] = unit
            df.at[idx, "容值_pf"] = parsed_pf
    return df


def should_skip_workbook(file_path):
    file_name = os.path.basename(file_path)
    lower_name = file_name.lower()
    if file_name.startswith("~$"):
        return True
    if ".backup_" in lower_name:
        return True
    if "backup" in lower_name:
        return True
    if "可查看版" in file_name or "view_only" in lower_name or "view-only" in lower_name:
        return True
    return False


def infer_default_component_type_from_source_path(source_path):
    if not source_path:
        return ""
    lower_name = os.path.basename(str(source_path)).lower()
    stem = os.path.splitext(lower_name)[0]
    candidates = [
        ("晶振", "晶振"),
        ("振荡器", "振荡器"),
        ("crystal", "晶振"),
        ("oscillator", "振荡器"),
        ("mlcc", "MLCC"),
        ("薄膜电容", "薄膜电容"),
        ("钽电容", "钽电容"),
        ("铝电解电容", "铝电解电容"),
        ("aluminum_electrolytic", "铝电解电容"),
        ("热敏电阻_ntc", "热敏电阻"),
        ("热敏电阻", "热敏电阻"),
        ("ntc", "热敏电阻"),
        ("压敏电阻_vdr", "压敏电阻"),
        ("压敏电阻", "压敏电阻"),
        ("vdr", "压敏电阻"),
        ("薄膜电阻", "薄膜电阻"),
        ("厚膜电阻", "厚膜电阻"),
        ("金属氧化膜电阻", "金属氧化膜电阻"),
        ("碳膜电阻", "碳膜电阻"),
        ("绕线电阻", "绕线电阻"),
        ("贴片电阻", "贴片电阻"),
        ("合金电阻", "合金电阻"),
        ("功率电感", "功率电感"),
        ("共模电感", "共模电感"),
        ("磁珠", "磁珠"),
    ]
    for needle, component_type in candidates:
        if needle in stem:
            return component_type
    return ""


def apply_component_type_inference_overrides_to_dataframe(df):
    if df is None or df.empty or "器件类型" not in df.columns:
        return df
    work = df.copy()
    current_types = work["器件类型"].astype("string").fillna("").apply(normalize_component_type)
    summary_text = work["规格摘要"].astype("string").fillna("").apply(clean_text) if "规格摘要" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
    note_text = work["校验备注"].astype("string").fillna("").apply(clean_text) if "校验备注" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
    candidate_mask = (
        current_types.eq("")
        | current_types.isin(CAPACITOR_COMPONENT_TYPES)
        | summary_text.str.contains("电阻", na=False)
        | note_text.str.contains("resistor", case=False, na=False)
    )
    if not candidate_mask.any():
        return work
    inferred = work.loc[candidate_mask].apply(infer_db_component_type, axis=1).astype("string").fillna("").apply(normalize_component_type)
    changed = inferred.ne("") & inferred.ne(current_types.loc[candidate_mask])
    if not changed.any():
        return work
    changed_idx = changed[changed].index
    work.loc[changed_idx, "器件类型"] = inferred.loc[changed_idx]
    if "容值_pf" in work.columns:
        non_cap_idx = inferred.loc[changed_idx][~inferred.loc[changed_idx].isin(CAPACITOR_COMPONENT_TYPES)].index
        if len(non_cap_idx) > 0:
            work.loc[non_cap_idx, "容值_pf"] = None
    return work


def normalize_capacitor_value_fields_from_pf(df):
    if df is None or df.empty or "器件类型" not in df.columns or "容值_pf" not in df.columns:
        return df
    work = df.copy()
    normalized_types = work["器件类型"].astype("string").fillna("").apply(normalize_component_type)
    pf_series = pd.to_numeric(work["容值_pf"], errors="coerce")
    current_units = work["容值单位"].astype("string").fillna("").apply(lambda value: clean_text(value).upper()) if "容值单位" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
    current_values = work["容值"].astype("string").fillna("").apply(clean_text) if "容值" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
    candidate_mask = (
        normalized_types.isin(CAPACITOR_COMPONENT_TYPES)
        & pf_series.notna()
        & (
            current_units.eq("")
            | ~current_units.isin({"PF", "NF", "UF"})
            | current_values.eq("")
        )
    )
    if not candidate_mask.any():
        return work
    for row_idx in candidate_mask[candidate_mask].index:
        pf_value = pf_series.at[row_idx]
        value, unit = pf_to_value_unit(float(pf_value))
        work.at[row_idx, "容值"] = value
        work.at[row_idx, "容值单位"] = unit
    return work


LIBRARY_COMMON_COLUMNS = [
    "器件类型", "安装方式", "封装代码", "尺寸（mm）", "规格摘要", "生产状态",
    "长度（mm）", "宽度（mm）", "高度（mm）",
    "官网链接", "数据来源", "数据状态", "校验时间", "校验备注",
    "直径（mm）", "脚距（mm）", "极性", "ESR", "纹波电流",
    "寿命（h）", "工作温度",
    "阻值@25C", "阻值单位", "阻值误差", "B值", "B值条件",
    "共模阻抗", "阻抗单位", "额定电流", "DCR", "回路数",
    "电感值", "电感单位", "电感误差", "饱和电流", "屏蔽类型", "阻抗@100MHz",
]


def normalize_library_value_unit(value):
    text = clean_text(value)
    if text == "":
        return ""
    upper = text.upper()
    lower = text.lower()
    if upper in {"PF", "NF", "UF", "MF", "F"}:
        return upper
    if upper in {"OHM", "OHMS", "Ω"}:
        return "Ω"
    if upper in {"KOHM", "KΩ"}:
        return "KΩ"
    if upper in {"MOHM", "MEGOHM", "MEGOHMS", "MΩ"} and text != "mΩ":
        return "MΩ"
    if lower in {"milliohm", "milliohms"} or text == "mΩ":
        return "mΩ"
    if text in {"mΩ", "MΩ"}:
        return text
    return text


def get_source_workbooks():
    source_files = []
    if os.path.exists(MASTER_XLSX_PATH):
        source_files.append(MASTER_XLSX_PATH)
    if os.path.exists(RESISTOR_LIBRARY_CACHE_PATH):
        source_files.append(RESISTOR_LIBRARY_CACHE_PATH)
    crystal_workbooks = [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "Crystal*", "*.xlsx"))
        if not should_skip_workbook(f)
    ]
    source_files.extend(sorted(crystal_workbooks))
    crystal_csvs = [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "Crystal*", "*.csv"))
        if not should_skip_workbook(f)
    ]
    source_files.extend(sorted(crystal_csvs))
    capacitor_workbooks = [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "Capacitor", "*.xlsx"))
        if not should_skip_workbook(f)
    ]
    source_files.extend(sorted(capacitor_workbooks))
    capacitor_csvs = [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "Capacitor", "*.csv"))
        if not should_skip_workbook(f)
    ]
    source_files.extend(sorted(capacitor_csvs))
    resistor_workbooks = [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "Resistor", "*.xlsx"))
        if not should_skip_workbook(f)
    ]
    source_files.extend(sorted(resistor_workbooks))
    resistor_csvs = [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "Resistor", "*.csv"))
        if not should_skip_workbook(f)
    ]
    source_files.extend(sorted(resistor_csvs))
    inductor_workbooks = [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "Inductor", "*.xlsx"))
        if not should_skip_workbook(f)
    ]
    source_files.extend(sorted(inductor_workbooks))
    inductor_csvs = [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "Inductor", "*.csv"))
        if not should_skip_workbook(f)
    ]
    source_files.extend(sorted(inductor_csvs))
    if source_files:
        # Preserve order while dropping duplicates from overlapping globs.
        return list(dict.fromkeys(source_files))
    return [
        f for f in glob.glob(os.path.join(DATA_FOLDER, "**", "*.xlsx"), recursive=True)
        if not should_skip_workbook(f)
    ]


def get_workbook_sheet_names(file_path, xls):
    normalized_path = os.path.normcase(file_path)
    is_inductor_workbook = os.path.normcase(os.path.join(DATA_FOLDER, "Inductor")) in normalized_path
    if is_inductor_workbook:
        data_sheets = [sheet for sheet in xls.sheet_names if sheet not in {"数据表", "字段说明"}]
        if data_sheets:
            return data_sheets
    if os.path.normcase(file_path) != os.path.normcase(MASTER_XLSX_PATH) and "数据表" in xls.sheet_names:
        return ["数据表"]
    return list(xls.sheet_names)


def normalize_imported_component_dataframe(df, source_path=""):
    if df is None or df.empty:
        return pd.DataFrame()
    work = importer_map_headers(df.copy())
    work = importer_ensure_standard_columns(work)
    for col in LIBRARY_COMMON_COLUMNS:
        if col not in work.columns:
            work[col] = ""

    if "官网链接" in work.columns and "备注2" in work.columns:
        blank_link_mask = work["备注2"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_link_mask, "备注2"] = work.loc[blank_link_mask, "官网链接"]

    for col in [
        "容值", "容值单位", "容值误差", "耐压（V）",
        "阻值", "阻值单位", "阻值误差", "阻值@25C",
        "电感值", "电感单位", "电感误差", "共模阻抗", "阻抗单位", "阻抗@100MHz",
        "输出频率", "频率", "频率单位", "频差（ppm）", "电源电压", "最高工作电压", "压敏电压",
    ]:
        if col in work.columns and not isinstance(work[col].dtype, pd.CategoricalDtype) and not pd.api.types.is_object_dtype(work[col]):
            work[col] = work[col].astype("object")

    if "器件类型" in work.columns:
        inductor_mask = work["器件类型"].astype(str).apply(normalize_component_type).isin(INDUCTOR_COMPONENT_TYPES)
        if inductor_mask.any():
            inferred_series = work.loc[inductor_mask].apply(lambda row: infer_official_inductor_fields_from_row(row), axis=1)
            inferred_df = pd.DataFrame(list(inferred_series), index=work.loc[inductor_mask].index)
            for col in inferred_df.columns:
                if col not in work.columns:
                    work[col] = ""
                elif not isinstance(work[col].dtype, pd.CategoricalDtype) and not pd.api.types.is_object_dtype(work[col]):
                    work[col] = work[col].astype("object")
                fill_values = inferred_df[col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
                if col == "系列":
                    current_series = work.loc[inductor_mask, col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
                    current_models = work.loc[inductor_mask, "型号"].astype(str).replace("nan", "").replace("None", "").apply(clean_text) if "型号" in work.columns else pd.Series("", index=current_series.index)
                    replace_mask = current_series.eq("") | current_series.eq(current_models)
                    target_idx = fill_values[replace_mask & fill_values.ne("")].index
                elif col == "系列说明":
                    current_desc = work.loc[inductor_mask, col].astype(str).replace("nan", "").replace("None", "").apply(clean_text)
                    current_models = work.loc[inductor_mask, "型号"].astype(str).replace("nan", "").replace("None", "").apply(clean_text) if "型号" in work.columns else pd.Series("", index=current_desc.index)
                    replace_mask = current_desc.eq("")
                    model_tokens = [re.escape(model) for model in current_models[current_models.ne("")].unique()]
                    if model_tokens:
                        replace_mask = replace_mask | current_desc.str.contains("|".join(model_tokens), case=False, na=False)
                    target_idx = fill_values[replace_mask & fill_values.ne("")].index
                else:
                    blank_mask = work.loc[inductor_mask, col].astype(str).replace("nan", "").replace("None", "").apply(clean_text).eq("")
                    target_idx = fill_values[blank_mask & fill_values.ne("")].index
                if len(target_idx) > 0:
                    work.loc[target_idx, col] = fill_values.loc[target_idx]
            if "容值" in work.columns and "电感值" in work.columns:
                value_unit_series = work["容值单位"].astype(str).replace("nan", "").replace("None", "").apply(normalize_library_value_unit)
                inductance_unit_mask = value_unit_series.isin({"NH", "UH", "MH"})
                blank_inductance_mask = work["电感值"].astype(str).replace("nan", "").replace("None", "").apply(clean_text).eq("")
                source_inductance_mask = (
                    inductor_mask
                    & inductance_unit_mask
                    & blank_inductance_mask
                    & work["容值"].astype(str).replace("nan", "").replace("None", "").apply(clean_text).ne("")
                )
                if source_inductance_mask.any():
                    work.loc[source_inductance_mask, "电感值"] = work.loc[source_inductance_mask, "容值"]
            if "容值单位" in work.columns and "电感单位" in work.columns:
                value_unit_series = work["容值单位"].astype(str).replace("nan", "").replace("None", "").apply(normalize_library_value_unit)
                blank_inductance_unit_mask = work["电感单位"].astype(str).replace("nan", "").replace("None", "").apply(clean_text).eq("")
                source_inductance_unit_mask = inductor_mask & value_unit_series.isin({"NH", "UH", "MH"}) & blank_inductance_unit_mask
                if source_inductance_unit_mask.any():
                    work.loc[source_inductance_unit_mask, "电感单位"] = value_unit_series.loc[source_inductance_unit_mask]

    if "阻值" in work.columns:
        blank_value_mask = work["容值"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_value_mask, "容值"] = work.loc[blank_value_mask, "阻值"]
    if "阻值单位" in work.columns:
        blank_unit_mask = work["容值单位"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_unit_mask, "容值单位"] = work.loc[blank_unit_mask, "阻值单位"]
    if "阻值误差" in work.columns:
        blank_tol_mask = work["容值误差"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_tol_mask, "容值误差"] = work.loc[blank_tol_mask, "阻值误差"]
    if "阻值@25C" in work.columns:
        blank_ntc_value_mask = work["容值"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_ntc_value_mask, "容值"] = work.loc[blank_ntc_value_mask, "阻值@25C"]
    if "电感值" in work.columns:
        blank_inductor_value_mask = work["容值"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_inductor_value_mask, "容值"] = work.loc[blank_inductor_value_mask, "电感值"]
    if "电感单位" in work.columns:
        blank_inductor_unit_mask = work["容值单位"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_inductor_unit_mask, "容值单位"] = work.loc[blank_inductor_unit_mask, "电感单位"]
    if "电感误差" in work.columns:
        blank_inductor_tol_mask = work["容值误差"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_inductor_tol_mask, "容值误差"] = work.loc[blank_inductor_tol_mask, "电感误差"]
    if "共模阻抗" in work.columns:
        blank_common_mode_value_mask = work["容值"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_common_mode_value_mask, "容值"] = work.loc[blank_common_mode_value_mask, "共模阻抗"]
    if "阻抗@100MHz" in work.columns:
        blank_bead_value_mask = work["容值"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_bead_value_mask, "容值"] = work.loc[blank_bead_value_mask, "阻抗@100MHz"]
    if "阻抗单位" in work.columns:
        blank_impedance_unit_mask = work["容值单位"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_impedance_unit_mask, "容值单位"] = work.loc[blank_impedance_unit_mask, "阻抗单位"]
    if "输出频率" in work.columns:
        blank_osc_value_mask = work["容值"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_osc_value_mask, "容值"] = work.loc[blank_osc_value_mask, "输出频率"]
    if "频率" in work.columns:
        blank_freq_value_mask = work["容值"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_freq_value_mask, "容值"] = work.loc[blank_freq_value_mask, "频率"]
    if "频率单位" in work.columns:
        blank_freq_unit_mask = work["容值单位"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_freq_unit_mask, "容值单位"] = work.loc[blank_freq_unit_mask, "频率单位"]
    if "频差（ppm）" in work.columns:
        blank_freq_tol_mask = work["容值误差"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_freq_tol_mask, "容值误差"] = work.loc[blank_freq_tol_mask, "频差（ppm）"]
    if "电源电压" in work.columns:
        blank_timing_volt_mask = work["耐压（V）"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_timing_volt_mask, "耐压（V）"] = work.loc[blank_timing_volt_mask, "电源电压"]
    if "最高工作电压" in work.columns:
        blank_resistor_volt_mask = work["耐压（V）"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_resistor_volt_mask, "耐压（V）"] = work.loc[blank_resistor_volt_mask, "最高工作电压"]
    if "压敏电压" in work.columns:
        blank_varistor_volt_mask = work["耐压（V）"].apply(lambda x: clean_text(x) == "")
        work.loc[blank_varistor_volt_mask, "耐压（V）"] = work.loc[blank_varistor_volt_mask, "压敏电压"]

    need_cols = list(dict.fromkeys(list(IMPORTER_STANDARD_COLUMNS) + LIBRARY_COMMON_COLUMNS + list(work.columns)))
    for col in need_cols:
        if col not in work.columns:
            work[col] = ""
    work = work[need_cols].copy()

    work["品牌"] = work["品牌"].apply(clean_brand)
    work["型号"] = work["型号"].apply(clean_model)
    work["系列"] = work["系列"].apply(clean_text)
    work["尺寸（inch）"] = work["尺寸（inch）"].apply(clean_size)
    work["材质（介质）"] = work["材质（介质）"].apply(clean_material)
    work["容值"] = work["容值"].apply(clean_text)
    work["容值单位"] = work["容值单位"].apply(normalize_library_value_unit)
    work["容值误差"] = work["容值误差"].apply(clean_tol_for_match)
    work["耐压（V）"] = work["耐压（V）"].apply(clean_voltage)
    work["特殊用途"] = work["特殊用途"].apply(clean_text)
    work["备注1"] = work["备注1"].apply(clean_text)
    work["备注2"] = work["备注2"].apply(clean_text)
    work["备注3"] = work["备注3"].apply(clean_text)
    for col in LIBRARY_COMMON_COLUMNS:
        work[col] = work[col].apply(clean_text)

    source_component_type = infer_default_component_type_from_source_path(source_path)
    if source_component_type and "器件类型" in work.columns:
        type_text = work["器件类型"].astype("string").fillna("").apply(clean_text)
        blank_type_mask = (
            type_text.eq("")
            | type_text.str.fullmatch(r"[?？]+", na=False)
            | type_text.str.lower().isin({"none", "nan"})
        )
        if blank_type_mask.any():
            work.loc[blank_type_mask, "器件类型"] = source_component_type

    if source_component_type and "系列" in work.columns:
        series_text = work["系列"].astype("string").fillna("").apply(clean_text)
        blank_series_mask = (
            series_text.eq("")
            | series_text.str.fullmatch(r"[?？]+", na=False)
            | series_text.isin(SERIES_PLACEHOLDER_TOKENS)
            | series_text.str.lower().isin({"none", "nan"})
        )
        if blank_series_mask.any():
            work.loc[blank_series_mask, "系列"] = source_component_type

    samsung_mask = work["品牌"].apply(
        lambda x: ("三星" in clean_brand(x)) or ("SAMSUNG" in clean_brand(x).upper())
    )
    if samsung_mask.any():
        missing_link_mask = work["备注2"].apply(lambda x: clean_text(x) == "")
        for row_idx in work[samsung_mask & missing_link_mask].index:
            part = clean_model(work.at[row_idx, "型号"])
            if part != "":
                work.at[row_idx, "备注2"] = f"https://product.samsungsem.com/mlcc/{part}.do"

    work["容值_pf"] = work.apply(lambda r: cap_to_pf(r["容值"], r["容值单位"]), axis=1)
    for col in ["器件类型", "尺寸（inch）", "材质（介质）", "容值", "容值单位", "容值误差", "耐压（V）", "_model_rule_authority"]:
        if col in work.columns and isinstance(work[col].dtype, pd.CategoricalDtype):
            work[col] = work[col].astype("object")
    work = fill_missing_series_from_model(work)
    if "器件类型" in work.columns:
        normalized_types = work["器件类型"].astype(str).apply(normalize_component_type)
        capacitor_like_mask = normalized_types.eq("MLCC")
        blank_type_mask = normalized_types.eq("")
    else:
        capacitor_like_mask = pd.Series([False] * len(work), index=work.index)
        blank_type_mask = pd.Series([True] * len(work), index=work.index)
    authority_blank_mask = pd.Series([False] * len(work), index=work.index)
    if "_model_rule_authority" in work.columns:
        authority_blank_mask = work["_model_rule_authority"].astype(str).apply(clean_text).eq("")
    backfill_mask = capacitor_like_mask | blank_type_mask | authority_blank_mask
    if backfill_mask.any():
        enriched = fill_missing_spec_fields_from_model(work.loc[backfill_mask].copy())
        for col in enriched.columns:
            work.loc[backfill_mask, col] = enriched[col]
    work = apply_model_rule_overrides_to_dataframe(work, override_conflicts=True)
    work = apply_component_type_inference_overrides_to_dataframe(work)
    work = normalize_capacitor_value_fields_from_pf(work)
    return work


def extract_workbook_hyperlink_values(workbook_path):
    try:
        wb = load_workbook(workbook_path, read_only=False, data_only=False)
    except Exception:
        return {}

    try:
        workbook_links = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
            status_col = next((idx + 1 for idx, value in enumerate(headers) if value == "生产状态"), None)
            if status_col is None:
                continue

            links = {}
            for excel_row in range(2, ws.max_row + 1):
                cell = ws.cell(excel_row, status_col)
                if cell.hyperlink is None:
                    continue
                target = cell.hyperlink.target or cell.hyperlink.location or ""
                target = clean_text(target)
                if target:
                    links[excel_row - 2] = target
            if links:
                workbook_links[sheet_name] = links
        return workbook_links
    finally:
        try:
            wb.close()
        except Exception:
            pass


def database_needs_refresh():
    if not os.path.exists(DB_PATH):
        return True
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='components'")
        if cur.fetchone() is None:
            conn.close()
            return True
        cur.execute("SELECT COUNT(*) FROM components")
        row_count = cur.fetchone()[0]
        conn.close()
        if row_count == 0:
            return True
    except Exception:
        return True
    source_files = get_source_workbooks()
    if not source_files:
        return False
    db_mtime = os.path.getmtime(DB_PATH)
    for f in source_files:
        try:
            if os.path.getmtime(f) > db_mtime:
                return True
        except Exception:
            continue
    return False


def database_has_component_rows():
    if not os.path.exists(DB_PATH):
        return False
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='components'")
        if cur.fetchone() is None:
            conn.close()
            return False
        cur.execute("SELECT COUNT(*) FROM components")
        row_count = int(cur.fetchone()[0] or 0)
        conn.close()
        return row_count > 0
    except Exception:
        return False


def get_source_normalized_cache_paths(source_path):
    cache_dir = SOURCE_NORMALIZED_CACHE_DIR
    if cache_dir:
        os.makedirs(cache_dir, exist_ok=True)
    abs_path = os.path.abspath(source_path)
    digest = hashlib.sha1(abs_path.encode("utf-8")).hexdigest()
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.splitext(os.path.basename(source_path))[0])
    if safe_name == "":
        safe_name = "source"
    cache_base = f"{safe_name}_{digest[:16]}"
    return (
        os.path.join(cache_dir, f"{cache_base}.parquet"),
        os.path.join(cache_dir, f"{cache_base}.json"),
    )


def get_source_normalized_signature(source_path):
    if not os.path.exists(source_path):
        return {"source_missing": True, "cache_version": SOURCE_NORMALIZED_CACHE_VERSION}
    stat = os.stat(source_path)
    return {
        "source_path": os.path.abspath(source_path),
        "source_mtime": round(stat.st_mtime, 6),
        "source_size": stat.st_size,
        "cache_version": SOURCE_NORMALIZED_CACHE_VERSION,
    }


def read_source_component_dataframe(source_path):
    ext = os.path.splitext(source_path)[1].lower()
    if ext == ".csv":
        try:
            df = pd.read_csv(source_path, dtype=str, encoding="utf-8-sig").dropna(how="all")
        except Exception:
            return pd.DataFrame()
        if df.empty:
            return df
        df = normalize_imported_component_dataframe(df, source_path=source_path)
        df = expand_hre_cga_rows_to_csa(df)
        if df is None or df.empty:
            return pd.DataFrame()
        return deduplicate_component_rows(df)

    try:
        xls = pd.ExcelFile(source_path)
        hyperlink_maps = extract_workbook_hyperlink_values(source_path)
    except Exception:
        return pd.DataFrame()

    all_dfs = []
    for sheet in get_workbook_sheet_names(source_path, xls):
        try:
            df = pd.read_excel(source_path, sheet_name=sheet, dtype=str)
            df = df.dropna(how="all")
            if df.empty:
                continue
            hyperlink_map = hyperlink_maps.get(sheet, {})
            if "备注2" in df.columns:
                df["备注2"] = df["备注2"].fillna("").astype(str)
                for row_idx, url in hyperlink_map.items():
                    if row_idx not in df.index:
                        continue
                    if clean_text(df.at[row_idx, "备注2"]) == "" and clean_text(url) != "":
                        df.at[row_idx, "备注2"] = url
            df = normalize_imported_component_dataframe(df, source_path=source_path)
            df = expand_hre_cga_rows_to_csa(df)
            if not df.empty:
                all_dfs.append(df)
        except Exception:
            continue
    if not all_dfs:
        return pd.DataFrame()
    combined = safe_concat_dataframes(all_dfs, ignore_index=True).drop_duplicates()
    combined = deduplicate_component_rows(combined)
    return combined


def load_or_build_normalized_source_dataframe(source_path, force=False):
    if not os.path.exists(source_path):
        return pd.DataFrame()
    cache_path, meta_path = get_source_normalized_cache_paths(source_path)
    signature = get_source_normalized_signature(source_path)
    if not force and os.path.exists(cache_path) and os.path.exists(meta_path):
        try:
            with open(meta_path, "r", encoding="utf-8") as handle:
                cached_signature = json.load(handle)
        except Exception:
            cached_signature = None
        if cached_signature == signature:
            try:
                if cache_path.lower().endswith(".parquet") and pq is not None:
                    return pd.read_parquet(cache_path)
                return pd.read_pickle(cache_path)
            except Exception:
                pass
    df = read_source_component_dataframe(source_path)
    if df is None:
        df = pd.DataFrame()
    try:
        if cache_path.lower().endswith(".parquet") and pq is not None:
            df.to_parquet(cache_path, index=False)
        else:
            df.to_pickle(cache_path)
        with open(meta_path, "w", encoding="utf-8") as handle:
            json.dump(signature, handle, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return df


def update_database(force=False):
    startup_trace(f"update_database:enter force={force}")
    all_files = get_source_workbooks()
    startup_trace(f"update_database:source-files={len(all_files)}")
    if not force and not database_needs_refresh():
        startup_trace("update_database:no-refresh-needed")
        return
    all_dfs = []
    for f in all_files:
        try:
            startup_trace(f"update_database:load-source:{os.path.basename(f)}")
            df = load_or_build_normalized_source_dataframe(f, force=force)
        except Exception:
            df = pd.DataFrame()
        if df is not None and not df.empty:
            all_dfs.append(df)
    startup_trace(f"update_database:frames={len(all_dfs)}")
    if all_dfs:
        df_all = safe_concat_dataframes(all_dfs, ignore_index=True).drop_duplicates()
        df_all = deduplicate_component_rows(df_all)
        if df_all.empty:
            startup_trace("update_database:empty-after-dedup")
            return
        df_all = enrich_mlcc_dimension_fields_in_dataframe(df_all, allow_online_lookup=False)
        prepared = prepare_search_dataframe(df_all)
        startup_trace(f"update_database:prepared-rows={len(prepared)}")
        conn = sqlite3.connect(DB_PATH)
        df_all.to_sql("components", conn, if_exists="replace", index=False)
        create_database_indexes(conn)
        try:
            refresh_search_index_table(conn, prepared)
        except Exception:
            pass
        conn.close()
        clear_data_load_caches()
        try:
            write_prepared_cache(prepared)
        except Exception:
            pass
        startup_trace("update_database:done")

if clean_text(os.getenv("COMPONENT_MATCHER_ENABLE_AUTO_UPDATE", "")) in {"1", "true", "yes", "on"}:
    update_database()

def auto_refresh_db(interval_sec=300):
    def loop():
        while True:
            try:
                update_database()
            except:
                pass
            time.sleep(interval_sec)
    threading.Thread(target=loop, daemon=True).start()

def deduplicate_component_rows(df):
    if df is None or df.empty:
        return df
    work = df.copy()
    link_col = "备注2" if "备注2" in work.columns else None
    work["_has_official_link"] = work[link_col].apply(lambda x: 1 if clean_text(x) != "" else 0) if link_col else 0
    dedup_key = [
        col for col in [
            "器件类型", "品牌", "型号", "系列", "安装方式", "封装代码",
            "尺寸（inch）", "材质（介质）", "容值", "容值单位", "容值误差",
            "耐压（V）", "特殊用途", "备注1", "备注3"
        ] if col in work.columns
    ]
    work = work.sort_values(by=["_has_official_link"], ascending=False, kind="stable")
    if dedup_key:
        work = work.drop_duplicates(subset=dedup_key, keep="first")
    else:
        work = work.drop_duplicates()
    return work.drop(columns=["_has_official_link"], errors="ignore").reset_index(drop=True)


def get_database_signature():
    if not os.path.exists(DB_PATH):
        return {"db_missing": True, "cache_version": PREPARED_CACHE_VERSION}
    stat = os.stat(DB_PATH)
    return {
        "db_path": DB_PATH,
        "db_mtime": round(stat.st_mtime, 6),
        "db_size": stat.st_size,
        "cache_version": PREPARED_CACHE_VERSION,
    }


def get_public_bundle_signature_for_cache():
    if not (is_public_mode() or is_streamlit_cloud_runtime()):
        return ""
    try:
        return get_streamlit_cloud_bundle_signature()
    except Exception:
        return ""


def get_search_index_signature():
    db_signature = get_database_signature()
    bundle_signature = get_public_bundle_signature_for_cache()
    if db_signature.get("db_missing"):
        signature = {
            "db_missing": True,
            "db_path": DB_PATH,
            "db_mtime": 0.0,
            "db_size": 0,
            "cache_version": PREPARED_CACHE_VERSION,
            "search_index_schema_version": SEARCH_INDEX_SCHEMA_VERSION,
        }
        if bundle_signature:
            signature["bundle_signature"] = bundle_signature
        return signature
    signature = {
        "db_path": db_signature.get("db_path", DB_PATH),
        "db_mtime": db_signature.get("db_mtime", 0),
        "db_size": db_signature.get("db_size", 0),
        "cache_version": PREPARED_CACHE_VERSION,
        "search_index_schema_version": SEARCH_INDEX_SCHEMA_VERSION,
    }
    if bundle_signature:
        signature["bundle_signature"] = bundle_signature
    return signature


def bundled_runtime_assets_are_current(required_paths=None):
    if not (is_public_mode() or is_streamlit_cloud_runtime()):
        return False
    if not ensure_streamlit_cloud_bundle_archive_available():
        return False
    try:
        return not streamlit_cloud_bundle_refresh_needed(required_paths=required_paths)
    except Exception:
        return False


def prepared_cache_meta_is_usable(meta):
    if not isinstance(meta, dict):
        return False
    try:
        return int(meta.get("cache_version", 0) or 0) == PREPARED_CACHE_VERSION
    except Exception:
        return False


def prepared_cache_is_current():
    if (
        (not os.path.exists(PREPARED_CACHE_PATH) and not os.path.exists(PREPARED_CACHE_FALLBACK_PATH))
        or not os.path.exists(PREPARED_CACHE_META_PATH)
    ):
        return False
    try:
        with open(PREPARED_CACHE_META_PATH, "r", encoding="utf-8") as handle:
            meta = json.load(handle)
    except Exception:
        return False
    if meta == get_database_signature():
        return True
    if not prepared_cache_meta_is_usable(meta):
        return False
    bundle_signature = get_public_bundle_signature_for_cache()
    if bundle_signature:
        meta_bundle_signature = clean_text(meta.get("bundle_signature", ""))
        if meta_bundle_signature != bundle_signature:
            return False
    try:
        cached = read_prepared_cache()
    except Exception:
        return False
    if not prepared_dataframe_looks_sane(cached):
        return False
    # Accept a sane prepared cache even when the main SQLite database has not
    # been restored yet. Cloud startups often have the cache available before
    # the larger database comes online, and treating the cache as invalid in
    # that state forces an unnecessary rebuild path.
    if not os.path.exists(DB_PATH):
        return True
    return prepared_dataframe_looks_sane(cached)


def optimize_prepared_dataframe_dtypes(df):
    if df is None or df.empty:
        return df
    work = df.copy()
    category_candidates = [
        "品牌", "系列", "尺寸（inch）", "材质（介质）", "容值单位", "容值误差", "耐压（V）",
        "特殊用途", "器件类型", "安装方式", "封装代码", "尺寸（mm）", "生产状态", "数据来源",
        "数据状态", "_size", "_mat", "_tol", "_volt", "_component_type", "_body_size",
        "_pitch", "_safety_class", "_disc_size",
    ]
    for col in category_candidates:
        if col not in work.columns:
            continue
        series = work[col]
        try:
            unique_count = int(series.nunique(dropna=False))
        except Exception:
            continue
        if unique_count == 0:
            continue
        if unique_count <= 4096 or unique_count / max(len(series), 1) <= 0.2:
            work[col] = series.astype("category")
    return work


def write_prepared_cache(df):
    if df is None or df.empty:
        return
    cache_dir = os.path.dirname(PREPARED_CACHE_PATH)
    if cache_dir:
        os.makedirs(cache_dir, exist_ok=True)
    prepared = prepare_search_dataframe(df)
    prepared = optimize_prepared_dataframe_dtypes(prepared)
    try:
        prepared.to_parquet(PREPARED_CACHE_PATH, index=False)
        if os.path.exists(PREPARED_CACHE_FALLBACK_PATH):
            os.remove(PREPARED_CACHE_FALLBACK_PATH)
    except Exception:
        prepared.to_pickle(PREPARED_CACHE_FALLBACK_PATH)
        if os.path.exists(PREPARED_CACHE_PATH):
            os.remove(PREPARED_CACHE_PATH)
    meta = dict(get_database_signature())
    bundle_signature = get_public_bundle_signature_for_cache()
    if bundle_signature:
        meta["bundle_signature"] = bundle_signature
    with open(PREPARED_CACHE_META_PATH, "w", encoding="utf-8") as handle:
        json.dump(meta, handle, ensure_ascii=False, indent=2)


def read_prepared_cache():
    if os.path.exists(PREPARED_CACHE_FALLBACK_PATH):
        return pd.read_pickle(PREPARED_CACHE_FALLBACK_PATH)
    if os.path.exists(PREPARED_CACHE_PATH):
        if pq is not None:
            try:
                return pd.read_parquet(PREPARED_CACHE_PATH)
            except Exception:
                if os.path.exists(PREPARED_CACHE_FALLBACK_PATH):
                    return pd.read_pickle(PREPARED_CACHE_FALLBACK_PATH)
        if os.path.exists(DB_PATH):
            conn = sqlite3.connect(DB_PATH)
            try:
                df = pd.read_sql("SELECT * FROM components", conn)
            finally:
                conn.close()
            rebuilt = prepare_search_dataframe(deduplicate_component_rows(df))
            try:
                rebuilt.to_pickle(PREPARED_CACHE_FALLBACK_PATH)
            except Exception:
                pass
            return rebuilt
    raise FileNotFoundError("prepared cache not found")


def create_database_indexes(conn):
    statements = [
        'CREATE INDEX IF NOT EXISTS idx_components_model ON components("型号")',
        'CREATE INDEX IF NOT EXISTS idx_components_brand_model ON components("品牌", "型号")',
        'CREATE INDEX IF NOT EXISTS idx_components_type_size ON components("器件类型", "尺寸（inch）")',
    ]
    cur = conn.cursor()
    for statement in statements:
        try:
            cur.execute(statement)
        except Exception:
            continue
    conn.commit()


def build_search_index_dataframe(prepared):
    if prepared is None or prepared.empty:
        return pd.DataFrame()
    if prepared_dataframe_has_required_columns(prepared):
        work = prepared.copy()
    else:
        work = prepare_search_dataframe(prepared)
    if work.empty:
        return pd.DataFrame()

    def null_if_blank(value):
        value = clean_text(value)
        return value if value != "" else None

    columns = {
        "品牌": work["品牌"].astype("string"),
        "型号": work["型号"].astype("string"),
        "_model_clean": work["_model_clean"].astype("string"),
        "_component_type": work["_component_type"].astype("string"),
        "_size": work["_size"].map(null_if_blank),
        "_mat": work["_mat"].map(null_if_blank),
        "_tol": work["_tol"].map(null_if_blank),
        "_volt_num": pd.to_numeric(work["_volt_num"], errors="coerce"),
        "_pf": pd.to_numeric(work["_pf"], errors="coerce"),
        "_res_ohm": pd.to_numeric(work["_res_ohm"], errors="coerce"),
        "_power_watt": pd.to_numeric(work["_power_watt"], errors="coerce"),
        "_body_size": work["_body_size"].map(null_if_blank),
        "_pitch": work["_pitch"].map(null_if_blank),
        "_safety_class": work["_safety_class"].map(null_if_blank),
        "_varistor_voltage": work["_varistor_voltage"].map(null_if_blank),
        "_disc_size": work["_disc_size"].map(null_if_blank),
        "_temp_low": pd.to_numeric(work["_temp_low"], errors="coerce"),
        "_temp_high": pd.to_numeric(work["_temp_high"], errors="coerce"),
        "_life_hours_num": pd.to_numeric(work["_life_hours_num"], errors="coerce"),
        "_mount_style": work["_mount_style"].map(null_if_blank),
        "_special_use_norm": work["_special_use_norm"].map(null_if_blank),
        "_unit_upper": work["容值单位"].map(lambda value: null_if_blank(value).upper() if null_if_blank(value) is not None else None),
        "_value_num": pd.to_numeric(work["容值"], errors="coerce"),
    }
    search_df = pd.DataFrame(columns)
    search_df["品牌"] = search_df["品牌"].astype("string")
    search_df["型号"] = search_df["型号"].astype("string")
    search_df["_model_clean"] = search_df["_model_clean"].astype("string")
    search_df["_component_type"] = search_df["_component_type"].astype("string")
    search_df["_size"] = search_df["_size"].astype("string")
    search_df["_mat"] = search_df["_mat"].astype("string")
    search_df["_tol"] = search_df["_tol"].astype("string")
    search_df["_body_size"] = search_df["_body_size"].astype("string")
    search_df["_pitch"] = search_df["_pitch"].astype("string")
    search_df["_safety_class"] = search_df["_safety_class"].astype("string")
    search_df["_varistor_voltage"] = search_df["_varistor_voltage"].map(lambda value: clean_voltage(value) or None)
    search_df["_disc_size"] = search_df["_disc_size"].astype("string")
    search_df["_mount_style"] = search_df["_mount_style"].astype("string")
    search_df["_special_use_norm"] = search_df["_special_use_norm"].astype("string")
    search_df["_unit_upper"] = search_df["_unit_upper"].astype("string")
    search_df = prioritize_component_rows_for_lookup(search_df)
    search_df = search_df.drop_duplicates(subset=["品牌", "型号"], keep="first").reset_index(drop=True)
    return search_df


def create_search_table_indexes(conn):
    statements = [
        f'CREATE INDEX IF NOT EXISTS idx_{COMPONENTS_SEARCH_TABLE}_brand_model ON {COMPONENTS_SEARCH_TABLE}("品牌", "型号")',
        f'CREATE INDEX IF NOT EXISTS idx_{COMPONENTS_SEARCH_TABLE}_model_clean ON {COMPONENTS_SEARCH_TABLE}("_model_clean")',
        f'CREATE INDEX IF NOT EXISTS idx_{COMPONENTS_SEARCH_TABLE}_resistor ON {COMPONENTS_SEARCH_TABLE}("_component_type", "_size", "_res_ohm", "_tol", "_power_watt")',
        f'CREATE INDEX IF NOT EXISTS idx_{COMPONENTS_SEARCH_TABLE}_mlcc ON {COMPONENTS_SEARCH_TABLE}("_component_type", "_size", "_mat", "_pf", "_tol", "_volt_num")',
        f'CREATE INDEX IF NOT EXISTS idx_{COMPONENTS_SEARCH_TABLE}_type_size ON {COMPONENTS_SEARCH_TABLE}("_component_type", "_size")',
        f'CREATE INDEX IF NOT EXISTS idx_{COMPONENTS_SEARCH_TABLE}_value_unit ON {COMPONENTS_SEARCH_TABLE}("_component_type", "_size", "_unit_upper", "_value_num", "_tol")',
        f'CREATE INDEX IF NOT EXISTS idx_{COMPONENTS_SEARCH_TABLE}_cap_body ON {COMPONENTS_SEARCH_TABLE}("_component_type", "_pf", "_volt_num", "_body_size", "_pitch")',
        f'CREATE INDEX IF NOT EXISTS idx_{COMPONENTS_SEARCH_TABLE}_varistor ON {COMPONENTS_SEARCH_TABLE}("_component_type", "_varistor_voltage", "_disc_size", "_pitch", "_tol")',
    ]
    cur = conn.cursor()
    for statement in statements:
        try:
            cur.execute(statement)
        except Exception:
            continue
    conn.commit()


def configure_sqlite_bulk_load(conn):
    pragmas = [
        "PRAGMA journal_mode = OFF",
        "PRAGMA synchronous = OFF",
        "PRAGMA temp_store = MEMORY",
        "PRAGMA cache_size = -200000",
        "PRAGMA locking_mode = EXCLUSIVE",
    ]
    for statement in pragmas:
        try:
            conn.execute(statement)
        except Exception:
            continue


def sqlite_bulk_insert_chunksize(df, max_variables=30000, hard_cap=1000):
    if df is None or getattr(df, "empty", True):
        return 1
    column_count = max(int(len(getattr(df, "columns", []))), 1)
    safe_by_variables = max(1, int(max_variables // column_count))
    return max(1, min(int(hard_cap), safe_by_variables))


def normalize_search_sidecar_value(value):
    text = clean_text(value)
    return text if text != "" else None


def normalize_search_sidecar_frame(df, columns, numeric_columns=(), dedupe_columns=("品牌", "型号")):
    if df is None or df.empty:
        return pd.DataFrame(columns=list(columns))
    work = df.copy()
    for column in columns:
        if column not in work.columns:
            work[column] = None
    work = work.loc[:, list(columns)].copy()
    numeric_columns = set(numeric_columns or [])
    for column in columns:
        if column in numeric_columns:
            work[column] = pd.to_numeric(work[column], errors="coerce")
        else:
            work[column] = work[column].map(normalize_search_sidecar_value)
    dedupe_columns = [column for column in (dedupe_columns or []) if column in work.columns]
    if {"品牌", "型号"}.issubset(work.columns) and {"品牌", "型号"}.issubset(set(dedupe_columns)):
        work = prioritize_component_rows_for_lookup(work)
    if dedupe_columns:
        work = work.drop_duplicates(subset=dedupe_columns, keep="first")
    else:
        work = work.drop_duplicates(keep="first")
    return work.reset_index(drop=True)


def search_index_table_for_component_type(component_type):
    component_type = clean_text(component_type)
    if component_type in VARISTOR_COMPONENT_TYPES:
        return COMPONENTS_SEARCH_VARISTOR_TABLE
    if component_type in RESISTOR_COMPONENT_TYPES or component_type == "热敏电阻":
        return COMPONENTS_SEARCH_RESISTOR_TABLE
    if component_type in CAPACITOR_COMPONENT_TYPES:
        return COMPONENTS_SEARCH_CAPACITOR_TABLE
    if component_type in INDUCTOR_COMPONENT_TYPES or component_type in TIMING_COMPONENT_TYPES:
        return COMPONENTS_SEARCH_VALUE_TABLE
    return COMPONENTS_SEARCH_CORE_TABLE


def get_search_sidecar_table_specs():
    return {
        COMPONENTS_SEARCH_CORE_TABLE: {
            "columns": ["品牌", "型号", "_model_clean", "_component_type"],
            "numeric": [],
            "indexes": [
                ("brand_model", ["品牌", "型号"]),
                ("model_clean", ["_model_clean"]),
            ],
        },
        COMPONENTS_SEARCH_RESISTOR_TABLE: {
            "columns": ["品牌", "型号", "_component_type", "_size", "_res_ohm", "_tol", "_power_watt"],
            "numeric": ["_res_ohm", "_power_watt"],
            "indexes": [
                ("brand_model", ["品牌", "型号"]),
                ("type_size", ["_component_type", "_size"]),
                ("resistor", ["_component_type", "_size", "_res_ohm", "_tol", "_power_watt"]),
            ],
        },
        COMPONENTS_SEARCH_CAPACITOR_TABLE: {
            "columns": ["品牌", "型号", "_component_type", "_size", "_mat", "_pf", "_tol", "_volt_num", "_body_size", "_pitch", "_safety_class", "_temp_low", "_temp_high", "_life_hours_num", "_mount_style", "_special_use_norm"],
            "numeric": ["_pf", "_volt_num", "_temp_low", "_temp_high", "_life_hours_num"],
            "indexes": [
                ("brand_model", ["品牌", "型号"]),
                ("type_size", ["_component_type", "_size"]),
                ("mlcc", ["_component_type", "_size", "_mat", "_pf", "_tol", "_volt_num"]),
                ("cap_body", ["_component_type", "_pf", "_volt_num", "_body_size", "_pitch"]),
                ("cap_env", ["_component_type", "_pf", "_volt_num", "_temp_high", "_life_hours_num", "_mount_style"]),
            ],
        },
        COMPONENTS_SEARCH_VALUE_TABLE: {
            "columns": ["品牌", "型号", "_component_type", "_size", "_unit_upper", "_value_num", "_tol", "_volt_num"],
            "numeric": ["_value_num", "_volt_num"],
            "indexes": [
                ("brand_model", ["品牌", "型号"]),
                ("type_size", ["_component_type", "_size"]),
                ("value_unit", ["_component_type", "_size", "_unit_upper", "_value_num", "_tol"]),
            ],
        },
        COMPONENTS_SEARCH_VARISTOR_TABLE: {
            "columns": ["品牌", "型号", "_component_type", "_varistor_voltage", "_disc_size", "_pitch", "_tol"],
            "numeric": [],
            "indexes": [
                ("brand_model", ["品牌", "型号"]),
                ("varistor", ["_component_type", "_varistor_voltage", "_disc_size", "_pitch", "_tol"]),
            ],
        },
    }


def build_search_sidecar_frames(prepared):
    if prepared is None or prepared.empty:
        return {}
    if prepared_dataframe_has_required_columns(prepared):
        work = prepared.copy()
    else:
        work = prepare_search_dataframe(prepared)
    if work.empty:
        return {}
    work["品牌"] = work["品牌"].astype("string")
    work["型号"] = work["型号"].astype("string")
    work["_model_clean"] = work["_model_clean"].astype("string")
    work["_component_type"] = work["_component_type"].astype("string")
    table_specs = get_search_sidecar_table_specs()
    frames = {}
    for table_name, spec in table_specs.items():
        if table_name == COMPONENTS_SEARCH_CORE_TABLE:
            frame = work.loc[:, ["品牌", "型号", "_model_clean", "_component_type"]].copy()
        elif table_name == COMPONENTS_SEARCH_RESISTOR_TABLE:
            mask = work["_component_type"].isin(RESISTOR_COMPONENT_TYPES | {"热敏电阻"})
            frame = work.loc[mask, spec["columns"]].copy()
        elif table_name == COMPONENTS_SEARCH_CAPACITOR_TABLE:
            mask = work["_component_type"].isin(CAPACITOR_COMPONENT_TYPES)
            frame = work.loc[mask, spec["columns"]].copy()
        elif table_name == COMPONENTS_SEARCH_VALUE_TABLE:
            mask = work["_component_type"].isin(INDUCTOR_COMPONENT_TYPES | TIMING_COMPONENT_TYPES)
            frame = work.loc[mask, spec["columns"]].copy()
        elif table_name == COMPONENTS_SEARCH_VARISTOR_TABLE:
            mask = work["_component_type"].isin(VARISTOR_COMPONENT_TYPES)
            frame = work.loc[mask, spec["columns"]].copy()
        else:
            continue
        if frame.empty:
            continue
        frames[table_name] = normalize_search_sidecar_frame(
            frame,
            spec["columns"],
            numeric_columns=spec.get("numeric", ()),
            dedupe_columns=("品牌", "型号"),
        )
    return frames


def drop_search_sidecar_tables(conn):
    tables = [
        COMPONENTS_SEARCH_LEGACY_TABLE,
        COMPONENTS_SEARCH_CORE_TABLE,
        COMPONENTS_SEARCH_RESISTOR_TABLE,
        COMPONENTS_SEARCH_CAPACITOR_TABLE,
        COMPONENTS_SEARCH_VALUE_TABLE,
        COMPONENTS_SEARCH_VARISTOR_TABLE,
        SEARCH_META_TABLE,
    ]
    cur = conn.cursor()
    for table in tables:
        try:
            cur.execute(f"DROP TABLE IF EXISTS {table}")
        except Exception:
            continue
    conn.commit()


def initialize_empty_search_sidecar_tables(conn):
    for table_name, spec in get_search_sidecar_table_specs().items():
        try:
            pd.DataFrame(columns=spec["columns"]).to_sql(
                table_name,
                conn,
                if_exists="replace",
                index=False,
            )
        except Exception:
            continue


def create_search_sidecar_indexes(conn):
    table_specs = get_search_sidecar_table_specs()
    cur = conn.cursor()
    for table_name, spec in table_specs.items():
        for suffix, columns in spec.get("indexes", []):
            column_expr = ", ".join([f'"{column}"' for column in columns])
            statement = f'CREATE INDEX IF NOT EXISTS idx_{table_name}_{suffix} ON {table_name}({column_expr})'
            try:
                cur.execute(statement)
            except Exception:
                continue
    conn.commit()


def write_search_sidecar_frames(conn, frames):
    if not frames:
        return 0
    total_rows = 0
    for table_name, frame in frames.items():
        if frame is None or frame.empty:
            continue
        frame.to_sql(
            table_name,
            conn,
            if_exists="replace",
            index=False,
            method="multi",
            chunksize=sqlite_bulk_insert_chunksize(frame),
        )
        total_rows += int(len(frame))
    return total_rows


def append_search_sidecar_frames(conn, frames, table_written, row_counts):
    appended_rows = 0
    for table_name, frame in frames.items():
        if frame is None or frame.empty:
            continue
        if table_name not in table_written:
            table_written[table_name] = False
        if table_name not in row_counts:
            row_counts[table_name] = 0
        frame.to_sql(
            table_name,
            conn,
            if_exists="replace" if not table_written[table_name] else "append",
            index=False,
            method="multi",
            chunksize=sqlite_bulk_insert_chunksize(frame),
        )
        table_written[table_name] = True
        row_counts[table_name] += int(len(frame))
        appended_rows += int(len(frame))
    return appended_rows


def refresh_search_index_table(conn, prepared):
    search_df = build_search_index_dataframe(prepared)
    cur = conn.cursor()
    try:
        cur.execute(f"DROP TABLE IF EXISTS {COMPONENTS_SEARCH_TABLE}")
        conn.commit()
    except Exception:
        pass
    if search_df.empty:
        return
    search_df.to_sql(
        COMPONENTS_SEARCH_TABLE,
        conn,
        if_exists="replace",
        index=False,
        method="multi",
        chunksize=sqlite_bulk_insert_chunksize(search_df),
    )
    create_search_table_indexes(conn)


def search_table_exists(conn, table_name=COMPONENTS_SEARCH_CORE_TABLE):
    try:
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table_name,),
        ).fetchone()
        return row is not None
    except Exception:
        return False


def search_table_has_columns(conn, required_columns, table_name=COMPONENTS_SEARCH_CORE_TABLE):
    try:
        rows = conn.execute(f'PRAGMA table_info({table_name})').fetchall()
        available = {clean_text(row[1]) for row in rows}
        return all(clean_text(column) in available for column in (required_columns or []))
    except Exception:
        return False


def write_search_index_meta(conn, row_counts):
    meta = dict(get_search_index_signature())
    if isinstance(row_counts, dict):
        normalized_counts = {
            clean_text(table_name): int(count or 0)
            for table_name, count in row_counts.items()
            if clean_text(table_name) != ""
        }
        meta["table_row_counts"] = normalized_counts
        meta["row_count"] = int(sum(normalized_counts.values()))
    else:
        meta["row_count"] = int(row_counts or 0)
        meta["table_row_counts"] = {COMPONENTS_SEARCH_TABLE: int(row_counts or 0)}
    conn.execute(f"DROP TABLE IF EXISTS {SEARCH_META_TABLE}")
    conn.execute(
        f"CREATE TABLE {SEARCH_META_TABLE} (meta_json TEXT NOT NULL)"
    )
    conn.execute(
        f"INSERT INTO {SEARCH_META_TABLE} (meta_json) VALUES (?)",
        (json.dumps(meta, ensure_ascii=False, sort_keys=True),),
    )
    conn.commit()


def read_search_index_meta(conn):
    try:
        row = conn.execute(f"SELECT meta_json FROM {SEARCH_META_TABLE} LIMIT 1").fetchone()
        if row is None or clean_text(row[0]) == "":
            return None
        return json.loads(row[0])
    except Exception:
        return None


def search_index_meta_is_usable(meta):
    if not isinstance(meta, dict):
        return False
    try:
        schema_version = int(meta.get("search_index_schema_version", 0) or 0)
        cache_version = int(meta.get("cache_version", 0) or 0)
    except Exception:
        return False
    return (
        schema_version == SEARCH_INDEX_SCHEMA_VERSION
        and cache_version == PREPARED_CACHE_VERSION
    )


def search_index_is_current(conn, required_columns=None, table_name=COMPONENTS_SEARCH_CORE_TABLE):
    if not search_table_exists(conn, table_name=table_name):
        return False
    if required_columns and not search_table_has_columns(conn, required_columns, table_name=table_name):
        return False
    meta = read_search_index_meta(conn)
    if not isinstance(meta, dict):
        return False
    if not search_index_meta_is_usable(meta):
        return False
    bundle_signature = get_public_bundle_signature_for_cache()
    if bundle_signature:
        meta_bundle_signature = clean_text(meta.get("bundle_signature", ""))
        if meta_bundle_signature != bundle_signature:
            return False
    # Do not compare source DB mtime/size here.
    # The source database can change size for unrelated data refreshes while the
    # search sidecar tables remain fully usable. Treating those file-level changes
    # as a hard invalidation forces an expensive rebuild on every search.
    return True


def search_index_can_serve_queries(conn, required_columns=None, table_name=COMPONENTS_SEARCH_CORE_TABLE, allow_without_database=False):
    if search_index_is_current(conn, required_columns=required_columns, table_name=table_name):
        return True
    if bundled_runtime_assets_are_current(required_paths=[SEARCH_DB_PATH]):
        if not search_table_exists(conn, table_name=table_name):
            return False
        if required_columns and not search_table_has_columns(conn, required_columns, table_name=table_name):
            return False
        meta = read_search_index_meta(conn)
        return search_index_meta_is_usable(meta)
    if allow_without_database and not os.path.exists(DB_PATH):
        if not search_table_exists(conn, table_name=table_name):
            return False
        if required_columns and not search_table_has_columns(conn, required_columns, table_name=table_name):
            return False
        meta = read_search_index_meta(conn)
        return isinstance(meta, dict)
    return False


def open_search_db_connection(timeout_sec=30):
    if not os.path.exists(SEARCH_DB_PATH):
        if is_public_mode() or is_streamlit_cloud_runtime():
            return None
        ensure_streamlit_cloud_data_bundle(required_paths=[SEARCH_DB_PATH])
    if not os.path.exists(SEARCH_DB_PATH):
        return None
    conn = sqlite3.connect(SEARCH_DB_PATH, timeout=float(timeout_sec))
    conn.execute(f"PRAGMA busy_timeout = {int(timeout_sec * 1000)}")
    return conn


def build_search_index_chunk_from_raw(raw_df):
    return build_search_index_dataframe(raw_df)


def rebuild_search_index_from_database_fast(chunk_rows=100000):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(DB_PATH)
    if os.path.exists(PREPARED_CACHE_PATH):
        try:
            if prepared_cache_is_current():
                rebuild_search_index_table_from_prepared_cache(PREPARED_CACHE_PATH)
                return
        except Exception:
            pass
    search_dir = os.path.dirname(SEARCH_DB_PATH)
    if search_dir:
        os.makedirs(search_dir, exist_ok=True)
    temp_path = SEARCH_DB_PATH + ".tmp"
    if os.path.exists(temp_path):
        os.remove(temp_path)
    read_conn = sqlite3.connect(DB_PATH, timeout=60)
    read_conn.execute("PRAGMA busy_timeout = 60000")
    write_conn = sqlite3.connect(temp_path, timeout=60)
    write_conn.execute("PRAGMA busy_timeout = 60000")
    configure_sqlite_bulk_load(write_conn)
    try:
        drop_search_sidecar_tables(write_conn)
        initialize_empty_search_sidecar_tables(write_conn)
        wrote_any = False
        table_written = {table_name: False for table_name in get_search_sidecar_table_specs().keys()}
        row_counts = {table_name: 0 for table_name in get_search_sidecar_table_specs().keys()}
        for chunk in pd.read_sql_query("SELECT * FROM components", read_conn, chunksize=int(chunk_rows)):
            if chunk is None or chunk.empty:
                continue
            search_frames = build_search_sidecar_frames(chunk)
            if not search_frames:
                continue
            appended_rows = append_search_sidecar_frames(write_conn, search_frames, table_written, row_counts)
            if appended_rows > 0:
                wrote_any = True
        if wrote_any:
            create_search_sidecar_indexes(write_conn)
            write_search_index_meta(write_conn, row_counts)
    finally:
        read_conn.close()
        write_conn.close()
    if not wrote_any:
        raise RuntimeError("no search index rows were written")
    replace_file_atomically(temp_path, SEARCH_DB_PATH)


@st.cache_data(ttl=3600)
def _load_data_cached(cache_signature):
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT * FROM components", conn)
    conn.close()
    return deduplicate_component_rows(df)


def load_data(cache_signature=None):
    if cache_signature is None:
        cache_signature = get_raw_data_cache_signature()
    return _load_data_cached(cache_signature)


def prepared_dataframe_has_required_columns(df):
    if df is None:
        return False
    return all(col in df.columns for col in PREPARED_SEARCH_REQUIRED_COLUMNS)


def prepared_dataframe_looks_sane(df):
    if not prepared_dataframe_has_required_columns(df):
        return False
    if "器件类型" not in df.columns or "_component_type" not in df.columns:
        return True

    raw_resistor_mask = df["器件类型"].isin(RESISTOR_COMPONENT_TYPES)
    raw_resistor_count = int(raw_resistor_mask.sum())
    if raw_resistor_count < 1000:
        return True

    prepared_types = df.loc[raw_resistor_mask, "_component_type"].astype("string").fillna("")
    resistor_family_count = int(prepared_types.isin(RESISTOR_COMPONENT_TYPES).sum())
    thermistor_count = int(prepared_types.eq("热敏电阻").sum())

    if resistor_family_count == 0:
        return False
    if thermistor_count / max(raw_resistor_count, 1) > 0.5 and resistor_family_count / max(raw_resistor_count, 1) < 0.4:
        return False
    return True


def replace_file_atomically(src_path, dst_path):
    last_error = None
    for attempt in range(6):
        try:
            os.replace(src_path, dst_path)
            return
        except PermissionError as exc:
            last_error = exc
            if attempt >= 5:
                break
            time.sleep(0.4 * (attempt + 1))
    raise last_error


def build_unique_temp_path(path, label="tmp"):
    return f"{path}.{label}.{os.getpid()}.{int(time.time() * 1000)}"


def rebuild_search_index_table_from_prepared_cache(prepared_path=None):
    if prepared_path is None:
        prepared_path = PREPARED_CACHE_PATH
    if not os.path.exists(prepared_path):
        raise FileNotFoundError(prepared_path)
    search_dir = os.path.dirname(SEARCH_DB_PATH)
    if search_dir:
        os.makedirs(search_dir, exist_ok=True)
    temp_path = build_unique_temp_path(SEARCH_DB_PATH, "tmp")
    conn = sqlite3.connect(temp_path, timeout=60)
    conn.execute("PRAGMA busy_timeout = 60000")
    configure_sqlite_bulk_load(conn)
    try:
        drop_search_sidecar_tables(conn)
        initialize_empty_search_sidecar_tables(conn)
        wrote_any = False
        table_written = {table_name: False for table_name in get_search_sidecar_table_specs().keys()}
        row_counts = {table_name: 0 for table_name in get_search_sidecar_table_specs().keys()}
        if pq is not None and prepared_path.lower().endswith(".parquet"):
            parquet_file = pq.ParquetFile(prepared_path)
            for row_group_index in range(parquet_file.num_row_groups):
                chunk = parquet_file.read_row_group(row_group_index).to_pandas()
                search_frames = build_search_sidecar_frames(chunk)
                if not search_frames:
                    continue
                appended_rows = append_search_sidecar_frames(conn, search_frames, table_written, row_counts)
                if appended_rows > 0:
                    wrote_any = True
        else:
            search_frames = build_search_sidecar_frames(read_prepared_cache())
            if search_frames:
                appended_rows = append_search_sidecar_frames(conn, search_frames, table_written, row_counts)
                if appended_rows > 0:
                    wrote_any = True
        if wrote_any:
            create_search_sidecar_indexes(conn)
            write_search_index_meta(conn, row_counts)
    finally:
        conn.close()
    if not wrote_any:
        raise RuntimeError("no search index rows were written")
    replace_file_atomically(temp_path, SEARCH_DB_PATH)


def rebuild_prepared_cache_from_database(chunk_rows=COMPONENTS_SEARCH_CHUNK_ROWS):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(DB_PATH)
    if pa is None or pq is None:
        conn = sqlite3.connect(DB_PATH)
        try:
            df = pd.read_sql("SELECT * FROM components", conn)
            df = deduplicate_component_rows(df)
        finally:
            conn.close()
        clear_data_load_caches()
        write_prepared_cache(df)
        rebuild_search_index_from_database_fast()
        return

    cache_dir = os.path.dirname(PREPARED_CACHE_PATH)
    if cache_dir:
        os.makedirs(cache_dir, exist_ok=True)
    temp_parquet_path = build_unique_temp_path(PREPARED_CACHE_PATH, "tmp")
    temp_meta_path = build_unique_temp_path(PREPARED_CACHE_META_PATH, "tmp")
    writer = None
    wrote_any = False
    prepared_columns = None
    conn = sqlite3.connect(DB_PATH)
    try:
        for chunk in pd.read_sql_query("SELECT * FROM components", conn, chunksize=int(chunk_rows)):
            if chunk is None or chunk.empty:
                continue
            prepared_chunk = prepare_search_dataframe(chunk)
            if prepared_chunk.empty:
                continue
            if prepared_columns is None:
                prepared_columns = list(prepared_chunk.columns)
            else:
                prepared_chunk = prepared_chunk.reindex(columns=prepared_columns, fill_value="")
            category_columns = list(prepared_chunk.select_dtypes(include=["category"]).columns)
            for col in category_columns:
                prepared_chunk[col] = prepared_chunk[col].astype("string")
            stable_float_columns = [
                "容值_pf", "_resistance_ohm", "_pf", "_tol_num",
                "_volt_num", "_res_ohm", "_power_watt",
            ]
            for col in stable_float_columns:
                if col in prepared_chunk.columns:
                    prepared_chunk[col] = pd.to_numeric(prepared_chunk[col], errors="coerce").astype("float64")
            string_columns = [col for col in prepared_chunk.columns if col not in stable_float_columns]
            for col in string_columns:
                prepared_chunk[col] = prepared_chunk[col].astype("string")
            table = pa.Table.from_pandas(prepared_chunk, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(temp_parquet_path, table.schema, compression="snappy")
            writer.write_table(table)
            wrote_any = True
    finally:
        if writer is not None:
            writer.close()
        conn.close()

    if not wrote_any:
        raise RuntimeError("no prepared cache rows were written")

    with open(temp_meta_path, "w", encoding="utf-8") as handle:
        json.dump(get_database_signature(), handle, ensure_ascii=False, indent=2)

    replace_file_atomically(temp_parquet_path, PREPARED_CACHE_PATH)
    replace_file_atomically(temp_meta_path, PREPARED_CACHE_META_PATH)
    if os.path.exists(PREPARED_CACHE_FALLBACK_PATH):
        os.remove(PREPARED_CACHE_FALLBACK_PATH)

    rebuild_search_index_table_from_prepared_cache(PREPARED_CACHE_PATH)
    clear_data_load_caches()


SERIES_BACKFILL_DB_COLUMNS = ("系列", "系列说明", "特殊用途")

COMPONENT_TYPE_BACKFILL_DB_COLUMNS = ("器件类型",)


def normalize_series_backfill_db_value(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def normalize_component_type_backfill_db_value(value):
    text = clean_text(value)
    normalized = normalize_component_type(text)
    return normalized if normalized != "" else text


def build_component_type_backfill_updates(chunk):
    if chunk is None or chunk.empty:
        return []
    if "__rowid__" not in chunk.columns:
        raise KeyError("__rowid__")
    row_ids = pd.to_numeric(chunk["__rowid__"], errors="coerce")
    current_types = chunk["器件类型"].astype("string").fillna("").apply(normalize_component_type) if "器件类型" in chunk.columns else pd.Series([""] * len(chunk), index=chunk.index, dtype="string")
    summary_text = chunk["规格摘要"].astype("string").fillna("").apply(clean_text) if "规格摘要" in chunk.columns else pd.Series([""] * len(chunk), index=chunk.index, dtype="string")
    note_text = chunk["校验备注"].astype("string").fillna("").apply(clean_text) if "校验备注" in chunk.columns else pd.Series([""] * len(chunk), index=chunk.index, dtype="string")
    candidate_mask = (
        current_types.eq("")
        | current_types.isin(CAPACITOR_COMPONENT_TYPES)
        | summary_text.str.contains("电阻", na=False)
        | note_text.str.contains("resistor", case=False, na=False)
    )
    if not candidate_mask.any():
        return []
    inferred = chunk.loc[candidate_mask].apply(infer_db_component_type, axis=1).astype("string").fillna("").apply(normalize_component_type)
    updates = []
    for row_index in inferred.index:
        new_type = normalize_component_type_backfill_db_value(inferred.at[row_index])
        current_type = normalize_component_type_backfill_db_value(current_types.at[row_index])
        if new_type == "" or new_type == current_type:
            continue
        row_id = row_ids.at[row_index]
        if pd.isna(row_id):
            continue
        updates.append((new_type, int(row_id)))
    return updates


def backfill_component_types_in_database_in_place(chunk_rows=100000):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(DB_PATH)

    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.execute("PRAGMA busy_timeout = 60000")
    try:
        conn.execute("PRAGMA journal_mode = WAL")
    except sqlite3.DatabaseError:
        pass

    try:
        ensure_sqlite_table_columns_exist(
            conn,
            "components",
            {
                "长度（mm）": "TEXT",
                "宽度（mm）": "TEXT",
                "高度（mm）": "TEXT",
                "尺寸来源": "TEXT",
            },
        )
        min_rowid, max_rowid = conn.execute("SELECT MIN(rowid), MAX(rowid) FROM components").fetchone()
        if min_rowid is None or max_rowid is None:
            raise RuntimeError("database contains no component rows")

        batch_size = max(int(chunk_rows), 1000)
        update_sql = (
            'UPDATE components SET "器件类型" = ?, "容值_pf" = NULL '
            'WHERE rowid = ?'
        )
        updated_rows = 0
        batch_start = int(min_rowid)
        final_rowid = int(max_rowid)
        while batch_start <= final_rowid:
            batch_end = batch_start + batch_size - 1
            chunk = pd.read_sql_query(
                'SELECT rowid AS "__rowid__", * FROM components '
                'WHERE rowid BETWEEN ? AND ? '
                'AND COALESCE("器件类型", "") LIKE "%MLCC%" '
                'AND (COALESCE("长度（mm）", "") = "" OR COALESCE("宽度（mm）", "") = "" OR COALESCE("高度（mm）", "") = "" OR COALESCE("尺寸来源", "") = "")',
                conn,
                params=[batch_start, batch_end],
            )
            if chunk is not None and not chunk.empty:
                updates = build_component_type_backfill_updates(chunk)
                if updates:
                    conn.executemany(update_sql, updates)
                    conn.commit()
                    updated_rows += len(updates)
            batch_start = batch_end + 1
        return updated_rows
    finally:
        conn.close()


def build_capacitor_value_unit_backfill_updates(chunk):
    if chunk is None or chunk.empty:
        return []
    if "__rowid__" not in chunk.columns:
        raise KeyError("__rowid__")
    normalized_types = chunk["器件类型"].astype("string").fillna("").apply(normalize_component_type) if "器件类型" in chunk.columns else pd.Series([""] * len(chunk), index=chunk.index, dtype="string")
    pf_series = pd.to_numeric(chunk["容值_pf"], errors="coerce") if "容值_pf" in chunk.columns else pd.Series([None] * len(chunk), index=chunk.index, dtype="float64")
    current_units = chunk["容值单位"].astype("string").fillna("").apply(lambda value: clean_text(value).upper()) if "容值单位" in chunk.columns else pd.Series([""] * len(chunk), index=chunk.index, dtype="string")
    current_values = chunk["容值"].astype("string").fillna("").apply(clean_text) if "容值" in chunk.columns else pd.Series([""] * len(chunk), index=chunk.index, dtype="string")
    row_ids = pd.to_numeric(chunk["__rowid__"], errors="coerce")
    candidate_mask = (
        normalized_types.isin(CAPACITOR_COMPONENT_TYPES)
        & pf_series.notna()
        & (
            current_units.eq("")
            | ~current_units.isin({"PF", "NF", "UF"})
            | current_values.eq("")
        )
    )
    if not candidate_mask.any():
        return []
    updates = []
    for row_index in candidate_mask[candidate_mask].index:
        row_id = row_ids.at[row_index]
        if pd.isna(row_id):
            continue
        value, unit = pf_to_value_unit(float(pf_series.at[row_index]))
        if value == "" or unit == "":
            continue
        if current_values.at[row_index] == value and current_units.at[row_index] == unit:
            continue
        updates.append((value, unit, int(row_id)))
    return updates


def backfill_capacitor_value_units_in_database_in_place(chunk_rows=100000):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(DB_PATH)

    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.execute("PRAGMA busy_timeout = 60000")
    try:
        conn.execute("PRAGMA journal_mode = WAL")
    except sqlite3.DatabaseError:
        pass

    try:
        min_rowid, max_rowid = conn.execute("SELECT MIN(rowid), MAX(rowid) FROM components").fetchone()
        if min_rowid is None or max_rowid is None:
            raise RuntimeError("database contains no component rows")

        batch_size = max(int(chunk_rows), 1000)
        update_sql = (
            'UPDATE components SET "容值" = ?, "容值单位" = ? '
            'WHERE rowid = ?'
        )
        updated_rows = 0
        batch_start = int(min_rowid)
        final_rowid = int(max_rowid)
        while batch_start <= final_rowid:
            batch_end = batch_start + batch_size - 1
            chunk = pd.read_sql_query(
                'SELECT rowid AS "__rowid__", * FROM components '
                'WHERE rowid BETWEEN ? AND ? '
                'AND COALESCE("器件类型", "") LIKE "%MLCC%" '
                'AND (COALESCE("长度（mm）", "") = "" OR COALESCE("宽度（mm）", "") = "" OR COALESCE("高度（mm）", "") = "" OR COALESCE("尺寸来源", "") = "")',
                conn,
                params=[batch_start, batch_end],
            )
            if chunk is not None and not chunk.empty:
                updates = build_capacitor_value_unit_backfill_updates(chunk)
                if updates:
                    conn.executemany(update_sql, updates)
                    conn.commit()
                    updated_rows += len(updates)
            batch_start = batch_end + 1
        return updated_rows
    finally:
        conn.close()


def build_series_backfill_updates(chunk):
    if chunk is None or chunk.empty:
        return []

    if "__rowid__" not in chunk.columns:
        raise KeyError("__rowid__")

    original_chunk = chunk.copy()
    row_ids = pd.to_numeric(original_chunk.pop("__rowid__"), errors="coerce")
    filled_chunk = fill_missing_series_from_model(original_chunk)
    updates = []
    for row_index, row_id in enumerate(row_ids.tolist()):
        if pd.isna(row_id):
            continue
        current_values = tuple(
            normalize_series_backfill_db_value(chunk.iloc[row_index].get(column, ""))
            for column in SERIES_BACKFILL_DB_COLUMNS
        )
        new_values = tuple(
            normalize_series_backfill_db_value(filled_chunk.iloc[row_index].get(column, ""))
            for column in SERIES_BACKFILL_DB_COLUMNS
        )
        if new_values != current_values:
            updates.append((*new_values, int(row_id)))
    return updates


def backfill_series_fields_in_database_in_place(chunk_rows=100000):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(DB_PATH)

    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.execute("PRAGMA busy_timeout = 60000")
    try:
        conn.execute("PRAGMA journal_mode = WAL")
    except sqlite3.DatabaseError:
        pass

    try:
        min_rowid, max_rowid = conn.execute("SELECT MIN(rowid), MAX(rowid) FROM components").fetchone()
        if min_rowid is None or max_rowid is None:
            raise RuntimeError("database contains no component rows")

        batch_size = max(int(chunk_rows), 1000)
        update_sql = (
            'UPDATE components SET "系列" = ?, "系列说明" = ?, "特殊用途" = ? '
            'WHERE rowid = ?'
        )
        updated_rows = 0
        batch_start = int(min_rowid)
        final_rowid = int(max_rowid)
        while batch_start <= final_rowid:
            batch_end = batch_start + batch_size - 1
            chunk = pd.read_sql_query(
                'SELECT rowid AS "__rowid__", * FROM components WHERE rowid BETWEEN ? AND ?',
                conn,
                params=[batch_start, batch_end],
            )
            if chunk is not None and not chunk.empty:
                updates = build_series_backfill_updates(chunk)
                if updates:
                    conn.executemany(update_sql, updates)
                    conn.commit()
                    updated_rows += len(updates)
            batch_start = batch_end + 1
        return updated_rows
    finally:
        conn.close()


MLCC_DIMENSION_BACKFILL_DB_COLUMNS = ["长度（mm）", "宽度（mm）", "高度（mm）", "尺寸来源"]


def ensure_sqlite_table_columns_exist(conn, table_name, column_definitions):
    if conn is None or clean_text(table_name) == "" or not isinstance(column_definitions, dict):
        return
    try:
        existing_columns = {
            clean_text(row[1])
            for row in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
        }
    except Exception:
        return
    for column_name, column_type in column_definitions.items():
        if clean_text(column_name) == "" or column_name in existing_columns:
            continue
        try:
            conn.execute(
                f'ALTER TABLE "{table_name}" ADD COLUMN "{column_name}" {column_type} DEFAULT \'\''
            )
            try:
                conn.commit()
            except Exception:
                pass
        except Exception:
            pass


def build_mlcc_dimension_backfill_updates(chunk, allow_online_lookup=False):
    if chunk is None or chunk.empty:
        return []
    if "__rowid__" not in chunk.columns:
        raise KeyError("__rowid__")

    original_chunk = chunk.copy()
    row_ids = pd.to_numeric(original_chunk.pop("__rowid__"), errors="coerce")
    filled_chunk = enrich_mlcc_dimension_fields_in_dataframe(original_chunk, allow_online_lookup=allow_online_lookup)
    updates = []
    for row_index, row_id in enumerate(row_ids.tolist()):
        if pd.isna(row_id):
            continue
        current_values = tuple(
            normalize_dimension_mm_value(chunk.iloc[row_index].get(column, "")) if column != "尺寸来源" else clean_text(chunk.iloc[row_index].get(column, ""))
            for column in MLCC_DIMENSION_BACKFILL_DB_COLUMNS
        )
        new_values = tuple(
            normalize_dimension_mm_value(filled_chunk.iloc[row_index].get(column, "")) if column != "尺寸来源" else clean_text(filled_chunk.iloc[row_index].get(column, ""))
            for column in MLCC_DIMENSION_BACKFILL_DB_COLUMNS
        )
        if new_values != current_values:
            updates.append((*new_values, int(row_id)))
    return updates


def backfill_mlcc_dimension_fields_in_database_in_place(chunk_rows=100000, allow_online_lookup=False, rebuild_cache=True):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(DB_PATH)

    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.execute("PRAGMA busy_timeout = 60000")
    try:
        conn.execute("PRAGMA journal_mode = WAL")
    except sqlite3.DatabaseError:
        pass

    try:
        batch_size = max(int(chunk_rows), 1000)
        update_sql = (
            'UPDATE components SET "长度（mm）" = ?, "宽度（mm）" = ?, "高度（mm）" = ?, "尺寸来源" = ? '
            'WHERE rowid = ?'
        )
        updated_rows = 0
        chunk = None
        query = (
            'SELECT rowid AS "__rowid__", * FROM components '
            'WHERE COALESCE("器件类型", "") LIKE "%MLCC%" '
            'AND (COALESCE("长度（mm）", "") = "" OR COALESCE("宽度（mm）", "") = "" OR COALESCE("高度（mm）", "") = "" OR COALESCE("尺寸来源", "") = "")'
        )
        for chunk in pd.read_sql_query(query, conn, chunksize=batch_size):
            if chunk is not None and not chunk.empty:
                updates = build_mlcc_dimension_backfill_updates(chunk, allow_online_lookup=allow_online_lookup)
                if updates:
                    conn.executemany(update_sql, updates)
                    conn.commit()
                    updated_rows += len(updates)
        clear_data_load_caches()
        if rebuild_cache:
            rebuild_prepared_cache_from_database()
        return updated_rows
    finally:
        conn.close()


def backfill_series_fields_in_database(chunk_rows=100000):
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(DB_PATH)

    temp_db_path = build_unique_temp_path(DB_PATH, "series_backfill")

    read_conn = sqlite3.connect(DB_PATH, timeout=60)
    read_conn.execute("PRAGMA busy_timeout = 60000")
    write_conn = sqlite3.connect(temp_db_path, timeout=60)
    write_conn.execute("PRAGMA busy_timeout = 60000")
    wrote_any = False

    try:
        for chunk in pd.read_sql_query("SELECT * FROM components", read_conn, chunksize=int(chunk_rows)):
            if chunk is None or chunk.empty:
                continue
            filled = fill_missing_series_from_model(chunk)
            filled.to_sql(
                "components",
                write_conn,
                if_exists="replace" if not wrote_any else "append",
                index=False,
            )
            wrote_any = True
        if not wrote_any:
            raise RuntimeError("no database rows were written")
        create_database_indexes(write_conn)
    finally:
        read_conn.close()
        write_conn.close()

    try:
        replace_file_atomically(temp_db_path, DB_PATH)
    except PermissionError:
        try:
            if os.path.exists(temp_db_path):
                os.remove(temp_db_path)
        except OSError:
            pass
        backfill_series_fields_in_database_in_place(chunk_rows=chunk_rows)
    else:
        if os.path.exists(temp_db_path):
            try:
                os.remove(temp_db_path)
            except OSError:
                pass
    clear_data_load_caches()
    rebuild_prepared_cache_from_database()


@st.cache_data(ttl=3600)
def _load_prepared_data_cached(cache_signature):
    if prepared_cache_is_current():
        try:
            cached = read_prepared_cache()
            if prepared_dataframe_looks_sane(cached):
                return cached
        except Exception:
            pass
    try:
        rebuild_prepared_cache_from_database()
        cached = read_prepared_cache()
        if prepared_dataframe_looks_sane(cached):
            return cached
    except Exception:
        pass
    prepared = prepare_search_dataframe(load_data())
    try:
        write_prepared_cache(prepared)
    except Exception:
        pass
    return prepared


def load_prepared_data(cache_signature=None):
    if cache_signature is None:
        cache_signature = get_data_cache_signature()
    return _load_prepared_data_cached(cache_signature)


def clear_data_load_caches():
    global SAMSUNG_MLCC_DIMENSION_LOOKUP, SAMSUNG_MLCC_DIMENSION_LOOKUP_SIGNATURE
    global MLCC_LCSC_DIMENSION_CACHE, MLCC_LCSC_DIMENSION_CACHE_SIGNATURE
    MODEL_REVERSE_LOOKUP_CACHE.clear()
    MLCC_REFERENCE_LOOKUP_CACHE.clear()
    clear_session_query_caches()
    SAMSUNG_MLCC_DIMENSION_LOOKUP = None
    SAMSUNG_MLCC_DIMENSION_LOOKUP_SIGNATURE = None
    MLCC_LCSC_DIMENSION_CACHE = None
    MLCC_LCSC_DIMENSION_CACHE_SIGNATURE = None
    for cached_func in (_load_data_cached, _load_prepared_data_cached):
        try:
            cached_func.clear()
        except Exception:
            continue


def maybe_update_database(force=False, min_interval_sec=60):
    startup_trace(f"maybe_update_database:enter force={force}")
    if force:
        update_database(force=True)
        startup_trace("maybe_update_database:forced-done")
        return
    # Keep app startup responsive: if a usable database already exists,
    # don't block the initial page render with a full source refresh.
    if database_has_component_rows():
        startup_trace("maybe_update_database:already-has-db")
        return
    try:
        now = time.time()
        last_check = float(st.session_state.get("_db_refresh_checked_at", 0.0))
        if now - last_check < float(min_interval_sec):
            startup_trace("maybe_update_database:skip-throttle")
            return
    except Exception:
        pass
    update_database(force=False)
    startup_trace("maybe_update_database:update-done")
    try:
        st.session_state["_db_refresh_checked_at"] = time.time()
    except Exception:
        pass


REGRESSION_CASE_COLUMNS = [
    "case_id", "query", "expected_mode", "expected_size", "expected_material",
    "expected_value", "expected_unit", "expected_tolerance", "expected_voltage",
    "min_match_count", "notes"
]

REGRESSION_MODE_MAP = {
    "料号": "part",
    "料号片段": "part_partial",
    "规格": "spec",
    "规格不足": "spec_insufficient",
    "无法识别": "unrecognized",
    "铝电解电容": "electrolytic",
    "薄膜电容": "film",
    "贴片电阻": "resistor",
    "厚膜电阻": "resistor",
    "薄膜电阻": "resistor",
    "合金电阻": "resistor",
    "碳膜电阻": "resistor",
    "热敏电阻": "thermistor",
    "压敏电阻": "varistor",
    "引线型压敏电阻": "varistor",
    "贴片压敏电阻": "varistor",
}

REGRESSION_MODE_LABELS = {
    "part": "料号",
    "part_partial": "料号片段",
    "spec": "规格",
    "spec_insufficient": "规格不足",
    "unrecognized": "无法识别",
    "electrolytic": "铝电解电容",
    "film": "薄膜电容",
    "resistor": "贴片电阻",
    "thermistor": "热敏电阻",
    "varistor": "压敏电阻",
}


@st.cache_data(ttl=60)
def load_regression_cases():
    if not os.path.exists(REGRESSION_CASES_PATH):
        return pd.DataFrame(columns=REGRESSION_CASE_COLUMNS)
    try:
        cases = pd.read_csv(REGRESSION_CASES_PATH, encoding="utf-8-sig")
    except Exception:
        return pd.DataFrame(columns=REGRESSION_CASE_COLUMNS)

    for col in REGRESSION_CASE_COLUMNS:
        if col not in cases.columns:
            cases[col] = ""
    cases = cases[REGRESSION_CASE_COLUMNS].copy()
    cases = cases.fillna("")
    return cases


def run_query_match(df, mode, spec):
    if spec is None:
        return pd.DataFrame()
    spec_type = infer_spec_component_type(spec)
    if spec_type in OTHER_PASSIVE_TYPES:
        return match_other_passive_spec(df, spec)
    if mode == "料号":
        return match_by_spec(df, spec)
    if mode in {"料号片段", "规格"}:
        return match_by_partial_spec(df, spec)
    if is_other_passive_mode(mode):
        return match_other_passive_spec(df, spec)
    return pd.DataFrame()


def add_path_signature_fields(signature, label, path):
    path_signature = get_path_signature(path)
    signature[f"{label}_path"] = path_signature.get("path", os.path.abspath(path))
    signature[f"{label}_exists"] = bool(path_signature.get("exists", False))
    if path_signature.get("exists", False):
        signature[f"{label}_mtime"] = path_signature.get("mtime", 0)
        signature[f"{label}_size"] = path_signature.get("size", 0)
    return signature


def build_data_cache_signature():
    signature = dict(get_database_signature())
    bundle_signature = get_public_bundle_signature_for_cache()
    if bundle_signature:
        signature["bundle_signature"] = bundle_signature
    add_path_signature_fields(signature, "prepared_cache", PREPARED_CACHE_PATH)
    add_path_signature_fields(signature, "prepared_meta", PREPARED_CACHE_META_PATH)
    return signature


def get_data_cache_signature():
    try:
        return json.dumps(build_data_cache_signature(), sort_keys=True, ensure_ascii=True)
    except Exception:
        return "db-unknown"


def get_raw_data_cache_signature():
    try:
        return json.dumps(get_database_signature(), sort_keys=True, ensure_ascii=True)
    except Exception:
        return "db-unknown"


def get_query_cache_signature():
    try:
        signature = build_data_cache_signature()
        signature["public_code_stamp"] = PUBLIC_CODE_STAMP
        for path_key, path_value in (
            ("search_db", SEARCH_DB_PATH),
            ("mlcc_lcsc_dimension_cache", MLCC_LCSC_DIMENSION_CACHE_PATH),
            ("pdc_findchips_cache", os.path.join(BASE_DIR, "cache", "pdc_findchips_cache.json")),
        ):
            add_path_signature_fields(signature, path_key, path_value)
        signature["search_index_schema_version"] = SEARCH_INDEX_SCHEMA_VERSION
        return json.dumps(signature, sort_keys=True, ensure_ascii=True)
    except Exception:
        return "db-unknown"


def get_session_query_cache():
    try:
        cache = st.session_state.get("_query_result_cache")
        if cache is None:
            cache = {}
            st.session_state["_query_result_cache"] = cache
        return cache
    except Exception:
        return {}


def get_session_query_dataframe_cache():
    try:
        cache = st.session_state.get("_query_dataframe_cache")
        if cache is None:
            cache = {}
            st.session_state["_query_dataframe_cache"] = cache
        return cache
    except Exception:
        return {}


def clear_session_query_caches():
    try:
        for key in ("_query_result_cache", "_query_dataframe_cache"):
            cache = st.session_state.get(key)
            if isinstance(cache, dict):
                cache.clear()
            elif key in st.session_state:
                try:
                    del st.session_state[key]
                except Exception:
                    pass
    except Exception:
        pass


def store_session_query_dataframe_cache(cache_key, frame, limit=96):
    if clean_text(cache_key) == "" or not isinstance(frame, pd.DataFrame) or frame.empty:
        return
    try:
        cache = get_session_query_dataframe_cache()
        if cache_key in cache:
            cache.pop(cache_key, None)
        cache[cache_key] = frame.copy()
        while len(cache) > max(int(limit), 1):
            oldest_key = next(iter(cache))
            if oldest_key == cache_key and len(cache) == 1:
                break
            cache.pop(oldest_key, None)
    except Exception:
        pass


def serialize_spec_for_cache(spec):
    if spec is None:
        return ""
    keys = [
        "器件类型",
        "品牌",
        "型号",
        "尺寸（inch）",
        "材质（介质）",
        "容值_pf",
        "容值",
        "容值单位",
        "容值误差",
        "耐压（V）",
        "工作温度",
        "寿命（h）",
        "安装方式",
        "特殊用途",
        "_resistance_ohm",
        "_body_size",
        "_pitch",
        "_safety_class",
        "_param_count",
        "_core_param_count",
        "规格摘要",
    ]
    parts = []
    for key in keys:
        value = spec.get(key, "")
        parts.append(f"{key}={clean_text(value)}")
    return "|".join(parts)


def make_query_cache_key(query_text, mode, spec=None):
    payload = "|".join([
        str(QUERY_RESULT_CACHE_VERSION),
        get_query_cache_signature(),
        clean_text(mode),
        clean_text(query_text),
        serialize_spec_for_cache(spec),
    ])
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def cached_run_query_match(df, mode, spec, query_text=""):
    if spec is None:
        return pd.DataFrame()
    cache = get_session_query_cache()
    cache_key = make_query_cache_key(query_text, mode, spec)
    cached = cache.get(cache_key)
    if isinstance(cached, pd.DataFrame):
        return cached.copy()
    matched = run_query_match(df, mode, spec)
    if isinstance(matched, pd.DataFrame):
        cache[cache_key] = matched.copy()
        return matched.copy()
    return matched


def make_table_widget_key(prefix, *parts):
    payload = "|".join([clean_text(part) for part in parts if clean_text(part) != ""])
    digest = hashlib.sha1(payload.encode("utf-8")).hexdigest()[:12]
    safe_prefix = re.sub(r"[^A-Za-z0-9_]+", "_", clean_text(prefix)) or "table"
    return f"{safe_prefix}_{digest}"


def build_regression_case_result(df, case_row):
    query = clean_text(case_row.get("query", ""))
    mode, spec = detect_query_mode_and_spec(df, query)
    mode_code = REGRESSION_MODE_MAP.get(mode, clean_text(mode).lower())
    matched = cached_run_query_match(df, mode, spec, query_text=query)

    actual_size = clean_size(spec.get("尺寸（inch）", "")) if spec else ""
    actual_material = clean_material(spec.get("材质（介质）", "")) if spec else ""
    actual_value, actual_unit = "", ""
    if spec:
        component_type = infer_spec_component_type(spec)
        resistance_ohm = spec.get("_resistance_ohm", None)
        if component_type in ALL_RESISTOR_TYPES and resistance_ohm is not None:
            actual_value, actual_unit = ohm_to_value_unit(resistance_ohm)
        else:
            actual_value, actual_unit = pf_to_value_unit(spec.get("容值_pf", None))
    actual_tol = clean_tol_for_display(spec.get("容值误差", "")) if spec else ""
    actual_volt = clean_voltage(spec.get("耐压（V）", "")) if spec else ""
    match_count = int(len(matched)) if isinstance(matched, pd.DataFrame) else 0

    checks = []

    expected_mode = clean_text(case_row.get("expected_mode", "")).lower()
    if expected_mode != "":
        checks.append(("模式", mode_code == expected_mode, REGRESSION_MODE_LABELS.get(expected_mode, expected_mode), REGRESSION_MODE_LABELS.get(mode_code, mode)))

    expected_size = clean_size(case_row.get("expected_size", ""))
    if expected_size != "":
        checks.append(("尺寸", actual_size == expected_size, expected_size, actual_size))

    expected_material = clean_material(case_row.get("expected_material", ""))
    if expected_material != "":
        checks.append(("材质", actual_material == expected_material, expected_material, actual_material))

    expected_value = clean_text(case_row.get("expected_value", ""))
    if expected_value != "":
        actual_value_text = clean_text(actual_value)
        value_ok = actual_value_text == expected_value
        try:
            value_ok = abs(float(actual_value_text) - float(expected_value)) <= 1e-9
        except Exception:
            pass
        checks.append(("容值", value_ok, expected_value, actual_value))

    expected_unit = clean_text(case_row.get("expected_unit", "")).upper()
    if expected_unit != "":
        checks.append(("容值单位", clean_text(actual_unit).upper() == expected_unit, expected_unit, actual_unit))

    expected_tol = clean_text(case_row.get("expected_tolerance", ""))
    if expected_tol != "":
        checks.append(("容值误差", tolerance_equal(expected_tol, spec.get("容值误差", "") if spec else ""), clean_tol_for_display(expected_tol), actual_tol))

    expected_volt = clean_voltage(case_row.get("expected_voltage", ""))
    if expected_volt != "":
        checks.append(("耐压", actual_volt == expected_volt, voltage_display(expected_volt), voltage_display(actual_volt)))

    min_match_raw = clean_text(case_row.get("min_match_count", ""))
    min_match_count = 0
    if min_match_raw != "":
        try:
            min_match_count = int(float(min_match_raw))
        except Exception:
            min_match_count = 0
        checks.append(("匹配数量", match_count >= min_match_count, f">={min_match_count}", str(match_count)))

    failed_items = [f"{name} 期望 {expected}，实际 {actual}" for name, ok, expected, actual in checks if not ok]
    status = "通过" if not failed_items else "失败"

    return {
        "样本ID": clean_text(case_row.get("case_id", "")),
        "查询内容": query,
        "状态": status,
        "期望模式": REGRESSION_MODE_LABELS.get(expected_mode, expected_mode),
        "实际模式": REGRESSION_MODE_LABELS.get(mode_code, mode),
        "实际尺寸": actual_size,
        "实际材质": actual_material,
        "实际容值": actual_value,
        "实际单位": actual_unit,
        "实际容差": actual_tol,
        "实际耐压": voltage_display(actual_volt),
        "匹配数量": match_count,
        "失败原因": "；".join(failed_items),
        "备注": clean_text(case_row.get("notes", "")),
    }


def run_regression_suite(df, cases_df):
    if df.empty or cases_df.empty:
        return pd.DataFrame()
    rows = [build_regression_case_result(df, row) for _, row in cases_df.iterrows()]
    return pd.DataFrame(rows)


def parser_can_enrich_model(model):
    parsed = reverse_spec_partial(model)
    return parsed is not None and parsed.get("_param_count", 0) >= 3


def build_search_text_series(frame, columns):
    if frame is None:
        return pd.Series(dtype="string")
    if getattr(frame, "empty", False):
        return pd.Series([""] * len(frame), index=frame.index, dtype="string")
    available = [col for col in columns if col in frame.columns]
    if not available:
        return pd.Series([""] * len(frame), index=frame.index, dtype="string")
    text_frame = frame.loc[:, available].astype("string").fillna("")
    if len(available) == 1:
        return text_frame.iloc[:, 0].str.replace(r"\s+", " ", regex=True).str.strip()
    return text_frame.agg(" ".join, axis=1).str.replace(r"\s+", " ", regex=True).str.strip()


def prepare_search_dataframe(df):
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    if all(col in df.columns for col in PREPARED_SEARCH_REQUIRED_COLUMNS):
        return df
    work = df.copy()
    work = apply_model_rule_overrides_to_dataframe(work, override_conflicts=True)
    if "_model_clean" not in work.columns:
        work["_model_clean"] = work["型号"].astype(str).apply(clean_model)
    if "_size" not in work.columns:
        work["_size"] = work["尺寸（inch）"].astype(str).apply(clean_size)
    if "_mat" not in work.columns:
        work["_mat"] = work["材质（介质）"].astype(str).apply(clean_material)
    if "_tol" not in work.columns:
        work["_tol"] = work["容值误差"].astype(str).apply(clean_tol_for_match)
    if "_volt" not in work.columns:
        work["_volt"] = work["耐压（V）"].astype(str).apply(clean_voltage)
    if "_pf" not in work.columns:
        work["_pf"] = pd.to_numeric(work["容值_pf"], errors="coerce")
    if "_tol_kind" not in work.columns or "_tol_num" not in work.columns:
        tol_keys = work["_tol"].apply(tolerance_sort_key)
        if "_tol_kind" not in work.columns:
            work["_tol_kind"] = tol_keys.apply(lambda item: item[0])
        if "_tol_num" not in work.columns:
            work["_tol_num"] = pd.to_numeric(tol_keys.apply(lambda item: item[1]), errors="coerce")
    if "_volt_num" not in work.columns:
        work["_volt_num"] = pd.to_numeric(work["_volt"], errors="coerce")
    if "_component_type" not in work.columns:
        work["_component_type"] = work.apply(infer_db_component_type, axis=1)
    if "_model_rule_authority" not in work.columns:
        work["_model_rule_authority"] = ""
    work = fill_missing_series_from_model(work)
    work = enrich_mlcc_dimension_fields_in_dataframe(work, allow_online_lookup=False)
    if "系列说明" not in work.columns:
        work["系列说明"] = ""
    if "特殊用途" not in work.columns:
        work["特殊用途"] = ""
    if "_mlcc_series_class" not in work.columns:
        work["_mlcc_series_class"] = ""
    else:
        work["_mlcc_series_class"] = work["_mlcc_series_class"].astype(str).apply(normalize_mlcc_series_class)
    mlcc_mask = work["_component_type"].astype(str).apply(normalize_component_type).eq("MLCC")
    if mlcc_mask.any():
        mlcc_rows = work.loc[mlcc_mask, ["品牌", "型号", "系列", "系列说明", "特殊用途"]].copy()
        mlcc_profiles = mlcc_rows.apply(
            lambda row: resolve_mlcc_series_profile(
                brand=row.get("品牌", ""),
                model=row.get("型号", ""),
                series=row.get("系列", ""),
                series_desc=row.get("系列说明", ""),
                special_use=row.get("特殊用途", ""),
            ),
            axis=1,
        )
        mlcc_profile_df = pd.DataFrame(list(mlcc_profiles), index=mlcc_profiles.index)
        current_series = work.loc[mlcc_mask, "系列"].astype(str).apply(clean_text)
        profile_series = mlcc_profile_df["系列"].astype(str).apply(clean_text)
        series_replace_mask = current_series.combine(profile_series, mlcc_series_should_replace)
        if series_replace_mask.any():
            replace_idx = series_replace_mask[series_replace_mask].index
            work.loc[replace_idx, "系列"] = profile_series.loc[replace_idx]
        for col in ["系列说明", "特殊用途"]:
            blank_mask = work.loc[mlcc_mask, col].astype(str).apply(clean_text).eq("")
            if blank_mask.any():
                blank_idx = blank_mask[blank_mask].index
                fill_values = mlcc_profile_df.loc[blank_idx, col].astype(str).apply(clean_text)
                valid_idx = fill_values[fill_values.ne("")].index
                if len(valid_idx) > 0:
                    work.loc[valid_idx, col] = fill_values.loc[valid_idx]
        work.loc[mlcc_mask, "_mlcc_series_class"] = [
            normalize_mlcc_series_class(
                "/".join(
                    part
                    for part in [
                        clean_text(work.at[idx, "_mlcc_series_class"]),
                        clean_text(work.at[idx, "系列说明"]),
                        clean_text(work.at[idx, "特殊用途"]),
                        clean_text(mlcc_profile_df.at[idx, "_mlcc_series_class"]) if idx in mlcc_profile_df.index else "",
                    ]
                    if clean_text(part) != ""
                )
            )
            for idx in work.loc[mlcc_mask].index
        ]
    if "_unit_upper" not in work.columns:
        if "容值单位" in work.columns:
            work["_unit_upper"] = work["容值单位"].apply(lambda value: normalize_search_sidecar_value(value).upper() if normalize_search_sidecar_value(value) is not None else None)
        else:
            work["_unit_upper"] = None
    else:
        work["_unit_upper"] = work["_unit_upper"].apply(lambda value: normalize_search_sidecar_value(value).upper() if normalize_search_sidecar_value(value) is not None else None)
    if "_value_num" not in work.columns:
        if "容值" in work.columns:
            work["_value_num"] = pd.to_numeric(work["容值"], errors="coerce")
        else:
            work["_value_num"] = None
    else:
        work["_value_num"] = pd.to_numeric(work["_value_num"], errors="coerce")
    if "_res_ohm" not in work.columns:
        work["_res_ohm"] = pd.Series([None] * len(work), index=work.index, dtype="object")
    if "_resistance_ohm" in work.columns:
        parsed_res = pd.to_numeric(work["_resistance_ohm"], errors="coerce")
        if work["_res_ohm"].dtype != "object":
            work["_res_ohm"] = work["_res_ohm"].astype("object")
        rule_res_mask = work["_model_rule_authority"].astype(str).apply(clean_text).ne("") & parsed_res.notna()
        work.loc[rule_res_mask, "_res_ohm"] = parsed_res[rule_res_mask]
        blank_res_mask = work["_res_ohm"].astype("string").fillna("").str.strip().eq("")
        work.loc[blank_res_mask & parsed_res.notna(), "_res_ohm"] = parsed_res[blank_res_mask & parsed_res.notna()]

    search_text_columns = ["器件类型", "品牌", "型号", "系列", "系列说明", "安装方式", "封装代码", "尺寸（inch）", "尺寸（mm）", "长度（mm）", "宽度（mm）", "高度（mm）", "材质（介质）", "规格摘要", "容值", "容值单位", "特殊用途", "备注1", "备注2", "备注3"]
    resistor_text_columns = ["器件类型", "品牌", "型号", "系列", "系列说明", "安装方式", "封装代码", "尺寸（inch）", "尺寸（mm）", "长度（mm）", "宽度（mm）", "高度（mm）", "规格摘要", "材质（介质）", "容值", "容值单位", "特殊用途", "备注1", "备注2", "备注3"]
    needs_search_text = any(
        col not in work.columns
        for col in [
            "_power", "_power_watt", "_body_size", "_pitch", "_safety_class", "_varistor_voltage", "_disc_size",
            "_temp_low", "_temp_high", "_life_hours_num", "_mount_style", "_special_use_norm",
        ]
    )
    search_text = build_search_text_series(work, search_text_columns) if needs_search_text else None
    if "_power" not in work.columns:
        work["_power"] = search_text.apply(find_power_in_text).apply(clean_text)
    if "_power_watt" not in work.columns:
        work["_power_watt"] = work["_power"].apply(parse_power_to_watts) if "_power" in work.columns else pd.Series([None] * len(work), index=work.index, dtype="object")
    if "_body_size" not in work.columns:
        work["_body_size"] = search_text.apply(extract_body_size_from_text)
    if "_pitch" not in work.columns:
        work["_pitch"] = search_text.apply(extract_pitch_from_text)
    if "_safety_class" not in work.columns:
        work["_safety_class"] = search_text.apply(find_safety_class)
    if "_varistor_voltage" not in work.columns:
        work["_varistor_voltage"] = search_text.apply(find_varistor_voltage_in_text)
    if "_disc_size" not in work.columns:
        work["_disc_size"] = search_text.apply(find_disc_size_code)
    if "_temp_low" not in work.columns or "_temp_high" not in work.columns:
        temp_source = work["工作温度"].astype(str) if "工作温度" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
        if search_text is not None:
            temp_source = temp_source.where(temp_source.astype(str).apply(clean_text).ne(""), search_text.apply(extract_working_temperature_from_text))
        temp_bounds = temp_source.apply(working_temperature_bounds)
        if "_temp_low" not in work.columns:
            work["_temp_low"] = pd.to_numeric(temp_bounds.apply(lambda item: item[0]), errors="coerce")
        if "_temp_high" not in work.columns:
            work["_temp_high"] = pd.to_numeric(temp_bounds.apply(lambda item: item[1]), errors="coerce")
    else:
        work["_temp_low"] = pd.to_numeric(work["_temp_low"], errors="coerce")
        work["_temp_high"] = pd.to_numeric(work["_temp_high"], errors="coerce")
    if "_life_hours_num" not in work.columns:
        life_source = work["寿命（h）"].astype(str) if "寿命（h）" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
        if search_text is not None:
            life_source = life_source.where(life_source.astype(str).apply(clean_text).ne(""), search_text.apply(parse_life_hours_from_text))
        work["_life_hours_num"] = pd.to_numeric(life_source.apply(normalize_life_hours_value), errors="coerce")
    else:
        work["_life_hours_num"] = pd.to_numeric(work["_life_hours_num"], errors="coerce")
    if "_mount_style" not in work.columns:
        mount_series = work["安装方式"].astype(str) if "安装方式" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
        package_series = work["封装代码"].astype(str) if "封装代码" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
        work["_mount_style"] = pd.Series(
            [normalize_mounting_style(mount, package) for mount, package in zip(mount_series.tolist(), package_series.tolist())],
            index=work.index,
            dtype="object",
        )
        if search_text is not None:
            blank_mount = work["_mount_style"].astype(str).apply(clean_text).eq("")
            work.loc[blank_mount, "_mount_style"] = search_text.loc[blank_mount].apply(extract_mounting_style_from_text)
    else:
        work["_mount_style"] = work["_mount_style"].astype(str).apply(normalize_mounting_style)
    if "_special_use_norm" not in work.columns:
        special_source = work["特殊用途"].astype(str) if "特殊用途" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="string")
        work["_special_use_norm"] = special_source.apply(normalize_special_use)
        if search_text is not None:
            blank_special = work["_special_use_norm"].astype(str).apply(clean_text).eq("")
            work.loc[blank_special, "_special_use_norm"] = search_text.loc[blank_special].apply(extract_special_use_from_text)
    else:
        work["_special_use_norm"] = work["_special_use_norm"].astype(str).apply(normalize_special_use)
    resistor_like_mask = work["_component_type"].isin(RESISTOR_COMPONENT_TYPES | {"热敏电阻"})
    if resistor_like_mask.any():
        res_ohm_blank_mask = work.loc[resistor_like_mask, "_res_ohm"].astype("string").fillna("").str.strip().eq("")
        if res_ohm_blank_mask.any():
            resistor_index = res_ohm_blank_mask[res_ohm_blank_mask].index
            if work["_res_ohm"].dtype != "object":
                work["_res_ohm"] = work["_res_ohm"].astype("object")
            resistor_text = build_search_text_series(work.loc[resistor_index], resistor_text_columns)
            work.loc[resistor_index, "_res_ohm"] = resistor_text.apply(find_resistance_in_text)
    return work


def compatible_component_types_for_search(target_type):
    if target_type in VARISTOR_COMPONENT_TYPES:
        if target_type == "压敏电阻":
            return sorted(VARISTOR_COMPONENT_TYPES)
        return [target_type, "压敏电阻"]
    if target_type in RESISTOR_COMPONENT_TYPES:
        if target_type == "贴片电阻":
            return sorted(RESISTOR_COMPONENT_TYPES)
        return [target_type, "贴片电阻"]
    if target_type in INDUCTOR_COMPONENT_TYPES:
        if target_type in {"功率电感", "射频电感"}:
            return ["功率电感", "射频电感"]
        return [target_type]
    if target_type:
        return [target_type]
    return []


def filter_base_by_candidate_pairs(base, candidate_pairs):
    if base is None or base.empty:
        return base
    if not candidate_pairs:
        return base.iloc[0:0]
    pair_df = pd.DataFrame(candidate_pairs, columns=["品牌", "型号"]).drop_duplicates()
    return base.merge(pair_df, on=["品牌", "型号"], how="inner")


def chunk_items(items, chunk_size):
    size = max(int(chunk_size), 1)
    for start in range(0, len(items), size):
        yield items[start:start + size]


def safe_concat_dataframes(frames, **kwargs):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r".*DataFrame concatenation with empty or all-NA entries.*",
            category=FutureWarning,
        )
        return pd.concat(frames, **kwargs)


def concat_component_frames(frames):
    valid = [
        frame for frame in frames
        if isinstance(frame, pd.DataFrame)
        and not frame.empty
        and not frame.dropna(how="all").empty
    ]
    if not valid:
        return pd.DataFrame()
    combined = safe_concat_dataframes(valid, ignore_index=True, sort=False)
    if combined.empty:
        return combined
    if {"品牌", "型号"}.issubset(combined.columns):
        combined["品牌"] = combined["品牌"].astype(str).apply(clean_text)
        combined["型号"] = combined["型号"].astype(str).apply(clean_text)
        combined = prioritize_component_rows_for_lookup(combined)
        combined = combined.drop_duplicates(subset=["品牌", "型号"], keep="first")
    return combined


def format_sidecar_numeric_display(value):
    try:
        number = float(value)
    except Exception:
        return ""
    if not math.isfinite(number):
        return ""
    rounded = round(number)
    if abs(number - rounded) < 1e-9:
        return str(int(rounded))
    return f"{number:.12g}"


def build_working_temperature_from_sidecar_bounds(low_value, high_value):
    low_text = format_sidecar_numeric_display(low_value)
    high_text = format_sidecar_numeric_display(high_value)
    if low_text != "" and high_text != "":
        return normalize_working_temperature_text(f"{low_text}~{high_text}℃")
    if high_text != "":
        return normalize_working_temperature_text(f"{high_text}℃")
    if low_text != "":
        return normalize_working_temperature_text(f"{low_text}℃")
    return ""


RESISTOR_POWER_BY_SIZE = {
    "01005": "1/32W",
    "0201": "1/20W",
    "0402": "1/16W",
    "0603": "1/10W",
    "0805": "1/8W",
    "1206": "1/4W",
    "1210": "1/2W",
    "2010": "3/4W",
    "2512": "1W",
    "3225": "2W",
}

RESISTOR_MAX_WORKING_VOLTAGE_BY_SIZE = {
    "01005": "15",
    "0201": "25",
    "0402": "50",
    "0603": "75",
    "0805": "150",
    "1206": "200",
    "1210": "200",
    "2010": "400",
    "2512": "500",
    "3225": "500",
}

RESISTOR_DEFAULT_WORKING_TEMPERATURE = "-55~155℃"


def infer_display_fallback_fields_from_record(record):
    if record is None:
        return {}

    component_type = normalize_component_type(record.get("器件类型", "")) or infer_db_component_type(record)
    row_text = build_component_context_text(record)
    size_code = clean_size(record.get("尺寸（inch）", "")) or infer_resistor_size_from_model(record.get("型号", ""))
    result = {}

    power_text = clean_text(record.get("功率", "")) or clean_text(record.get("_power", ""))
    if power_text == "":
        power_watt = record.get("_power_watt", "")
        if power_watt is not None and not pd.isna(power_watt):
            power_text = f"{format_sidecar_numeric_display(power_watt)}W"
    if power_text == "":
        power_text = find_power_in_text(row_text)
    if power_text == "" and component_type in RESISTOR_COMPONENT_TYPES:
        power_text = RESISTOR_POWER_BY_SIZE.get(size_code, "")
    if power_text != "":
        result["功率"] = format_power_display(power_text)

    voltage_text = clean_text(record.get("耐压（V）", "")) or clean_text(record.get("最高工作电压", ""))
    if voltage_text == "" and component_type in VARISTOR_COMPONENT_TYPES:
        voltage_text = clean_text(record.get("_varistor_voltage", ""))
    if voltage_text == "":
        volt_num = record.get("_volt_num", "")
        if volt_num is not None and not pd.isna(volt_num):
            voltage_text = format_sidecar_numeric_display(volt_num)
    if voltage_text == "":
        voltage_text = parse_voltage_from_text(row_text)
    if voltage_text == "" and component_type in RESISTOR_COMPONENT_TYPES:
        voltage_text = RESISTOR_MAX_WORKING_VOLTAGE_BY_SIZE.get(size_code, "")
    if voltage_text != "":
        result["耐压（V）"] = clean_voltage(voltage_text)

    work_temp_text = normalize_working_temperature_text(record.get("工作温度", ""))
    if work_temp_text == "":
        work_temp_text = extract_working_temperature_from_text(row_text)
    if work_temp_text == "" and component_type in RESISTOR_COMPONENT_TYPES:
        work_temp_text = RESISTOR_DEFAULT_WORKING_TEMPERATURE
    if work_temp_text != "":
        result["工作温度"] = work_temp_text

    special_use_text = clean_text(record.get("特殊用途", "")) or clean_text(record.get("_special_use_norm", ""))
    if special_use_text == "":
        special_use_text = extract_special_use_from_text(row_text)
    if special_use_text != "":
        result["特殊用途"] = normalize_special_use(special_use_text)

    mount_text = normalize_mounting_style(record.get("安装方式", ""), record.get("封装代码", ""))
    if mount_text == "":
        mount_text = extract_mounting_style_from_text(row_text)
    if mount_text != "":
        result["安装方式"] = normalize_mounting_style(mount_text, record.get("封装代码", ""))

    status_text = clean_text(record.get("生产状态", "")) or clean_text(record.get("量产状态", ""))
    if status_text == "":
        for candidate in [record.get("备注1", ""), record.get("备注3", ""), row_text]:
            if looks_like_official_status_text(candidate):
                status_text = extract_official_status(candidate)
                if status_text != "":
                    break
    if status_text != "":
        result["生产状态"] = status_text

    return result


def query_search_sidecar_table_by_models(conn, table_name, models):
    model_list = [clean_text(model) for model in (models or []) if clean_text(model) != ""]
    if conn is None or not model_list:
        return pd.DataFrame()
    frames = []
    try:
        for model_chunk in chunk_items(model_list, SEARCH_DB_FETCH_CHUNK):
            placeholders = ",".join(["?"] * len(model_chunk))
            query = f'SELECT * FROM {table_name} WHERE "型号" IN ({placeholders})'
            frames.append(pd.read_sql_query(query, conn, params=model_chunk))
    except Exception:
        return pd.DataFrame()
    return concat_component_frames(frames)


def build_lightweight_component_row_from_search_sidecar(core_row, detail_row=None, include_model_rule=True):
    core_row = core_row or {}
    detail_row = detail_row or {}
    brand = clean_text(core_row.get("品牌", detail_row.get("品牌", "")))
    model = clean_text(core_row.get("型号", detail_row.get("型号", "")))
    if model == "":
        return {}

    component_type = normalize_component_type(
        detail_row.get("_component_type", core_row.get("_component_type", ""))
    )
    record = {
        "品牌": brand,
        "型号": model,
        "器件类型": component_type,
        "系列": "",
        "系列说明": "",
        "尺寸（inch）": clean_size(detail_row.get("_size", "")),
        "尺寸（mm）": clean_text(detail_row.get("_body_size", "")),
        "材质（介质）": clean_material(detail_row.get("_mat", "")),
        "容值误差": clean_tol_for_display(detail_row.get("_tol", "")),
        "耐压（V）": voltage_display(format_sidecar_numeric_display(detail_row.get("_volt_num", ""))),
        "工作温度": build_working_temperature_from_sidecar_bounds(detail_row.get("_temp_low", ""), detail_row.get("_temp_high", "")),
        "寿命（h）": format_life_hours_display(format_sidecar_numeric_display(detail_row.get("_life_hours_num", ""))),
        "安装方式": normalize_mounting_style(detail_row.get("_mount_style", "")),
        "特殊用途": normalize_special_use(detail_row.get("_special_use_norm", "")),
        "_mlcc_series_class": clean_text(detail_row.get("_mlcc_series_class", core_row.get("_mlcc_series_class", ""))),
        "脚距": clean_text(detail_row.get("_pitch", "")),
        "安规": clean_text(detail_row.get("_safety_class", "")),
        "压敏电压": voltage_display(clean_voltage(detail_row.get("_varistor_voltage", ""))),
        "规格": clean_text(detail_row.get("_disc_size", "")) or clean_text(detail_row.get("_body_size", "")),
        "数据来源": "搜索索引轻量回退",
        "容值_pf": None,
        "_resistance_ohm": None,
        "_model_rule_authority": "search_sidecar_light",
    }

    pf_value = pd.to_numeric(pd.Series([detail_row.get("_pf", None)]), errors="coerce").iloc[0]
    res_ohm = pd.to_numeric(pd.Series([detail_row.get("_res_ohm", None)]), errors="coerce").iloc[0]
    value_num = pd.to_numeric(pd.Series([detail_row.get("_value_num", None)]), errors="coerce").iloc[0]
    if pd.notna(pf_value):
        record["容值_pf"] = float(pf_value)
        value, unit = pf_to_value_unit(float(pf_value))
        record["容值"] = value
        record["容值单位"] = unit
    elif pd.notna(res_ohm):
        record["_resistance_ohm"] = float(res_ohm)
        value, unit = ohm_to_library_value_unit(float(res_ohm))
        record["容值"] = value
        record["容值单位"] = unit
    elif pd.notna(value_num):
        record["容值"] = format_sidecar_numeric_display(value_num)
        record["容值单位"] = clean_text(detail_row.get("_unit_upper", "")).upper()

    if include_model_rule:
        parsed_rule = parse_model_rule(model, brand=brand, component_type=component_type)
        if isinstance(parsed_rule, dict) and parsed_rule:
            record = merge_parsed_rule_into_record(record, parsed_rule, override_conflicts=False)
    return record


def load_component_rows_by_brand_model_pairs(candidate_pairs, preferred_component_type=""):
    pairs = [
        (clean_text(brand), clean_text(model))
        for brand, model in (candidate_pairs or [])
        if clean_text(model) != ""
    ]
    if not pairs:
        return pd.DataFrame()
    combined = pd.DataFrame()
    models = sorted({model for _, model in pairs if model != ""})
    if os.path.exists(DB_PATH) and models:
        conn = sqlite3.connect(DB_PATH)
        try:
            frames = []
            for model_chunk in chunk_items(models, SEARCH_DB_FETCH_CHUNK):
                placeholders = ",".join(["?"] * len(model_chunk))
                query = f'SELECT * FROM components WHERE "型号" IN ({placeholders})'
                frames.append(pd.read_sql_query(query, conn, params=model_chunk))
            combined = concat_component_frames(frames)
        except Exception:
            combined = pd.DataFrame()
        finally:
            conn.close()
    if combined.empty:
        combined = load_search_sidecar_rows_by_brand_model_pairs(pairs, preferred_component_type=preferred_component_type)
    seed_rows = load_jianghai_seed_rows() + load_sunlord_mcl_seed_rows()
    if seed_rows:
        combined = concat_component_frames([combined, pd.DataFrame(seed_rows)])
    if combined.empty:
        return combined
    combined = filter_base_by_candidate_pairs(combined, pairs)
    if combined.empty:
        return combined
    return prepare_search_dataframe(combined)


def load_search_sidecar_rows_by_brand_model_pairs(candidate_pairs, preferred_component_type=""):
    pairs = [
        (clean_text(brand), clean_text(model))
        for brand, model in (candidate_pairs or [])
        if clean_text(model) != ""
    ]
    if not pairs:
        return pd.DataFrame()

    required_columns = {"品牌", "型号", "_model_clean", "_component_type"}
    conn = open_search_db_connection(timeout_sec=10)
    try:
        if conn is None or not search_index_can_serve_queries(
            conn,
            required_columns=required_columns,
            table_name=COMPONENTS_SEARCH_CORE_TABLE,
            allow_without_database=True,
        ):
            if os.path.exists(DB_PATH):
                if conn is not None:
                    conn.close()
                rebuild_search_index_from_database_fast()
                conn = open_search_db_connection(timeout_sec=10)
            if conn is None or not search_index_can_serve_queries(
                conn,
                required_columns=required_columns,
                table_name=COMPONENTS_SEARCH_CORE_TABLE,
                allow_without_database=True,
            ):
                prepared = None
                try:
                    prepared = load_prepared_data()
                except Exception:
                    prepared = None
                if prepared is None or prepared.empty:
                    return pd.DataFrame()
                prepared_work = prepare_search_dataframe(prepared)
                if prepared_work.empty or not {"品牌", "型号", "_model_clean"}.issubset(prepared_work.columns):
                    return pd.DataFrame()
                pair_set = set(pairs)
                filtered = prepared_work[
                    prepared_work.apply(
                        lambda row: (
                            clean_text(row.get("品牌", "")),
                            clean_text(row.get("型号", "")),
                        ) in pair_set,
                        axis=1,
                    )
                ]
                if filtered.empty:
                    return pd.DataFrame()
                return prepare_search_dataframe(filtered)

        models = sorted({model for _, model in pairs if model != ""})
        pair_set = set(pairs)
        preferred_table = search_index_table_for_component_type(clean_text(preferred_component_type))

        if preferred_table and preferred_table != COMPONENTS_SEARCH_CORE_TABLE:
            detail_df = query_search_sidecar_table_by_models(conn, preferred_table, models)
            if not detail_df.empty and {"品牌", "型号"}.issubset(detail_df.columns):
                detail_rows = [
                    row
                    for row in detail_df.to_dict("records")
                    if (
                        clean_text(row.get("品牌", "")),
                        clean_text(row.get("型号", "")),
                    ) in pair_set
                ]
                if detail_rows:
                    records = []
                    for detail_row in detail_rows:
                        record = build_lightweight_component_row_from_search_sidecar(
                            detail_row,
                            detail_row,
                            include_model_rule=False,
                        )
                        if record:
                            records.append(record)
                    if records:
                        return prepare_search_dataframe(pd.DataFrame(records))

        core_df = query_search_sidecar_table_by_models(conn, COMPONENTS_SEARCH_CORE_TABLE, models)
        if core_df.empty:
            return pd.DataFrame()
        if not {"品牌", "型号"}.issubset(core_df.columns):
            return pd.DataFrame()
        core_rows = [
            row
            for row in core_df.to_dict("records")
            if (
                clean_text(row.get("品牌", "")),
                clean_text(row.get("型号", "")),
            ) in pair_set
        ]
        if not core_rows:
            return pd.DataFrame()

        fallback_table_order = [
            COMPONENTS_SEARCH_CAPACITOR_TABLE,
            COMPONENTS_SEARCH_RESISTOR_TABLE,
            COMPONENTS_SEARCH_VALUE_TABLE,
            COMPONENTS_SEARCH_VARISTOR_TABLE,
        ]
        table_scan_order = []
        if preferred_table and preferred_table != COMPONENTS_SEARCH_CORE_TABLE:
            table_scan_order.append(preferred_table)
        table_scan_order.extend([table_name for table_name in fallback_table_order if table_name not in table_scan_order])

        side_maps = {}
        for table_name in table_scan_order:
            table_df = query_search_sidecar_table_by_models(conn, table_name, models)
            if table_df.empty or not {"品牌", "型号"}.issubset(table_df.columns):
                side_maps[table_name] = {}
                continue
            side_maps[table_name] = {
                (clean_text(row.get("品牌", "")), clean_text(row.get("型号", ""))): row
                for row in table_df.to_dict("records")
                if (
                    clean_text(row.get("型号", "")) != ""
                    and (
                        clean_text(row.get("品牌", "")),
                        clean_text(row.get("型号", "")),
                    ) in pair_set
                )
            }
    except Exception:
        return pd.DataFrame()
    finally:
        if conn is not None:
            conn.close()

    records = []
    for core_row in core_rows:
        pair_key = (
            clean_text(core_row.get("品牌", "")),
            clean_text(core_row.get("型号", "")),
        )
        detail_row = {}
        row_preferred_table = search_index_table_for_component_type(clean_text(core_row.get("_component_type", "")))
        if row_preferred_table in side_maps:
            detail_row = side_maps[row_preferred_table].get(pair_key, {})
        if not detail_row:
            for table_name in table_scan_order:
                detail_row = side_maps.get(table_name, {}).get(pair_key, {})
                if detail_row:
                    break
        record = build_lightweight_component_row_from_search_sidecar(core_row, detail_row)
        if record:
            records.append(record)

    return prepare_search_dataframe(pd.DataFrame(records)) if records else pd.DataFrame()


def load_component_rows_by_clean_models_map(models):
    model_clean_list = [
        clean_model(model)
        for model in (models or [])
        if clean_model(model) != ""
    ]
    unique_models = list(dict.fromkeys(model_clean_list))
    if not unique_models:
        return {}
    required_columns = {"品牌", "型号", "_model_clean"}
    conn = open_search_db_connection(timeout_sec=10)
    try:
        if conn is None or not search_index_can_serve_queries(
            conn,
            required_columns=required_columns,
            table_name=COMPONENTS_SEARCH_CORE_TABLE,
            allow_without_database=True,
        ):
            if conn is not None:
                conn.close()
            if os.path.exists(DB_PATH):
                rebuild_search_index_from_database_fast()
                conn = open_search_db_connection(timeout_sec=10)
            if conn is None or not search_index_can_serve_queries(
                conn,
                required_columns=required_columns,
                table_name=COMPONENTS_SEARCH_CORE_TABLE,
                allow_without_database=True,
            ):
                prepared = None
                try:
                    prepared = load_prepared_data()
                except Exception:
                    prepared = None
                if prepared is None or prepared.empty:
                    return {model_clean: pd.DataFrame() for model_clean in unique_models}
                prepared_work = prepare_search_dataframe(prepared)
                if prepared_work.empty or "_model_clean" not in prepared_work.columns:
                    return {model_clean: pd.DataFrame() for model_clean in unique_models}
                result_map = {}
                model_series = prepared_work["_model_clean"].astype(str)
                for model_clean in unique_models:
                    result_map[model_clean] = prepared_work[model_series.eq(model_clean)].copy()
                return result_map
        rows = []
        for model_chunk in chunk_items(unique_models, SEARCH_DB_FETCH_CHUNK):
            placeholders = ",".join(["?"] * len(model_chunk))
            chunk_rows = conn.execute(
                f'SELECT DISTINCT "_model_clean", "品牌", "型号" FROM {COMPONENTS_SEARCH_CORE_TABLE} WHERE "_model_clean" IN ({placeholders})',
                model_chunk,
            ).fetchall()
            rows.extend(chunk_rows)
    except Exception:
        return {model_clean: pd.DataFrame() for model_clean in unique_models}
    finally:
        if conn is not None:
            conn.close()

    candidate_pairs = []
    for row in rows:
        row_model_clean = clean_model(row[0] if len(row) >= 1 else "")
        brand = clean_text(row[1] if len(row) >= 2 else "")
        model = clean_text(row[2] if len(row) >= 3 else "")
        if row_model_clean == "" or model == "":
            continue
        pair_key = (brand, model)
        candidate_pairs.append(pair_key)

    result_map = {model_clean: pd.DataFrame() for model_clean in unique_models}
    seed_df = pd.DataFrame(load_jianghai_seed_rows() + load_sunlord_mcl_seed_rows())
    if not seed_df.empty:
        seed_df = prepare_search_dataframe(seed_df)

    work_frames = []
    prepared_exact = load_component_rows_by_brand_model_pairs(candidate_pairs)
    if isinstance(prepared_exact, pd.DataFrame) and not prepared_exact.empty:
        work_frames.append(prepared_exact)
    if isinstance(seed_df, pd.DataFrame) and not seed_df.empty:
        work_frames.append(seed_df)
    if not work_frames:
        return result_map

    work = concat_component_frames(work_frames)
    if "型号" not in work.columns:
        return result_map
    work["_clean_model_key"] = work["型号"].astype(str).apply(clean_model)

    for model_clean in unique_models:
        result_map[model_clean] = work[work["_clean_model_key"].eq(model_clean)].drop(columns=["_clean_model_key"], errors="ignore").copy()
        if result_map[model_clean].empty:
            fallback = build_rule_fallback_row_from_model(model_clean)
            if isinstance(fallback, pd.DataFrame) and not fallback.empty:
                result_map[model_clean] = fallback.copy()
    return result_map


def build_rule_fallback_row_from_model(model, brand=""):
    parsed = parse_model_rule(model, brand=brand, component_type="")
    if not isinstance(parsed, dict) or not parsed:
        return pd.DataFrame()
    fallback_defaults = {
        "品牌": "",
        "型号": clean_model(model),
        "器件类型": "",
        "系列": "",
        "系列说明": "",
        "尺寸（inch）": "",
        "尺寸（mm）": "",
        "长度（mm）": "",
        "宽度（mm）": "",
        "高度（mm）": "",
        "材质（介质）": "",
        "容值": "",
        "容值单位": "",
        "容值_pf": None,
        "容值误差": "",
        "耐压（V）": "",
        "安装方式": "",
        "封装代码": "",
        "特殊用途": "",
        "规格摘要": "",
        "备注1": "",
        "备注2": "",
        "备注3": "",
    }
    fallback_defaults.update(parsed)
    parsed = fallback_defaults
    fallback = pd.DataFrame([parsed])
    try:
        fallback = prepare_search_dataframe(fallback)
    except Exception:
        pass
    return fallback


def load_component_rows_by_clean_model(model):
    model_clean = clean_model(model)
    if model_clean == "":
        return pd.DataFrame()
    return load_component_rows_by_clean_models_map([model]).get(model_clean, pd.DataFrame())


def load_component_rows_by_typed_spec(spec):
    if spec is None:
        return pd.DataFrame()
    component_type = infer_spec_component_type(spec)
    if not os.path.exists(DB_PATH):
        candidate_pairs = fetch_search_candidate_pairs(spec)
        if candidate_pairs is not None:
            if candidate_pairs:
                return load_component_rows_by_brand_model_pairs(candidate_pairs, preferred_component_type=component_type)
            return pd.DataFrame()
        prepared = None
        try:
            prepared = load_prepared_data()
        except Exception:
            prepared = None
        if prepared is not None and not prepared.empty:
            scoped = scope_search_dataframe(prepared, spec)
            if isinstance(scoped, pd.DataFrame) and not scoped.empty:
                return scoped
        return pd.DataFrame()
    supported_types = (
        INDUCTOR_COMPONENT_TYPES
        | TIMING_COMPONENT_TYPES
        | {"薄膜电容", "铝电解电容", "引线型陶瓷电容", "热敏电阻"}
        | VARISTOR_COMPONENT_TYPES
    )
    if component_type not in supported_types:
        return pd.DataFrame()
    size = clean_size(spec.get("尺寸（inch）", ""))
    unit = clean_text(spec.get("容值单位", "")).upper()
    type_values = compatible_component_types_for_search(component_type) or [component_type]
    where_clauses = [f'[器件类型] IN ({",".join(["?"] * len(type_values))})']
    params = list(type_values)
    if size != "" and component_type in (INDUCTOR_COMPONENT_TYPES | TIMING_COMPONENT_TYPES | {"热敏电阻"}):
        where_clauses.append('[尺寸（inch）] = ?')
        params.append(size)
    if unit != "" and component_type in (INDUCTOR_COMPONENT_TYPES | TIMING_COMPONENT_TYPES | {"薄膜电容", "铝电解电容", "引线型陶瓷电容"}):
        where_clauses.append("UPPER(IFNULL([容值单位], '')) = ?")
        params.append(unit)
    query = f'SELECT * FROM [components] WHERE {" AND ".join(where_clauses)}'
    conn = sqlite3.connect(DB_PATH)
    try:
        frame = pd.read_sql_query(query, conn, params=params)
    except Exception:
        return pd.DataFrame()
    finally:
        conn.close()

    if component_type == "铝电解电容":
        seed_rows = pd.DataFrame(load_jianghai_seed_rows() + load_sunlord_mcl_seed_rows())
        if not seed_rows.empty:
            frame = concat_component_frames([frame, seed_rows])

    if frame.empty:
        return frame

    frame = prepare_search_dataframe(frame)
    if frame.empty:
        return frame

    spec_value = clean_text(spec.get("容值", ""))
    spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
    if component_type in (INDUCTOR_COMPONENT_TYPES | TIMING_COMPONENT_TYPES):
        if spec_value != "" and "容值" in frame.columns:
            row_value = frame["容值"].astype(str).apply(clean_text)
            exact_text = row_value.eq(spec_value)
            try:
                target_value = float(spec_value)
                numeric_mask = pd.to_numeric(row_value, errors="coerce").sub(target_value).abs().lt(1e-9)
                same_value = exact_text | numeric_mask.fillna(False)
            except Exception:
                same_value = exact_text
            frame = frame[same_value]
        if not frame.empty and spec_tol != "" and "容值误差" in frame.columns:
            tol_mask = frame["容值误差"].astype(str).apply(clean_tol_for_match).eq(spec_tol)
            if tol_mask.any():
                frame = frame[tol_mask]
        if component_type in TIMING_COMPONENT_TYPES:
            spec_volt = clean_voltage(spec.get("耐压（V）", ""))
            if spec_volt != "" and "_volt_num" in frame.columns:
                same_volt = frame["_volt_num"].notna() & (pd.to_numeric(frame["_volt_num"], errors="coerce") - float(spec_volt)).abs().lt(1e-9)
                if same_volt.any():
                    frame = frame[same_volt]
    elif component_type in {"薄膜电容", "铝电解电容", "引线型陶瓷电容"}:
        spec_pf = spec.get("容值_pf", None)
        if spec_pf is not None and "_pf" in frame.columns:
            pf_mask = frame["_pf"].notna() & ((pd.to_numeric(frame["_pf"], errors="coerce") - float(spec_pf)).abs() < 1e-6)
            frame = frame[pf_mask]
        if not frame.empty and spec_tol != "" and "_tol" in frame.columns:
            tol_mask = frame["_tol"].astype(str).apply(clean_tol_for_match).eq(spec_tol)
            if tol_mask.any():
                frame = frame[tol_mask]
        spec_volt = clean_voltage(spec.get("耐压（V）", ""))
        if not frame.empty and spec_volt != "" and "_volt_num" in frame.columns:
            try:
                spec_volt_num = float(spec_volt)
                volt_mask = frame["_volt_num"].notna() & pd.to_numeric(frame["_volt_num"], errors="coerce").ge(spec_volt_num)
                frame = frame[volt_mask]
            except Exception:
                pass
        if component_type == "薄膜电容":
            spec_material = clean_material(spec.get("材质（介质）", ""))
            if not frame.empty and spec_material != "" and "_mat" in frame.columns:
                mat_mask = frame["_mat"].astype(str).apply(clean_material).eq(spec_material)
                if mat_mask.any():
                    frame = frame[mat_mask]
            spec_safety = clean_text(spec.get("_safety_class", ""))
            if not frame.empty and spec_safety != "" and "_safety_class" in frame.columns:
                safety_mask = frame["_safety_class"].astype(str).apply(clean_text).eq(spec_safety)
                if safety_mask.any():
                    frame = frame[safety_mask]
        if component_type == "引线型陶瓷电容":
            spec_material = clean_material(spec.get("材质（介质）", ""))
            if not frame.empty and spec_material != "" and "_mat" in frame.columns:
                mat_mask = frame["_mat"].astype(str).apply(clean_material).eq(spec_material)
                if mat_mask.any():
                    frame = frame[mat_mask]
        spec_body_size = clean_text(spec.get("_body_size", ""))
        if not frame.empty and spec_body_size != "" and "_body_size" in frame.columns:
            body_mask = frame["_body_size"].astype(str).apply(clean_text).eq(spec_body_size)
            frame = frame[body_mask]
        if component_type == "铝电解电容":
            spec_temp = normalize_working_temperature_text(spec.get("工作温度", ""))
            if not frame.empty and spec_temp != "":
                same_temp = frame["工作温度"].astype(str).apply(normalize_working_temperature_text).eq(spec_temp) if "工作温度" in frame.columns else pd.Series(False, index=frame.index)
                if same_temp.any():
                    frame = frame[same_temp]
                elif "_temp_high" in frame.columns:
                    target_low, target_high = working_temperature_bounds(spec_temp)
                    cover_mask = pd.Series(True, index=frame.index)
                    if target_high is not None:
                        cover_mask &= pd.to_numeric(frame["_temp_high"], errors="coerce").notna() & pd.to_numeric(frame["_temp_high"], errors="coerce").ge(target_high)
                    if target_low is not None:
                        cover_mask &= pd.to_numeric(frame["_temp_low"], errors="coerce").notna() & pd.to_numeric(frame["_temp_low"], errors="coerce").le(target_low)
                    frame = frame[cover_mask]
                else:
                    frame = frame.iloc[0:0]
            spec_life = normalize_life_hours_value(spec.get("寿命（h）", ""))
            if not frame.empty and spec_life != "":
                same_life = frame["寿命（h）"].astype(str).apply(normalize_life_hours_value).eq(spec_life) if "寿命（h）" in frame.columns else pd.Series(False, index=frame.index)
                if same_life.any():
                    frame = frame[same_life]
                elif "_life_hours_num" in frame.columns:
                    life_target = life_hours_to_number(spec_life)
                    life_mask = pd.to_numeric(frame["_life_hours_num"], errors="coerce").notna() & pd.to_numeric(frame["_life_hours_num"], errors="coerce").ge(life_target)
                    frame = frame[life_mask]
                else:
                    frame = frame.iloc[0:0]
            spec_mount = normalize_mounting_style(spec.get("安装方式", ""))
            if not frame.empty and spec_mount != "":
                mount_mask = frame["_mount_style"].astype(str).apply(normalize_mounting_style).eq(spec_mount) if "_mount_style" in frame.columns else (
                    frame["安装方式"].astype(str).apply(normalize_mounting_style).eq(spec_mount) if "安装方式" in frame.columns else pd.Series(False, index=frame.index)
                )
                frame = frame[mount_mask]
            spec_special = normalize_special_use(spec.get("特殊用途", ""))
            if not frame.empty and spec_special != "":
                special_mask = frame["特殊用途"].astype(str).apply(lambda value: special_use_matches(value, spec_special)) if "特殊用途" in frame.columns else (
                    frame["_special_use_norm"].astype(str).apply(lambda value: special_use_matches(value, spec_special)) if "_special_use_norm" in frame.columns else pd.Series(False, index=frame.index)
                )
                frame = frame[special_mask]
        spec_pitch = clean_text(spec.get("_pitch", ""))
        if not frame.empty and spec_pitch != "" and "_pitch" in frame.columns:
            pitch_mask = frame["_pitch"].astype(str).apply(clean_text).eq(spec_pitch)
            frame = frame[pitch_mask]
    elif component_type in VARISTOR_COMPONENT_TYPES:
        spec_varistor_voltage = clean_voltage(spec.get("_varistor_voltage", "")) or clean_voltage(spec.get("耐压（V）", ""))
        if spec_varistor_voltage != "" and "_varistor_voltage" in frame.columns:
            voltage_mask = frame["_varistor_voltage"].astype(str).apply(clean_voltage).eq(spec_varistor_voltage)
            if voltage_mask.any():
                frame = frame[voltage_mask]
        if not frame.empty and spec_tol != "" and "_tol" in frame.columns:
            tol_mask = frame["_tol"].astype(str).apply(clean_tol_for_match).eq(spec_tol)
            if tol_mask.any():
                frame = frame[tol_mask]
        spec_disc_size = clean_text(spec.get("_disc_size", ""))
        if not frame.empty and spec_disc_size != "" and "_disc_size" in frame.columns:
            disc_mask = frame["_disc_size"].astype(str).apply(clean_text).eq(spec_disc_size)
            if disc_mask.any():
                frame = frame[disc_mask]
        spec_pitch = clean_text(spec.get("_pitch", ""))
        if not frame.empty and spec_pitch != "" and "_pitch" in frame.columns:
            pitch_mask = frame["_pitch"].astype(str).apply(clean_text).eq(spec_pitch)
            if pitch_mask.any():
                frame = frame[pitch_mask]
    elif component_type == "热敏电阻":
        resistance_ohm = spec.get("_resistance_ohm", None)
        if resistance_ohm is not None and "_res_ohm" in frame.columns:
            res_mask = frame["_res_ohm"].notna() & ((pd.to_numeric(frame["_res_ohm"], errors="coerce") - float(resistance_ohm)).abs() < 1e-9)
            frame = frame[res_mask]
        if not frame.empty and spec_tol != "" and "_tol" in frame.columns:
            tol_mask = frame["_tol"].astype(str).apply(clean_tol_for_match).eq(spec_tol)
            if tol_mask.any():
                frame = frame[tol_mask]
    return frame


def can_use_fast_search_dataframe(spec):
    if spec is None:
        return False
    component_type = infer_spec_component_type(spec)
    return (
        component_type == "MLCC"
        or component_type in RESISTOR_COMPONENT_TYPES
        or component_type == "热敏电阻"
        or component_type in VARISTOR_COMPONENT_TYPES
        or component_type in {"薄膜电容", "铝电解电容", "引线型陶瓷电容"}
        or component_type in INDUCTOR_COMPONENT_TYPES
        or component_type in TIMING_COMPONENT_TYPES
    )


def load_search_dataframe_for_query(mode, spec, query_text="", exact_part_rows=None):
    if spec is None:
        return None
    frames = []
    used_fast_path = False
    if mode == "料号":
        if isinstance(exact_part_rows, pd.DataFrame):
            part_rows = exact_part_rows
        else:
            part_rows = load_component_rows_by_clean_model(query_text)
        if isinstance(part_rows, pd.DataFrame):
            frames.append(part_rows)
            used_fast_path = True
    component_type = infer_spec_component_type(spec)
    if can_use_fast_search_dataframe(spec):
        candidate_pairs = fetch_search_candidate_pairs(spec)
        if candidate_pairs is not None:
            if len(candidate_pairs) > 0:
                frames.append(load_component_rows_by_brand_model_pairs(candidate_pairs, preferred_component_type=component_type))
                used_fast_path = True
            else:
                return concat_component_frames(frames) if frames else pd.DataFrame()
        else:
            if component_type in (
                INDUCTOR_COMPONENT_TYPES
                | TIMING_COMPONENT_TYPES
                | {"薄膜电容", "铝电解电容", "引线型陶瓷电容", "热敏电阻"}
                | VARISTOR_COMPONENT_TYPES
            ):
                frames.append(load_component_rows_by_typed_spec(spec))
                used_fast_path = True
    if not used_fast_path:
        return None
    return concat_component_frames(frames)


def resolve_prefetched_exact_part_rows(query_text, exact_part_rows=None):
    if isinstance(exact_part_rows, pd.DataFrame) and not exact_part_rows.empty:
        return exact_part_rows
    if not looks_like_compact_part_query(query_text):
        return pd.DataFrame()
    rows = load_component_rows_by_clean_model(query_text)
    if isinstance(rows, pd.DataFrame):
        return rows
    return pd.DataFrame()


def filter_mlcc_candidate_pairs_by_series_class(pairs, target_class):
    if not pairs or not mlcc_series_class_requires_filter(target_class):
        return pairs

    rows = load_component_rows_by_brand_model_pairs(pairs, preferred_component_type="MLCC")
    if not isinstance(rows, pd.DataFrame) or rows.empty:
        return pairs

    filtered_pairs = []
    seen = set()
    for row in rows.to_dict(orient="records"):
        brand_key = clean_brand(row.get("品牌", ""))
        model_key = clean_model(row.get("型号", ""))
        if brand_key == "" or model_key == "":
            continue
        profile = resolve_mlcc_series_profile(
            brand=brand_key,
            model=model_key,
            series=row.get("系列", ""),
            series_desc=row.get("系列说明", ""),
            special_use=row.get("特殊用途", ""),
        )
        if not mlcc_series_class_matches(profile.get("_mlcc_series_class", ""), target_class):
            continue
        key = (brand_key, model_key)
        if key in seen:
            continue
        seen.add(key)
        filtered_pairs.append(key)
    return filtered_pairs


def fetch_search_candidate_pairs(spec):
    if spec is None:
        return None
    target_type = infer_spec_component_type(spec)
    target_mlcc_class = infer_mlcc_series_class_from_spec(spec) if target_type == "MLCC" else ""
    compatible_types = compatible_component_types_for_search(target_type)
    search_table_name = search_index_table_for_component_type(target_type)
    if search_table_name == COMPONENTS_SEARCH_CORE_TABLE:
        return None
    where_clauses = []
    params = []
    exact_query = ""
    exact_query_params = []
    required_search_columns = {"品牌", "型号", "_component_type"}

    if target_type == "MLCC":
        size = clean_size(spec.get("尺寸（inch）", ""))
        material = clean_material(spec.get("材质（介质）", ""))
        pf = spec.get("容值_pf", None)
        tol = clean_tol_for_match(spec.get("容值误差", ""))
        volt = clean_voltage(spec.get("耐压（V）", ""))
        if size == "" or pf is None:
            return None
        required_search_columns.update({"_size", "_mat", "_pf", "_tol", "_volt_num"})
        if compatible_types:
            placeholders = ",".join(["?"] * len(compatible_types))
            where_clauses.append(f"_component_type IN ({placeholders})")
            params.extend(compatible_types)
        where_clauses.append("_size = ?")
        params.append(size)
        where_clauses.append("_pf IS NOT NULL AND ABS(_pf - ?) < 1e-6")
        params.append(float(pf))
        if material != "":
            where_clauses.append("_mat = ?")
            params.append(material)
        if tol != "":
            where_clauses.append("_tol = ?")
            params.append(tol)
        if volt != "":
            exact_voltage_clauses = list(where_clauses)
            exact_voltage_params = list(params)
            exact_voltage_clauses.append("_volt_num IS NOT NULL AND ABS(_volt_num - ?) < 1e-9")
            exact_voltage_params.append(float(volt))
            exact_query = (
                f'SELECT DISTINCT "品牌", "型号" FROM {search_table_name} '
                f'WHERE {" AND ".join(exact_voltage_clauses)}'
            )
            exact_query_params = exact_voltage_params
            where_clauses.append("_volt_num IS NOT NULL AND _volt_num >= ?")
            params.append(float(volt))
    elif target_type in RESISTOR_COMPONENT_TYPES or target_type == "热敏电阻":
        size = clean_size(spec.get("尺寸（inch）", ""))
        resistance_ohm = spec.get("_resistance_ohm", None)
        if resistance_ohm is None:
            return None
        required_search_columns.update({"_size", "_res_ohm", "_tol"})
        if compatible_types:
            placeholders = ",".join(["?"] * len(compatible_types))
            where_clauses.append(f"_component_type IN ({placeholders})")
            params.extend(compatible_types)
        elif target_type != "":
            where_clauses.append("_component_type = ?")
            params.append(target_type)
        if size != "":
            where_clauses.append("_size = ?")
            params.append(size)
        where_clauses.append("_res_ohm IS NOT NULL AND ABS(_res_ohm - ?) < 1e-12")
        params.append(float(resistance_ohm))
        tol = clean_tol_for_match(spec.get("容值误差", ""))
        if tol != "":
            where_clauses.append("_tol = ?")
            params.append(tol)
        power = clean_text(spec.get("_power", ""))
        power_watt = parse_power_to_watts(power) if power != "" else None
        if power_watt is not None:
            required_search_columns.add("_power_watt")
            where_clauses.append("_power_watt IS NOT NULL AND ABS(_power_watt - ?) < 1e-12")
            params.append(float(power_watt))
    elif target_type in INDUCTOR_COMPONENT_TYPES:
        size = clean_size(spec.get("尺寸（inch）", ""))
        value_text = clean_text(spec.get("容值", ""))
        unit = clean_text(spec.get("容值单位", "")).upper()
        tol = clean_tol_for_match(spec.get("容值误差", ""))
        if value_text == "" or unit == "":
            return None
        required_search_columns.update({"_size", "_value_num", "_unit_upper", "_tol"})
        placeholders = ",".join(["?"] * len(compatible_types or [target_type]))
        where_clauses.append(f"_component_type IN ({placeholders})")
        params.extend(compatible_types or [target_type])
        if size != "":
            where_clauses.append("_size = ?")
            params.append(size)
        where_clauses.append("_unit_upper = ?")
        params.append(unit)
        try:
            where_clauses.append("_value_num IS NOT NULL AND ABS(_value_num - ?) < 1e-9")
            params.append(float(value_text))
        except Exception:
            return None
        if tol != "":
            where_clauses.append("_tol = ?")
            params.append(tol)
    elif target_type in TIMING_COMPONENT_TYPES:
        size = clean_size(spec.get("尺寸（inch）", ""))
        value_text = clean_text(spec.get("容值", ""))
        unit = clean_text(spec.get("容值单位", "")).upper()
        tol = clean_tol_for_match(spec.get("容值误差", ""))
        volt = clean_voltage(spec.get("耐压（V）", ""))
        if value_text == "" or unit == "":
            return None
        required_search_columns.update({"_size", "_value_num", "_unit_upper", "_tol", "_volt_num"})
        placeholders = ",".join(["?"] * len(compatible_types or [target_type]))
        where_clauses.append(f"_component_type IN ({placeholders})")
        params.extend(compatible_types or [target_type])
        if size != "":
            where_clauses.append("_size = ?")
            params.append(size)
        where_clauses.append("_unit_upper = ?")
        params.append(unit)
        try:
            where_clauses.append("_value_num IS NOT NULL AND ABS(_value_num - ?) < 1e-9")
            params.append(float(value_text))
        except Exception:
            return None
        if tol != "":
            where_clauses.append("_tol = ?")
            params.append(tol)
        if volt != "":
            where_clauses.append("_volt_num IS NOT NULL AND ABS(_volt_num - ?) < 1e-9")
            params.append(float(volt))
    elif target_type in {"薄膜电容", "铝电解电容", "引线型陶瓷电容"}:
        pf = spec.get("容值_pf", None)
        tol = clean_tol_for_match(spec.get("容值误差", ""))
        volt = clean_voltage(spec.get("耐压（V）", ""))
        body_size = clean_text(spec.get("_body_size", ""))
        pitch = clean_text(spec.get("_pitch", ""))
        work_temp = normalize_working_temperature_text(spec.get("工作温度", ""))
        life_hours = normalize_life_hours_value(spec.get("寿命（h）", ""))
        mount_style = normalize_mounting_style(spec.get("安装方式", ""))
        if pf is None:
            return None
        required_search_columns.update({"_component_type", "_pf", "_tol", "_volt_num", "_body_size", "_pitch", "_temp_low", "_temp_high", "_life_hours_num", "_mount_style"})
        where_clauses.append("_component_type = ?")
        params.append(target_type)
        where_clauses.append("_pf IS NOT NULL AND ABS(_pf - ?) < 1e-6")
        params.append(float(pf))
        if tol != "":
            where_clauses.append("_tol = ?")
            params.append(tol)
        if volt != "":
            where_clauses.append("_volt_num IS NOT NULL AND _volt_num >= ?")
            params.append(float(volt))
        if body_size != "":
            where_clauses.append("_body_size = ?")
            params.append(body_size)
        if pitch != "":
            where_clauses.append("_pitch = ?")
            params.append(pitch)
        if target_type == "铝电解电容":
            temp_low, temp_high = working_temperature_bounds(work_temp)
            if temp_high is not None:
                where_clauses.append("_temp_high IS NOT NULL AND _temp_high >= ?")
                params.append(float(temp_high))
            if temp_low is not None:
                where_clauses.append("_temp_low IS NOT NULL AND _temp_low <= ?")
                params.append(float(temp_low))
            if life_hours != "":
                life_target = life_hours_to_number(life_hours)
                if life_target is not None:
                    where_clauses.append("_life_hours_num IS NOT NULL AND _life_hours_num >= ?")
                    params.append(float(life_target))
            if mount_style != "":
                where_clauses.append("_mount_style = ?")
                params.append(mount_style)
        if target_type == "薄膜电容":
            material = clean_material(spec.get("材质（介质）", ""))
            safety_class = clean_text(spec.get("_safety_class", ""))
            required_search_columns.update({"_mat", "_safety_class"})
            if material != "":
                where_clauses.append("_mat = ?")
                params.append(material)
            if safety_class != "":
                where_clauses.append("_safety_class = ?")
                params.append(safety_class)
        elif target_type == "引线型陶瓷电容":
            material = clean_material(spec.get("材质（介质）", ""))
            required_search_columns.add("_mat")
            if material != "":
                where_clauses.append("_mat = ?")
                params.append(material)
    elif target_type in VARISTOR_COMPONENT_TYPES:
        tol = clean_tol_for_match(spec.get("容值误差", ""))
        varistor_voltage = clean_voltage(spec.get("_varistor_voltage", "")) or clean_voltage(spec.get("耐压（V）", ""))
        disc_size = clean_text(spec.get("_disc_size", ""))
        pitch = clean_text(spec.get("_pitch", ""))
        if varistor_voltage == "" and disc_size == "" and pitch == "":
            return None
        required_search_columns.update({"_varistor_voltage", "_disc_size", "_pitch", "_tol"})
        placeholders = ",".join(["?"] * len(compatible_types or [target_type]))
        where_clauses.append(f"_component_type IN ({placeholders})")
        params.extend(compatible_types or [target_type])
        if varistor_voltage != "":
            where_clauses.append("_varistor_voltage = ?")
            params.append(varistor_voltage)
        if disc_size != "":
            where_clauses.append("_disc_size = ?")
            params.append(disc_size)
        if pitch != "":
            where_clauses.append("_pitch = ?")
            params.append(pitch)
        if tol != "":
            where_clauses.append("_tol = ?")
            params.append(tol)
    else:
        return None

    query = (
        f'SELECT DISTINCT "品牌", "型号" FROM {search_table_name} '
        f'WHERE {" AND ".join(where_clauses)}'
    )
    conn = open_search_db_connection(timeout_sec=10)
    try:
        if conn is None or not search_index_can_serve_queries(
            conn,
            required_columns=required_search_columns,
            table_name=search_table_name,
            allow_without_database=True,
        ):
            if conn is not None:
                conn.close()
            if os.path.exists(DB_PATH):
                rebuild_search_index_from_database_fast()
            conn = open_search_db_connection(timeout_sec=10)
            if conn is None or not search_index_can_serve_queries(
                conn,
                required_columns=required_search_columns,
                table_name=search_table_name,
                allow_without_database=True,
            ):
                return None
        if exact_query:
            exact_rows = conn.execute(exact_query, exact_query_params).fetchall()
            exact_result = [
                (clean_text(row[0]), clean_text(row[1]))
                for row in exact_rows
                if clean_text(row[1]) != ""
            ]
            if exact_result:
                if target_type == "MLCC" and mlcc_series_class_requires_filter(target_mlcc_class):
                    exact_result = filter_mlcc_candidate_pairs_by_series_class(
                        exact_result,
                        target_mlcc_class,
                    )
                return exact_result
        rows = conn.execute(query, params).fetchall()
        result = [(clean_text(row[0]), clean_text(row[1])) for row in rows if clean_text(row[1]) != ""]
        if target_type == "MLCC" and mlcc_series_class_requires_filter(target_mlcc_class):
            result = filter_mlcc_candidate_pairs_by_series_class(result, target_mlcc_class)
        return result if result else []
    except Exception:
        return None
    finally:
        if conn is not None:
            conn.close()


def scope_search_dataframe(df, spec):
    if df is None:
        return pd.DataFrame()
    if spec is None:
        return prepare_search_dataframe(df)
    base = df
    search_candidate_pairs = fetch_search_candidate_pairs(spec)
    if search_candidate_pairs is not None:
        base = filter_base_by_candidate_pairs(base, search_candidate_pairs)
        if base.empty:
            return base
    base = prepare_search_dataframe(base)
    if base.empty:
        return base

    target_type = infer_spec_component_type(spec)
    mask = pd.Series(True, index=base.index)
    if target_type != "" and "_component_type" in base.columns:
        if target_type in VARISTOR_COMPONENT_TYPES:
            compatible_types = VARISTOR_COMPONENT_TYPES if target_type == "压敏电阻" else {target_type, "压敏电阻"}
            same_type_mask = base["_component_type"].isin(compatible_types)
        elif target_type in RESISTOR_COMPONENT_TYPES:
            compatible_types = RESISTOR_COMPONENT_TYPES if target_type == "贴片电阻" else {target_type, "贴片电阻"}
            same_type_mask = base["_component_type"].isin(compatible_types)
        elif target_type in INDUCTOR_COMPONENT_TYPES:
            compatible_types = compatible_component_types_for_search(target_type) or [target_type]
            same_type_mask = base["_component_type"].isin(compatible_types)
        else:
            same_type_mask = base["_component_type"].eq(target_type)
        if target_type in OTHER_PASSIVE_TYPES:
            if not same_type_mask.any():
                return base.iloc[0:0]
            mask &= same_type_mask
        elif same_type_mask.any():
            mask &= same_type_mask

    size = clean_size(spec.get("尺寸（inch）", ""))
    if size != "":
        mask &= base["_size"].eq(size)
        if not mask.any():
            return base.iloc[0:0]

    if target_type == "MLCC":
        material = clean_material(spec.get("材质（介质）", ""))
        if material != "":
            mask &= base["_mat"].eq(material)
            if not mask.any():
                return base.iloc[0:0]
        pf = spec.get("容值_pf", None)
        if pf is not None:
            target_pf = float(pf)
            mask &= base["_pf"].notna() & ((base["_pf"] - target_pf).abs() < 1e-6)
            if not mask.any():
                return base.iloc[0:0]
        spec_mlcc_class = infer_mlcc_series_class_from_spec(spec)
        if mlcc_series_class_requires_filter(spec_mlcc_class):
            candidate_class = (
                base["_mlcc_series_class"].astype(str)
                if "_mlcc_series_class" in base.columns
                else pd.Series([""] * len(base), index=base.index, dtype="object")
            )
            class_mask = candidate_class.apply(lambda value: mlcc_series_class_matches(value, spec_mlcc_class))
            if not class_mask.any():
                return base.iloc[0:0]
            mask &= class_mask
            if not mask.any():
                return base.iloc[0:0]
    else:
        if target_type in VARISTOR_COMPONENT_TYPES:
            spec_varistor_voltage = clean_voltage(spec.get("_varistor_voltage", "")) or clean_voltage(spec.get("耐压（V）", ""))
            spec_disc_size = clean_text(spec.get("_disc_size", ""))
            if spec_varistor_voltage != "" and "_varistor_voltage" in base.columns:
                same_varistor_voltage = base["_varistor_voltage"].astype("string").fillna("").eq(spec_varistor_voltage)
                mask &= same_varistor_voltage
                if not mask.any():
                    return base.iloc[0:0]
            if spec_disc_size != "" and "_disc_size" in base.columns:
                same_disc_size = base["_disc_size"].astype("string").fillna("").str.strip().eq(spec_disc_size)
                if same_disc_size.any():
                    mask &= same_disc_size
                    if not mask.any():
                        return base.iloc[0:0]
        resistance_ohm = spec.get("_resistance_ohm", None)
        if resistance_ohm is not None and "_res_ohm" in base.columns:
            target_ohm = float(resistance_ohm)
            mask &= base["_res_ohm"].notna() & ((base["_res_ohm"] - target_ohm).abs() < 1e-6)
            if not mask.any():
                return base.iloc[0:0]
        elif spec.get("容值_pf", None) is not None:
            target_pf = float(spec.get("容值_pf"))
            mask &= base["_pf"].notna() & ((base["_pf"] - target_pf).abs() < 1e-6)
            if not mask.any():
                return base.iloc[0:0]

    return base if bool(mask.all()) else base[mask]


def match_other_passive_spec(df, spec):
    if spec is None:
        return pd.DataFrame()
    base = scope_search_dataframe(df, spec)
    if base.empty:
        return pd.DataFrame()

    component_type = infer_spec_component_type(spec)
    if component_type == "铝电解电容":
        work = base.copy()
        spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
        spec_volt = clean_voltage(spec.get("耐压（V）", ""))
        spec_temp = normalize_working_temperature_text(spec.get("工作温度", ""))
        spec_life = normalize_life_hours_value(spec.get("寿命（h）", ""))
        spec_mount = normalize_mounting_style(spec.get("安装方式", ""))
        spec_special = normalize_special_use(spec.get("特殊用途", ""))
        spec_body_size = clean_text(spec.get("_body_size", ""))
        spec_pitch = clean_text(spec.get("_pitch", ""))
        if spec_tol != "":
            work = work[tolerance_allows_series(work, spec_tol)]
        if spec_volt != "":
            try:
                spec_volt_num = float(spec_volt)
                work = work[work["_volt_num"].notna() & work["_volt_num"].ge(spec_volt_num)]
            except Exception:
                work = work[work["_volt"].eq(spec_volt)]
        if spec_body_size != "" and "_body_size" in work.columns:
            same_body = work["_body_size"].astype(str).apply(clean_text).eq(spec_body_size)
            work = work[same_body]
        if spec_temp != "":
            same_temp = work["工作温度"].astype(str).apply(normalize_working_temperature_text).eq(spec_temp) if "工作温度" in work.columns else pd.Series(False, index=work.index)
            if same_temp.any():
                work = work[same_temp]
            elif "_temp_high" in work.columns:
                temp_low, temp_high = working_temperature_bounds(spec_temp)
                temp_mask = pd.Series(True, index=work.index)
                if temp_high is not None:
                    temp_mask &= pd.to_numeric(work["_temp_high"], errors="coerce").notna() & pd.to_numeric(work["_temp_high"], errors="coerce").ge(temp_high)
                if temp_low is not None:
                    temp_mask &= pd.to_numeric(work["_temp_low"], errors="coerce").notna() & pd.to_numeric(work["_temp_low"], errors="coerce").le(temp_low)
                work = work[temp_mask]
            else:
                work = work.iloc[0:0]
        if spec_life != "":
            same_life = work["寿命（h）"].astype(str).apply(normalize_life_hours_value).eq(spec_life) if "寿命（h）" in work.columns else pd.Series(False, index=work.index)
            if same_life.any():
                work = work[same_life]
            elif "_life_hours_num" in work.columns:
                life_target = life_hours_to_number(spec_life)
                life_mask = pd.to_numeric(work["_life_hours_num"], errors="coerce").notna() & pd.to_numeric(work["_life_hours_num"], errors="coerce").ge(life_target)
                work = work[life_mask]
            else:
                work = work.iloc[0:0]
        if spec_mount != "":
            same_mount = work["_mount_style"].astype(str).apply(normalize_mounting_style).eq(spec_mount) if "_mount_style" in work.columns else (
                work["安装方式"].astype(str).apply(normalize_mounting_style).eq(spec_mount) if "安装方式" in work.columns else pd.Series(False, index=work.index)
            )
            work = work[same_mount]
        if spec_special != "":
            same_special = work["特殊用途"].astype(str).apply(lambda value: special_use_matches(value, spec_special)) if "特殊用途" in work.columns else (
                work["_special_use_norm"].astype(str).apply(lambda value: special_use_matches(value, spec_special)) if "_special_use_norm" in work.columns else pd.Series(False, index=work.index)
            )
            work = work[same_special]
        if spec_pitch != "" and "_pitch" in work.columns:
            same_pitch = work["_pitch"].astype(str).apply(clean_text).eq(spec_pitch)
            work = work[same_pitch]
        if work.empty:
            return pd.DataFrame()
        work = work.copy()
        work["推荐等级"] = "完全匹配"
        if "_model_rule_authority" in work.columns:
            work["_seed_rank"] = work["_model_rule_authority"].astype(str).apply(lambda value: 0 if clean_text(value) == "jianghai_seed" else 1)
        else:
            work["_seed_rank"] = 1
        work["_brand_rank"] = work["品牌"].apply(lambda value: brand_priority_value(value, component_type=component_type)) if "品牌" in work.columns else 99
        sort_cols = ["_seed_rank", "_brand_rank", "品牌", "型号"] if "品牌" in work.columns else ["_seed_rank"]
        ascending = [True] * len(sort_cols)
        work = work.sort_values(by=sort_cols, ascending=ascending)
        work = work.drop(columns=["_seed_rank", "_brand_rank"], errors="ignore")
        return work

    if component_type == "薄膜电容":
        work = base.copy()
        spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
        spec_volt = clean_voltage(spec.get("耐压（V）", ""))
        spec_material = clean_material(spec.get("材质（介质）", ""))
        spec_body_size = clean_text(spec.get("_body_size", ""))
        spec_pitch = clean_text(spec.get("_pitch", ""))
        spec_safety = clean_text(spec.get("_safety_class", ""))

        if spec_material != "" and "_mat" in work.columns:
            same_material = work["_mat"].eq(spec_material)
            if same_material.any():
                work = work[same_material]
        if spec_tol != "":
            same_tol = tolerance_equal_series(work, spec_tol)
            if same_tol.any():
                work = work[same_tol]
        if spec_volt != "":
            try:
                spec_volt_num = float(spec_volt)
                work = work[work["_volt_num"].notna() & work["_volt_num"].ge(spec_volt_num)]
            except Exception:
                work = work[work["_volt"].eq(spec_volt)]
        if spec_body_size != "" and "_body_size" in work.columns:
            same_body = work["_body_size"].astype(str).apply(clean_text).eq(spec_body_size)
            if same_body.any():
                work = work[same_body]
        if spec_pitch != "" and "_pitch" in work.columns:
            same_pitch = work["_pitch"].astype(str).apply(clean_text).eq(spec_pitch)
            if same_pitch.any():
                work = work[same_pitch]
        if spec_safety != "" and "_safety_class" in work.columns:
            same_safety = work["_safety_class"].astype(str).apply(clean_text).eq(spec_safety)
            if same_safety.any():
                work = work[same_safety]
        if work.empty:
            return pd.DataFrame()
        work = work.copy()
        work["推荐等级"] = "完全匹配"
        return work

    if component_type in VARISTOR_COMPONENT_TYPES:
        work = base.copy()
        spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
        spec_varistor_voltage = clean_voltage(spec.get("_varistor_voltage", "")) or clean_voltage(spec.get("耐压（V）", ""))
        spec_disc_size = clean_text(spec.get("_disc_size", ""))
        spec_pitch = clean_text(spec.get("_pitch", ""))
        if spec_varistor_voltage != "" and "_varistor_voltage" in work.columns:
            same_varistor_voltage = work["_varistor_voltage"].astype(str).apply(clean_voltage).eq(spec_varistor_voltage)
            if same_varistor_voltage.any():
                work = work[same_varistor_voltage]
        if spec_tol != "":
            same_tol = tolerance_equal_series(work, spec_tol)
            if same_tol.any():
                work = work[same_tol]
        if spec_disc_size != "" and "_disc_size" in work.columns:
            same_disc_size = work["_disc_size"].astype(str).apply(clean_text).eq(spec_disc_size)
            if same_disc_size.any():
                work = work[same_disc_size]
        if spec_pitch != "" and "_pitch" in work.columns:
            same_pitch = work["_pitch"].astype(str).apply(clean_text).eq(spec_pitch)
            if same_pitch.any():
                work = work[same_pitch]
        if work.empty:
            return pd.DataFrame()
        work = work.copy()
        work["推荐等级"] = "完全匹配"
        return work

    if component_type in (RESISTOR_COMPONENT_TYPES | {"热敏电阻"}):
        work = base.copy()
        spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
        spec_power_watt = parse_power_to_watts(spec.get("_power", ""))
        if spec_tol != "":
            same_tol = tolerance_equal_series(work, spec_tol)
            if same_tol.any():
                work = work[same_tol]
        if spec_power_watt is not None and "_power_watt" in work.columns:
            same_power = work["_power_watt"].notna() & ((pd.to_numeric(work["_power_watt"], errors="coerce") - spec_power_watt).abs() < 1e-9)
            if same_power.any():
                work = work[same_power]
        if work.empty:
            return pd.DataFrame()
        return apply_match_levels_and_sort(work, spec)

    if component_type in INDUCTOR_COMPONENT_TYPES:
        work = base.copy()
        spec_value = clean_text(spec.get("容值", ""))
        spec_unit = clean_text(spec.get("容值单位", "")).upper()
        spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
        if spec_unit != "" and "容值单位" in work.columns:
            same_unit = work["容值单位"].astype(str).apply(lambda x: clean_text(x).upper()).eq(spec_unit)
            if same_unit.any():
                work = work[same_unit]
        if spec_value != "" and "容值" in work.columns:
            row_value = work["容值"].astype(str).apply(clean_text)
            exact_text = row_value.eq(spec_value)
            try:
                target_value = float(spec_value)
                numeric_mask = pd.to_numeric(row_value, errors="coerce").sub(target_value).abs().lt(1e-9)
                same_value = exact_text | numeric_mask.fillna(False)
            except Exception:
                same_value = exact_text
            if same_value.any():
                work = work[same_value]
        if spec_tol != "":
            same_tol = tolerance_equal_series(work, spec_tol)
            if same_tol.any():
                work = work[same_tol]
        if work.empty:
            return pd.DataFrame()
        work = work.copy()
        work["推荐等级"] = "完全匹配"
        return work

    if component_type in TIMING_COMPONENT_TYPES:
        work = base.copy()
        spec_value = clean_text(spec.get("容值", ""))
        spec_unit = clean_text(spec.get("容值单位", "")).upper()
        spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
        spec_volt = clean_voltage(spec.get("耐压（V）", ""))
        if spec_unit != "" and "容值单位" in work.columns:
            same_unit = work["容值单位"].astype(str).apply(lambda x: clean_text(x).upper()).eq(spec_unit)
            if same_unit.any():
                work = work[same_unit]
        if spec_value != "" and "容值" in work.columns:
            row_value = work["容值"].astype(str).apply(clean_text)
            exact_text = row_value.eq(spec_value)
            try:
                target_value = float(spec_value)
                numeric_mask = pd.to_numeric(row_value, errors="coerce").sub(target_value).abs().lt(1e-9)
                same_value = exact_text | numeric_mask.fillna(False)
            except Exception:
                same_value = exact_text
            if same_value.any():
                work = work[same_value]
        if spec_tol != "":
            same_tol = tolerance_equal_series(work, spec_tol)
            if same_tol.any():
                work = work[same_tol]
        if spec_volt != "":
            same_volt = work["耐压（V）"].astype(str).apply(clean_voltage).eq(spec_volt)
            if same_volt.any():
                work = work[same_volt]
        if work.empty:
            return pd.DataFrame()
        work = work.copy()
        work["推荐等级"] = "完全匹配"
        return work

    return pd.DataFrame()


def build_data_quality_report(df):
    if df.empty:
        return [], {}

    work = prepare_search_dataframe(df).copy()
    work["品牌"] = work["品牌"].astype(str).apply(clean_brand)
    work["型号"] = work["_model_clean"]
    work["尺寸（inch）"] = work["_size"]
    work["材质（介质）"] = work["_mat"]
    work["容值误差"] = work["_tol"]
    work["耐压（V）"] = work["_volt"]
    work["_型号规范"] = work["_model_clean"]

    critical_blank = (
        work["尺寸（inch）"].eq("")
        | work["材质（介质）"].eq("")
        | work["_pf"].isna()
        | work["容值误差"].eq("")
        | work["耐压（V）"].eq("")
    )

    parser_fillable_mask = critical_blank & work["_型号规范"].apply(parser_can_enrich_model)

    duplicate_mask = work["_型号规范"].ne("") & work["_型号规范"].duplicated(keep=False)

    base_cols = ["品牌", "型号", "系列", "尺寸（inch）", "材质（介质）", "容值", "容值单位", "容值误差", "耐压（V）", "特殊用途"]
    summary = []
    details = {}

    def add_issue(key, title, description, issue_df):
        summary.append({
            "检查项": title,
            "问题数量": int(len(issue_df)),
            "说明": description,
        })
        details[key] = issue_df[base_cols].head(200).copy() if not issue_df.empty else pd.DataFrame(columns=base_cols)

    add_issue(
        "missing_size",
        "缺尺寸",
        "数据库存在型号但尺寸（inch）为空，后续会直接影响匹配精度。",
        work[work["尺寸（inch）"].eq("")],
    )
    add_issue(
        "missing_material",
        "缺材质",
        "材质（介质）为空，会影响同规格筛选和替代判断。",
        work[work["材质（介质）"].eq("")],
    )
    add_issue(
        "missing_cap",
        "缺容值",
        "容值_pf 无法计算，通常意味着容值或单位缺失。",
        work[work["_pf"].isna()],
    )
    add_issue(
        "missing_tolerance",
        "缺容差",
        "容值误差为空，会影响完全匹配和高代低判断。",
        work[work["容值误差"].eq("")],
    )
    add_issue(
        "missing_voltage",
        "缺耐压",
        "耐压为空，会影响替代方向判断。",
        work[work["耐压（V）"].eq("")],
    )
    add_issue(
        "duplicate_model",
        "重复型号",
        "同一清洗后型号重复出现，可能是重复导入或品牌名不统一导致。",
        work[duplicate_mask].sort_values(by=["型号", "品牌"]),
    )
    add_issue(
        "parser_fillable",
        "规则可回填但库里仍缺字段",
        "这些型号按当前规则本可反推出更多参数，但数据库字段仍然留空，优先建议回填。",
        work[parser_fillable_mask],
    )

    return pd.DataFrame(summary), details


def get_model_reverse_lookup(df, cache_signature=None):
    if df is None or df.empty:
        return pd.DataFrame(columns=MODEL_REVERSE_LOOKUP_COLUMNS).set_index("_model_clean", drop=False)
    if cache_signature is None:
        cache_signature = get_data_cache_signature()
    cached = MODEL_REVERSE_LOOKUP_CACHE.get(cache_signature)
    if cached is not None:
        return cached
    work = prepare_search_dataframe(df)
    available_cols = [col for col in MODEL_REVERSE_LOOKUP_COLUMNS if col in work.columns]
    lookup = work[available_cols].copy()
    lookup = lookup[lookup["_model_clean"].astype(str).ne("")]
    if not lookup.empty:
        lookup = lookup.copy()
        lookup["_reverse_priority"] = work.loc[lookup.index].apply(reverse_lookup_row_priority, axis=1)
        lookup["_reverse_order"] = range(len(lookup))
        lookup = lookup.sort_values(
            by=["_model_clean", "_reverse_priority", "_reverse_order"],
            ascending=[True, False, True],
            kind="mergesort",
        )
    lookup = lookup.drop_duplicates(subset=["_model_clean"], keep="first")
    lookup = lookup.drop(columns=["_reverse_priority", "_reverse_order"], errors="ignore")
    lookup = lookup.set_index("_model_clean", drop=False)
    MODEL_REVERSE_LOOKUP_CACHE.clear()
    MODEL_REVERSE_LOOKUP_CACHE[cache_signature] = lookup
    return lookup


def lookup_model_reverse_row(df, model, cache_signature=None):
    m = clean_model(model)
    if m == "":
        return None
    if df is not None and not df.empty:
        lookup = get_model_reverse_lookup(df, cache_signature=cache_signature)
        if not lookup.empty and m in lookup.index:
            row = lookup.loc[m]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            return row
    db_rows = load_component_rows_by_clean_model(m)
    if db_rows.empty:
        return None
    lookup = get_model_reverse_lookup(db_rows, cache_signature=f"model:{m}")
    if lookup.empty or m not in lookup.index:
        return None
    row = lookup.loc[m]
    if isinstance(row, pd.DataFrame):
        row = row.iloc[0]
    return row


def reverse_spec(df, model):
    m = clean_model(model)
    if m == "":
        return None
    row = lookup_model_reverse_row(df, m)
    row_brand = clean_brand(row.get("品牌", "")) if row is not None else ""
    row_type = normalize_component_type(row.get("器件类型", "")) if row is not None else ""
    parsed_rule = parse_model_rule(m, brand=row_brand, component_type=row_type)
    if row is not None:
        row_text = " ".join([
            clean_text(row.get("器件类型", "")),
            clean_text(row.get("品牌", "")),
            clean_text(row.get("型号", "")),
            clean_text(row.get("系列", "")),
            clean_text(row.get("安装方式", "")),
            clean_text(row.get("封装代码", "")),
            clean_text(row.get("尺寸（inch）", "")),
            clean_text(row.get("尺寸（mm）", "")),
            clean_text(row.get("材质（介质）", "")),
            clean_text(row.get("规格摘要", "")),
            clean_text(row.get("容值", "")),
            clean_text(row.get("容值单位", "")),
            clean_text(row.get("工作温度", "")),
            clean_text(row.get("寿命（h）", "")),
            clean_text(row.get("特殊用途", "")),
            clean_text(row.get("备注1", "")),
            clean_text(row.get("备注2", "")),
            clean_text(row.get("备注3", "")),
        ])
        component_type = infer_db_component_type(row) or normalize_component_type(row.get("器件类型", ""))
        raw_value = clean_text(row.get("容值", ""))
        raw_unit = clean_text(row.get("容值单位", "")).upper()
        if component_type in INDUCTOR_COMPONENT_TYPES and (raw_value == "" or raw_unit == ""):
            inductance = find_inductance_in_text(row_text)
            inductance_match = re.fullmatch(r"(\d+(?:\.\d+)?)(NH|UH|MH)", inductance, flags=re.I)
            if inductance_match:
                raw_value = raw_value or inductance_match.group(1)
                raw_unit = raw_unit or inductance_match.group(2).upper()
        if component_type in TIMING_COMPONENT_TYPES and (raw_value == "" or raw_unit == ""):
            frequency = find_frequency_in_text(row_text)
            frequency_match = re.fullmatch(r"(\d+(?:\.\d+)?)(HZ|KHZ|MHZ)", frequency, flags=re.I)
            if frequency_match:
                raw_value = raw_value or frequency_match.group(1)
                raw_unit = raw_unit or frequency_match.group(2).upper()
        pf = cap_to_pf(raw_value, raw_unit) if component_type in CAPACITOR_COMPONENT_TYPES else None
        db_spec = {
            "品牌": clean_brand(row["品牌"]),
            "型号": clean_model(row["型号"]),
            "器件类型": component_type,
            "尺寸（inch）": clean_size(row["尺寸（inch）"]),
            "材质（介质）": clean_material(row["材质（介质）"]),
            "容值": raw_value,
            "容值单位": raw_unit,
            "容值_pf": pf,
            "容值误差": clean_tol_for_match(row["容值误差"]),
            "耐压（V）": clean_voltage(row["耐压（V）"]),
            "系列": clean_text(row.get("系列", "")),
            "系列说明": clean_text(row.get("系列说明", "")),
            "工作温度": normalize_working_temperature_text(row.get("工作温度", "") or extract_working_temperature_from_text(row_text)),
            "寿命（h）": normalize_life_hours_value(row.get("寿命（h）", "") or parse_life_hours_from_text(row_text)),
            "安装方式": normalize_mounting_style(row.get("安装方式", ""), row.get("封装代码", "")),
            "特殊用途": normalize_special_use(row.get("特殊用途", "") or extract_special_use_from_text(row_text)),
            "_mlcc_series_class": clean_text(row.get("_mlcc_series_class", "")),
            "封装代码": clean_text(row.get("封装代码", "")),
            "尺寸（mm）": clean_text(row.get("尺寸（mm）", "")),
            "规格摘要": clean_text(row.get("规格摘要", "")),
            "_resistance_ohm": find_resistance_in_text(row_text) if component_type in (RESISTOR_COMPONENT_TYPES | {"热敏电阻"}) else None,
            "_power": find_power_in_text(row_text) if component_type in (RESISTOR_COMPONENT_TYPES | VARISTOR_COMPONENT_TYPES) else "",
            "_body_size": extract_body_size_from_text(row_text) if component_type in {"铝电解电容", "薄膜电容", "引线型陶瓷电容"} else clean_text(row.get("尺寸（mm）", "")),
            "_pitch": extract_pitch_from_text(row_text) if component_type in ({"铝电解电容", "薄膜电容", "引线型陶瓷电容"} | VARISTOR_COMPONENT_TYPES) else "",
            "_safety_class": find_safety_class(row_text) if component_type == "薄膜电容" else "",
            "_varistor_voltage": find_varistor_voltage_in_text(row_text) if component_type in VARISTOR_COMPONENT_TYPES else "",
            "_disc_size": find_disc_size_code(row_text) if component_type in VARISTOR_COMPONENT_TYPES else "",
        }
        if parsed_rule is not None:
            row_is_jianghai = (
                "JIANGHAI" in clean_text(row_brand).upper()
                or "江海" in row_brand
                or jianghai_series_code_from_model(m) != ""
            )
            db_spec = merge_parsed_rule_into_record(
                db_spec,
                parsed_rule,
                override_conflicts=(not row_is_jianghai),
            )
        if component_type == "MLCC":
            mlcc_profile = resolve_mlcc_series_profile(
                brand=db_spec.get("品牌", ""),
                model=db_spec.get("型号", ""),
                series=db_spec.get("系列", ""),
                series_desc=db_spec.get("系列说明", ""),
                special_use=db_spec.get("特殊用途", ""),
            )
            if mlcc_series_should_replace(db_spec.get("系列", ""), mlcc_profile.get("系列", "")):
                db_spec["系列"] = clean_text(mlcc_profile.get("系列", ""))
            db_spec["系列说明"] = clean_text(db_spec.get("系列说明", "")) or clean_text(mlcc_profile.get("系列说明", ""))
            db_spec["特殊用途"] = normalize_special_use(
                db_spec.get("特殊用途", "") or clean_text(mlcc_profile.get("特殊用途", ""))
            )
            db_spec["_mlcc_series_class"] = clean_text(db_spec.get("_mlcc_series_class", "")) or clean_text(
                mlcc_profile.get("_mlcc_series_class", "")
            )
        return db_spec
    return parsed_rule

def exclude_same_brand(df, source_brand):
    source_brand = clean_brand(source_brand)
    source_brand_upper = source_brand.upper()
    if source_brand == "":
        return df
    if "信昌" in source_brand or "PDC" in source_brand_upper:
        return df[~df["品牌"].astype(str).str.contains("信昌|PDC", case=False, na=False)]
    if "三星" in source_brand or "SAMSUNG" in source_brand_upper or "SAMSU" in source_brand_upper:
        return df[~df["品牌"].astype(str).str.contains("三星|SAMSUNG|SAMSU", case=False, na=False)]
    if "TDK" in source_brand_upper or "东电化" in source_brand:
        return df[~df["品牌"].astype(str).str.contains("TDK|东电化", case=False, na=False)]
    if "MURATA" in source_brand_upper or "村田" in source_brand:
        return df[~df["品牌"].astype(str).str.contains("MURATA|村田", case=False, na=False)]
    if "YAGEO" in source_brand_upper or "国巨" in source_brand or "YEGO" in source_brand_upper:
        return df[~df["品牌"].astype(str).str.contains("YAGEO|YEGO|国巨", case=False, na=False)]
    if "SAMWHA" in source_brand_upper or "三和" in source_brand:
        return df[~df["品牌"].astype(str).str.contains("SAMWHA|三和", case=False, na=False)]
    if "CCTC" in source_brand_upper or "三环" in source_brand:
        return df[~df["品牌"].astype(str).str.contains("CCTC|三环", case=False, na=False)]
    if "WALSIN" in source_brand_upper or "华新科" in source_brand:
        return df[~df["品牌"].astype(str).str.contains("WALSIN|华新科", case=False, na=False)]
    if "TAIYO" in source_brand_upper or "太阳诱电" in source_brand:
        return df[~df["品牌"].astype(str).str.contains("TAIYO|太阳诱电", case=False, na=False)]
    return df[df["品牌"].astype(str).apply(clean_brand) != source_brand]


def extract_brand_model_pairs(df):
    if not isinstance(df, pd.DataFrame) or df.empty:
        return []
    if "品牌" not in df.columns or "型号" not in df.columns:
        return []
    pairs = []
    seen = set()
    for brand, model in df[["品牌", "型号"]].itertuples(index=False, name=None):
        brand_key = clean_brand(brand)
        model_key = clean_model(model)
        if brand_key == "" or model_key == "":
            continue
        key = (brand_key, model_key)
        if key in seen:
            continue
        seen.add(key)
        pairs.append(key)
    return pairs


def exclude_brand_model_pairs(df, pairs):
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df
    if "品牌" not in df.columns or "型号" not in df.columns:
        return df
    normalized_pairs = {
        (clean_brand(brand), clean_model(model))
        for brand, model in pairs
        if clean_brand(brand) != "" and clean_model(model) != ""
    }
    if not normalized_pairs:
        return df
    brand_series = df["品牌"].astype(str).apply(clean_brand)
    model_series = df["型号"].astype(str).apply(clean_model)
    keep_mask = pd.Series(True, index=df.index)
    for brand_key, model_key in normalized_pairs:
        keep_mask &= ~((brand_series == brand_key) & (model_series == model_key))
    return df.loc[keep_mask].copy()

def match_by_spec(df, spec):
    if spec is None:
        return pd.DataFrame()
    work = scope_search_dataframe(df, spec)
    if work.empty:
        return pd.DataFrame()

    spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
    spec_volt = clean_voltage(spec.get("耐压（V）", ""))
    spec_pf = spec.get("容值_pf", None)

    if spec_pf is None:
        return pd.DataFrame()

    mask = pd.Series(True, index=work.index)

    def volt_num(value):
        try:
            return float(value)
        except:
            return None

    if spec_tol != "":
        mask &= tolerance_allows_series(work, spec_tol)

    if spec_volt != "":
        spec_volt_num = volt_num(spec_volt)
        mask &= work["_volt_num"].notna() & (spec_volt_num is not None) & work["_volt_num"].ge(spec_volt_num)

    out = work[mask].copy()
    out = exclude_same_brand(out, spec.get("品牌", ""))
    out = apply_match_levels_and_sort(out, spec)
    return out.drop(columns=[c for c in ["_size", "_mat", "_tol", "_volt", "_pf", "_tol_kind", "_tol_num", "_volt_num", "_component_type", "_res_ohm"] if c in out.columns])


def style_exact_match_rows(df, spec=None):
    def row_style(row):
        level = clean_text(row.get("推荐等级", "")) or clean_text(row.get("首选推荐等级", ""))
        if level == "完全匹配":
            return ["background-color: #fff59d; color: #111111;" for _ in row]
        if level == "部分参数匹配":
            return ["background-color: #fde2e1; color: #7f1d1d;" for _ in row]
        if level in {"高代低", "可直接替代"}:
            return ["background-color: #dbeafe; color: #1d4ed8;" for _ in row]
        return ["" for _ in row]
    try:
        return df.style.apply(row_style, axis=1)
    except:
        return df


def format_display_df(show_df):
    show_df = show_df.copy()
    for col in [
        "品牌","型号","品牌1","型号1","品牌2","型号2","品牌3","型号3",
        "推荐品牌","推荐型号","推荐品牌1","推荐型号1","推荐品牌2","推荐型号2","推荐品牌3","推荐型号3",
        "信昌料号","华科料号","前5个其他品牌型号","其他品牌型号",
        "器件类别","系列","系列说明","尺寸（inch）","封装代码","尺寸(mm)","长度（mm）","宽度（mm）","高度（mm）","直径（mm）",
        "材质（介质）","容值","容值单位","工作温度","寿命（h）","安装方式","特殊用途",
        "备注1","备注2","备注3","规格参数明细","匹配参数明细",
        "功率","脚距","安规","规格","压敏电压","生产状态","极性","ESR","纹波电流",
        "阻值@25C","阻值单位","阻值误差","B值","B值条件",
        "共模阻抗","阻抗单位","额定电流","DCR","回路数","电感值","电感单位","电感误差","饱和电流","屏蔽类型","阻抗@100MHz",
        "负载电容（pF）","驱动电平","输出类型","占空比",
    ]:
        if col in show_df.columns:
            show_df[col] = show_df[col].astype(str).replace("nan", "").replace("None", "")
    for dim_col in ["长度（mm）", "宽度（mm）", "高度（mm）", "直径（mm）"]:
        if dim_col in show_df.columns:
            show_df[dim_col] = show_df[dim_col].apply(normalize_dimension_mm_value)
    if "容值" in show_df.columns:
        show_df["容值"] = show_df["容值"].apply(normalize_numeric_range_text)
    if "容值单位" in show_df.columns:
        show_df["容值单位"] = show_df["容值单位"].apply(normalize_library_value_unit)
    if "容值误差" in show_df.columns:
        show_df["容值误差"] = show_df["容值误差"].apply(clean_tol_for_display)
    if "耐压（V）" in show_df.columns:
        show_df["耐压（V）"] = show_df["耐压（V）"].apply(voltage_display)
    if "工作温度" in show_df.columns:
        show_df["工作温度"] = show_df["工作温度"].apply(normalize_working_temperature_text)
    if "寿命（h）" in show_df.columns:
        show_df["寿命（h）"] = show_df["寿命（h）"].apply(format_life_hours_display)
    if "安装方式" in show_df.columns:
        show_df["安装方式"] = show_df["安装方式"].apply(normalize_mounting_style)
    if "特殊用途" in show_df.columns:
        show_df["特殊用途"] = show_df["特殊用途"].apply(normalize_special_use)
    if "脚距" in show_df.columns:
        show_df["脚距"] = show_df["脚距"].apply(format_pitch_display)
    if "压敏电压" in show_df.columns:
        show_df["压敏电压"] = show_df["压敏电压"].apply(voltage_display)
    if "功率" in show_df.columns:
        show_df["功率"] = show_df["功率"].apply(format_power_display)
    if "阻值单位" in show_df.columns:
        show_df["阻值单位"] = show_df["阻值单位"].apply(normalize_library_value_unit)
    if "阻值误差" in show_df.columns:
        show_df["阻值误差"] = show_df["阻值误差"].apply(clean_tol_for_display)
    if "B值" in show_df.columns:
        show_df["B值"] = show_df["B值"].apply(normalize_b_value_text)
    if "阻抗单位" in show_df.columns:
        show_df["阻抗单位"] = show_df["阻抗单位"].apply(normalize_library_value_unit)
    if "电感单位" in show_df.columns:
        show_df["电感单位"] = show_df["电感单位"].apply(lambda value: clean_text(value).upper())
    if "电感误差" in show_df.columns:
        show_df["电感误差"] = show_df["电感误差"].apply(clean_tol_for_display)
    for current_col in ["纹波电流", "额定电流", "饱和电流"]:
        if current_col in show_df.columns:
            show_df[current_col] = show_df[current_col].apply(format_current_display)
    if "推荐等级" in show_df.columns:
        show_df["推荐等级"] = show_df["推荐等级"].astype(str).replace("nan", "").replace("None", "")
    return show_df


def annotate_match_display_gaps(show_df, spec):
    out = show_df.copy()
    tol_hint = clean_tol_for_display(spec.get("容值误差", ""))
    volt_hint = voltage_display(spec.get("耐压（V）", ""))
    temp_hint = normalize_working_temperature_text(spec.get("工作温度", ""))
    life_hint = format_life_hours_display(spec.get("寿命（h）", ""))
    power_hint = format_power_display(spec.get("_power", ""))
    pitch_hint = clean_text(spec.get("_pitch", ""))
    mounting_hint = normalize_mounting_style(spec.get("安装方式", ""))
    special_use_hint = normalize_special_use(spec.get("特殊用途", ""))
    safety_hint = clean_text(spec.get("_safety_class", ""))
    varistor_hint = voltage_display(spec.get("_varistor_voltage", spec.get("耐压（V）", "")))
    body_size_hint = clean_text(spec.get("_body_size", spec.get("尺寸（mm）", "")))
    spec_hint = clean_text(spec.get("_disc_size", "")) or body_size_hint

    if "容值误差" in out.columns and tol_hint != "":
        tol_mask = out["容值误差"].astype(str).apply(clean_text).eq("")
        out.loc[tol_mask, "容值误差"] = f"{tol_hint}（库中缺失）"

    if "耐压（V）" in out.columns and volt_hint != "":
        volt_mask = out["耐压（V）"].astype(str).apply(clean_text).eq("")
        out.loc[volt_mask, "耐压（V）"] = f"{volt_hint}（库中缺失）"

    if "工作温度" in out.columns and temp_hint != "":
        temp_mask = out["工作温度"].astype(str).apply(clean_text).eq("")
        out.loc[temp_mask, "工作温度"] = f"{temp_hint}（库中缺失）"

    if "寿命（h）" in out.columns and life_hint != "":
        life_mask = out["寿命（h）"].astype(str).apply(clean_text).eq("")
        out.loc[life_mask, "寿命（h）"] = f"{life_hint}（库中缺失）"

    if "功率" in out.columns and power_hint != "":
        power_mask = out["功率"].astype(str).apply(clean_text).eq("")
        out.loc[power_mask, "功率"] = f"{power_hint}（库中缺失）"

    if "安装方式" in out.columns and mounting_hint != "":
        mount_mask = out["安装方式"].astype(str).apply(clean_text).eq("")
        out.loc[mount_mask, "安装方式"] = f"{mounting_hint}（库中缺失）"

    if "特殊用途" in out.columns and special_use_hint != "":
        special_mask = out["特殊用途"].astype(str).apply(clean_text).eq("")
        out.loc[special_mask, "特殊用途"] = f"{special_use_hint}（库中缺失）"

    if "脚距" in out.columns and pitch_hint != "":
        pitch_mask = out["脚距"].astype(str).apply(clean_text).eq("")
        out.loc[pitch_mask, "脚距"] = f"{pitch_hint}（库中缺失）"

    if "安规" in out.columns and safety_hint != "":
        safety_mask = out["安规"].astype(str).apply(clean_text).eq("")
        out.loc[safety_mask, "安规"] = f"{safety_hint}（库中缺失）"

    if "压敏电压" in out.columns and varistor_hint != "":
        varistor_mask = out["压敏电压"].astype(str).apply(clean_text).eq("")
        out.loc[varistor_mask, "压敏电压"] = f"{varistor_hint}（库中缺失）"

    if "尺寸(mm)" in out.columns and body_size_hint != "":
        body_mask = out["尺寸(mm)"].astype(str).apply(clean_text).eq("")
        out.loc[body_mask, "尺寸(mm)"] = f"{body_size_hint}（库中缺失）"

    if "规格" in out.columns and spec_hint != "":
        spec_mask = out["规格"].astype(str).apply(clean_text).eq("")
        out.loc[spec_mask, "规格"] = f"{spec_hint}（库中缺失）"

    return out


def matched_display_columns(row, spec):
    if spec is None:
        return set()

    schema_columns = {source for source, _ in get_component_display_schema(spec)}
    hits = set()

    spec_size = clean_size(spec.get("尺寸（inch）", ""))
    if "尺寸（inch）" in schema_columns and spec_size and clean_size(row.get("尺寸（inch）", "")) == spec_size:
        hits.add("尺寸（inch）")

    spec_mat = clean_material(spec.get("材质（介质）", ""))
    if "材质（介质）" in schema_columns and spec_mat and clean_material(row.get("材质（介质）", "")) == spec_mat:
        hits.add("材质（介质）")

    spec_value, spec_unit = spec_display_value_unit(spec)
    if "容值" in schema_columns and spec_value != "" and clean_text(row.get("容值", "")) == clean_text(spec_value):
        hits.add("容值")
    if "容值单位" in schema_columns and spec_unit != "" and clean_text(row.get("容值单位", "")).upper() == clean_text(spec_unit).upper():
        hits.add("容值单位")

    spec_tol = clean_tol_for_display(spec.get("容值误差", ""))
    if "容值误差" in schema_columns and spec_tol != "" and clean_text(row.get("容值误差", "")) == spec_tol:
        hits.add("容值误差")

    spec_volt = voltage_display(spec.get("耐压（V）", ""))
    if "耐压（V）" in schema_columns and spec_volt != "" and clean_text(row.get("耐压（V）", "")) == spec_volt:
        hits.add("耐压（V）")

    spec_temp = normalize_working_temperature_text(spec.get("工作温度", ""))
    if "工作温度" in schema_columns and spec_temp != "" and normalize_working_temperature_text(row.get("工作温度", "")) == spec_temp:
        hits.add("工作温度")

    spec_life = format_life_hours_display(spec.get("寿命（h）", ""))
    if "寿命（h）" in schema_columns and spec_life != "" and format_life_hours_display(row.get("寿命（h）", "")) == spec_life:
        hits.add("寿命（h）")

    spec_power = format_power_display(spec.get("_power", ""))
    if "功率" in schema_columns and spec_power != "" and clean_text(row.get("功率", "")) == spec_power:
        hits.add("功率")

    spec_mount = normalize_mounting_style(spec.get("安装方式", ""))
    if "安装方式" in schema_columns and spec_mount != "" and normalize_mounting_style(row.get("安装方式", "")) == spec_mount:
        hits.add("安装方式")

    spec_special_use = normalize_special_use(spec.get("特殊用途", ""))
    if "特殊用途" in schema_columns and spec_special_use != "" and special_use_matches(row.get("特殊用途", ""), spec_special_use):
        hits.add("特殊用途")

    spec_pitch = clean_text(spec.get("_pitch", ""))
    if "脚距" in schema_columns and spec_pitch != "" and clean_text(row.get("脚距", "")) == spec_pitch:
        hits.add("脚距")

    spec_safety = clean_text(spec.get("_safety_class", ""))
    if "安规" in schema_columns and spec_safety != "" and clean_text(row.get("安规", "")) == spec_safety:
        hits.add("安规")

    spec_varistor_voltage = voltage_display(spec.get("_varistor_voltage", spec.get("耐压（V）", "")))
    if "压敏电压" in schema_columns and spec_varistor_voltage != "" and clean_text(row.get("压敏电压", "")) == spec_varistor_voltage:
        hits.add("压敏电压")

    spec_body_size = clean_text(spec.get("_body_size", spec.get("尺寸（mm）", "")))
    if "尺寸(mm)" in schema_columns and spec_body_size != "" and clean_text(row.get("尺寸(mm)", "")) == spec_body_size:
        hits.add("尺寸(mm)")

    spec_disc_size = clean_text(spec.get("_disc_size", "")) or spec_body_size
    if "规格" in schema_columns and spec_disc_size != "" and clean_text(row.get("规格", "")) == spec_disc_size:
        hits.add("规格")

    return hits


def extract_official_url(value):
    text = clean_text(value)
    if text == "":
        return ""
    lower = text.lower()
    for prefix in ["official_url=", "链接=", "url=", "网址="]:
        if lower.startswith(prefix.lower()):
            text = clean_text(text.split("=", 1)[1] if "=" in text else "")
            break
    if not (text.startswith("https://") or text.startswith("http://")):
        return ""
    if "product.samsungsem.com/mlcc/" in text and not text.endswith(".do"):
        text = text + ".do"
    return text


def extract_official_status(value):
    text = clean_text(value)
    if text == "":
        return ""
    lower = text.lower()
    if lower.startswith("source=") or lower.startswith("来源="):
        return ""
    if "status=" in lower:
        match = re.search(r"status\s*=\s*([^;]+)", text, flags=re.I)
        if match:
            return clean_text(match.group(1))
    if lower.startswith("verified_at=") or lower.startswith("校验时间="):
        return ""
    return text


def looks_like_official_status_text(value):
    text = clean_text(value)
    if text == "":
        return False
    upper = text.upper()
    if "PRODUCTION" in upper:
        return True
    if upper in {"NRND", "EOL"}:
        return True
    if "量产" in text:
        return True
    if "STATUS=" in upper:
        return True
    return False


def append_official_status_column(show_df):
    out = show_df.copy()
    statuses = []
    urls = []
    for row_idx, row in out.iterrows():
        raw_note1 = row.get("备注1", "")
        raw_note3 = row.get("备注3", "")
        status = ""
        if looks_like_official_status_text(raw_note1):
            status = extract_official_status(raw_note1)
            if "备注1" in out.columns:
                out.at[row_idx, "备注1"] = ""
        if status == "" and looks_like_official_status_text(raw_note3):
            status = extract_official_status(raw_note3)
        url = extract_official_url(row.get("备注2", ""))
        if url == "":
            url = extract_official_url(row.get("备注3", ""))
        if status == "" and url != "":
            status = "官方详情"
        statuses.append(status)
        urls.append(url)
    if any(statuses) or any(urls):
        insert_at = out.columns.get_loc("特殊用途") + 1 if "特殊用途" in out.columns else len(out.columns)
        out.insert(insert_at, "量产状态", statuses)
        out["_量产状态链接"] = urls
    return out


def result_row_level(row):
    return clean_text(row.get("推荐等级", "")) or clean_text(row.get("首选推荐等级", ""))


def split_model_list(value):
    text = clean_text(value)
    if text == "":
        return []
    items = []
    seen = set()
    for piece in re.split(r"\s*(?:\||\r?\n)+\s*", text):
        model = clean_text(piece)
        if model == "" or model in seen:
            continue
        seen.add(model)
        items.append(model)
    return items


def split_brand_model_list(value):
    items = split_model_list(value)
    parsed = []
    seen = set()
    for item in items:
        brand = ""
        model = clean_text(item)
        if ":" in item:
            brand_part, model_part = item.split(":", 1)
            brand = clean_brand(brand_part)
            model = clean_text(model_part)
        if model == "":
            continue
        key = (brand, model)
        if key in seen:
            continue
        seen.add(key)
        parsed.append((brand, model))
    return parsed


def render_grouped_brand_model_cell(value, preview_count=2):
    items = split_brand_model_list(value)
    if not items:
        return "&nbsp;"

    grouped = []
    brand_index = {}
    for brand, model in items:
        if brand not in brand_index:
            brand_index[brand] = len(grouped)
            grouped.append({"brand": brand, "models": [model]})
        else:
            grouped[brand_index[brand]]["models"].append(model)

    def render_group_body():
        blocks = []
        for group in grouped:
            brand = html.escape(group["brand"])
            model_lines = "".join(
                f'<div class="brand-model-item">{html.escape(model)}</div>'
                for model in group["models"]
            )
            if brand != "":
                blocks.append(
                    f'<div class="brand-model-group">'
                    f'<div class="brand-model-brand">{brand}:</div>'
                    f'<div class="brand-model-items">{model_lines}</div>'
                    f'</div>'
                )
            else:
                blocks.append(
                    f'<div class="brand-model-group">'
                    f'<div class="brand-model-items">{model_lines}</div>'
                    f'</div>'
                )
        return "".join(blocks)

    if len(items) <= 1:
        return f'<div class="multi-model-cell brand-model-cell">{render_group_body()}</div>'

    summary = f"共 {len(items)} 个，点击展开"
    return (
        f'<details class="multi-model-details">'
        f'<summary>{html.escape(summary)}</summary>'
        f'<div class="multi-model-cell brand-model-cell">{render_group_body()}</div>'
        f"</details>"
    )


def render_model_list_cell(value, preview_count=2, grouped=False):
    if grouped:
        return render_grouped_brand_model_cell(value, preview_count=preview_count)
    items = split_model_list(value)
    if not items:
        return "&nbsp;"
    if len(items) <= 1:
        return "<div class=\"multi-model-cell\">" + "<br>".join(html.escape(item) for item in items) + "</div>"

    summary = f"共 {len(items)} 个，点击展开"
    full_list = "<br>".join(html.escape(item) for item in items)
    return (
        f'<details class="multi-model-details">'
        f'<summary>{html.escape(summary)}</summary>'
        f'<div class="multi-model-cell">{full_list}</div>'
        f"</details>"
    )


def estimate_result_table_column_width(col, values, header_label):
    col_text = clean_text(col)
    header_text = clean_text(header_label)
    if col_text in {
        clean_text("信昌料号"),
        clean_text("华科料号"),
        clean_text("前5个其他品牌型号"),
        clean_text("其他品牌型号"),
    }:
        return estimate_model_list_width(values)

    sample_lengths = [display_text_width(header_text)]
    for value in values:
        text = clean_text(value)
        if text != "":
            sample_lengths.append(display_text_width(text))
    longest = max(sample_lengths) if sample_lengths else display_text_width(header_text)
    width = longest * 8.4 + 24
    if col_text in {clean_text("推荐等级"), clean_text("品牌"), clean_text("推荐品牌"), clean_text("品牌1"), clean_text("推荐品牌1"), clean_text("品牌2"), clean_text("推荐品牌2"), clean_text("品牌3"), clean_text("推荐品牌3"), clean_text("系列"), clean_text("容值单位")}:
        width = max(width, 92)
    if col_text in {clean_text("型号"), clean_text("推荐型号"), clean_text("型号1"), clean_text("推荐型号1"), clean_text("型号2"), clean_text("推荐型号2"), clean_text("型号3"), clean_text("推荐型号3")}:
        width = max(width, 112)
    if col_text in {clean_text("尺寸（inch）"), clean_text("容值"), clean_text("容值误差"), clean_text("耐压（V）")}:
        width = max(width, 92)
    if col_text in {clean_text("BOM规格"), clean_text("BOM品名"), clean_text("其他品牌型号"), clean_text("关键规格"), clean_text("差异说明"), clean_text("解析输入"), clean_text("规格参数明细"), clean_text("匹配参数明细")}:
        width = max(width, 160)
    if col_text in {clean_text("规格参数明细"), clean_text("匹配参数明细")}:
        width = max(width, 200)
    return max(78, min(width, 320))


def estimate_model_list_width(values, min_width=90, char_width=7.4, max_width=150, preview_count=2):
    if values is None:
        return min_width
    if not isinstance(values, (list, tuple, pd.Series, pd.Index)):
        values = [values]

    single_like_max = 0
    has_multi = False

    for value in values:
        items = split_model_list(value)
        items = [clean_text(item) for item in items if clean_text(item) != ""]
        if not items:
            text = clean_text(value)
            if text != "":
                single_like_max = max(single_like_max, display_text_width(text))
            continue

        if len(items) > preview_count:
            has_multi = True
        else:
            single_like_max = max(
                single_like_max,
                max(display_text_width(item) for item in items),
            )

    if single_like_max == 0 and not has_multi:
        return min_width

    if has_multi:
        summary_text = f"共 {max(3, preview_count + 1)} 个，点击展开"
        width = display_text_width(summary_text) * 7.8 + 16
    else:
        width = single_like_max * char_width + 18

    return max(min_width, min(width, max_width))


def estimate_result_table_iframe_height(row_count, show_official_status=True, compact=False):
    row_count = max(0, int(row_count or 0))
    use_compact = compact or row_count <= 12
    visible_rows = min(max(row_count, 1), 10)
    if use_compact:
        base = 80 if show_official_status else 72
        per_row = 48
        min_height = 160
        max_height = 560
    else:
        base = 92 if show_official_status else 84
        per_row = 42
        min_height = 180
        max_height = 600
    height = base + visible_rows * per_row
    return max(min_height, min(max_height, height))

def estimate_bom_result_iframe_height(row_count):
    row_count = max(0, int(row_count or 0))
    visible_rows = min(max(row_count, 1), 10)
    base = 92
    per_row = 60
    min_height = 320
    max_height = 700
    height = base + visible_rows * per_row
    return max(min_height, min(max_height, height))


def estimate_bom_preview_iframe_height(row_count):
    row_count = max(0, int(row_count or 0))
    visible_rows = min(max(row_count, 1), 5)
    base = 86
    per_row = 42
    min_height = 220
    max_height = 320
    height = base + visible_rows * per_row
    return max(min_height, min(max_height, height))


def estimate_match_card_iframe_height(part_row_count, result_row_count):
    part_row_count = max(0, int(part_row_count or 0))
    result_row_count = max(0, int(result_row_count or 0))
    part_visible_rows = max(1, min(part_row_count, 2))
    result_visible_rows = max(1, min(result_row_count, 10))
    # Keep the part-info table compact while letting the result table expose about
    # ten visible rows before scrolling internally, then reserve extra space for
    # a visible rounded footer so the single-model bubble closes cleanly.
    part_height = 124 + (part_visible_rows - 1) * 34
    footer_height = 28
    chrome_height = 28
    if result_row_count <= 0:
        compact_total = part_height + footer_height + chrome_height - 18
        return max(178, min(210, compact_total))
    result_height = 118 + result_visible_rows * 34
    total = part_height + result_height + footer_height + chrome_height
    return max(576, min(676, total))


def paginate_result_dataframe(df, table_key, page_size=12):
    if df is None or df.empty:
        return df, None
    total_rows = len(df)
    if total_rows <= page_size:
        return df, None

    page_count = int(math.ceil(total_rows / float(page_size)))
    safe_key = re.sub(r"[^A-Za-z0-9_]+", "_", clean_text(table_key)) or "table"
    page_widget_key = f"{safe_key}_page"

    current_page = 1
    try:
        current_page = int(st.session_state.get(page_widget_key, 1))
    except Exception:
        current_page = 1
    current_page = min(max(current_page, 1), page_count)

    control_cols = st.columns([1, 3])
    with control_cols[0]:
        selected_page = st.selectbox(
            "分页",
            list(range(1, page_count + 1)),
            index=current_page - 1,
            key=page_widget_key,
            label_visibility="collapsed",
        )
    with control_cols[1]:
        start = (selected_page - 1) * page_size + 1
        end = min(selected_page * page_size, total_rows)
        st.caption(f"当前第 {selected_page}/{page_count} 页，显示第 {start}-{end} 行，共 {total_rows} 行")

    start_idx = (selected_page - 1) * page_size
    end_idx = start_idx + page_size
    return df.iloc[start_idx:end_idx].copy(), (selected_page, page_count)


def build_result_table_iframe_html(table_fragment, outer_footer_html=""):
    return f"""
<style>
html, body {{
    margin: 0;
    padding: 0;
    font-family: "Segoe UI", "Microsoft YaHei", sans-serif;
    background: #ffffff;
}}
.result-table-wrap {{
    overflow: auto;
    max-height: min(560px, 52vh);
    margin-bottom: 0;
    position: relative;
}}
.bom-result-table-wrap {{
    overflow: auto;
    max-height: min(560px, 52vh);
    margin-bottom: 0;
    position: relative;
}}
.result-section-card {{
    display: flex;
    flex-direction: column;
    gap: 0;
    padding: 4px;
    border: 1px solid rgba(191, 219, 254, 0.90);
    border-radius: 18px;
    background: linear-gradient(180deg, #ffffff 0%, #fbfdff 100%);
    box-shadow: 0 10px 24px rgba(15, 23, 42, 0.06);
    box-sizing: border-box;
}}
.result-section-card .result-table-wrap,
.result-section-card .bom-result-table-wrap {{
    max-height: min(560px, 52vh);
    margin-bottom: 0;
}}
.result-table {{
    width: max-content;
    min-width: max-content;
    border-collapse: collapse;
    font-size: 14px;
    background: #ffffff;
    table-layout: fixed;
}}
.result-table th,
.result-table td {{
    border: 1px solid #e6e6e6;
    padding: 8px 10px;
    text-align: center;
    vertical-align: middle;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    box-sizing: border-box;
}}
.result-table th {{
    background: #f7f8fa;
    font-weight: 600;
    position: sticky;
    top: 0;
    z-index: 6;
    box-shadow: 0 1px 0 #e6e6e6;
}}
.result-table td.model-list-cell,
.result-table th.model-list-cell {{
    overflow: visible;
    text-overflow: clip;
    text-align: left;
    vertical-align: middle;
}}
.result-table td.detail-cell,
.result-table th.detail-cell {{
    white-space: normal;
    overflow: visible;
    text-overflow: clip;
    text-align: left;
    vertical-align: middle;
}}
.result-table td.model-list-cell {{
    vertical-align: middle;
}}
.result-table td.detail-cell .detail-cell-text {{
    white-space: nowrap;
    line-height: 1.45;
    display: inline-block;
}}
.result-table td.model-list-cell summary {{
    cursor: pointer;
    color: #1565c0;
    font-weight: 600;
    outline: none;
    white-space: nowrap;
}}
.result-table td.model-list-cell summary::-webkit-details-marker {{
    display: none;
}}
.result-table td.model-list-cell .multi-model-cell {{
    max-height: 96px;
    overflow-y: auto;
    overflow-x: auto;
    white-space: nowrap;
    word-break: normal;
    overflow-wrap: normal;
    line-height: 1.35;
}}
.result-table td.model-list-cell details[open] .multi-model-cell {{
    max-height: 260px;
}}
.result-table td.model-list-cell .brand-model-cell {{
    white-space: normal;
}}
.result-table td.model-list-cell .brand-model-group {{
    margin-bottom: 6px;
}}
.result-table td.model-list-cell .brand-model-brand {{
    font-weight: 700;
    white-space: nowrap;
    margin-bottom: 2px;
}}
.result-table td.model-list-cell .brand-model-items {{
    padding-left: 0;
}}
.result-table td.model-list-cell .brand-model-item {{
    white-space: normal;
    overflow-wrap: anywhere;
}}
.result-table tr.exact-match td {{
    background: #fff59d;
    color: #111111;
}}
.result-table tr.partial-match-row td {{
    background: #fde2e1;
    color: #7f1d1d;
}}
.result-table tr.substitute-row td {{
    background: #dbeafe;
    color: #1d4ed8;
}}
.result-table tr.parse-fail-row td {{
    background: #fde2e1;
}}
.result-table tr.warn-row td {{
    background: #fff4db;
}}
.result-table td.param-hit {{
    color: #c62828;
    font-weight: 600;
}}
.result-table a {{
    color: #1565c0;
    text-decoration: none;
}}
.result-table a:hover {{
    text-decoration: underline;
}}
.result-table-footer {{
    display: flex;
    justify-content: flex-end;
    align-items: center;
    gap: 10px;
    margin-top: 6px;
    padding: 0 2px 2px 2px;
}}
.bom-download-footer-outside {{
    display: flex;
    justify-content: flex-end;
    align-items: center;
    gap: 10px;
    margin-top: 0;
    padding: 2px 2px 0 2px;
}}
.match-card-footer {{
    height: 28px;
    margin-top: 6px;
    border-top: 1px solid rgba(191, 219, 254, 0.58);
    border-radius: 0 0 16px 16px;
    background: linear-gradient(180deg, rgba(255, 255, 255, 0.98) 0%, rgba(247, 250, 255, 0.94) 65%, rgba(235, 244, 255, 0.84) 100%);
}}
.no-alt-match-alert {{
    margin: 6px 8px 0 8px;
    padding: 16px 22px;
    border-radius: 14px;
    border: 1px solid rgba(147, 197, 253, 0.88);
    background: linear-gradient(180deg, rgba(239, 246, 255, 0.98) 0%, rgba(219, 234, 254, 0.96) 100%);
    color: #0f4c9a;
    font-size: 16px;
    font-weight: 600;
    line-height: 1.4;
    box-shadow: 0 6px 14px rgba(37, 99, 235, 0.08);
}}
.result-query-inline {{
    padding: 6px 2px 0 2px;
    font-size: 18px;
    font-weight: 700;
    color: #1f2937;
    line-height: 1.35;
}}
.bom-download-btn {{
    display: inline-flex;
    align-items: center;
    justify-content: center;
    min-width: 190px;
    padding: 12px 18px;
    border-radius: 12px;
    border: 1px solid rgba(203, 213, 225, 0.98);
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
    color: #1f2937;
    font-size: 16px;
    font-weight: 700;
    text-decoration: none;
    box-shadow: 0 8px 18px rgba(31, 41, 55, 0.08);
}}
.bom-download-btn:hover {{
    border-color: rgba(148, 163, 184, 0.95);
    box-shadow: 0 12px 24px rgba(31, 41, 55, 0.12);
    transform: translateY(-1px);
    text-decoration: none;
}}
.bom-download-btn:active {{
    transform: translateY(0);
    box-shadow: 0 6px 12px rgba(31, 41, 55, 0.08);
}}
.bom-download-hint {{
    color: #64748b;
    font-size: 13px;
    line-height: 1.4;
}}
.col-resizer {{
    position: absolute;
    top: 0;
    right: -7px;
    width: 14px;
    height: 100%;
    cursor: col-resize;
    user-select: none;
    touch-action: none;
    z-index: 12;
}}
.col-resizer:hover,
.col-resizer.active {{
    background: rgba(21, 101, 192, 0.12);
}}
.col-resizer::after {{
    content: "";
    position: absolute;
    top: 0;
    left: 50%;
    width: 2px;
    height: 100%;
    transform: translateX(-50%);
    background: rgba(117, 117, 117, 0.35);
}}
</style>
<div class="result-section-card">
{table_fragment}
</div>
{outer_footer_html}
<script>
(function() {{
    function toNumber(value) {{
        var n = parseFloat(value);
        return Number.isFinite(n) ? n : 0;
    }}

    function getVisibleText(node) {{
        if (!node) {{
            return '';
        }}
        var summary = node.querySelector ? node.querySelector('summary') : null;
        if (summary) {{
            return (summary.innerText || summary.textContent || '').replace(/\\s+/g, ' ').trim();
        }}
        var modelCell = node.querySelector ? node.querySelector('.multi-model-cell') : null;
        if (modelCell) {{
            return (modelCell.innerText || modelCell.textContent || '').replace(/\\s+/g, ' ').trim();
        }}
        return (node.innerText || node.textContent || '').replace(/\\s+/g, ' ').trim();
    }}

    function measureTextWidth(ctx, text, font) {{
        ctx.font = font;
        return ctx.measureText(text || '').width;
    }}

    function initTable(table) {{
        var wrapper = table.closest('.result-table-wrap, .bom-result-table-wrap');
        if (!wrapper) {{
            return;
        }}
        var cols = Array.from(table.querySelectorAll('colgroup col'));
        var headers = Array.from(table.querySelectorAll('thead th'));
        if (!headers.length || !cols.length) {{
            return;
        }}

        function splitDisplayLines(text) {{
            return (text || '')
                .replace(/\\r/g, '')
                .split(/\\n+/)
                .map(function(line) {{
                    return line.replace(/\\s+/g, ' ').trim();
                }})
                .filter(function(line) {{
                    return line !== '';
                }});
        }}

        function applyWidth(index, width) {{
            var px = Math.max(84, Math.round(width));
            var col = cols[index];
            var header = headers[index];
            if (col) {{
                col.style.width = px + 'px';
                col.style.minWidth = px + 'px';
                col.style.maxWidth = px + 'px';
            }}
            if (header) {{
                header.style.width = px + 'px';
                header.style.minWidth = px + 'px';
                header.style.maxWidth = px + 'px';
            }}
            var total = cols.reduce(function(sum, item) {{
                return sum + toNumber(item.style.width || item.getAttribute('width') || item.getBoundingClientRect().width);
            }}, 0);
            table.style.width = Math.max(1, Math.round(total)) + 'px';
            table.style.minWidth = Math.max(1, Math.round(total)) + 'px';
        }}

        function measureCellWidth(ctx, cell, font) {{
            if (!cell) {{
                return 0;
            }}
            var details = cell.querySelector ? cell.querySelector('details') : null;
            var summary = cell.querySelector ? cell.querySelector('summary') : null;
            var multiModelCell = cell.querySelector ? cell.querySelector('.multi-model-cell') : null;
            if (summary && details && !details.open) {{
                var summaryText = (summary.innerText || summary.textContent || '').trim();
                return measureTextWidth(ctx, summaryText, font);
            }}
            if (multiModelCell) {{
                var lines = splitDisplayLines(multiModelCell.innerText || multiModelCell.textContent || '');
                if (!lines.length) {{
                    lines = splitDisplayLines(cell.innerText || cell.textContent || '');
                }}
                var maxWidth = 0;
                lines.forEach(function(line) {{
                    maxWidth = Math.max(maxWidth, measureTextWidth(ctx, line, font));
                }});
                return maxWidth;
            }}
            if (summary) {{
                var summaryText2 = (summary.innerText || summary.textContent || '').trim();
                return measureTextWidth(ctx, summaryText2, font);
            }}
            var lines = splitDisplayLines(cell.innerText || cell.textContent || '');
            var maxWidth = 0;
            lines.forEach(function(line) {{
                maxWidth = Math.max(maxWidth, measureTextWidth(ctx, line, font));
            }});
            return maxWidth;
        }}

    function autoFitWidths() {{
            var canvas = document.createElement('canvas');
            var ctx = canvas.getContext && canvas.getContext('2d');
            if (!ctx) {{
                return;
            }}
            var bodyFont = '14px "Segoe UI", "Microsoft YaHei", sans-serif';
            var headerFont = '600 14px "Segoe UI", "Microsoft YaHei", sans-serif';
            var minWidthMap = {{
                '品牌': 92,
                '推荐等级': 92,
                '系列': 92,
                '容值单位': 92,
                '阻值单位': 92,
                '状态': 92,
                '器件类型': 92,
                '尺寸（inch）': 92,
                '尺寸(mm)': 92,
                '长度(mm)': 108,
                '宽度(mm)': 108,
                '厚度(mm)': 108,
                '容值': 92,
                '阻值': 92,
                '容值误差': 92,
                '误差': 92,
                '耐压（V）': 92,
                '压敏电压': 92,
                '功率': 92,
                '脚距': 92,
                '安规': 92,
                '规格': 92,
                '匹配数量': 92,
                '推荐品牌': 92,
                '推荐型号': 112,
                '推荐品牌1': 92,
                '推荐型号1': 112,
                '推荐品牌2': 92,
                '推荐型号2': 112,
                '推荐品牌3': 92,
                '推荐型号3': 112,
                '信昌料号': 108,
                '华科料号': 108,
                '其他品牌型号': 112,
                'BOM规格': 160,
                'BOM品名': 160,
                '关键规格': 160,
                '规格参数明细': 200,
                '匹配参数明细': 200,
                '差异说明': 160,
                '解析输入': 160
            }};

            headers.forEach(function(th, index) {{
                var headerText = getVisibleText(th);
                var maxWidth = measureTextWidth(ctx, headerText, headerFont);
                table.querySelectorAll('tbody tr').forEach(function(tr) {{
                    var cell = tr.children[index];
                    if (!cell) {{
                        return;
                    }}
                    var cellWidth = measureCellWidth(ctx, cell, bodyFont);
                    if (cellWidth) {{
                        maxWidth = Math.max(maxWidth, cellWidth);
                    }}
                }});
                var minWidth = minWidthMap[headerText] || 78;
                applyWidth(index, Math.max(minWidth, Math.ceil(maxWidth + 24)));
            }});
        }}

        function reportFrameHeight() {{
            try {{
                if (window.Streamlit && typeof window.Streamlit.setFrameHeight === 'function') {{
                    var height = Math.max(0, Math.ceil(document.documentElement.scrollHeight || document.body.scrollHeight || 0));
                    window.Streamlit.setFrameHeight(height);
                }}
            }} catch (err) {{
                // Keep the iframe usable even if the host API is unavailable.
            }}
        }}

        headers.forEach(function(th, index) {{
            var handle = document.createElement('span');
            handle.className = 'col-resizer';
            th.appendChild(handle);

            var startX = 0;
            var startWidth = 0;
            var active = false;

            function onMove(event) {{
                if (!active) {{
                    return;
                }}
                applyWidth(index, startWidth + (event.clientX - startX));
            }}

            function onUp() {{
                if (!active) {{
                    return;
                }}
                active = false;
                document.removeEventListener('mousemove', onMove);
                document.removeEventListener('mouseup', onUp);
                handle.classList.remove('active');
            }}

            handle.addEventListener('mousedown', function(event) {{
                event.preventDefault();
                event.stopPropagation();
                active = true;
                startX = event.clientX;
                startWidth = th.getBoundingClientRect().width;
                handle.classList.add('active');
                document.addEventListener('mousemove', onMove);
                document.addEventListener('mouseup', onUp);
            }});
        }});

        autoFitWidths();
        reportFrameHeight();
        window.setTimeout(reportFrameHeight, 0);
        window.setTimeout(reportFrameHeight, 120);
        window.addEventListener('resize', reportFrameHeight);

        var initialTotal = cols.reduce(function(sum, item, index) {{
            return sum + toNumber(item.style.width || headers[index].getBoundingClientRect().width);
        }}, 0);
        table.style.width = Math.max(1, Math.round(initialTotal)) + 'px';
        table.style.minWidth = Math.max(1, Math.round(initialTotal)) + 'px';
        reportFrameHeight();
    }}

    document.querySelectorAll('table.result-table').forEach(initTable);
    window.addEventListener('load', function() {{
        window.setTimeout(function() {{
            if (window.Streamlit && typeof window.Streamlit.setFrameHeight === 'function') {{
                window.Streamlit.setFrameHeight(Math.max(0, Math.ceil(document.documentElement.scrollHeight || document.body.scrollHeight || 0)));
            }}
        }}, 80);
    }});
}})();
</script>
"""


def render_clickable_result_table(show_df, hide_columns=None, wrapper_class="result-table-wrap", spec=None, show_official_status=True, footer_html="", outer_footer_html="", wrap_iframe=True):
    if show_official_status:
        display_df = append_official_status_column(show_df)
        if "_量产状态链接" not in display_df.columns:
            display_df["_量产状态链接"] = ""
    else:
        display_df = show_df.copy()

    hidden = set(hide_columns or [])
    visible_columns = [col for col in display_df.columns if col != "_量产状态链接" and col not in hidden]
    header_labels = get_component_header_labels(spec)
    multi_model_columns = {"信昌料号", "华科料号", "前5个其他品牌型号", "其他品牌型号"}
    grouped_brand_model_columns = {"前5个其他品牌型号", "其他品牌型号"}
    detail_columns = {"规格参数明细", "匹配参数明细"}
    multi_model_columns_norm = {clean_text(col) for col in multi_model_columns}
    grouped_brand_model_columns_norm = {clean_text(col) for col in grouped_brand_model_columns}
    detail_columns_norm = {clean_text(col) for col in detail_columns}
    column_widths = {}
    for col in visible_columns:
        sample_values = display_df[col].fillna("").astype(str).head(20).tolist() if col in display_df.columns else []
        column_widths[col] = estimate_result_table_column_width(col, sample_values, header_labels.get(col, col))
    parts = [f'<div class="{html.escape(wrapper_class, quote=True)}"><table class="result-table"><colgroup>']
    for col in visible_columns:
        width = column_widths.get(col, 120)
        parts.append(f'<col style="width: {width}px;" />')
    parts.append('</colgroup><thead><tr>')
    for col in visible_columns:
        norm_col = clean_text(col)
        matched_width = column_widths.get(col, 120)
        th_style = f' style="width: {matched_width}px; min-width: {matched_width}px; max-width: {matched_width}px;"'
        th_class = ""
        if norm_col in multi_model_columns_norm:
            th_class = ' class="model-list-cell"'
        elif norm_col in detail_columns_norm:
            th_class = ' class="detail-cell"'
        parts.append(f'<th{th_style}{th_class}>{html.escape(header_labels.get(col, col))}</th>')
    parts.append("</tr></thead><tbody>")

    for _, row in display_df.iterrows():
        row_class_name = ""
        level = result_row_level(row)
        if clean_text(row.get("解析状态", "")) == "解析失败":
            row_class_name = "parse-fail-row"
        elif level == "完全匹配":
            row_class_name = "exact-match"
        elif level == "部分参数匹配":
            row_class_name = "partial-match-row"
        elif level in {"高代低", "可直接替代"}:
            row_class_name = "substitute-row"
        row_class = f' class="{row_class_name}"' if row_class_name else ""
        parts.append(f"<tr{row_class}>")
        official_link = clean_text(row.get("_量产状态链接", "")) if show_official_status else ""
        if level == "部分参数匹配":
            raw_hit_columns = clean_text(row.get("匹配命中列", ""))
            if raw_hit_columns != "":
                hit_columns = {clean_text(x) for x in raw_hit_columns.split("|") if clean_text(x) != ""}
            else:
                hit_columns = matched_display_columns(row, spec)
        else:
            hit_columns = set()
        for col in visible_columns:
            value = clean_text(row.get(col, ""))
            norm_col = clean_text(col)
            if norm_col in multi_model_columns_norm:
                cell = render_model_list_cell(value, grouped=norm_col in grouped_brand_model_columns_norm)
                cell_class = ' class="model-list-cell brand-model-cell"' if norm_col in grouped_brand_model_columns_norm else ' class="model-list-cell"'
                matched_width = column_widths.get(col, 240)
                td_style = f' style="width: {matched_width}px; min-width: {matched_width}px; max-width: {matched_width}px;"'
            elif norm_col in detail_columns_norm:
                cell = f'<div class="detail-cell-text">{html.escape(value).replace(chr(10), "<br>")}</div>' if value != "" else "&nbsp;"
                cell_class = ' class="detail-cell"'
                matched_width = column_widths.get(col, 220)
                td_style = f' style="width: {matched_width}px; min-width: {matched_width}px; max-width: {matched_width}px;"'
            elif col == "量产状态" and value != "" and official_link != "":
                cell = f'<a href="{html.escape(official_link, quote=True)}" target="_blank">{html.escape(value)}</a>'
                cell_class = ' class="param-hit"' if col in hit_columns else ""
                td_style = ""
            else:
                cell = html.escape(value) if value != "" else "&nbsp;"
                cell_class = ' class="param-hit"' if col in hit_columns else ""
                td_style = ""
            parts.append(f"<td{cell_class}{td_style}>{cell}</td>")
        parts.append("</tr>")

    parts.append("</tbody></table></div>")
    if footer_html:
        parts.append(footer_html)
    fragment = "".join(parts)
    if wrap_iframe:
        return build_result_table_iframe_html(fragment, outer_footer_html=outer_footer_html)
    return fragment


def render_static_preview_table(show_df, wrapper_class="result-table-wrap", outer_footer_html="", wrap_iframe=True):
    if show_df is None or show_df.empty:
        return ""

    display_df = show_df.copy()
    visible_columns = [clean_text(col) for col in display_df.columns]
    column_widths = {}
    for idx, col in enumerate(visible_columns):
        series = display_df.iloc[:, idx] if idx < display_df.shape[1] else pd.Series(dtype=object)
        sample_values = series.fillna("").astype(str).head(20).tolist()
        column_widths[idx] = estimate_result_table_column_width(col, sample_values, col)

    parts = [f'<div class="{html.escape(wrapper_class, quote=True)}"><table class="result-table"><colgroup>']
    for idx, _ in enumerate(visible_columns):
        width = column_widths.get(idx, 120)
        parts.append(f'<col style="width: {width}px;" />')
    parts.append("</colgroup><thead><tr>")
    for idx, col in enumerate(visible_columns):
        matched_width = column_widths.get(idx, 120)
        parts.append(
            f'<th style="width: {matched_width}px; min-width: {matched_width}px; max-width: {matched_width}px; text-align: left;">'
            f"{html.escape(col)}</th>"
        )
    parts.append("</tr></thead><tbody>")

    for _, row in display_df.iterrows():
        parts.append("<tr>")
        for idx, _ in enumerate(visible_columns):
            value = clean_text(row.iloc[idx] if idx < len(row) else "")
            cell = html.escape(value).replace(chr(10), "<br>") if value != "" else "&nbsp;"
            parts.append(f'<td style="text-align: left;">{cell}</td>')
        parts.append("</tr>")

    parts.append("</tbody></table></div>")
    fragment = "".join(parts)
    if wrap_iframe:
        return build_result_table_iframe_html(fragment, outer_footer_html=outer_footer_html)
    return fragment

def build_part_info_df(df, spec, query_model):
    if spec is None:
        return pd.DataFrame()
    hit = pd.DataFrame()
    if df is not None and not df.empty and "型号" in df.columns:
        hit = df[df["型号"].astype(str).apply(clean_model) == clean_model(query_model)].copy()
    if hit.empty:
        hit = load_component_rows_by_clean_model(query_model)
    if not hit.empty:
        show_df = ensure_component_display_columns(hit)
        show_df["材质（介质）"] = show_df["材质（介质）"].apply(clean_material)
        show_df["容值误差"] = show_df["容值误差"].apply(clean_tol_for_match)
        show_df["耐压（V）"] = show_df["耐压（V）"].apply(clean_voltage)
        show_df = fill_component_display_blanks(show_df, spec)
        prioritized_hit = prioritize_component_rows_for_lookup(show_df)
        display_target = prioritized_hit.iloc[0].to_dict() if prioritized_hit is not None and not prioritized_hit.empty else spec
        suffix_columns = ["官网链接", "数据来源"]
        display_component_type = resolve_component_display_type(display_target)
        if display_component_type in INDUCTOR_COMPONENT_TYPES:
            for extra_col in ["规格摘要", "特殊用途"]:
                if extra_col in show_df.columns:
                    has_value = (
                        show_df[extra_col]
                        .astype(str)
                        .replace("nan", "")
                        .replace("None", "")
                        .apply(clean_text)
                        .ne("")
                        .any()
                    )
                    if has_value and extra_col not in suffix_columns:
                        suffix_columns.insert(len(suffix_columns) - 2 if len(suffix_columns) >= 2 else 0, extra_col)
        allow_online_lookup = display_component_type == "MLCC" and len(show_df) <= 20
        show_df = select_component_display_columns(
            show_df,
            display_target,
            prefix_columns=["品牌", "型号", "器件类别"],
            suffix_columns=suffix_columns,
            allow_online_lookup=allow_online_lookup,
        )
        return format_display_df(show_df)
    row = pd.DataFrame([{
        "品牌": spec.get("品牌", ""),
        "型号": spec.get("型号", query_model),
        **build_component_display_row(spec),
    }])
    row_target = row.iloc[0].to_dict() if not row.empty else spec
    row = select_component_display_columns(
        row,
        row_target,
        prefix_columns=["品牌", "型号", "器件类别"],
        suffix_columns=["官网链接", "数据来源"],
    )
    return format_display_df(row)



def brand_priority_value(brand, component_type=""):
    b = clean_brand(brand).upper()
    ctype = normalize_component_type(component_type)
    if ctype in (RESISTOR_COMPONENT_TYPES | {"热敏电阻"}):
        priorities = [
            ("华新科", 1), ("WALSIN", 1), ("华科", 1),
            ("信昌", 2), ("PDC", 2),
            ("厚声", 3), ("UNI-ROYAL", 3), ("UNIROYAL", 3),
            ("大毅", 4), ("TA-I", 4), ("TAI", 4),
            ("光颉", 5), ("VIKING", 5),
            ("国巨", 6), ("YAGEO", 6), ("YEGO", 6),
            ("TE", 7), ("泰科", 7), ("TE CONNECTIVITY", 7),
            ("RESI", 8), ("开步睿思", 8),
            ("威世", 9), ("VISHAY", 9),
            ("旺诠", 10), ("RALEC", 10),
            ("KOA", 11), ("BOURNS", 12), ("STACKPOLE", 13), ("SUSUMU", 14), ("PANASONIC", 15),
        ]
    elif ctype in CAPACITOR_COMPONENT_TYPES:
        priorities = [
            ("信昌", 1), ("PDC", 1),
            ("华新科", 2), ("WALSIN", 2), ("华科", 2),
            ("芯声微", 3), ("HRE", 3),
            ("风华", 4), ("FENGHUA", 4), ("FENGHUA ADVANCED", 4),
            ("村田", 5), ("MURATA", 5),
            ("TDK", 6), ("东电化", 6),
            ("国巨", 7), ("YAGEO", 7), ("YEGO", 7),
            ("三环", 8), ("CCTC", 8),
            ("太阳诱电", 9), ("TAIYO", 9),
            ("三星", 10), ("SAMSUNG", 10),
            ("三和", 11), ("SAMWHA", 11),
            ("江海", 12), ("JIANGHAI", 12), ("NANTONG JIANGHAI", 12),
        ]
    else:
        priorities = [
            ("信昌", 1), ("PDC", 1),
            ("华新科", 2), ("WALSIN", 2),
            ("风华", 2), ("FENGHUA", 2), ("FENGHUA ADVANCED", 2),
            ("村田", 3), ("MURATA", 3),
            ("TDK", 4), ("东电化", 4),
            ("国巨", 5), ("YAGEO", 5), ("YEGO", 5),
            ("三环", 6), ("CCTC", 6),
            ("太阳诱电", 7), ("TAIYO", 7),
            ("三星", 8), ("SAMSUNG", 8),
            ("三和", 9), ("SAMWHA", 9),
            ("江海", 10), ("JIANGHAI", 10), ("NANTONG JIANGHAI", 10),
        ]
    for key, val in priorities:
        if key in b:
            return val
    return 99

def classify_match_level(row, spec):
    # 比较时区分数据库原始值和查询补齐值，避免把空白字段伪装成“完全匹配”
    row_size_raw = clean_size(row.get("尺寸（inch）", ""))
    row_mat_raw = clean_material(row.get("材质（介质）", ""))
    row_tol_raw = clean_tol_for_match(row.get("容值误差", ""))
    row_volt_raw = clean_voltage(row.get("耐压（V）", ""))
    row_mat = row_mat_raw or clean_material(spec.get("材质（介质）", ""))
    row_pf = row.get("容值_pf", None)

    spec_size = clean_size(spec.get("尺寸（inch）", ""))
    spec_mat = clean_material(spec.get("材质（介质）", ""))
    spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
    spec_volt = clean_voltage(spec.get("耐压（V）", ""))
    spec_pf = spec.get("容值_pf", None)

    def volt_num(v):
        try:
            return float(v)
        except:
            return None

    def pf_equal(a, b):
        try:
            return abs(float(a) - float(b)) < 1e-6
        except:
            return False

    def tol_rank(value):
        rank_map = {"1": 1, "2": 2, "5": 3, "10": 4, "20": 5, "+80/-20": 6}
        return rank_map.get(value, 99)

    target_type = infer_spec_component_type(spec)

    if target_type in (RESISTOR_COMPONENT_TYPES | {"热敏电阻"}):
        row_res = row.get("_res_ohm", None)
        try:
            row_res_value = float(row_res) if row_res is not None and not pd.isna(row_res) else None
        except Exception:
            row_res_value = None
        spec_res = spec.get("_resistance_ohm", None)
        try:
            spec_res_value = float(spec_res) if spec_res is not None else None
        except Exception:
            spec_res_value = None
        row_power_watt = parse_power_to_watts(row.get("_power", ""))
        spec_power_watt = parse_power_to_watts(spec.get("_power", ""))

        query_complete = (
            spec_size != "" and
            spec_res_value is not None and
            spec_tol != "" and
            spec_power_watt is not None
        )

        same_core = True
        if spec_size and row_size_raw != spec_size:
            same_core = False
        if spec_res_value is not None:
            same_core = row_res_value is not None and abs(row_res_value - spec_res_value) < 1e-6

        exact = same_core
        if spec_size and row_size_raw == "":
            exact = False
        if spec_res_value is not None and row_res_value is None:
            exact = False
        if spec_tol and (row_tol_raw == "" or not tolerance_equal(row_tol_raw, spec_tol)):
            exact = False
        if spec_power_watt is not None and (row_power_watt is None or abs(row_power_watt - spec_power_watt) > 1e-9):
            exact = False

        if query_complete and exact:
            return "完全匹配", 1

        if not query_complete and same_core:
            tol_ok_partial = (spec_tol == "" or (row_tol_raw != "" and tolerance_equal(row_tol_raw, spec_tol)))
            power_ok_partial = (spec_power_watt is None or (row_power_watt is not None and abs(row_power_watt - spec_power_watt) <= 1e-9))
            if tol_ok_partial and power_ok_partial:
                return "部分参数匹配", 2

        if query_complete and same_core:
            tol_ok = tolerance_allows(row_tol_raw, spec_tol)
            power_ok = row_power_watt is not None and row_power_watt >= spec_power_watt - 1e-9
            strictly_better = tolerance_strictly_better(row_tol_raw, spec_tol) or (
                row_power_watt is not None and row_power_watt > spec_power_watt + 1e-9
            )
            if tol_ok and power_ok and strictly_better:
                return "高代低", 3

        return "需确认替代", 4

    # MLCC 必须以真实存在的尺寸和容值为前提，不能拿查询值回填后伪装成可匹配。
    if target_type == "MLCC":
        if spec_size == "" or spec_pf is None:
            return "需确认替代", 4
        if row_size_raw == "" or row_pf is None:
            return "需确认替代", 4

    # 查询条件完整：五大关键参数都明确
    query_complete = (
        spec_size != "" and
        spec_mat != "" and
        spec_pf is not None and
        spec_tol != "" and
        spec_volt != ""
    )

    # 核心参数相同
    same_core = True
    if spec_size and row_size_raw != spec_size:
        same_core = False
    if spec_mat and (row_mat_raw == "" or row_mat != spec_mat):
        same_core = False
    if spec_pf is not None and not pf_equal(row_pf, spec_pf):
        same_core = False

    # 完全匹配必须以数据库原始字段真实存在且一致为准
    exact = same_core
    if spec_size and row_size_raw == "":
        exact = False
    if spec_mat and row_mat_raw == "":
        exact = False
    if spec_tol and (row_tol_raw == "" or not tolerance_equal(row_tol_raw, spec_tol)):
        exact = False
    if spec_volt and (row_volt_raw == "" or row_volt_raw != spec_volt):
        exact = False

    # 只有查询条件完整时，才允许判“完全匹配”
    if query_complete and exact:
        return "完全匹配", 1

    # 如果查询条件不完整，但已输入参数都能对上，统一叫“部分参数匹配”
    if not query_complete and same_core:
        tol_ok_partial = (spec_tol == "" or tolerance_equal(row_tol_raw, spec_tol))
        volt_ok_partial = (spec_volt == "" or (row_volt_raw != "" and row_volt_raw == spec_volt))
        if tol_ok_partial and volt_ok_partial:
            return "部分参数匹配", 2

    # 查询条件完整时，同尺寸/同介质/同容值，且容差更严或耐压更高 => 高代低
    if query_complete and same_core:
        s_tol_rank = tol_rank(spec_tol)
        r_tol_rank = tol_rank(row_tol_raw)

        s_v = volt_num(spec_volt)
        r_v = volt_num(row_volt_raw)

        tol_ok = tolerance_allows(row_tol_raw, spec_tol)
        volt_ok = (row_volt_raw != "" and s_v is not None and r_v is not None and r_v >= s_v)
        strictly_better = tolerance_strictly_better(row_tol_raw, spec_tol) or (r_v > s_v)

        if tol_ok and volt_ok and strictly_better:
            return "高代低", 3

    return "需确认替代", 4

def apply_match_levels_and_sort(df, spec):
    if df.empty:
        return df
    work = df.copy()
    target_type = infer_spec_component_type(spec)
    if target_type == "MLCC":
        spec_size = clean_size(spec.get("尺寸（inch）", ""))
        spec_mat = clean_material(spec.get("材质（介质）", ""))
        spec_tol = clean_tol_for_match(spec.get("容值误差", ""))
        spec_volt = clean_voltage(spec.get("耐压（V）", ""))
        spec_pf = spec.get("容值_pf", None)
        spec_mlcc_class = infer_mlcc_series_class_from_spec(spec)

        row_size_raw = work["_size"] if "_size" in work.columns else work["尺寸（inch）"].astype(str).apply(clean_size)
        row_mat_raw = work["_mat"] if "_mat" in work.columns else work["材质（介质）"].astype(str).apply(clean_material)
        row_tol_raw = work["_tol"] if "_tol" in work.columns else work["容值误差"].astype(str).apply(clean_tol_for_match)
        row_volt_raw = work["_volt"] if "_volt" in work.columns else work["耐压（V）"].astype(str).apply(clean_voltage)
        row_pf = work["_pf"] if "_pf" in work.columns else pd.to_numeric(work["容值_pf"], errors="coerce")
        row_mlcc_class = work["_mlcc_series_class"] if "_mlcc_series_class" in work.columns else pd.Series([""] * len(work), index=work.index, dtype="object")

        level_series = pd.Series("需确认替代", index=work.index, dtype="object")
        rank_series = pd.Series(4, index=work.index, dtype="int64")

        if spec_size != "" and spec_pf is not None:
            same_core = pd.Series(True, index=work.index)
            same_core &= row_size_raw.eq(spec_size)
            if spec_mat != "":
                same_core &= row_mat_raw.ne("") & row_mat_raw.eq(spec_mat)
            target_pf = float(spec_pf)
            same_core &= row_pf.notna() & ((row_pf - target_pf).abs() < 1e-6)

            query_complete = (
                spec_size != "" and
                spec_mat != "" and
                spec_pf is not None and
                spec_tol != "" and
                spec_volt != ""
            )

            if query_complete:
                exact_mask = same_core.copy()
                exact_mask &= row_size_raw.ne("")
                exact_mask &= row_mat_raw.ne("")
                exact_mask &= row_tol_raw.ne("") & tolerance_equal_series(work, spec_tol)
                exact_mask &= row_volt_raw.ne("") & row_volt_raw.eq(spec_volt)
                level_series.loc[exact_mask] = "完全匹配"
                rank_series.loc[exact_mask] = 1

                remaining_same_core = same_core & ~exact_mask
                tol_ok = tolerance_allows_series(work, spec_tol)
                tol_kind, tol_value = tolerance_sort_key(spec_tol)
                if tol_kind in {"percent", "pf"}:
                    tol_strictly_better_mask = work["_tol_kind"].eq(tol_kind) & work["_tol_num"].notna() & work["_tol_num"].lt(float(tol_value))
                else:
                    tol_strictly_better_mask = pd.Series(False, index=work.index)

                try:
                    spec_volt_num = float(spec_volt)
                    volt_ok = row_volt_raw.ne("") & work["_volt_num"].notna() & work["_volt_num"].ge(spec_volt_num)
                    volt_strictly_better_mask = work["_volt_num"].notna() & work["_volt_num"].gt(spec_volt_num)
                except Exception:
                    volt_ok = row_volt_raw.ne("") & row_volt_raw.eq(spec_volt)
                    volt_strictly_better_mask = pd.Series(False, index=work.index)

                high_mask = remaining_same_core & tol_ok & volt_ok & (tol_strictly_better_mask | volt_strictly_better_mask)
                level_series.loc[high_mask] = "高代低"
                rank_series.loc[high_mask] = 3
            else:
                partial_mask = same_core.copy()
                if spec_tol != "":
                    partial_mask &= tolerance_equal_series(work, spec_tol)
                if spec_volt != "":
                    partial_mask &= row_volt_raw.ne("") & row_volt_raw.eq(spec_volt)
                level_series.loc[partial_mask] = "部分参数匹配"
                rank_series.loc[partial_mask] = 2

        work["推荐等级"] = level_series
        work["_level_rank"] = rank_series
        work["_mlcc_class_rank"] = row_mlcc_class.astype(str).apply(lambda value: mlcc_series_class_sort_rank(value, spec_mlcc_class))
    else:
        levels = work.apply(lambda r: classify_match_level(r, spec), axis=1)
        work["推荐等级"] = [x[0] for x in levels]
        work["_level_rank"] = [x[1] for x in levels]
    if "_model_rule_authority" in work.columns:
        work["_seed_rank"] = work["_model_rule_authority"].astype(str).apply(lambda value: 0 if clean_text(value) == "jianghai_seed" else 1)
    else:
        work["_seed_rank"] = 1
    work["_brand_rank"] = work["品牌"].apply(lambda value: brand_priority_value(value, component_type=target_type))
    sort_cols = ["_seed_rank", "_level_rank"]
    ascending = [True, True]
    if "_mlcc_class_rank" in work.columns:
        sort_cols.append("_mlcc_class_rank")
        ascending.append(True)
    if target_type in CAPACITOR_COMPONENT_TYPES:
        sort_cols.append("_brand_rank")
        ascending.append(True)
        if "_matched_param_count" in work.columns:
            sort_cols.append("_matched_param_count")
            ascending.append(False)
        sort_cols.extend(["品牌", "型号"])
        ascending.extend([True, True])
    else:
        if "_matched_param_count" in work.columns:
            sort_cols.append("_matched_param_count")
            ascending.append(False)
        sort_cols.extend(["_brand_rank", "品牌", "型号"])
        ascending.extend([True, True, True])
    work = work.sort_values(by=sort_cols, ascending=ascending)
    return work.drop(columns=["_seed_rank", "_level_rank", "_brand_rank", "_mlcc_class_rank"], errors="ignore")

def detect_query_mode_and_spec(df, line):
    other_spec = parse_other_passive_query(line)
    if other_spec is not None:
        if count_query_params(other_spec) < other_passive_min_required_params(other_spec):
            return "规格不足", other_spec
        return normalized_other_passive_mode(other_spec), other_spec

    if looks_like_compact_part_query(line):
        spec = reverse_spec(df, line)
        if spec is not None and count_core_params(spec) >= 3:
            return "料号", spec

        part_spec = reverse_spec_partial(line)
        if part_spec is not None and part_spec.get("_param_count", 0) >= 3:
            return "料号片段", part_spec

    if looks_like_spec_query(line):
        spec = parse_spec_query(line)
        if spec is None:
            return "无法识别", None
        if spec.get("_param_count", 0) < 3:
            return "规格不足", spec
        return "规格", spec

    spec = reverse_spec(df, line)
    if spec is not None:
        return "料号", spec

    part_spec = reverse_spec_partial(line)
    if part_spec is not None:
        return "料号片段", part_spec

    other_spec = parse_other_passive_query(line)
    if other_spec is not None:
        if count_core_params(other_spec) < 2:
            return "规格不足", other_spec
        return normalized_other_passive_mode(other_spec), other_spec

    spec = parse_spec_query(line)
    if spec is not None:
        if spec.get("_param_count", 0) < 3:
            return "规格不足", spec
        return "规格", spec

    return "无法识别", None

def make_unique_column_names(columns):
    seen = {}
    out = []
    for idx, col in enumerate(columns, start=1):
        base = clean_text(col) or f"未命名列{idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        out.append(base if count == 1 else f"{base} ({count})")
    return out


def get_uploaded_file_bytes(uploaded_file):
    if uploaded_file is None:
        return b""
    for method_name in ("getvalue", "read"):
        try:
            method = getattr(uploaded_file, method_name, None)
            if method is None:
                continue
            data = method()
            if data is not None:
                return data
        except Exception:
            continue
    try:
        uploaded_file.seek(0)
        return uploaded_file.read()
    except Exception:
        return b""


def read_uploaded_bom_workbook(uploaded_file):
    if uploaded_file is None:
        return {
            "kind": "empty",
            "file_name": "",
            "file_bytes": b"",
            "sheet_frames": [],
            "sheet_names": [],
        }

    file_name = clean_text(getattr(uploaded_file, "name", ""))
    raw_bytes = get_uploaded_file_bytes(uploaded_file)
    lower_name = file_name.lower()

    if lower_name.endswith(".csv"):
        try:
            bom_df = pd.read_csv(BytesIO(raw_bytes), dtype=str)
        except Exception:
            bom_df = pd.DataFrame()
        if bom_df is None or bom_df.empty:
            return {
                "kind": "csv",
                "file_name": file_name,
                "file_bytes": raw_bytes,
                "sheet_frames": [],
                "sheet_names": [],
            }
        bom_df = bom_df.dropna(how="all").copy()
        bom_df.columns = make_unique_column_names(bom_df.columns)
        return {
            "kind": "csv",
            "file_name": file_name,
            "file_bytes": raw_bytes,
            "sheet_frames": [{"sheet_name": os.path.splitext(file_name or "CSV")[0] or "CSV", "df": bom_df}],
            "sheet_names": [os.path.splitext(file_name or "CSV")[0] or "CSV"],
        }

    try:
        xls = pd.ExcelFile(BytesIO(raw_bytes))
    except Exception:
        try:
            fallback_df = pd.read_excel(BytesIO(raw_bytes), dtype=str)
        except Exception:
            fallback_df = pd.DataFrame()
        if fallback_df is None or fallback_df.empty:
            return {
                "kind": "excel",
                "file_name": file_name,
                "file_bytes": raw_bytes,
                "sheet_frames": [],
                "sheet_names": [],
            }
        fallback_df = fallback_df.dropna(how="all").copy()
        fallback_df.columns = make_unique_column_names(fallback_df.columns)
        return {
            "kind": "excel",
            "file_name": file_name,
            "file_bytes": raw_bytes,
            "sheet_frames": [{"sheet_name": "Sheet1", "df": fallback_df}],
            "sheet_names": ["Sheet1"],
        }

    sheet_frames = []
    for sheet_name in xls.sheet_names:
        try:
            bom_df = xls.parse(sheet_name=sheet_name, dtype=str)
        except Exception:
            continue
        if bom_df is None or bom_df.empty:
            continue
        bom_df = bom_df.dropna(how="all").copy()
        if bom_df.empty:
            continue
        bom_df.columns = make_unique_column_names(bom_df.columns)
        sheet_frames.append({"sheet_name": sheet_name, "df": bom_df})

    return {
        "kind": "excel",
        "file_name": file_name,
        "file_bytes": raw_bytes,
        "sheet_frames": sheet_frames,
        "sheet_names": [item["sheet_name"] for item in sheet_frames],
    }


def read_uploaded_bom(uploaded_file):
    workbook = read_uploaded_bom_workbook(uploaded_file)
    if workbook.get("sheet_frames"):
        return workbook["sheet_frames"][0]["df"].copy()
    return pd.DataFrame()

def normalize_bom_header_name(col):
    s = clean_text(col).lower()
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"[\s\-_()/\\\[\]{}:：,.，;；]+", "", s)
    return s

def is_probable_quantity_value(value):
    s = clean_text(value).replace(",", "")
    if s == "":
        return False
    return re.fullmatch(r"\d+(?:\.\d+)?", s) is not None


def is_probable_price_header(col_name):
    header = normalize_bom_header_name(col_name)
    return any(keyword in header for keyword in ["单价", "單價", "price", "金额", "rmb", "含税", "tax"])


def is_probable_quantity_column(upload_df, col_name):
    header = normalize_bom_header_name(col_name)
    if is_probable_price_header(col_name):
        return False
    if any(keyword in header for keyword in ["数量", "用量", "需求数量", "qty", "quantity", "用数"]):
        return True
    sample = upload_df[col_name].dropna().astype(str).head(25).tolist()
    if not sample:
        return False
    numeric_ratio = sum(1 for x in sample if is_probable_quantity_value(x)) / len(sample)
    return numeric_ratio >= 0.9

def guess_bom_column(upload_df, role, used_columns=None):
    if upload_df is None or upload_df.empty:
        return None
    used_columns = set(used_columns or [])
    best_col = None
    best_score = 0

    for pos, col in enumerate(upload_df.columns):
        if col in used_columns:
            continue

        score = 0
        header = normalize_bom_header_name(col)
        for keyword, weight in BOM_COLUMN_KEYWORDS.get(role, []):
            if keyword in header:
                score += weight
        for keyword, weight in BOM_COLUMN_NEGATIVE_KEYWORDS.get(role, []):
            if keyword in header:
                score -= weight

        sample = upload_df[col].dropna().astype(str).head(25).tolist()
        if sample:
            if role == "quantity":
                numeric_ratio = sum(1 for x in sample if is_probable_quantity_value(x)) / len(sample)
                score += int(numeric_ratio * 30)
            elif role == "spec":
                spec_ratio = sum(1 for x in sample if looks_like_spec_query(x)) / len(sample)
                other_passive_ratio = sum(1 for x in sample if detect_component_type_hint(x) != "" or find_resistance_in_text(x) is not None) / len(sample)
                score += int(spec_ratio * 18)
                score += int(other_passive_ratio * 16)
            elif role == "model":
                model_ratio = 0
                for x in sample:
                    raw = clean_text(x)
                    compact = raw.upper().replace(" ", "")
                    if re.fullmatch(r"[A-Z0-9._/\-]+", compact or "") and re.search(r"[A-Z]", compact or "") and re.search(r"\d", compact or ""):
                        model_ratio += 1
                score += int((model_ratio / len(sample)) * 18)

        if score > best_score or (score == best_score and best_col is not None and pos < list(upload_df.columns).index(best_col)):
            best_col = col
            best_score = score

    if role == "quantity":
        if best_col is None or best_score <= 0:
            return None
        return best_col if is_probable_quantity_column(upload_df, best_col) else None
    return best_col if best_score > 0 else None

def guess_bom_column_mapping(upload_df):
    mapping = {"model": None, "spec": None, "name": None, "quantity": None}
    used = set()
    for role in ["model", "spec", "name", "quantity"]:
        guessed = guess_bom_column(upload_df, role, used)
        if guessed:
            mapping[role] = guessed
            used.add(guessed)
    return mapping

def get_bom_selected_value(record, column_name):
    if not column_name or column_name == BOM_NONE_OPTION:
        return ""
    return clean_text(record.get(column_name, ""))

def format_bom_quantity(value):
    raw = clean_text(value)
    if raw == "":
        return ""
    normalized = raw.replace(",", "")
    try:
        num = float(normalized)
        if num.is_integer():
            return str(int(num))
        text = f"{num:.6f}".rstrip("0").rstrip(".")
        return text
    except:
        return raw

def join_bom_parts(*parts):
    tokens = [clean_text(x) for x in parts if clean_text(x) != ""]
    return re.sub(r"\s+", " ", " ".join(tokens)).strip()


def collect_bom_extra_spec_values(record, column_mapping):
    selected_columns = {
        column_mapping.get("model"),
        column_mapping.get("spec"),
        column_mapping.get("name"),
        column_mapping.get("quantity"),
    }
    extras = []
    seen = set()
    for col_name, raw_value in record.items():
        if col_name in selected_columns:
            continue
        if is_probable_price_header(col_name):
            continue
        header = normalize_bom_header_name(col_name)
        if any(token in header for token in ["moq", "lt", "leadtime", "有效期", "日期", "date"]):
            continue
        value = clean_text(raw_value)
        if value == "":
            continue
        score = 0
        if looks_like_compact_part_query(value):
            score += 3
        if looks_like_spec_query(value):
            score += 3
        if detect_component_type_hint(value) != "":
            score += 3
        if find_embedded_size(value) != "":
            score += 2
        if find_embedded_material(value) != "":
            score += 2
        if re.search(r"\d+(?:\.\d+)?\s*(?:PF|NF|UF|V)\b", value, flags=re.I):
            score += 2
        if find_resistance_in_text(value) is not None:
            score += 2
        if score <= 0:
            continue
        compact_value = clean_text(value)
        if compact_value in seen:
            continue
        seen.add(compact_value)
        extras.append((score, compact_value))
    extras.sort(key=lambda item: (-item[0], len(item[1])))
    return [item[1] for item in extras[:2]]


def build_bom_query_candidates(model_value, spec_value, name_value, extra_values=None):
    candidates = []
    seen = set()
    extra_values = [clean_text(x) for x in (extra_values or []) if clean_text(x) != ""]

    def add_candidate(text, source):
        query = join_bom_parts(text)
        if query and query not in seen:
            candidates.append({"query": query, "source": source})
            seen.add(query)

    add_candidate(model_value, "型号列")
    add_candidate(join_bom_parts(spec_value, name_value), "规格列+品名列")
    add_candidate(spec_value, "规格列")
    add_candidate(name_value, "品名列")
    add_candidate(join_bom_parts(model_value, spec_value), "型号列+规格列")
    add_candidate(join_bom_parts(model_value, spec_value, name_value), "型号列+规格列+品名列")
    if extra_values:
        add_candidate(join_bom_parts(spec_value, name_value, *extra_values), "规格列+品名列+其他列")
        add_candidate(join_bom_parts(spec_value, *extra_values), "规格列+其他列")
        add_candidate(join_bom_parts(name_value, *extra_values), "品名列+其他列")
    return candidates

def describe_bom_result(candidate_result):
    status = candidate_result["status"]
    mode = candidate_result["mode"]
    source = candidate_result["source"]
    top_level = candidate_result.get("top_match_level", "")
    spec = candidate_result.get("spec") or {}

    if status == "无法识别":
        return f"{source}内容无法识别为有效料号或规格"
    if status == "规格不足":
        min_required = other_passive_min_required_params(spec) if clean_text(spec.get("器件类型", "")) != "" else 3
        return f"{source}已识别出部分规格，但关键参数不足 {min_required} 项"
    if status == "无匹配":
        if is_other_passive_mode(mode):
            component_type = normalized_other_passive_mode(spec, mode)
            return f"已识别为{component_type}规格，但当前数据库未找到可匹配结果"
        if mode == "料号":
            return "已按料号反推规格，但数据库未找到可匹配结果"
        if mode == "料号片段":
            return "已按料号片段反推规格，但数据库暂无匹配结果"
        return "已识别规格，但数据库未找到匹配结果"

    if top_level == "完全匹配":
        return f"使用{source}解析，首选结果为完全匹配"
    if top_level == "部分参数匹配":
        return f"使用{source}解析，首选结果为部分参数匹配"
    if top_level in {"高代低", "可直接替代"}:
        return f"使用{source}解析，首选结果为高代低可替代"
    if top_level == "需确认替代":
        return f"使用{source}解析，首选结果需人工确认替代关系"
    return f"使用{source}解析成功"

def evaluate_bom_candidate(df, query_text, source_label, candidate_index, query_cache=None, full_df_provider=None):
    cached = query_cache.get(query_text) if query_cache is not None else None
    if cached is None:
        full_df = df
        exact_part_rows = resolve_prefetched_exact_part_rows(query_text)
        detect_df = exact_part_rows if isinstance(exact_part_rows, pd.DataFrame) and not exact_part_rows.empty else None
        mode, spec = detect_query_mode_and_spec(detect_df, query_text)
        query_df = None
        source_text = clean_text(source_label)
        model_token_rows = pd.DataFrame()
        model_token_candidates = []
        matched_model_token = ""
        if mode == "无法识别":
            model_token_rows, model_token_candidates, matched_model_token = load_component_rows_by_query_model_tokens(query_text)
        if mode == "无法识别" and source_text == "型号列":
            if isinstance(exact_part_rows, pd.DataFrame) and not exact_part_rows.empty:
                query_df = exact_part_rows
                mode, spec = detect_query_mode_and_spec(query_df, query_text)
        if (
            mode == "无法识别"
            and query_df is None
            and isinstance(model_token_rows, pd.DataFrame)
            and not model_token_rows.empty
        ):
            query_df = model_token_rows
            mode, spec = detect_query_mode_and_spec(query_df, matched_model_token or query_text)
        if mode != "无法识别" and spec is not None:
            query_df = (
                load_search_dataframe_for_query(mode, spec, query_text, exact_part_rows=exact_part_rows)
                if query_df is None
                else query_df
            )
        if query_df is None:
            allow_heavy_fallback = not (
                mode == "无法识别"
                and (
                    source_text == "型号列"
                    or bool(model_token_candidates)
                )
            )
            if allow_heavy_fallback and full_df is None and callable(full_df_provider):
                full_df = full_df_provider()
            query_df = full_df if allow_heavy_fallback else None
            if query_df is not None and (mode == "无法识别" or spec is None):
                mode, spec = detect_query_mode_and_spec(query_df, query_text)
        result = {
            "mode": mode,
            "spec": spec,
            "matched": pd.DataFrame(),
            "parse_status": "解析失败",
            "failure_reason": "",
            "status": "无法识别",
            "top_match_level": "",
            "query_df": query_df,
        }

        if mode == "无法识别" or spec is None:
            result["failure_reason"] = "无法识别型号或规格"
        elif mode == "规格不足":
            result["status"] = "规格不足"
            min_required = other_passive_min_required_params(spec) if clean_text(spec.get("器件类型", "")) != "" else 3
            result["failure_reason"] = f"请至少提供{min_required}个关键规格参数"
        else:
            result["parse_status"] = "解析成功"
            matched = cached_run_query_match(query_df if query_df is not None else pd.DataFrame(), mode, spec, query_text=query_text)
            result["matched"] = matched
            if matched.empty:
                result["status"] = "无匹配"
            else:
                result["status"] = "匹配成功"
                if "推荐等级" in matched.columns and not matched.empty:
                    result["top_match_level"] = clean_text(matched.iloc[0].get("推荐等级", ""))
        cached = result
        if query_cache is not None:
            query_cache[query_text] = cached

    result = {
        "query": query_text,
        "source": source_label,
        "mode": cached.get("mode", "无法识别"),
        "spec": cached.get("spec"),
        "matched": cached.get("matched", pd.DataFrame()),
        "parse_status": cached.get("parse_status", "解析失败"),
        "failure_reason": cached.get("failure_reason", ""),
        "status": cached.get("status", "无法识别"),
        "top_match_level": cached.get("top_match_level", ""),
        "query_df": cached.get("query_df"),
        "_candidate_index": candidate_index,
    }
    result["difference_note"] = describe_bom_result(result)
    return result

def bom_candidate_priority(candidate_result):
    status_rank = {"匹配成功": 4, "无匹配": 3, "规格不足": 2, "无法识别": 1}
    match_rank = {"完全匹配": 4, "部分参数匹配": 3, "高代低": 2, "可直接替代": 2, "需确认替代": 1}
    mode_value = clean_text(candidate_result.get("mode", ""))
    mode_rank = {
        "料号": 3,
        "规格": 2,
        "料号片段": 1,
        "规格不足": 0,
        "无法识别": 0,
    }
    if is_other_passive_mode(mode_value):
        mode_score = 2
    else:
        mode_score = mode_rank.get(mode_value, 0)
    return (
        status_rank.get(candidate_result.get("status", ""), 0),
        match_rank.get(candidate_result.get("top_match_level", ""), 0),
        mode_score,
        -candidate_result.get("_candidate_index", 999),
    )

def bom_candidate_source_bonus(candidate_result):
    source = clean_text(candidate_result.get("source", ""))
    if source == "型号列":
        return 3
    if "规格列+品名列" in source:
        return 2
    if "品名列" in source:
        return 1
    return 0

def bom_candidate_core_count(candidate_result):
    return count_query_params(candidate_result.get("spec"))

def bom_candidate_prefers_richer_spec(candidate_result, current_best):
    if candidate_result is None or current_best is None:
        return False

    candidate_core = bom_candidate_core_count(candidate_result)
    current_core = bom_candidate_core_count(current_best)
    if candidate_core <= current_core:
        return False

    candidate_source_bonus = bom_candidate_source_bonus(candidate_result)
    current_source_bonus = bom_candidate_source_bonus(current_best)
    if candidate_source_bonus <= current_source_bonus:
        return False

    candidate_mode = clean_text(candidate_result.get("mode", ""))
    current_mode = clean_text(current_best.get("mode", ""))
    if candidate_mode != "规格" and not is_other_passive_mode(candidate_mode):
        return False
    if current_mode not in {"规格", "规格不足"} and not is_other_passive_mode(current_mode):
        return False

    current_status = clean_text(current_best.get("status", ""))
    candidate_status = clean_text(candidate_result.get("status", ""))
    current_level = clean_text(current_best.get("top_match_level", ""))

    if current_status == "匹配成功" and current_level == "完全匹配":
        return False
    if candidate_status == "匹配成功":
        return True
    if candidate_status == "无匹配" and current_status in {"匹配成功", "规格不足", "无法识别"}:
        return True
    return False

def should_replace_best_bom_candidate(best, candidate_result):
    if best is None:
        return True
    if bom_candidate_prefers_richer_spec(candidate_result, best):
        return True
    if bom_candidate_prefers_richer_spec(best, candidate_result):
        return False
    return bom_candidate_priority(candidate_result) > bom_candidate_priority(best)


def bom_candidate_good_enough(candidate_result):
    if candidate_result is None:
        return False
    if is_other_passive_mode(candidate_result.get("mode")) and "品名列" in clean_text(candidate_result.get("source", "")) and bom_candidate_core_count(candidate_result) >= 2:
        return True
    if candidate_result.get("status") != "匹配成功":
        return False
    level = clean_text(candidate_result.get("top_match_level", ""))
    source = clean_text(candidate_result.get("source", ""))
    spec = candidate_result.get("spec")
    if source == "型号列":
        return True
    if level == "完全匹配":
        return True
    if "规格列+品名列" in source and count_core_params(spec) >= 4 and level in {"部分参数匹配", "高代低"}:
        return True
    return False


def choose_best_bom_candidate(df, candidates, query_cache=None, full_df_provider=None):
    best = None
    for idx, candidate in enumerate(candidates):
        result = evaluate_bom_candidate(df, candidate["query"], candidate["source"], idx, query_cache=query_cache, full_df_provider=full_df_provider)
        if should_replace_best_bom_candidate(best, result):
            best = result
        if bom_candidate_good_enough(best):
            break
    return best

def build_bom_result_row(df, line):
    mode, spec = detect_query_mode_and_spec(df, line)
    row = {
        "原始输入": line,
        "识别模式": mode,
        "品牌": "",
        "型号": "",
        "推荐品牌": "",
        "推荐型号": "",
        "品牌1": "",
        "型号1": "",
        "品牌2": "",
        "型号2": "",
        "品牌3": "",
        "型号3": "",
        "尺寸（inch）": "",
        "材质（介质）": "",
        "容值": "",
        "容值单位": "",
        "容值误差": "",
        "耐压（V）": "",
        "匹配数量": 0,
        "前5个其他品牌型号": "",
        "状态": ""
    }

    if mode == "无法识别" or spec is None:
        row["状态"] = "无法识别"
        return row

    if mode == "规格不足":
        min_required = other_passive_min_required_params(spec) if clean_text(spec.get("器件类型", "")) != "" else 3
        if min_required <= 1:
            row["状态"] = "请至少输入一个关键规格参数"
        elif min_required == 2:
            row["状态"] = "请至少输入两个关键规格参数"
        else:
            row["状态"] = "请最少输入三个规格参数"
        value, unit = spec_display_value_unit(spec)
        row["尺寸（inch）"] = spec.get("尺寸（inch）", "")
        row["材质（介质）"] = spec.get("材质（介质）", "")
        row["容值"] = value
        row["容值单位"] = unit
        row["容值误差"] = clean_tol_for_display(spec.get("容值误差", ""))
        row["耐压（V）"] = voltage_display(spec.get("耐压（V）", ""))
        return row

    value, unit = spec_display_value_unit(spec)
    row["品牌"] = spec.get("品牌", "")
    row["型号"] = spec.get("型号", "") if mode == "料号" else ""
    row["尺寸（inch）"] = spec.get("尺寸（inch）", "")
    row["材质（介质）"] = spec.get("材质（介质）", "")
    row["容值"] = value
    row["容值单位"] = unit
    row["容值误差"] = clean_tol_for_display(spec.get("容值误差", ""))
    row["耐压（V）"] = voltage_display(spec.get("耐压（V）", ""))

    matched = cached_run_query_match(df, mode, spec, query_text=line)

    if matched.empty:
        row["状态"] = "无匹配"
        return row

    matched = matched.copy()
    matched["品牌"] = matched["品牌"].astype(str).fillna("")
    matched["型号"] = matched["型号"].astype(str).fillna("")
    display_match = choose_bom_display_match(matched)
    row["匹配数量"] = int(len(matched))
    if display_match is not None:
        row["推荐品牌"] = clean_text(display_match.get("品牌", ""))
        row["推荐型号"] = clean_text(display_match.get("型号", ""))
    row.update(
        build_bom_preferred_brand_slots(
            matched,
            exclude_pairs=[(row["推荐品牌"], row["推荐型号"])],
        )
    )
    row["前5个其他品牌型号"] = format_other_brand_models(
        matched,
        exclude_aliases=BOM_PREFERRED_BRAND_EXCLUDE_ALIASES,
        exclude_pairs=[(row["推荐品牌"], row["推荐型号"])],
    )
    row["状态"] = "匹配成功"
    return row

def build_bom_upload_result_row(df, row_index, record, column_mapping, query_cache=None, full_df_provider=None):
    model_value = get_bom_selected_value(record, column_mapping.get("model"))
    spec_value = get_bom_selected_value(record, column_mapping.get("spec"))
    name_value = get_bom_selected_value(record, column_mapping.get("name"))
    quantity_value = format_bom_quantity(get_bom_selected_value(record, column_mapping.get("quantity")))

    result_row = {
        "BOM行号": int(row_index) + 2,
        "BOM型号": model_value,
        "BOM规格": spec_value,
        "BOM品名": name_value,
        "器件类型": "",
        "关键规格": "",
        "规格参数明细": "",
        "BOM数量": quantity_value,
        "解析来源": "",
        "解析输入": "",
        "识别模式": "无法识别",
        "解析状态": "解析失败",
        "失败原因": "",
        "差异说明": "",
        "品牌": "",
        "型号": "",
        "推荐品牌": "",
        "推荐型号": "",
        "品牌1": "",
        "型号1": "",
        "品牌2": "",
        "型号2": "",
        "品牌3": "",
        "型号3": "",
        "尺寸（inch）": "",
        "材质（介质）": "",
        "容值": "",
        "容值单位": "",
        "容值误差": "",
        "耐压（V）": "",
        "匹配数量": 0,
        "首选推荐等级": "",
        "匹配参数明细": "",
        "前5个其他品牌型号": "",
        "匹配命中列": "",
        "备注1": "",
        "备注2": "",
        "备注3": "",
        "状态": "无法识别",
    }

    extra_values = collect_bom_extra_spec_values(record, column_mapping)
    candidates = build_bom_query_candidates(model_value, spec_value, name_value, extra_values=extra_values)
    if not candidates:
        result_row["失败原因"] = "指定的 BOM 列为空"
        result_row["差异说明"] = "当前行在已指定列中没有可用于解析的内容"
        return result_row

    best = choose_best_bom_candidate(df, candidates, query_cache=query_cache, full_df_provider=full_df_provider)
    if best is None:
        result_row["失败原因"] = "BOM 解析失败"
        result_row["差异说明"] = "未能生成有效解析结果"
        return result_row

    spec = best.get("spec")
    query_df = best.get("query_df")
    result_row["解析来源"] = best["source"]
    result_row["解析输入"] = best["query"]
    display_mode = normalized_other_passive_mode(spec, best["mode"]) if spec is not None else best["mode"]
    result_row["识别模式"] = display_mode
    result_row["解析状态"] = best["parse_status"]
    result_row["失败原因"] = best["failure_reason"]
    result_row["差异说明"] = best.get("difference_note", "")
    result_row["状态"] = best["status"]

    if spec is not None:
        inferred_type = infer_spec_component_type(spec)
        raw_type = clean_text(spec.get("器件类型", ""))
        result_row["器件类型"] = inferred_type or raw_type or ("MLCC" if not is_other_passive_mode(best["mode"]) else display_mode)
        result_row["关键规格"] = clean_text(spec.get("规格摘要", "")) or build_component_summary_from_spec(spec)
        result_row["规格参数明细"] = build_component_spec_detail_from_spec(spec)
        value, unit = spec_display_value_unit(spec)
        result_row["品牌"] = spec.get("品牌", "")
        result_row["型号"] = spec.get("型号", "") if best["mode"] == "料号" else ""
        result_row["尺寸（inch）"] = spec.get("尺寸（inch）", "")
        result_row["材质（介质）"] = spec.get("材质（介质）", "")
        result_row["容值"] = value
        result_row["容值单位"] = unit
        result_row["容值误差"] = clean_tol_for_display(spec.get("容值误差", ""))
        result_row["耐压（V）"] = voltage_display(spec.get("耐压（V）", ""))

    matched = best.get("matched")
    if isinstance(matched, pd.DataFrame) and not matched.empty:
        matched = matched.copy()
        matched["品牌"] = matched["品牌"].astype(str).fillna("")
        matched["型号"] = matched["型号"].astype(str).fillna("")
        display_match = choose_bom_display_match(matched)
        detail_match = display_match if display_match is not None else matched.iloc[0]
        result_row["匹配数量"] = int(len(matched))
        result_row["首选推荐等级"] = clean_text(display_match.get("推荐等级", "")) if display_match is not None else best.get("top_match_level", "")
        result_row["推荐品牌"] = clean_text(display_match.get("品牌", "")) if display_match is not None else ""
        result_row["推荐型号"] = clean_text(display_match.get("型号", "")) if display_match is not None else ""
        result_row.update(
            build_bom_preferred_brand_slots(
                matched,
                exclude_pairs=[(result_row["推荐品牌"], result_row["推荐型号"])],
            )
        )
        result_row["匹配参数明细"] = build_component_spec_detail_from_row(detail_match, result_row.get("器件类型", ""))
        result_row["前5个其他品牌型号"] = format_other_brand_models(
            matched,
            exclude_aliases=BOM_PREFERRED_BRAND_EXCLUDE_ALIASES,
            exclude_pairs=[(result_row["推荐品牌"], result_row["推荐型号"])],
        )
        hit_columns = []
        if clean_text(result_row["尺寸（inch）"]) != "" and clean_size(detail_match.get("尺寸（inch）", "")) == clean_size(result_row["尺寸（inch）"]):
            hit_columns.append("尺寸（inch）")
        if clean_text(result_row["材质（介质）"]) != "" and clean_material(detail_match.get("材质（介质）", "")) == clean_material(result_row["材质（介质）"]):
            hit_columns.append("材质（介质）")
        if clean_text(result_row["容值"]) != "" and clean_text(detail_match.get("容值", "")) == clean_text(result_row["容值"]):
            hit_columns.append("容值")
        if clean_text(result_row["容值单位"]) != "" and clean_text(detail_match.get("容值单位", "")).upper() == clean_text(result_row["容值单位"]).upper():
            hit_columns.append("容值单位")
        if clean_text(result_row["容值误差"]) != "" and clean_tol_for_display(detail_match.get("容值误差", "")) == clean_text(result_row["容值误差"]):
            hit_columns.append("容值误差")
        if clean_text(result_row["耐压（V）"]) != "" and voltage_display(detail_match.get("耐压（V）", "")) == clean_text(result_row["耐压（V）"]):
            hit_columns.append("耐压（V）")
        result_row["匹配命中列"] = "|".join(hit_columns)
        result_row["备注1"] = clean_text(detail_match.get("备注1", ""))
        result_row["备注2"] = clean_text(detail_match.get("备注2", ""))
        result_row["备注3"] = clean_text(detail_match.get("备注3", ""))

    if spec is not None and infer_spec_component_type(spec) == "MLCC":
        refs = resolve_mlcc_brand_references(
            query_df if isinstance(query_df, pd.DataFrame) else df,
            spec,
            matched=matched if isinstance(matched, pd.DataFrame) else None,
            current_model=result_row.get("型号", "") or spec.get("型号", ""),
        )
        result_row["信昌料号"] = refs.get("信昌料号", "")
        result_row["华科料号"] = refs.get("华科料号", "")

    if BOM_MATCH_DEBUG:
        spec_type = infer_spec_component_type(spec) if spec is not None else ""
        db_type = infer_db_component_type(spec) if spec is not None else ""
        bom_match_debug_log(
            f"row={result_row.get('BOM行号', '')}",
            f"query={clean_text(result_row.get('解析输入', ''))}",
            f"source={clean_text(result_row.get('解析来源', ''))}",
            f"mode={clean_text(result_row.get('识别模式', ''))}",
            f"best_type={clean_text(result_row.get('器件类型', ''))}",
            f"spec_type={clean_text(spec_type)}",
            f"db_type={clean_text(db_type)}",
            f"status={clean_text(result_row.get('解析状态', ''))}",
            f"spec={clean_text(spec_value)}",
            f"name={clean_text(name_value)}",
        )

    return result_row

def format_bom_elapsed_time(seconds):
    try:
        seconds = max(0.0, float(seconds))
    except Exception:
        seconds = 0.0
    if seconds < 10:
        return f"{seconds:.1f}s"
    total_seconds = int(round(seconds))
    if total_seconds < 60:
        return f"{total_seconds}s"
    minutes, sec = divmod(total_seconds, 60)
    if minutes < 60:
        return f"{minutes}m {sec:02d}s"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}h {minutes:02d}m"

def shorten_bom_progress_text(text, max_len=72):
    text = clean_text(text)
    if not text:
        return ""
    if len(text) <= max_len:
        return html.escape(text)
    return html.escape(text[: max(0, max_len - 1)].rstrip() + "…")

def build_bom_progress_card_html(progress_state):
    progress_state = progress_state or {}
    title = clean_text(progress_state.get("title", "BOM批量上传匹配"))
    subtitle = clean_text(progress_state.get("subtitle", ""))
    current_text = clean_text(progress_state.get("current_text", ""))
    current_label = clean_text(progress_state.get("current_label", "当前行")) or "当前行"
    total_rows = max(0, int(progress_state.get("total_rows") or 0))
    processed_rows = max(0, int(progress_state.get("processed_rows") or 0))
    percent = float(progress_state.get("percent") or 0.0)
    percent = max(0.0, min(100.0, percent))
    done = bool(progress_state.get("done"))
    pill_text = clean_text(progress_state.get("pill_text", "")) or f"{percent:.0f}%"
    elapsed_text = clean_text(progress_state.get("elapsed_text", ""))
    if not elapsed_text:
        elapsed_text = format_bom_elapsed_time(progress_state.get("elapsed_seconds", 0.0))

    chips = progress_state.get("chips") or []
    chip_html = []
    for chip in chips:
        if not isinstance(chip, dict):
            continue
        label = clean_text(chip.get("label", ""))
        value = clean_text(chip.get("value", ""))
        tone = clean_text(chip.get("tone", ""))
        chip_class = "bom-progress-chip"
        if tone in {"success", "warn", "fail"}:
            chip_class += f" {tone}"
        chip_html.append(
            f'<span class="{chip_class}"><strong>{html.escape(label)}</strong>{html.escape(value)}</span>'
        )

    if processed_rows and total_rows:
        summary_text = f"已处理 {processed_rows}/{total_rows} 行"
    elif total_rows:
        summary_text = f"共 {total_rows} 行"
    else:
        summary_text = clean_text(progress_state.get("summary", ""))

    subtitle_parts = [part for part in [summary_text, subtitle] if part]
    subtitle_html = " · ".join(html.escape(part) for part in subtitle_parts)
    current_html = ""
    if current_text:
        current_html = f'<div class="bom-progress-current"><strong>{html.escape(current_label)}：</strong>{shorten_bom_progress_text(current_text, 120)}</div>'

    fill_class = "bom-progress-fill is-done" if done else "bom-progress-fill"
    pill_class = "bom-progress-pill is-done" if done else "bom-progress-pill"
    current_line = ""
    if current_html:
        current_line = current_html

    summary_lines = progress_state.get("summary_lines") or []
    summary_html = ""
    if done and summary_lines:
        rendered_lines = []
        for line in summary_lines:
            clean_line = clean_text(line)
            if clean_line == "":
                continue
            rendered_lines.append(
                f'<div class="bom-progress-summary-line">{html.escape(clean_line)}</div>'
            )
        if rendered_lines:
            summary_html = f'<div class="bom-progress-summary">{"".join(rendered_lines)}</div>'

    return f"""
<div class="bom-progress-card">
  <div class="bom-progress-head">
    <div>
      <div class="bom-progress-title">{html.escape(title)}</div>
      <div class="bom-progress-subtitle">{subtitle_html}</div>
    </div>
    <div class="{pill_class}">{html.escape(pill_text)}</div>
  </div>
  <div class="bom-progress-track">
    <div class="{fill_class}" style="width: {percent:.1f}%;"></div>
  </div>
  <div class="bom-progress-meta">
    {''.join(chip_html)}
    <span class="bom-progress-chip"><strong>耗时</strong>{html.escape(elapsed_text)}</span>
  </div>
  {current_line}
  {summary_html}
</div>
"""

def render_bom_progress_card(progress_placeholder, progress_state):
    if progress_placeholder is None:
        return
    progress_placeholder.markdown(build_bom_progress_card_html(progress_state), unsafe_allow_html=True)


def build_search_progress_state(
    total_queries,
    completed_queries,
    current_text="",
    stage_text="",
    note="",
    elapsed_seconds=0.0,
    stage_step=0,
    done=False,
    extra_chips=None,
    summary_lines=None,
):
    total_queries = max(0, int(total_queries or 0))
    completed_queries = max(0, int(completed_queries or 0))
    stage_step = max(0, min(SEARCH_PROGRESS_STAGE_COUNT, int(stage_step or 0)))
    if done:
        percent = 100.0
    elif total_queries > 0:
        numerator = (completed_queries * SEARCH_PROGRESS_STAGE_COUNT) + stage_step
        percent = max(0.0, min(100.0, (numerator / float(total_queries * SEARCH_PROGRESS_STAGE_COUNT)) * 100.0))
    else:
        percent = 0.0

    chips = [{"label": "阶段", "value": clean_text(stage_text) or "准备中"}]
    if total_queries > 0:
        current_index = completed_queries if done or clean_text(current_text) == "" else min(total_queries, completed_queries + 1)
        chips.append({"label": "进度", "value": f"{current_index}/{total_queries}"})
    for chip in extra_chips or []:
        if not isinstance(chip, dict):
            continue
        label = clean_text(chip.get("label", ""))
        value = clean_text(chip.get("value", ""))
        if label == "" or value == "":
            continue
        chips.append(
            {
                "label": label,
                "value": value,
                "tone": clean_text(chip.get("tone", "")),
            }
        )

    return {
        "title": "料号搜索匹配",
        "subtitle": clean_text(note),
        "current_text": current_text,
        "current_label": "当前输入",
        "total_rows": total_queries,
        "processed_rows": completed_queries,
        "percent": percent,
        "done": done,
        "pill_text": "已完成" if done else "匹配中",
        "elapsed_seconds": elapsed_seconds,
        "chips": chips,
        "summary_lines": summary_lines or [],
    }

def build_bom_download_footer_html(data_bytes, filename, label="下载 BOM 匹配后 Excel", container_class="result-table-footer"):
    if not data_bytes:
        return f'<div class="{html.escape(container_class, quote=True)}"><span class="bom-download-hint">下载文件尚未生成</span></div>'
    href = "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64," + base64.b64encode(data_bytes).decode("ascii")
    safe_name = html.escape(clean_text(filename) or "bom_匹配后.xlsx", quote=True)
    return f'''
<div class="{html.escape(container_class, quote=True)}">
  <a class="bom-download-btn" href="{href}" download="{safe_name}">{html.escape(label)}</a>
</div>
'''


def get_query_param_text(name):
    try:
        value = st.query_params.get(name, "")
    except Exception:
        return ""
    if isinstance(value, list):
        return clean_text(value[0] if value else "")
    return clean_text(value)


def consume_bom_manual_mapping_toggle(workbook_signature):
    marker = get_query_param_text("bom_manual_mapping_toggle")
    if marker == "" or marker != clean_text(workbook_signature):
        return
    st.session_state["_bom_manual_mapping_open"] = not bool(st.session_state.get("_bom_manual_mapping_open", False))
    try:
        del st.query_params["bom_manual_mapping_toggle"]
    except Exception:
        pass
    st.rerun()


def build_bom_manual_mapping_toggle_html(workbook_signature, label="找不到规格手动定位匹配位置"):
    href = f'?bom_manual_mapping_toggle={urllib.parse.quote(clean_text(workbook_signature), safe="")}#bom-preview-toggle-anchor'
    return f'''
<div id="bom-preview-toggle-anchor" class="bom-preview-toggle-anchor"></div>
<div class="bom-download-footer-outside bom-preview-toggle-footer">
  <a class="bom-download-btn" href="{html.escape(href, quote=True)}">{html.escape(label)}</a>
</div>
'''

def bom_dataframe_from_upload(df, upload_df, column_mapping=None, allow_full_fallback=False, progress_callback=None):
    if upload_df is None or upload_df.empty:
        return pd.DataFrame()

    work = upload_df.copy()
    if column_mapping is None:
        first_col = work.columns[0]
        column_mapping = {"model": first_col, "spec": None, "name": None, "quantity": None}

    full_df_cache = {"loaded": df is not None, "prepared": False, "df": df}

    def get_full_bom_df():
        if not full_df_cache["loaded"]:
            loaded_df = load_search_dataframe_for_action("BOM 匹配")
            full_df_cache["df"] = prepare_search_dataframe(loaded_df) if loaded_df is not None and not loaded_df.empty else loaded_df
            full_df_cache["loaded"] = True
            full_df_cache["prepared"] = True
        elif not full_df_cache["prepared"] and full_df_cache["df"] is not None and not full_df_cache["df"].empty:
            full_df_cache["df"] = prepare_search_dataframe(full_df_cache["df"])
            full_df_cache["prepared"] = True
        return full_df_cache["df"]

    query_cache = {}
    full_df_provider = get_full_bom_df if allow_full_fallback else None
    records = work.to_dict(orient="records")
    total_rows = len(records)
    rows = []
    exact_count = 0
    partial_count = 0
    substitute_count = 0
    matched_count = 0
    no_match_count = 0
    fail_count = 0
    start_ts = time.perf_counter()
    last_emit_ts = start_ts
    last_emit_index = 0
    emit_every = max(1, total_rows // 80) if total_rows else 1

    def emit_progress(processed_rows, result_row=None, done=False):
        if progress_callback is None:
            return
        elapsed_seconds = max(0.0, time.perf_counter() - start_ts)
        percent = 100.0 if total_rows <= 0 else min(100.0, (processed_rows / total_rows) * 100.0)
        current_text = ""
        if isinstance(result_row, dict):
            row_no = clean_text(result_row.get("BOM行号", ""))
            row_hint = ""
            for key in ["BOM型号", "BOM规格", "BOM品名", "解析输入"]:
                row_hint = clean_text(result_row.get(key, ""))
                if row_hint:
                    break
            if row_no and row_hint:
                current_text = f"第 {row_no} 行 · {row_hint}"
            elif row_no:
                current_text = f"第 {row_no} 行"
            elif row_hint:
                current_text = row_hint
        if not current_text and total_rows:
            current_text = f"正在处理第 {processed_rows}/{total_rows} 行"
        elif not current_text:
            current_text = "正在处理 BOM 数据"
        if done:
            percent = 100.0
        progress_callback({
            "title": "BOM 匹配完成" if done else "BOM 正在匹配",
            "subtitle": "正在解析并匹配上传的 BOM 数据" if not done else "已完成本次 BOM 上传匹配",
            "current_text": current_text,
            "processed_rows": processed_rows,
            "total_rows": total_rows,
            "percent": percent,
            "done": done,
            "elapsed_seconds": elapsed_seconds,
            "chips": [
                {"label": "匹配成功", "value": str(matched_count), "tone": "success"},
                {"label": "完全匹配", "value": str(exact_count), "tone": "success"},
                {"label": "部分参数", "value": str(partial_count), "tone": "warn"},
                {"label": "高代低", "value": str(substitute_count), "tone": ""},
                {"label": "无匹配", "value": str(no_match_count), "tone": "warn"},
                {"label": "失败", "value": str(fail_count), "tone": "fail"},
            ],
        })

    emit_progress(0, {"BOM行号": "", "解析输入": "准备开始 BOM 匹配"}, done=False)
    for idx, record in enumerate(records):
        result_row = build_bom_upload_result_row(
            None,
            idx,
            record,
            column_mapping,
            query_cache=query_cache,
            full_df_provider=full_df_provider,
        )
        rows.append(result_row)

        status_text = clean_text(result_row.get("状态", ""))
        level_text = clean_text(result_row.get("首选推荐等级", ""))
        if status_text == "匹配成功":
            matched_count += 1
            if level_text == "完全匹配":
                exact_count += 1
            elif level_text == "部分参数匹配":
                partial_count += 1
            elif level_text in {"高代低", "可直接替代"}:
                substitute_count += 1
        elif status_text == "无匹配":
            no_match_count += 1
        else:
            fail_count += 1

        now_ts = time.perf_counter()
        processed_rows = idx + 1
        if processed_rows == 1 or processed_rows == total_rows or (processed_rows - last_emit_index) >= emit_every or (now_ts - last_emit_ts) >= 0.2:
            emit_progress(processed_rows, result_row=result_row, done=False)
            last_emit_ts = now_ts
            last_emit_index = processed_rows
    result_df = pd.DataFrame(rows)
    emit_progress(total_rows, result_row=rows[-1] if rows else {"BOM行号": "", "解析输入": "BOM 匹配已完成"}, done=True)
    return move_reference_model_columns_after_rank(result_df)

def style_bom_result_rows(df):
    def row_style(row):
        level = clean_text(row.get("首选推荐等级", ""))
        if level == "完全匹配":
            return ["background-color: #fff59d; color: #111111;" for _ in row]
        if level == "部分参数匹配":
            return ["background-color: #fde2e1; color: #7f1d1d;" for _ in row]
        if level in {"高代低", "可直接替代"}:
            return ["background-color: #dbeafe; color: #1d4ed8;" for _ in row]
        return ["" for _ in row]
    try:
        return df.style.apply(row_style, axis=1)
    except:
        return df

def format_bom_reference_models_for_export(row):
    if row is None:
        return ""

    row_index = list(getattr(row, "index", []))
    signal_col = None
    hwa_col = None
    for col in row_index:
        col_text = clean_text(str(col))
        if signal_col is None and "信昌" in col_text and "料号" in col_text:
            signal_col = col
        elif hwa_col is None and "华科" in col_text and "料号" in col_text:
            hwa_col = col

    parts = []
    for label, source_col in [("信昌料号", signal_col), ("华科料号", hwa_col)]:
        if source_col is None:
            continue
        value = clean_text(row.get(source_col, ""))
        models = split_model_list(value)
        if not models:
            continue
        block_lines = [f"{label}:"]
        block_lines.extend(models)
        parts.append("\n".join(block_lines))

    return "\n\n".join(parts)


def build_bom_matched_export_df(upload_df, result_df):
    if upload_df is None:
        return pd.DataFrame()

    export_df = upload_df.copy().reset_index(drop=True)
    export_col_name = "信昌/华科匹配型号"
    if export_col_name in export_df.columns:
        export_col_name = "信昌/华科匹配型号(导出)"

    if result_df is None or result_df.empty:
        export_df[export_col_name] = ""
        return export_df

    result_work = result_df.reset_index(drop=True).copy()
    max_len = len(export_df)
    matched_values = []
    for idx in range(max_len):
        if idx < len(result_work):
            matched_values.append(format_bom_reference_models_for_export(result_work.iloc[idx]))
        else:
            matched_values.append("")
    export_df[export_col_name] = matched_values
    return export_df


def copy_excel_cell_style(src_cell, dst_cell):
    if src_cell is None or dst_cell is None:
        return
    try:
        dst_cell._style = copy(src_cell._style)
    except Exception:
        pass
    try:
        dst_cell.font = copy(src_cell.font)
    except Exception:
        pass
    try:
        dst_cell.fill = copy(src_cell.fill)
    except Exception:
        pass
    try:
        dst_cell.border = copy(src_cell.border)
    except Exception:
        pass
    try:
        dst_cell.number_format = src_cell.number_format
    except Exception:
        pass
    try:
        dst_cell.protection = copy(src_cell.protection)
    except Exception:
        pass
    try:
        dst_cell.alignment = copy(src_cell.alignment)
    except Exception:
        pass


def build_bom_workbook_run_signature(uploaded_file, sheet_mappings):
    return json.dumps(
        {
            "file": build_uploaded_file_signature(uploaded_file),
            "sheet_mappings": sheet_mappings or {},
        },
        sort_keys=True,
        ensure_ascii=True,
    )


def append_export_columns_to_worksheet(ws, source_df, export_df):
    if ws is None or export_df is None or export_df.empty:
        return

    source_columns = list(source_df.columns) if source_df is not None and not source_df.empty else []
    added_columns = [col for col in export_df.columns if col not in source_columns]
    if not added_columns:
        return

    base_col = max(1, len(source_columns) if source_columns else int(ws.max_column or 1))
    header_template_col = max(1, min(base_col, int(ws.max_column or base_col)))
    max_rows = len(export_df) + 1

    for offset, column_name in enumerate(added_columns, start=1):
        target_col = base_col + offset
        header_cell = ws.cell(row=1, column=target_col)
        header_cell.value = column_name
        copy_excel_cell_style(ws.cell(row=1, column=header_template_col), header_cell)
        header_cell.alignment = Alignment(wrap_text=True, vertical="center")

        values = export_df[column_name].tolist()
        width_hint = max([len(clean_text(column_name))] + [len(clean_text(x)) for x in values[:50]] + [12])
        ws.column_dimensions[get_column_letter(target_col)].width = min(max(width_hint * 1.2, 16), 40)

        for row_idx in range(2, max_rows + 1):
            cell = ws.cell(row=row_idx, column=target_col)
            value_index = row_idx - 2
            cell.value = values[value_index] if value_index < len(values) else ""
            source_row_idx = min(max(1, row_idx), len(export_df) + 1)
            copy_excel_cell_style(ws.cell(row=source_row_idx, column=header_template_col), cell)
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_bom_workbook_sheet_results(bom_workbook, sheet_mappings=None, progress_callback=None):
    sheet_frames = list((bom_workbook or {}).get("sheet_frames", []) or [])
    if not sheet_frames:
        return []

    sheet_mappings = sheet_mappings or {}
    total_rows = sum(len(item.get("df", pd.DataFrame())) for item in sheet_frames)
    processed_rows = 0
    sheet_results = []
    start_ts = time.perf_counter()

    def emit_workbook_progress(state, sheet_name="", sheet_index=0, sheet_count=0, sheet_offset=0):
        if progress_callback is None:
            return
        payload = dict(state or {})
        local_processed = int(payload.get("processed_rows", 0) or 0)
        global_processed = min(total_rows, sheet_offset + local_processed)
        payload["processed_rows"] = global_processed
        payload["total_rows"] = total_rows
        payload["percent"] = 100.0 if total_rows <= 0 else min(100.0, (global_processed / total_rows) * 100.0)
        is_last_sheet = sheet_count > 0 and sheet_index >= sheet_count
        payload["done"] = bool(payload.get("done", False)) and is_last_sheet
        payload["title"] = "BOM 匹配完成" if payload["done"] else "BOM 正在匹配"
        current_text = clean_text(payload.get("current_text", ""))
        if sheet_count > 1:
            sheet_prefix = f"分页 {sheet_index}/{sheet_count}"
            current_text = f"{sheet_prefix} · {current_text}" if current_text else sheet_prefix
            subtitle = clean_text(payload.get("subtitle", ""))
            payload["subtitle"] = f"{sheet_name} · {subtitle}" if subtitle else sheet_name
        else:
            if current_text and sheet_name:
                payload["current_text"] = f"{sheet_name} · {current_text}"
            payload["subtitle"] = clean_text(payload.get("subtitle", "")) or sheet_name
        payload["elapsed_seconds"] = max(0.0, time.perf_counter() - start_ts)
        progress_callback(payload)

    for sheet_index, sheet_info in enumerate(sheet_frames, start=1):
        sheet_name = clean_text(sheet_info.get("sheet_name", f"Sheet{sheet_index}")) or f"Sheet{sheet_index}"
        sheet_df = sheet_info.get("df", pd.DataFrame())
        sheet_mapping = sheet_mappings.get(sheet_name)
        if sheet_mapping is None:
            sheet_mapping = guess_bom_column_mapping(sheet_df)

        sheet_offset = processed_rows

        def sheet_progress(progress_state, _sheet_name=sheet_name, _sheet_index=sheet_index, _sheet_count=len(sheet_frames), _sheet_offset=sheet_offset):
            emit_workbook_progress(progress_state, sheet_name=_sheet_name, sheet_index=_sheet_index, sheet_count=_sheet_count, sheet_offset=_sheet_offset)

        result_df = bom_dataframe_from_upload(None, sheet_df, sheet_mapping, progress_callback=sheet_progress if progress_callback is not None else None)
        sheet_results.append({
            "sheet_name": sheet_name,
            "source_df": sheet_df,
            "result_df": result_df,
            "mapping": sheet_mapping,
        })
        processed_rows += len(sheet_df)
        if progress_callback is not None:
            emit_workbook_progress({
                "title": "BOM 正在匹配",
                "subtitle": f"已完成 {sheet_name}",
                "current_text": f"{sheet_name} 匹配完成",
                "processed_rows": len(sheet_df),
                "total_rows": len(sheet_df),
                "percent": 100.0 if total_rows <= 0 else min(100.0, processed_rows / total_rows * 100.0),
                "done": False,
                "elapsed_seconds": max(0.0, time.perf_counter() - start_ts),
                "chips": [
                    {"label": "阶段", "value": "匹配中", "tone": "warn"},
                    {"label": "分页", "value": f"{sheet_index}/{len(sheet_frames)}", "tone": "warn"},
                ],
            }, sheet_name=sheet_name, sheet_index=sheet_index, sheet_count=len(sheet_frames), sheet_offset=sheet_offset)

    return sheet_results


def sanitize_excel_formula_value(value):
    if value is None:
        return value
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped.startswith(("=", "+", "-", "@")):
            return "'" + value
    return value


def sanitize_dataframe_for_excel_export(df):
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df.copy()
    work = df.copy()
    for col in work.columns:
        work[col] = work[col].map(sanitize_excel_formula_value)
    return work


def bom_to_excel_bytes(result_df, source_df=None, source_workbook=None, sheet_results=None):
    if source_workbook is not None and sheet_results is not None:
        raw_bytes = source_workbook.get("file_bytes", b"")
        if raw_bytes:
            try:
                wb = load_workbook(BytesIO(raw_bytes))
                result_map = {clean_text(item.get("sheet_name", "")): item for item in (sheet_results or [])}
                for ws in wb.worksheets:
                    payload = result_map.get(clean_text(ws.title))
                    if not payload:
                        continue
                    safe_source_df = sanitize_dataframe_for_excel_export(payload.get("source_df"))
                    safe_result_df = sanitize_dataframe_for_excel_export(payload.get("result_df"))
                    export_df = build_bom_matched_export_df(safe_source_df, safe_result_df)
                    append_export_columns_to_worksheet(ws, safe_source_df, export_df)
                    try:
                        ws.freeze_panes = ws.freeze_panes or "A2"
                    except Exception:
                        pass
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                try:
                    wb.close()
                except Exception:
                    pass
                return output.getvalue()
            except Exception:
                pass

    if source_df is not None and not source_df.empty:
        export_df = build_bom_matched_export_df(
            sanitize_dataframe_for_excel_export(source_df),
            sanitize_dataframe_for_excel_export(result_df),
        )
    else:
        export_df = sanitize_dataframe_for_excel_export(result_df) if result_df is not None else pd.DataFrame()
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="BOM匹配结果")
        sheet = writer.sheets["BOM匹配结果"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        export_col_name = "信昌/华科匹配型号"
        if export_col_name not in export_df.columns and "信昌/华科匹配型号(导出)" in export_df.columns:
            export_col_name = "信昌/华科匹配型号(导出)"
        if export_col_name in export_df.columns:
            export_col_idx = export_df.columns.get_loc(export_col_name) + 1
            export_letter = get_column_letter(export_col_idx)
            sheet.column_dimensions[export_letter].width = 28
            for excel_row in range(2, len(export_df) + 2):
                cell = sheet.cell(row=excel_row, column=export_col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    output.seek(0)
    return output.getvalue()



SSL_CTX = ssl.create_default_context()

def http_get_text(url, timeout=12):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0",
            "Accept-Language": "en-US,en;q=0.9,zh-CN;q=0.8"
        }
    )
    with urllib.request.urlopen(req, timeout=timeout, context=SSL_CTX) as resp:
        charset = resp.headers.get_content_charset() or "utf-8"
        return resp.read().decode(charset, errors="ignore")

def infer_official_brand(part):
    p = clean_model(part)
    if p.startswith(("CL","CA","CH","CJ","CK","CN")):
        return "Samsung"
    if p.startswith(("KGM","KGF","KGL","KGA","KAL","KGU","KGQ")):
        return "KYOCERA AVX"
    if p.startswith("C") and len(p) >= 5:
        size4 = p[1:5]
        if size4 in {"0402","0603","0805","1206","1210","1812","2220"}:
            return "KEMET"
        if size4 in {"1005","1608","2012","3216","3225","4532","5750"} or p.startswith(("CGA","CGJ","CGB","CGAF")):
            return "TDK"
    return ""

def samsung_base_part(part):
    p = clean_model(part)
    if len(p) >= 2 and p[-1].isalnum():
        return p[:-1]
    return p

def verify_samsung_official(part):
    p = clean_model(part)
    candidates = [f"https://product.samsungsem.com/mlcc/{p}.do"]
    base = samsung_base_part(p)
    if base != p:
        candidates.append(f"https://product.samsungsem.com/mlcc/{base}.do")

    last_error = ""
    for url in candidates:
        try:
            html = http_get_text(url)
            up = html.upper()
            if p in up:
                return {
                    "输入料号": p,
                    "校验品牌": "Samsung",
                    "官网校验状态": "官网确认存在",
                    "官网校验备注": "Samsung 官方页面命中料号",
                    "官方查询链接": url
                }
            if base in up and ("PACKAGE CODE" in up or "LIST OF PART NUMBERS WITH PACKAGE CODES" in up):
                return {
                    "输入料号": p,
                    "校验品牌": "Samsung",
                    "官网校验状态": "官网确认存在",
                    "官网校验备注": "Samsung 官方基础料号页命中，包装码列表可见",
                    "官方查询链接": url
                }
        except Exception as e:
            last_error = str(e)

    return {
        "输入料号": p,
        "校验品牌": "Samsung",
        "官网校验状态": "官网未查到，需人工核查",
        "官网校验备注": "Samsung 官方页面未命中或访问受限" + (f"；{last_error}" if last_error else ""),
        "官方查询链接": "https://product.samsungsem.com/mlcc/basic-search.do"
    }

def verify_tdk_official(part):
    p = clean_model(part)
    search_url = "https://product.tdk.com/en/search/capacitor/ceramic/mlcc/part_no"
    try:
        html = http_get_text(search_url)
        up = html.upper()
        if "SEARCH BY PART NUMBER" in up and "UP TO 50 PART NUMBERS CAN BE SEARCHED" in up:
            return {
                "输入料号": p,
                "校验品牌": "TDK",
                "官网校验状态": "品牌支持自动校验，但当前版本未直连结果页",
                "官网校验备注": "TDK 官网支持 Part Number Search；建议人工在官方页粘贴料号确认",
                "官方查询链接": search_url
            }
    except Exception as e:
        return {
            "输入料号": p,
            "校验品牌": "TDK",
            "官网校验状态": "查询失败，稍后重试",
            "官网校验备注": f"TDK 官网访问失败：{e}",
            "官方查询链接": search_url
        }
    return {
        "输入料号": p,
        "校验品牌": "TDK",
        "官网校验状态": "官网未查到，需人工核查",
        "官网校验备注": "TDK 官方页未返回可判定结果",
        "官方查询链接": search_url
    }

def verify_kemet_official(part):
    p = clean_model(part)
    search_url = "https://www.yageogroup.com/en/us/support/design-and-product-information/how-do-i-find-a-part-number.html"
    try:
        html = http_get_text(search_url)
        up = html.upper()
        if "SEARCH FOR A SPECIFIC PART NUMBER" in up or "YOUR PART WILL BE DISPLAYED IN THE SEARCH RESULTS" in up:
            return {
                "输入料号": p,
                "校验品牌": "KEMET",
                "官网校验状态": "品牌支持自动校验，但当前版本未直连结果页",
                "官网校验备注": "KEMET/Yageo 官网说明支持按 specific part number 搜索，建议人工点官方搜索页确认",
                "官方查询链接": search_url
            }
    except Exception as e:
        return {
            "输入料号": p,
            "校验品牌": "KEMET",
            "官网校验状态": "查询失败，稍后重试",
            "官网校验备注": f"KEMET 官网访问失败：{e}",
            "官方查询链接": search_url
        }
    return {
        "输入料号": p,
        "校验品牌": "KEMET",
        "官网校验状态": "官网未查到，需人工核查",
        "官网校验备注": "KEMET 官方页未返回可判定结果",
        "官方查询链接": search_url
    }

def verify_kyocera_avx_official(part):
    p = clean_model(part)
    query_url = "https://www.kyocera-avx.com/?s=" + urllib.parse.quote(p)
    try:
        html = http_get_text(query_url)
        up = html.upper()
        if p in up:
            return {
                "输入料号": p,
                "校验品牌": "KYOCERA AVX",
                "官网校验状态": "官网确认存在",
                "官网校验备注": "KYOCERA AVX 官方站点搜索命中",
                "官方查询链接": query_url
            }
        if "PART NUMBER INFORMATION" in up or "PARAMETRIC SEARCH TOOL" in up:
            return {
                "输入料号": p,
                "校验品牌": "KYOCERA AVX",
                "官网校验状态": "品牌支持自动校验，但当前结果未确认",
                "官网校验备注": "KYOCERA AVX 官方站有系列页与 Part Number Information，建议人工核查",
                "官方查询链接": query_url
            }
    except Exception as e:
        return {
            "输入料号": p,
            "校验品牌": "KYOCERA AVX",
            "官网校验状态": "查询失败，稍后重试",
            "官网校验备注": f"KYOCERA AVX 官网访问失败：{e}",
            "官方查询链接": query_url
        }
    return {
        "输入料号": p,
        "校验品牌": "KYOCERA AVX",
        "官网校验状态": "官网未查到，需人工核查",
        "官网校验备注": "KYOCERA AVX 官方站未命中该料号",
        "官方查询链接": query_url
    }

def verify_official_part(part, brand_mode="自动判断"):
    p = clean_model(part)
    if p == "":
        return None

    brand = brand_mode if brand_mode != "自动判断" else infer_official_brand(p)

    if brand == "Samsung":
        return verify_samsung_official(p)
    if brand == "TDK":
        return verify_tdk_official(p)
    if brand == "KEMET":
        return verify_kemet_official(p)
    if brand == "KYOCERA AVX":
        return verify_kyocera_avx_official(p)

    return {
        "输入料号": p,
        "校验品牌": brand if brand else "未识别",
        "官网校验状态": "品牌暂不支持自动校验",
        "官网校验备注": "当前版本仅内建 Samsung / TDK / KEMET / KYOCERA AVX",
        "官方查询链接": ""
    }

def official_verify_dataframe(lines, brand_mode="自动判断"):
    rows = []
    for line in lines:
        row = verify_official_part(line, brand_mode)
        if row:
            rows.append(row)
    return pd.DataFrame(rows)


def load_search_dataframe_for_action(action_label):
    cache_signature = get_data_cache_signature()
    with st.spinner(f"正在加载元件库，准备{action_label}..."):
        return load_prepared_data(cache_signature)


SEARCH_PROGRESS_STAGE_COUNT = 4


def resolve_search_query_dataframe_and_spec(
    line,
    get_full_search_df=None,
    progress_callback=None,
    exact_part_rows=None,
    query_frame_cache=None,
):
    def emit(stage_step, stage_text, note="", source_label="", source_tone="", candidate_rows=None):
        if progress_callback is None:
            return
        progress_callback(
            {
                "stage_step": stage_step,
                "stage_text": stage_text,
                "note": note,
                "source_label": source_label,
                "source_tone": source_tone,
                "candidate_rows": candidate_rows,
            }
        )

    emit(1, "正在解析输入", "先按命名规则和规格关键词识别当前输入")
    prefetched_exact_rows = resolve_prefetched_exact_part_rows(line, exact_part_rows=exact_part_rows)
    detect_df = prefetched_exact_rows if isinstance(prefetched_exact_rows, pd.DataFrame) and not prefetched_exact_rows.empty else None
    mode, spec = detect_query_mode_and_spec(detect_df, line)
    query_df = None
    candidate_rows = 0
    query_frame_cache_key = ""

    if mode != "无法识别" and spec is not None:
        if isinstance(query_frame_cache, dict):
            query_frame_cache_key = make_query_cache_key("", f"query_df::{mode}", spec)
            cached_query_df = query_frame_cache.get(query_frame_cache_key)
        if query_frame_cache_key == "":
            query_frame_cache_key = make_query_cache_key("", f"query_df::{mode}", spec)
        cached_query_df = query_frame_cache.get(query_frame_cache_key) if isinstance(query_frame_cache, dict) else None
        cache_source = "本轮缓存"
        if not isinstance(cached_query_df, pd.DataFrame) or cached_query_df.empty:
            session_query_df_cache = get_session_query_dataframe_cache()
            session_cached_df = session_query_df_cache.get(query_frame_cache_key)
            if isinstance(session_cached_df, pd.DataFrame) and not session_cached_df.empty:
                cached_query_df = session_cached_df.copy()
                cache_source = "会话缓存"
        if isinstance(cached_query_df, pd.DataFrame) and not cached_query_df.empty:
            query_df = cached_query_df
            if mode == "料号" and isinstance(exact_part_rows, pd.DataFrame) and not exact_part_rows.empty:
                query_df = concat_component_frames([exact_part_rows, query_df])
            candidate_rows = len(query_df)
            emit(
                2,
                "已复用候选缓存",
                "相同输入或同规格结果已缓存，无需再次查库",
                cache_source,
                candidate_rows=candidate_rows,
            )
            return {
                "query_df": query_df,
                "mode": mode,
                "spec": spec,
                "resolution_path": "cached_fast_query",
                "used_full_df": False,
                "candidate_rows": candidate_rows,
            }
        candidate_note = "按识别到的规格从索引缩小候选范围"
        if not database_has_component_rows() and ensure_streamlit_cloud_bundle_archive_available():
            candidate_note += "；公网首搜可能需要 5-15 秒预热索引"
        emit(2, "正在载入候选库", candidate_note, "快速索引")
        query_df = load_search_dataframe_for_query(mode, spec, line, exact_part_rows=prefetched_exact_rows)
        if isinstance(query_df, pd.DataFrame):
            candidate_rows = len(query_df)
            if isinstance(query_frame_cache, dict) and not query_df.empty and query_frame_cache_key != "":
                query_frame_cache[query_frame_cache_key] = query_df
            if not query_df.empty and query_frame_cache_key != "":
                store_session_query_dataframe_cache(query_frame_cache_key, query_df)
        if isinstance(query_df, pd.DataFrame) and not query_df.empty:
            emit(
                2,
                "候选范围已锁定",
                "本次搜索无需加载完整元件库",
                "快速索引",
                candidate_rows=candidate_rows,
            )
            return {
                "query_df": query_df,
                "mode": mode,
                "spec": spec,
                "resolution_path": "fast_query",
                "used_full_df": False,
                "candidate_rows": candidate_rows,
            }

    if looks_like_compact_part_query(line):
        emit(2, "正在按料号直查数据库", "命名规则未完整命中时，先尝试数据库精确料号直查", "数据库直查")
        exact_df = prefetched_exact_rows
        if isinstance(exact_df, pd.DataFrame) and not exact_df.empty:
            exact_mode, exact_spec = detect_query_mode_and_spec(exact_df, line)
            if exact_mode != "无法识别" and exact_spec is not None:
                candidate_rows = len(exact_df)
                emit(
                    2,
                    "已锁定数据库原始料号",
                    "直接用数据库原始行反推规格，避免整库加载",
                    "数据库直查",
                    candidate_rows=candidate_rows,
                )
                return {
                    "query_df": exact_df,
                    "mode": exact_mode,
                    "spec": exact_spec,
                    "resolution_path": "exact_model_lookup",
                    "used_full_df": False,
                    "candidate_rows": candidate_rows,
                }
        if mode == "无法识别" and spec is None:
            emit(
                2,
                "未命中数据库原始料号",
                "命名规则未覆盖且数据库没有完全相同料号，已跳过整库回退",
                "快速失败",
                "warn",
            )
            return {
                "query_df": pd.DataFrame(),
                "mode": mode,
                "spec": spec,
                "resolution_path": "unknown_compact_part",
                "used_full_df": False,
                "candidate_rows": 0,
            }

    if query_df is not None:
        emit(
            2,
            "候选范围已确定",
            "本次无需加载完整元件库",
            "快速索引",
            candidate_rows=candidate_rows,
        )
        return {
            "query_df": query_df,
            "mode": mode,
            "spec": spec,
            "resolution_path": "fast_query_empty",
            "used_full_df": False,
            "candidate_rows": candidate_rows,
        }

    emit(
        2,
        "正在加载完整元件库",
        "这条输入需要整库回退，首次可能需要 30-60 秒",
        "整库回退",
        "warn",
    )
    query_df = get_full_search_df() if callable(get_full_search_df) else load_search_dataframe_for_action("搜索")
    if query_df is None:
        query_df = pd.DataFrame()
    candidate_rows = len(query_df) if isinstance(query_df, pd.DataFrame) else 0
    if not query_df.empty:
        emit(
            2,
            "完整元件库已载入",
            "已切换到整库回退模式继续匹配",
            "整库回退",
            "warn",
            candidate_rows=candidate_rows,
        )
        mode, spec = detect_query_mode_and_spec(query_df, line)
    return {
        "query_df": query_df,
        "mode": mode,
        "spec": spec,
        "resolution_path": "full_dataframe",
        "used_full_df": True,
        "candidate_rows": candidate_rows,
    }


def build_uploaded_file_signature(uploaded_file):
    if uploaded_file is None:
        return ""
    raw_bytes = get_uploaded_file_bytes(uploaded_file)
    return json.dumps(
        {
            "name": clean_text(getattr(uploaded_file, "name", "")),
            "size": int(getattr(uploaded_file, "size", 0) or 0),
            "sha256": hashlib.sha256(raw_bytes).hexdigest() if raw_bytes else "",
        },
        sort_keys=True,
        ensure_ascii=True,
    )


def build_bom_run_signature(uploaded_file, selected_mapping):
    return json.dumps(
        {
            "file": build_uploaded_file_signature(uploaded_file),
            "mapping": selected_mapping or {},
        },
        sort_keys=True,
        ensure_ascii=True,
    )


if __name__ == "__main__" and "--rebuild-search-index" in sys.argv:
    rebuild_search_index_from_database_fast()
    raise SystemExit(0)


if __name__ == "__main__" and "--rebuild-db" in sys.argv:
    update_database(force=True)
    raise SystemExit(0)


if __name__ == "__main__" and "--rebuild-prepared-cache" in sys.argv:
    rebuild_prepared_cache_from_database()
    raise SystemExit(0)

if __name__ == "__main__" and "--backfill-series" in sys.argv:
    backfill_series_fields_in_database()
    raise SystemExit(0)


require_app_access()
startup_trace("after_require_app_access")

if get_configured_access_code() != "" and st.session_state.get("_app_access_granted", False):
    with st.sidebar:
        if st.button("退出访问码", use_container_width=True):
            st.session_state.pop("_app_access_granted", None)
            st.rerun()
STARTUP_MAINTENANCE_ENABLED = clean_text(os.getenv("COMPONENT_MATCHER_STARTUP_MAINTENANCE", "")) in {"1", "true", "yes", "on"}

if STARTUP_MAINTENANCE_ENABLED and not is_component_matcher_build_mode() and not (is_public_mode() or is_streamlit_cloud_runtime()) and database_has_component_rows():
    try:
        startup_trace("before_maybe_update_database_on_startup")
        maybe_update_database(force=False)
        startup_trace("after_maybe_update_database_on_startup")
    except Exception as exc:
        startup_trace(f"startup_init_error:{type(exc).__name__}")
        st.error("云端初始化失败，请查看下方错误详情。")
        st.exception(exc)
        st.code(traceback.format_exc())
        st.stop()
elif STARTUP_MAINTENANCE_ENABLED and not is_component_matcher_build_mode() and not (is_public_mode() or is_streamlit_cloud_runtime()):
    try:
        startup_trace("before_cloud_search_warmup")
        maybe_start_cloud_search_asset_warmup()
        startup_trace("after_cloud_search_warmup")
    except Exception as exc:
        startup_trace(f"startup_init_error:{type(exc).__name__}")
        st.error("云端初始化失败，请查看下方错误详情。")
        st.exception(exc)
        st.code(traceback.format_exc())
        st.stop()


logo_b64 = image_to_base64(LOGO_PATH)
startup_trace(f"logo_base64:{'yes' if logo_b64 else 'no'}")
if logo_b64:
    st.markdown(
        f'''
        <div style="text-align:center; padding-top:18px; margin-bottom:10px;">
            <img src="data:image/png;base64,{logo_b64}" style="width:210px; display:block; margin:0 auto;" />
        </div>
        ''',
        unsafe_allow_html=True
    )
    startup_trace("after_logo_markdown")

st.markdown('<div class="main-title">富临通元器件匹配系统</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">输入料号自动匹配所有同规格品牌型号</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title-2">（输入多个或单个料号或规格参数，例如 FP31X333K631EEG、1206 X7R 333K 630V 或 0402 10K 1% 1/16W；规格参数至少需包含尺寸和关键参数，电容看容值/耐压，电阻看阻值/功率，并满足三个关键参数后才能进行匹配）</div>', unsafe_allow_html=True)
startup_trace("after_intro_markdown")

with st.form("manual_query_search_form", clear_on_submit=False):
    query_input = st.text_area("查询输入", placeholder="请输入料号，可多行输入", label_visibility="collapsed")
    search_clicked = st.form_submit_button("搜索")
startup_trace("after_search_form")

if search_clicked:
    if not query_input.strip():
        st.warning("请输入料号或规格参数")
    else:
        lines = [x.strip() for x in query_input.splitlines() if x.strip()]
        security_limits = get_runtime_security_limits()
        total_query_chars = sum(len(line) for line in lines)
        if len(lines) > security_limits["max_search_lines"]:
            st.error(f"单次最多允许 {security_limits['max_search_lines']} 行查询，当前输入了 {len(lines)} 行。")
            st.stop()
        if total_query_chars > security_limits["max_search_total_chars"]:
            st.error(f"单次查询总字符数不能超过 {security_limits['max_search_total_chars']}，当前为 {total_query_chars}。")
            st.stop()
        too_long_lines = [line for line in lines if len(line) > security_limits["max_search_line_chars"]]
        if too_long_lines:
            st.error(f"单行查询最多允许 {security_limits['max_search_line_chars']} 个字符，请先拆分后再搜。")
            st.stop()
        if any(re.search(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", line) for line in lines):
            st.error("查询内容包含不可见控制字符，请清理后再搜索。")
            st.stop()
        full_search_df_cache = {"loaded": False, "df": pd.DataFrame()}
        database_empty_warned = False
        search_started_at = time.time()
        progress_placeholder = st.empty()
        search_stats = {"success": 0, "no_match": 0, "warning": 0}
        aborted_reason = ""
        exact_part_prefetch_map = {}
        resolved_line_cache = {}
        query_frame_cache = {}

        def get_full_search_df():
            if not full_search_df_cache["loaded"]:
                full_search_df_cache["df"] = load_search_dataframe_for_action("搜索")
                full_search_df_cache["loaded"] = True
            return full_search_df_cache["df"]

        def render_search_progress(
            completed_queries,
            stage_step=0,
            current_text="",
            stage_text="",
            note="",
            extra_chips=None,
            done=False,
            summary_lines=None,
        ):
            render_bom_progress_card(
                progress_placeholder,
                build_search_progress_state(
                    len(lines),
                    completed_queries,
                    current_text=current_text,
                    stage_text=stage_text,
                    note=note,
                    elapsed_seconds=time.time() - search_started_at,
                    stage_step=stage_step,
                    done=done,
                    extra_chips=extra_chips,
                    summary_lines=summary_lines,
                ),
            )

        render_search_progress(0, stage_step=0, stage_text="准备开始", note="已收到搜索请求，正在准备匹配")
        if not database_has_component_rows():
            render_search_progress(
                0,
                stage_step=0,
                stage_text="正在准备数据",
                note="首次使用当前环境，正在加载搜索所需的数据包",
            )
            if not ensure_component_data_ready("搜索"):
                render_search_progress(
                    0,
                    stage_step=SEARCH_PROGRESS_STAGE_COUNT,
                    stage_text="数据准备失败",
                    note="当前环境缺少可用数据库，无法继续搜索",
                    done=True,
                    summary_lines=["未能准备搜索数据库，请检查部署数据包或本地数据库是否完整。"],
                )
                st.error("当前环境缺少可用数据库，暂时无法执行搜索。")
                st.stop()
        exact_prefetch_lines = [line for line in dict.fromkeys(lines) if looks_like_compact_part_query(line)]
        if exact_prefetch_lines:
            render_search_progress(
                0,
                stage_step=0,
                stage_text="正在预取精确料号",
                note="先批量锁定输入料号的数据库原始行，减少逐条查库开销",
                extra_chips=[{"label": "预取数", "value": str(len(exact_prefetch_lines))}],
            )
            exact_part_prefetch_map = load_component_rows_by_clean_models_map(exact_prefetch_lines)

        for line_index, line in enumerate(lines, start=1):
            resolved_state = {
                "stage_step": 1,
                "stage_text": "正在解析输入",
                "note": "识别料号 / 料号片段 / 规格条件",
                "source_label": "",
                "source_tone": "",
                "candidate_rows": None,
            }

            def update_line_progress(payload):
                if isinstance(payload, dict):
                    resolved_state.update(payload)
                extra_chips = []
                source_label = clean_text(resolved_state.get("source_label", ""))
                source_tone = clean_text(resolved_state.get("source_tone", ""))
                candidate_rows = resolved_state.get("candidate_rows", None)
                if source_label != "":
                    extra_chips.append({"label": "路径", "value": source_label, "tone": source_tone})
                try:
                    candidate_value = int(candidate_rows)
                except Exception:
                    candidate_value = 0
                if candidate_value > 0:
                    extra_chips.append({"label": "候选数", "value": f"{candidate_value:,}"})
                render_search_progress(
                    line_index - 1,
                    stage_step=resolved_state.get("stage_step", 1),
                    current_text=line,
                    stage_text=resolved_state.get("stage_text", "正在解析输入"),
                    note=resolved_state.get("note", ""),
                    extra_chips=extra_chips,
                )

            line_cache_key = clean_text(line)
            prefetched_exact_rows = exact_part_prefetch_map.get(clean_model(line), pd.DataFrame()) if looks_like_compact_part_query(line) else None
            if line_cache_key in resolved_line_cache:
                resolved = resolved_line_cache[line_cache_key]
                cached_rows = 0 if resolved.get("query_df") is None else len(resolved.get("query_df"))
                render_search_progress(
                    line_index - 1,
                    stage_step=2,
                    current_text=line,
                    stage_text="已复用相同输入缓存",
                    note="同一轮内重复输入不再重新解析和查库",
                    extra_chips=[
                        {"label": "路径", "value": "本轮缓存"},
                        {"label": "候选数", "value": f"{cached_rows:,}"} if cached_rows > 0 else None,
                    ],
                )
            else:
                render_search_progress(line_index - 1, stage_step=1, current_text=line, stage_text="正在解析输入", note="识别料号 / 料号片段 / 规格条件")
                resolved = resolve_search_query_dataframe_and_spec(
                    line,
                    get_full_search_df=get_full_search_df,
                    progress_callback=update_line_progress,
                    exact_part_rows=prefetched_exact_rows,
                    query_frame_cache=query_frame_cache,
                )
                resolved_line_cache[line_cache_key] = resolved
            query_df = resolved.get("query_df")
            mode = resolved.get("mode")
            spec = resolved.get("spec")
            resolution_path = clean_text(resolved.get("resolution_path", ""))
            try:
                candidate_rows = int(resolved.get("candidate_rows") or 0)
            except Exception:
                candidate_rows = 0
            if mode in {"无法识别", "规格不足"} or spec is None:
                forced_exact_rows = resolve_prefetched_exact_part_rows(line, exact_part_rows=prefetched_exact_rows)
                if isinstance(forced_exact_rows, pd.DataFrame) and not forced_exact_rows.empty:
                    forced_mode, forced_spec = detect_query_mode_and_spec(forced_exact_rows, line)
                    if forced_mode == "料号" and forced_spec is not None:
                        query_df = forced_exact_rows
                        mode = forced_mode
                        spec = forced_spec
                        resolution_path = "forced_exact_model_lookup"
                        candidate_rows = len(forced_exact_rows)
                        resolved_state.update(
                            {
                                "stage_step": 2,
                                "stage_text": "已锁定数据库原始料号",
                                "note": "输入已命中正式库原厂料号，按原厂资料继续展示",
                                "source_label": "数据库直查",
                                "source_tone": "success",
                                "candidate_rows": candidate_rows,
                            }
                        )
            source_label = clean_text(resolved_state.get("source_label", ""))
            source_tone = clean_text(resolved_state.get("source_tone", ""))
            base_chips = []
            if source_label != "":
                base_chips.append({"label": "路径", "value": source_label, "tone": source_tone})
            if candidate_rows > 0:
                base_chips.append({"label": "候选数", "value": f"{candidate_rows:,}"})

            if resolution_path == "full_dataframe" and (query_df is None or query_df.empty):
                render_search_progress(
                    line_index - 1,
                    stage_step=2,
                    current_text=line,
                    stage_text="整库回退未就绪",
                    note="当前输入未命中快速索引，且当前环境未加载整库回退数据；已跳过这一条并继续后续输入",
                    extra_chips=base_chips or [{"label": "路径", "value": "整库回退", "tone": "warn"}],
                )
                if not database_empty_warned:
                    st.warning("当前环境未加载整库回退数据；本条输入已跳过，后续输入会继续处理。")
                    database_empty_warned = True
                search_stats["warning"] += 1
                continue

            if mode == "无法识别" or spec is None:
                render_search_progress(
                    line_index - 1,
                    stage_step=3,
                    current_text=line,
                    stage_text="输入无法识别",
                    note="请检查料号是否完整，或补充规格参数后再试",
                    extra_chips=base_chips,
                )
                st.warning("无法识别输入内容")
                search_stats["warning"] += 1
                continue

            if mode == "规格不足":
                render_search_progress(
                    line_index - 1,
                    stage_step=3,
                    current_text=line,
                    stage_text="规格参数不足",
                    note="当前输入少于最小匹配条件，请补充至少三个关键参数",
                    extra_chips=base_chips,
                )
                st.warning("请最少输入三个规格参数")
                search_stats["warning"] += 1
                continue

            render_search_progress(
                line_index - 1,
                stage_step=3,
                current_text=line,
                stage_text="正在执行匹配",
                note="正在比对候选料号并计算推荐等级",
                extra_chips=base_chips,
            )
            if mode == "料号":
                part_info_df = build_part_info_df(query_df, spec, line)
                matched = cached_run_query_match(query_df, mode, spec, query_text=line)
                matched = exclude_brand_model_pairs(matched, extract_brand_model_pairs(part_info_df))
                part_info_fragment = render_clickable_result_table(
                    part_info_df,
                    show_official_status=False,
                    footer_html="",
                    wrap_iframe=False,
                )
                match_card_header_html = (
                    '<div style="display:flex; align-items:center; justify-content:flex-start; gap:10px; '
                    'margin:0 2px 4px 2px; padding:0;">'
                    '<div style="font-size:20px; font-weight:800; color:#1f2937; line-height:1.2;">匹配料号资料</div>'
                    f'<div style="max-width:48%; padding:8px 14px; border-radius:999px; '
                    'border:1px solid rgba(59,130,246,0.28); background:linear-gradient(180deg, rgba(59,130,246,0.10) 0%, rgba(59,130,246,0.04) 100%); '
                    'color:#1d4ed8; font-size:15px; font-weight:700; line-height:1.35; word-break:break-all; text-align:left;">'
                    f'{html.escape(line)}'
                    '</div>'
                    '</div>'
                )
                if not matched.empty:
                    render_search_progress(
                        line_index - 1,
                        stage_step=4,
                        current_text=line,
                        stage_text="正在整理结果",
                        note=f"已命中 {len(matched)} 条其他品牌匹配结果，正在生成展示内容",
                        extra_chips=base_chips + [{"label": "命中数", "value": str(len(matched)), "tone": "success"}],
                    )
                    matched = matched.copy()
                    matched["尺寸（inch）"] = matched["尺寸（inch）"].apply(clean_size)
                    matched["材质（介质）"] = matched["材质（介质）"].apply(clean_material)
                    matched["容值误差"] = matched["容值误差"].apply(clean_tol_for_match)
                    matched["耐压（V）"] = matched["耐压（V）"].apply(clean_voltage)
                    matched = ensure_component_display_columns(matched)
                    allow_online_lookup = infer_spec_component_type(spec) == "MLCC" and len(matched) <= 20
                    show_df = select_component_display_columns(
                        matched,
                        spec,
                        prefix_columns=["推荐等级", "品牌", "型号", "器件类别", "系列"],
                        suffix_columns=["备注1", "备注2", "备注3"],
                        allow_online_lookup=allow_online_lookup,
                    )
                    show_df = format_display_df(show_df)
                    show_df = annotate_match_display_gaps(show_df, spec)
                    if infer_spec_component_type(spec) == "MLCC" and "特殊用途" in show_df.columns:
                        show_df = show_df.drop(columns=["特殊用途"])

                    result_fragment = render_clickable_result_table(
                        show_df,
                        spec=spec,
                        footer_html="",
                        wrap_iframe=False,
                    )
                    match_card_html = (
                        f'{match_card_header_html}'
                        f'{part_info_fragment}'
                        '<div style="height:1px; margin:8px 0 6px 0; background:rgba(191,219,254,0.78);"></div>'
                        '<div style="font-size:20px; font-weight:800; color:#1f2937; line-height:1.2; margin:0 0 4px 2px;">匹配结果</div>'
                        f'{result_fragment}'
                        '<div class="match-card-footer"></div>'
                    )
                    components.html(
                        build_result_table_iframe_html(match_card_html),
                        height=estimate_match_card_iframe_height(len(part_info_df), len(show_df)),
                        scrolling=False,
                    )
                    st.markdown('<div style="height:0px;"></div>', unsafe_allow_html=True)
                    search_stats["success"] += 1
                else:
                    render_search_progress(
                        line_index - 1,
                        stage_step=4,
                        current_text=line,
                        stage_text="当前输入已完成",
                        note="已定位到原厂料号资料，但暂未找到其他品牌匹配结果",
                        extra_chips=base_chips + [{"label": "命中数", "value": "0", "tone": "warn"}],
                    )
                    match_card_html = (
                        f'{match_card_header_html}'
                        f'{part_info_fragment}'
                        '<div class="match-card-footer"></div>'
                        '<div class="no-alt-match-alert">已找到原厂料号资料，暂未找到其他品牌替代结果</div>'
                    )
                    components.html(
                        build_result_table_iframe_html(match_card_html),
                        height=estimate_match_card_iframe_height(len(part_info_df), 0) + 72,
                        scrolling=False,
                    )
                    if part_info_df is not None and not part_info_df.empty:
                        st.markdown('<div style="height:18px;"></div>', unsafe_allow_html=True)
                        search_stats["success"] += 1
                    else:
                        st.warning("未找到其他品牌匹配结果")
                        search_stats["no_match"] += 1
                continue
            elif mode == "料号片段":
                st.markdown('<div class="section-title">料号片段反推规格</div>', unsafe_allow_html=True)
                spec_info_df = build_spec_info_df(spec)
                st.dataframe(
                    spec_info_df,
                    use_container_width=True,
                    hide_index=True,
                    height=85,
                    column_config=build_component_column_config(spec_info_df.columns, spec)
                )
                st.markdown(f'<div class="section-title">{build_component_section_title(spec, "匹配结果（含推荐等级）")}</div>', unsafe_allow_html=True)
                matched = cached_run_query_match(query_df, mode, spec, query_text=line)
            else:
                st.markdown(f'<div class="section-title">{build_component_section_title(spec, "规格条件")}</div>', unsafe_allow_html=True)
                spec_info_df = build_spec_info_df(spec)
                st.dataframe(
                    spec_info_df,
                    use_container_width=True,
                    hide_index=True,
                    height=85,
                    column_config=build_component_column_config(spec_info_df.columns, spec)
                )
                st.markdown(f'<div class="section-title">{build_component_section_title(spec, "匹配结果（含推荐等级）")}</div>', unsafe_allow_html=True)
                matched = cached_run_query_match(query_df, mode, spec, query_text=line)

            if not matched.empty:
                render_search_progress(
                    line_index - 1,
                    stage_step=4,
                    current_text=line,
                    stage_text="正在整理结果",
                    note=f"已命中 {len(matched)} 条匹配结果，正在生成展示内容",
                    extra_chips=base_chips + [{"label": "命中数", "value": str(len(matched)), "tone": "success"}],
                )
                matched = matched.copy()
                matched["尺寸（inch）"] = matched["尺寸（inch）"].apply(clean_size)
                matched["材质（介质）"] = matched["材质（介质）"].apply(clean_material)
                matched["容值误差"] = matched["容值误差"].apply(clean_tol_for_match)
                matched["耐压（V）"] = matched["耐压（V）"].apply(clean_voltage)
                matched = ensure_component_display_columns(matched)
                allow_online_lookup = infer_spec_component_type(spec) == "MLCC" and len(matched) <= 20
                show_df = select_component_display_columns(
                    matched,
                    spec,
                    prefix_columns=["推荐等级", "品牌", "型号", "器件类别", "系列"],
                    suffix_columns=["备注1", "备注2", "备注3"],
                    allow_online_lookup=allow_online_lookup,
                )
                show_df = format_display_df(show_df)
                show_df = annotate_match_display_gaps(show_df, spec)
                if infer_spec_component_type(spec) == "MLCC" and "特殊用途" in show_df.columns:
                    show_df = show_df.drop(columns=["特殊用途"])
                clickable_table_html = render_clickable_result_table(
                    show_df,
                    spec=spec,
                    footer_html="",
                )
                if clickable_table_html:
                    components.html(
                        clickable_table_html,
                        height=estimate_result_table_iframe_height(len(show_df), show_official_status=True),
                        scrolling=False,
                    )
                else:
                    styled_show_df = style_exact_match_rows(show_df, spec=spec)
                    st.dataframe(
                        styled_show_df,
                        use_container_width=True,
                        hide_index=True,
                        height=420,
                        column_config=build_component_column_config(show_df.columns, spec)
                    )
                search_stats["success"] += 1
            else:
                render_search_progress(
                    line_index - 1,
                    stage_step=4,
                    current_text=line,
                    stage_text="当前输入已完成",
                    note="未找到符合条件的匹配结果",
                    extra_chips=base_chips + [{"label": "命中数", "value": "0", "tone": "warn"}],
                )
                if mode == "料号":
                    st.warning("未找到其他品牌匹配结果")
                elif mode == "料号片段":
                    st.warning("未找到符合片段规格的品牌料号（含推荐等级），请确认数据库里已有这些规格资料")
                else:
                    st.warning("未找到符合规格的品牌料号（含推荐等级），请确认数据库里已有这些规格资料")
                search_stats["no_match"] += 1

        processed_queries = search_stats["success"] + search_stats["no_match"] + search_stats["warning"]
        summary_lines = [f"成功返回匹配结果 {search_stats['success']} 条"]
        if search_stats["no_match"] > 0:
            summary_lines.append(f"未找到匹配结果 {search_stats['no_match']} 条")
        if search_stats["warning"] > 0:
            summary_lines.append(f"需要补充或修正输入 {search_stats['warning']} 条")
        remaining_queries = max(0, len(lines) - processed_queries)
        if remaining_queries > 0:
            summary_lines.append(f"未处理输入 {remaining_queries} 条")
        if aborted_reason:
            summary_lines.append(aborted_reason)
        render_search_progress(
            processed_queries,
            stage_step=SEARCH_PROGRESS_STAGE_COUNT,
            current_text=lines[min(processed_queries, len(lines)) - 1] if processed_queries > 0 and lines else "",
            stage_text="搜索已完成" if aborted_reason == "" else "搜索提前结束",
            note="所有输入已处理完成" if aborted_reason == "" else aborted_reason,
            extra_chips=[
                {"label": "成功", "value": str(search_stats["success"]), "tone": "success"},
                {"label": "无匹配", "value": str(search_stats["no_match"]), "tone": "warn"} if search_stats["no_match"] > 0 else None,
                {"label": "提示", "value": str(search_stats["warning"]), "tone": "warn"} if search_stats["warning"] > 0 else None,
            ],
            done=True,
            summary_lines=summary_lines,
        )

st.markdown('<div class="result-title">BOM批量上传匹配</div>', unsafe_allow_html=True)
startup_trace("after_bom_title")
uploaded_file = st.file_uploader("上传 BOM Excel/CSV", type=["xlsx", "xls", "csv"])
startup_trace("after_file_uploader")

if uploaded_file is not None:
    progress_placeholder = st.empty()
    try:
        render_bom_progress_card(
            progress_placeholder,
            {
                "title": "BOM 文件读取中",
                "subtitle": "正在解析上传的 Excel / CSV 文件",
                "current_text": getattr(uploaded_file, "name", "正在读取上传文件"),
                "processed_rows": 0,
                "total_rows": 0,
                "percent": 3.0,
                "done": False,
                "elapsed_seconds": 0.0,
                "chips": [
                    {"label": "阶段", "value": "读取文件", "tone": "warn"},
                    {"label": "状态", "value": "等待中", "tone": "warn"},
                ],
            },
        )
        bom_workbook = read_uploaded_bom_workbook(uploaded_file)
        bom_sheet_frames = bom_workbook.get("sheet_frames", [])
        security_limits = get_runtime_security_limits()
        bom_total_rows = sum(len(item.get("df", pd.DataFrame())) for item in bom_sheet_frames)
        bom_total_sheets = len(bom_sheet_frames)
        if bom_total_rows > security_limits["max_bom_rows"]:
            render_bom_progress_card(
                progress_placeholder,
                {
                    "title": "BOM 文件过大",
                    "subtitle": "当前文件行数超过了系统安全上限",
                    "current_text": f"行数：{bom_total_rows}，上限：{security_limits['max_bom_rows']}",
                    "processed_rows": 0,
                    "total_rows": bom_total_rows,
                    "percent": 100.0,
                    "done": True,
                    "elapsed_seconds": 0.0,
                    "chips": [
                        {"label": "限制", "value": "行数过大", "tone": "fail"},
                        {"label": "当前", "value": str(bom_total_rows), "tone": "fail"},
                        {"label": "上限", "value": str(security_limits["max_bom_rows"]), "tone": "warn"},
                    ],
                },
            )
            st.error(f"单次 BOM 处理行数不能超过 {security_limits['max_bom_rows']} 行。")
            st.stop()
        if bom_total_sheets > security_limits["max_bom_sheets"]:
            render_bom_progress_card(
                progress_placeholder,
                {
                    "title": "BOM 文件过大",
                    "subtitle": "当前文件分页过多，已触发安全上限",
                    "current_text": f"分页数：{bom_total_sheets}，上限：{security_limits['max_bom_sheets']}",
                    "processed_rows": 0,
                    "total_rows": bom_total_rows,
                    "percent": 100.0,
                    "done": True,
                    "elapsed_seconds": 0.0,
                    "chips": [
                        {"label": "限制", "value": "分页过多", "tone": "fail"},
                        {"label": "当前", "value": str(bom_total_sheets), "tone": "fail"},
                        {"label": "上限", "value": str(security_limits["max_bom_sheets"]), "tone": "warn"},
                    ],
                },
            )
            st.error(f"单次 BOM 分页数不能超过 {security_limits['max_bom_sheets']} 页。")
            st.stop()

        if not bom_sheet_frames:
            render_bom_progress_card(
                progress_placeholder,
                {
                    "title": "BOM 读取失败",
                    "subtitle": "上传文件内容为空，未能生成可匹配数据",
                    "current_text": getattr(uploaded_file, "name", "空文件"),
                    "processed_rows": 0,
                    "total_rows": 0,
                    "percent": 100.0,
                    "done": True,
                    "elapsed_seconds": 0.0,
                    "chips": [
                        {"label": "阶段", "value": "读取完成", "tone": "fail"},
                        {"label": "状态", "value": "空文件", "tone": "fail"},
                    ],
                },
            )
            st.warning("上传文件为空")
        else:
            workbook_signature = build_uploaded_file_signature(uploaded_file)
            if st.session_state.get("_bom_workbook_signature") != workbook_signature:
                for key in [
                    "_bom_result_signature",
                    "_bom_result_df",
                    "_bom_export_bytes",
                    "_bom_sheet_results",
                    "_bom_sheet_mappings",
                ]:
                    st.session_state.pop(key, None)
                st.session_state["_bom_workbook_signature"] = workbook_signature
                st.session_state["_bom_manual_mapping_open"] = False

            st.session_state["_bom_workbook_state"] = bom_workbook
            if "_bom_sheet_mappings" not in st.session_state:
                st.session_state["_bom_sheet_mappings"] = {}
            if "_bom_sheet_results" not in st.session_state:
                st.session_state["_bom_sheet_results"] = {}

            sheet_names = [item.get("sheet_name", f"Sheet{idx + 1}") for idx, item in enumerate(bom_sheet_frames)]
            selected_sheet_name = sheet_names[0]
            sheet_selector_key = f"bom_sheet_selector_{workbook_signature}"
            if len(sheet_names) > 1:
                selected_sheet_name = st.selectbox("分页", sheet_names, key=sheet_selector_key)
                st.caption(f"本次上传共 {len(sheet_names)} 个分页，系统会逐页匹配并在下载时保留原分页结构。")

            selected_sheet = next(
                (item for item in bom_sheet_frames if clean_text(item.get("sheet_name", "")) == clean_text(selected_sheet_name)),
                bom_sheet_frames[0],
            )
            bom_df = selected_sheet.get("df", pd.DataFrame()).copy()
            total_workbook_rows = sum(len(item.get("df", pd.DataFrame())) for item in bom_sheet_frames)

            cached_bom_result_df = pd.DataFrame()
            cached_bom_sheet_results = st.session_state.get("_bom_sheet_results", {})
            if (
                st.session_state.get("_bom_workbook_signature") == workbook_signature
                and isinstance(cached_bom_sheet_results, dict)
                and cached_bom_sheet_results
            ):
                cached_bom_result_df = cached_bom_sheet_results.get(
                    selected_sheet_name,
                    st.session_state.get("_bom_result_df", pd.DataFrame()),
                )
            if isinstance(cached_bom_result_df, pd.DataFrame) and not cached_bom_result_df.empty:
                cached_fail_count = int((cached_bom_result_df["解析状态"] == "解析失败").sum()) if "解析状态" in cached_bom_result_df.columns else 0
                cached_no_match_count = int((cached_bom_result_df["状态"] == "无匹配").sum()) if "状态" in cached_bom_result_df.columns else 0
                cached_success_count = int((cached_bom_result_df["状态"] == "匹配成功").sum()) if "状态" in cached_bom_result_df.columns else 0
                cached_component_distribution_text = build_bom_component_distribution_text(cached_bom_result_df)
                cached_summary_lines = [
                    f"解析完成：匹配成功 {cached_success_count} 行，解析失败 {cached_fail_count} 行，无匹配 {cached_no_match_count} 行。",
                ]
                if cached_component_distribution_text:
                    cached_summary_lines.append(cached_component_distribution_text)
                render_bom_progress_card(
                    progress_placeholder,
                    {
                        "title": "BOM 匹配完成",
                        "subtitle": f"已生成当前分页匹配结果，共 {len(bom_sheet_frames)} 个分页，下载文件已保留原分页结构",
                        "current_text": f"当前分页：{selected_sheet_name}",
                        "processed_rows": total_workbook_rows,
                        "total_rows": total_workbook_rows,
                        "percent": 100.0,
                        "done": True,
                        "elapsed_seconds": 0.0,
                            "chips": [
                                {"label": "阶段", "value": "完成", "tone": "success"},
                                {"label": "状态", "value": "可下载", "tone": "success"},
                                {"label": "匹配成功", "value": str(cached_success_count), "tone": "success"},
                                {"label": "解析失败", "value": str(cached_fail_count), "tone": "success" if cached_fail_count == 0 else "fail"},
                                {"label": "无匹配", "value": str(cached_no_match_count), "tone": "success" if cached_no_match_count == 0 else "warn"},
                            ],
                            "summary_lines": cached_summary_lines,
                        },
                    )
            else:
                render_bom_progress_card(
                    progress_placeholder,
                    {
                        "title": "BOM 文件读取完成",
                        "subtitle": f"已加载 {len(bom_sheet_frames)} 个分页，共 {total_workbook_rows} 行原始数据，正在准备列识别",
                        "current_text": f"当前分页：{selected_sheet_name}",
                        "processed_rows": 0,
                        "total_rows": total_workbook_rows,
                        "percent": 8.0,
                        "done": False,
                        "elapsed_seconds": 0.0,
                        "chips": [
                            {"label": "阶段", "value": "读取完成", "tone": "success"},
                            {"label": "分页", "value": f"{len(bom_sheet_frames)}", "tone": "success"},
                            {"label": "行数", "value": str(total_workbook_rows), "tone": "success"},
                        ],
                    },
                )

            st.markdown('<div class="section-title">BOM原始内容预览</div>', unsafe_allow_html=True)
            preview_df = bom_df.head(20).copy()
            preview_df = preview_df.astype(object).where(pd.notna(preview_df), "")
            preview_html = render_static_preview_table(
                preview_df,
                wrapper_class="bom-result-table-wrap",
            )
            if preview_html:
                components.html(
                    preview_html,
                    height=estimate_bom_preview_iframe_height(len(preview_df)),
                    scrolling=False,
                )

            guessed_mapping = guess_bom_column_mapping(bom_df)
            bom_column_options = [BOM_NONE_OPTION] + list(bom_df.columns)
            sheet_mapping_store = st.session_state["_bom_sheet_mappings"]
            for item in bom_sheet_frames:
                sheet_name = clean_text(item.get("sheet_name", ""))
                if sheet_name == "":
                    continue
                if sheet_name not in sheet_mapping_store or not isinstance(sheet_mapping_store.get(sheet_name), dict):
                    sheet_mapping_store[sheet_name] = guess_bom_column_mapping(item.get("df", pd.DataFrame()))
            if "_bom_manual_mapping_open" not in st.session_state:
                st.session_state["_bom_manual_mapping_open"] = False
            consume_bom_manual_mapping_toggle(workbook_signature)
            manual_mapping_open = bool(st.session_state.get("_bom_manual_mapping_open", False))
            stored_manual_mapping = sheet_mapping_store.get(selected_sheet_name, guessed_mapping)
            st.markdown(build_bom_manual_mapping_toggle_html(workbook_signature), unsafe_allow_html=True)

            def resolve_bom_mapping_value(role, fallback_mapping):
                value = clean_text(stored_manual_mapping.get(role, ""))
                if value not in bom_column_options:
                    value = clean_text((fallback_mapping or {}).get(role, ""))
                if value not in bom_column_options:
                    value = BOM_NONE_OPTION
                return None if value == BOM_NONE_OPTION else value

            selected_mapping = {
                role: resolve_bom_mapping_value(role, guessed_mapping)
                for role in ["model", "spec", "name", "quantity"]
            }

            if manual_mapping_open:
                st.caption("系统会先自动猜测常见表头，你也可以手动改成正确的型号列、规格列、品名列、数量列。再次点击按钮即可收起。")
                mapping_cols = st.columns(4)
                selected_mapping = {}
                for col_ui, role in zip(mapping_cols, ["model", "spec", "name", "quantity"]):
                    default_value = resolve_bom_mapping_value(role, guessed_mapping)
                    widget_value = col_ui.selectbox(
                        BOM_ROLE_LABELS[role],
                        bom_column_options,
                        index=bom_column_options.index(default_value if default_value in bom_column_options else BOM_NONE_OPTION),
                        key=f"bom_{workbook_signature}_{selected_sheet_name}_{role}_column",
                    )
                    selected_mapping[role] = None if widget_value == BOM_NONE_OPTION else widget_value
                sheet_mapping_store[selected_sheet_name] = selected_mapping
            else:
                sheet_mapping_store[selected_sheet_name] = selected_mapping

            parse_columns = [selected_mapping.get(x) for x in ["model", "spec", "name"] if selected_mapping.get(x)]
            if not parse_columns:
                st.warning("请至少指定一个用于解析的列（型号列、规格列、品名列三者至少选一个）。")
            else:
                current_bom_signature = build_bom_workbook_run_signature(uploaded_file, sheet_mapping_store)
                stored_bom_signature = st.session_state.get("_bom_result_signature", "")
                if stored_bom_signature != current_bom_signature:
                    st.session_state.pop("_bom_result_signature", None)
                    st.session_state.pop("_bom_result_df", None)
                    st.session_state.pop("_bom_export_bytes", None)
                    st.session_state.pop("_bom_sheet_results", None)

                if len(parse_columns) != len(set(parse_columns)):
                    st.info("当前有重复列被同时用于多个角色，系统会按你的选择继续解析。")

                if stored_bom_signature != current_bom_signature:
                    if not ensure_component_data_ready("BOM 匹配"):
                        render_bom_progress_card(
                            progress_placeholder,
                            {
                                "title": "BOM 匹配失败",
                                "subtitle": "当前环境缺少可用数据库，无法开始匹配",
                                "current_text": "请先确认部署数据包或本地数据库是否完整",
                                "processed_rows": 0,
                                "total_rows": total_workbook_rows,
                                "percent": 100.0,
                                "done": True,
                                "elapsed_seconds": 0.0,
                                "chips": [
                                    {"label": "阶段", "value": "匹配终止", "tone": "fail"},
                                    {"label": "状态", "value": "数据不可用", "tone": "fail"},
                                ],
                            },
                        )
                        st.warning("当前环境缺少可用数据库，请先确认部署数据包或本地数据库。")
                    else:
                        progress_state_holder = {"state": {}}

                        def update_bom_progress(progress_state):
                            progress_state_holder["state"] = progress_state or {}
                            render_bom_progress_card(progress_placeholder, progress_state_holder["state"])

                        update_bom_progress({
                            "title": "BOM 正在匹配",
                            "subtitle": f"正在解析上传文件并匹配元器件库（分页 {selected_sheet_name}）",
                            "current_text": "准备开始 BOM 匹配",
                            "processed_rows": 0,
                            "total_rows": total_workbook_rows,
                            "percent": 8.0,
                            "done": False,
                            "elapsed_seconds": 0.0,
                            "chips": [
                                {"label": "阶段", "value": "匹配中", "tone": "warn"},
                                {"label": "分页", "value": f"{len(bom_sheet_frames)}", "tone": "warn"},
                                {"label": "当前", "value": selected_sheet_name, "tone": "warn"},
                            ],
                        })
                        sheet_results = build_bom_workbook_sheet_results(bom_workbook, sheet_mapping_store, progress_callback=update_bom_progress)
                        sheet_result_map = {item["sheet_name"]: item.get("result_df", pd.DataFrame()) for item in sheet_results}

                        final_match_state = dict(progress_state_holder["state"] or {})
                        base_match_chips = [
                            chip for chip in (progress_state_holder["state"] or {}).get("chips", [])
                            if clean_text(chip.get("label", "")) not in {"阶段", "状态"}
                        ]
                        final_match_state.update({
                            "title": "BOM 正在生成下载文件",
                            "subtitle": "匹配已完成，正在生成导出 Excel",
                            "current_text": "请稍候，下载文件正在生成",
                            "processed_rows": total_workbook_rows,
                            "total_rows": total_workbook_rows,
                            "percent": 96.0,
                            "done": False,
                            "chips": [
                                {"label": "阶段", "value": "导出中", "tone": "warn"},
                                {"label": "状态", "value": "准备下载", "tone": "warn"},
                            ] + base_match_chips,
                        })
                        render_bom_progress_card(progress_placeholder, final_match_state)
                        st.session_state["_bom_result_signature"] = current_bom_signature
                        st.session_state["_bom_sheet_results"] = sheet_result_map
                        current_bom_result_df = sheet_result_map.get(selected_sheet_name, pd.DataFrame())
                        st.session_state["_bom_result_df"] = current_bom_result_df
                        st.session_state["_bom_export_bytes"] = bom_to_excel_bytes(
                            st.session_state["_bom_result_df"],
                            bom_df,
                            source_workbook=bom_workbook,
                            sheet_results=sheet_results,
                        )
                        component_distribution_text = build_bom_component_distribution_text(
                            current_bom_result_df
                        )
                        fail_count = int((current_bom_result_df["解析状态"] == "解析失败").sum()) if isinstance(current_bom_result_df, pd.DataFrame) and "解析状态" in current_bom_result_df.columns else 0
                        no_match_count = int((current_bom_result_df["状态"] == "无匹配").sum()) if isinstance(current_bom_result_df, pd.DataFrame) and "状态" in current_bom_result_df.columns else 0
                        success_count = int((current_bom_result_df["状态"] == "匹配成功").sum()) if isinstance(current_bom_result_df, pd.DataFrame) and "状态" in current_bom_result_df.columns else 0
                        final_done_state = dict(progress_state_holder["state"] or {})
                        base_done_chips = [
                            chip for chip in (progress_state_holder["state"] or {}).get("chips", [])
                            if clean_text(chip.get("label", "")) not in {"阶段", "状态"}
                        ]
                        final_done_state.update({
                            "title": "BOM 匹配完成",
                            "subtitle": "已生成匹配结果和下载文件",
                            "current_text": "可以继续查看下方结果表或下载 Excel",
                            "processed_rows": total_workbook_rows,
                            "total_rows": total_workbook_rows,
                            "percent": 100.0,
                            "done": True,
                            "chips": [
                                {"label": "阶段", "value": "完成", "tone": "success"},
                                {"label": "状态", "value": "可下载", "tone": "success"},
                            ] + base_done_chips,
                            "summary_lines": [
                                f"解析完成：匹配成功 {success_count} 行，解析失败 {fail_count} 行，无匹配 {no_match_count} 行。",
                            ] + ([component_distribution_text] if component_distribution_text else []),
                        })
                        render_bom_progress_card(progress_placeholder, final_done_state)

                bom_sheet_results = st.session_state.get("_bom_sheet_results", {})
                bom_result_df = bom_sheet_results.get(selected_sheet_name, st.session_state.get("_bom_result_df", pd.DataFrame()))
                if isinstance(bom_result_df, pd.DataFrame) and not bom_result_df.empty:
                    bom_display_df = build_bom_display_df(bom_result_df)
                    bom_view_df = bom_display_df.copy()
                    styled_bom_result_df = style_bom_result_rows(bom_view_df)

                    fail_count = int((bom_result_df["解析状态"] == "解析失败").sum()) if "解析状态" in bom_result_df.columns else 0
                    no_match_count = int((bom_result_df["状态"] == "无匹配").sum()) if "状态" in bom_result_df.columns else 0
                    success_count = int((bom_result_df["状态"] == "匹配成功").sum()) if "状态" in bom_result_df.columns else 0
                    component_distribution_text = build_bom_component_distribution_text(bom_result_df)

                    st.markdown(f'<div class="section-title">BOM匹配结果 · {html.escape(selected_sheet_name)}</div>', unsafe_allow_html=True)
                    display_bom_result_df = format_display_df(build_bom_display_df(bom_result_df))
                    export_name_root = os.path.splitext(getattr(uploaded_file, "name", "bom"))[0] or "bom"
                    export_filename = f"{export_name_root}_匹配后.xlsx"
                    download_footer_html = build_bom_download_footer_html(
                        st.session_state.get("_bom_export_bytes", b""),
                        export_filename,
                        container_class="bom-download-footer-outside",
                    )
                    clickable_bom_html = render_clickable_result_table(
                        display_bom_result_df,
                        hide_columns=[],
                        show_official_status=False,
                        wrapper_class="bom-result-table-wrap",
                        footer_html="",
                        outer_footer_html=download_footer_html,
                    )
                    if clickable_bom_html:
                        components.html(
                            clickable_bom_html,
                            height=estimate_bom_result_iframe_height(
                                len(display_bom_result_df),
                            ),
                            scrolling=False,
                        )
                    else:
                        st.markdown(download_footer_html, unsafe_allow_html=True)
                else:
                    st.info("正在等待当前自动匹配结果生成。")

    except Exception as e:
        try:
            render_bom_progress_card(
                progress_placeholder,
                {
                    "title": "BOM 处理失败",
                    "subtitle": "上传文件或匹配流程发生异常",
                    "current_text": str(e),
                    "processed_rows": 0,
                    "total_rows": 0,
                    "percent": 100.0,
                    "done": True,
                    "elapsed_seconds": 0.0,
                    "chips": [
                        {"label": "阶段", "value": "异常", "tone": "fail"},
                        {"label": "状态", "value": "请重试", "tone": "fail"},
                    ],
                },
            )
        except Exception:
            pass
        st.error(f"BOM 处理失败：{e}")


if should_render_inline_footer():
    st.markdown(
        '''
        <div class="app-footer-shell">
        <div class="app-footer">
            网站管理员：Terry Wu　　
            系统问题请与管理员联系：
            <a href="mailto:terry@fruition-sz.com" style="color: #1565c0; text-decoration: none;">terry@fruition-sz.com</a>
        </div>
        </div>
        ''',
        unsafe_allow_html=True
    )
startup_trace("after_footer")




